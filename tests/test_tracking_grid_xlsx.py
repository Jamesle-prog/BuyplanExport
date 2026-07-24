"""Tracking grid ⇄ Excel round-trip (build_tracking_grid_xlsx / parse)."""
from __future__ import annotations

import io

import openpyxl
import pytest

from po_extractor.exporters.tracking_grid_xlsx import (
    build_tracking_grid_xlsx, parse_tracking_grid_xlsx,
)
from po_extractor.store._factory_progress_schema import MILESTONE_STAGES

_S0 = MILESTONE_STAGES[0][0]        # first milestone stage key (e.g. fabric_purchase)
_S1 = MILESTONE_STAGES[1][0]


def _record(po="PO-1", style="STY-A", **stage_dates):
    """A tracking record with the given ``{stage}_planned/_actual`` values."""
    rec = {"po_number": po, "style": style, "factory": "Factory X"}
    for stage, _ in MILESTONE_STAGES:
        rec.setdefault(f"{stage}_planned", "")
        rec.setdefault(f"{stage}_actual", "")
    rec.update(stage_dates)
    return rec


def test_roundtrip_preserves_planned_and_actual():
    rec = _record(**{f"{_S0}_planned": "2026-08-01",
                     f"{_S0}_actual":  "2026-08-03",
                     f"{_S1}_planned": "2026-08-10"})
    parsed = parse_tracking_grid_xlsx(build_tracking_grid_xlsx([rec]))
    assert len(parsed["rows"]) == 1
    fields = parsed["rows"][0]["fields"]
    assert fields[f"{_S0}_planned"] == "2026-08-01"
    assert fields[f"{_S0}_actual"] == "2026-08-03"
    assert fields[f"{_S1}_planned"] == "2026-08-10"
    assert parsed["rows"][0]["po_number"] == "PO-1"
    assert parsed["rows"][0]["style"] == "STY-A"


def test_actual_date_marks_stage_done():
    rec = _record(**{f"{_S0}_actual": "2026-08-03"})
    fields = parse_tracking_grid_xlsx(build_tracking_grid_xlsx([rec]))["rows"][0]["fields"]
    assert fields[f"{_S0}_status"] == "Done"


def test_planned_only_never_sets_status():
    rec = _record(**{f"{_S0}_planned": "2026-08-01"})
    fields = parse_tracking_grid_xlsx(build_tracking_grid_xlsx([rec]))["rows"][0]["fields"]
    assert f"{_S0}_status" not in fields


def test_blank_cells_never_appear_as_updates():
    """An all-empty record yields no rows — a blank export can't wipe data."""
    parsed = parse_tracking_grid_xlsx(build_tracking_grid_xlsx([_record()]))
    assert parsed["rows"] == []
    assert parsed["rows_seen"] == 1        # the row was seen, just had nothing


def test_only_filled_stage_is_emitted():
    rec = _record(**{f"{_S1}_planned": "2026-09-01"})
    fields = parse_tracking_grid_xlsx(build_tracking_grid_xlsx([rec]))["rows"][0]["fields"]
    assert set(fields) == {f"{_S1}_planned"}      # nothing for the untouched _S0


def test_column_reorder_still_parses():
    """Columns are matched on header text, so swapping two of them in Excel
    must not scramble which stage a date belongs to."""
    rec = _record(**{f"{_S0}_planned": "2026-08-01",
                     f"{_S1}_planned": "2026-08-10"})
    wb = openpyxl.load_workbook(io.BytesIO(build_tracking_grid_xlsx([rec])))
    ws = wb.active
    # Find the header row, then swap the two first date columns' header+value.
    hr = next(r for r in range(1, 10)
              if str(ws.cell(r, 1).value or "").startswith("PO号"))
    c1, c2 = 4, 5
    for rr in (hr, hr + 1):
        ws.cell(rr, c1).value, ws.cell(rr, c2).value = (
            ws.cell(rr, c2).value, ws.cell(rr, c1).value)
    buf = io.BytesIO(); wb.save(buf)
    fields = parse_tracking_grid_xlsx(buf.getvalue())["rows"][0]["fields"]
    assert fields[f"{_S0}_planned"] == "2026-08-01"
    assert fields[f"{_S1}_planned"] == "2026-08-10"


def test_po_column_found_by_header_when_reordered():
    """The PO column is matched by header text, not fixed index — moving it
    in Excel must not break re-import (regression: PO was read at column 1)."""
    rec = _record(po="PO-9", style="STY-Z",
                  **{f"{_S0}_planned": "2026-08-01"})
    wb = openpyxl.load_workbook(io.BytesIO(build_tracking_grid_xlsx([rec])))
    ws = wb.active
    hr = next(r for r in range(1, 10)
              if str(ws.cell(r, 1).value or "").startswith("PO号"))
    # Swap columns 1 (PO) and 3 (Factory) — header + every data row.
    for rr in range(hr, ws.max_row + 1):
        ws.cell(rr, 1).value, ws.cell(rr, 3).value = (
            ws.cell(rr, 3).value, ws.cell(rr, 1).value)
    buf = io.BytesIO(); wb.save(buf)
    parsed = parse_tracking_grid_xlsx(buf.getvalue())
    assert len(parsed["rows"]) == 1
    assert parsed["rows"][0]["po_number"] == "PO-9"
    assert parsed["rows"][0]["style"] == "STY-Z"


def test_bad_date_is_reported_not_applied():
    wb = openpyxl.load_workbook(io.BytesIO(build_tracking_grid_xlsx([_record()])))
    ws = wb.active
    hr = next(r for r in range(1, 10)
              if str(ws.cell(r, 1).value or "").startswith("PO号"))
    ws.cell(hr + 1, 4, value="not-a-date")     # first planned date column
    buf = io.BytesIO(); wb.save(buf)
    parsed = parse_tracking_grid_xlsx(buf.getvalue())
    assert parsed["rows"] == []
    assert any("isn't YYYY-MM-DD" in i for i in parsed["issues"])


def test_unreadable_file_raises():
    with pytest.raises(ValueError):
        parse_tracking_grid_xlsx(b"not a workbook")


def test_missing_header_raises():
    wb = openpyxl.Workbook()
    wb.active["A1"] = "something else"
    buf = io.BytesIO(); wb.save(buf)
    with pytest.raises(ValueError):
        parse_tracking_grid_xlsx(buf.getvalue())


def test_multiple_records_each_row_independent():
    recs = [
        _record(po="PO-1", style="A", **{f"{_S0}_planned": "2026-08-01"}),
        _record(po="PO-2", style="B", **{f"{_S0}_actual": "2026-08-05"}),
    ]
    rows = {(r["po_number"], r["style"]): r["fields"]
            for r in parse_tracking_grid_xlsx(build_tracking_grid_xlsx(recs))["rows"]}
    assert rows[("PO-1", "A")][f"{_S0}_planned"] == "2026-08-01"
    assert rows[("PO-2", "B")][f"{_S0}_actual"] == "2026-08-05"
    assert rows[("PO-2", "B")][f"{_S0}_status"] == "Done"
