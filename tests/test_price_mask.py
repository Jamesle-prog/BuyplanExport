"""Tests for price masking — PDF token pattern + Excel column masking."""
from __future__ import annotations

import openpyxl
import pytest

from po_extractor.utils.price_mask import _PRICE_RE, mask_prices_excel


# ── PDF token pattern ────────────────────────────────────────────────────────

@pytest.mark.parametrize("token", [
    "4.17", "69.00", "1234.00",
    "1,234.00", "12,345.67",          # thousands separators (was missed before)
    "$4.17", "€1,000.00", "£99.99", "¥50.00",   # currency prefixes
    "0.75%", "000.00", "12.50%",      # discount / tariff percentages
    ".75%", ".00",                    # leading-dot form (discount in prose)
])
def test_price_regex_matches_prices(token):
    assert _PRICE_RE.match(token), f"{token!r} should be treated as a price"


@pytest.mark.parametrize("token", [
    "69",          # bare integer — quantity/PO#, must NOT mask
    "123456789012",  # UPC
    "2026",        # year
    "4.1", "4.175",  # not exactly two decimals
    "abc", "", "%",   # non-numeric / bare percent sign (a label, not a value)
])
def test_price_regex_rejects_non_prices(token):
    assert not _PRICE_RE.match(token), f"{token!r} should NOT be masked"


# ── PDF keep-set: retail MSRP must never be redacted ─────────────────────────

def _make_pdf(path, *tokens):
    import fitz
    doc = fitz.open()
    page = doc.new_page()
    y = 72
    for tok in tokens:
        page.insert_text((72, y), tok, fontsize=11)
        y += 24
    doc.save(str(path))
    doc.close()


def _pdf_text(path):
    import fitz
    doc = fitz.open(path)
    try:
        return "\n".join(p.get_text() for p in doc)
    finally:
        doc.close()


def test_pdf_keep_leaves_msrp_visible(tmp_path):
    """A retail price passed in `keep` survives even though the regex would
    otherwise mask it; the confidential cost is still redacted. '$59' matches
    the '59.00' token (numeric compare, trailing zeros ignored)."""
    from po_extractor.utils.price_mask import mask_prices
    pdf = tmp_path / "po.pdf"
    _make_pdf(pdf, "Unit Cost 4.17", "MSRP 59.00")
    out = mask_prices(str(pdf), str(tmp_path), keep=["$59"])
    txt = _pdf_text(out)
    assert "59.00" in txt        # retail MSRP kept
    assert "4.17" not in txt     # confidential cost redacted


def test_pdf_without_keep_or_label_masks_every_price(tmp_path):
    """Sanity check: an unlabelled price with no keep-set IS masked — a value is
    only spared when it's a known/labelled retail price."""
    from po_extractor.utils.price_mask import mask_prices
    pdf = tmp_path / "po.pdf"
    _make_pdf(pdf, "Unit Cost 4.17", "Extended Cost 59.00")   # no MSRP label
    out = mask_prices(str(pdf), str(tmp_path))       # no keep
    txt = _pdf_text(out)
    assert "59.00" not in txt and "4.17" not in txt


def test_pdf_keeps_labelled_msrp_without_passed_keep(tmp_path):
    """The PO labels its retail price ('MSRP: $54.00'); the masker reads that
    label from the file itself and keeps the value — even when the caller passes
    NO keep-set (the DKNY/DUKHSP parser captures no MSRP). The cost is masked."""
    from po_extractor.utils.price_mask import mask_prices, _retail_values_from_text
    assert _retail_values_from_text("MSRP: $54.00  SRP $59  Cost 4.17") == {54.0, 59.0}
    pdf = tmp_path / "po.pdf"
    _make_pdf(pdf, "MSRP: $54.00", "Unit Cost 4.17", "Extended Cost 11,793.60")
    out = mask_prices(str(pdf), str(tmp_path))       # no keep argument
    txt = _pdf_text(out)
    assert "54.00" in txt                            # labelled MSRP kept
    assert "4.17" not in txt and "11,793.60" not in txt   # costs masked


def test_ai_price_prompt_excludes_retail():
    """The AI is instructed to keep retail prices visible, not mask them."""
    from po_extractor.utils.price_mask import _AI_PRICE_SYSTEM, _AI_COLUMN_SYSTEM
    for p in (_AI_PRICE_SYSTEM, _AI_COLUMN_SYSTEM):
        assert "MSRP" in p and ("public" in p.lower() or "retail" in p.lower())


# ── Excel column masking ─────────────────────────────────────────────────────

def _build_xlsx(tmp_path, rows):
    p = tmp_path / "in.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    wb.save(p)
    return str(p)


def _read_masked(out_path):
    wb = openpyxl.load_workbook(out_path)
    ws = wb.active
    return [[c.value for c in row] for row in ws.iter_rows()]


def test_excel_masks_keyword_columns_and_keeps_others(tmp_path):
    src = _build_xlsx(tmp_path, [
        ["Style", "Qty", "FOB Price", "MSRP", "Line Total ($)"],
        ["ST1",   100,   4.17,        "$69.00", "1,234.00"],
        ["ST2",   50,    3.75,        59.00,    617.00],
    ])
    out = mask_prices_excel(src, str(tmp_path))
    rows = _read_masked(out)

    # Header row untouched
    assert rows[0] == ["Style", "Qty", "FOB Price", "MSRP", "Line Total ($)"]
    # Style + Qty preserved (Qty has no price keyword)
    assert rows[1][0] == "ST1" and rows[1][1] == 100
    assert rows[2][0] == "ST2" and rows[2][1] == 50
    # FOB Price + Line Total masked; MSRP is PUBLIC retail → kept
    for r in (1, 2):
        assert rows[r][2] == "***"   # FOB Price (confidential cost)
        assert rows[r][4] == "***"   # Line Total ($) (confidential)
    assert rows[1][3] == "$69.00" and rows[2][3] == 59.00   # MSRP not masked


def test_excel_never_masks_retail_prices(tmp_path):
    """MSRP / SRP / RRP / Suggested Retail are public — never masked, even when
    the header also contains a generic price word."""
    src = _build_xlsx(tmp_path, [
        ["Style", "FOB", "MSRP", "SRP", "Suggested Retail Price", "Unit Cost"],
        ["ST1",   4.17,  69.00,  75.00, 79.99,                    3.50],
    ])
    out = mask_prices_excel(src, str(tmp_path))
    rows = _read_masked(out)
    assert rows[1][1] == "***"      # FOB masked
    assert rows[1][5] == "***"      # Unit Cost masked
    assert rows[1][2] == 69.00      # MSRP kept
    assert rows[1][3] == 75.00      # SRP kept
    assert rows[1][4] == 79.99      # "Suggested Retail Price" kept (despite "price")


def test_excel_leaves_text_cells_in_price_columns(tmp_path):
    # A text cell in a genuine price column ("Unit Price") stays; the number masks.
    src = _build_xlsx(tmp_path, [
        ["Vendor", "Unit Price"],
        ["Macy's", 12.50],
    ])
    out = mask_prices_excel(src, str(tmp_path))
    rows = _read_masked(out)
    assert rows[1][0] == "Macy's"   # text preserved
    assert rows[1][1] == "***"      # numeric price masked


def test_excel_masks_formula_cells_in_price_columns(tmp_path):
    """A formula in a price column recalculates on open, leaking the price —
    it must be masked, not left as the (string) formula text."""
    p = tmp_path / "f.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Style", "Qty", "Unit Price", "Line Total"])
    ws.append(["ST1", 100, 4.17, "=B2*C2"])   # formula in a price column
    wb.save(p)
    out = mask_prices_excel(str(p), str(tmp_path))
    rows = _read_masked(out)
    assert rows[1][2] == "***"        # numeric price masked
    assert rows[1][3] == "***"        # formula masked, not "=B2*C2"


def test_excel_no_price_columns_is_noop(tmp_path):
    src = _build_xlsx(tmp_path, [["Style", "Color", "Qty"], ["ST1", "BLACK", 100]])
    out = mask_prices_excel(src, str(tmp_path))
    rows = _read_masked(out)
    assert rows[1] == ["ST1", "BLACK", 100]


def test_excel_rejects_legacy_xls(tmp_path):
    p = tmp_path / "old.xls"
    p.write_bytes(b"not a real xls")
    with pytest.raises(ValueError, match="legacy .xls"):
        mask_prices_excel(str(p), str(tmp_path))


# ── AI-assisted detection (DeepSeek call mocked) ─────────────────────────────

import po_extractor.utils.price_mask as pm


def test_detect_prices_ai_parses_and_normalizes(monkeypatch):
    monkeypatch.setattr(pm, "_deepseek_json",
                        lambda *a, **k: {"prices": ["$1,200.00", "69", "  "]})
    got = pm.detect_prices_ai("irrelevant", api_key="k")
    assert got == {"1200.00", "69"}   # currency/commas stripped, blanks dropped


def test_detect_prices_ai_no_key_is_empty(monkeypatch):
    # No API call attempted without a key → empty set, regex still the floor.
    called = {"n": 0}
    def _spy(*a, **k):
        called["n"] += 1
        return {}
    monkeypatch.setattr(pm, "_deepseek_json", _spy)
    assert pm.detect_prices_ai("text", api_key="") == set()
    # _deepseek_json itself guards empty key, but detect_prices_ai passes through;
    # the real guard is inside _deepseek_json (tested below).


def test_deepseek_json_returns_empty_without_key():
    assert pm._deepseek_json("sys", "user", api_key="", model="m") == {}


def test_ai_columns_union_with_keywords(tmp_path, monkeypatch):
    # A price column whose header has NO keyword ("Deal") — only AI can catch it.
    monkeypatch.setattr(pm, "_deepseek_json",
                        lambda *a, **k: {"price_headers": ["Deal"]})
    src = _build_xlsx(tmp_path, [
        ["Style", "Deal", "Qty"],
        ["ST1",   4.17,   100],
    ])
    out = mask_prices_excel(str(src) if not isinstance(src, str) else src,
                            str(tmp_path), api_key="k")
    rows = _read_masked(out)
    assert rows[1][1] == "***"   # "Deal" column masked via AI
    assert rows[1][2] == 100     # Qty untouched


def test_ai_failure_falls_back_to_keywords(tmp_path, monkeypatch):
    # AI returns nothing (as if the call failed) — keyword detection still works.
    monkeypatch.setattr(pm, "_deepseek_json", lambda *a, **k: {})
    src = _build_xlsx(tmp_path, [["Style", "FOB"], ["ST1", 4.17]])
    out = mask_prices_excel(src, str(tmp_path), api_key="k")
    rows = _read_masked(out)
    assert rows[1][1] == "***"   # FOB still masked by keyword


# ── batch: parallel AI detection + progress ───────────────────────────────────

def test_mask_batch_runs_ai_concurrently_and_reports_progress(monkeypatch, tmp_path):
    from po_extractor.utils import price_mask as pm
    seen_text, mask_calls = [], []
    monkeypatch.setattr(pm, "_pdf_text", lambda p: "text:" + p)

    def _detect(text, key, model):
        seen_text.append(text)
        return {"9.99"}
    monkeypatch.setattr(pm, "detect_prices_ai", _detect)

    def _mask(path, out, api_key=None, model="", ai_prices=None, keep=None,
              full_text=None):
        mask_calls.append((path, ai_prices))
        return path + ".masked"
    monkeypatch.setattr(pm, "mask_prices", _mask)

    prog = []
    paths = [f"f{i}.pdf" for i in range(5)]
    out = pm.mask_prices_batch(paths, str(tmp_path), api_key="k",
                               on_progress=lambda d, t: prog.append((d, t)))

    assert out == [f"{p}.masked" for p in paths]
    assert sorted(seen_text) == sorted("text:" + p for p in paths)  # AI once per file
    # the redaction pass gets the PRECOMPUTED set (never re-calls AI)
    assert all(ai == {"9.99"} for _, ai in mask_calls)
    assert prog[-1] == (5, 5)                       # progress reached total


def test_mask_batch_serial_without_ai_still_reports_progress(monkeypatch, tmp_path):
    from po_extractor.utils import price_mask as pm
    monkeypatch.setattr(pm, "mask_prices",
                        lambda p, o, api_key=None, model="", ai_prices=None, keep=None,
                        full_text=None: p + ".m")
    prog = []
    out = pm.mask_prices_batch(["a.pdf", "b.pdf"], str(tmp_path),
                               on_progress=lambda d, t: prog.append((d, t)))
    assert out == ["a.pdf.m", "b.pdf.m"] and prog[-1] == (2, 2)
