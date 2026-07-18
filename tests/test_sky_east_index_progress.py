"""Tests for the Sky East buy plan Index sheet's 🏭 Tracking enrichment —
best-effort fill of the factory/schedule columns (previously always blank
template headers) from production_tracking, joined on (Zalando PO, style).
"""
from __future__ import annotations

import pandas as pd
import pytest
from openpyxl import load_workbook

from po_extractor.exporters.sky_east_buyplan_export import (
    export_sky_east_buyplan,
    _enrich_sheet_meta_with_progress,
    _INDEX_MILESTONE_MAP,
)
from po_extractor.store.production_tracking_store import ProductionTrackingStore
from po_extractor.lookups.progress_lookup import PCColorMatch


@pytest.fixture(autouse=True)
def _isolated_sky_east_store(tmp_path, monkeypatch):
    from po_extractor.store.sky_east_store import SkyEastStore
    store = SkyEastStore(str(tmp_path / "_isolated_sky_east.db"))
    monkeypatch.setattr("po_extractor.store.get_sky_east_store", lambda: store)
    return store


@pytest.fixture
def two_style_df():
    return pd.DataFrame([
        {
            "pc_no": "HHPPC048", "style": "DR5124", "brand": "Anna Field",
            "contract_no": "26302-ZA7148", "article_name": "LACE DRESS",
            "zalando_po": "PO001", "config_sku": "C1", "color_name": "dark blue",
            "fabric_item_no": "HHN-JA-01715", "fabrication": "cotton",
            "ex_fty_date": "2026-08-11",
            "xs": 30, "s": 82, "m": 119, "l": 102, "xl": 67, "xxl": 0,
        },
        {
            "pc_no": "HHPPC048", "style": "DR4578", "brand": "Anna Field",
            "contract_no": "26302-ZA7149", "article_name": "LONG SLEEVE DRESS",
            "zalando_po": "PO002", "config_sku": "C2", "color_name": "green",
            "fabric_item_no": "HHN-JSRSR-04068", "fabrication": "poly",
            "ex_fty_date": "2026-08-11",
            "xs": 22, "s": 61, "m": 90, "l": 77, "xl": 50, "xxl": 0,
        },
    ])


@pytest.fixture
def pc_color_lookup():
    return {
        ("HHPPC048", "DR5124", "Dark Blue"): PCColorMatch("藏青", "503", ""),
        ("HHPPC048", "DR4578", "Green"):     PCColorMatch("绿色", "602", ""),
    }


def _pt_store(tmp_path):
    return ProductionTrackingStore(str(tmp_path / "po_history.db"))


# ── Unit tests: _enrich_sheet_meta_with_progress ────────────────────────────

def test_enrich_fills_mapped_fields_from_planned_dates(tmp_path, monkeypatch):
    store = _pt_store(tmp_path)
    monkeypatch.setattr(
        "po_extractor.store.get_production_tracking_store", lambda: store)
    store.upsert(
        po_number="PO001", style="DR5124", factory="Factory Alpha",
        company="Sky East", updated_by="t", overall_notes="",
        use_substitute_materials=1,
        stage_fields={
            "fabric_purchase_planned": "2026-01-05",
            "trim_purchase_planned":   "2026-01-06",
            "pp_sample_planned":       "2026-01-20",
            "cutting_planned":         "2026-02-01",
            "sewing_planned":          "2026-02-10",
        },
        dep_fields={}, qc_fields={},
    )

    meta = [{"po_no": "PO001", "style": "DR5124"}]
    _enrich_sheet_meta_with_progress(meta)

    assert meta[0]["factory"] == "Factory Alpha"
    assert meta[0]["fabric_arrival"] == "2026-01-05"
    assert meta[0]["trim_arrival"] == "2026-01-06"
    assert meta[0]["sample_confirm"] == "2026-01-20"
    assert meta[0]["cutting_complete"] == "2026-02-01"
    assert meta[0]["sewing_complete"] == "2026-02-10"
    # Unset stages -> blank, not missing/KeyError.
    assert meta[0]["factory_delivery"] == ""
    assert meta[0]["bulk_pattern"] == ""


def test_enrich_leaves_blank_for_untracked_style(tmp_path, monkeypatch):
    store = _pt_store(tmp_path)
    monkeypatch.setattr(
        "po_extractor.store.get_production_tracking_store", lambda: store)

    meta = [{"po_no": "PO-NOT-TRACKED", "style": "STYX"}]
    _enrich_sheet_meta_with_progress(meta)

    assert all(meta[0][f] == "" for f in _INDEX_MILESTONE_MAP)


def test_enrich_is_best_effort_on_store_failure(tmp_path, monkeypatch):
    def _boom():
        raise RuntimeError("db unavailable")
    monkeypatch.setattr(
        "po_extractor.store.get_production_tracking_store", _boom)

    meta = [{"po_no": "PO001", "style": "DR5124", "existing": "keep-me"}]
    _enrich_sheet_meta_with_progress(meta)  # must not raise

    assert meta[0]["existing"] == "keep-me"
    assert all(meta[0][f] == "" for f in _INDEX_MILESTONE_MAP)


def test_enrich_mutates_multiple_entries_independently(tmp_path, monkeypatch):
    store = _pt_store(tmp_path)
    monkeypatch.setattr(
        "po_extractor.store.get_production_tracking_store", lambda: store)
    store.upsert(
        po_number="PO001", style="A", factory="F1", company="Sky East",
        updated_by="t", overall_notes="", use_substitute_materials=1,
        stage_fields={"cutting_planned": "2026-03-01"},
        dep_fields={}, qc_fields={},
    )
    store.upsert(
        po_number="PO001", style="B", factory="F2", company="Sky East",
        updated_by="t", overall_notes="", use_substitute_materials=1,
        stage_fields={"cutting_planned": "2026-04-01"},
        dep_fields={}, qc_fields={},
    )

    meta = [{"po_no": "PO001", "style": "A"}, {"po_no": "PO001", "style": "B"}]
    _enrich_sheet_meta_with_progress(meta)

    assert meta[0]["factory"] == "F1" and meta[0]["cutting_complete"] == "2026-03-01"
    assert meta[1]["factory"] == "F2" and meta[1]["cutting_complete"] == "2026-04-01"


# ── End-to-end: real export threads the enrichment into the Index sheet ────

def test_index_sheet_shows_tracked_style_milestones(
    two_style_df, pc_color_lookup, tmp_path, monkeypatch,
):
    store = ProductionTrackingStore(str(tmp_path / "po_history_pt.db"))
    monkeypatch.setattr(
        "po_extractor.store.get_production_tracking_store", lambda: store)
    store.upsert(
        po_number="PO001", style="DR5124", factory="Factory Alpha",
        company="Sky East", updated_by="t", overall_notes="",
        use_substitute_materials=1,
        stage_fields={"cutting_planned": "2026-02-15", "sewing_planned": "2026-02-25"},
        dep_fields={}, qc_fields={},
    )

    path, _totals = export_sky_east_buyplan(
        two_style_df, cn_lookup={}, output_dir=str(tmp_path),
        label_lookup={}, cn_code_lookup={}, cn_by_pc_lookup=pc_color_lookup,
    )
    ws = load_workbook(path)["Index"]
    headers = [c.value for c in ws[1]]
    row2 = {headers[i]: ws.cell(2, i + 1).value for i in range(len(headers))}
    row3 = {headers[i]: ws.cell(3, i + 1).value for i in range(len(headers))}

    # DR5124 is tracked (row order follows sheet_meta_list insertion == df order)
    assert row2["生产工厂"] == "Factory Alpha"
    assert row2["裁剪（计划）完成时间"] == "2026-02-15"
    assert row2["车位（计划）完成时间"] == "2026-02-25"
    # DR4578 (PO002) was never tracked -> blank, not an error, not "None" text.
    assert row3["生产工厂"] in ("", None)
    assert row3["裁剪（计划）完成时间"] in ("", None)
    # Manual-only columns are never populated by this feature.
    assert row2["裁剪数"] in ("", None)
    assert row2["出货数"] in ("", None)
