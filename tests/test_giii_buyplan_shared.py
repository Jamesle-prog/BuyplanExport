"""Shared GIII buy-plan builder (ui.giii._buyplan) — one production-plan path."""
from __future__ import annotations

import io

import openpyxl
import pandas as pd

from po_extractor.store.po_store import POStore
from po_extractor.models.po_data import POData, POMetadata, SizeRow
from ui.giii._buyplan import build_giii_production_plan, resolve_reqs


def _store(tmp_path):
    s = POStore(str(tmp_path / "t.db"))
    po = POData(metadata=POMetadata(
        po_number="DW1", style="ST1", company="GIII",
        division_name="DKNY Sportswear", destination_code="UC"),
        size_rows=[SizeRow("DW1", "ST1", "BLACK", "M", 10, "700948471565")])
    s.save_many_checked([po])
    return s


def test_builder_produces_production_plan_format(tmp_path):
    data, warns, preview = build_giii_production_plan(["DW1"], _store(tmp_path),
                                                      cprs=None)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    # the 生产计划单 format — NOT the old "Style × PO × Color" grid
    assert wb.sheetnames[0] == "Summary 汇总"
    assert "款号对比 By Style" in wb.sheetnames or "ST1" in wb.sheetnames
    assert "ST1" in wb.sheetnames
    assert preview == []                                  # no CPRS → no preview
    assert any("cprs" in w.lower() for w in warns)        # CPRS-not-configured note


def test_builder_empty_selection_returns_empty():
    data, warns, preview = build_giii_production_plan([], object(), cprs=None)
    assert data == b"" and warns == [] and preview == []


class _FakeCprs:
    def evaluate_po(self, raw):
        brand = str(raw.get("brand", "")).upper()
        if "DKNY" not in brand:
            return {"decoded": {}, "evaluation": {"results": []}}
        return {"decoded": {
                    "clientId": "a1", "clientName": brand, "channel": "WHOLESALE",
                    "accountCode": "", "warehouseCode": raw.get("warehouseCode", ""),
                    "warehouseInfo": {"region": "US", "rfid_default": True,
                                      "msrp_required_default": False},
                    "warnings": []},
                "evaluation": {"results": []}}
    def manual_image(self, image_id):
        return None


def test_resolve_reqs_builds_preview_and_reqs_from_cprs():
    meta = pd.DataFrame([{
        "po_number": "DW1", "division_name": "DKNY Sportswear",
        "destination_code": "UC", "ship_to": "ROSS DC", "buyer": "MACYS",
        "packaging": "PPK", "hanger": "",
    }])
    reqs, warns, preview = resolve_reqs(_FakeCprs(), meta)
    assert "DW1" in reqs
    assert reqs["DW1"].warehouse == "UC"
    assert reqs["DW1"].rfid == "Y" and reqs["DW1"].msrp == "N"
    assert preview and preview[0]["PO"] == "DW1" and preview[0]["品牌"] == "DKNY Sportswear"
