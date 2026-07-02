"""Direct coverage for ``export_sky_east_nukuryou`` (核料 / Template_P workbooks).

Before this file the 核料 exporter had no direct tests — only the main buy plan
path was covered.  These lock in:
  • one workbook per distinct fabric, one sheet per style within it,
  • size headers + colour-quantity aggregation,
  • the colour-miss behaviour that mirrors the buy plan (English colour shown
    alone — never ``Black(未找到)`` — plus a diagnostic comment and a
    colour-miss-log entry),
  • that the produced workbook actually re-opens (regression: the old
    ``copy.deepcopy`` path emitted a workbook whose stylesheet openpyxl and
    Excel's repair check rejected).

Every export call passes an explicit ``sky_east_store`` (a tmp DB or ``None``)
so no test writes colour-miss rows into the shared dev database.
"""
from __future__ import annotations

import pandas as pd
import pytest
from openpyxl import load_workbook

from po_extractor.exporters.sky_east_buyplan_export import (
    export_sky_east_nukuryou, _SE_TEMPLATE_P,
)
from po_extractor.lookups.progress_lookup import PCColorMatch
from po_extractor.store.color_translation_store import _normalize_color_name as _nz

pytestmark = pytest.mark.skipif(
    not _SE_TEMPLATE_P.exists(),
    reason="Sky_East_P.xlsx (Template_P) not present",
)


def _row(**kw) -> dict:
    base = dict(
        pc_no="PC1", style="DR9001", brand="Anna Field",
        contract_no="C-1", zalando_po="PO1", config_sku="S1",
        color_name="Dark Blue", fabric_item_no="HHN-A",
        xs=1, s=2, m=3, l=0, xl=0, xxl=0,
    )
    base.update(kw)
    return base


def _find_cell(ws, needle):
    for row in ws.iter_rows():
        for c in row:
            if c.value is not None and needle in str(c.value):
                return c
    return None


def test_one_workbook_per_fabric_reloads_cleanly(tmp_path):
    df = pd.DataFrame([
        _row(fabric_item_no="HHN-A", style="DR9001"),
        _row(fabric_item_no="HHN-B", style="DR9002"),
    ])
    paths = export_sky_east_nukuryou(
        df, cn_lookup={}, output_dir=str(tmp_path),
        cn_code_lookup={}, cn_by_pc_lookup=None, ai_enhance=False,
        ai_api_key="", ai_model="deepseek-chat", sky_east_store=None,
    )
    assert len(paths) == 2
    # Regression: deepcopy output was unreadable — every workbook must re-open.
    for p in paths:
        load_workbook(p)


def test_multiple_styles_same_fabric_are_separate_sheets(tmp_path):
    df = pd.DataFrame([
        _row(fabric_item_no="HHN-A", style="DR9001"),
        _row(fabric_item_no="HHN-A", style="DR9002"),
    ])
    paths = export_sky_east_nukuryou(
        df, cn_lookup={}, output_dir=str(tmp_path),
        cn_code_lookup={}, cn_by_pc_lookup=None, ai_enhance=False,
        ai_api_key="", ai_model="deepseek-chat", sky_east_store=None,
    )
    assert len(paths) == 1
    wb = load_workbook(paths[0])
    names = [n for n in wb.sheetnames if n != "Empty"]
    assert "DR9001" in names and "DR9002" in names


def test_size_headers_written(tmp_path):
    df = pd.DataFrame([_row()])
    paths = export_sky_east_nukuryou(
        df, cn_lookup={}, output_dir=str(tmp_path),
        cn_code_lookup={}, cn_by_pc_lookup=None, ai_enhance=False,
        ai_api_key="", ai_model="deepseek-chat", sky_east_store=None,
    )
    wb = load_workbook(paths[0])
    ws = wb[wb.sheetnames[0]]
    headers = {str(c.value).upper() for row in ws.iter_rows() for c in row if c.value}
    # The Template_P's detected size columns (S/M/L/XL for the shipped template,
    # more if a wider template is uploaded) must be written as header labels.
    size_vocab = {"XS", "S", "M", "L", "XL", "2XL", "XXL"}
    assert len(headers & size_vocab) >= 3


def test_colour_quantities_aggregate_across_rows(tmp_path):
    # Put quantities across S/M (present in the shipped Template_P) with unique
    # per-size sums so aggregation — not passthrough — is what we assert.
    df = pd.DataFrame([
        _row(color_name="Dark Blue", s=2, m=10, l=0, xl=0),
        _row(color_name="Dark Blue", s=3, m=20, l=0, xl=0),
    ])
    paths = export_sky_east_nukuryou(
        df, cn_lookup={}, output_dir=str(tmp_path),
        cn_code_lookup={}, cn_by_pc_lookup=None, ai_enhance=False,
        ai_api_key="", ai_model="deepseek-chat", sky_east_store=None,
    )
    wb = load_workbook(paths[0])
    ws = wb[wb.sheetnames[0]]
    cell = _find_cell(ws, "Dark Blue")
    assert cell is not None
    row_vals = [c.value for c in ws[cell.row]]
    assert 5 in row_vals and 30 in row_vals   # s 2+3=5, m 10+20=30


def test_resolved_colour_uses_label_format(tmp_path):
    df = pd.DataFrame([_row(color_name="Dark Blue")])
    cn_by_pc = {("PC1", "DR9001", _nz("Dark Blue")): PCColorMatch("藏青", "503", "")}
    paths = export_sky_east_nukuryou(
        df, cn_lookup={}, output_dir=str(tmp_path),
        cn_code_lookup={}, cn_by_pc_lookup=cn_by_pc, ai_enhance=False,
        ai_api_key="", ai_model="deepseek-chat", sky_east_store=None,
    )
    wb = load_workbook(paths[0])
    ws = wb[wb.sheetnames[0]]
    cell = _find_cell(ws, "Dark Blue")
    assert cell is not None
    assert "503|藏青" in str(cell.value)   # EN(code|cn)
    assert cell.comment is None


def test_colour_miss_shows_en_only_with_comment_and_logs(tmp_path):
    from po_extractor.store.sky_east_store import SkyEastStore

    store = SkyEastStore(str(tmp_path / "se.db"))
    df = pd.DataFrame([_row(color_name="Dark Blue")])
    # cn_by_pc_lookup is a (non-None) dict → 大货进度表 is the selected source,
    # but it has no matching key → a genuine miss.
    paths = export_sky_east_nukuryou(
        df, cn_lookup={}, output_dir=str(tmp_path),
        cn_code_lookup={}, cn_by_pc_lookup={}, ai_enhance=False,
        ai_api_key="", ai_model="deepseek-chat", sky_east_store=store,
    )
    wb = load_workbook(paths[0])
    ws = wb[wb.sheetnames[0]]
    cell = _find_cell(ws, "Dark Blue")
    assert cell is not None
    assert str(cell.value) == "Dark Blue"        # English alone, no "(未找到)"
    assert "未找到" not in str(cell.value)
    assert cell.comment is not None
    assert "Dark Blue" in cell.comment.text      # client PO colour shown

    misses = store.list_color_misses()
    assert len(misses) == 1
    assert misses.iloc[0]["style"] == "DR9001"
    assert misses.iloc[0]["source"] == "progress"


def test_colour_miss_logging_disabled_when_store_none(tmp_path):
    # sky_east_store=None must fully disable logging (no auto-fetch).
    df = pd.DataFrame([_row(color_name="Dark Blue")])
    paths = export_sky_east_nukuryou(
        df, cn_lookup={}, output_dir=str(tmp_path),
        cn_code_lookup={}, cn_by_pc_lookup={}, ai_enhance=False,
        ai_api_key="", ai_model="deepseek-chat", sky_east_store=None,
    )
    wb = load_workbook(paths[0])
    ws = wb[wb.sheetnames[0]]
    cell = _find_cell(ws, "Dark Blue")
    # Behaviour still correct even with logging disabled: EN alone + comment.
    assert str(cell.value) == "Dark Blue"
    assert cell.comment is not None
