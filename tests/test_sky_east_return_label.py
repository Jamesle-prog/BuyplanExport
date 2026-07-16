"""Sky East order parser: Return Label column extraction.

The client PO's Return Label column indicates (per style/color/PO line,
same granularity as size quantities) whether that item needs a return
label. parse() normalizes the raw cell value to "Yes" / "No" / "NA".
"""
from __future__ import annotations

import openpyxl

from po_extractor.parsers.sky_east_order import parse, _normalize_return_label

_COLS = [
    "Item", "Style No.", "PO Number", "Config SKU", "Article Name", "Picture",
    "Fabric No.", "Fabrication", "Brand", "Color Name", "Color Code",
    "Launch Date", "XS", "S", "M", "L", "XL", "Total Qty", "FOB USD",
    "Total Cost", "Ex-Fty", "Return Label",
]
_CONTRACT_LABELS = [
    ("PC No.",        "PCTEST01"),
    ("Date",          "2026-01-01"),
    ("Party A",       "BuyerCo"),
    ("Party B",       "SellerCo"),
    ("Currency",      "USD"),
    ("Payment Terms", "TT"),
    ("Trade Term",    "FOB"),
]
_HEADER_ROW = 9
_DATA_START = 10


def _make_order_file(tmp_path, rows: list[dict]):
    wb = openpyxl.Workbook()
    ws = wb.active

    for ri, (label, val) in enumerate(_CONTRACT_LABELS, start=1):
        ws.cell(row=ri, column=1, value=label)
        ws.cell(row=ri, column=5, value=val)

    for ci, label in enumerate(_COLS, start=1):
        ws.cell(row=_HEADER_ROW, column=ci, value=label)

    col_idx = {label: i + 1 for i, label in enumerate(_COLS)}
    for ri, row in enumerate(rows, start=_DATA_START):
        for key, val in row.items():
            ws.cell(row=ri, column=col_idx[key], value=val)

    path = tmp_path / "order.xlsx"
    wb.save(str(path))
    wb.close()
    return str(path)


def test_return_label_yes_and_no_are_extracted(tmp_path):
    path = _make_order_file(tmp_path, [
        {"Item": 1, "Style No.": "STYLE-A", "PO Number": "PO001",
         "XS": 10, "Return Label": "Yes"},
        {"Item": 2, "Style No.": "STYLE-B", "PO Number": "PO002",
         "XS": 10, "Return Label": "No"},
    ])
    contract = parse(path)
    by_style = {i.style: i.return_label for i in contract.items}
    assert by_style == {"STYLE-A": "Yes", "STYLE-B": "No"}


def test_return_label_blank_or_missing_column_defaults_to_na(tmp_path):
    path = _make_order_file(tmp_path, [
        {"Item": 1, "Style No.": "STYLE-A", "PO Number": "PO001", "XS": 10},
    ])
    contract = parse(path)
    assert contract.items[0].return_label == "NA"


def test_return_label_unrecognized_value_defaults_to_na(tmp_path):
    path = _make_order_file(tmp_path, [
        {"Item": 1, "Style No.": "STYLE-A", "PO Number": "PO001",
         "XS": 10, "Return Label": "Maybe"},
    ])
    contract = parse(path)
    assert contract.items[0].return_label == "NA"


def test_normalize_return_label_accepts_common_variants():
    assert _normalize_return_label("Yes") == "Yes"
    assert _normalize_return_label("y") == "Yes"
    assert _normalize_return_label("TRUE") == "Yes"
    assert _normalize_return_label("1") == "Yes"
    assert _normalize_return_label("是") == "Yes"
    assert _normalize_return_label("No") == "No"
    assert _normalize_return_label("n") == "No"
    assert _normalize_return_label("FALSE") == "No"
    assert _normalize_return_label("0") == "No"
    assert _normalize_return_label("否") == "No"
    assert _normalize_return_label(None) == "NA"
    assert _normalize_return_label("") == "NA"
    assert _normalize_return_label("n/a") == "NA"
