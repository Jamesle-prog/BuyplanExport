"""Tests for the enriched GIII production plan (生产计划单) — the buy plan.

Covers the v2.40.0 enrichment: 面料 rows from the 款式面料表格, 合同号 from
the 大货进度表, 进度表 CN colours, CPRS 红色箱贴纸/主箱唛 (text + artwork),
NaN cleanup, and the all-zero-row filter in export_buyplan.
"""
from __future__ import annotations

import io

import numpy as np
import openpyxl
import pandas as pd
import pytest

from po_extractor.exporters.giii_production_plan import generate_giii_production_plan
from po_extractor.lookups.progress_lookup import _norm_key
from po_extractor.models.fabric_part import FabricPart


# ── stub store ────────────────────────────────────────────────────────────────

class _Store:
    def __init__(self, df_pos, df_sizes, fabric_parts=None):
        self._pos = df_pos
        self._sizes = df_sizes
        self._parts = fabric_parts or {}

    def list_pos(self, companies=None):
        return self._pos

    def load_size_rows(self, selected):
        return self._sizes[self._sizes["PO Number"].isin(selected)]

    def load_fabric_parts_for_styles(self, styles, source=None):
        return {s: self._parts.get(s, []) for s in styles}


def _pos_df(**overrides):
    base = {
        "po_number": ["PO1", "PO2"],
        "style": ["ST1", "ST1"],
        "seller": ["FACTORY CO", "FACTORY CO"],
        "fabric": ["", ""],
        "description_code": ["DESC", "DESC"],
        "style_description": ["", ""],
        "cpo": ["CPO1", np.nan],
        "destination_code": ["UC", np.nan],
        "customer": ["ROSS STORES", "ROSS STORES"],
        "buyer": ["", ""],
        "factory_ship_date": ["7/30/2026", np.nan],
        "xport_date": ["", "8/15/2026"],
        "ship_to": ["ROSS DC, CARLISLE PA", np.nan],
        "packaging": ["PPK", np.nan],
        "hanger": ["HANGER (1-2-2-1)", np.nan],
    }
    base.update(overrides)
    return pd.DataFrame(base)


def _sizes_df():
    return pd.DataFrame({
        "PO Number": ["PO1", "PO1", "PO2"],
        "Style": ["ST1", "ST1", "ST1"],
        "Color": ["JET BLACK", "JET BLACK", "PINE"],
        "Size": ["S", "M", "S"],
        "Units": [100, 200, 50],
    })


def _sizes_df_upc():
    return pd.DataFrame({
        "PO Number": ["PO1", "PO1", "PO2"],
        "Style": ["ST1", "ST1", "ST1"],
        "Color": ["JET BLACK", "JET BLACK", "PINE"],
        "Size": ["S", "M", "S"],
        "Units": [100, 200, 50],
        "UPC": ["700948471565", "700948471534", "700948471507"],
    })


def _sheet(data: bytes, name="ST1"):
    return openpyxl.load_workbook(io.BytesIO(data))[name]


def _upc_headers(ws):
    return [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]


def _upc_ttl_row(ws):
    for r in range(3, ws.max_row + 1):
        if ws.cell(r, 1).value == "TTL":
            return r
    return ws.max_row


def _all_values(ws):
    return [c.value for row in ws.iter_rows() for c in row if c.value is not None]


# ── fabric block ──────────────────────────────────────────────────────────────

def test_fabric_rows_from_style_fabric_table():
    parts = [FabricPart(seq=1, body_part="大身", hhn_no="HHN-DB-1",
                        composition="86%Polyester 14%Spandex",
                        weight_gsm=200, width_cm=170),
             FabricPart(seq=2, body_part="口袋布", hhn_no="HHN-CJS-2",
                        composition="100%Cotton")]
    store = _Store(_pos_df(), _sizes_df(), {"ST1": parts})
    ws = _sheet(generate_giii_production_plan(["PO1", "PO2"], store, {}))

    assert ws.cell(5, 1).value == "面料/FIBER:"
    assert "HHN-DB-1" in ws.cell(5, 2).value
    assert "86%Polyester 14%Spandex" in ws.cell(5, 2).value
    assert "200gsm" in ws.cell(5, 2).value and "有效170cm" in ws.cell(5, 2).value
    assert ws.cell(6, 1).value == "面料_其他1:"
    assert "HHN-CJS-2" in ws.cell(6, 2).value
    # one extra fabric row shifts everything below down by one
    assert ws.cell(7, 1).value == "品名/Description:"
    assert ws.cell(9, 1).value == "合同号"


def test_no_fabric_parts_keeps_classic_layout():
    store = _Store(_pos_df(), _sizes_df())
    ws = _sheet(generate_giii_production_plan(["PO1", "PO2"], store, {}))
    assert ws.cell(5, 1).value == "面料/FIBER:"
    assert ws.cell(6, 1).value == "品名/Description:"
    assert ws.cell(8, 1).value == "合同号"


# ── CPRS requirements ─────────────────────────────────────────────────────────

class _Req:
    def __init__(self, warehouse="", red_sticker="", carton_mark="",
                 red_img=None, mark_img=None, prepack_ratio="", pcs_box="",
                 msrp="", rfid="", carton_weight=""):
        self.warehouse = warehouse
        self.red_sticker = red_sticker
        self.carton_mark = carton_mark
        self.red_img = red_img
        self.mark_img = mark_img
        self.prepack_ratio = prepack_ratio
        self.pcs_box = pcs_box
        self.msrp = msrp
        self.rfid = rfid
        self.carton_weight = carton_weight


def _png_bytes() -> bytes:
    from PIL import Image
    img = Image.new("RGB", (40, 20), "red")
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


def test_requirements_fill_red_sticker_and_carton_mark():
    reqs = {"PO1": _Req(red_sticker="MY", carton_mark="见箱唛要求"),
            "PO2": _Req(warehouse="DN", red_sticker="无需")}
    store = _Store(_pos_df(), _sizes_df())
    ws = _sheet(generate_giii_production_plan(["PO1", "PO2"], store, {},
                                              requirements=reqs))
    vals = _all_values(ws)
    assert "MY" in vals and "无需" in vals and "见箱唛要求" in vals
    # PO2 had no destination_code — CPRS-resolved warehouse fills 仓库代码
    assert "DN" in vals


def test_requirement_artwork_embeds_thumbnails_and_full_size_block():
    png = _png_bytes()
    reqs = {"PO1": _Req(red_sticker="MY", red_img=png, mark_img=_png_bytes()),
            "PO2": _Req(red_sticker="MY", red_img=png)}   # same red artwork
    store = _Store(_pos_df(), _sizes_df())
    data = generate_giii_production_plan(["PO1", "PO2"], store, {},
                                         requirements=reqs)
    ws = _sheet(data)
    # thumbnails: PO1 red+mark, PO2 red = 3; full-size: red (deduped) + mark = 2
    assert len(ws._images) == 5
    vals = [str(v) for v in _all_values(ws)]
    # deduped label lists both POs sharing the artwork
    assert any(v.startswith("红色箱贴纸图示") and "PO1" in v and "PO2" in v
               for v in vals)
    assert any(v.startswith("主箱唛图示") and "PO1" in v for v in vals)


def test_bad_image_bytes_do_not_break_export():
    reqs = {"PO1": _Req(red_sticker="MY", red_img=b"not-an-image")}
    store = _Store(_pos_df(), _sizes_df())
    ws = _sheet(generate_giii_production_plan(["PO1", "PO2"], store, {},
                                              requirements=reqs))
    assert "MY" in _all_values(ws)


def test_no_requirements_leaves_brand_dependent_cells_empty():
    """No CPRS resolution → 红色箱贴纸/主箱唛 stay EMPTY (never a claim like
    无), and brand-less POs are flagged ⚠ 无品牌 in 备注."""
    store = _Store(_pos_df(), _sizes_df())        # no division_name → no brand
    ws = _sheet(generate_giii_production_plan(["PO1", "PO2"], store, {}))
    vals = [str(v) for v in _all_values(ws)]
    assert not any(v in ("无", "无需") for v in vals)
    assert any(v == "⚠ 无品牌" for v in vals)


def test_branded_pos_are_not_flagged():
    store = _Store(_pos_df(division_name=["DW", "DW"]), _sizes_df())
    data = generate_giii_production_plan(["PO1", "PO2"], store, {})
    ws = _sheet(data)
    assert not any(str(v) == "⚠ 无品牌" for v in _all_values(ws))
    wb = openpyxl.load_workbook(io.BytesIO(data))
    sm = wb["Summary 汇总"]
    assert sm.cell(2, 3).value == "品牌"
    assert sm.cell(3, 3).value == "DW"


def test_brand_derived_from_po_prefix_when_division_missing():
    """CSKHHN… POs carry the CK division code in the PO number itself —
    decoded (documented prefix), so no flag and the 品牌 column fills."""
    df = _pos_df(po_number=["CSKHHN015R", "CSKHHN016R"])
    sizes = _sizes_df().assign(**{"PO Number": ["CSKHHN015R", "CSKHHN015R",
                                                "CSKHHN016R"]})
    store = _Store(df, sizes)
    data = generate_giii_production_plan(["CSKHHN015R", "CSKHHN016R"], store, {})
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert wb["Summary 汇总"].cell(3, 3).value == "Calvin Klein"
    vals = [str(v) for v in _all_values(wb["ST1"])]
    assert not any(v == "⚠ 无品牌" for v in vals)


def test_brandless_pos_flagged_in_summary():
    store = _Store(_pos_df(), _sizes_df())
    wb = openpyxl.load_workbook(io.BytesIO(
        generate_giii_production_plan(["PO1", "PO2"], store, {})))
    assert "⚠ 无品牌" in str(wb["Summary 汇总"].cell(3, 3).value)


# ── style-sheet 目的地 / 包装方式 / 离厂时间 ──────────────────────────────────

def test_style_sheet_has_destination_and_packing_columns():
    store = _Store(_pos_df(), _sizes_df())
    ws = _sheet(generate_giii_production_plan(["PO1", "PO2"], store, {}))
    hdr_row = 8    # no fabric parts → classic layout
    headers = [ws.cell(hdr_row, c).value for c in range(1, 30)]
    assert "目的地" in headers and "包装方式" in headers and "备注" in headers
    vals = [str(v) for v in _all_values(ws)]
    assert any("CARLISLE" in v for v in vals)     # ship-to in 目的地
    assert any(v == "PPK" for v in vals)          # packaging in 包装方式


def test_ex_factory_date_is_etd_minus_10_days():
    store = _Store(_pos_df(), _sizes_df())
    data = generate_giii_production_plan(["PO1", "PO2"], store, {})
    ws = _sheet(data)
    vals = [str(v) for v in _all_values(ws)]
    # PO1 ETD 7/30/2026 → 07/20/2026; PO2 ETD 8/15/2026 → 08/05/2026
    assert any("07/20/2026" in v for v in vals)
    assert any("08/05/2026" in v for v in vals)
    assert not any("7/30/2026" in v for v in vals)
    # summary inherits the adjusted date
    wb = openpyxl.load_workbook(io.BytesIO(data))
    svals = [str(v) for v in _all_values(wb["Summary 汇总"])]
    assert any("07/20/2026" in v for v in svals)


def test_unparseable_etd_passes_through():
    store = _Store(_pos_df(factory_ship_date=["TBD", np.nan]), _sizes_df())
    ws = _sheet(generate_giii_production_plan(["PO1", "PO2"], store, {}))
    assert any(str(v) == "TBD" for v in _all_values(ws))


# ── 合同号 / CN colours / NaN cleanup ─────────────────────────────────────────

def test_contract_from_progress_maps():
    store = _Store(_pos_df(), _sizes_df())
    ws = _sheet(generate_giii_production_plan(
        ["PO1", "PO2"], store, {},
        contract_by_po={_norm_key("PO1"): "26302-ZA1"},
        contract_by_style={_norm_key("ST1"): "26302-ZA9"}))
    vals = _all_values(ws)
    assert "26302-ZA1" in vals          # by-PO hit
    assert "26302-ZA9" in vals          # PO2 falls back to by-style


def test_cn_color_from_progress_lookup():
    store = _Store(_pos_df(), _sizes_df())
    data = generate_giii_production_plan(
        ["PO1", "PO2"], store, {},
        color_lookup_en={"JET BLACK": "煤黑色"})
    assert "煤黑色" in _all_values(_sheet(data))
    # CN colour reaches BOTH summary sheets too
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert "煤黑色" in [str(v) for v in _all_values(wb["Summary 汇总"])]
    sp = wb["简明汇总"]
    assert any(sp.cell(r, 4).value == "煤黑色" for r in range(3, sp.max_row + 1))


def test_style_sheet_packing_broken_into_columns():
    """包装方式 splits into 包装方式/衣架/是否预包/每箱件数/MSRP/RFID columns."""
    reqs = {"PO1": _Req(prepack_ratio="1-2-2-1", pcs_box="36", msrp="Y", rfid="N",
                        carton_weight="40 lbs / 18 kg per carton")}
    store = _Store(_pos_df(), _sizes_df())
    ws = _sheet(generate_giii_production_plan(["PO1", "PO2"], store, {},
                                              requirements=reqs))
    headers = [ws.cell(8, c).value for c in range(1, 27)]
    for h in ("品牌", "包装方式", "衣架", "是否预包", "每箱件数", "箱重限制",
              "MSRP", "RFID", "备注"):
        assert h in headers, f"missing style-sheet column {h}"
    vals = [str(v) for v in _all_values(ws)]
    assert any(v == "PPK" for v in vals)                  # packing only
    assert any(v == "HANGER" for v in vals)               # hanger, ratio stripped
    assert any(v == "Y 1-2-2-1" for v in vals)            # prepack + ratio
    assert "36" in vals and "Y" in vals and "N" in vals   # pcs / MSRP / RFID
    assert "40 lbs / 18 kg per carton" in vals            # 箱重限制


def test_msrp_column_shows_actual_price_over_flag():
    """A PO that prints an MSRP shows the price ($59.00) in the MSRP column;
    CPRS's Y/N 'MSRP required' flag is only the fallback for POs without one."""
    df = _pos_df(msrp=["59.00", np.nan])
    reqs = {"PO1": _Req(msrp="N"), "PO2": _Req(msrp="Y")}
    store = _Store(df, _sizes_df())
    data = generate_giii_production_plan(["PO1", "PO2"], store, {},
                                         requirements=reqs)
    ws = _sheet(data)
    headers = [ws.cell(8, c).value for c in range(1, 30)]
    mi = headers.index("MSRP") + 1
    col = [ws.cell(r, mi).value for r in range(9, ws.max_row + 1)]
    col = [str(v) for v in col if v is not None]
    assert "$59.00" in col     # PO1 price wins over its 'N' flag
    assert "Y" in col          # PO2 has no price → falls back to the flag
    assert "N" not in col      # PO1's flag was overridden, not shown
    # the actual price reaches the Summary 汇总 too (joined per style with 、)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert any("$59.00" in str(v) for v in _all_values(wb["Summary 汇总"]))


def _col_idx(ws, name, hdr_row=8):
    for c in range(1, 30):
        if ws.cell(hdr_row, c).value == name:
            return c
    raise AssertionError(f"column {name!r} not found")


def _data_cells(ws, col, start=9):
    return [ws.cell(r, col) for r in range(start, ws.max_row + 1)]


def test_empty_requirement_cells_get_explanatory_comments():
    # branded POs but no CPRS resolution → requirement cells blank AND commented
    store = _Store(_pos_df(division_name=["DW", "DW"]), _sizes_df())
    ws = _sheet(generate_giii_production_plan(["PO1", "PO2"], store, {}))
    ri = _col_idx(ws, "RFID")
    commented = [ws.cell(r, ri) for r in range(9, ws.max_row + 1)
                 if ws.cell(r, ri).comment is not None]
    assert len(commented) == 2                    # one blank-cell note per PO
    assert all(not c.value for c in commented)
    assert all("CPRS" in c.comment.text for c in commented)


def test_msrp_required_without_price_gets_comment():
    reqs = {"PO1": _Req(msrp="Y"), "PO2": _Req(msrp="N")}
    store = _Store(_pos_df(division_name=["DW", "DW"]), _sizes_df())
    ws = _sheet(generate_giii_production_plan(["PO1", "PO2"], store, {},
                                             requirements=reqs))
    mi = _col_idx(ws, "MSRP")
    y = [c for c in _data_cells(ws, mi) if c.value == "Y"]
    assert y and y[0].comment is not None and "MSRP" in y[0].comment.text
    n = [c for c in _data_cells(ws, mi) if c.value == "N"]
    assert n and n[0].comment is None                       # N = not required → no note


def test_actual_msrp_price_has_no_comment():
    df = _pos_df(division_name=["DW", "DW"], msrp=["59.00", "59.00"])
    reqs = {"PO1": _Req(msrp="Y"), "PO2": _Req(msrp="Y")}
    ws = _sheet(generate_giii_production_plan(
        ["PO1", "PO2"], _Store(df, _sizes_df()), {}, requirements=reqs))
    mi = _col_idx(ws, "MSRP")
    priced = [c for c in _data_cells(ws, mi) if c.value == "$59.00"]
    assert priced and all(c.comment is None for c in priced)  # price shown → no note


def test_populated_requirement_cell_has_no_comment():
    reqs = {"PO1": _Req(pcs_box="36"), "PO2": _Req(pcs_box="36")}
    store = _Store(_pos_df(division_name=["DW", "DW"]), _sizes_df())
    ws = _sheet(generate_giii_production_plan(["PO1", "PO2"], store, {},
                                             requirements=reqs))
    pi = _col_idx(ws, "每箱件数")
    cells = [c for c in _data_cells(ws, pi) if c.value == "36"]
    assert cells and all(c.comment is None for c in cells)


def test_prepack_falls_back_to_cprs_ratio_when_po_silent():
    df = _pos_df(division_name=["DW", "DW"],
                 packaging=[np.nan, np.nan], hanger=[np.nan, np.nan])
    reqs = {"PO1": _Req(prepack_ratio="1-2-2-1"),
            "PO2": _Req(prepack_ratio="1-2-2-1")}
    ws = _sheet(generate_giii_production_plan(
        ["PO1", "PO2"], _Store(df, _sizes_df()), {}, requirements=reqs))
    ci = _col_idx(ws, "是否预包")
    vals = [str(c.value) for c in _data_cells(ws, ci) if c.value]
    assert any(v.startswith("Y 1-2-2-1") for v in vals)      # CPRS ratio → prepack Y


def test_fmt_msrp_formats_only_bare_numbers():
    from po_extractor.exporters.giii_production_plan import _fmt_msrp
    assert _fmt_msrp("59.00") == "$59.00"
    assert _fmt_msrp("59") == "$59"
    assert _fmt_msrp("$59.00") == "$59.00"     # already a price — unchanged
    assert _fmt_msrp("N") == "N"               # Y/N flag passes through
    assert _fmt_msrp("") == "" and _fmt_msrp(None) == "" and _fmt_msrp(np.nan) == ""


def test_dest_country_from_address_markers():
    from po_extractor.exporters.giii_production_plan import _dest_country
    assert _dest_country("ROSS STORES / 3404 INDIAN AVE / PERRIS,CA 92571") == "US"
    assert _dest_country("ROSS DC, CARLISLE PA") == "US"
    assert _dest_country("NEWTON 4-5, ALMELO NL 7609RR") == "EU"
    assert _dest_country("G-III c/o BLECKMANN, NETHERLANDS") == "EU"
    assert _dest_country("TJX AUSTRALIA PROCESSING CENTRE") == "AU"
    assert _dest_country("TORONTO, ONTARIO, CANADA") == "CA"
    assert _dest_country("BTB DIRECT SHIPMENT") == ""   # no marker → honest blank
    assert _dest_country("") == ""
    # word-like state codes never match as bare words — only after a comma
    assert _dest_country("GOODS ON HOLD IN TRANSIT") == ""
    assert _dest_country("WAREHOUSE, INDIANAPOLIS, IN") == "US"


def test_dest_address_keeps_segment_with_extra_info():
    from po_extractor.exporters.giii_production_plan import _dest_address
    # exact buyer name (or an abbreviation of it) → dropped
    assert _dest_address("ROSS STORES / 3404 INDIAN AVE / PERRIS,CA",
                         "ROSS STORES") == "3404 INDIAN AVE / PERRIS,CA"
    assert _dest_address("ROSS / 3404 INDIAN AVE", "ROSS STORES") == "3404 INDIAN AVE"
    # segment carrying MORE than the buyer name (the DC designation) → kept
    assert _dest_address("ROSS STORES DC#4 / 123 MAIN RD", "ROSS STORES") == \
        "ROSS STORES DC#4 / 123 MAIN RD"


def test_style_sheet_and_summary_have_dest_country():
    store = _Store(_pos_df(), _sizes_df())
    data = generate_giii_production_plan(["PO1", "PO2"], store, {})
    ws = _sheet(data)
    headers = [ws.cell(8, c).value for c in range(1, 28)]
    assert "目的地国家" in headers
    vals = [str(v) for v in _all_values(ws)]
    assert "US" in vals                        # from "ROSS DC, CARLISLE PA"
    wb = openpyxl.load_workbook(io.BytesIO(data))
    sm = wb["Summary 汇总"]
    sm_headers = [sm.cell(2, c).value for c in range(1, 25)]
    assert "目的地国家" in sm_headers
    ci = sm_headers.index("目的地国家") + 1
    assert sm.cell(3, ci).value == "US"


def test_destination_strips_duplicated_buyer_name():
    """'ROSS STORES / 3404 INDIAN AVE / PERRIS,CA' → 目的地 shows only the
    address; the consignee name lives in the 买家 column."""
    df = _pos_df(ship_to=["ROSS STORES / 3404 INDIAN AVE / DISTRIBUTION "
                          "CENTER / PERRIS,CA 92571", np.nan])
    store = _Store(df, _sizes_df())
    ws = _sheet(generate_giii_production_plan(["PO1", "PO2"], store, {}))
    vals = [str(v) for v in _all_values(ws)]
    assert any(v.startswith("3404 INDIAN AVE") for v in vals)
    assert not any("ROSS STORES / 3404" in v for v in vals)
    assert "ROSS STORES" in vals           # 买家 column keeps the buyer


def test_ratio_in_packing_text_means_prepack_without_ppk_marker():
    """Ross fax POs print 'FLAT PACK + HANGER (1-2-2-1)' with no PPK marker —
    the ratio itself means prepack, and it moves out of 衣架 into 是否预包."""
    df = _pos_df(packaging=["FLAT PACK", "FLAT PACK"],
                 hanger=["HANGER (1-2-2-1)", "HANGER (2-2-2)"])
    store = _Store(df, _sizes_df())
    ws = _sheet(generate_giii_production_plan(["PO1", "PO2"], store, {}))
    vals = [str(v) for v in _all_values(ws)]
    assert any(v == "Y 1-2-2-1" for v in vals)
    assert any(v == "Y 2-2-2" for v in vals)
    assert any(v == "HANGER" for v in vals)
    assert not any("(1-2-2-1)" in v for v in vals)


def test_style_sheet_brand_column_with_flag():
    df = _pos_df(po_number=["CSKHHN015R", "PO2"])
    sizes = _sizes_df().assign(**{"PO Number": ["CSKHHN015R", "CSKHHN015R", "PO2"]})
    store = _Store(df, sizes)
    ws = _sheet(generate_giii_production_plan(["CSKHHN015R", "PO2"], store, {}))
    headers = [ws.cell(8, c).value for c in range(1, 6)]
    assert headers[:4] == ["合同号", "款号", "品牌", "PO号"]
    vals = [str(v) for v in _all_values(ws)]
    assert "Calvin Klein" in vals          # decoded from the CS prefix
    assert "⚠ 无品牌" in vals              # PO2 has no brand → flagged in 品牌


def test_nan_metadata_never_renders_as_nan():
    store = _Store(_pos_df(), _sizes_df())
    ws = _sheet(generate_giii_production_plan(["PO1", "PO2"], store, {}))
    for v in _all_values(ws):
        assert "nan" not in str(v).lower().split(), f"NaN leaked into cell: {v!r}"


# ── Summary 汇总 sheet ────────────────────────────────────────────────────────

def test_summary_sheet_first_with_one_row_per_style_and_ttl():
    df_pos = _pos_df(style=["ST1", "ST2"])
    sizes = pd.DataFrame({
        "PO Number": ["PO1", "PO1", "PO2"],
        "Style": ["ST1", "ST1", "ST2"],
        "Color": ["JET BLACK", "JET BLACK", "PINE"],
        "Size": ["S", "M", "S"],
        "Units": [100, 200, 50],
    })
    store = _Store(df_pos, sizes)
    wb = openpyxl.load_workbook(io.BytesIO(
        generate_giii_production_plan(["PO1", "PO2"], store, {})))

    assert wb.sheetnames[0] == "Summary 汇总"
    assert "ST1" in wb.sheetnames and "ST2" in wb.sheetnames

    ws = wb["Summary 汇总"]
    assert ws.cell(2, 1).value == "No." and ws.cell(2, 2).value == "款号"
    # row 3 = ST1 (hyperlink formula to its sheet), row 4 = ST2, row 5 = TTL
    assert ws.cell(3, 1).value == 1
    assert "ST1" in str(ws.cell(3, 2).value) and "HYPERLINK" in str(ws.cell(3, 2).value)
    assert ws.cell(3, 11).value == 300         # ST1 total (总数量)
    assert "PO1" in str(ws.cell(3, 7).value)
    assert "JET BLACK" in str(ws.cell(3, 8).value)
    assert ws.cell(4, 1).value == 2
    assert ws.cell(4, 11).value == 50          # ST2 total
    assert ws.cell(5, 1).value == "TTL"
    assert ws.cell(5, 11).value == 350         # grand total


def test_summary_has_size_breakdown_packing_destination():
    store = _Store(_pos_df(), _sizes_df())
    wb = openpyxl.load_workbook(io.BytesIO(
        generate_giii_production_plan(["PO1", "PO2"], store, {})))
    ws = wb["Summary 汇总"]
    headers = [ws.cell(2, c).value for c in range(1, 24)]
    for h in ("尺码明细", "包装方式", "目的地", "衣架", "是否预包",
              "每箱件数", "MSRP", "RFID", "颜色(中文)"):
        assert h in headers, f"missing summary column {h}"
    vals = [str(v) for v in _all_values(ws)]
    assert any("S 150" in v and "M 200" in v for v in vals)   # size breakdown w/ qty
    assert any("PPK" in v for v in vals)                      # packing
    assert any(v == "HANGER" for v in vals)                   # hanger (ratio stripped)
    assert any(v == "Y 1-2-2-1" for v in vals)                # prepack + PO ratio
    assert any("CARLISLE" in v for v in vals)                 # destination


def test_simple_summary_sheet_style_color_fabric_sizes():
    parts = [FabricPart(seq=1, body_part="大身", hhn_no="HHN-DB-1",
                        composition="86%Polyester 14%Spandex")]
    store = _Store(_pos_df(), _sizes_df(), {"ST1": parts})
    wb = openpyxl.load_workbook(io.BytesIO(
        generate_giii_production_plan(["PO1", "PO2"], store, {})))

    assert wb.sheetnames[1] == "简明汇总"
    ws = wb["简明汇总"]
    headers = [ws.cell(2, c).value for c in range(1, 9)]
    assert headers == ["No.", "款号", "颜色", "颜色(中文)", "面料", "S", "M", "总数量"]
    # row 3: JET BLACK  S=100 M=200 total=300; row 4: PINE S=50 total=50
    assert ws.cell(3, 1).value == 1
    assert ws.cell(3, 2).value == "ST1"
    assert "HHN-DB-1" in ws.cell(3, 5).value
    assert ws.cell(3, 3).value == "JET BLACK"
    assert ws.cell(3, 6).value == 100 and ws.cell(3, 7).value == 200
    assert ws.cell(3, 8).value == 300
    assert ws.cell(4, 3).value == "PINE" and ws.cell(4, 8).value == 50
    # TTL row: per-size sums + grand total
    assert ws.cell(5, 1).value == "TTL"
    assert ws.cell(5, 6).value == 150 and ws.cell(5, 7).value == 200
    assert ws.cell(5, 8).value == 350


def test_consumption_on_summary_not_style_sheet():
    """单耗/排版 appear on Summary 汇总 (per style), not on the style sheet;
    kg is derived from cm via the gross width (150+5)."""
    store = _Store(_pos_df(), _sizes_df())
    data = generate_giii_production_plan(
        ["PO1", "PO2"], store, {},
        consumption={"ST1": {"cons_kg": None, "cons_cm": 165, "util": 82,
                             "marker_pcs": 24, "width_cm": 150, "gsm": 200}})
    wb = openpyxl.load_workbook(io.BytesIO(data))

    sm = wb["Summary 汇总"]
    hdr = [sm.cell(2, c).value for c in range(1, sm.max_column + 1)]
    for h in ("单耗(kg)", "单耗(cm)", "排版利用率", "排版件数",
              "排版有效门幅(cm)", "排版面料克重(g/m²)"):
        assert h in hdr, f"summary missing consumption column {h}"
    ci = {h: i + 1 for i, h in enumerate(hdr)}
    assert sm.cell(3, ci["单耗(kg)"]).value == 0.5115   # 165*155*200/1e7
    assert sm.cell(3, ci["单耗(cm)"]).value == 165
    assert sm.cell(3, ci["排版面料克重(g/m²)"]).value == 200

    # style sheet must NOT carry the consumption columns
    ws = wb["ST1"]
    style_vals = {str(c.value) for row in ws.iter_rows() for c in row if c.value}
    assert "单耗(kg)" not in style_vals and "排版利用率" not in style_vals


def test_summary_carries_requirement_texts():
    reqs = {"PO1": _Req(red_sticker="MY", carton_mark="见箱唛要求"),
            "PO2": _Req(red_sticker="无需")}
    store = _Store(_pos_df(), _sizes_df())
    wb = openpyxl.load_workbook(io.BytesIO(
        generate_giii_production_plan(["PO1", "PO2"], store, {},
                                      requirements=reqs)))
    ws = wb["Summary 汇总"]
    vals = _all_values(ws)
    assert any("MY" in str(v) for v in vals)
    assert any("无需" in str(v) for v in vals)
    assert any("见箱唛要求" in str(v) for v in vals)


# ── UPC 汇总 sheet ─────────────────────────────────────────────────────────────

def test_upc_summary_two_lines_units_then_upc_under_each_size():
    store = _Store(_pos_df(), _sizes_df_upc())
    data = generate_giii_production_plan(["PO1", "PO2"], store, {})
    wb = openpyxl.load_workbook(io.BytesIO(data))

    assert wb.sheetnames[2] == "UPC 汇总"        # after Summary 汇总 + 简明汇总
    ws = wb["UPC 汇总"]
    headers = _upc_headers(ws)
    assert headers[:7] == ["No.", "款号", "品牌", "合同号", "PO号",
                           "颜色(英文)", "颜色(中文)"]
    assert "项目" in headers                       # 数量 / UPC line label
    assert "S" in headers and "M" in headers      # sizes are their own columns
    assert headers[-1] == "总数量"
    ii = headers.index("项目") + 1
    si, mi = headers.index("S") + 1, headers.index("M") + 1
    ti = headers.index("总数量") + 1

    # PO1 JET BLACK block: row 3 = 数量, row 4 = UPC (two lines, aligned by size)
    assert ws.cell(3, 5).value == "PO1" and ws.cell(3, 6).value == "JET BLACK"
    assert ws.cell(3, ii).value == "数量"
    assert ws.cell(3, si).value == 100 and ws.cell(3, mi).value == 200
    assert ws.cell(4, ii).value == "UPC"
    assert ws.cell(4, si).value == "700948471565"
    assert ws.cell(4, mi).value == "700948471534"
    assert isinstance(ws.cell(4, si).value, str)   # UPC kept as text, not float
    assert ws.cell(3, ti).value == 300             # 总数量 merged over the block

    # PO2 PINE block: rows 5 (数量) / 6 (UPC); only S carries a UPC
    assert ws.cell(5, ii).value == "数量" and ws.cell(5, si).value == 50
    assert ws.cell(6, ii).value == "UPC" and ws.cell(6, si).value == "700948471507"
    assert not ws.cell(6, mi).value                # M blank for PINE

    ttl_r = _upc_ttl_row(ws)
    assert ttl_r == 7                              # 2 blocks × 2 rows + header offset
    assert ws.cell(ttl_r, ti).value == 350


def test_upc_detail_sheet_flat_per_size_list():
    store = _Store(_pos_df(), _sizes_df_upc())
    data = generate_giii_production_plan(["PO1", "PO2"], store, {})
    wb = openpyxl.load_workbook(io.BytesIO(data))
    # 4th sheet, after Summary 汇总 / 简明汇总 / UPC 汇总
    assert wb.sheetnames[:4] == ["Summary 汇总", "简明汇总", "UPC 汇总", "UPC 明细"]
    ws = wb["UPC 明细"]
    headers = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
    assert headers == ["No.", "款号", "品牌", "合同号", "PO号",
                       "颜色(英文)", "颜色(中文)", "尺码", "UPC", "数量"]
    # one row per (PO, colour, size): PO1 JET BLACK S/M + PO2 PINE S = 3
    rows = []
    r = 3
    while ws.cell(r, 1).value not in (None, "TTL"):
        rows.append((ws.cell(r, 8).value, ws.cell(r, 9).value, ws.cell(r, 10).value))
        r += 1
    assert len(rows) == 3
    assert ("S", "700948471565", 100) in rows       # size / UPC / units
    assert isinstance(ws.cell(3, 9).value, str)      # UPC kept as text
    assert ws.cell(r, 10).value == 350               # TTL sums units


def test_upc_summary_carries_style_and_po_context():
    store = _Store(_pos_df(), _sizes_df_upc())
    data = generate_giii_production_plan(
        ["PO1", "PO2"], store, {},
        contract_by_po={_norm_key("PO1"): "26302-ZA1"})
    ws = openpyxl.load_workbook(io.BytesIO(data))["UPC 汇总"]
    vals = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
    assert "ST1" in vals and "PO1" in vals          # 款号 / PO号
    assert "JET BLACK" in vals and "PINE" in vals    # colours
    assert "26302-ZA1" in vals                       # 合同号 threaded through


def test_upc_summary_blank_when_po_has_no_upc():
    """No UPC captured for a size → the UPC line stays blank; never fabricated.
    The 数量 line still carries the units."""
    store = _Store(_pos_df(), _sizes_df())            # _sizes_df has no UPC column
    data = generate_giii_production_plan(["PO1", "PO2"], store, {})
    ws = openpyxl.load_workbook(io.BytesIO(data))["UPC 汇总"]
    headers = _upc_headers(ws)
    ii = headers.index("项目") + 1
    ti = headers.index("总数量") + 1
    size_cols = [i for i, h in enumerate(headers, start=1) if h in ("S", "M")]
    ttl_r = _upc_ttl_row(ws)
    assert ttl_r > 3
    units_seen = False
    for r in range(3, ttl_r):
        item = ws.cell(r, ii).value
        if item == "UPC":
            for c in size_cols:
                assert not ws.cell(r, c).value        # every UPC cell blank
        elif item == "数量":
            units_seen = units_seen or any(ws.cell(r, c).value for c in size_cols)
    assert units_seen                                 # 数量 lines still populated
    assert ws.cell(ttl_r, ti).value == 350            # quantities still total


# ── export_buyplan zero-row filter ────────────────────────────────────────────

def test_export_buyplan_drops_all_zero_color_rows(tmp_path):
    from po_extractor.exporters.buyplan_export import export_buyplan

    data = pd.DataFrame({
        "PO Number": ["PO1"] * 4,
        "Style": ["ST1"] * 4,
        "Color": ["JVS/JET BLACK", "JVS/JET BLACK", "GHOST/UNORDERED", "GHOST/UNORDERED"],
        "Size": ["S", "M", "S", "M"],
        "Units": [100, 200, 0, 0],
    })
    path = export_buyplan(data, pd.DataFrame(), str(tmp_path))
    wb = openpyxl.load_workbook(path)
    vals = [str(c.value) for ws in wb.worksheets
            for row in ws.iter_rows() for c in row if c.value is not None]
    assert any("JVS" in v for v in vals)
    assert not any("GHOST" in v for v in vals), "all-zero colour row leaked"
