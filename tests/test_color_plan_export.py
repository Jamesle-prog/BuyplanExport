"""Tests for the color-plan exporter (Color x Size pivot per style)."""
from __future__ import annotations

import pandas as pd
from openpyxl import load_workbook

from po_extractor.exporters.color_plan_export import export_color_plan


def test_color_plan_cn_map_matches_numeric_color_keys(tmp_path):
    """df_size['Color'] may be a numeric dtype (e.g. all-digit color codes).
    cn_map used to be built with the native (numeric) dtype as the key while
    the lookup always coerced with str(color) — so a numeric-keyed map never
    matched and Color (CN) rendered blank for every row."""
    df_size = pd.DataFrame({
        "Style": ["ST1", "ST1"],
        "Color": [101, 202],          # numeric dtype, not strings
        "Color (CN)": ["红色", "蓝色"],
        "Size": ["S", "S"],
        "Units": [10, 20],
    })
    path = export_color_plan(df_size, str(tmp_path))
    wb = load_workbook(path)
    ws = wb["ST1"]
    rows = {ws.cell(r, 1).value: ws.cell(r, 2).value for r in range(3, ws.max_row + 1)}
    assert rows.get(101) == "红色"
    assert rows.get(202) == "蓝色"


def test_color_plan_cn_map_still_works_with_string_colors(tmp_path):
    df_size = pd.DataFrame({
        "Style": ["ST1", "ST1"],
        "Color": ["BLACK", "WHITE"],
        "Color (CN)": ["黑色", "白色"],
        "Size": ["S", "S"],
        "Units": [5, 6],
    })
    path = export_color_plan(df_size, str(tmp_path))
    ws = load_workbook(path)["ST1"]
    rows = {ws.cell(r, 1).value: ws.cell(r, 2).value for r in range(3, ws.max_row + 1)}
    assert rows == {"BLACK": "黑色", "WHITE": "白色"}


def test_color_plan_without_cn_column_has_no_cn_column(tmp_path):
    df_size = pd.DataFrame({
        "Style": ["ST1"],
        "Color": ["BLACK"],
        "Size": ["S"],
        "Units": [5],
    })
    path = export_color_plan(df_size, str(tmp_path))
    ws = load_workbook(path)["ST1"]
    assert ws.cell(1, 1).value == "Color"
    assert ws.cell(1, 2).value == "Sizes"    # no Color (CN) column inserted
