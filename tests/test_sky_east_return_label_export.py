"""Sky East buy-plan export: Return Label column.

Per-row value (same granularity as size quantities) sourced from the
client PO's Return Label column via the parser -- "Yes" / "No" / "NA".
Written to a new column past the template's existing last column, with a
programmatically-added header when the template itself has no matching
header text.
"""
from __future__ import annotations

import pandas as pd
import pytest

from po_extractor.exporters.sky_east_buyplan_export import export_sky_east_buyplan
from po_extractor.exporters._sky_east_helpers import _detect_buyplan_layout
from po_extractor.lookups.progress_lookup import PCColorMatch


@pytest.fixture(autouse=True)
def _isolated_sky_east_store(tmp_path, monkeypatch):
    from po_extractor.store.sky_east_store import SkyEastStore

    store = SkyEastStore(str(tmp_path / "_isolated_sky_east.db"))
    monkeypatch.setattr("po_extractor.store.get_sky_east_store", lambda: store)
    return store


@pytest.fixture
def two_style_df():
    return pd.DataFrame([
        {
            "pc_no": "HHPPC048", "style": "DR5124", "brand": "Anna Field",
            "contract_no": "26302-ZA7148", "article_name": "LACE DRESS",
            "zalando_po": "PO001", "config_sku": "C1", "color_name": "dark blue",
            "fabric_item_no": "HHN-JA-01715", "fabrication": "cotton",
            "ex_fty_date": "2026-08-11", "return_label": "Yes",
            "xs": 30, "s": 82, "m": 119, "l": 102, "xl": 67, "xxl": 0,
        },
        {
            "pc_no": "HHPPC048", "style": "DR4578", "brand": "Anna Field",
            "contract_no": "26302-ZA7149", "article_name": "LONG SLEEVE DRESS",
            "zalando_po": "PO001", "config_sku": "C2", "color_name": "green",
            "fabric_item_no": "HHN-JSRSR-04068", "fabrication": "poly",
            "ex_fty_date": "2026-08-11", "return_label": "No",
            "xs": 22, "s": 61, "m": 90, "l": 77, "xl": 50, "xxl": 0,
        },
    ])


@pytest.fixture
def pc_color_lookup():
    return {
        ("HHPPC048", "DR5124", "Dark Blue"): PCColorMatch("藏青", "503", ""),
        ("HHPPC048", "DR4578", "Green"):     PCColorMatch("绿色", "602", ""),
    }


def _style_sheet_rows(wb, suffix):
    ws = next(s for s in wb.sheetnames[2:] if s.endswith(suffix))
    return wb[ws]


def _return_label_col_and_values(ws):
    """Locate the Return Label header (via the same layout-detection the
    exporter itself uses) and return (header_row, col_num, [values])."""
    col, data_row = _detect_buyplan_layout(ws)
    header_row = data_row - 1
    header_text = ws.cell(header_row, col["return_label"]).value
    values = [ws.cell(r, col["return_label"]).value
              for r in range(data_row, ws.max_row + 1)
              if ws.cell(r, col["return_label"]).value not in (None, "")]
    return header_text, values


def test_return_label_column_has_header_and_per_row_values(
    two_style_df, pc_color_lookup, tmp_path,
):
    path, _totals = export_sky_east_buyplan(
        two_style_df, cn_lookup={}, output_dir=str(tmp_path),
        label_lookup={}, cn_code_lookup={}, cn_by_pc_lookup=pc_color_lookup,
    )
    from openpyxl import load_workbook
    wb = load_workbook(path)

    ws1 = _style_sheet_rows(wb, "_DR5124")
    header_text, values = _return_label_col_and_values(ws1)
    assert str(header_text).strip().lower() == "return label"
    assert values == ["Yes"]

    ws2 = _style_sheet_rows(wb, "_DR4578")
    _, values2 = _return_label_col_and_values(ws2)
    assert values2 == ["No"]


def test_return_label_defaults_to_na_when_column_absent_from_df(
    two_style_df, pc_color_lookup, tmp_path,
):
    df = two_style_df.drop(columns=["return_label"])
    path, _totals = export_sky_east_buyplan(
        df, cn_lookup={}, output_dir=str(tmp_path),
        label_lookup={}, cn_code_lookup={}, cn_by_pc_lookup=pc_color_lookup,
    )
    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws1 = _style_sheet_rows(wb, "_DR5124")
    _, values = _return_label_col_and_values(ws1)
    assert values == ["NA"]
