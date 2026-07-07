"""Tests for the GIII/Sky East pipeline structural convergence.

Phase 1 — both pipelines go through the same front door: config format id,
universal file detection, and the parsers facade.
"""
from __future__ import annotations

import openpyxl
import pytest


# ── Phase 1: facade ──────────────────────────────────────────────────────────

def test_facade_exports_canonical_sky_east_parser():
    from po_extractor.parsers import parse_sky_east_order
    from po_extractor.parsers.sky_east_order import parse
    assert callable(parse_sky_east_order)
    # Facade delegates to the canonical (dynamic-layout) parser module.
    assert parse_sky_east_order.__module__ == "po_extractor.parsers"
    assert callable(parse)


def test_legacy_sky_east_parser_no_longer_exported():
    import po_extractor.parsers as parsers
    assert "parse_sky_east" not in parsers.__all__
    assert not hasattr(parsers, "parse_sky_east")


def test_format_sky_east_constant_exists():
    from po_extractor.config import FORMAT_SKY_EAST
    assert FORMAT_SKY_EAST == "sky_east"


# ── Phase 1: detection ───────────────────────────────────────────────────────

def _make_sky_east_xlsx(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contract"
    ws["A1"] = "SKY EAST INTERNATIONAL TRADING LIMITED"
    ws["A3"] = "PURCHASE CONTRACT"
    ws["A4"] = "PC NO."
    ws["B4"] = "HHPPC038"
    wb.save(path)


def _make_plain_xlsx(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "just some spreadsheet"
    wb.save(path)


def test_detect_file_recognises_sky_east_contract(tmp_path):
    from po_extractor.detectors import detect_file
    p = tmp_path / "se_contract.xlsx"
    _make_sky_east_xlsx(str(p))
    result = detect_file(str(p))
    assert result.format_id == "sky_east"
    assert result.file_type == "excel"
    assert result.confidence == "high"
    assert "Sky East" in result.companies


def test_detect_file_plain_excel_still_unknown(tmp_path):
    from po_extractor.detectors import detect_file
    p = tmp_path / "plain.xlsx"
    _make_plain_xlsx(str(p))
    result = detect_file(str(p))
    assert result.format_id == "excel_unknown"


def test_looks_like_sky_east_contract_readonly_mode(tmp_path):
    """The signature helper must work on read-only workbooks (detector mode)."""
    from po_extractor.parsers.sky_east_order import looks_like_sky_east_contract
    p = tmp_path / "se.xlsx"
    _make_sky_east_xlsx(str(p))
    wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
    try:
        assert looks_like_sky_east_contract(wb) is True
    finally:
        wb.close()

    p2 = tmp_path / "plain.xlsx"
    _make_plain_xlsx(str(p2))
    wb2 = openpyxl.load_workbook(str(p2), read_only=True, data_only=True)
    try:
        assert looks_like_sky_east_contract(wb2) is False
    finally:
        wb2.close()


def test_facade_parses_sky_east_end_to_end(tmp_path):
    """detect → facade parse, the same flow GIII PDFs get via parse_pdf."""
    from po_extractor.detectors import detect_file
    from po_extractor.parsers import parse_sky_east_order
    p = tmp_path / "se_contract.xlsx"
    _make_sky_east_xlsx(str(p))
    assert detect_file(str(p)).format_id == "sky_east"
    contract = parse_sky_east_order(str(p), processed_by="test")
    # Minimal fixture: header-only contract parses without raising and
    # carries traceability fields.
    assert contract.processed_by == "test"
    assert contract.source_file_hash


# ── Phase 2: parser-level quality grading, GIII semantics ────────────────────

def _make_item(**over):
    from po_extractor.models.sky_east_data import SkyEastItem
    base = dict(
        pc_no="PC1", zalando_po="PO1", style="ST1", config_sku="SKU-1",
        article_name="A", brand="Anna Field", color_name="Blue",
        colour_code="Q11", launch_date="", fabric_item_no="HHP-JS-12345",
        fabrication="", contract_no="", sizes={"S": 1}, total_qty=1,
        fob_usd=1.0, total_cost_usd=1.0,
    )
    base.update(over)
    return SkyEastItem(**base)


def _make_contract(items, **over):
    from po_extractor.models.sky_east_data import SkyEastContract
    base = dict(pc_no="PC1", pc_date="2026-01-01", buyer="B", seller="S",
                currency="USD", payment_terms="TT", trade_term="FOB")
    base.update(over)
    return SkyEastContract(items=items, **base)


def test_sky_east_contract_has_validation_status_field():
    c = _make_contract([_make_item()])
    assert hasattr(c, "validation_status")


def test_parsed_header_only_contract_grades_as_exception(tmp_path):
    """No line items → 'exception', matching the GIII parsers' convention."""
    from po_extractor.parsers import parse_sky_east_order
    p = tmp_path / "se_contract.xlsx"
    _make_sky_east_xlsx(str(p))
    contract = parse_sky_east_order(str(p))
    assert contract.items == []
    assert contract.validation_status == "exception"
    assert isinstance(contract.parse_confidence, int)


def test_backend_validation_reports_item_issues():
    from po_extractor.parsers.sky_east_validation import validate_contracts
    good = _make_item()
    bad = _make_item(style="", color_name="", sizes={"S": -5},
                     fabric_item_no="WRONGFORMAT", ex_fty_date="not-a-date",
                     config_sku="")
    report = validate_contracts([_make_contract([good, bad])])
    assert report["total_items"] == 2
    assert report["sku_covered"] == 1
    assert report["sku_coverage_pct"] == 50.0
    assert len(report["issues"]["missing_style_color"]) == 2   # style + color
    assert len(report["issues"]["negative_qty"]) == 1
    assert len(report["issues"]["bad_hhn"]) == 1
    assert len(report["issues"]["bad_ex_fty_date"]) == 1


def test_backend_validation_clean_contract_reports_no_issues():
    from po_extractor.parsers.sky_east_validation import validate_contracts
    report = validate_contracts([_make_contract([_make_item(ex_fty_date="2026-05-01")])])
    assert report["sku_coverage_pct"] == 100.0
    assert all(not v for v in report["issues"].values())


# ── Phase 3: shared exception queue accepts Sky East failures ────────────────

def test_exception_queue_stores_and_filters_sky_east_failures(tmp_path):
    from po_extractor.store.po_store import POStore
    store = POStore(str(tmp_path / "po.db"))
    store.save_exception(po_number="", file_name="bad_contract.xlsx",
                         company="Sky East",
                         reason="Sky East parse failed: boom",
                         processed_by="tester")
    store.save_exception(po_number="PO1", file_name="po1.pdf",
                         company="GIII", reason="giii failure")

    se_df = store.list_exceptions(companies=["Sky East"])
    assert len(se_df) == 1
    assert se_df.iloc[0]["file_name"] == "bad_contract.xlsx"
    assert se_df.iloc[0]["status"] == "pending"

    all_df = store.list_exceptions()
    assert set(all_df["company"]) == {"Sky East", "GIII"}


# ── Phase 5: shared read-side row shape ──────────────────────────────────────

def test_load_standard_orders_combines_both_stores(tmp_path):
    """One backend call returns both pipelines' orders in the standard shape."""
    from po_extractor.models.po_data import POData, POMetadata, SizeRow
    from po_extractor.store.po_store import POStore
    from po_extractor.store.sky_east_store import SkyEastStore
    from po_extractor.ui_helpers.combined_summary import (
        STANDARD_KEYS, load_standard_orders,
    )

    po_store = POStore(str(tmp_path / "po.db"))
    se_store = SkyEastStore(str(tmp_path / "po.db"))

    po_store.check_and_save(POData(
        metadata=POMetadata(po_number="PO777", style="ST7", company="GIII",
                            po_date="2026-01-05", xport_date="2026-04-01"),
        size_rows=[SizeRow("PO777", "ST7", "Black", "M", 40, "")],
    ))
    se_store.save_contract_checked(_make_contract(
        [_make_item(pc_no="PC9", zalando_po="PO888", total_qty=9)], pc_no="PC9"))

    df = load_standard_orders(po_store, se_store)
    assert list(df.columns) == STANDARD_KEYS
    assert set(df["po_number"]) == {"PO777", "PO888"}
    assert set(df["source"]) == {None, "sky_east"} or "sky_east" in set(df["source"])

    # Pipeline toggles mirror the access-control gating consumers need.
    only_giii = load_standard_orders(po_store, se_store, include_sky_east=False)
    assert set(only_giii["po_number"]) == {"PO777"}
    only_se = load_standard_orders(po_store, se_store, include_giii=False)
    assert set(only_se["po_number"]) == {"PO888"}


def test_sky_east_items_to_size_rows_matches_po_size_rows_grain():
    import pandas as pd
    from po_extractor.ui_helpers.combined_summary import (
        SIZE_ROW_KEYS, sky_east_items_to_size_rows,
    )
    items = pd.DataFrame([{
        "pc_no": "PC1", "zalando_po": "PO1", "style": "ST1",
        "color_name": "Blue", "xs": 0, "s": 10, "m": 20, "l": 0,
        "xl": 5, "xxl": 0, "total_qty": 35,
    }])
    rows = sky_east_items_to_size_rows(items)
    assert list(rows.columns) == SIZE_ROW_KEYS
    # Zero buckets dropped; one row per used size, GIII po_size_rows grain.
    assert len(rows) == 3
    assert set(rows["size"]) == {"S", "M", "XL"}
    assert rows["units"].sum() == 35
    assert (rows["po_number"] == "PO1").all()
    assert (rows["color"] == "Blue").all()


def test_sky_east_items_to_size_rows_empty_input():
    import pandas as pd
    from po_extractor.ui_helpers.combined_summary import (
        SIZE_ROW_KEYS, sky_east_items_to_size_rows,
    )
    out = sky_east_items_to_size_rows(pd.DataFrame())
    assert list(out.columns) == SIZE_ROW_KEYS
    assert out.empty


# ── Phase 4: one output-format catalogue for both pipelines ──────────────────

def test_sky_east_exporters_registered_alongside_giii():
    import po_extractor.exporters  # noqa: F401 — triggers registration
    from po_extractor.exporters.registry import all_formats, get

    ids = {f.format_id for f in all_formats()}
    # GIII formats (pre-existing)
    assert {"buy_plan", "color_plan", "po_summary", "cross_check"} <= ids
    # Sky East formats (phase 4)
    assert {"se_buy_plan", "se_nukuryou"} <= ids

    se_bp = get("se_buy_plan")
    assert se_bp.extension == ".xlsx"
    assert callable(se_bp.export_fn)
    assert get("se_nukuryou").export_fn is not None
