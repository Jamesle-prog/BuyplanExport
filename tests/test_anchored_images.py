"""Tests for reading photos that FLOAT over a cell rather than living in it.

A photo reaches a spreadsheet cell two different ways, and a single client PO
routinely contains both — one row's picture pasted as a WPS cell image (a
``=DISPIMG("ID_…")`` formula), the next row's inserted with Excel's own
Insert ▸ Picture, which anchors a drawing over the cell and leaves the cell
itself empty.

Only the DISPIMG form was ever read, so a style whose photo was inserted the
ordinary way arrived with no picture at all — and was then reported as having
no photo, with the picture sitting visibly in the file the user had uploaded.
"""
from __future__ import annotations

import io
import zipfile

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from PIL import Image

from po_extractor.utils.image_extractor import (
    ImageCache, extract_anchored_positions, extract_dispimg_positions,
    extract_images_from_xlsx,
)


def _png(color=(10, 20, 30), size=(24, 24)) -> bytes:
    buf = io.BytesIO()
    Image.new("RGB", size, color).save(buf, format="PNG")
    return buf.getvalue()


def _book_with_anchored(path, cell: str, data: bytes) -> str:
    """An .xlsx whose only picture is anchored over *cell*."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Style No."
    ws.add_image(XLImage(io.BytesIO(data)), cell)
    wb.save(path)
    return str(path)


# ── The anchored form is read at all ────────────────────────────────────────

def test_anchored_picture_is_located_at_its_cell(tmp_path):
    data = _png()
    book = _book_with_anchored(tmp_path / "a.xlsx", "G18", data)

    pos = extract_anchored_positions(book, sheet_index=0)
    assert list(pos) == [(18, 7)]                      # G18, 1-based (row, col)
    assert extract_images_from_xlsx(book)[pos[(18, 7)]] == data


def test_anchored_picture_is_invisible_to_the_dispimg_reader(tmp_path):
    """The reason it was missed: the cell holds no formula to find."""
    book = _book_with_anchored(tmp_path / "a.xlsx", "G18", _png())
    assert extract_dispimg_positions(book, sheet_index=0) == {}


def test_image_cache_picks_up_anchored_pictures(tmp_path):
    book = _book_with_anchored(tmp_path / "a.xlsx", "C3", _png())
    cache = ImageCache()
    assert cache.add_file(book) == 1
    assert len(cache) == 1
    assert cache.get(cache.all_ids()[0]) is not None


# ── The generated id ────────────────────────────────────────────────────────

def test_the_same_photo_gets_the_same_id_in_a_different_workbook(tmp_path):
    """An anchored picture has no id of its own, so one is derived from its
    content: re-uploading the same photo must not file it twice."""
    data = _png()
    a = _book_with_anchored(tmp_path / "a.xlsx", "B2", data)
    b = _book_with_anchored(tmp_path / "b.xlsx", "H9", data)
    assert list(extract_images_from_xlsx(a)) == list(extract_images_from_xlsx(b))


def test_two_photos_both_called_image1_do_not_collide(tmp_path):
    """Every workbook names its first picture xl/media/image1.png. A picture_id
    becomes a filename in a shared folder downstream, so deriving the id from
    that name would attach one client's photo to another's style."""
    a = _book_with_anchored(tmp_path / "a.xlsx", "B2", _png((1, 2, 3)))
    b = _book_with_anchored(tmp_path / "b.xlsx", "B2", _png((250, 240, 230)))
    assert set(extract_images_from_xlsx(a)) != set(extract_images_from_xlsx(b))


def test_generated_ids_are_safe_to_use_as_filenames(tmp_path):
    book = _book_with_anchored(tmp_path / "a.xlsx", "B2", _png())
    for img_id in extract_images_from_xlsx(book):
        assert img_id.replace("_", "").isalnum()


# ── Files that are not well-formed must not raise ───────────────────────────

def test_a_linked_picture_with_no_bytes_is_skipped(tmp_path):
    """Real files carry Target="NULL" TargetMode="External" placeholders — a
    relationship pointing at nothing this workbook contains."""
    book = _book_with_anchored(tmp_path / "a.xlsx", "B2", _png())
    broken = str(tmp_path / "broken.xlsx")
    with zipfile.ZipFile(book) as src, zipfile.ZipFile(broken, "w") as dst:
        for item in src.infolist():
            blob = src.read(item.filename)
            if item.filename.endswith("drawing1.xml.rels"):
                before = blob
                blob = blob.replace(b'Target="/xl/media/image1.png"',
                                    b'Target="NULL" TargetMode="External"')
                assert blob != before, "test fixture no longer matches"
            dst.writestr(item, blob)

    assert extract_anchored_positions(broken, sheet_index=0) == {}
    assert extract_images_from_xlsx(broken) == {}


def test_a_workbook_with_no_pictures_yields_nothing(tmp_path):
    path = str(tmp_path / "plain.xlsx")
    wb = openpyxl.Workbook()
    wb.active["A1"] = "no pictures here"
    wb.save(path)
    assert extract_anchored_positions(path, sheet_index=0) == {}
    assert extract_images_from_xlsx(path) == {}


def test_an_unreadable_file_is_skipped_not_raised(tmp_path):
    junk = tmp_path / "not.xlsx"
    junk.write_bytes(b"this is not a zip")
    assert extract_anchored_positions(str(junk), sheet_index=0) == {}
    assert extract_images_from_xlsx(str(junk)) == {}


def test_a_missing_file_is_skipped_not_raised(tmp_path):
    missing = str(tmp_path / "nope.xlsx")
    assert extract_anchored_positions(missing, sheet_index=0) == {}
    assert extract_images_from_xlsx(missing) == {}


# ── Reading from bytes, as the progress-sheet loader does ───────────────────

def test_positions_can_be_read_from_bytes(tmp_path):
    book = _book_with_anchored(tmp_path / "a.xlsx", "D5", _png())
    with open(book, "rb") as fh:
        raw = fh.read()
    assert extract_anchored_positions(raw, sheet_index=0) == \
        extract_anchored_positions(book, sheet_index=0)
    assert extract_anchored_positions(io.BytesIO(raw), sheet_index=0) == \
        extract_anchored_positions(book, sheet_index=0)


def test_only_the_named_sheet_is_searched(tmp_path):
    """Image positions were once always read from sheet 1 even when the
    contract lived on a later sheet, attaching wrong photos by row collision."""
    path = str(tmp_path / "two.xlsx")
    wb = openpyxl.Workbook()
    wb.active.title = "cover"
    ws2 = wb.create_sheet("contract")
    ws2.add_image(XLImage(io.BytesIO(_png())), "G18")
    wb.save(path)

    assert extract_anchored_positions(path, sheet_index=0) == {}
    assert list(extract_anchored_positions(path, sheet_index=1)) == [(18, 7)]
    # …but the byte extractor sweeps the whole workbook, so the cache still
    # holds the picture whichever sheet it was anchored on.
    assert len(extract_images_from_xlsx(path)) == 1
