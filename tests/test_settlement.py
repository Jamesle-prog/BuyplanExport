"""Tests for the settlement statistics module (结算统计表).

The workbook is maintained by hand and its sheets disagree with each other:
column order differs between clients and years, one book is in GBP, and two
facts — discount risk and contract status — are carried only as cell fills.
The fixtures below reproduce those disagreements deliberately.
"""
from __future__ import annotations

import io
import sqlite3

import openpyxl
import pytest
from openpyxl.styles import PatternFill

from po_extractor.parsers.settlement import (
    SettlementParseError, cost_total, margin, outstanding, parse_settlement,
    sheet_identity,
)
from po_extractor.store.settlement_store import SettlementStore

RED = "FFFF0000"


def _bulk_sheet(ws, *, order: str = "zalando2026", currency: str = "",
                rows: list[list] | None = None) -> None:
    """Write a settlement sheet with a two-row header.

    *order* picks one of the real column orders — Zalando2026 puts 辅助列/款号/
    PO# where every other sheet puts PO#/辅助列/款号.
    """
    ws["A1"] = "合同已寄工厂"
    ws["A2"] = "红底的是有折扣风险的"
    ws["A2"].fill = PatternFill("solid", fgColor=RED)

    if order == "zalando2026":
        head = ["INVOICE NO.", "辅助列", "款号", "PO#"]
    else:
        head = ["INVOICE NO.", "PO#", "辅助列", "款号"]
    head += ["合同号", "合同数量", "翻单还是新款", "工厂", "加工费",
             f"FOB{currency}", f"合同金额{currency}", None, "离厂时间",
             "出货数", "溢短装", "发票金额\n(报关金额）", "实际收汇", "日期",
             "支付", None, None, None, None, None, "费用"]
    for c, v in enumerate(head, start=1):
        ws.cell(4, c, v)
    for c, v in [(19, "面料款"), (20, "日期"), (21, "辅料款"), (22, "日期"),
                 (23, "加工费"), (24, "日期"), (25, "港杂费(含运费)"),
                 (26, "其他一"), (27, "其他二(税金）")]:
        ws.cell(5, c, v)
    for i, row in enumerate(rows or [], start=6):
        for c, v in enumerate(row, start=1):
            if v is not None:
                ws.cell(i, c, v)


def _book(**sheets) -> io.BytesIO:
    """An .xlsx in memory. ``sheets`` maps a title to a writer callable."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for title, writer in sheets.items():
        writer(wb.create_sheet(title))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _zalando2026(ws):
    _bulk_sheet(ws, order="zalando2026", rows=[
        # inv, helper, style, po, contract, qty, new, factory, cmt, fob, amt,
        # _, exfty, shipped, oversh, invamt, recvd, date, then pay/fee
        ["S267009", None, "DR4319", "PO2360361C", "26302-ZA7002", 400, "New",
         "XZY", 18, 7.9, 3160, None, "2026-03-17", 413, 0.0325, 3262.7,
         3262.7, "2026-06-05", 900, "2026-04-01", 100, "2026-04-02",
         1200, "2026-04-03", 60, 10, 25],
        ["S267010", None, "DR4313", "PO2360362C", "26302-ZA7003", 100, "New",
         "XZY", 18, 7.9, 790, None, "2026-03-17", 100, 0, 790,
         None, "", None, "", None, "", None, "", None, None, None],
    ])
    ws.cell(7, 1).fill = PatternFill("solid", fgColor=RED)   # discount risk


@pytest.fixture
def store(tmp_path):
    SettlementStore._checked_paths.clear()
    return SettlementStore(str(tmp_path / "po_history.db"))


# ── Parser: headings, not positions ─────────────────────────────────────────

def test_columns_are_found_by_heading_not_position():
    """Zalando2026 puts 辅助列/款号/PO# where Zalando2025 puts PO#/辅助列/款号.
    Reading either by position gets the other one's data."""
    def y2025(ws):
        _bulk_sheet(ws, order="other", rows=[
            ["S257025", "PO2091447C", None, "ZLD114", "25302-ZA7060", 400,
             "翻", "XZY", 13, 13.5, 5400, None, "2025-08-28", 418, 0.045,
             5643, 5643, "2026-02-11"]])
    d = parse_settlement(_book(Zalando2026=_zalando2026, Zalando2025=y2025))
    by_inv = {r["invoice_no"]: r for r in d["rows"]}
    assert by_inv["S267009"]["style"] == "DR4319"
    assert by_inv["S267009"]["po_number"] == "PO2360361C"
    assert by_inv["S257025"]["style"] == "ZLD114"
    assert by_inv["S257025"]["po_number"] == "PO2091447C"


def test_each_payment_takes_the_date_column_on_its_right():
    """Three payment dates all read 日期 — only position says which is which."""
    r = parse_settlement(_book(Zalando2026=_zalando2026))["rows"][0]
    assert (r["pay_fabric"], r["pay_fabric_date"]) == (900, "2026-04-01")
    assert (r["pay_trim"],   r["pay_trim_date"])   == (100, "2026-04-02")
    assert (r["pay_cmt"],    r["pay_cmt_date"])    == (1200, "2026-04-03")


def test_fees_are_read_from_the_second_group():
    r = parse_settlement(_book(Zalando2026=_zalando2026))["rows"][0]
    assert (r["fee_port"], r["fee_other1"], r["fee_other2"]) == (60, 10, 25)


def test_currency_comes_from_the_heading():
    def sc(ws):
        _bulk_sheet(ws, order="other", currency="（£）", rows=[
            ["", "PJX260001", None, "SC-55", "YMX2026", 770, None, "XZY",
             None, 6.59, 5074.3]])
    d = parse_settlement(_book(SC2026=sc))
    assert d["rows"][0]["currency"] == "GBP"


def test_percentages_and_amounts_survive_their_formatting():
    def s(ws):
        _bulk_sheet(ws, order="other", rows=[
            ["S1", "PO1", None, "ST1", "C1", "1,200", None, "F", None,
             "$7.90", "9,480", None, "2026-01-02", 1200, "3.25%"]])
    r = parse_settlement(_book(X2026=s))["rows"][0]
    assert (r["contract_qty"], r["fob"], r["contract_amount"]) == (1200, 7.9, 9480)
    assert r["over_short_pct"] == pytest.approx(0.0325)


def test_an_ex_factory_revision_is_kept_verbatim():
    """A moved ship date is written 2025/8/28->9/4 — that is information, and
    dropping it for not parsing as a date loses it."""
    def s(ws):
        _bulk_sheet(ws, order="other", rows=[
            ["S1", "PO1", None, "ST1", "C1", 400, None, "F", None, 9.1, 3640,
             None, "2025/8/28->9/4"]])
    assert parse_settlement(_book(X2025=s))["rows"][0]["ex_factory"] \
        == "2025/8/28->9/4"


def test_rows_with_no_identifying_field_are_not_lines():
    """The sheets carry stray totals and notes below the data."""
    def s(ws):
        _bulk_sheet(ws, order="other", rows=[
            ["S1", "PO1", None, "ST1", "C1", 400],
            [None] * 21 + [2665.01],            # a stray fee total
            [None, None, None, None, None, None],
        ])
    assert len(parse_settlement(_book(X2026=s))["rows"]) == 1


# ── Parser: colour is data ──────────────────────────────────────────────────

def test_discount_risk_is_read_from_the_sheets_own_legend():
    """The flag exists only as a red fill on the invoice number. Reading the
    legend rather than hardcoding a colour keeps a re-themed book working."""
    rows = {r["invoice_no"]: r for r in
            parse_settlement(_book(Zalando2026=_zalando2026))["rows"]}
    assert rows["S267009"]["discount_risk"] == 0
    assert rows["S267010"]["discount_risk"] == 1


def test_a_workbook_with_no_legend_flags_nothing():
    def s(ws):
        _bulk_sheet(ws, order="other", rows=[["S1", "PO1", None, "ST1", "C1", 4]])
        ws["A2"] = None                       # legend removed
        ws.cell(6, 1).fill = PatternFill("solid", fgColor=RED)
    assert parse_settlement(_book(X2026=s))["rows"][0]["discount_risk"] == 0


# ── Parser: which sheets ────────────────────────────────────────────────────

def test_derived_sheets_are_skipped_not_imported():
    """到期未付款明细 and 折扣风险明细 are views of the same rows; a stored
    copy would go stale the moment a payment lands."""
    def derived(ws):
        ws["D2"], ws["E2"] = "客人", "金额（$)"
        ws["D3"], ws["E3"] = "Zalando", 197730.88
    d = parse_settlement(_book(**{"Zalando2026": _zalando2026,
                                  "到期未付款明细 2.4": derived}))
    assert [s["sheet"] for s in d["sheets"]] == ["Zalando2026"]
    assert any("到期未付款明细" in s for s in d["skipped"])


def test_a_scratch_sheet_is_skipped_rather_than_half_read():
    def scratch(ws):
        for i, v in enumerate(["DR4319", "DR4313"], start=2):
            ws.cell(i, 8, v)
    d = parse_settlement(_book(Zalando2026=_zalando2026, Sheet1=scratch))
    assert any("Sheet1" in s for s in d["skipped"])


def test_a_workbook_with_no_settlement_sheet_is_rejected():
    with pytest.raises(SettlementParseError):
        parse_settlement(_book(Sheet1=lambda ws: ws.__setitem__("A1", "hi")))


def test_samples_sheet_is_read_per_style():
    def samples(ws):
        for c, v in enumerate(["款号", "客人", "工厂", "面料", "里布", "螺纹",
                               "辅料", "合计"], start=1):
            ws.cell(4, c, v)
        ws.append([])
        for c, v in enumerate(["Rider Jacket", "SCRT", "朱师傅", 264, 108,
                               150, None, 522], start=1):
            ws.cell(6, c, v)
    d = parse_settlement(_book(Zalando2026=_zalando2026, 样品=samples))
    assert d["samples"] == [{"style": "Rider Jacket", "client": "SCRT",
                             "factory": "朱师傅", "fabric": 264.0,
                             "lining": 108.0, "rib": 150.0, "trim": None,
                             "total": 522.0, "row_no": 6}]


@pytest.mark.parametrize("name,expect", [
    ("Zalando2026", ("Zalando", "2026")),
    ("SC2025 ", ("SC", "2025")),
    ("2025其它 客人大货", ("其它", "2025")),
    ("misc", ("misc", "")),
])
def test_sheet_identity(name, expect):
    assert sheet_identity(name) == expect


# ── Derived money ───────────────────────────────────────────────────────────

def test_outstanding_falls_back_to_the_contract_amount():
    """A line that shipped but hasn't been invoiced is still owed; treating a
    missing invoice as nothing billed would read as settled."""
    assert outstanding({"contract_amount": 5000, "received": None}) == 5000
    assert outstanding({"invoice_amount": 5100, "contract_amount": 5000,
                        "received": 5100}) == 0


def test_margin_is_blank_until_something_has_been_paid_out():
    """0 would read as break-even, which is a different claim from unknown."""
    assert margin({"received": 3262.7}) is None
    assert margin({"received": 3262.7, "pay_cmt": 1200, "fee_port": 60}) \
        == pytest.approx(2002.7)
    assert cost_total({"pay_fabric": 900, "pay_trim": 100, "pay_cmt": 1200,
                       "fee_port": 60, "fee_other1": 10, "fee_other2": 25}) \
        == 2295


# ── Store ───────────────────────────────────────────────────────────────────

def test_import_replaces_only_the_sheets_it_carries(store):
    def y2025(ws):
        _bulk_sheet(ws, order="other", rows=[
            ["S257001", "PO1", None, "ZLD1", "25302-1", 400]])
    store.import_parsed(parse_settlement(
        _book(Zalando2026=_zalando2026, Zalando2025=y2025)))
    assert store.rows_by_sheet() == {"Zalando2026": 2, "Zalando2025": 1}

    # A later book carrying only 2026 must not delete 2025.
    res = store.import_parsed(parse_settlement(_book(Zalando2026=_zalando2026)))
    assert store.rows_by_sheet() == {"Zalando2026": 2, "Zalando2025": 1}
    assert res["untouched"] == ["Zalando2025"]


def test_reimport_does_not_duplicate_rows(store):
    for _ in range(3):
        store.import_parsed(parse_settlement(_book(Zalando2026=_zalando2026)))
    assert store.rows_by_sheet() == {"Zalando2026": 2}


def test_derived_columns_are_computed_on_read(store):
    store.import_parsed(parse_settlement(_book(Zalando2026=_zalando2026)))
    df = store.list_rows().set_index("invoice_no")
    assert df.loc["S267009", "outstanding"] == 0
    assert df.loc["S267009", "cost_total"] == 2295
    assert df.loc["S267009", "margin"] == pytest.approx(967.7)
    # Nothing received and nothing paid out: still owed, margin unknown.
    assert df.loc["S267010", "outstanding"] == 790
    assert pd_isna(df.loc["S267010", "margin"])


def pd_isna(v) -> bool:
    import pandas as pd
    return bool(pd.isna(v))


def test_filters_and_search(store):
    store.import_parsed(parse_settlement(_book(Zalando2026=_zalando2026)))
    assert store.distinct("client") == ["Zalando"]
    assert len(store.list_rows(clients=["Zalando"])) == 2
    assert len(store.list_rows(clients=["nobody"])) == 0
    assert list(store.list_rows(search="DR4313")["invoice_no"]) == ["S267010"]


def test_import_history_records_what_was_replaced(store):
    store.import_parsed(parse_settlement(_book(Zalando2026=_zalando2026)),
                        source_file="结算统计表.xlsx", file_bytes=b"x",
                        imported_by="angel")
    row = store.list_imports().iloc[0]
    assert row["source_file"] == "结算统计表.xlsx"
    assert row["imported_by"] == "angel"
    assert row["n_rows"] == 2 and row["sheets"] == "Zalando2026"


# ── Cross-reference with the orders already in the system ───────────────────

def _seed_items(db_path, rows) -> None:
    conn = sqlite3.connect(db_path)
    try:
        conn.execute(
            """CREATE TABLE IF NOT EXISTS sky_east_items (
                   id INTEGER PRIMARY KEY AUTOINCREMENT,
                   pc_no TEXT, zalando_po TEXT, style TEXT,
                   color_name TEXT DEFAULT '', total_qty INTEGER DEFAULT 0,
                   ex_fty_date TEXT DEFAULT '')""")
        conn.executemany(
            "INSERT INTO sky_east_items "
            "(pc_no, zalando_po, style, total_qty, ex_fty_date) "
            "VALUES (?,?,?,?,?)", rows)
        conn.commit()
    finally:
        conn.close()


def test_lines_match_on_po_and_style_together(store):
    """A PO covers several styles, each invoiced separately — matching on the
    PO alone fans one settlement line across all of them."""
    _seed_items(store.db_path, [
        ("PC1", "PO2360361C", "DR4319", 400, "2026-03-17"),
        ("PC1", "PO2360361C", "DR9999", 250, "2026-03-17"),
    ])
    store.import_parsed(parse_settlement(_book(Zalando2026=_zalando2026)))
    m = store.match_orders().set_index("invoice_no")
    assert len(m) == 2                              # not 3
    assert m.loc["S267009", "se_qty"] == 400


def test_a_line_naming_an_unknown_po_is_reported(store):
    _seed_items(store.db_path, [("PC1", "PO2360361C", "DR4319", 400, "")])
    store.import_parsed(parse_settlement(_book(Zalando2026=_zalando2026)))
    assert list(store.unmatched_pos()["invoice_no"]) == ["S267010"]


def test_matching_survives_a_database_with_no_order_tables(store):
    """Same guard the cut-plan store uses: a fresh install must report, not
    raise."""
    store.import_parsed(parse_settlement(_book(Zalando2026=_zalando2026)))
    assert store.match_orders().empty
    assert store.unmatched_pos().empty


def test_reading_an_empty_store_gives_the_derived_columns_anyway(store):
    df = store.list_rows()
    assert df.empty
    for col in ("outstanding", "cost_total", "margin"):
        assert col in df.columns
