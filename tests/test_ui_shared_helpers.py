"""ui/shared.py helpers added in v2.125.4: df_to_xlsx_bytes and the
multiselect-with-Select-all picker's stale-value handling; ui/schema_labels."""
import io

import openpyxl
import pandas as pd

from ui.shared import df_to_xlsx_bytes


def _sheets(data: bytes):
    return openpyxl.load_workbook(io.BytesIO(data))


def test_df_to_xlsx_bytes_single_sheet():
    df = pd.DataFrame({"a": [1, 2], "b": ["x", "y"]})
    wb = _sheets(df_to_xlsx_bytes(df, sheet_name="Issues"))
    assert wb.sheetnames == ["Issues"]
    ws = wb["Issues"]
    assert [c.value for c in ws[1]] == ["a", "b"] and ws["A3"].value == 2


def test_df_to_xlsx_bytes_multi_sheet_skips_none():
    wb = _sheets(df_to_xlsx_bytes({"Summary": pd.DataFrame({"k": [1]}),
                                   "GIII POs": None,
                                   "Sky East Items": pd.DataFrame({"z": ["q"]})}))
    assert wb.sheetnames == ["Summary", "Sky East Items"]


def test_df_to_xlsx_bytes_autofit_widths():
    df = pd.DataFrame({"short": ["a"], "long": ["x" * 100]})
    ws = _sheets(df_to_xlsx_bytes(df, autofit=True, max_width=60)).active
    assert ws.column_dimensions["A"].width == len("short") + 4
    assert ws.column_dimensions["B"].width == 60


def test_live_label_falls_back():
    from ui.schema_labels import live_label
    assert live_label("no_such_column_xyz", "Fallback") == "Fallback"
    assert isinstance(live_label("po_number", "PO"), str)
