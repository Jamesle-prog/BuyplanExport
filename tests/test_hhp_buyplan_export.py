"""Tests for the HHP / Zalando buy-plan exporter's per-sheet data writers.

Covers three fixes to ``_fill_template_sheet`` / ``_fill_blank_sheet``:
  * grouping no longer silently narrows (and merges different colors/SKUs
    together) when an optional column is missing from the input;
  * stale template rows past the last written group get cleared;
  * the fabric block (B/D/E) clears stale values instead of leaving them
    when a style has no position/details for a given row.

The real BuyPlan_Template.xlsx is not present in this checkout (data/ is
gitignored runtime state), so these exercise the private row-writing
functions directly against a plain openpyxl worksheet rather than going
through ``export_hhp_buyplan``'s template-file lookup.
"""
from __future__ import annotations

import pandas as pd
import pytest
from openpyxl import Workbook

from po_extractor.exporters.hhp_buyplan_export import (
    DATA_START_ROW, SIZE_COL_START, _fill_fabric_block, _fill_template_sheet,
    _safe_group_rows,
)


# ── Fix: grouping must not silently narrow when a column is missing ──────────

def test_safe_group_rows_groups_by_all_three_when_present():
    df = pd.DataFrame({
        "Purchase Order Number": ["PO1", "PO1", "PO2"],
        "Config SKU": ["SKU1", "SKU1", "SKU2"],
        "Main Supplier Color Description": ["RED", "RED", "BLUE"],
        "S": [10, 5, 3],
    })
    groups = _safe_group_rows(df)
    assert len(groups) == 2
    sizes = sorted(len(grp) for _key, grp in groups)
    assert sizes == [1, 2]   # PO1/SKU1/RED rows merged, PO2/SKU2/BLUE alone


def test_safe_group_rows_falls_back_to_per_row_when_color_col_missing():
    """Without 'Main Supplier Color Description', two rows sharing PO+SKU
    used to be merged into one group and have their sizes summed together —
    even if they're actually different colors the column just isn't there
    to prove it. Must now keep them as separate groups instead of guessing."""
    df = pd.DataFrame({
        "Purchase Order Number": ["PO1", "PO1"],
        "Config SKU": ["SKU1", "SKU1"],
        "S": [10, 20],
    })
    with pytest.warns(UserWarning, match="grouping column"):
        groups = _safe_group_rows(df)
    assert len(groups) == 2
    assert sorted(grp["S"].iloc[0] for _key, grp in groups) == [10, 20]


def test_fill_template_sheet_does_not_merge_rows_when_color_col_missing():
    """End-to-end through _fill_template_sheet: two rows with different
    quantities but no color/SKU column present must land on separate
    output rows, never summed into one."""
    wb = Workbook()
    ws = wb.active
    sub = pd.DataFrame({
        "Purchase Order Number": ["PO1", "PO1"],
        "合同号": ["C1", "C1"],
        "XS": [10, 999],   # if merged, XS would wrongly read 1009
    })
    with pytest.warns(UserWarning, match="grouping column"):
        _fill_template_sheet(ws, "STYLE1", sub, fm_cache={})

    xs_values = [ws.cell(row=r, column=SIZE_COL_START).value
                 for r in (DATA_START_ROW, DATA_START_ROW + 1)]
    assert sorted(v for v in xs_values if v) == [10, 999]


# ── Fix: stale template rows past the last written group must be cleared ────

def test_fill_template_sheet_clears_stale_rows_beyond_last_group():
    wb = Workbook()
    ws = wb.active
    # Simulate leftover content below the data area (e.g. a template mockup
    # row, or a previous run's data that a smaller run must not inherit).
    ws.cell(row=12, column=1).value = "STALE"
    ws.cell(row=12, column=6).value = "STALE_SKU"
    ws.merge_cells("A12:C12")

    sub = pd.DataFrame({
        "Purchase Order Number": ["PO1"],
        "Config SKU": ["SKU1"],
        "Main Supplier Color Description": ["RED"],
        "合同号": ["C1"],
    })
    _fill_template_sheet(ws, "STYLE1", sub, fm_cache={})

    assert ws.cell(row=DATA_START_ROW, column=1).value == "C1"   # new data written
    assert ws.cell(row=12, column=1).value is None, "stale row 12 was not cleared"
    assert ws.cell(row=12, column=6).value is None
    assert "A12:C12" not in {str(m) for m in ws.merged_cells.ranges}


def test_fill_template_sheet_keeps_all_written_rows_intact():
    """Clearing past last_row must never touch the rows just written."""
    wb = Workbook()
    ws = wb.active
    sub = pd.DataFrame({
        "Purchase Order Number": ["PO1", "PO2"],
        "Config SKU": ["SKU1", "SKU2"],
        "Main Supplier Color Description": ["RED", "BLUE"],
        "合同号": ["C1", "C2"],
    })
    _fill_template_sheet(ws, "STYLE1", sub, fm_cache={})
    assert ws.cell(row=DATA_START_ROW, column=1).value == "C1"
    assert ws.cell(row=DATA_START_ROW + 1, column=1).value == "C2"


# ── Fix: fabric block (B/D/E) must clear stale values, not just skip them ──

def test_fill_fabric_block_clears_stale_values_when_data_absent():
    wb = Workbook()
    ws = wb.active
    ws.cell(row=2, column=2).value = "STALE POSITION"
    ws.cell(row=2, column=4).value = "STALE|DETAILS|1|2"
    ws.cell(row=2, column=5).value = "STALE E"

    empty_row = pd.Series(dtype=object)   # no 面料_ columns at all
    _fill_fabric_block(ws, empty_row, fm_cache={})

    assert ws.cell(row=2, column=2).value is None
    assert ws.cell(row=2, column=4).value is None
    assert ws.cell(row=2, column=5).value is None


def test_fill_fabric_block_writes_position_and_details_when_present():
    wb = Workbook()
    ws = wb.active
    row = pd.Series({
        "面料_面料_部位": "大身",
        "面料_面料_编号": "HHN-1",
        "面料_面料_成分": "100%Cotton",
    })
    _fill_fabric_block(ws, row, fm_cache={})
    assert ws.cell(row=2, column=2).value == "大身"
    assert ws.cell(row=2, column=4).value == "HHN-1|100%Cotton||"
    assert ws.cell(row=2, column=5).value is None
