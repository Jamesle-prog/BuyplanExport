"""Fabric-list upload peer review: propose → review → approve/reject.

Uploads never touch fabric_master directly -- they are parsed, validated,
diffed, and staged in fabric_pending_import/fabric_pending_rows until a
reviewer approves (two-person rule). Single-record deletes stay immediate
(already versioned) by design. Restore-a-version goes through the same gate.
"""
from __future__ import annotations

import io

import openpyxl
import pytest

from po_extractor.store.fabric_master_store import FabricMasterStore

# Canonical fallback columns (see _COL_MAP_FALLBACK) -- same convention as
# tests/test_fabric_versioning.py so values never cross-map into text fields.
_FIXTURE_COLS = {
    "quality_no": 1, "composition_en": 5, "weight_gsm": 10, "cuttable_width_cm": 11,
}
_FIXTURE_HEADERS = {
    "quality_no": "公司面料编号", "composition_en": "面料成分(英文)",
    "weight_gsm": "克重(gsm)", "cuttable_width_cm": "有效门幅(cm)",
}


def _make_fabric_xlsx(rows: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "all"
    for field, col in _FIXTURE_COLS.items():
        ws.cell(row=1, column=col, value=_FIXTURE_HEADERS[field])
    for ri, r in enumerate(rows, start=2):
        for field, col in _FIXTURE_COLS.items():
            ws.cell(row=ri, column=col, value=r.get(field))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _propose(store, tmp_path, name: str, rows: list[dict], **kw) -> dict:
    path = tmp_path / name
    path.write_bytes(_make_fabric_xlsx(rows))
    return store.propose_import(str(path), source_file_name=name, **kw)


def _import_direct(store, tmp_path, name: str, rows: list[dict]) -> dict:
    path = tmp_path / name
    path.write_bytes(_make_fabric_xlsx(rows))
    return store.import_from_xlsx(str(path), source_file_name=name)


@pytest.fixture
def store(tmp_path):
    return FabricMasterStore(str(tmp_path / "fabric_master.db"))


_ROWS_V1 = [
    {"quality_no": "A", "composition_en": "100%Cotton", "weight_gsm": 200, "cuttable_width_cm": 140},
    {"quality_no": "B", "composition_en": "100%Polyester", "weight_gsm": 150, "cuttable_width_cm": 150},
]


def test_propose_stages_without_touching_live_table(store, tmp_path):
    result = _propose(store, tmp_path, "v1.xlsx", _ROWS_V1, proposed_by="fabric")
    assert "pending_id" in result
    assert result["diff_added"] == 2

    assert store.count() == 0                    # live table untouched
    assert store.list_versions() == []           # no version minted

    pending = store.get_pending()
    assert pending is not None
    assert pending["proposed_by"] == "fabric"
    assert pending["row_count"] == 2
    assert pending["status"] == "pending"


def test_approve_applies_and_mints_version_with_reviewer(store, tmp_path):
    r = _propose(store, tmp_path, "v1.xlsx", _ROWS_V1, proposed_by="fabric")
    outcome = store.approve_pending(r["pending_id"], reviewed_by="admin",
                                    comment="checked against supplier quotes")

    assert outcome["unchanged"] is False
    assert outcome["version_id"] == 1
    assert store.count() == 2

    v = store.list_versions()[0]
    assert v["uploaded_by"] == "fabric"
    assert v["approved_by"] == "admin"
    assert v["review_comment"] == "checked against supplier quotes"
    assert store.get_pending() is None           # resolved
    assert store.get_diff_summary(1) == {"added": 2, "removed": 0, "changed": 0}


def test_reject_discards_without_applying(store, tmp_path):
    r = _propose(store, tmp_path, "v1.xlsx", _ROWS_V1, proposed_by="fabric")
    store.reject_pending(r["pending_id"], reviewed_by="admin",
                         comment="wrong file — this is last month's list")

    assert store.count() == 0
    assert store.list_versions() == []
    assert store.get_pending() is None


def test_reject_requires_a_reason(store, tmp_path):
    r = _propose(store, tmp_path, "v1.xlsx", _ROWS_V1, proposed_by="fabric")
    with pytest.raises(ValueError):
        store.reject_pending(r["pending_id"], reviewed_by="admin", comment="   ")


def test_self_approval_requires_comment(store, tmp_path):
    r = _propose(store, tmp_path, "v1.xlsx", _ROWS_V1, proposed_by="fabric")
    with pytest.raises(ValueError):
        store.approve_pending(r["pending_id"], reviewed_by="fabric", comment="")
    # With a justification comment the (admin-gated) self-approve path works.
    outcome = store.approve_pending(r["pending_id"], reviewed_by="fabric",
                                    comment="urgent — sole operator today")
    assert outcome["version_id"] == 1


def test_second_proposal_blocked_while_one_is_pending(store, tmp_path):
    _propose(store, tmp_path, "v1.xlsx", _ROWS_V1, proposed_by="fabric")
    result = _propose(store, tmp_path, "v2.xlsx", _ROWS_V1, proposed_by="admin")
    assert "blocked_by_pending" in result
    assert result["pending_proposed_by"] == "fabric"


def test_cancel_pending_allows_new_proposal(store, tmp_path):
    r = _propose(store, tmp_path, "v1.xlsx", _ROWS_V1, proposed_by="fabric")
    store.cancel_pending(r["pending_id"], by="fabric")
    assert store.get_pending() is None
    r2 = _propose(store, tmp_path, "v2.xlsx", _ROWS_V1, proposed_by="fabric")
    assert "pending_id" in r2


def test_noop_proposal_returns_unchanged_and_stages_nothing(store, tmp_path):
    _import_direct(store, tmp_path, "v1.xlsx", _ROWS_V1)   # baseline v1
    result = _propose(store, tmp_path, "same.xlsx", _ROWS_V1, proposed_by="fabric")
    assert result.get("unchanged") is True
    assert store.get_pending() is None


def test_pending_diff_recomputed_for_review_panel(store, tmp_path):
    _import_direct(store, tmp_path, "v1.xlsx", _ROWS_V1)
    changed = [dict(_ROWS_V1[0], weight_gsm=220), _ROWS_V1[1]]
    r = _propose(store, tmp_path, "v2.xlsx", changed, proposed_by="fabric")

    diff = store.get_pending_diff(r["pending_id"])
    assert len(diff) == 1
    assert diff[0]["quality_no"] == "A"
    assert diff[0]["change_type"] == "changed"
    assert diff[0]["field"] == "weight_gsm"


def test_approve_clear_first_replaces_whole_table(store, tmp_path):
    _import_direct(store, tmp_path, "v1.xlsx", _ROWS_V1)
    replacement = [
        {"quality_no": "A", "composition_en": "100%Cotton", "weight_gsm": 200, "cuttable_width_cm": 140},
        {"quality_no": "C", "composition_en": "100%Wool", "weight_gsm": 300, "cuttable_width_cm": 145},
    ]
    r = _propose(store, tmp_path, "v2.xlsx", replacement,
                 proposed_by="fabric", clear_first=True)
    store.approve_pending(r["pending_id"], reviewed_by="admin")

    assert store.count() == 2
    assert store.get_diff_summary(2) == {"added": 1, "removed": 1, "changed": 0}


def test_quality_warnings_surface_in_proposal(store, tmp_path):
    bad = [{"quality_no": "BAD-X-1", "composition_en": "60%Cotton 30%Polyester",
            "weight_gsm": 5000, "cuttable_width_cm": 140}]
    r = _propose(store, tmp_path, "bad.xlsx", bad, proposed_by="fabric")
    warnings = r["warnings"]
    assert any("100%" in w for w in warnings)          # composition sums to 90%
    assert any("5000" in w for w in warnings)          # weight out of range


def test_high_risk_flag_on_mass_removal(store, tmp_path):
    many = [{"quality_no": f"Q{i}", "composition_en": "100%Cotton",
             "weight_gsm": 200, "cuttable_width_cm": 140} for i in range(15)]
    _import_direct(store, tmp_path, "v1.xlsx", many)

    r = _propose(store, tmp_path, "v2.xlsx", many[:2],
                 proposed_by="fabric", clear_first=True)
    assert r["high_risk"] is True
    assert r["diff_removed"] == 13


def test_restore_version_goes_through_review_gate(store, tmp_path):
    _import_direct(store, tmp_path, "v1.xlsx", _ROWS_V1)
    changed = [dict(_ROWS_V1[0], weight_gsm=999), _ROWS_V1[1]]
    _import_direct(store, tmp_path, "v2.xlsx", changed)
    assert store.get_batch_enrichment(["A"])["A"]["weight_gsm"] == 999

    r = store.propose_restore(1, proposed_by="fabric")
    assert "pending_id" in r
    assert store.get_batch_enrichment(["A"])["A"]["weight_gsm"] == 999  # not yet applied

    outcome = store.approve_pending(r["pending_id"], reviewed_by="admin",
                                    comment="rolling back bad weights")
    assert outcome["version_id"] == 3
    assert store.get_batch_enrichment(["A"])["A"]["weight_gsm"] == 200  # restored

    v3 = store.list_versions()[0]
    assert v3["source_file"] == "restore v1"
    assert v3["approved_by"] == "admin"


def test_restore_of_pruned_version_raises(store, tmp_path):
    for i in range(1, 7):   # 6 versions -> v1's snapshot pruned (keep 4)
        _import_direct(store, tmp_path, f"v{i}.xlsx", [
            {"quality_no": "A", "composition_en": "100%Cotton",
             "weight_gsm": 100 + i, "cuttable_width_cm": 140},
        ])
    with pytest.raises(ValueError):
        store.propose_restore(1, proposed_by="fabric")


def test_approve_stale_or_resolved_pending_raises(store, tmp_path):
    r = _propose(store, tmp_path, "v1.xlsx", _ROWS_V1, proposed_by="fabric")
    store.approve_pending(r["pending_id"], reviewed_by="admin")
    with pytest.raises(ValueError):
        store.approve_pending(r["pending_id"], reviewed_by="admin")
