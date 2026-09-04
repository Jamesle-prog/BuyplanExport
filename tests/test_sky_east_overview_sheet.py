"""Tests for the Sky East Buy Plan's "Overview" sheet — a flat, item-level
cross-check table (one row per style/PO/colour) inserted right after Index,
mirroring the Contract History item-browser preview plus the Chinese colour
name and colour code alongside the English name.
"""
from __future__ import annotations

import io
import re

import pandas as pd
import pytest
from openpyxl import load_workbook
from PIL import Image

from po_extractor.exporters.sky_east_buyplan_export import export_sky_east_buyplan
from po_extractor.lookups.progress_lookup import PCColorMatch


def _png_bytes(size=(20, 20), color=(200, 50, 50)) -> bytes:
    buf = io.BytesIO()
    Image.new("RGB", size, color=color).save(buf, format="PNG")
    return buf.getvalue()


@pytest.fixture(autouse=True)
def _isolated_sky_east_store(tmp_path, monkeypatch):
    """Colour-lookup misses in this file are logged via
    ``get_sky_east_store().log_color_miss(...)`` (a real DB write). Point that
    factory at a throwaway tmp-path DB for every test here so a NOT_FOUND
    scenario never writes diagnostic rows into the shared dev database.
    """
    from po_extractor.store.sky_east_store import SkyEastStore

    store = SkyEastStore(str(tmp_path / "_isolated_sky_east.db"))
    monkeypatch.setattr("po_extractor.store.get_sky_east_store", lambda: store)
    return store


@pytest.fixture
def two_style_df():
    return pd.DataFrame([
        {
            "pc_no": "HHPPC048", "style": "DR5124", "brand": "Anna Field",
            "contract_no": "26302-ZA7148", "article_name": "LACE DRESS",
            "zalando_po": "PO001", "config_sku": "C1", "color_name": "dark blue",
            "fabric_item_no": "HHN-JA-01715", "fabrication": "cotton",
            "ex_fty_date": "2026-08-11",
            "xs": 30, "s": 82, "m": 119, "l": 102, "xl": 67, "xxl": 0,
        },
        {
            "pc_no": "HHPPC048", "style": "DR4578", "brand": "Anna Field",
            "contract_no": "26302-ZA7149", "article_name": "LONG SLEEVE DRESS",
            "zalando_po": "PO001", "config_sku": "C2", "color_name": "green",
            "fabric_item_no": "HHN-JSRSR-04068", "fabrication": "poly",
            "ex_fty_date": "2026-08-11",
            "xs": 22, "s": 61, "m": 90, "l": 77, "xl": 50, "xxl": 0,
        },
    ])


@pytest.fixture
def pc_color_lookup():
    return {
        ("HHPPC048", "DR5124", "Dark Blue"): PCColorMatch("藏青", "503", ""),
        ("HHPPC048", "DR4578", "Green"):     PCColorMatch("绿色", "602", ""),
    }


def test_overview_sheet_created_right_after_index(two_style_df, pc_color_lookup, tmp_path):
    path, _totals = export_sky_east_buyplan(
        two_style_df, cn_lookup={}, output_dir=str(tmp_path),
        label_lookup={}, cn_code_lookup={}, cn_by_pc_lookup=pc_color_lookup,
    )
    wb = load_workbook(path)
    assert wb.sheetnames[0] == "Index"
    assert wb.sheetnames[1] == "Overview"
    # Sheet names are "<index>_<style>" -- verify by suffix, not exact name.
    style_sheets = wb.sheetnames[2:]
    assert len(style_sheets) == 2
    assert any(s.endswith("_DR5124") for s in style_sheets)
    assert any(s.endswith("_DR4578") for s in style_sheets)


def test_overview_rows_include_english_and_chinese_color_plus_code(
    two_style_df, pc_color_lookup, tmp_path,
):
    path, _totals = export_sky_east_buyplan(
        two_style_df, cn_lookup={}, output_dir=str(tmp_path),
        label_lookup={}, cn_code_lookup={}, cn_by_pc_lookup=pc_color_lookup,
    )
    ov = load_workbook(path)["Overview"]
    headers = [c.value for c in ov[1]]
    assert "Color (EN)" in headers
    assert "Color (CN)" in headers
    assert "Color Code" in headers

    def _row_dict(ri):
        return dict(zip(headers, [ov.cell(ri, c + 1).value for c in range(len(headers))]))

    rows = [_row_dict(r) for r in (2, 3)]
    by_style = {r["Style"]: r for r in rows}

    # Colour (EN) echoes the order file verbatim ("dark blue", not title-cased);
    # the resolved CN/code prove the normalised form still drove the lookup.
    assert by_style["DR5124"]["Color (EN)"]   == "dark blue"
    assert by_style["DR5124"]["Color (CN)"]   == "藏青"
    assert by_style["DR5124"]["Color Code"]   == "503"
    assert by_style["DR5124"]["Contract No."] == "26302-ZA7148"
    assert by_style["DR5124"]["Total Qty"]    == 30 + 82 + 119 + 102 + 67

    assert by_style["DR4578"]["Color (EN)"] == "green"
    assert by_style["DR4578"]["Color (CN)"] == "绿色"
    assert by_style["DR4578"]["Color Code"] == "602"


def test_overview_includes_return_label_column(two_style_df, pc_color_lookup, tmp_path):
    df = two_style_df.copy()
    df.loc[df["style"] == "DR5124", "return_label"] = "Yes"
    df.loc[df["style"] == "DR4578", "return_label"] = "No"

    path, _totals = export_sky_east_buyplan(
        df, cn_lookup={}, output_dir=str(tmp_path),
        label_lookup={}, cn_code_lookup={}, cn_by_pc_lookup=pc_color_lookup,
    )
    ov = load_workbook(path)["Overview"]
    headers = [c.value for c in ov[1]]
    assert "Return Label" in headers

    def _row_dict(ri):
        return dict(zip(headers, [ov.cell(ri, c + 1).value for c in range(len(headers))]))

    rows = [_row_dict(r) for r in (2, 3)]
    by_style = {r["Style"]: r for r in rows}
    assert by_style["DR5124"]["Return Label"] == "Yes"
    assert by_style["DR4578"]["Return Label"] == "No"


def test_overview_return_label_defaults_to_na_when_absent(two_style_df, pc_color_lookup, tmp_path):
    path, _totals = export_sky_east_buyplan(
        two_style_df, cn_lookup={}, output_dir=str(tmp_path),
        label_lookup={}, cn_code_lookup={}, cn_by_pc_lookup=pc_color_lookup,
    )
    ov = load_workbook(path)["Overview"]
    headers = [c.value for c in ov[1]]

    def _row_dict(ri):
        return dict(zip(headers, [ov.cell(ri, c + 1).value for c in range(len(headers))]))

    assert _row_dict(2)["Return Label"] == "NA"
    assert _row_dict(3)["Return Label"] == "NA"


def test_index_sheet_includes_return_label_column(two_style_df, pc_color_lookup, tmp_path):
    df = two_style_df.copy()
    df.loc[df["style"] == "DR5124", "return_label"] = "Yes"
    df.loc[df["style"] == "DR4578", "return_label"] = "No"

    path, _totals = export_sky_east_buyplan(
        df, cn_lookup={}, output_dir=str(tmp_path),
        label_lookup={}, cn_code_lookup={}, cn_by_pc_lookup=pc_color_lookup,
    )
    idx = load_workbook(path)["Index"]
    headers = [c.value for c in idx[1]]
    assert "Return Label" in headers

    def _row_dict(ri):
        return dict(zip(headers, [idx.cell(ri, c + 1).value for c in range(len(headers))]))

    rows = [_row_dict(r) for r in range(2, idx.max_row + 1)]
    by_style = {r["款号"]: r for r in rows}
    assert by_style["DR5124"]["Return Label"] == "Yes"
    assert by_style["DR4578"]["Return Label"] == "No"


def _fabric_store_with_versions(tmp_path, monkeypatch, n_versions: int = 1):
    """Point the fabric-master factory at a tmp store holding *n_versions*
    real versions, so the Fabric Version sheet has deterministic content."""
    import io as _io
    import openpyxl as _oxl
    from po_extractor.store.fabric_master_store import FabricMasterStore
    import po_extractor.store as _store_pkg

    store = FabricMasterStore(str(tmp_path / "fabric_version_sheet.db"))
    for i in range(1, n_versions + 1):
        wb = _oxl.Workbook()
        ws = wb.active
        ws.title = "all"
        ws.cell(row=1, column=1, value="公司面料编号")
        ws.cell(row=1, column=10, value="克重(gsm)")
        ws.cell(row=2, column=1, value="HHN-TEST-0001")
        ws.cell(row=2, column=10, value=100 + i)
        buf = _io.BytesIO()
        wb.save(buf)
        p = tmp_path / f"fv{i}.xlsx"
        p.write_bytes(buf.getvalue())
        store.import_from_xlsx(str(p), source_file_name=f"fv{i}.xlsx")
    monkeypatch.setattr(_store_pkg, "get_fabric_master_store", lambda: store)
    return store


def test_index_sheet_carries_fabric_version_note(
    two_style_df, pc_color_lookup, tmp_path, monkeypatch,
):
    """The buy plan records which fabric-list version enriched it as a
    single discreet cell on the Index sheet -- deliberately NOT a separate
    tab and deliberately terse (version + date only): these files go to
    clients/factories, so internal detail (uploader, internal filename,
    change counts) stays out of the workbook entirely."""
    _fabric_store_with_versions(tmp_path, monkeypatch, n_versions=2)
    path, _totals = export_sky_east_buyplan(
        two_style_df, cn_lookup={}, output_dir=str(tmp_path),
        label_lookup={}, cn_code_lookup={}, cn_by_pc_lookup=pc_color_lookup,
    )
    wb = load_workbook(path)
    assert "Fabric Version" not in wb.sheetnames   # no separate tab
    ws = wb["Index"]
    values = [str(c.value) for row in ws.iter_rows() for c in row if c.value]
    note = next((v for v in values if "Fabric list version" in v), None)
    assert note is not None, values
    assert "v2" in note
    # Internal housekeeping detail must NOT leak into an outside-facing file.
    assert "fv2.xlsx" not in note
    assert not any("fv2.xlsx" in v for v in values)


def test_index_sheet_fabric_version_note_pinned_variant(
    two_style_df, pc_color_lookup, tmp_path, monkeypatch,
):
    _fabric_store_with_versions(tmp_path, monkeypatch, n_versions=2)
    path, _totals = export_sky_east_buyplan(
        two_style_df, cn_lookup={}, output_dir=str(tmp_path),
        label_lookup={}, cn_code_lookup={}, cn_by_pc_lookup=pc_color_lookup,
        fabric_version_id=1,
    )
    ws = load_workbook(path)["Index"]
    values = [str(c.value) for row in ws.iter_rows() for c in row if c.value]
    note = next((v for v in values if "Fabric list version" in v), None)
    assert note is not None, values
    assert "v1" in note and "pinned" in note


def test_index_sheet_return_label_defaults_to_na_when_absent(two_style_df, pc_color_lookup, tmp_path):
    path, _totals = export_sky_east_buyplan(
        two_style_df, cn_lookup={}, output_dir=str(tmp_path),
        label_lookup={}, cn_code_lookup={}, cn_by_pc_lookup=pc_color_lookup,
    )
    idx = load_workbook(path)["Index"]
    headers = [c.value for c in idx[1]]

    def _row_dict(ri):
        return dict(zip(headers, [idx.cell(ri, c + 1).value for c in range(len(headers))]))

    rows = [_row_dict(r) for r in range(2, idx.max_row + 1)]
    assert all(r["Return Label"] == "NA" for r in rows)


def test_buyplan_brand_sourced_from_progress(two_style_df, pc_color_lookup, tmp_path):
    """The 品牌 column comes from 大货进度表 (brand_by_pc_lookup) when present,
    keyed by PC No · style; styles absent from the lookup keep the order-file
    brand."""
    from po_extractor.lookups.progress_lookup import _norm_key
    brand_lkup = {(_norm_key("HHPPC048"), _norm_key("DR5124")): "Only"}   # DR4578 absent
    path, _ = export_sky_east_buyplan(
        two_style_df, cn_lookup={}, output_dir=str(tmp_path),
        label_lookup={}, cn_code_lookup={}, cn_by_pc_lookup=pc_color_lookup,
        brand_by_pc_lookup=brand_lkup,
    )
    wb = load_workbook(path)
    dr5124 = wb[next(s for s in wb.sheetnames if s.endswith("_DR5124"))]
    dr4578 = wb[next(s for s in wb.sheetnames if s.endswith("_DR4578"))]
    v5124 = [c.value for row in dr5124.iter_rows() for c in row]
    v4578 = [c.value for row in dr4578.iter_rows() for c in row]
    assert "Only" in v5124 and "Anna Field" not in v5124   # progress brand overrides
    assert "Anna Field" in v4578                            # no progress brand → order file


def test_overview_style_cell_hyperlinks_to_its_own_sheet(
    two_style_df, pc_color_lookup, tmp_path,
):
    path, _totals = export_sky_east_buyplan(
        two_style_df, cn_lookup={}, output_dir=str(tmp_path),
        label_lookup={}, cn_code_lookup={}, cn_by_pc_lookup=pc_color_lookup,
    )
    wb = load_workbook(path)
    ov = wb["Overview"]
    headers = [c.value for c in ov[1]]
    style_col = headers.index("Style") + 1
    for ri in (2, 3):
        cell = ov.cell(ri, style_col)
        assert cell.hyperlink is not None
        # Internal links use `location` (not `target`) -- a plain-string
        # `target="#'Sheet'!A1"` is written as an *external* relationship,
        # which Excel follows internally as a leniency but WPS does not.
        assert cell.hyperlink.target is None
        m = re.search(r"^'([^']+)'!A1$", cell.hyperlink.location)
        assert m, cell.hyperlink.location
        # Sheet names are "<index>_<style>" (e.g. "1_DR5124"), so the cell's
        # displayed style name is only a suffix of the linked sheet name now.
        target_sheet = m.group(1)
        assert target_sheet in wb.sheetnames
        assert target_sheet.endswith(f"_{cell.value}")


def test_overview_includes_fabric_and_display_key_columns(
    two_style_df, pc_color_lookup, tmp_path,
):
    path, _totals = export_sky_east_buyplan(
        two_style_df, cn_lookup={}, output_dir=str(tmp_path),
        label_lookup={}, cn_code_lookup={}, cn_by_pc_lookup=pc_color_lookup,
    )
    ov = load_workbook(path)["Overview"]
    headers = [c.value for c in ov[1]]
    assert "Fabric 1" in headers
    assert "综合标识 Key 1" in headers

    fab_col = headers.index("Fabric 1") + 1
    key_col = headers.index("综合标识 Key 1") + 1
    fab_values = [ov.cell(r, fab_col).value for r in (2, 3)]
    key_values = [ov.cell(r, key_col).value for r in (2, 3)]
    assert any(v and "HHN-JA-01715" in v for v in fab_values)
    assert all(v for v in key_values)   # display key always non-empty (falls back to bare HHN)


def test_overview_embeds_style_photo_when_available(two_style_df, pc_color_lookup, tmp_path):
    photo_map = {"DR5124": [_png_bytes()]}
    path, _totals = export_sky_east_buyplan(
        two_style_df, cn_lookup={}, output_dir=str(tmp_path),
        label_lookup={}, cn_code_lookup={}, cn_by_pc_lookup=pc_color_lookup,
        style_image_map=photo_map,
    )
    ov = load_workbook(path)["Overview"]
    headers = [c.value for c in ov[1]]
    assert "Photo" in headers
    assert len(ov._images) == 1   # only DR5124 has a photo; DR4578 doesn't


def test_overview_omits_photo_column_when_no_images_supplied(
    two_style_df, pc_color_lookup, tmp_path,
):
    path, _totals = export_sky_east_buyplan(
        two_style_df, cn_lookup={}, output_dir=str(tmp_path),
        label_lookup={}, cn_code_lookup={}, cn_by_pc_lookup=pc_color_lookup,
    )
    ov = load_workbook(path)["Overview"]
    headers = [c.value for c in ov[1]]
    assert "Photo" not in headers


def test_overview_not_created_when_no_items(tmp_path):
    empty_df = pd.DataFrame(columns=[
        "pc_no", "style", "brand", "contract_no", "article_name", "zalando_po",
        "config_sku", "color_name", "xs", "s", "m", "l", "xl", "xxl", "ex_fty_date",
    ])
    path, _totals = export_sky_east_buyplan(
        empty_df, cn_lookup={}, output_dir=str(tmp_path),
        label_lookup={}, cn_code_lookup={},
    )
    wb = load_workbook(path)
    assert "Overview" not in wb.sheetnames


# ---------------------------------------------------------------------------
# Decorative colour brackets — "(dark blue)" style values from some order files
# ---------------------------------------------------------------------------

def test_strip_color_brackets_single_wrap():
    from po_extractor.exporters.sky_east_buyplan_export import _strip_color_brackets
    assert _strip_color_brackets("(dark blue)") == "dark blue"
    assert _strip_color_brackets("(black)") == "black"


def test_strip_color_brackets_concatenated_two_tone():
    from po_extractor.exporters.sky_east_buyplan_export import _strip_color_brackets
    assert _strip_color_brackets("(black)(white)") == "black / white"
    assert _strip_color_brackets("(dark blue)(white)") == "dark blue / white"


def test_strip_color_brackets_passthrough_when_no_brackets():
    from po_extractor.exporters.sky_east_buyplan_export import _strip_color_brackets
    assert _strip_color_brackets("BLACK") == "BLACK"
    assert _strip_color_brackets("BLACK WITH WHITE STRAP AT WAIST AND BOTTOM") == \
        "BLACK WITH WHITE STRAP AT WAIST AND BOTTOM"
    assert _strip_color_brackets("") == ""
    assert _strip_color_brackets(None) is None


def test_buyplan_shows_client_colour_verbatim_but_matches_on_the_stripped_form(tmp_path):
    """The colour cell shows the client's PO string exactly as it arrived —
    brackets, casing and all — because the buy plan is read side by side with
    that PO and has to be greppable against it.

    Stripping stays in force for MATCHING: "(dark blue)" still has to resolve
    against 大货进度表 / the colour DB, which store plain "Dark Blue".  This
    test pins both halves at once, since they pull in opposite directions.
    """
    df = pd.DataFrame([{
        "pc_no": "HHPPC048", "style": "DR5124", "brand": "Anna Field",
        "contract_no": "26302-ZA7148", "article_name": "LACE INSERT DRESS",
        "zalando_po": "PO001", "config_sku": "C1", "color_name": "(dark blue)",
        "xs": 30, "s": 82, "m": 0, "l": 0, "xl": 0, "xxl": 0,
    }])
    # The DB is keyed on the stripped, title-cased name — matching only
    # succeeds if the bracketed value is still normalised before lookup.
    cn_lookup = {("Sky East", "Anna Field", "Dark Blue"): "深蓝色"}
    path, _totals = export_sky_east_buyplan(
        df, cn_lookup=cn_lookup, output_dir=str(tmp_path), label_lookup={},
        cn_code_lookup={},
    )
    wb = load_workbook(path)
    ov = wb["Overview"]
    headers = [c.value for c in ov[1]]
    color_col = headers.index("Color (EN)") + 1
    assert ov.cell(2, color_col).value == "(dark blue)"

    # ...and the bracketed value still matched, so the Chinese name resolved.
    cn_col = headers.index("Color (CN)") + 1
    assert ov.cell(2, cn_col).value == "深蓝色"

    # ...and the same on the style's own sheet.  Find the data row by its
    # contract number rather than hardcoding one: an earlier version of this
    # test scanned rows 5-7, which hold headers, so its assertion passed
    # without ever looking at a colour.
    style_ws = wb[next(s for s in wb.sheetnames if s.endswith("_DR5124"))]
    data_row = next(r for r in range(1, style_ws.max_row + 1)
                    if style_ws.cell(r, 1).value == "26302-ZA7148")
    assert style_ws.cell(data_row, 7).value == "(dark blue)"


# ---------------------------------------------------------------------------
# Colour source is authoritative — no silent cross-source fallback
# ---------------------------------------------------------------------------

def test_progress_source_miss_does_not_fall_back_to_internal_db(tmp_path):
    """When 大货进度表 is the selected source (cn_by_pc_lookup provided) and a
    style/colour isn't in it, the internal DB must NOT be consulted even if
    it has a matching entry -- that would silently show data from a source
    other than the one the user picked.
    """
    from po_extractor.exporters.sky_east_buyplan_export import _COLOR_NOT_FOUND

    df = pd.DataFrame([{
        "pc_no": "HHPPC048", "style": "DR5124", "brand": "Anna Field",
        "contract_no": "26302-ZA7148", "article_name": "LACE DRESS",
        "zalando_po": "PO001", "config_sku": "C1", "color_name": "dark blue",
        "xs": 30, "s": 82, "m": 0, "l": 0, "xl": 0, "xxl": 0,
    }])
    # cn_lookup/cn_code_lookup DO have a matching entry -- but must be ignored
    # because cn_by_pc_lookup (the selected source) is provided and misses.
    poisoned_cn      = {("Sky East", "Anna Field", "Dark Blue"): "藏青"}
    poisoned_cn_code = {("Sky East", "Anna Field", "Dark Blue"): "999"}
    path, _totals = export_sky_east_buyplan(
        df, cn_lookup=poisoned_cn, output_dir=str(tmp_path),
        label_lookup={}, cn_code_lookup=poisoned_cn_code,
        cn_by_pc_lookup={},   # progress source selected, but empty/no match
    )
    ov = load_workbook(path)["Overview"]
    headers = [c.value for c in ov[1]]
    row = dict(zip(headers, [ov.cell(2, c + 1).value for c in range(len(headers))]))
    assert row["Color (CN)"] == _COLOR_NOT_FOUND
    assert row["Color Code"] == _COLOR_NOT_FOUND


def test_internal_db_source_miss_shows_error_not_blank(tmp_path):
    """When the internal DB is the selected source (cn_by_pc_lookup is None)
    and neither cn_lookup nor cn_code_lookup has an entry, that must show as
    an explicit error -- not silent blank/"NA" as if nothing was even tried.
    """
    from po_extractor.exporters.sky_east_buyplan_export import _COLOR_NOT_FOUND

    df = pd.DataFrame([{
        "pc_no": "HHPPC048", "style": "DR9999", "brand": "Anna Field",
        "contract_no": "C1", "article_name": "TEST",
        "zalando_po": "PO001", "config_sku": "C1", "color_name": "chartreuse",
        "xs": 10, "s": 0, "m": 0, "l": 0, "xl": 0, "xxl": 0,
    }])
    path, _totals = export_sky_east_buyplan(
        df, cn_lookup={}, output_dir=str(tmp_path),
        label_lookup={}, cn_code_lookup={},
        cn_by_pc_lookup=None,   # internal DB is the selected source
    )
    ov = load_workbook(path)["Overview"]
    headers = [c.value for c in ov[1]]
    row = dict(zip(headers, [ov.cell(2, c + 1).value for c in range(len(headers))]))
    assert row["Color (CN)"] == _COLOR_NOT_FOUND
    assert row["Color Code"] == _COLOR_NOT_FOUND


def test_partial_match_still_uses_na_placeholder_not_error(tmp_path):
    """A genuine match whose colour-code field happens to be blank in the
    source row is a different situation from a full miss -- it must keep
    showing the existing "NA" placeholder, not the new error marker.
    """
    from po_extractor.lookups.progress_lookup import PCColorMatch

    df = pd.DataFrame([{
        "pc_no": "HHPPC048", "style": "DR5124", "brand": "Anna Field",
        "contract_no": "26302-ZA7148", "article_name": "LACE DRESS",
        "zalando_po": "PO001", "config_sku": "C1", "color_name": "dark blue",
        "xs": 30, "s": 82, "m": 0, "l": 0, "xl": 0, "xxl": 0,
    }])
    # Matched in 大货进度表, but its colour-code field is blank.
    cn_by_pc = {("HHPPC048", "DR5124", "Dark Blue"): PCColorMatch("藏青", "", "")}
    path, _totals = export_sky_east_buyplan(
        df, cn_lookup={}, output_dir=str(tmp_path),
        label_lookup={}, cn_code_lookup={}, cn_by_pc_lookup=cn_by_pc,
    )
    ov = load_workbook(path)["Overview"]
    headers = [c.value for c in ov[1]]
    row = dict(zip(headers, [ov.cell(2, c + 1).value for c in range(len(headers))]))
    assert row["Color (CN)"] == "藏青"
    # Overview blanks the internal "NA" sentinel -- openpyxl round-trips an
    # empty-string cell write as None on reload, which is the correct
    # "blank cell" outcome in Excel.
    assert not row["Color Code"]


# ---------------------------------------------------------------------------
# Sheet naming — "<running index>_<style>", no fabric code
# ---------------------------------------------------------------------------

def test_sheet_names_are_index_then_style_no_fabric_code(tmp_path):
    """Per-style sheet names are "<n>_<style>" -- a running 1-based index
    matching the Index sheet's own "No." column -- with no fabric/HHN code
    in the name.
    """
    from po_extractor.models.fabric_part import FabricPart

    df = pd.DataFrame([
        {"style": "DR5124", "brand": "Anna Field", "zalando_po": "PO1",
         "config_sku": "C1", "color_name": "dark blue", "contract_no": "K1",
         "article_name": "LACE", "xs": 1, "s": 0, "m": 0, "l": 0, "xl": 0, "xxl": 0},
        {"style": "DR4578", "brand": "Anna Field", "zalando_po": "PO1",
         "config_sku": "C2", "color_name": "green", "contract_no": "K2",
         "article_name": "SLEEVE", "xs": 1, "s": 0, "m": 0, "l": 0, "xl": 0, "xxl": 0},
    ])
    fabric_parts_by_style = {
        "DR5124": [FabricPart(combo_idx=0, seq=1, body_part="Main Body", hhn_no="HHN-JA-01715")],
        "DR4578": [FabricPart(combo_idx=0, seq=1, body_part="Main Body", hhn_no="HHN-JSRSR-04068")],
    }
    path, _totals = export_sky_east_buyplan(
        df, cn_lookup={}, output_dir=str(tmp_path), label_lookup={}, cn_code_lookup={},
        fabric_parts_by_style=fabric_parts_by_style,
    )
    wb = load_workbook(path)
    _summary_tabs = {"Index", "Overview", "Fabric Version"}
    style_sheets = [s for s in wb.sheetnames if s not in _summary_tabs]
    assert style_sheets == ["1_DR5124", "2_DR4578"]
    assert not any("HHN" in s for s in style_sheets)


def test_sheet_names_disambiguate_multiple_combos_of_same_style(tmp_path):
    """A style with two fabric combos gets two sheets, disambiguated purely
    by the running index (e.g. "1_DR5009", "2_DR5009") -- no fabric code
    needed since the index itself is already unique.
    """
    from po_extractor.models.fabric_part import FabricPart

    df = pd.DataFrame([{
        "style": "DR5009", "brand": "Anna Field", "zalando_po": "PO1",
        "config_sku": "C1", "color_name": "dark blue", "contract_no": "K1",
        "article_name": "MAXI", "xs": 1, "s": 0, "m": 0, "l": 0, "xl": 0, "xxl": 0,
    }])
    fabric_parts_by_style = {
        "DR5009": [
            FabricPart(combo_idx=0, seq=1, body_part="Main Body", hhn_no="HHN-A"),
            FabricPart(combo_idx=1, seq=1, body_part="Main Body", hhn_no="HHN-B"),
        ],
    }
    path, _totals = export_sky_east_buyplan(
        df, cn_lookup={}, output_dir=str(tmp_path), label_lookup={}, cn_code_lookup={},
        fabric_parts_by_style=fabric_parts_by_style,
    )
    wb = load_workbook(path)
    _summary_tabs = {"Index", "Overview", "Fabric Version"}
    style_sheets = [s for s in wb.sheetnames if s not in _summary_tabs]
    assert style_sheets == ["1_DR5009", "2_DR5009"]


def test_multi_combo_index_links_and_totals_resolve(tmp_path):
    """A style with two fabric combos yields two Index rows; each row's style
    hyperlink is WPS-safe (internal, ``target is None``) and points at a real,
    distinct sheet, and the same row's 订单数合计 formula references that sheet.
    """
    from po_extractor.models.fabric_part import FabricPart

    df = pd.DataFrame([{
        "style": "DR5009", "brand": "Anna Field", "zalando_po": "PO1",
        "config_sku": "C1", "color_name": "dark blue", "contract_no": "K1",
        "article_name": "MAXI", "xs": 1, "s": 0, "m": 0, "l": 0, "xl": 0, "xxl": 0,
    }])
    fabric_parts_by_style = {
        "DR5009": [
            FabricPart(combo_idx=0, seq=1, body_part="Main Body", hhn_no="HHN-A"),
            FabricPart(combo_idx=1, seq=1, body_part="Main Body", hhn_no="HHN-B"),
        ],
    }
    path, _ = export_sky_east_buyplan(
        df, cn_lookup={}, output_dir=str(tmp_path), label_lookup={}, cn_code_lookup={},
        fabric_parts_by_style=fabric_parts_by_style, sky_east_store=None,
    )
    wb = load_workbook(path)
    idx = wb["Index"]
    linked_rows = [idx.cell(r, 2) for r in range(2, idx.max_row + 1)
                   if idx.cell(r, 2).hyperlink is not None]
    assert len(linked_rows) == 2               # one Index row per combo sheet

    sheets_pointed_at = set()
    for cell in linked_rows:
        assert cell.hyperlink.target is None                 # internal (WPS-safe)
        sheet = cell.hyperlink.location.split("!")[0].strip("'")
        assert sheet in wb.sheetnames                        # points at a real sheet
        sheets_pointed_at.add(sheet)
        # the 订单数合计 total formula on the same row references the same sheet
        row_formulas = [c.value for c in idx[cell.row]
                        if isinstance(c.value, str) and c.value.startswith("=")]
        assert any(f"'{sheet}'!" in f for f in row_formulas)
    assert sheets_pointed_at == {"1_DR5009", "2_DR5009"}      # two distinct sheets


# ---------------------------------------------------------------------------
# Order-file multi-colour resolution — "(dark blue)(white)" style cells
# ---------------------------------------------------------------------------

def test_resolve_pc_color_multi_passes_through_single_color_unchanged():
    """A plain single colour (the common case, no ' / ' separator) must
    resolve identically through the multi-aware wrapper and the original
    single-colour resolver.
    """
    from po_extractor.exporters.sky_east_buyplan_export import (
        _resolve_pc_color, _resolve_pc_color_multi,
    )
    from po_extractor.lookups.progress_lookup import PCColorMatch

    cn_by_pc = {("HHPPC048", "DR5124", "Dark Blue"): PCColorMatch("藏青", "503", "")}
    row = {"pc_no": "HHPPC048"}
    expected = _resolve_pc_color(row, "DR5124", "Dark Blue", "Anna Field", {}, {}, cn_by_pc)
    actual   = _resolve_pc_color_multi(row, "DR5124", "Dark Blue", "Anna Field", {}, {}, cn_by_pc)
    assert actual == expected
    assert actual[0] == "藏青"


def test_resolve_pc_color_multi_uses_first_matching_component():
    from po_extractor.exporters.sky_east_buyplan_export import _resolve_pc_color_multi
    from po_extractor.lookups.progress_lookup import PCColorMatch

    cn_by_pc = {("HHPPC048", "DR5124", "Dark Blue"): PCColorMatch("藏青", "503", "")}
    row = {"pc_no": "HHPPC048"}
    result = _resolve_pc_color_multi(
        row, "DR5124", "Dark Blue / White", "Anna Field", {}, {}, cn_by_pc,
    )
    assert result[0] == "藏青"
    assert result[1] == "503"


def test_resolve_pc_color_multi_falls_back_to_second_component():
    """First component ("Dark Blue") isn't in the lookup, but the second
    ("White") is -- detect-and-separate must try each component in turn
    rather than giving up after the first miss.
    """
    from po_extractor.exporters.sky_east_buyplan_export import _resolve_pc_color_multi
    from po_extractor.lookups.progress_lookup import PCColorMatch

    cn_by_pc = {("HHPPC048", "DR5124", "White"): PCColorMatch("白色", "003", "")}
    row = {"pc_no": "HHPPC048"}
    result = _resolve_pc_color_multi(
        row, "DR5124", "Dark Blue / White", "Anna Field", {}, {}, cn_by_pc,
    )
    assert result[0] == "白色"
    assert result[1] == "003"


def test_resolve_pc_color_multi_neither_component_matches_shows_not_found():
    from po_extractor.exporters.sky_east_buyplan_export import (
        _COLOR_NOT_FOUND, _resolve_pc_color_multi,
    )

    row = {"pc_no": "HHPPC048"}
    result = _resolve_pc_color_multi(
        row, "DR5124", "Dark Blue / White", "Anna Field", {}, {}, {},
    )
    assert result[0] == _COLOR_NOT_FOUND
    assert result[1] == _COLOR_NOT_FOUND


def test_order_file_concatenated_bracket_colors_resolve_via_second_component(tmp_path):
    """End-to-end: an order-file colour cell like "(dark blue)(white)" is
    stripped to "Dark Blue / White" and, since only "White" is in the
    selected 大货进度表 source, the Overview sheet must show White's
    translation rather than an outright miss.
    """
    from po_extractor.lookups.progress_lookup import PCColorMatch

    df = pd.DataFrame([{
        "pc_no": "HHPPC048", "style": "DR5124", "brand": "Anna Field",
        "contract_no": "26302-ZA7148", "article_name": "TWO TONE DRESS",
        "zalando_po": "PO001", "config_sku": "C1", "color_name": "(dark blue)(white)",
        "xs": 30, "s": 82, "m": 0, "l": 0, "xl": 0, "xxl": 0,
    }])
    cn_by_pc = {("HHPPC048", "DR5124", "White"): PCColorMatch("白色", "003", "")}
    path, _totals = export_sky_east_buyplan(
        df, cn_lookup={}, output_dir=str(tmp_path),
        label_lookup={}, cn_code_lookup={}, cn_by_pc_lookup=cn_by_pc,
    )
    ov = load_workbook(path)["Overview"]
    headers = [c.value for c in ov[1]]
    row = dict(zip(headers, [ov.cell(2, c + 1).value for c in range(len(headers))]))
    # Shown verbatim as the client wrote it; the resolved CN/code below are
    # what prove the stripped "Dark Blue / White" form still drove matching.
    assert row["Color (EN)"] == "(dark blue)(white)"
    assert row["Color (CN)"] == "白色"
    assert row["Color Code"] == "003"


# ---------------------------------------------------------------------------
# "Local + AI Enhance" — API is only ever a last resort after a local miss
# ---------------------------------------------------------------------------

def test_ai_enhance_not_called_when_disabled(monkeypatch):
    """ai_enhance=False (the default) must never reach the API, even on a
    total local miss -- prevents any surprise token spend for users who
    haven't opted in.
    """
    import po_extractor.lookups.color_ai_enhance as _ai
    from po_extractor.exporters.sky_east_buyplan_export import (
        _COLOR_NOT_FOUND, _resolve_pc_color_multi,
    )

    def _boom(*a, **k):
        raise AssertionError("recognize_colors must not be called when ai_enhance is False")

    monkeypatch.setattr(_ai, "recognize_colors", _boom)

    row = {"pc_no": "HHPPC048"}
    result = _resolve_pc_color_multi(
        row, "DR5124", "Mystery Hue", "Anna Field", {}, {}, {},
        ai_enhance=False, ai_api_key="sk-fake",
    )
    assert result[0] == _COLOR_NOT_FOUND


def test_ai_enhance_not_called_when_local_component_already_resolves(monkeypatch):
    """A colour that already resolves locally must never trigger the API --
    the whole point of "only on unresolved colour" is that a working match
    costs zero tokens.
    """
    import po_extractor.lookups.color_ai_enhance as _ai
    from po_extractor.exporters.sky_east_buyplan_export import _resolve_pc_color_multi
    from po_extractor.lookups.progress_lookup import PCColorMatch

    def _boom(*a, **k):
        raise AssertionError("recognize_colors must not be called on a local hit")

    monkeypatch.setattr(_ai, "recognize_colors", _boom)

    cn_by_pc = {("HHPPC048", "DR5124", "Dark Blue"): PCColorMatch("藏青", "503", "")}
    row = {"pc_no": "HHPPC048"}
    result = _resolve_pc_color_multi(
        row, "DR5124", "Dark Blue", "Anna Field", {}, {}, cn_by_pc,
        ai_enhance=True, ai_api_key="sk-fake",
    )
    assert result[0] == "藏青"


def test_ai_enhance_used_only_after_local_miss(monkeypatch):
    """When every local component misses and ai_enhance is on, the raw
    (unsplit) colour string is sent to recognize_colors(), and its returned
    candidates are retried against the same local lookup.
    """
    import po_extractor.lookups.color_ai_enhance as _ai
    from po_extractor.exporters.sky_east_buyplan_export import _resolve_pc_color_multi
    from po_extractor.lookups.progress_lookup import PCColorMatch

    calls = []

    def _fake_recognize(raw_color, api_key, model="deepseek-chat"):
        calls.append((raw_color, api_key, model))
        return ("Navy",)

    monkeypatch.setattr(_ai, "recognize_colors", _fake_recognize)

    cn_by_pc = {("HHPPC048", "DR5124", "Navy"): PCColorMatch("藏青", "503", "")}
    row = {"pc_no": "HHPPC048"}
    result = _resolve_pc_color_multi(
        row, "DR5124", "Chocolate Brown With Navy Trim", "Anna Field", {}, {}, cn_by_pc,
        ai_enhance=True, ai_api_key="sk-fake", ai_model="deepseek-chat",
    )
    assert result[0] == "藏青"
    assert result[1] == "503"
    assert calls == [("Chocolate Brown With Navy Trim", "sk-fake", "deepseek-chat")]


def test_ai_enhance_falls_back_to_local_miss_when_api_also_fails(monkeypatch):
    """If the API returns no usable candidates, the original not-found
    result is kept -- never crash, never fabricate a match.
    """
    import po_extractor.lookups.color_ai_enhance as _ai
    from po_extractor.exporters.sky_east_buyplan_export import (
        _COLOR_NOT_FOUND, _resolve_pc_color_multi,
    )

    monkeypatch.setattr(_ai, "recognize_colors", lambda *a, **k: ())

    row = {"pc_no": "HHPPC048"}
    result = _resolve_pc_color_multi(
        row, "DR5124", "Unmappable Hue", "Anna Field", {}, {}, {},
        ai_enhance=True, ai_api_key="sk-fake",
    )
    assert result[0] == _COLOR_NOT_FOUND


# ---------------------------------------------------------------------------
# Both components resolving — Chinese name / code combine, not just the first
# ---------------------------------------------------------------------------

def test_resolve_pc_color_multi_combines_both_components_when_both_resolve():
    """A two-tone colour whose BOTH components have a translation must show
    both in Color (CN) / Color Code, e.g. "藏青 / 白色" and "52# / 3#" --
    not just whichever component happened to resolve first.
    """
    from po_extractor.exporters.sky_east_buyplan_export import _resolve_pc_color_multi
    from po_extractor.lookups.progress_lookup import PCColorMatch

    cn_by_pc = {
        ("HHPPC048", "DR5009", "Dark Blue"): PCColorMatch("藏青", "52#", ""),
        ("HHPPC048", "DR5009", "White"):     PCColorMatch("白色", "3#", ""),
    }
    row = {"pc_no": "HHPPC048"}
    result = _resolve_pc_color_multi(
        row, "DR5009", "Dark Blue / White", "Anna Field", {}, {}, cn_by_pc,
    )
    assert result[0] == "藏青 / 白色"
    assert result[1] == "52# / 3#"


def test_resolve_pc_color_multi_combines_label_color_from_resolved_component():
    """label_color (主标颜色) isn't itself split/combined -- it's a single
    physical label -- but must still be carried through from whichever
    component actually resolved it.
    """
    from po_extractor.exporters.sky_east_buyplan_export import _resolve_pc_color_multi
    from po_extractor.lookups.progress_lookup import PCColorMatch

    cn_by_pc = {
        ("HHPPC048", "DR5009", "Dark Blue"): PCColorMatch("藏青", "52#", "黑色"),
        ("HHPPC048", "DR5009", "White"):     PCColorMatch("白色", "3#", ""),
    }
    row = {"pc_no": "HHPPC048"}
    result = _resolve_pc_color_multi(
        row, "DR5009", "Dark Blue / White", "Anna Field", {}, {}, cn_by_pc,
    )
    assert result[2] == "黑色"


def test_ai_enhance_recovers_the_missing_component_in_a_two_tone_pair(monkeypatch):
    """When one of two components misses locally but AI enhance recognises
    it, the recovered component's translation must still be combined with
    the component that already resolved locally -- not discarded.
    """
    import po_extractor.lookups.color_ai_enhance as _ai
    from po_extractor.exporters.sky_east_buyplan_export import _resolve_pc_color_multi
    from po_extractor.lookups.progress_lookup import PCColorMatch

    def _fake_recognize(raw_color, api_key, model="deepseek-chat"):
        assert raw_color == "White"   # only the missing component is sent
        return ("Ivory",)

    monkeypatch.setattr(_ai, "recognize_colors", _fake_recognize)

    cn_by_pc = {
        ("HHPPC048", "DR5009", "Dark Blue"): PCColorMatch("藏青", "52#", ""),
        ("HHPPC048", "DR5009", "Ivory"):     PCColorMatch("米白", "3#", ""),
    }
    row = {"pc_no": "HHPPC048"}
    result = _resolve_pc_color_multi(
        row, "DR5009", "Dark Blue / White", "Anna Field", {}, {}, cn_by_pc,
        ai_enhance=True, ai_api_key="sk-fake",
    )
    assert result[0] == "藏青 / 米白"
    assert result[1] == "52# / 3#"


# ── AI colour prefetch (v2.75.2 perf fix) ────────────────────────────────────
# "Local + AI Enhance" used to make one BLOCKING network call per unresolved
# colour INLINE inside the serial row-writing loop -- for a run with a dozen
# such colours (~1s+ each), that dominated total export time. These tests lock
# _prefetch_ai_color_cache's own logic: it must skip colours that already
# resolve locally, dedupe repeated raw strings across many rows into ONE job
# each, and report progress as jobs complete -- all BEFORE the row-writing
# loop runs, off the network-call critical path.

# ── Skip AI entirely for a row that isn't a real item (v2.75.3 fix) ──────────
# Production data had a row whose every field was literally the source
# file's own column-header text (pc_no a real PC No., but style="Style No.",
# color_name="Color name", every size column 0) -- a template row that leaked
# into the items table. Its "colour" was never a colour, so AI must never be
# asked to guess one for it.

def test_row_has_real_quantity_true_when_any_size_nonzero():
    from po_extractor.exporters.sky_east_buyplan_export import _row_has_real_quantity
    assert _row_has_real_quantity(
        {"xs": 0, "s": 5, "m": 0, "l": 0, "xl": 0, "xxl": 0})
    assert _row_has_real_quantity(
        {"xs": 0, "s": 0, "m": 0, "l": 0, "xl": 0, "xxl": 1})


def test_row_has_real_quantity_false_only_when_every_size_col_present_and_zero():
    from po_extractor.exporters.sky_east_buyplan_export import _row_has_real_quantity
    assert not _row_has_real_quantity(
        {"xs": 0, "s": 0, "m": 0, "l": 0, "xl": 0, "xxl": 0})


def test_row_has_real_quantity_true_when_size_cols_simply_absent():
    """No size-column info at all must NOT be treated as confirmed-empty --
    many callers (and existing tests) pass a minimal row without size data,
    and AI must not be silently blocked for them."""
    from po_extractor.exporters.sky_east_buyplan_export import _row_has_real_quantity
    assert _row_has_real_quantity({})
    assert _row_has_real_quantity({"pc_no": "HHPPC048"})


def test_ai_prefetch_skips_zero_quantity_artifact_row(monkeypatch):
    """A row with pc_no/style but zero quantity in every size must never
    queue an AI job, even when its colour would otherwise miss locally."""
    import po_extractor.lookups.color_ai_enhance as _ai
    from po_extractor.exporters.sky_east_buyplan_export import _prefetch_ai_color_cache

    def _boom(*a, **k):
        raise AssertionError("AI must not be called for a zero-quantity artifact row")
    monkeypatch.setattr(_ai, "recognize_colors", _boom)
    monkeypatch.setattr(_ai, "match_color_to_candidates", _boom)

    df = pd.DataFrame([{
        "pc_no": "HHPPC046", "style": "Style No.", "color_name": "Color name",
        "brand": "Brand", "xs": 0, "s": 0, "m": 0, "l": 0, "xl": 0, "xxl": 0,
    }])
    _prefetch_ai_color_cache(df, {}, {}, None, True, "sk-fake", "deepseek-chat")


def test_resolve_pc_color_multi_never_calls_ai_for_zero_quantity_row(monkeypatch):
    """Same guard at the real per-row resolution path (_resolve_pc_color_multi),
    not just the prefetch pass -- covers a run without the prefetch too."""
    import po_extractor.lookups.color_ai_enhance as _ai
    from po_extractor.exporters.sky_east_buyplan_export import (
        _COLOR_NOT_FOUND, _resolve_pc_color_multi,
    )

    def _boom(*a, **k):
        raise AssertionError("AI must not be called for a zero-quantity artifact row")
    monkeypatch.setattr(_ai, "recognize_colors", _boom)

    row = {"pc_no": "HHPPC046", "xs": 0, "s": 0, "m": 0, "l": 0, "xl": 0, "xxl": 0}
    result = _resolve_pc_color_multi(
        row, "STYLENO", "Color Name", "Brand", {}, {}, None,
        ai_enhance=True, ai_api_key="sk-fake",
    )
    assert result[0] == _COLOR_NOT_FOUND   # stays a miss -- never fabricated


def test_ai_prefetch_noop_without_enhance_or_key():
    from po_extractor.exporters.sky_east_buyplan_export import _prefetch_ai_color_cache

    df = pd.DataFrame([{"style": "DR1", "color_name": "Fuchsia", "brand": "", "pc_no": ""}])
    calls = []
    # ai_enhance=False -- must return immediately, no lookup work at all.
    _prefetch_ai_color_cache(df, {}, {}, None, False, "sk-fake", "deepseek-chat")
    # ai_enhance=True but no key -- same.
    _prefetch_ai_color_cache(df, {}, {}, None, True, "", "deepseek-chat")
    assert calls == []   # nothing to assert on calls, but neither call should raise


def test_ai_prefetch_skips_colours_that_resolve_locally(monkeypatch):
    """A colour present in cn_lookup must never trigger an AI job."""
    import po_extractor.lookups.color_ai_enhance as _ai
    from po_extractor.exporters.sky_east_buyplan_export import _prefetch_ai_color_cache

    def _boom(*a, **k):
        raise AssertionError("AI must not be called for a colour that resolves locally")
    monkeypatch.setattr(_ai, "recognize_colors", _boom)
    monkeypatch.setattr(_ai, "match_color_to_candidates", _boom)

    cn_lookup = {("Sky East", "Anna Field", "NAVY"): "藏青"}
    df = pd.DataFrame([
        {"style": "DR1", "color_name": "Navy", "brand": "Anna Field", "pc_no": "P1"},
    ])
    _prefetch_ai_color_cache(df, cn_lookup, {}, None, True, "sk-fake", "deepseek-chat")


def test_ai_prefetch_dedupes_repeated_colour_across_rows(monkeypatch):
    """The same unresolved raw colour string appearing on many rows must
    trigger exactly ONE recognize_colors job, not one per row."""
    import po_extractor.lookups.color_ai_enhance as _ai
    from po_extractor.exporters.sky_east_buyplan_export import _prefetch_ai_color_cache

    calls = []
    def _fake_recognize(raw, key, model="deepseek-chat"):
        calls.append(raw)
        return ("Navy",)
    monkeypatch.setattr(_ai, "recognize_colors", _fake_recognize)

    df = pd.DataFrame([
        {"style": "DR1", "color_name": "Unresolvable Color Xyz", "brand": "Anna Field", "pc_no": "P1"},
        {"style": "DR2", "color_name": "Unresolvable Color Xyz", "brand": "Anna Field", "pc_no": "P1"},
        {"style": "DR3", "color_name": "Unresolvable Color Xyz", "brand": "Anna Field", "pc_no": "P1"},
    ])
    _prefetch_ai_color_cache(df, {}, {}, None, True, "sk-fake", "deepseek-chat")
    assert calls == ["Unresolvable Color Xyz"]   # one job, not three


def test_ai_prefetch_progress_callback_counts_to_total(monkeypatch):
    """on_progress(done, total) must be called once before any job starts
    (done=0) and once per completed job, ending at done==total."""
    import po_extractor.lookups.color_ai_enhance as _ai
    from po_extractor.exporters.sky_east_buyplan_export import _prefetch_ai_color_cache

    monkeypatch.setattr(_ai, "recognize_colors", lambda raw, key, model="deepseek-chat": ("X",))

    df = pd.DataFrame([
        {"style": "DR1", "color_name": "__TestAI_A__", "brand": "", "pc_no": ""},
        {"style": "DR2", "color_name": "__TestAI_B__", "brand": "", "pc_no": ""},
        {"style": "DR3", "color_name": "__TestAI_C__", "brand": "", "pc_no": ""},
    ])
    ticks = []
    _prefetch_ai_color_cache(df, {}, {}, None, True, "sk-fake", "deepseek-chat",
                             on_progress=lambda d, t: ticks.append((d, t)))
    assert ticks[0] == (0, 3)
    assert ticks[-1] == (3, 3)
    assert [d for d, _ in ticks] == sorted(d for d, _ in ticks)   # monotonic
    assert all(t == 3 for _, t in ticks)


def test_export_sky_east_buyplan_on_progress_reaches_zero_and_one():
    """End-to-end (ai_enhance off, so no network calls, keeps the test fast):
    on_progress must fire with 0.0 first and 1.0 last, strictly increasing
    in between, covering the whole export call."""
    import tempfile
    from po_extractor.exporters.sky_east_buyplan_export import export_sky_east_buyplan

    df = pd.DataFrame([
        {"pc_no": "P1", "style": "DR1", "contract_no": "C1", "brand": "Anna Field",
         "article_name": "DRESS", "zalando_po": "PO1", "config_sku": "K1",
         "color_name": "Navy", "colour_code": "",
         "xs": 1, "s": 1, "m": 1, "l": 1, "xl": 1, "xxl": 0,
         "fabric_item_no": "", "fabrication": ""},
        {"pc_no": "P1", "style": "DR2", "contract_no": "C1", "brand": "Anna Field",
         "article_name": "DRESS", "zalando_po": "PO1", "config_sku": "K1",
         "color_name": "Black", "colour_code": "",
         "xs": 1, "s": 1, "m": 1, "l": 1, "xl": 1, "xxl": 0,
         "fabric_item_no": "", "fabrication": ""},
    ])
    ticks = []
    with tempfile.TemporaryDirectory() as out_dir:
        export_sky_east_buyplan(
            df, {}, out_dir,
            sky_east_store=None,
            on_progress=lambda f, l: ticks.append(f),
        )
    assert ticks[0] == 0.0
    assert ticks[-1] == 1.0
    assert ticks == sorted(ticks)   # monotonically non-decreasing
    assert all(0.0 <= f <= 1.0 for f in ticks)


def test_order_file_two_tone_shows_combined_chinese_names_end_to_end(tmp_path):
    """Overview sheet for a two-tone order-file colour whose both components
    resolve must show both Chinese names combined, not just one.
    """
    from po_extractor.lookups.progress_lookup import PCColorMatch

    df = pd.DataFrame([{
        "pc_no": "HHPPC048", "style": "DR5009", "brand": "Anna Field",
        "contract_no": "26302-ZA7156", "article_name": "MAXI DRESS",
        "zalando_po": "PO001", "config_sku": "C1", "color_name": "(dark blue)(white)",
        "xs": 22, "s": 61, "m": 90, "l": 77, "xl": 50, "xxl": 0,
    }])
    cn_by_pc = {
        ("HHPPC048", "DR5009", "Dark Blue"): PCColorMatch("藏青", "52#", ""),
        ("HHPPC048", "DR5009", "White"):     PCColorMatch("白色", "3#", ""),
    }
    path, _totals = export_sky_east_buyplan(
        df, cn_lookup={}, output_dir=str(tmp_path),
        label_lookup={}, cn_code_lookup={}, cn_by_pc_lookup=cn_by_pc,
    )
    ov = load_workbook(path)["Overview"]
    headers = [c.value for c in ov[1]]
    row = dict(zip(headers, [ov.cell(2, c + 1).value for c in range(len(headers))]))
    assert row["Color (EN)"] == "(dark blue)(white)"
    assert row["Color (CN)"] == "藏青 / 白色"
    assert row["Color Code"] == "52# / 3#"


# ---------------------------------------------------------------------------
# Overview sheet — 客人PC NO and 主标颜色 columns
# ---------------------------------------------------------------------------

def test_overview_includes_pc_no_next_to_contract_no_and_label_color(tmp_path):
    from po_extractor.lookups.progress_lookup import PCColorMatch

    df = pd.DataFrame([{
        "pc_no": "HHPPC048", "style": "DR5124", "brand": "Anna Field",
        "contract_no": "26302-ZA7148", "article_name": "LACE DRESS",
        "zalando_po": "PO001", "config_sku": "C1", "color_name": "dark blue",
        "xs": 30, "s": 82, "m": 0, "l": 0, "xl": 0, "xxl": 0,
    }])
    cn_by_pc = {("HHPPC048", "DR5124", "Dark Blue"): PCColorMatch("藏青", "503", "黑色")}
    path, _totals = export_sky_east_buyplan(
        df, cn_lookup={}, output_dir=str(tmp_path),
        label_lookup={}, cn_code_lookup={}, cn_by_pc_lookup=cn_by_pc,
    )
    ov = load_workbook(path)["Overview"]
    headers = [c.value for c in ov[1]]
    # 客人PC NO sits immediately after Contract No.
    assert headers[headers.index("Contract No.") + 1] == "客人PC NO"
    row = dict(zip(headers, [ov.cell(2, c + 1).value for c in range(len(headers))]))
    assert row["客人PC NO"] == "HHPPC048"


# ---------------------------------------------------------------------------
# Colour-miss diagnostics — Excel comment + Sky East colour-miss log
# ---------------------------------------------------------------------------

def test_not_found_cell_gets_comment_with_clients_po_color(tmp_path):
    """A 未找到 cell in the Overview sheet must carry a comment showing the
    client's PO colour text -- e.g. it can differ from Color (EN) if the
    combined-string display omits some detail from the raw order file.
    """
    df = pd.DataFrame([{
        "pc_no": "HHPPC048", "style": "BL4257", "brand": "Anna Field",
        "contract_no": "26302-ZA7158", "article_name": "LONG SLEEVE BLOUSE",
        "zalando_po": "PO2338263C", "config_sku": "C1", "color_name": "Dark Brown",
        "xs": 28, "s": 69, "m": 90, "l": 67, "xl": 46, "xxl": 0,
    }])
    path, _totals = export_sky_east_buyplan(
        df, cn_lookup={}, output_dir=str(tmp_path),
        label_lookup={}, cn_code_lookup={}, cn_by_pc_lookup={},
        sky_east_store=None,
    )
    ov = load_workbook(path)["Overview"]
    headers = [c.value for c in ov[1]]
    cn_col   = headers.index("Color (CN)") + 1
    code_col = headers.index("Color Code") + 1
    cn_cell   = ov.cell(2, cn_col)
    code_cell = ov.cell(2, code_col)
    assert cn_cell.value == "未找到"
    assert cn_cell.comment is not None
    assert "Dark Brown" in cn_cell.comment.text
    assert code_cell.comment is not None
    assert "Dark Brown" in code_cell.comment.text


def test_found_cell_has_no_comment(tmp_path):
    """A successfully-resolved colour must not carry a stray comment."""
    from po_extractor.lookups.progress_lookup import PCColorMatch

    df = pd.DataFrame([{
        "pc_no": "HHPPC048", "style": "DR5124", "brand": "Anna Field",
        "contract_no": "26302-ZA7148", "article_name": "LACE DRESS",
        "zalando_po": "PO001", "config_sku": "C1", "color_name": "dark blue",
        "xs": 30, "s": 82, "m": 0, "l": 0, "xl": 0, "xxl": 0,
    }])
    cn_by_pc = {("HHPPC048", "DR5124", "Dark Blue"): PCColorMatch("藏青", "503", "")}
    path, _totals = export_sky_east_buyplan(
        df, cn_lookup={}, output_dir=str(tmp_path),
        label_lookup={}, cn_code_lookup={}, cn_by_pc_lookup=cn_by_pc,
        sky_east_store=None,
    )
    ov = load_workbook(path)["Overview"]
    headers = [c.value for c in ov[1]]
    cn_cell = ov.cell(2, headers.index("Color (CN)") + 1)
    assert cn_cell.comment is None


def test_color_miss_is_logged_to_sky_east_store(_isolated_sky_east_store, tmp_path):
    """A colour that ends up 未找到 must be logged via
    SkyEastStore.log_color_miss() with the client's raw PO colour text --
    this is what powers the diagnostic log, separate from the Excel comment.
    """
    df = pd.DataFrame([{
        "pc_no": "HHPPC048", "style": "BL4257", "brand": "Anna Field",
        "contract_no": "26302-ZA7158", "article_name": "LONG SLEEVE BLOUSE",
        "zalando_po": "PO2338263C", "config_sku": "C1", "color_name": "Dark Brown",
        "xs": 28, "s": 69, "m": 90, "l": 67, "xl": 46, "xxl": 0,
    }])
    export_sky_east_buyplan(
        df, cn_lookup={}, output_dir=str(tmp_path),
        label_lookup={}, cn_code_lookup={}, cn_by_pc_lookup={},
    )
    misses = _isolated_sky_east_store.list_color_misses()
    assert len(misses) == 1
    row = misses.iloc[0]
    assert row["pc_no"] == "HHPPC048"
    assert row["style"] == "BL4257"
    assert row["contract_no"] == "26302-ZA7158"
    assert row["po_no"] == "PO2338263C"
    assert row["client_po_color"] == "Dark Brown"
    assert row["source"] == "progress"
    assert row["progress_colors"] == ""   # nothing on file for this PC/style


def test_color_miss_log_captures_progress_colors_on_naming_mismatch(_isolated_sky_east_store, tmp_path):
    """When 大货进度表 DOES have colour(s) on file for this PC/style (just
    under a different English name), the log must capture them too -- the
    same comparison already shown in the Excel cell comment.
    """
    from po_extractor.lookups.progress_lookup import PCColorMatch

    df = pd.DataFrame([{
        "pc_no": "HHPPC048", "style": "BL4257", "brand": "Anna Field",
        "contract_no": "26302-ZA7158", "article_name": "LONG SLEEVE BLOUSE",
        "zalando_po": "PO2338263C", "config_sku": "C1", "color_name": "Dark Brown",
        "xs": 28, "s": 69, "m": 90, "l": 67, "xl": 46, "xxl": 0,
    }])
    cn_by_pc = {("HHPPC048", "BL4257", "Chocolate"): PCColorMatch("巧克力色", "12#", "")}
    export_sky_east_buyplan(
        df, cn_lookup={}, output_dir=str(tmp_path),
        label_lookup={}, cn_code_lookup={}, cn_by_pc_lookup=cn_by_pc,
    )
    row = _isolated_sky_east_store.list_color_misses().iloc[0]
    assert row["progress_colors"] == "Chocolate"


def test_resolved_color_is_not_logged_as_a_miss(_isolated_sky_east_store, tmp_path):
    from po_extractor.lookups.progress_lookup import PCColorMatch

    df = pd.DataFrame([{
        "pc_no": "HHPPC048", "style": "DR5124", "brand": "Anna Field",
        "contract_no": "26302-ZA7148", "article_name": "LACE DRESS",
        "zalando_po": "PO001", "config_sku": "C1", "color_name": "dark blue",
        "xs": 30, "s": 82, "m": 0, "l": 0, "xl": 0, "xxl": 0,
    }])
    cn_by_pc = {("HHPPC048", "DR5124", "Dark Blue"): PCColorMatch("藏青", "503", "")}
    export_sky_east_buyplan(
        df, cn_lookup={}, output_dir=str(tmp_path),
        label_lookup={}, cn_code_lookup={}, cn_by_pc_lookup=cn_by_pc,
    )
    assert _isolated_sky_east_store.list_color_misses().empty


def test_sky_east_store_none_disables_logging_without_breaking_export(tmp_path):
    """Passing sky_east_store=None must still produce a correct export --
    logging is purely diagnostic and never required for the export itself.
    """
    df = pd.DataFrame([{
        "pc_no": "HHPPC048", "style": "BL4257", "brand": "Anna Field",
        "contract_no": "26302-ZA7158", "article_name": "LONG SLEEVE BLOUSE",
        "zalando_po": "PO2338263C", "config_sku": "C1", "color_name": "Dark Brown",
        "xs": 28, "s": 69, "m": 90, "l": 67, "xl": 46, "xxl": 0,
    }])
    path, _totals = export_sky_east_buyplan(
        df, cn_lookup={}, output_dir=str(tmp_path),
        label_lookup={}, cn_code_lookup={}, cn_by_pc_lookup={},
        sky_east_store=None,
    )
    ov = load_workbook(path)["Overview"]
    headers = [c.value for c in ov[1]]
    assert ov.cell(2, headers.index("Color (CN)") + 1).value == "未找到"


# ---------------------------------------------------------------------------
# Not-found comment also lists 大货进度表's own colour(s) for that PC/Style
# ---------------------------------------------------------------------------

def test_color_miss_comment_text_includes_progress_colors_when_present():
    from po_extractor.exporters._sky_east_helpers import _color_miss_comment_text

    text = _color_miss_comment_text("Dark Blue", ["Navy", "Sky Blue"])
    assert 'Client\'s PO colour: "Dark Blue"' in text
    assert "大货进度表 colour(s) on file: Navy, Sky Blue" in text


def test_color_miss_comment_text_notes_no_colors_on_file():
    from po_extractor.exporters._sky_east_helpers import _color_miss_comment_text

    text = _color_miss_comment_text("Dark Blue", [])
    assert "大货进度表 has no colour recorded for this PC No./Style" in text


def test_color_miss_comment_text_omits_second_line_for_internal_db_source():
    """None means the internal DB (not 大货进度表) was the selected source --
    there's no "大货进度表 colour" to show, so the second line is skipped
    entirely rather than showing a misleading empty list.
    """
    from po_extractor.exporters._sky_east_helpers import _color_miss_comment_text

    text = _color_miss_comment_text("Dark Blue", None)
    assert text == 'Client\'s PO colour: "Dark Blue"'


def test_available_progress_colors_scoped_to_pc_and_style():
    from po_extractor.exporters.sky_east_buyplan_export import _available_progress_colors

    cn_by_pc = {
        ("HHPPC048", "DR5124", "Navy"):  object(),
        ("HHPPC048", "DR5124", "Green"): object(),
        ("HHPPC048", "BL4257", "Black"): object(),   # different style — excluded
        ("HHPPC099", "DR5124", "White"): object(),   # different PC — excluded
    }
    result = _available_progress_colors(cn_by_pc, "HHPPC048", "DR5124")
    assert result == ["Green", "Navy"]   # sorted


def test_available_progress_colors_empty_when_no_lookup():
    from po_extractor.exporters.sky_east_buyplan_export import _available_progress_colors

    assert _available_progress_colors(None, "HHPPC048", "DR5124") == []
    assert _available_progress_colors({}, "HHPPC048", "DR5124") == []


def test_not_found_comment_lists_progress_colors_for_naming_mismatch(tmp_path):
    """End-to-end: client's PO says "Dark Brown" but 大货进度表 has this
    style under "Chocolate" -- the comment must show 大货进度表's own colour
    so the mismatch is obvious without opening the source file.
    """
    from po_extractor.lookups.progress_lookup import PCColorMatch

    df = pd.DataFrame([{
        "pc_no": "HHPPC048", "style": "BL4257", "brand": "Anna Field",
        "contract_no": "26302-ZA7158", "article_name": "LONG SLEEVE BLOUSE",
        "zalando_po": "PO2338263C", "config_sku": "C1", "color_name": "Dark Brown",
        "xs": 28, "s": 69, "m": 90, "l": 67, "xl": 46, "xxl": 0,
    }])
    cn_by_pc = {("HHPPC048", "BL4257", "Chocolate"): PCColorMatch("巧克力色", "12#", "")}
    path, _totals = export_sky_east_buyplan(
        df, cn_lookup={}, output_dir=str(tmp_path),
        label_lookup={}, cn_code_lookup={}, cn_by_pc_lookup=cn_by_pc,
        sky_east_store=None,
    )
    ov = load_workbook(path)["Overview"]
    headers = [c.value for c in ov[1]]
    cn_cell = ov.cell(2, headers.index("Color (CN)") + 1)
    assert cn_cell.value == "未找到"
    assert 'Dark Brown' in cn_cell.comment.text
    assert "Chocolate" in cn_cell.comment.text


def test_not_found_comment_omits_progress_line_for_internal_db_source(tmp_path):
    """When the internal DB is the active source (cn_by_pc_lookup=None), the
    comment must not mention 大货进度表 at all -- that source wasn't consulted.
    """
    df = pd.DataFrame([{
        "pc_no": "HHPPC048", "style": "DR9999", "brand": "Anna Field",
        "contract_no": "C1", "article_name": "TEST",
        "zalando_po": "PO001", "config_sku": "C1", "color_name": "chartreuse",
        "xs": 10, "s": 0, "m": 0, "l": 0, "xl": 0, "xxl": 0,
    }])
    path, _totals = export_sky_east_buyplan(
        df, cn_lookup={}, output_dir=str(tmp_path),
        label_lookup={}, cn_code_lookup={},
        cn_by_pc_lookup=None,   # internal DB is the selected source
        sky_east_store=None,
    )
    ov = load_workbook(path)["Overview"]
    headers = [c.value for c in ov[1]]
    cn_cell = ov.cell(2, headers.index("Color (CN)") + 1)
    assert "大货进度表" not in cn_cell.comment.text


# ---------------------------------------------------------------------------
# AI candidate-matching — bridge order-file/大货进度表 colour-name mismatches
# ---------------------------------------------------------------------------

def test_ai_candidate_match_resolves_naming_mismatch(monkeypatch):
    """Order says "Dark Blue"; 大货进度表 has this exact PC+style under
    "Navy" only. The exact colour match misses, but AI candidate-matching
    picks "Navy" from the colours on file and resolves to its Chinese
    name/code -- the values still come from the trusted file.
    """
    import po_extractor.lookups.color_ai_enhance as _ai
    from po_extractor.exporters.sky_east_buyplan_export import _resolve_pc_color_multi
    from po_extractor.lookups.progress_lookup import PCColorMatch

    seen = []

    def _fake_match(client_color, candidates, api_key, model="deepseek-chat"):
        seen.append((client_color, list(candidates)))
        return "Navy" if "Navy" in candidates else ""

    monkeypatch.setattr(_ai, "match_color_to_candidates", _fake_match)
    # recognize_colors must NOT be needed once the candidate match succeeds.
    monkeypatch.setattr(_ai, "recognize_colors",
                        lambda *a, **k: (_ for _ in ()).throw(AssertionError("should not be called")))

    cn_by_pc = {("HHPPC048", "DR5124", "Navy"): PCColorMatch("藏青", "52#", "")}
    row = {"pc_no": "HHPPC048"}
    result = _resolve_pc_color_multi(
        row, "DR5124", "Dark Blue", "Anna Field", {}, {}, cn_by_pc,
        ai_enhance=True, ai_api_key="sk-fake",
    )
    assert result[0] == "藏青"
    assert result[1] == "52#"
    assert seen == [("Dark Blue", ["Navy"])]


def test_ai_candidate_match_not_used_when_disabled(monkeypatch):
    import po_extractor.lookups.color_ai_enhance as _ai
    from po_extractor.exporters.sky_east_buyplan_export import (
        _COLOR_NOT_FOUND, _resolve_pc_color_multi,
    )
    from po_extractor.lookups.progress_lookup import PCColorMatch

    monkeypatch.setattr(_ai, "match_color_to_candidates",
                        lambda *a, **k: (_ for _ in ()).throw(AssertionError("must not call")))

    cn_by_pc = {("HHPPC048", "DR5124", "Navy"): PCColorMatch("藏青", "52#", "")}
    row = {"pc_no": "HHPPC048"}
    result = _resolve_pc_color_multi(
        row, "DR5124", "Dark Blue", "Anna Field", {}, {}, cn_by_pc,
        ai_enhance=False, ai_api_key="sk-fake",
    )
    assert result[0] == _COLOR_NOT_FOUND


def test_ai_candidate_match_end_to_end_overview(tmp_path, monkeypatch):
    """Full export: a bracketed "(dark blue)" order colour, stripped to
    "Dark Blue", resolves via AI candidate-matching against the "Navy" row
    on file and shows the Chinese name/code in the Overview sheet.
    """
    import po_extractor.lookups.color_ai_enhance as _ai
    from po_extractor.lookups.progress_lookup import PCColorMatch

    monkeypatch.setattr(
        _ai, "match_color_to_candidates",
        lambda client_color, candidates, api_key, model="deepseek-chat":
            "Navy" if "Navy" in candidates else "",
    )

    df = pd.DataFrame([{
        "pc_no": "HHPPC048", "style": "DR5124", "brand": "Anna Field",
        "contract_no": "26302-ZA7148", "article_name": "LACE DRESS",
        "zalando_po": "PO2333438C", "config_sku": "C1", "color_name": "(dark blue)",
        "xs": 30, "s": 82, "m": 0, "l": 0, "xl": 0, "xxl": 0,
    }])
    cn_by_pc = {("HHPPC048", "DR5124", "Navy"): PCColorMatch("藏青", "52#", "")}
    path, _totals = export_sky_east_buyplan(
        df, cn_lookup={}, output_dir=str(tmp_path),
        label_lookup={}, cn_code_lookup={}, cn_by_pc_lookup=cn_by_pc,
        ai_enhance=True, ai_api_key="sk-fake",
        sky_east_store=None,
    )
    ov = load_workbook(path)["Overview"]
    headers = [c.value for c in ov[1]]
    row = dict(zip(headers, [ov.cell(2, c + 1).value for c in range(len(headers))]))
    assert row["Color (CN)"] == "藏青"
    assert row["Color Code"] == "52#"
