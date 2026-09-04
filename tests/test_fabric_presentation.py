"""Tests for the fabric presentation module (面料推荐单).

Covers the three things that would be expensive to get wrong:
  * the quoted price matches the workbook this feature replaces, exactly
  * the internal RMB cost cannot leak into a customer-facing export
  * a QR scan is recorded against the right sheet
"""
from __future__ import annotations

import io

import openpyxl
import pytest

from po_extractor.exporters.fabric_presentation_export import (
    PRICE_BOTH, PRICE_RMB, PRICE_USD, build_presentation_workbook,
)
from po_extractor.store.fabric_presentation_store import FabricPresentationStore
from po_extractor.utils.fabric_quote import usd_per_yard


# Every (RMB/M, USD/Y) pair from the real HHN Presentation GIII 5.13 workbook.
# These are the numbers a customer was actually quoted, so the formula must
# keep reproducing them exactly.
_TEMPLATE_PRICES = [
    (12.16, 1.85), (16.37, 2.50), (17.33, 2.65), (10.03, 1.55),
    (13.65, 2.05), (15.86, 2.40), (16.32, 2.50), (12.50, 1.90),
    (11.80, 1.80), (10.50, 1.60), (9.50, 1.45), (9.80, 1.50),
    (19.50, 2.95), (27.00, 4.10), (23.00, 3.50), (17.80, 2.70),
    (16.00, 2.45), (17.60, 2.65),
]


@pytest.fixture()
def store(tmp_path):
    return FabricPresentationStore(str(tmp_path / "pres.db"))


@pytest.fixture()
def sample_lines():
    return [
        {"quality_no": "HHN-JS-03494", "content": "57%Polyester 40%Rayon 3%Spandex",
         "description": "Hacci Jersey", "weight_gsm": 200, "width_in": "63/61",
         "moq_y": 3600, "mcq_y": 1400, "price_rmb_m": 12.16},
        {"quality_no": "HHN-JA-00274", "content": "67%Polyester 27%Rayon 6%Spandex",
         "description": "Jacquard", "weight_gsm": 385, "width_in": "62/60",
         "moq_y": 1800, "price_rmb_m": 29.20},
    ]


# ── pricing ─────────────────────────────────────────────────────────────────

@pytest.mark.parametrize("rmb,expected", _TEMPLATE_PRICES)
def test_quote_matches_source_workbook(rmb, expected):
    assert usd_per_yard(rmb) == pytest.approx(expected)


def test_quote_rounds_up_never_down():
    # 10.00 RMB/M → 1.50126… USD/Y → must land on 1.55, not down on 1.50:
    # rounding a quote down would sell below the intended margin.
    assert usd_per_yard(10.00) == pytest.approx(1.55)


def test_quote_without_price_is_none_not_zero():
    # A fabric with no cost must print blank; quoting 0.00 would be a
    # commitment to sell it for nothing.
    for missing in (None, 0, "", "n/a", -1):
        assert usd_per_yard(missing) is None


def test_quote_parameters_are_applied():
    base = usd_per_yard(20.0, markup=1.0, fx_rate=7.0, round_step=0)
    # abs=1e-4: the helper rounds to 4 dp to keep binary-float dust out of
    # printed prices.
    assert base == pytest.approx(20.0 / 7.0 * 0.9144, abs=1e-4)
    assert usd_per_yard(20.0, markup=1.2, fx_rate=7.0, round_step=0) > base


def test_zero_fx_rate_is_rejected():
    with pytest.raises(ValueError):
        usd_per_yard(10.0, fx_rate=0)


# ── store ───────────────────────────────────────────────────────────────────

def test_create_snapshots_price_and_assigns_token(store, sample_lines):
    pres = store.create(lines=sample_lines, title="GIII-SWIMMING", customer="GIII")
    assert pres["token"]
    lines = store.lines(pres["id"])
    assert [l["price_usd_y"] for l in lines] == [pytest.approx(1.85),
                                                 pytest.approx(4.40)]
    assert [l["line_no"] for l in lines] == [1, 2]


def test_stored_quote_uses_the_sheets_own_parameters(store, sample_lines):
    pres = store.create(lines=sample_lines, fx_rate=7.5, markup=1.0,
                        round_step=0.05)
    assert pres["fx_rate"] == 7.5
    # Same cost, different FX → a different (lower) quote than the default.
    assert store.lines(pres["id"])[0]["price_usd_y"] < 1.85


def test_empty_presentation_is_rejected(store):
    with pytest.raises(ValueError):
        store.create(lines=[])


def test_tokens_are_unique(store, sample_lines):
    tokens = {store.create(lines=sample_lines)["token"] for _ in range(15)}
    assert len(tokens) == 15


def test_delete_removes_lines_and_scans(store, sample_lines):
    pres = store.create(lines=sample_lines)
    store.log_scan(pres["token"])
    assert store.delete(pres["id"]) is True
    assert store.count() == 0
    assert store.lines(pres["id"]) == []
    assert store.scans(pres["id"]) == []


# ── scan tracking ───────────────────────────────────────────────────────────

def test_scan_is_recorded_against_the_sheet(store, sample_lines):
    pres = store.create(lines=sample_lines)
    got = store.log_scan(pres["token"], client_ip="192.168.0.7", user_agent="UA")
    assert got["id"] == pres["id"]
    scans = store.scans(pres["id"])
    assert len(scans) == 1 and scans[0]["client_ip"] == "192.168.0.7"


def test_unknown_token_logs_nothing_and_does_not_raise(store, sample_lines):
    store.create(lines=sample_lines)
    assert store.log_scan("not-a-real-token") is None


def test_history_reports_scan_counts(store, sample_lines):
    pres = store.create(lines=sample_lines, customer="GIII")
    store.log_scan(pres["token"])
    store.log_scan(pres["token"])
    row = store.list_all().iloc[0]
    assert row["n_lines"] == 2 and row["n_scans"] == 2
    assert row["first_scan"] and row["last_scan"]


def test_list_all_on_empty_db_has_columns(store):
    df = store.list_all()
    assert df.empty and "token" in df.columns


# ── export ──────────────────────────────────────────────────────────────────

def _sheet_values(data: bytes) -> list[str]:
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active
    return [str(ws.cell(r, c).value)
            for r in range(1, ws.max_row + 1)
            for c in range(1, ws.max_column + 1)]


def test_customer_export_never_contains_internal_cost(store, sample_lines):
    pres = store.create(lines=sample_lines)
    data = build_presentation_workbook(pres, store.lines(pres["id"]),
                                       price_mode=PRICE_USD)
    values = _sheet_values(data)
    assert "12.16" not in values and "29.2" not in values, \
        "internal RMB/M cost leaked into the customer-facing export"
    assert "1.85" in values


def test_internal_export_shows_both_prices(store, sample_lines):
    pres = store.create(lines=sample_lines)
    data = build_presentation_workbook(pres, store.lines(pres["id"]),
                                       price_mode=PRICE_BOTH)
    values = _sheet_values(data)
    assert "12.16" in values and "1.85" in values


def test_rmb_only_export_omits_the_quote(store, sample_lines):
    pres = store.create(lines=sample_lines)
    data = build_presentation_workbook(pres, store.lines(pres["id"]),
                                       price_mode=PRICE_RMB)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    headers = [wb.active.cell(7, c).value for c in range(1, wb.active.max_column + 1)]
    assert not any("USD" in str(h) for h in headers)


def test_unknown_price_mode_is_rejected(store, sample_lines):
    pres = store.create(lines=sample_lines)
    with pytest.raises(ValueError):
        build_presentation_workbook(pres, store.lines(pres["id"]),
                                    price_mode="everything")


def test_export_without_scanner_url_still_works(store, sample_lines):
    """A site that has not configured the scanner must still get its sheet."""
    pres = store.create(lines=sample_lines)
    data = build_presentation_workbook(pres, store.lines(pres["id"]),
                                       scan_base_url="")
    assert len(data) > 0
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert not wb.active._images


def test_export_carries_the_title_block(store, sample_lines):
    pres = store.create(lines=sample_lines, title="GIII-SWIMMING",
                        customer="GIII", submission_date="2026-05-13",
                        fabric_type="New fabric (HHN-Initiated)")
    values = _sheet_values(build_presentation_workbook(
        pres, store.lines(pres["id"])))
    assert "HIGH HOPE NEWEST" in values
    assert "GIII" in values and "2026-05-13" in values
    assert "New fabric (HHN-Initiated)" in values


def test_unpriced_fabric_prints_blank_not_zero(store):
    pres = store.create(lines=[{"quality_no": "HHN-X", "price_rmb_m": None}])
    wb = openpyxl.load_workbook(io.BytesIO(build_presentation_workbook(
        pres, store.lines(pres["id"]))))
    ws = wb.active
    usd_col = [c for c in range(1, ws.max_column + 1)
               if ws.cell(7, c).value == "USD/Y"][0]
    assert ws.cell(8, usd_col).value is None
