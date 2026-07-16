"""Regression tests for ColorTranslationStore bug fixes:

  • import_from_progress_xlsx must release the workbook's file handle even
    when an exception is raised mid-import (previously skipped wb.close(),
    leaking the open zip handle and locking the file on Windows).
  • load_from_po_data(skip_existing=False) must actually touch rows that
    already exist (previously always fell through to `skipped += 1`
    regardless of the flag's value).
  • the "Replace all data" import path in ui/color_translation_view.py must
    use a connection method that actually exists on the store (Fix 4,
    sub-issue A -- see tests at the bottom of this file).
"""
from __future__ import annotations

import pandas as pd
import pytest
import openpyxl

from po_extractor.store import color_translation_store as cts
from po_extractor.store.color_translation_store import ColorTranslationStore


def _make_progress_xlsx(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test"
    headers = ["颜色", "中文颜色"]
    for ci, h in enumerate(headers, 1):
        ws.cell(1, ci, h)
    ws.cell(2, 1, "NAVY")
    ws.cell(2, 2, "藏青色")
    wb.save(path)
    wb.close()


def test_import_from_progress_xlsx_closes_workbook_on_exception(tmp_path, monkeypatch):
    """A forced exception mid-import must still release the workbook's file
    handle -- otherwise the open zip handle keeps the file locked on
    Windows (can't be deleted or re-uploaded) until the process exits."""
    p = tmp_path / "progress.xlsx"
    _make_progress_xlsx(p)

    store = ColorTranslationStore(str(tmp_path / "ct.db"))

    def _boom(en_color):
        raise RuntimeError("forced failure mid-import")

    monkeypatch.setattr(cts, "_derive_shade_and_label", _boom)

    with pytest.raises(RuntimeError):
        store.import_from_progress_xlsx(str(p), client="__t_co__")

    # If wb.close() were skipped (the bug), the open zip handle would keep
    # this file locked on Windows and unlink() would raise PermissionError.
    p.unlink()


def test_load_from_po_data_skip_existing_false_updates_existing_row(tmp_path):
    """skip_existing=False must actually mark existing rows as 'updated',
    not silently treat them the same as skip_existing=True (previously the
    flag had zero effect once a row already existed -- it always fell
    through to `skipped += 1`)."""
    from po_extractor.store.po_store import POStore
    from po_extractor.store.sky_east_store import SkyEastStore
    from po_extractor.models.po_data import POData, POMetadata, SizeRow
    from auth.companies import COMPANY_GIII

    db_path = str(tmp_path / "shared.db")
    po_store = POStore(db_path)
    po_store.check_and_save(POData(
        metadata=POMetadata(po_number="PO900", style="ST9", company=COMPANY_GIII),
        size_rows=[SizeRow("PO900", "ST9", "Teal", "M", 5, "")],
    ))
    # load_from_po_data() also scans sky_east_items in the same DB -- create
    # that schema too (empty is fine) since ColorTranslationStore itself
    # never creates it.
    SkyEastStore(db_path)

    ct_store = ColorTranslationStore(db_path)

    first = ct_store.load_from_po_data(skip_existing=True)
    assert first["inserted"] == 1
    assert first["sources"]["giii"] == 1

    # Re-scanning with skip_existing=False must report the row as updated,
    # not inserted (still absent) and not silently skipped (the bug).
    second = ct_store.load_from_po_data(skip_existing=False)
    assert second["inserted"] == 0
    assert second["updated"] == 1


# ---------------------------------------------------------------------------
# Fix 4, sub-issue A: "Replace all data" called store._connect(), which
# BaseSQLiteStore never defines (only _conn() exists) -- every attempt raised
# AttributeError before the DELETE ever ran. These lock in that _conn() is
# the real method and that the exact pattern the view now uses works.
# ---------------------------------------------------------------------------

def test_store_has_no_connect_method(tmp_path):
    store = ColorTranslationStore(str(tmp_path / "ct2.db"))
    assert not hasattr(store, "_connect")


def test_conn_context_manager_deletes_all_rows(tmp_path):
    """Exact pattern used by ui/color_translation_view.py's "Replace all"
    branch after the fix -- must not raise AttributeError, and must commit
    the delete."""
    store = ColorTranslationStore(str(tmp_path / "ct3.db"))
    store.upsert_from_df(pd.DataFrame([
        {"Client": "GIII", "Brand": "Karl Lagerfeld", "English Color": "Navy", "Chinese Color": "藏青"},
        {"Client": "Sky East", "Brand": "Anna Field", "English Color": "Black", "Chinese Color": "黑色"},
    ]))
    assert store.count() == 2

    with store._conn() as conn:
        conn.execute("DELETE FROM color_translations")

    assert store.count() == 0
