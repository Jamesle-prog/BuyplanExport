"""Golden style tests for the GIII fax-PO Excel builders.

Pin the workbook styling (header fills, fonts, borders, widths) so the
shared style-kit refactor — and any future one — provably does not change
the delivered files.
"""
from __future__ import annotations

import io

import openpyxl
import pytest


def _po(**over):
    base = dict(
        po_number='CSKHHA001R', style='G5DTN93C', po_date='1/1/26',
        ship_date='1/2/26', etd='1/3/26', vendor='VENDOR CO', factory='FTY',
        fob_price='4.50', description='JACKET', customer_name='CUST',
        ship_to='CUST / STREET', hanger_info='HANGER (1-2-2-1)',
        pack_ratio='1-2-2-1', hts_num='6110.20.2079', cpo='TBD', msrp='59.00',
        source_file='a.pdf',
        line_items=[{
            'ln': '001', 'style': 'G5DTN93C', 'color': 'BLACK',
            'sizes': [('S', 10, '123456789012', 4.5),
                      ('M', 20, '123456789013', None)],
        }],
    )
    base.update(over)
    return base


def _first_sheet(xlsx_bytes: bytes):
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    return wb, wb.worksheets[0]


def _assert_hdr_style(cell, expected_bg: str):
    assert cell.fill.fgColor.rgb == expected_bg
    assert cell.font.name == 'Arial'
    assert cell.font.bold is True
    assert cell.font.size == 10
    assert cell.font.color.rgb == 'FFFFFFFF'
    assert cell.alignment.horizontal == 'center'
    assert cell.alignment.wrap_text is True
    assert cell.border.left.style == 'thin'


def _assert_data_style(cell):
    assert cell.font.name == 'Arial'
    assert cell.font.size == 10
    assert cell.border.left.style == 'thin'
    assert cell.border.bottom.style == 'thin'


@pytest.mark.parametrize("builder_path, hdr_bg", [
    ("ui.giii.msg_extraction:_build_excel_bytes",  'FF1F3864'),   # navy
    ("ui.giii.kl_extraction:_build_kl_excel",      'FF1F3864'),   # navy
    ("ui.giii.tk_eu_extraction:_build_tk_eu_excel", 'FF1F6B75'),  # teal
])
def test_builder_styles_are_stable(builder_path, hdr_bg):
    mod_name, fn_name = builder_path.split(":")
    import importlib
    builder = getattr(importlib.import_module(mod_name), fn_name)

    wb, ws = _first_sheet(builder([_po()]))

    # Header row styling
    _assert_hdr_style(ws.cell(1, 1), hdr_bg)
    # First data cell styling
    _assert_data_style(ws.cell(2, 1))
    # Autofit set a bounded width on column A
    width = ws.column_dimensions['A'].width
    assert width is not None and 8 <= width <= 60
    # Two sheets (detail + summary)
    assert len(wb.worksheets) == 2
    # The PO's data actually landed
    assert ws.cell(2, 1).value == 'CSKHHA001R'
