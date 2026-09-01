"""The GIII 生产计划单 carries style photos, like the Sky East buy plan.

Both now read the same curated photo library through
``ui.shared.load_style_photo_map`` and take the same
``{style: [front, back]}`` map, so a style photographed once shows up on
whichever buy plan the style appears in.

Placement differs on purpose: Sky East stretches its photos into a merged box
its template reserves (twoCellAnchor editAs="twoCell"), while this layout's
free header region is wide and short, so a stretch would squash a portrait
garment shot. Here the photo is scaled to fit with its aspect ratio kept, into
the blank row-3 header band.
"""
from __future__ import annotations

import io

import pytest
from openpyxl import load_workbook

from po_extractor.exporters.giii_production_plan import generate_giii_production_plan
from po_extractor.models.po_data import POData, POMetadata, SizeRow
from po_extractor.store.po_store import POStore


def _png(w: int = 120, h: int = 160, colour=(200, 30, 30)) -> bytes:
    """A real PNG — openpyxl decodes the bytes, so a stub would not embed."""
    from PIL import Image
    buf = io.BytesIO()
    Image.new("RGB", (w, h), colour).save(buf, "PNG")
    return buf.getvalue()


@pytest.fixture
def store(tmp_path):
    s = POStore(str(tmp_path / "po.db"))
    s.save_many_checked([POData(
        metadata=POMetadata(po_number="PO1", style="ST1", company="GIII",
                            seller="A Factory", buyer="ROSS"),
        size_rows=[SizeRow("PO1", "ST1", "BLACK", "M", 100, "700948471565"),
                   SizeRow("PO1", "ST1", "BLACK", "L", 50, "700948471534")])])
    return s


def _sheet(data: bytes, style: str = "ST1"):
    """The style's own sheet — the workbook also carries four summary sheets
    (Summary 汇总 / 简明汇总 / UPC 汇总 / UPC 明细) and the photo belongs to
    the per-style one."""
    wb = load_workbook(io.BytesIO(data))
    assert style in wb.sheetnames, f"{style} not in {wb.sheetnames}"
    return wb[style]


def test_a_style_photo_is_embedded_in_the_plan(store):
    front = _png()
    data = generate_giii_production_plan(
        ["PO1"], store, style_image_map={"ST1": [front, None]})
    assert data

    ws = _sheet(data)
    assert len(ws._images) == 1, "the front photo should be embedded"


def test_front_and_back_are_both_embedded(store):
    data = generate_giii_production_plan(
        ["PO1"], store, style_image_map={"ST1": [_png(), _png(colour=(0, 0, 200))]})
    assert len(_sheet(data)._images) == 2


def test_the_photo_keeps_its_aspect_ratio(store):
    """A garment shot is portrait. Sky East's stretch-to-fill anchor is right
    for its reserved box and wrong here, so this layout scales instead."""
    data = generate_giii_production_plan(
        ["PO1"], store, style_image_map={"ST1": [_png(120, 160), None]})
    img = _sheet(data)._images[0]
    assert img.width / img.height == pytest.approx(120 / 160, rel=0.02)
    assert img.height <= 180 and img.width <= 150      # capped to the band


def test_no_photo_means_no_image_and_no_crash(store):
    """The plan predates photos entirely — a style with none must still
    export, unchanged."""
    for m in (None, {}, {"ST1": []}, {"ST1": [None, None]}, {"OTHER": [_png()]}):
        data = generate_giii_production_plan(["PO1"], store, style_image_map=m)
        assert data
        assert _sheet(data)._images == []


def test_a_corrupt_photo_never_breaks_the_export(store):
    """Photos are a secondary feature layered on real PO data — bad bytes
    lose the picture, never the buy plan."""
    data = generate_giii_production_plan(
        ["PO1"], store, style_image_map={"ST1": [b"not a png at all", None]})
    assert data
    assert _sheet(data)._images == []


def test_the_photo_sits_in_the_blank_header_band_not_over_the_labels(store):
    """Row 3 is the layout's spacer and the size columns carry nothing until
    the table header, so the photo covers neither the 面料 / 品名 labels on
    the left nor the 日期 fields on the right."""
    data = generate_giii_production_plan(
        ["PO1"], store, style_image_map={"ST1": [_png(), None]})
    ws = _sheet(data)
    anchor = ws._images[0].anchor
    col = anchor._from.col if hasattr(anchor, "_from") else None
    row = anchor._from.row if hasattr(anchor, "_from") else None
    assert row == 2, "row 3 (0-based 2) — the blank spacer"
    assert col >= 11, "at or right of the first size column (L, 0-based 11)"


# ── the builder wires the loader in, so every caller gets photos ────────────

def test_the_builder_loads_photos_without_being_asked(store, monkeypatch):
    """All three UI entry points call build_giii_production_plan; none of them
    should have to know about photos."""
    import ui.giii._buyplan as bp

    seen = {}

    def _fake_generate(selected, st, **kw):
        seen["map"] = kw.get("style_image_map")
        return b"xlsx"

    monkeypatch.setattr(bp, "generate_giii_production_plan", _fake_generate)
    monkeypatch.setattr(bp, "_style_photos", lambda meta: {"ST1": [b"front", None]})

    data, _warns, _preview = bp.build_giii_production_plan(["PO1"], store)
    assert data == b"xlsx"
    assert seen["map"] == {"ST1": [b"front", None]}


def test_an_unreachable_image_folder_loses_photos_not_the_plan(store, monkeypatch):
    """A folder on a share this machine cannot reach must not fail the run."""
    import ui.giii._buyplan as bp
    import ui.shared as sh

    def _boom(*a, **k):
        raise OSError("network path not found")

    monkeypatch.setattr(sh, "load_style_photo_map", _boom)
    data, _warns, _preview = bp.build_giii_production_plan(["PO1"], store)
    assert data, "the buy plan must still be produced"
