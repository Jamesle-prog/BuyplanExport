"""Regression tests for buy-plan template generation and header detection."""
import io

import openpyxl

from po_extractor.ui_helpers.template_detect import (
    detect_template_header_row, make_sample_buyplan_template,
)


def test_make_sample_template_returns_valid_xlsx_bytes():
    data = make_sample_buyplan_template()
    assert isinstance(data, bytes) and len(data) > 1000
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert "Template" in wb.sheetnames


def test_sample_template_contains_data_start_marker():
    data = make_sample_buyplan_template()
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["Template"]
    assert ws["A5"].value == "{{data_start}}"


def test_sample_template_has_metadata_placeholders():
    data = make_sample_buyplan_template()
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["Template"]
    assert ws["B1"].value == "{{factory}}"
    assert ws["B2"].value == "{{style}}"
    assert ws["B4"].value == "{{xfactory_date}}"
    assert ws["H4"].value == "{{coo}}"
    assert ws["K4"].value == "{{division}}"


def test_detect_header_row_finds_marker():
    data = make_sample_buyplan_template()
    row = detect_template_header_row(data)
    assert row == 5


def test_detect_header_row_returns_none_when_marker_absent():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Header"
    ws["A2"] = "Data"
    buf = io.BytesIO()
    wb.save(buf)
    assert detect_template_header_row(buf.getvalue()) is None


def test_detect_header_row_handles_invalid_bytes():
    assert detect_template_header_row(b"not an xlsx file") is None


def test_detect_header_row_case_insensitive():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A3"] = "{{DATA_START}}"  # uppercase
    buf = io.BytesIO()
    wb.save(buf)
    assert detect_template_header_row(buf.getvalue()) == 3
