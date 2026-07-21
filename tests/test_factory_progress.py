"""Tests for the factory progress module (工厂进度回报):
FactoryProgressStore (dated report log, derived totals, order-qty lookup)
and the Excel round-trip (request-form build + returned-form parse).
"""
from __future__ import annotations

import sqlite3

import pytest

from po_extractor.store.factory_progress_store import FactoryProgressStore
from po_extractor.exporters.factory_progress_form import (
    build_progress_request_xlsx,
    parse_progress_report_xlsx,
)


@pytest.fixture
def store(tmp_path):
    return FactoryProgressStore(str(tmp_path / "po_history.db"))


# ── Store: report log + totals ──────────────────────────────────────────────

def test_add_and_totals_are_derived_by_summing(store):
    store.add_report("PO1", "STY1", "cutting", "2026-07-10", 300, factory="F1")
    store.add_report("PO1", "STY1", "cutting", "2026-07-15", 200, factory="F1")
    store.add_report("PO1", "STY1", "sewing",  "2026-07-15", 150, factory="F1")

    totals = store.totals_for_pairs([("PO1", "STY1")])
    assert totals[("PO1", "STY1")] == {"cutting": 500, "sewing": 150, "packing": 0}

    dates = store.last_report_dates([("PO1", "STY1")])
    assert dates[("PO1", "STY1")] == "2026-07-15"


def test_add_report_validation(store):
    with pytest.raises(ValueError):
        store.add_report("", "STY1", "cutting", "2026-07-10", 10)
    with pytest.raises(ValueError):
        store.add_report("PO1", "STY1", "washing", "2026-07-10", 10)  # unknown stage
    with pytest.raises(ValueError):
        store.add_report("PO1", "STY1", "cutting", "2026-07-10", 0)   # zero units
    with pytest.raises(ValueError):
        store.add_report("PO1", "STY1", "cutting", "2026-07-10", -5)  # negative
    with pytest.raises(ValueError):
        store.add_report("PO1", "STY1", "cutting", "", 10)            # no date


def test_delete_reports_is_the_correction_path(store):
    rid = store.add_report("PO1", "STY1", "cutting", "2026-07-10", 999)
    store.add_report("PO1", "STY1", "cutting", "2026-07-10", 100)
    assert store.delete_reports([rid]) == 1
    totals = store.totals_for_pairs([("PO1", "STY1")])
    assert totals[("PO1", "STY1")]["cutting"] == 100


def test_list_reports_filters(store):
    store.add_report("PO1", "STY1", "cutting", "2026-07-10", 10, factory="F1")
    store.add_report("PO2", "STY2", "sewing",  "2026-07-11", 20, factory="F2")

    assert len(store.list_reports()) == 2
    assert len(store.list_reports(po_number="PO1")) == 1
    assert len(store.list_reports(factory="F2")) == 1
    assert store.list_reports(factory="F2")[0]["stage"] == "sewing"


def test_order_qty_for_pairs_reads_both_pipelines(store, tmp_path):
    # Create the GIII + Sky East source tables in the same DB file.
    from po_extractor.store.po_store import POStore
    from po_extractor.store.sky_east_store import SkyEastStore
    db = store.db_path
    POStore(db)
    SkyEastStore(db)
    with sqlite3.connect(db) as conn:
        conn.execute("INSERT INTO po_size_rows (po_number, style, color, size, units) "
                     "VALUES ('PO-G1', 'STYG', 'Black', 'M', 400)")
        conn.execute("INSERT INTO po_size_rows (po_number, style, color, size, units) "
                     "VALUES ('PO-G1', 'STYG', 'Black', 'L', 600)")
        conn.execute("INSERT INTO sky_east_items (pc_no, zalando_po, style, xs, s, m, l, xl, xxl) "
                     "VALUES ('PC1', 'PO-SE1', 'STYS', 10, 20, 30, 20, 10, 0)")

    qty = store.order_qty_for_pairs([("PO-G1", "STYG"), ("PO-SE1", "STYS"),
                                     ("PO-NONE", "X")])
    assert qty[("PO-G1", "STYG")] == 1000
    assert qty[("PO-SE1", "STYS")] == 90
    assert ("PO-NONE", "X") not in qty


# ── Excel round-trip ────────────────────────────────────────────────────────

def _form_rows():
    return [
        {"po_number": "PO1", "style": "STY1", "order_qty": 1000,
         "cut": 300, "sewn": 0, "packed": 0},
        {"po_number": "PO2", "style": "STY2", "order_qty": 500,
         "cut": 0, "sewn": 0, "packed": 0},
    ]


def test_form_round_trip_parses_filled_quantities(tmp_path):
    content = build_progress_request_xlsx("Factory Alpha", _form_rows())

    # Simulate the factory filling the form in.
    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    # Header at row 4 -> data rows 5, 6. New Cut=col7, New Sewn=8, Date=10, Notes=11.
    ws.cell(5, 7, value=200)
    ws.cell(5, 10, value="2026-07-18")
    ws.cell(6, 8, value=50)
    ws.cell(6, 10, value="2026-07-18")
    ws.cell(6, 11, value="first sewing batch")
    buf = io.BytesIO()
    wb.save(buf)

    parsed = parse_progress_report_xlsx(buf.getvalue(), factory="Factory Alpha")
    assert parsed["issues"] == []
    assert parsed["rows_seen"] == 2
    reports = {(r["po_number"], r["stage"]): r for r in parsed["reports"]}
    assert reports[("PO1", "cutting")]["units"] == 200
    assert reports[("PO1", "cutting")]["report_date"] == "2026-07-18"
    assert reports[("PO2", "sewing")]["units"] == 50
    assert reports[("PO2", "sewing")]["notes"] == "first sewing batch"
    assert all(r["factory"] == "Factory Alpha" for r in parsed["reports"])


def test_parse_skips_bad_rows_with_issues_not_exceptions(tmp_path):
    content = build_progress_request_xlsx("F", _form_rows())
    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    ws.cell(5, 7, value="abc")          # non-numeric -> issue
    ws.cell(5, 10, value="2026-07-18")
    ws.cell(6, 7, value=-10)            # negative -> issue
    ws.cell(6, 10, value="2026-07-18")
    buf = io.BytesIO()
    wb.save(buf)

    parsed = parse_progress_report_xlsx(buf.getvalue())
    assert parsed["reports"] == []
    assert len(parsed["issues"]) == 2


def test_parse_requires_date_when_quantities_given(tmp_path):
    content = build_progress_request_xlsx("F", _form_rows())
    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    ws.cell(5, 7, value=100)            # quantity but NO date
    buf = io.BytesIO()
    wb.save(buf)

    parsed = parse_progress_report_xlsx(buf.getvalue())
    assert parsed["reports"] == []
    assert any("Report Date" in i for i in parsed["issues"])


def test_parse_rejects_a_non_form_file():
    import io
    import openpyxl
    wb = openpyxl.Workbook()
    wb.active.append(["random", "spreadsheet"])
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(ValueError):
        parse_progress_report_xlsx(buf.getvalue())


# ── Buy plan Index sheet: 裁剪数 from cutting reports ────────────────────────

def test_index_cut_qty_enrichment(tmp_path, monkeypatch):
    from po_extractor.exporters.sky_east_buyplan_export import (
        _enrich_sheet_meta_with_progress,
    )
    store = FactoryProgressStore(str(tmp_path / "fp.db"))
    monkeypatch.setattr(
        "po_extractor.store.get_factory_progress_store", lambda: store)
    # No production_tracking record needed -- cut_qty fills independently.
    from po_extractor.store.production_tracking_store import ProductionTrackingStore
    pt = ProductionTrackingStore(str(tmp_path / "pt.db"))
    monkeypatch.setattr(
        "po_extractor.store.get_production_tracking_store", lambda: pt)

    store.add_report("PO001", "DR5124", "cutting", "2026-07-10", 300)
    store.add_report("PO001", "DR5124", "cutting", "2026-07-15", 250)

    meta = [{"po_no": "PO001", "style": "DR5124"},
            {"po_no": "PO002", "style": "DR9999"}]
    _enrich_sheet_meta_with_progress(meta)
    assert meta[0]["cut_qty"] == 550
    assert meta[1]["cut_qty"] == ""     # no reports -> blank cell


# ── Milestone round-trip (form sheet 2 + production_tracking updates) ────────

def test_form_milestone_sheet_round_trip(tmp_path):
    from po_extractor.store._factory_progress_schema import MILESTONE_STAGES
    ms_in = [
        {"po_number": "PO1", "style": "STY1", "stage": "fabric_purchase",
         "expected": "2026-08-01", "note": "on order", "completed": ""},
        {"po_number": "PO1", "style": "STY1", "stage": "shipping",
         "expected": "", "note": "", "completed": ""},
    ]
    content = build_progress_request_xlsx("F", _form_rows(), milestones=ms_in)

    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content))
    assert "里程碑 Milestones" in wb.sheetnames
    ms = wb["里程碑 Milestones"]
    # Header at row 3 -> data rows 4+. Factory fills row 4's completed date
    # and row 5's expected date + note.
    ms.cell(4, 7, value="2026-08-03")            # fabric arrived
    ms.cell(5, 5, value="2026-09-15")            # delivery expected
    ms.cell(5, 6, value="rush order")
    buf = io.BytesIO()
    wb.save(buf)

    parsed = parse_progress_report_xlsx(buf.getvalue())
    got = {(m["stage"]): m for m in parsed["milestones"]}
    assert got["fabric_purchase"]["completed"] == "2026-08-03"
    assert got["fabric_purchase"]["expected"] == "2026-08-01"   # pre-fill kept
    assert got["shipping"]["expected"] == "2026-09-15"
    assert got["shipping"]["note"] == "rush order"


def test_form_milestone_unknown_stage_is_an_issue_not_a_crash(tmp_path):
    ms_in = [{"po_number": "PO1", "style": "S", "stage": "fabric_purchase",
              "expected": "2026-08-01", "note": "", "completed": ""}]
    content = build_progress_request_xlsx("F", _form_rows(), milestones=ms_in)
    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ms = wb["里程碑 Milestones"]
    ms.cell(4, 4, value="bogus_stage")           # factory mangled the code col
    buf = io.BytesIO()
    wb.save(buf)
    parsed = parse_progress_report_xlsx(buf.getvalue())
    assert parsed["milestones"] == []
    assert any("unknown stage" in i for i in parsed["issues"])


def test_update_stage_fields_partial_and_done(tmp_path):
    from po_extractor.store.production_tracking_store import ProductionTrackingStore
    store = ProductionTrackingStore(str(tmp_path / "pt.db"))
    store.upsert(
        po_number="PO1", style="STY1", factory="F1", company="sky_east",
        updated_by="t", overall_notes="", use_substitute_materials=1,
        stage_fields={}, dep_fields={}, qc_fields={},
    )
    ok = store.update_stage_fields("PO1", "STY1", {
        "fabric_purchase_planned": "2026-08-01",
        "fabric_purchase_notes": "on order",
        "fabric_purchase_actual": "2026-08-03",
        "fabric_purchase_status": "Done",
    }, updated_by="factory")
    assert ok is True
    rec = store.get_batch_by_po_styles([("PO1", "STY1")])[("PO1", "STY1")]
    assert rec["fabric_purchase_planned"] == "2026-08-01"
    assert rec["fabric_purchase_actual"] == "2026-08-03"
    assert rec["fabric_purchase_status"] == "Done"
    assert rec["fabric_purchase_notes"] == "on order"

    # Untracked pair -> False, nothing raised.
    assert store.update_stage_fields("NOPE", "X", {"cutting_notes": "hi"}) is False
    # Bogus column -> hard error (guards SQL injection via stage names).
    import pytest as _pytest
    with _pytest.raises(ValueError):
        store.update_stage_fields("PO1", "STY1", {"evil; DROP": "x"})
