"""po_extractor/parsers/_sheet.py — the worksheet scaffold shared by the
settlement and fabric-condition parsers (v2.125.4)."""
import openpyxl

from po_extractor.parsers._sheet import cell_getter, find_header_row, header_index


def _ws():
    ws = openpyxl.Workbook().active
    ws["A1"] = "结算统计表 2026"                 # title row
    ws["A3"] = " 合同号 "; ws["B3"] = "款号\n(Style)"; ws["C3"] = "Invoice No."
    ws["A4"] = "C-1"; ws["B4"] = "S24DDR010"; ws["C4"] = "INV-9"
    return ws


def test_find_header_row_uses_normalised_headings():
    ws = _ws()
    assert find_header_row(ws, lambda h: "合同号" in h and "款号(style)" in h, max_rows=10) == 3
    assert find_header_row(ws, lambda h: "合同号" in h, max_rows=2) == -1     # beyond max_rows
    assert find_header_row(ws, lambda h: "missing" in h, max_rows=10) == -1


def test_find_header_row_custom_norm():
    ws = _ws()
    assert find_header_row(ws, lambda h: "Invoice No." in h, max_rows=10,
                           norm=lambda v: "" if v is None else str(v).strip()) == 3


def test_header_index_and_cell_getter():
    ws = _ws()
    idx = header_index(ws, 3)
    assert idx == {"合同号": 1, "款号(style)": 2, "invoiceno.": 3}
    get = cell_getter(ws, {"contract": idx["合同号"], "style": idx["款号(style)"]})
    assert get(4, "contract") == "C-1" and get(4, "style") == "S24DDR010"
    assert get(4, "invoice") is None                                   # unmapped field


def test_header_index_first_column_wins_for_duplicate_heading():
    ws = openpyxl.Workbook().active
    ws["A1"] = "日期"; ws["B1"] = "支付"; ws["C1"] = "日期"
    assert header_index(ws, 1) == {"日期": 1, "支付": 2}
