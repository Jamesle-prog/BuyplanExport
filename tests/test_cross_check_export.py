"""Tests for the cross-check exporter (buy plan / color plan / PO summary totals)."""
from __future__ import annotations

import pandas as pd
from openpyxl import Workbook

from po_extractor.exporters.cross_check_export import export_cross_check


def _make_buyplan_with_header_at_row(path, header_row: int, total: int):
    """A buy-plan-shaped workbook whose 'Total' header sits below the
    previously hardcoded 10-row scan cap (client templates make the header
    row configurable with no documented upper bound).

    ``_buyplan_totals`` sums rows header_row+1 .. max_row-1 — the sheet's
    OWN last row is treated as a trailing grand-total row and excluded — so
    a trailing footer row is added after the single data row.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "ST1"
    ws.cell(row=header_row, column=1, value="Style")
    ws.cell(row=header_row, column=2, value="Total")
    ws.cell(row=header_row + 1, column=1, value="ST1")
    ws.cell(row=header_row + 1, column=2, value=total)
    ws.cell(row=header_row + 2, column=1, value="TOTAL")   # excluded footer row
    ws.cell(row=header_row + 2, column=2, value=total)
    wb.save(path)


def _make_colorplan(path, total: int):
    """A color-plan-shaped workbook — ``_colorplan_totals`` just sums every
    numeric cell from row 3 onward (no header search), so a single data row
    with the given total is all that's needed."""
    wb = Workbook()
    ws = wb.active
    ws.title = "ST1"
    ws.cell(row=1, column=1, value="Color")
    ws.cell(row=3, column=1, value="BLACK")
    ws.cell(row=3, column=2, value=total)
    wb.save(path)


def _po_summary(path, style: str, total: int):
    wb = Workbook()
    ws = wb.active
    ws.title = "PO Summary"
    ws.cell(1, 1, "Style"); ws.cell(1, 2, "Total")
    ws.cell(2, 1, style); ws.cell(2, 2, total)
    wb.save(path)


def test_total_header_found_beyond_old_10_row_cap(tmp_path):
    """A 'Total' header at row 15 (below the old min(10, ...) cap) must still
    be found — the previous hardcoded cap made every style read 0 and the
    sheet flag as MISMATCH even though the numbers actually agreed."""
    bp_path = str(tmp_path / "buyplan.xlsx")
    cp_path = str(tmp_path / "colorplan.xlsx")
    ps_path = str(tmp_path / "po_summary.xlsx")
    _make_buyplan_with_header_at_row(bp_path, header_row=15, total=100)
    _make_colorplan(cp_path, total=100)
    _po_summary(ps_path, "ST1", 100)

    df_size = pd.DataFrame({"Style": ["ST1"], "Units": [100]})
    out = export_cross_check(df_size, bp_path, cp_path, ps_path, str(tmp_path))

    from openpyxl import load_workbook
    ws = load_workbook(out)["Cross Check"]
    rows = {ws.cell(r, 1).value: [ws.cell(r, c).value for c in range(2, 7)]
            for r in range(2, ws.max_row)}
    assert rows["ST1"][:4] == [100, 100, 100, 100]
    assert rows["ST1"][4] == "OK"


def test_total_header_still_found_at_classic_row_5(tmp_path):
    """Regression guard: the canonical row-5 header must keep working."""
    bp_path = str(tmp_path / "buyplan.xlsx")
    cp_path = str(tmp_path / "colorplan.xlsx")
    ps_path = str(tmp_path / "po_summary.xlsx")
    _make_buyplan_with_header_at_row(bp_path, header_row=5, total=50)
    _make_colorplan(cp_path, total=50)
    _po_summary(ps_path, "ST1", 50)

    df_size = pd.DataFrame({"Style": ["ST1"], "Units": [50]})
    out = export_cross_check(df_size, bp_path, cp_path, ps_path, str(tmp_path))

    from openpyxl import load_workbook
    ws = load_workbook(out)["Cross Check"]
    assert ws.cell(2, 6).value == "OK"
