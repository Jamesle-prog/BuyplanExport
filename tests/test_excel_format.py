"""Regression tests for the Excel header-row writer."""
import openpyxl

from po_extractor.ui_helpers.excel_format import write_excel_header_row


def _new_ws():
    wb = openpyxl.Workbook()
    return wb, wb.active


def test_writes_values_in_row_1():
    wb, ws = _new_ws()
    write_excel_header_row(ws, ["A", "B", "C"])
    assert ws.cell(1, 1).value == "A"
    assert ws.cell(1, 2).value == "B"
    assert ws.cell(1, 3).value == "C"


def test_styles_applied():
    wb, ws = _new_ws()
    write_excel_header_row(ws, ["Header"])
    c = ws.cell(1, 1)
    assert c.font.bold is True
    assert c.font.color.rgb.endswith("FFFFFF")
    # Default fill hex
    assert c.fill.start_color.rgb.endswith("4472C4")
    assert c.alignment.horizontal == "center"
    assert c.alignment.wrap_text is True


def test_custom_fill_color_applied():
    wb, ws = _new_ws()
    write_excel_header_row(ws, ["X"], fill_hex="FF0000")
    assert ws.cell(1, 1).fill.start_color.rgb.endswith("FF0000")


def test_column_widths_grow_with_label():
    wb, ws = _new_ws()
    write_excel_header_row(ws, ["A", "Very Long Column Header"])
    assert ws.column_dimensions["A"].width == max(10, len("A") + 2)
    assert ws.column_dimensions["B"].width == max(10, len("Very Long Column Header") + 2)


def test_row_height_fixed():
    wb, ws = _new_ws()
    write_excel_header_row(ws, ["x"])
    assert ws.row_dimensions[1].height == 28


def test_handles_iterable_input():
    wb, ws = _new_ws()
    write_excel_header_row(ws, (str(i) for i in range(3)))   # generator
    assert ws.cell(1, 1).value == "0"
    assert ws.cell(1, 3).value == "2"
