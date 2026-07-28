"""Cut-plan house cleanup — port of the hand-run Excel macro.

Pins the macro's behaviour, including the parts that look odd but are
deliberate (substring matching, rule order), so a later "tidy-up" can't
silently change what the cutting room reads.
"""
from __future__ import annotations

import io

import openpyxl
import pytest

from po_extractor.exporters.cutting_plan_clean import (
    build_color_map, build_reverse_color_map, clean_value, clean_workbook,
    find_color_conflicts, has_chinese, strip_path_and_ext, translate_text,
)


# ── translate ────────────────────────────────────────────────────────────────

@pytest.mark.parametrize("src,want", [
    ("Plies number", "层数"),
    ("Colors", "颜色"),
    ("Marker Ratio", "裁剪配比"),
    ("Marker", "版"),
    ("Fabric Length", "面料长度"),
    ("Fabric Weight", "面料重量"),
    ("Order", "订单数"),
    ("Real", "裁剪数"),
])
def test_exact_labels(src, want):
    assert translate_text(src) == want


def test_specific_marker_rules_win_over_the_bare_one():
    """'Marker Length'/'Marker Ratio' must be replaced before plain 'Marker'.

    If the short rule ran first these would come out as '版 Length' / '版 Ratio'
    — the macro's rule order exists precisely to prevent that.
    """
    assert translate_text("Marker Length") == "版长"
    assert translate_text("Marker Ratio") == "裁剪配比"
    # and with the unit suffix the exporter actually writes
    assert translate_text("Marker Length,cm") == "版长,cm"


def test_substring_and_case_insensitive():
    """LookAt:=xlPart + MatchCase:=False — matches inside words, any case."""
    assert translate_text("MARKER") == "版"
    assert translate_text("colors") == "颜色"
    assert translate_text("MATERIAL COST,CNY") == "材料成本,CNY"


def test_every_fixed_heading_is_translated():
    """No fixed heading may reach the cutting room in English. Unit suffixes
    (cm / m / %) stay — they are units, not words."""
    import re
    for label in ["Material", "Solution", "Spreading Plies", "Total Efficiency",
                  "Min Plies", "Max Plies", "Client", "Date", "Time", "Sum",
                  "Sizes", "Quantity", "Total Quantity", "File Name",
                  "N Markers", "Order demands", "Marker Definition",
                  "Cut plan operator", "Average Length", "Total Tables",
                  "Width, cm", "Length, cm", "Efficiency,%"]:
        out = translate_text(label)
        left = re.sub(r"[^A-Za-z]", "", out).replace("cm", "").replace("m", "")
        assert not left, f"{label!r} -> {out!r} still English"


def test_cut_length_and_material_cost_labels():
    assert translate_text("Cut Length,m") == "裁剪长度,m"
    assert translate_text("Material Cost,CNY") == "材料成本,CNY"


def test_total_cut_length_is_not_half_translated():
    """'Total cut length' must be replaced before the 'Cut Length' rule, or it
    would come out as 'Total 裁剪长度'."""
    assert translate_text("Total cut length") == "总裁剪长度"


def test_material_cost_wins_over_bare_material():
    """'Material Cost' must be replaced before the bare 'Material' rule, or
    the cost heading comes out as '面料 Cost,CNY'."""
    assert translate_text("Material Cost,CNY") == "材料成本,CNY"
    assert translate_text("Material") == "面料"


def test_replacements_do_not_cascade():
    """Every replacement is Chinese and every search term ASCII, so no rule
    can re-match an earlier rule's output."""
    assert translate_text("Marker Length") == "版长"
    assert translate_text(translate_text("Marker Length")) == "版长"


# ── path / extension ─────────────────────────────────────────────────────────

def test_strips_windows_path_and_mrk():
    assert strip_path_and_ext(r"D:\Markers\SS26\ABC-1234.mrk") == "ABC-1234"


def test_strips_path_without_extension():
    assert strip_path_and_ext(r"D:\Markers\SS26\ABC-1234") == "ABC-1234"


def test_strips_extension_without_path():
    assert strip_path_and_ext("ABC-1234.mrk") == "ABC-1234"


def test_extension_check_is_case_insensitive():
    """The macro's Right()= comparison missed .MRK; this does not."""
    assert strip_path_and_ext("ABC.MRK") == "ABC"


def test_plain_text_untouched():
    assert strip_path_and_ext("ABC-1234") == "ABC-1234"
    assert strip_path_and_ext("") == ""


def test_only_last_backslash_segment_kept():
    assert strip_path_and_ext(r"\\server\share\deep\path\M1.mrk") == "M1"


# ── cell-level ───────────────────────────────────────────────────────────────

def test_non_text_values_pass_through():
    for v in (None, 12, 30.5, True):
        assert clean_value(v) is v


def test_formulas_are_never_rewritten():
    assert clean_value("=SUM(A1:A9)") == "=SUM(A1:A9)"


def test_translate_then_strip_on_one_cell():
    assert clean_value(r"D:\Markers\SS26\ABC.mrk") == "ABC"


# ── colour translation from the PO colour data ───────────────────────────────

# Shape returned by ColorTranslationStore.build_lookup_dict().
_LOOKUP = {
    ("GIII", "DKNY", "Navy"): "藏青",
    ("GIII", "DKNY", "Clay"): "泥粉",
    ("Sky East", "Zalando", "Navy"): "深蓝",
    ("GIII", "DKNY", "Blank"): "",          # half-filled row
}


def test_build_color_map_flattens_and_normalises():
    m = build_color_map(_LOOKUP)
    assert m["navy"] in {"藏青", "深蓝"}
    assert m["clay"] == "泥粉"
    assert "blank" not in m                 # empty cn never mapped


def test_client_rows_win_over_other_clients():
    assert build_color_map(_LOOKUP, client="GIII")["navy"] == "藏青"
    assert build_color_map(_LOOKUP, client="Sky East")["navy"] == "深蓝"


def test_colour_cell_translated_case_insensitively():
    m = build_color_map(_LOOKUP, client="GIII")
    for src in ("NAVY", "navy", " Navy "):
        assert clean_value(src, m) == "藏青"


def test_colour_already_chinese_is_left_alone():
    m = build_color_map(_LOOKUP, client="GIII")
    assert clean_value("藏青", m) == "藏青"


def test_unknown_colour_falls_through_unchanged():
    m = build_color_map(_LOOKUP, client="GIII")
    assert clean_value("PUCE", m) == "PUCE"


def test_colour_match_is_whole_cell_only():
    """A substring rule here would corrupt any text containing a colour word."""
    m = build_color_map(_LOOKUP, client="GIII")
    assert clean_value("Navy Blazer 2pc", m) == "Navy Blazer 2pc"


def test_colour_mapping_does_not_disturb_labels():
    m = build_color_map(_LOOKUP, client="GIII")
    assert clean_value("Marker Ratio", m) == "裁剪配比"
    assert clean_value(r"D:\Markers\M1.mrk", m) == "M1"


def test_has_chinese():
    assert has_chinese("藏青") and not has_chinese("Navy")


# ── plan-vs-PO colour disagreement (report, don't rewrite) ───────────────────

def _wb_with(*values):
    wb = openpyxl.Workbook()
    ws = wb.active
    for i, v in enumerate(values, start=1):
        ws.cell(i, 1, v)
    return wb


def test_reverse_map_recognises_each_clients_spelling():
    rev = build_reverse_color_map(_LOOKUP)
    assert rev["藏青"] == "navy" and rev["深蓝"] == "navy"
    assert rev["泥粉"] == "clay"


def test_conflict_reported_when_plan_uses_another_spelling():
    """Plan says 深蓝; for GIII the PO calls Navy 藏青 — flag it."""
    cmap = build_color_map(_LOOKUP, client="GIII")
    rev = build_reverse_color_map(_LOOKUP)
    got = find_color_conflicts(_wb_with("深蓝"), cmap, rev)
    assert len(got) == 1
    assert got[0]["plan_cn"] == "深蓝"
    assert got[0]["po_cn"] == "藏青"
    assert got[0]["english"] == "navy"


def test_no_conflict_when_names_agree():
    cmap = build_color_map(_LOOKUP, client="GIII")
    rev = build_reverse_color_map(_LOOKUP)
    assert find_color_conflicts(_wb_with("藏青"), cmap, rev) == []


def test_unknown_chinese_text_is_not_a_conflict():
    """Chinese that isn't a known colour (labels, notes) must never be flagged."""
    cmap = build_color_map(_LOOKUP, client="GIII")
    rev = build_reverse_color_map(_LOOKUP)
    assert find_color_conflicts(_wb_with("裁剪配比", "层数"), cmap, rev) == []


def test_conflicts_are_deduped():
    cmap = build_color_map(_LOOKUP, client="GIII")
    rev = build_reverse_color_map(_LOOKUP)
    assert len(find_color_conflicts(_wb_with("深蓝", "深蓝", "深蓝"),
                                    cmap, rev)) == 1


def test_detection_does_not_modify_the_workbook():
    """The whole point: reporting must leave the plan's own names in place."""
    cmap = build_color_map(_LOOKUP, client="GIII")
    rev = build_reverse_color_map(_LOOKUP)
    wb = _wb_with("深蓝")
    find_color_conflicts(wb, cmap, rev)
    assert wb.active["A1"].value == "深蓝"


def test_cleaning_alone_never_rewrites_a_chinese_colour():
    cmap = build_color_map(_LOOKUP, client="GIII")
    wb = _wb_with("深蓝")
    clean_workbook(wb, cmap)                    # no overrides approved
    assert wb.active["A1"].value == "深蓝"


def test_approved_override_rewrites_it():
    wb = _wb_with("深蓝")
    clean_workbook(wb, None, {"深蓝": "藏青"})
    assert wb.active["A1"].value == "藏青"


def test_cleaned_copy_leaves_the_source_bytes_untouched():
    """The PDF cleanup works on a throwaway copy — the stored workbook that
    the user can still download must be byte-identical afterwards."""
    from ui.cutting_plan._shared import _cleaned_copy
    wb = _wb_with("Marker Ratio", r"D:\Markers\M1.mrk")
    buf = io.BytesIO()
    wb.save(buf)
    original = buf.getvalue()

    cleaned = _cleaned_copy(original)
    assert original == buf.getvalue()          # source not mutated
    ws = openpyxl.load_workbook(io.BytesIO(cleaned)).active
    assert ws["A1"].value == "裁剪配比"
    assert ws["A2"].value == "M1"


def test_cleaned_copy_falls_back_to_the_original_on_failure():
    """A cleanup problem must cost the translation, never the PDF."""
    from ui.cutting_plan._shared import _cleaned_copy
    junk = b"not a workbook"
    assert _cleaned_copy(junk) is junk


def test_apply_color_overrides_round_trips_through_bytes():
    from po_extractor.exporters.cutting_plan_clean import apply_color_overrides
    wb = _wb_with("深蓝", "Material")
    buf = io.BytesIO()
    wb.save(buf)
    out = apply_color_overrides(buf.getvalue(), {"深蓝": "藏青"})
    ws = openpyxl.load_workbook(io.BytesIO(out)).active
    assert ws["A1"].value == "藏青"
    # apply_color_overrides re-runs the whole cleanup, so headings translate too
    assert ws["A2"].value == "面料"


# ── workbook pass ────────────────────────────────────────────────────────────

def test_clean_workbook_reports_and_applies():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Marker Ratio"
    ws["A2"] = r"D:\Markers\M1.mrk"
    ws["A3"] = "Material"          # a fixed heading — also translated now
    ws["A4"] = 42                  # untouched non-text
    changed = clean_workbook(wb)
    assert ws["A1"].value == "裁剪配比"
    assert ws["A2"].value == "M1"
    assert ws["A3"].value == "面料"
    assert ws["A4"].value == 42
    assert changed == 3


# ── integration with the exporter ────────────────────────────────────────────

def _build(clean: bool, color_map=None) -> list[str]:
    from po_extractor.exporters.cutting_plan_export import build_standard_cut_plan
    data = build_standard_cut_plan(
        header={"order_name": "PO-99", "client": "GIII", "operator": "LI",
                "output_folder": r"D:\Markers\SS26",
                "styles": [{"name": "ST1", "file": r"D:\Styles\ST1.sty"}]},
        groups=[("ST1", ["S", "M"])],
        colors=["NAVY"],
        demand_qty={("ST1", "NAVY", "S"): 10, ("ST1", "NAVY", "M"): 20},
        materials=[{
            "material": "Twill",
            "markers": [{"marker_no": 1,
                         "file_name": r"D:\Markers\SS26\ABC-1234.mrk",
                         "ratio": [{"style": "ST1", "size": "S", "qty": 1}]}],
            "spreads": [{"marker_no": 1,
                         "file_name": r"D:\Markers\SS26\ABC-1234.mrk",
                         "rows": []}],
        }],
        clean=clean, color_map=color_map,
    )
    ws = openpyxl.load_workbook(io.BytesIO(data)).active
    return [str(c.value) for r in ws.iter_rows() for c in r
            if isinstance(c.value, str)]


def test_export_cleaned_when_requested():
    vals = _build(clean=True)
    assert "颜色" in vals and "裁剪配比" in vals and "层数" in vals
    # Cut Length / Material Cost are no longer emitted at all (see _METRICS),
    # so the cleaned sheet must not carry them in either language.
    assert not any(v.startswith(("Cut Length", "裁剪长度,",
                                 "Material Cost", "材料成本")) for v in vals)
    assert "面料长度,m" in vals and "版长,cm" in vals   # the ones kept
    assert "ABC-1234" in vals                       # marker path reduced
    assert not any(".mrk" in v for v in vals)
    assert not any(v == "Marker Ratio" for v in vals)


def test_export_translates_colours_from_the_po_data():
    """NAVY is a demand-block colour row; with the PO colour map it ships in
    Chinese, without one it stays as it was."""
    cmap = build_color_map(_LOOKUP, client="GIII")
    assert "藏青" in _build(clean=True, color_map=cmap)
    assert "NAVY" in _build(clean=True, color_map=None)


def test_cleaned_header_keeps_only_what_the_cutting_room_reads():
    vals = _build(clean=True)
    # Kept — and translated, since every fixed heading is Chinese now.
    assert "款式文件 1" in vals and "裁剪计划员" in vals and "客户" in vals
    for gone in ("Date", "Time", "Order name", "Style name 1", "Style name 2",
                 "Output folder path", "款式名称 1", "订单名称"):
        assert gone not in vals, gone


def test_canonical_header_keeps_every_row():
    """The unclean sheet must stay complete — the parser reads Order name /
    Date / Time back out of it and the app round-trips its own export."""
    vals = _build(clean=False)
    for kept in ("Date", "Order name", "Style file 1", "Style name 1",
                 "Cut plan operator", "Client"):
        assert kept in vals, kept


def test_header_rows_are_blanked_not_deleted():
    """Deleting a row pulls everything below it up and slides the marker blocks
    out from under their headings. The cells are emptied in place instead."""
    from po_extractor.exporters.cutting_plan_clean import drop_header_rows
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(1, 1, "Date"); ws.cell(1, 2, "2026-07-23")
    ws.cell(2, 1, "Colors"); ws.cell(3, 1, "Marker Ratio")
    before_rows = ws.max_row
    assert drop_header_rows(ws) == 1
    assert ws.cell(1, 1).value is None and ws.cell(1, 2).value is None
    # everything else stays exactly where it was
    assert ws.max_row == before_rows
    assert ws.cell(2, 1).value == "Colors"
    assert ws.cell(3, 1).value == "Marker Ratio"


def test_row_blanking_only_scans_the_header():
    """A data cell reading 'Date' further down must not be cleared."""
    from po_extractor.exporters.cutting_plan_clean import drop_header_rows
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(1, 1, "Date")
    for r in range(2, 45):
        ws.cell(r, 1, "Colors")
    ws.cell(44, 1, "Date")            # far below the header block
    assert drop_header_rows(ws) == 1          # only the header one
    assert ws.cell(44, 1).value == "Date"     # untouched, and still on row 44


def test_export_default_stays_english_and_reparseable():
    """The canonical layout must stay English — the parser finds its blocks by
    English anchor text and the app round-trips its own export."""
    vals = _build(clean=False)
    assert "Colors" in vals and "Marker Ratio" in vals
    assert any(".mrk" in v for v in vals)           # raw path retained
    assert not any("颜色" == v for v in vals)


# ── optional output folder ───────────────────────────────────────────────────

def test_save_copy_writes_the_file(tmp_path):
    from ui.cutting_plan._shared import save_copy_to_folder
    save_copy_to_folder(b"data", "plan.xlsx", str(tmp_path))
    assert (tmp_path / "plan.xlsx").read_bytes() == b"data"


def test_save_copy_is_a_no_op_without_a_folder(tmp_path):
    from ui.cutting_plan._shared import save_copy_to_folder
    save_copy_to_folder(b"data", "plan.xlsx", "")     # must not raise
    assert list(tmp_path.iterdir()) == []


def test_save_copy_cannot_escape_the_folder(tmp_path):
    """A filename carrying separators must not write outside the folder."""
    from ui.cutting_plan._shared import save_copy_to_folder
    save_copy_to_folder(b"x", r"..\..\escaped.xlsx", str(tmp_path))
    assert (tmp_path / "escaped.xlsx").exists()
    assert not (tmp_path.parent.parent / "escaped.xlsx").exists()


def test_missing_folder_does_not_raise(tmp_path):
    from ui.cutting_plan._shared import save_copy_to_folder
    save_copy_to_folder(b"x", "p.xlsx", str(tmp_path / "nope"))   # warns only
