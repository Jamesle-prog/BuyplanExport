"""Tests for the Cutting Plan module: the Optitex-export parser, the standard
layout exporter (round-trip), the store (plans, PO links, demand comparison)
and the module-permission wiring.

The sample plans these fixtures mirror are real Optitex exports: one where the
markers hit the ordered quantity exactly (single Solution row per colour) and
one where they overcut (Solution splits into 'Order' / 'Real' sub-rows).
"""
from __future__ import annotations

import io

import pytest
from openpyxl import load_workbook

from po_extractor.exporters.cutting_plan_export import (
    build_standard_cut_plan, plan_header_from_parsed, today_header,
)
from po_extractor.parsers.cutting_plan import (
    CuttingPlanParseError, achieved_rows, parse_cut_plan, parse_cut_plan_grid,
)
from po_extractor.store.cutting_plan_store import (
    CuttingPlanStore, consumption, cut_vs_po_pct, summarise_plan,
)


# ── Fixtures: hand-built grids matching the real export layout ──────────────

def _exact_grid() -> list[list]:
    """Two styles, one colour, one material, demand met exactly."""
    return [
        ["Date", "July 27 - 2026", None, "Time", "18:04"],
        ["Order name", "003_060_20260723_JH"],
        ["Style file 1", r"D:\x\S24DTR003.PDS"],
        ["Style name 1", "S24DTR003"],
        ["Style file 2", r"D:\x\ZLD060.PDS"],
        ["Style name 2", "ZLD060"],
        ["Cut plan operator", "JL"],
        ["Client", "ZR"],
        [],
        ["Output folder path", r"D:\x\20260723"],
        [],
        [None, "Order demands"],
        [None, "Colors", "Sizes"],
        [None, None, "S24DTR003", None, "ZLD060"],
        [None, None, "S", "M", "S", "M"],
        [None, "Wine", 300, 200, 300, 200],
        [None, "Sum", 300, 200, 300, 200],
        [],
        ["Marker Definition"],
        ["Material", "N Markers", "Spreading", "Width, cm", "Length, cm",
         "Min Plies", "Max Plies", "Waste Limits, cm"],
        [],
        ["A", 1, "Single", 157, 600, 1, 200, "0, 0, 0, 0"],
        [],
        ["Marker Ratio"],
        ["Marker", "File Name", "Sizes"],
        [None, None, "S24DTR003", None, "ZLD060"],
        [None, None, "S", "M", "S", "M"],
        [1, r"D:\x\A1.mrk", 3, 2, 3, 2],
        [],
        ["Spreading Plies"],
        ["Marker 1", r"D:\x\A1.mrk", "Sizes", None, None, None,
         "Fabric Length,m", "Efficiency,%", "Cut Length,m",
         "Marker Length,cm", "Material Cost,CNY"],
        ["Colors", "Plies number", "S24DTR003", None, "ZLD060"],
        [None, None, "S", "M", "S", "M"],
        ["Wine", 100, 300, 200, 300, 200, 512.5, 87.5, 110.6, 512.5, 512.5],
        ["Sum", 100, 300, 200, 300, 200, 512.5],
        [],
        ["Solution"],
        [None, "Colors", "Sizes", None, None, None,
         "Total Quantity", "Fabric Length,m"],
        [None, None, "S24DTR003", None, "ZLD060"],
        [None, None, "S", "M", "S", "M"],
        [None, "Wine", 300, 200, 300, 200, 1000, 512.5],
        [],
        ["Total Efficiency", 87.5, "%"],
        ["Total Cost", 512.5, "CNY"],
        ["Total cut length", 110.6, "m"],
        ["Total Tables", 1],
        ["Average Length", 0.95, "m"],
    ]


def _overcut_grid() -> list[list]:
    """One style, one material, markers overcut the order (Order/Real rows)."""
    return [
        ["Date", "July 03 - 2026", None, "Time", "16:29"],
        ["Order name", "003_060_20260703"],
        ["Style name 1", "S24DTR003"],
        ["Cut plan operator", "JL"],
        ["Client", "ZR"],
        [],
        [None, "Order demands"],
        [None, "Colors", "Sizes"],
        [None, None, "S24DTR003"],
        [None, None, "S", "M"],
        [None, "Black", 599, 347],
        [None, "Sum", 599, 347],
        [],
        ["Marker Definition"],
        ["Material", "N Markers", "Spreading", "Width, cm", "Length, cm",
         "Min Plies", "Max Plies", "Waste Limits, cm"],
        [],
        ["A", 1, "Single", 157, 600, 1, 200, "0, 0, 0, 0"],
        [],
        ["Marker Ratio"],
        ["Marker", "File Name", "Sizes"],
        [None, None, "S24DTR003"],
        [None, None, "S", "M"],
        [1, r"D:\x\A1.mrk", 2, 1],
        [],
        ["Spreading Plies"],
        ["Marker 1", r"D:\x\A1.mrk", "Sizes", None,
         "Fabric Length,m", "Efficiency,%", "Cut Length,m",
         "Marker Length,cm", "Material Cost,CNY"],
        ["Colors", "Plies number", "S24DTR003"],
        [None, None, "S", "M"],
        ["Black", 350, 600, 350, 875.0, 86.8, 583.9, 583.9, 875.0],
        ["Sum", 350, 600, 350, 875.0],
        [],
        ["Solution"],
        [None, "Colors", "Quantity", "Sizes", None,
         "Total Quantity", "Fabric Length,m"],
        [None, None, None, "S24DTR003"],
        [None, None, None, "S", "M"],
        [None, "Black", "Order", 599, 347, 946],
        [None, None, "Real", 600, 350, 950, 875.0],
        [],
        ["Total Efficiency", 86.8, "%"],
        ["Total Cost", 875.0, "CNY"],
        ["Total cut length", 583.9, "m"],
        ["Total Tables", 3],
        ["Average Length", 0.06, "m"],
        [],
        ["Total Tables", 3],
    ]


@pytest.fixture
def exact():
    return parse_cut_plan_grid(_exact_grid())


@pytest.fixture
def overcut():
    return parse_cut_plan_grid(_overcut_grid())


@pytest.fixture
def store(tmp_path):
    CuttingPlanStore._checked_paths.clear()
    return CuttingPlanStore(str(tmp_path / "po_history.db"))


# ── Parser ──────────────────────────────────────────────────────────────────

def test_header_fields(exact):
    assert exact["order_name"] == "003_060_20260723_JH"
    assert exact["operator"] == "JL"
    assert exact["client"] == "ZR"
    assert exact["plan_date"] == "July 27 - 2026"
    assert exact["output_folder"] == r"D:\x\20260723"
    assert [s["name"] for s in exact["styles"]] == ["S24DTR003", "ZLD060"]


def test_demands_parsed_per_style_and_size(exact):
    demands = {(d["style"], d["size"]): d["qty"] for d in exact["demands"]}
    assert demands[("S24DTR003", "S")] == 300
    assert demands[("ZLD060", "M")] == 200
    assert exact["colors"] == ["Wine"]


def test_order_qty_is_per_style_not_summed(exact):
    # The same garment is cut from two style files; summing both groups would
    # double the order.
    assert exact["style_totals"] == {"S24DTR003": 500, "ZLD060": 500}
    assert exact["total_qty"] == 500


def test_marker_definition_and_ratio(exact):
    mat = exact["materials"][0]
    assert mat["material"] == "A"
    assert mat["width_cm"] == 157
    assert mat["max_plies"] == 200
    ratio = {(c["style"], c["size"]): c["qty"]
             for c in mat["markers"][0]["ratio"]}
    assert ratio[("S24DTR003", "S")] == 3
    assert ratio[("ZLD060", "M")] == 2


def test_spreading_plies_rows_and_metrics(exact):
    spread = exact["materials"][0]["spreads"][0]
    assert spread["marker_no"] == 1
    row = spread["rows"][0]
    assert row["color"] == "Wine"
    assert row["plies"] == 100
    assert row["fabric_length_m"] == 512.5
    assert row["efficiency_pct"] == 87.5
    assert row["marker_length_cm"] == 512.5
    assert sum(p["qty"] for p in row["pieces"]) == 1000


def test_material_totals(exact):
    mat = exact["materials"][0]
    assert mat["total_tables"] == 1
    assert mat["total_efficiency_pct"] == 87.5
    assert mat["total_cut_length_m"] == 110.6
    assert mat["cut_qty"] == 1000


def test_overcut_solution_splits_order_and_real(overcut):
    sol = overcut["materials"][0]["solution"]
    assert [s["kind"] for s in sol] == ["order", "real"]
    # The colour label appears only on the first sub-row but carries down.
    assert {s["color"] for s in sol} == {"Black"}
    real = achieved_rows(sol)
    assert len(real) == 1 and real[0]["total_qty"] == 950
    assert real[0]["fabric_length_m"] == 875.0


def test_overcut_cut_qty_uses_real_row(overcut):
    assert overcut["materials"][0]["cut_qty"] == 950
    assert overcut["total_qty"] == 946          # ordered
    assert summarise_plan(overcut)["cut_qty"] == 950


def test_trailing_grand_total_does_not_overwrite_material_tables(overcut):
    # 'Total Tables' appears twice at the end of a one-material plan; the
    # material's own count must survive.
    assert overcut["materials"][0]["total_tables"] == 3
    assert overcut["total_tables"] == 3


def test_non_cut_plan_workbook_is_rejected():
    with pytest.raises(CuttingPlanParseError):
        parse_cut_plan_grid([["Invoice"], ["Item", "Qty"], ["Shirt", 10]])


def test_empty_grid_is_rejected():
    with pytest.raises(CuttingPlanParseError):
        parse_cut_plan_grid([])


# ── Summary ─────────────────────────────────────────────────────────────────

def test_summarise_plan_efficiency_is_fabric_weighted():
    plan = {
        "style_totals": {"S": 100},
        "materials": [
            {"material": "A", "n_markers": 2, "fabric_length_m": 900.0,
             "total_efficiency_pct": 90.0, "total_tables": 2, "solution": []},
            {"material": "L", "n_markers": 1, "fabric_length_m": 100.0,
             "total_efficiency_pct": 50.0, "total_tables": 1, "solution": []},
        ],
        "total_tables": 3,
    }
    s = summarise_plan(plan)
    assert s["fabric_length_m"] == 1000.0
    assert s["efficiency_pct"] == pytest.approx(86.0)   # not the 70.0 mean
    assert s["total_markers"] == 3


def test_summarise_plan_cut_qty_not_multiplied_by_materials(exact):
    # Shell and lining each cut the same garments; cut qty must not double.
    two_materials = dict(exact)
    two_materials["materials"] = exact["materials"] + exact["materials"]
    assert summarise_plan(two_materials)["cut_qty"] == 500


# ── Exporter round-trip ─────────────────────────────────────────────────────

def _groups_and_qty(plan):
    groups: dict[str, list[str]] = {}
    qty: dict[tuple[str, str, str], int] = {}
    colors: list[str] = []
    for d in plan["demands"]:
        groups.setdefault(d["style"], [])
        if d["size"] not in groups[d["style"]]:
            groups[d["style"]].append(d["size"])
        if d["color"] not in colors:
            colors.append(d["color"])
        qty[(d["style"], d["color"], d["size"])] = d["qty"]
    return list(groups.items()), colors, qty


@pytest.mark.parametrize("fixture_name", ["exact", "overcut"])
def test_standard_export_round_trips(fixture_name, request):
    plan = request.getfixturevalue(fixture_name)
    groups, colors, qty = _groups_and_qty(plan)
    data = build_standard_cut_plan(
        header=plan_header_from_parsed(plan), groups=groups, colors=colors,
        demand_qty=qty, materials=plan["materials"])
    again = parse_cut_plan(io.BytesIO(data))

    assert again["order_name"] == plan["order_name"]
    assert again["style_totals"] == plan["style_totals"]
    assert again["total_tables"] == plan["total_tables"]
    assert len(again["materials"]) == len(plan["materials"])
    for before, after in zip(plan["materials"], again["materials"]):
        assert after["material"] == before["material"]
        assert after["n_markers"] == before["n_markers"]
        assert after["total_tables"] == before["total_tables"]
        assert after["cut_qty"] == before["cut_qty"]
        assert len(after["markers"]) == len(before["markers"])
        assert ([len(s["rows"]) for s in after["spreads"]]
                == [len(s["rows"]) for s in before["spreads"]])


def test_export_without_materials_writes_blank_scaffold():
    data = build_standard_cut_plan(
        header=today_header("PC-1", client="ZR",
                            styles=[{"name": "STY1", "file": ""}]),
        groups=[("STY1", ["S", "M"])], colors=["Black"],
        demand_qty={("STY1", "Black", "S"): 10, ("STY1", "Black", "M"): 20},
        materials=[])
    ws = load_workbook(io.BytesIO(data)).worksheets[0]
    labels = {str(c.value).strip() for row in ws.iter_rows()
              for c in row if c.value}
    # The standard sections are present even with nothing to put in them.
    for section in ("Order demands", "Marker Definition", "Marker Ratio",
                    "Spreading Plies", "Solution", "Total Tables"):
        assert section in labels
    # And the PO quantities made it in.
    assert 10 in {c.value for row in ws.iter_rows() for c in row}


def test_parser_records_each_materials_column_order(exact):
    assert exact["materials"][0]["groups"] == [
        ("S24DTR003", ["S", "M"]), ("ZLD060", ["S", "M"])]


def test_material_size_order_survives_unrelated_po_style_codes(exact):
    """The PO uses the client's style codes, the plan uses the CAD style
    names, so the demand matrix can't order the plan's sizes — the material's
    own recorded order must be used instead of the cell-encounter order."""
    data = build_standard_cut_plan(
        header=plan_header_from_parsed(exact),
        groups=[("TP5016", ["S", "M"])],          # PO style code, not the plan's
        colors=["Wine"],
        demand_qty={("TP5016", "Wine", "S"): 10, ("TP5016", "Wine", "M"): 20},
        materials=exact["materials"])
    again = parse_cut_plan(io.BytesIO(data))
    assert again["materials"][0]["groups"] == [
        ("S24DTR003", ["S", "M"]), ("ZLD060", ["S", "M"])]
    ratio = {(c["style"], c["size"]): c["qty"]
             for c in again["materials"][0]["markers"][0]["ratio"]}
    assert ratio == {("S24DTR003", "S"): 3, ("S24DTR003", "M"): 2,
                     ("ZLD060", "S"): 3, ("ZLD060", "M"): 2}


def test_material_groups_fall_back_to_canonical_size_order():
    """A plan stored before the column order was recorded still exports with
    sizes in a sane order rather than whichever marker was read first."""
    from po_extractor.exporters.cutting_plan_export import _material_groups
    mat = {
        "markers": [{"marker_no": 1, "ratio": [
            {"style": "STY", "size": "XL", "qty": 1},
            {"style": "STY", "size": "S", "qty": 1},
            {"style": "STY", "size": "M", "qty": 1},
        ]}],
        "spreads": [], "solution": [],
    }
    assert _material_groups(mat, fallback=[("OTHER", ["S", "M"])]) == [
        ("STY", ["S", "M", "XL"])]


def test_export_demands_come_from_the_po_not_the_plan(exact):
    """The whole point of standardising: the sheet shows what was ordered."""
    data = build_standard_cut_plan(
        header=plan_header_from_parsed(exact),
        groups=[("S24DTR003", ["S"])], colors=["Wine"],
        demand_qty={("S24DTR003", "Wine", "S"): 999},
        materials=exact["materials"])
    again = parse_cut_plan(io.BytesIO(data))
    assert again["demands"] == [
        {"style": "S24DTR003", "color": "Wine", "size": "S", "qty": 999}]


# ── Store ───────────────────────────────────────────────────────────────────

def test_save_and_get_plan(store, exact):
    pid = store.save_plan(exact, source_file="plan.xlsx", file_bytes=b"xyz",
                          uploaded_by="jl", notes="first run",
                          links=[{"pc_no": "PC1", "po_no": "PO1",
                                  "style": "TP5016"}])
    rec = store.get_plan(pid)
    assert rec["plan_name"] == "003_060_20260723_JH"
    assert rec["order_qty"] == 500
    assert rec["notes"] == "first run"
    assert rec["parsed"]["materials"][0]["material"] == "A"
    assert rec["links"] == [
        {"source": "sky_east", "pc_no": "PC1", "po_no": "PO1",
         "style": "TP5016", "linked_at": rec["links"][0]["linked_at"],
         "linked_by": "jl"}]


def test_list_plans_counts_distinct_pos_not_link_rows(store, exact):
    # One PO linked for three styles is still one PO.
    store.save_plan(exact, source_file="a.xlsx", links=[
        {"pc_no": "PC1", "po_no": "PO1", "style": "TP5016"},
        {"pc_no": "PC1", "po_no": "PO1", "style": "DR5004"},
        {"pc_no": "PC1", "po_no": "PO1", "style": "DR5118"},
        {"pc_no": "PC1", "po_no": "PO2", "style": "TP5016"}])
    row = store.list_plans().iloc[0]
    assert int(row["linked_pos"]) == 2
    assert sorted(row["linked_styles"].split(", ")) == [
        "DR5004", "DR5118", "TP5016"]


def test_one_plan_links_to_many_pos_and_lookup_works_either_way(store, exact):
    pid = store.save_plan(exact, source_file="a.xlsx", links=[
        {"pc_no": "PC1", "po_no": "PO1", "style": "TP5016"},
        {"pc_no": "PC2", "po_no": "PO2", "style": "DR5004"},
    ])
    assert store.plans_for_pos(pc_nos=["PC2"]).iloc[0]["plan_id"] == pid
    assert store.plans_for_pos(po_nos=["PO1"]).iloc[0]["plan_id"] == pid
    assert store.plans_for_pos(styles=["DR5004"]).iloc[0]["plan_id"] == pid
    assert store.plans_for_pos(pc_nos=["NOPE"]).empty
    assert store.plans_for_pos(styles=["NOPE"]).empty


def test_style_narrows_a_po_lookup(store, exact):
    """Two plans on the same PO, one per style — asking for a style must not
    return the other style's plan."""
    tops = store.save_plan(exact, source_file="tops.xlsx", links=[
        {"pc_no": "PC1", "po_no": "PO1", "style": "TP5016"}])
    dresses = store.save_plan(exact, source_file="dresses.xlsx", links=[
        {"pc_no": "PC1", "po_no": "PO1", "style": "DR5004"}])

    both = store.plans_for_pos(pc_nos=["PC1"])
    assert set(both["plan_id"]) == {tops, dresses}

    only_tops = store.plans_for_pos(pc_nos=["PC1"], styles=["TP5016"])
    assert set(only_tops["plan_id"]) == {tops}


def test_whole_po_links_still_match_a_style_query(store, exact):
    """A link with no style means the whole PO, so it covers every style —
    including ones recorded before styles existed."""
    pid = store.save_plan(exact, source_file="a.xlsx",
                          links=[{"pc_no": "PC1", "po_no": "PO1"}])
    hit = store.plans_for_pos(pc_nos=["PC1"], styles=["ANYTHING"])
    assert hit.iloc[0]["plan_id"] == pid


def test_set_links_replaces_previous_links(store, exact):
    pid = store.save_plan(exact, source_file="a.xlsx",
                          links=[{"pc_no": "PC1", "po_no": "PO1",
                                  "style": "TP5016"}])
    store.set_links(pid, [{"pc_no": "PC9", "po_no": "PO9",
                           "style": "DR5004"}], "jl")
    links = store.get_plan(pid)["links"]
    assert [(l["pc_no"], l["po_no"], l["style"]) for l in links] == [
        ("PC9", "PO9", "DR5004")]


def test_same_po_different_styles_are_separate_links(store, exact):
    pid = store.save_plan(exact, source_file="a.xlsx", links=[
        {"pc_no": "PC1", "po_no": "PO1", "style": "TP5016"},
        {"pc_no": "PC1", "po_no": "PO1", "style": "DR5004"},
    ])
    assert len(store.get_plan(pid)["links"]) == 2


def test_duplicate_links_are_collapsed(store, exact):
    pid = store.save_plan(exact, source_file="a.xlsx", links=[
        {"pc_no": "PC1", "po_no": "PO1", "style": "TP5016"},
        {"pc_no": "PC1", "po_no": "PO1", "style": "TP5016"},
    ])
    assert len(store.get_plan(pid)["links"]) == 1


def test_links_table_migrates_from_the_pre_style_schema(tmp_path):
    """A DB written before v2.106.0 keeps its links, gains the style column,
    and can then hold two styles for the same PO."""
    import sqlite3

    db = str(tmp_path / "po_history.db")
    conn = sqlite3.connect(db)
    conn.executescript(
        """
        CREATE TABLE cutting_plan_links (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            plan_id   INTEGER NOT NULL,
            source    TEXT NOT NULL DEFAULT 'sky_east',
            pc_no     TEXT DEFAULT '',
            po_no     TEXT DEFAULT '',
            linked_at TEXT,
            linked_by TEXT,
            UNIQUE(plan_id, source, pc_no, po_no)
        );
        CREATE INDEX idx_cpl_plan ON cutting_plan_links(plan_id);
        CREATE INDEX idx_cpl_pc   ON cutting_plan_links(pc_no);
        CREATE INDEX idx_cpl_po   ON cutting_plan_links(po_no);
        INSERT INTO cutting_plan_links
               (plan_id, source, pc_no, po_no, linked_at, linked_by)
        VALUES (7, 'sky_east', 'PC1', 'PO1', '2026-07-01T00:00:00', 'jl');
        """
    )
    conn.commit()
    conn.close()

    CuttingPlanStore._checked_paths.clear()
    store = CuttingPlanStore(db)

    with store._conn() as c:
        rows = [dict(r) for r in c.execute(
            "SELECT plan_id, pc_no, po_no, style, linked_by "
            "FROM cutting_plan_links")]
    assert rows == [{"plan_id": 7, "pc_no": "PC1", "po_no": "PO1",
                     "style": "", "linked_by": "jl"}]

    store.set_links(7, [{"pc_no": "PC1", "po_no": "PO1", "style": "A"},
                        {"pc_no": "PC1", "po_no": "PO1", "style": "B"}], "jl")
    with store._conn() as c:
        assert c.execute("SELECT COUNT(*) FROM cutting_plan_links "
                         "WHERE plan_id=7").fetchone()[0] == 2


def test_demands_table_records_ordered_and_cut(store, overcut):
    pid = store.save_plan(overcut, source_file="a.xlsx")
    df = store.list_demands(pid)
    row = df[(df["style"] == "S24DTR003") & (df["size"] == "M")].iloc[0]
    assert int(row["qty"]) == 347        # ordered
    assert int(row["cut_qty"]) == 350    # cut


def test_original_file_round_trips(store, exact):
    pid = store.save_plan(exact, source_file="plan.xlsx", file_bytes=b"binary")
    assert store.get_plan_file(pid) == ("plan.xlsx", b"binary")


def test_find_by_hash_flags_a_re_upload(store, exact):
    store.save_plan(exact, source_file="plan.xlsx", file_bytes=b"binary",
                    uploaded_by="jl")
    hit = store.find_by_hash(b"binary")
    assert hit and hit["uploaded_by"] == "jl"
    assert store.find_by_hash(b"other") is None


def test_materials_are_stored_one_row_per_fabric(store, exact):
    pid = store.save_plan(exact, source_file="a.xlsx")
    mats = store.list_plan_materials(pid)
    assert list(mats["material"]) == ["A"]
    row = mats.iloc[0]
    assert row["width_cm"] == 157
    assert row["n_markers"] == 1
    assert row["total_tables"] == 1
    assert row["total_plies"] == 100          # summed over the spreads
    assert row["fabric_length_m"] == 512.5
    assert row["efficiency_pct"] == 87.5


def test_plan_list_expands_to_one_row_per_fabric(store, exact):
    """Shell and lining are different fabrics at different widths — their
    metres and efficiency must never be shown combined."""
    two = dict(exact)
    lining = dict(exact["materials"][0])
    lining.update({"material": "里", "width_cm": 173.0,
                   "fabric_length_m": 64.2, "total_efficiency_pct": 84.2,
                   "n_markers": 2, "total_tables": 2})
    two["materials"] = [exact["materials"][0], lining]

    pid = store.save_plan(two, source_file="a.xlsx")
    df = store.list_plans_by_material()
    assert len(df) == 2
    assert set(df["id"]) == {pid}                      # same plan, two rows
    assert list(df["material"]) == ["A", "里"]
    assert list(df["width_cm"]) == [157, 173.0]
    assert list(df["mat_fabric_m"]) == [512.5, 64.2]
    assert list(df["mat_efficiency_pct"]) == [87.5, 84.2]
    # Plan-level columns repeat on every fabric row.
    assert list(df["plan_name"]) == [two["order_name"]] * 2


def test_plan_with_unreadable_materials_still_appears_in_the_list(store):
    plan = {"order_name": "no materials", "demands": [], "materials": [],
            "style_totals": {}, "colors": []}
    pid = store.save_plan(plan, source_file="a.xlsx")
    df = store.list_plans_by_material()
    assert list(df["id"]) == [pid]
    assert df.iloc[0]["material"] == ""


def test_materials_are_backfilled_for_plans_saved_before_the_table(store, exact):
    """A plan uploaded before per-fabric rows existed must not vanish from the
    per-fabric list — the figures were always in parsed_json."""
    pid = store.save_plan(exact, source_file="a.xlsx")
    with store._conn() as conn:
        conn.execute("DELETE FROM cutting_plan_materials WHERE plan_id=?",
                     (pid,))
    assert store.list_plan_materials(pid).empty

    CuttingPlanStore._checked_paths.clear()
    reopened = CuttingPlanStore(store.db_path)
    assert list(reopened.list_plan_materials(pid)["material"]) == ["A"]


def test_delete_plan_removes_its_material_rows(store, exact):
    pid = store.save_plan(exact, source_file="a.xlsx")
    store.delete_plan(pid)
    assert store.list_plan_materials(pid).empty
    assert store.list_plans_by_material().empty


def test_po_qty_comes_from_the_linked_po(store, exact):
    """The plan states what the cutting room was told to cut; the PO qty is
    what was actually ordered.  Both are shown so a gap is visible."""
    _seed_sky_east_items(store.db_path, [
        ("PC1", "PO1", "TP5016", 300),
        ("PC1", "PO1", "DR5004", 200),
        ("PC1", "PO2", "TP5016", 150),
    ])
    pid = store.save_plan(exact, source_file="a.xlsx", links=[
        {"pc_no": "PC1", "po_no": "PO1", "style": "TP5016"}])
    assert store.po_qty_by_plan()[pid] == 300

    store.set_links(pid, [{"pc_no": "PC1", "po_no": "PO1", "style": "TP5016"},
                          {"pc_no": "PC1", "po_no": "PO1", "style": "DR5004"}])
    assert store.po_qty_by_plan()[pid] == 500

    store.set_links(pid, [{"pc_no": "PC1", "po_no": "", "style": ""}])
    assert store.po_qty_by_plan()[pid] == 650      # every style, both POs


@pytest.mark.parametrize("metres,qty,expected", [
    (2020.6236, 1057, 1.9117),      # shell, per garment
    (64.2141, 1057, 0.0608),        # lining
    (0, 1057, None),                # no fabric recorded — blank, not 0
    (100.0, 0, None),               # nothing cut
    (100.0, None, None),
    (None, 100, None),
])
def test_consumption(metres, qty, expected):
    assert consumption(metres, qty) == expected


def test_consumption_is_per_garment_not_per_piece(store, exact):
    """512.5 m over 500 garments — not over the 1000 pieces those garments
    are cut from, which would halve the figure fabric is ordered against."""
    pid = store.save_plan(exact, source_file="a.xlsx")
    assert store.list_plans_by_material().iloc[0]["m_per_unit"] == 1.025
    assert store.list_plan_materials(pid).iloc[0]["m_per_unit"] == 1.025


@pytest.mark.parametrize("po,cut,expected", [
    (1000, 1057, 5.7),        # overcut
    (1000, 950, -5.0),        # short
    (1000, 1000, 0.0),
    (0, 1057, None),          # nothing linked — no baseline
    (None, 1057, None),
    (1000, 0, -100.0),
])
def test_cut_vs_po_pct(po, cut, expected):
    assert cut_vs_po_pct(po, cut) == expected


def test_diff_pct_uses_unit_cut_qty_not_the_per_fabric_piece_count(store):
    """A co-ord set cuts trousers *and* a top from the shell, so that fabric's
    piece count is double the units. Measuring it against the PO's unit
    quantity would read +111 % when the plan is only 5.7 % over."""
    plan = {
        "order_name": "set", "colors": ["Wine"],
        "style_totals": {"TROUSER": 1057, "TOP": 1057},
        "demands": [{"style": "TROUSER", "color": "Wine", "size": "M",
                     "qty": 1057},
                    {"style": "TOP", "color": "Wine", "size": "M",
                     "qty": 1057}],
        "materials": [{
            "material": "A", "n_markers": 1, "total_tables": 1,
            "fabric_length_m": 100.0, "total_efficiency_pct": 87.0,
            "cut_qty": 2114, "markers": [], "spreads": [],
            "solution": [{"color": "Wine", "kind": "", "total_qty": 2114,
                          "fabric_length_m": 100.0, "cells": [
                              {"style": "TROUSER", "size": "M", "qty": 1057},
                              {"style": "TOP", "size": "M", "qty": 1057}]}],
        }],
        "total_tables": 1,
    }
    _seed_sky_east_items(store.db_path, [("PC1", "PO1", "SET1", 1000)])
    pid = store.save_plan(plan, source_file="a.xlsx",
                          links=[{"pc_no": "PC1", "po_no": "PO1",
                                  "style": "SET1"}])
    row = store.list_plans_by_material().iloc[0]
    assert int(row["mat_cut_qty"]) == 2114      # pieces off the shell
    assert int(row["cut_qty"]) == 1057          # units
    assert int(row["po_qty"]) == 1000
    assert row["diff_pct"] == 5.7               # not 111.4
    assert store.get_plan(pid)["cut_qty"] == 1057


def test_diff_pct_is_blank_when_no_po_is_linked(store, exact):
    store.save_plan(exact, source_file="a.xlsx")
    assert store.list_plans_by_material().iloc[0]["diff_pct"] is None


def test_po_qty_counts_each_item_once(store, exact):
    """A whole-PO link plus a per-style link on the same PO must not add the
    style's quantity twice."""
    _seed_sky_east_items(store.db_path, [("PC1", "PO1", "TP5016", 300)])
    pid = store.save_plan(exact, source_file="a.xlsx", links=[
        {"pc_no": "PC1", "po_no": "PO1", "style": ""},
        {"pc_no": "PC1", "po_no": "PO1", "style": "TP5016"},
    ])
    assert store.po_qty_by_plan()[pid] == 300


def test_po_qty_is_empty_without_the_po_tables(store, exact):
    store.save_plan(exact, source_file="a.xlsx",
                    links=[{"pc_no": "PC1", "po_no": "PO1"}])
    assert store.po_qty_by_plan() == {}
    # The list still renders; the column is simply zero.
    assert list(store.list_plans_by_material()["po_qty"]) == [0]


def _seed_sky_east_items(db_path: str, rows) -> None:
    """Minimal sky_east_items table: ``(pc, po, style, qty[, ex_fty_date])``."""
    import sqlite3

    conn = sqlite3.connect(db_path)
    try:
        conn.execute(
            """CREATE TABLE IF NOT EXISTS sky_east_items (
                   id        INTEGER PRIMARY KEY AUTOINCREMENT,
                   pc_no     TEXT, zalando_po TEXT, style TEXT,
                   total_qty INTEGER DEFAULT 0,
                   ex_fty_date TEXT DEFAULT '')""")
        conn.executemany(
            "INSERT INTO sky_east_items "
            "(pc_no, zalando_po, style, total_qty, ex_fty_date) "
            "VALUES (?,?,?,?,?)",
            [tuple(r) + ("",) * (5 - len(r)) for r in rows])
        conn.commit()
    finally:
        conn.close()


def test_link_rows_expand_to_concrete_pc_po_style_triples():
    """The UI helper turns a selection into one row per (PC, PO, style) so a
    later lookup by any one of them finds the plan."""
    import pandas as pd

    from ui.cutting_plan._shared import POSelection, link_rows

    items = pd.DataFrame([
        {"pc_no": "PC1", "zalando_po": "PO1", "style": "TP5016"},
        {"pc_no": "PC1", "zalando_po": "PO1", "style": "DR5004"},
        {"pc_no": "PC1", "zalando_po": "PO2", "style": "TP5016"},
        {"pc_no": "PC1", "zalando_po": "PO2", "style": "TP5016"},   # dupe
    ])
    rows = link_rows(POSelection(["PC1"], [], [], items))
    assert rows == [
        {"pc_no": "PC1", "po_no": "PO1", "style": "TP5016"},
        {"pc_no": "PC1", "po_no": "PO1", "style": "DR5004"},
        {"pc_no": "PC1", "po_no": "PO2", "style": "TP5016"},
    ]


def test_link_rows_fall_back_to_the_raw_selection_without_items():
    import pandas as pd

    from ui.cutting_plan._shared import POSelection, link_rows

    rows = link_rows(POSelection(["PC1"], ["PO1"], ["TP5016"], pd.DataFrame()))
    assert rows == [{"pc_no": "PC1", "po_no": "PO1", "style": "TP5016"}]


def test_delete_plan_removes_links_and_demands(store, exact):
    pid = store.save_plan(exact, source_file="a.xlsx",
                          links=[{"pc_no": "PC1", "po_no": "PO1",
                                  "style": "TP5016"}])
    assert store.delete_plan(pid) is True
    assert store.get_plan(pid) is None
    assert store.list_demands(pid).empty
    assert store.plans_for_pos(pc_nos=["PC1"]).empty
    assert store.count() == 0


# ── Permissions ─────────────────────────────────────────────────────────────

def test_cutting_plan_is_its_own_module():
    from auth.users import (
        ALL_MODULES, MODULE_CUTTING_PLAN, MODULE_LABELS, MODULE_SKY_EAST,
        MODULE_SKY_EAST_BUYPLAN,
    )
    assert MODULE_CUTTING_PLAN in ALL_MODULES
    assert MODULE_CUTTING_PLAN in MODULE_LABELS
    assert MODULE_CUTTING_PLAN not in (MODULE_SKY_EAST, MODULE_SKY_EAST_BUYPLAN)


def test_sky_east_modules_do_not_grant_cutting_plan():
    """A Buy Plan user must not see cut plans (marker cost/efficiency)."""
    from auth.users import (
        MODULE_CUTTING_PLAN, MODULE_SKY_EAST, MODULE_SKY_EAST_BUYPLAN,
    )

    def allowed(module_key: str, user_modules: list[str]) -> bool:
        # Mirrors app.py's _allowed().
        if not user_modules:
            return True
        if module_key == "sky_east":
            return (MODULE_SKY_EAST in user_modules
                    or MODULE_SKY_EAST_BUYPLAN in user_modules)
        return module_key in user_modules

    buyplan_user = [MODULE_SKY_EAST_BUYPLAN]
    assert allowed("sky_east", buyplan_user) is True
    assert allowed(MODULE_CUTTING_PLAN, buyplan_user) is False

    full_se_user = [MODULE_SKY_EAST]
    assert allowed(MODULE_CUTTING_PLAN, full_se_user) is False

    granted = [MODULE_SKY_EAST_BUYPLAN, MODULE_CUTTING_PLAN]
    assert allowed(MODULE_CUTTING_PLAN, granted) is True


# ── X-factory date: the deadline the cutting room is actually working to ────

def test_x_factory_date_comes_from_the_linked_po(store, exact):
    _seed_sky_east_items(store.db_path, [
        ("PC1", "PO1", "TP5016", 300, "2026-09-02"),
        ("PC1", "PO2", "DR5004", 200, "2026-10-15"),
    ])
    pid = store.save_plan(exact, source_file="a.xlsx", links=[
        {"pc_no": "PC1", "po_no": "PO1", "style": "TP5016"}])
    assert store.x_factory_by_plan()[pid] == "2026-09-02"
    assert store.list_plans_by_material().iloc[0]["ex_fty"] == "2026-09-02"


def test_x_factory_shows_the_span_when_linked_pos_ship_on_different_days(
        store, exact):
    """One date picked silently out of several would misstate the deadline —
    early if the later one is chosen, and hiding that a later one exists if
    the earlier is."""
    _seed_sky_east_items(store.db_path, [
        ("PC1", "PO1", "TP5016", 300, "2026-10-15"),
        ("PC1", "PO2", "DR5004", 200, "2026-09-02"),
    ])
    pid = store.save_plan(exact, source_file="a.xlsx",
                          links=[{"pc_no": "PC1", "po_no": "", "style": ""}])
    assert store.x_factory_by_plan()[pid] == "2026-09-02 → 2026-10-15"


def test_x_factory_ignores_items_that_carry_no_date(store, exact):
    """An undated item must not drag the span back to an empty string."""
    _seed_sky_east_items(store.db_path, [
        ("PC1", "PO1", "TP5016", 300, ""),
        ("PC1", "PO2", "DR5004", 200, "2026-09-02"),
    ])
    pid = store.save_plan(exact, source_file="a.xlsx",
                          links=[{"pc_no": "PC1", "po_no": "", "style": ""}])
    assert store.x_factory_by_plan()[pid] == "2026-09-02"


def test_x_factory_is_blank_for_an_unlinked_plan(store, exact):
    _seed_sky_east_items(store.db_path,
                         [("PC1", "PO1", "TP5016", 300, "2026-09-02")])
    pid = store.save_plan(exact, source_file="a.xlsx")
    assert pid not in store.x_factory_by_plan()
    assert store.list_plans_by_material().iloc[0]["ex_fty"] == ""


def test_x_factory_survives_a_database_with_no_po_tables(store, exact):
    """Same guard as po_qty: a database without the Sky East tables (a test
    DB, a fresh install) must list plans, not raise."""
    store.save_plan(exact, source_file="a.xlsx")
    assert store.x_factory_by_plan() == {}
    assert store.list_plans_by_material().iloc[0]["ex_fty"] == ""


# ── Single-style plans ──────────────────────────────────────────────────────
#
# A plan covering ONE style is laid out differently in two ways, and both were
# read as if it were a multi-style plan:
#
#   * the header writes plain "Style file" / "Style name", with no number;
#   * there is no band of style names between the "Sizes" heading and the size
#     labels, because there is only one style to name.
#
# Read with a band assumed, every size label was taken for a style name and
# the first data row for the sizes — which left the plan with no style totals
# at all (Plan qty 0) and its quantities filed under size names, so Cut qty
# came out as the sum of two mis-keyed buckets instead of the units cut.

def _single_style_grid() -> list[list]:
    """One style, one colour, one material, markers overcut the order.

    Mirrors a real export: 3457 ordered, 3499 cut.
    """
    return [
        ["Date", "July 29 - 2026", None, "Time", "09:46"],
        ["Order name", "1237_mix"],
        ["Style file", r"D:\x\S25DDR1237-XL.PDS"],       # no number
        ["Style name", "S25DDR1237"],                     # no number
        ["Cut plan operator", "LLL"],
        ["Client", "ZR"],
        [],
        ["Output folder path", r"D:\x\2026_1273"],
        [],
        [None, "Order demands"],
        [None, "Colors", "Sizes"],
        [None, None, "XS", "S", "M", "L", "XL"],          # no style band
        [None, "Black", 269, 727, 1033, 865, 563],
        [None, "Sum", 269, 727, 1033, 865, 563],
        [],
        ["Marker Definition"],
        ["Material", "N Markers", "Spreading", "Width, cm", "Length, cm",
         "Min Plies", "Max Plies", "Waste Limits, cm"],
        [],
        ["Shell", 7, "Single", 159, 650, 1, 110, "0, 0, 0, 0"],
        [],
        ["Marker Ratio"],
        ["Marker", "File Name", "Sizes"],
        [None, None, "XS", "S", "M", "L", "XL"],
        [1, r"D:\x\Shell1.mrk", 1, None, 2, 1, 1],
        [],
        ["Spreading Plies"],
        ["Marker 1", r"D:\x\Shell1.mrk", "Sizes", None, None, None, None,
         "Fabric Length,m", "Efficiency,%", "Cut Length,m",
         "Marker Length,cm", "Material Cost,CNY"],
        # Colors / Plies number share the size row — there is no band to
        # carry them.
        ["Colors", "Plies number", "XS", "S", "M", "L", "XL"],
        ["Black", 110, 110, None, 220, 110, 110,
         594.57, 85.84, 100.0, 540.52, 594.57],
        ["Sum", 110, 110, 0, 220, 110, 110, 594.57],
        [],
        ["Solution"],
        [None, "Colors", "Quantity", "Sizes", None, None, None, None,
         "Total Quantity", "Fabric Length,m"],
        [None, None, None, "XS", "S", "M", "L", "XL"],
        [None, "Black", "Order", 269, 727, 1033, 865, 563, 3457],
        [None, None, "Real", 269, 729, 1054, 884, 563, 3499, 3787.88],
        [],
        ["Total Efficiency", 85.72, "%"],
        ["Total Cost", 3787.88, "CNY"],
        ["Total cut length", 773.15, "m"],
        ["Total Tables", 7],
        ["Average Length", 1.08, "m"],
    ]


@pytest.fixture
def single():
    return parse_cut_plan_grid(_single_style_grid())


def test_unnumbered_style_file_and_name_are_one_style(single):
    """Numbering them by arrival order split the pair across two slots — one
    style with only a filename, one with only a name."""
    assert single["styles"] == [
        {"name": "S25DDR1237", "file": r"D:\x\S25DDR1237-XL.PDS"}]


def test_single_style_demands_carry_the_style_name(single):
    """The name comes from the header block; there is no band to read it off."""
    assert len(single["demands"]) == 5
    assert {d["style"] for d in single["demands"]} == {"S25DDR1237"}
    assert {(d["size"], d["qty"]) for d in single["demands"]} == {
        ("XS", 269), ("S", 727), ("M", 1033), ("L", 865), ("XL", 563)}
    assert single["colors"] == ["Black"]


def test_single_style_plan_qty_is_the_ordered_total(single):
    """Was 0: with no style band the demand rows never parsed at all."""
    assert single["style_totals"] == {"S25DDR1237": 3457}
    assert single["total_qty"] == 3457


def test_single_style_cut_qty_is_the_achieved_total(single):
    """Was the sum of two mis-keyed size buckets rather than the units cut."""
    assert [m["cut_qty"] for m in single["materials"]] == [3499]


def test_single_style_size_groups_are_read_in_order(single):
    assert single["materials"][0]["groups"] == [
        ("S25DDR1237", ["XS", "S", "M", "L", "XL"])]


def test_plies_column_is_not_read_as_a_size(single):
    """Colors / Plies number share the size row when there is no band, so
    without a left bound the ply count was filed as a garment quantity."""
    pieces = single["materials"][0]["spreads"][0]["rows"][0]["pieces"]
    assert {p["size"] for p in pieces} == {"XS", "M", "L", "XL"}
    assert single["materials"][0]["spreads"][0]["rows"][0]["plies"] == 110


def test_single_style_summary_matches_the_saved_columns(single):
    s = summarise_plan(single)
    assert s["order_qty"] == 3457
    assert s["cut_qty"] == 3499
    assert s["styles"] == "S25DDR1237"


def test_multi_style_geometry_is_untouched(exact, overcut):
    """The band still governs where a plan has one — these are the layouts
    that already worked."""
    assert exact["style_totals"] == {"S24DTR003": 500, "ZLD060": 500}
    assert [m["cut_qty"] for m in exact["materials"]] == [1000]
    assert overcut["style_totals"] == {"S24DTR003": 946}
    assert [m["cut_qty"] for m in overcut["materials"]] == [950]


# ── Re-reading a stored plan ────────────────────────────────────────────────

def test_reparse_refreshes_the_saved_figures(store, single, exact):
    """A parser fix otherwise reaches only plans uploaded after it. Simulated
    by saving one plan's figures against another's file."""
    from po_extractor.exporters.cutting_plan_export import (
        build_standard_cut_plan, plan_header_from_parsed,
    )
    groups = single["materials"][0]["groups"]
    demand = {(d["style"], d["color"], d["size"]): d["qty"]
              for d in single["demands"]}
    xlsx = build_standard_cut_plan(
        header=plan_header_from_parsed(single), groups=groups,
        colors=single["colors"], demand_qty=demand,
        materials=single["materials"])

    pid = store.save_plan(exact, source_file="a.xlsx", file_bytes=xlsx)
    assert store.get_plan(pid)["order_qty"] == 500      # the wrong figures

    assert store.reparse_plan(pid) is True
    assert store.get_plan(pid)["order_qty"] == 3457     # read from the file
    assert list(store.list_plan_materials(pid)["material"]) == ["Shell"]


def test_reparse_leaves_the_row_alone_when_there_is_no_stored_file(store, exact):
    pid = store.save_plan(exact, source_file="a.xlsx")   # no file_bytes
    assert store.reparse_plan(pid) is False
    assert store.get_plan(pid)["order_qty"] == 500


def test_reparse_keeps_the_links(store, exact):
    xlsx = build_standard_cut_plan(
        header=today_header("x"), groups=[("S24DTR003", ["S"])],
        colors=["Wine"], demand_qty={("S24DTR003", "Wine", "S"): 10})
    pid = store.save_plan(exact, source_file="a.xlsx", file_bytes=xlsx,
                          links=[{"pc_no": "PC1", "po_no": "PO1",
                                  "style": "TP5016"}])
    store.reparse_plan(pid)
    assert [l["pc_no"] for l in store.get_plan(pid)["links"]] == ["PC1"]


def test_a_plan_saved_with_no_links_is_fully_usable(store, exact):
    """The upload screen offers "save without linking to a PO" for a plan
    with no matching PO yet — passing links=[] must behave exactly like
    omitting the argument, not like an error or a half-saved row."""
    pid = store.save_plan(exact, source_file="a.xlsx", links=[])
    assert store.get_plan(pid)["links"] == []
    row = store.list_plans_by_material().iloc[0]
    assert row["id"] == pid and row["po_qty"] == 0
    # It can still be linked afterwards, same as a plan saved with no
    # links argument at all (the pre-existing path).
    store.set_links(pid, [{"pc_no": "PC1", "po_no": "PO1", "style": "TP5016"}])
    assert [l["pc_no"] for l in store.get_plan(pid)["links"]] == ["PC1"]


def test_blank_template_build_does_not_crash_on_style_summary(monkeypatch):
    """Building the standard template with NO linked cut plan used to clobber
    the ``styles`` parameter (the user's selected PO style codes) with header
    dicts, and the style_summary join then raised TypeError on every blank
    build. The header now uses its own local; the summary keeps the codes."""
    import ui.cutting_plan.standard as std

    class _St:
        def __init__(self):
            self.session_state = {}
        def success(self, *a, **k): pass

    fake = _St()
    monkeypatch.setattr(std, "st", fake)
    monkeypatch.setattr(std, "save_copy_to_folder", lambda *a, **k: None)
    monkeypatch.setattr(std, "pdf_export_invalidate", lambda *a, **k: None)

    std._build(store=None, plan_ids=[], pc_nos=["PC1"], po_nos=[],
               styles=["TP5016"], items=None,
               groups=[("S24DTR003", ["S", "M"])], colors=["Wine"],
               qty={("S24DTR003", "Wine", "S"): 300,
                    ("S24DTR003", "Wine", "M"): 200},
               clean=False, folder="")

    from ui.session_keys import SK
    data = fake.session_state[SK.CP_STD_BYTES]
    assert data and fake.session_state[SK.CP_STD_FNAME].endswith(".xlsx")
    ws = load_workbook(io.BytesIO(data)).active
    texts = {str(ws.cell(r, c).value) for r in range(1, 15)
             for c in range(1, 6) if ws.cell(r, c).value is not None}
    assert "TP5016" in texts          # the user's style codes, not repr(dict)
