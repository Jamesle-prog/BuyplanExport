"""Tests for the worksheet → PDF renderer.

The contract the cutting room depends on: every column lands on one page
width, page edges are minimal, and nothing is silently dropped — not a CJK
colour name, not a metric header, not a marker's file path.
"""
from __future__ import annotations

import io

import fitz
import pytest
from openpyxl import Workbook, load_workbook

from po_extractor.exporters.xlsx_to_pdf import (
    MINIMAL_MARGIN_PT, PAGE_SIZES, PdfRenderError, _format_value, _runs,
    plan_layout, workbook_to_pdf, xlsx_bytes_to_pdf,
)


def _book(rows, *, widths=None, merges=None, title="Sheet1") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = title
    for r, row in enumerate(rows, start=1):
        for c, value in enumerate(row, start=1):
            if value is not None:
                ws.cell(row=r, column=c, value=value)
    for letter, width in (widths or {}).items():
        ws.column_dimensions[letter].width = width
    for rng in (merges or []):
        ws.merge_cells(rng)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _pdf_text(data: bytes) -> str:
    doc = fitz.open(stream=data, filetype="pdf")
    try:
        return "".join(page.get_text() for page in doc)
    finally:
        doc.close()


def _pages(data: bytes) -> int:
    doc = fitz.open(stream=data, filetype="pdf")
    try:
        return doc.page_count
    finally:
        doc.close()


def _content_extent(data: bytes, page_no: int = 0) -> tuple[float, float, float]:
    """(min_x, max_x, page_width) of the drawn words on a page."""
    doc = fitz.open(stream=data, filetype="pdf")
    try:
        page = doc[page_no]
        words = page.get_text("words")
        if not words:
            return (0.0, 0.0, page.rect.width)
        return (min(w[0] for w in words), max(w[2] for w in words),
                page.rect.width)
    finally:
        doc.close()


# ── The core promise: columns fit the page width ────────────────────────────

@pytest.mark.parametrize("n_cols", [5, 20, 40, 80])
def test_every_column_fits_one_page_width(n_cols):
    rows = [[f"Header{c}" for c in range(n_cols)],
            [c * 1000 for c in range(n_cols)]]
    data = workbook_to_pdf(io.BytesIO(_book(rows)))
    min_x, max_x, page_w = _content_extent(data)
    assert min_x >= MINIMAL_MARGIN_PT - 0.5
    assert max_x <= page_w - MINIMAL_MARGIN_PT + 0.5


def test_wide_sheet_is_one_page_across_not_split():
    rows = [[f"C{c}" for c in range(60)], [c for c in range(60)]]
    data = workbook_to_pdf(io.BytesIO(_book(rows)))
    # 60 narrow columns and two rows: width is the only constraint, so a
    # correct fit-to-width render is a single page.
    assert _pages(data) == 1


def test_long_sheet_flows_onto_more_pages():
    rows = [[f"row {i}", i] for i in range(400)]
    data = workbook_to_pdf(io.BytesIO(_book(rows)))
    assert _pages(data) > 1


def test_margins_are_minimal_by_default():
    data = workbook_to_pdf(io.BytesIO(_book([["A", "B"], [1, 2]])))
    min_x, _max_x, _w = _content_extent(data)
    # 2 mm plus the in-cell padding — an order of magnitude tighter than
    # Excel's 'Normal' 1.9 cm margins.
    assert min_x < 12.0


def test_margin_is_configurable():
    book = _book([["A", "B"], [1, 2]])
    wide = workbook_to_pdf(io.BytesIO(book), margin_pt=72.0)
    assert _content_extent(wide)[0] >= 72.0


def test_landscape_is_the_default_orientation():
    data = workbook_to_pdf(io.BytesIO(_book([["A"], [1]])))
    doc = fitz.open(stream=data, filetype="pdf")
    try:
        assert doc[0].rect.width > doc[0].rect.height
        assert round(doc[0].rect.width) == round(max(PAGE_SIZES["A4"]))
    finally:
        doc.close()


def test_portrait_and_page_size_are_honoured():
    data = workbook_to_pdf(io.BytesIO(_book([["A"], [1]])),
                           orientation="portrait", page_size="A3")
    doc = fitz.open(stream=data, filetype="pdf")
    try:
        assert doc[0].rect.height > doc[0].rect.width
        assert round(doc[0].rect.height) == round(max(PAGE_SIZES["A3"]))
    finally:
        doc.close()


def test_a3_gives_a_wide_sheet_a_larger_font_than_a4():
    rows = [[f"Header {c}" for c in range(40)], [c for c in range(40)]]
    ws = load_workbook(io.BytesIO(_book(rows))).worksheets[0]
    a4 = plan_layout(ws, avail_w=max(PAGE_SIZES["A4"]) - 2 * MINIMAL_MARGIN_PT)
    a3 = plan_layout(ws, avail_w=max(PAGE_SIZES["A3"]) - 2 * MINIMAL_MARGIN_PT)
    assert a3["font_size"] > a4["font_size"]


def test_narrow_sheet_is_not_scaled_up():
    ws = load_workbook(io.BytesIO(_book([["A", "B"], [1, 2]]))).worksheets[0]
    layout = plan_layout(ws, avail_w=2000.0)
    assert layout["scale"] == 1.0


# ── Nothing gets silently dropped ───────────────────────────────────────────

def test_cjk_text_is_rendered_not_replaced():
    data = workbook_to_pdf(io.BytesIO(_book([["颜色", "面料"], ["酒红", "里"]])))
    text = _pdf_text(data)
    for word in ("颜色", "面料", "酒红", "里"):
        assert word in text


def test_mixed_latin_and_cjk_splits_into_separate_font_runs():
    # One run per script, so Latin isn't set full-width in the CJK face.
    runs = _runs(r"D:\LJ\S24DTR003-裤子.PDS", bold=False)
    assert len(runs) == 3
    assert [r[0] for r in runs] == [r"D:\LJ\S24DTR003-", "裤子", ".PDS"]
    assert runs[0][1] == runs[2][1] != runs[1][1]


def test_vertically_merged_header_is_not_truncated():
    """A merge spanning only rows still occupies one column, so autofit must
    size that column for its text — this is the 'Material Cost,CNY' → 'Mat…'
    regression."""
    rows = [["Material Cost,CNY", "Efficiency,%"], [None, None], [1.0, 2.0]]
    data = workbook_to_pdf(io.BytesIO(
        _book(rows, widths={"A": 4.0, "B": 4.0}, merges=["A1:A2", "B1:B2"])))
    text = _pdf_text(data)
    assert "Material Cost,CNY" in text
    assert "Efficiency,%" in text
    assert "…" not in text


def test_content_wider_than_the_stored_column_is_still_shown():
    """The Optitex export ships 12-character columns holding 50-character
    marker paths; the printout must not inherit that clipping."""
    path = r"D:\LJ\zalando\060_003\20260723\S24DTR003-A1.mrk"
    data = workbook_to_pdf(io.BytesIO(
        _book([["File Name", "Qty"], [path, 176]], widths={"A": 12.625})))
    assert path in _pdf_text(data)


def test_text_spills_over_empty_neighbours_like_excel():
    long_text = "A very long label that needs several columns of room"
    data = workbook_to_pdf(io.BytesIO(
        _book([[long_text, None, None, None], ["x", 1, 2, 3]],
              widths={"A": 6.0})))
    assert long_text in _pdf_text(data)


def test_an_extreme_value_cannot_starve_the_other_columns():
    rows = [["ID", "Note", "Qty"], [1, "x" * 4000, 999]]
    data = workbook_to_pdf(io.BytesIO(_book(rows)))
    text = _pdf_text(data)
    # The outlier is capped and clipped, but its neighbours survive intact.
    assert "999" in text and "ID" in text and "Qty" in text


# ── Value formatting ────────────────────────────────────────────────────────

@pytest.mark.parametrize("value,fmt,expected", [
    (1057, "General", "1057"),
    (87.71107699120981, "0.00", "87.71"),
    (0.9558295391, "0.0000", "0.9558"),
    (2020.6236455879, "General", "2020.623646"),
    (157, "0", "157"),
    (1022.16, "#,##0.00", "1,022.16"),
    (0.8771, "0.0%", "87.7%"),
    (None, "General", ""),
    (True, "General", "TRUE"),
    ("酒红", "General", "酒红"),
])
def test_format_value(value, fmt, expected):
    assert _format_value(value, fmt) == expected


def test_general_format_does_not_show_float_noise():
    assert _format_value(0.1 + 0.2, "General") == "0.3"


# ── Robustness ──────────────────────────────────────────────────────────────

def test_empty_sheet_still_produces_a_page():
    data = workbook_to_pdf(io.BytesIO(_book([[None]], title="Blank")))
    assert _pages(data) == 1
    assert "Blank" in _pdf_text(data)


def test_styled_but_valueless_cells_do_not_stretch_the_scale():
    """max_row/max_column count touched cells; using them would squeeze a
    two-column sheet as if it were hundreds wide."""
    wb = Workbook()
    ws = wb.active
    ws["A1"], ws["B1"] = "A", "B"
    ws.cell(row=1, column=200).value = None
    ws.cell(row=200, column=200).number_format = "0.00"
    buf = io.BytesIO()
    wb.save(buf)
    layout = plan_layout(load_workbook(io.BytesIO(buf.getvalue())).worksheets[0],
                         avail_w=800.0)
    assert layout["last_col"] == 2
    assert layout["scale"] == 1.0


def test_each_sheet_starts_on_its_own_page():
    wb = Workbook()
    wb.active.title = "One"
    wb.active["A1"] = "first"
    second = wb.create_sheet("Two")
    second["A1"] = "second"
    buf = io.BytesIO()
    wb.save(buf)
    data = workbook_to_pdf(io.BytesIO(buf.getvalue()))
    assert _pages(data) == 2
    assert "first" in _pdf_text(data) and "second" in _pdf_text(data)


def test_sheet_selection():
    wb = Workbook()
    wb.active.title = "One"
    wb.active["A1"] = "first"
    wb.create_sheet("Two")["A1"] = "second"
    buf = io.BytesIO()
    wb.save(buf)
    data = workbook_to_pdf(io.BytesIO(buf.getvalue()), sheets=["Two"])
    assert _pages(data) == 1
    assert "second" in _pdf_text(data) and "first" not in _pdf_text(data)


def test_unknown_sheet_name_is_an_error():
    with pytest.raises(PdfRenderError):
        workbook_to_pdf(io.BytesIO(_book([["A"], [1]])), sheets=["Nope"])


def test_unreadable_input_raises_pdf_render_error():
    with pytest.raises(PdfRenderError):
        xlsx_bytes_to_pdf(b"this is not a workbook")


# ── The real thing: a cutting plan end to end ───────────────────────────────

def test_standard_cut_plan_renders_with_every_column_on_the_page():
    from po_extractor.exporters.cutting_plan_export import (
        build_standard_cut_plan, today_header,
    )

    groups = [(f"STY{i}", ["XS", "S", "M", "L", "XL"]) for i in range(6)]
    colors = ["酒红", "黑色"]
    qty = {(style, color, size): 100
           for style, sizes in groups for color in colors for size in sizes}
    xlsx = build_standard_cut_plan(
        header=today_header("HHPPC052", client="ZR"),
        groups=groups, colors=colors, demand_qty=qty, materials=[])

    data = xlsx_bytes_to_pdf(xlsx)
    min_x, max_x, page_w = _content_extent(data)
    assert max_x <= page_w - MINIMAL_MARGIN_PT + 0.5
    text = _pdf_text(data)
    for probe in ("Order demands", "Marker Definition", "酒红", "黑色",
                  "STY0", "STY5"):
        assert probe in text
