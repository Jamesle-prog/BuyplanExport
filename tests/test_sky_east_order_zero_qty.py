"""Sky East order parser: zero-unit rows are dropped and reported, not imported.

A cancelled/blanked-out order line (strikethrough style, PO, price wiped but
the row left in the sheet) has 0 total units. Importing it creates a phantom
item with no real quantity, so parse() drops it and records it in
SkyEastContract.skipped_zero_qty for the upload log instead.
"""
from __future__ import annotations

import openpyxl
import pytest

from po_extractor.parsers.sky_east_order import parse

# Column layout for the fixture header row (1-based).
_COLS = [
    "Item", "Style No.", "PO Number", "Config SKU", "Article Name", "Picture",
    "Fabric No.", "Fabrication", "Brand", "Color Name", "Color Code",
    "Launch Date", "XS", "S", "M", "L", "XL", "Total Qty", "FOB USD",
    "Total Cost", "Ex-Fty",
]
# Contract-header labels occupy rows 1-7 (label in col A, value in col E),
# well clear of the column-header/data block — parse() indexes the header
# dict directly (hdr["pc_date"], etc.) so every field needs a real label
# match here, not just a same-row fallback that a blank data cell can miss.
_CONTRACT_LABELS = [
    ("PC No.",        "PCTEST01"),
    ("Date",          "2026-01-01"),
    ("Party A",       "BuyerCo"),
    ("Party B",       "SellerCo"),
    ("Currency",      "USD"),
    ("Payment Terms", "TT"),
    ("Trade Term",    "FOB"),
]
_HEADER_ROW = 9
_DATA_START = 10


def _make_order_file(tmp_path, rows: list[dict]):
    wb = openpyxl.Workbook()
    ws = wb.active

    for ri, (label, val) in enumerate(_CONTRACT_LABELS, start=1):
        ws.cell(row=ri, column=1, value=label)
        ws.cell(row=ri, column=5, value=val)

    for ci, label in enumerate(_COLS, start=1):
        ws.cell(row=_HEADER_ROW, column=ci, value=label)

    col_idx = {label: i + 1 for i, label in enumerate(_COLS)}
    for ri, row in enumerate(rows, start=_DATA_START):
        for key, val in row.items():
            ws.cell(row=ri, column=col_idx[key], value=val)

    path = tmp_path / "order.xlsx"
    wb.save(str(path))
    wb.close()
    return str(path)


def test_zero_unit_row_is_dropped_and_reported(tmp_path):
    path = _make_order_file(tmp_path, [
        {"Item": 1, "Style No.": "STYLE-A", "PO Number": "PO001",
         "XS": 10, "S": 20, "M": 30, "L": 20, "XL": 20},
        {"Item": 2, "Style No.": "STYLE-B", "PO Number": "PO002",
         "Total Qty": 0},
    ])
    contract = parse(path)

    assert [i.style for i in contract.items] == ["STYLE-A"]
    assert contract.items[0].total_qty == 100

    assert len(contract.skipped_zero_qty) == 1
    skipped = contract.skipped_zero_qty[0]
    assert skipped["style"] == "STYLE-B"
    assert skipped["po"] == "PO002"


def test_row_with_all_size_columns_blank_is_also_dropped(tmp_path):
    # No explicit Total Qty and no size quantities at all — falls back to
    # sum(sizes) == 0, must be treated the same as an explicit 0.
    path = _make_order_file(tmp_path, [
        {"Item": 1, "Style No.": "STYLE-C", "PO Number": "PO003"},
    ])
    contract = parse(path)

    assert contract.items == []
    assert len(contract.skipped_zero_qty) == 1
    assert contract.skipped_zero_qty[0]["style"] == "STYLE-C"


def test_no_zero_qty_rows_leaves_skip_list_empty(tmp_path):
    path = _make_order_file(tmp_path, [
        {"Item": 1, "Style No.": "STYLE-A", "PO Number": "PO001",
         "XS": 5, "S": 5},
    ])
    contract = parse(path)

    assert len(contract.items) == 1
    assert contract.skipped_zero_qty == []
