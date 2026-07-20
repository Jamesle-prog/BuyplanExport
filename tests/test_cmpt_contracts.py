"""Tests for the CMPT contract module: store (contracts/lines/payments with
computed agreed/paid/balance), the RMB capital-amount util, and template-based
document generation ({{placeholder}} substitution + line-row replication).
"""
from __future__ import annotations

import io

import openpyxl
import pytest

from po_extractor.store.cmpt_contract_store import CmptContractStore
from po_extractor.utils.rmb_amount import rmb_capital
from po_extractor.exporters.cmpt_contract_doc import generate_cmpt_contract_xlsx


@pytest.fixture
def store(tmp_path):
    return CmptContractStore(str(tmp_path / "po_history.db"))


_LINES = [
    {"po_number": "PO1", "style": "STY1", "color": "Black",
     "description": "", "qty": 1000, "unit_price": 12.50},
    {"po_number": "PO2", "style": "STY2", "color": "Navy",
     "description": "", "qty": 500, "unit_price": 11.00},
]


# ── Store ───────────────────────────────────────────────────────────────────

def test_create_contract_computes_agreed_total(store):
    cid = store.create_contract("HT-2026-001", "Factory Alpha",
                                company="Sky East", contract_date="2026-07-19",
                                lines=_LINES, created_by="james")
    c = store.get_contract(cid)
    assert c["agreed_total"] == 1000 * 12.50 + 500 * 11.00   # 18,000
    assert c["total_qty"] == 1500
    assert c["paid_total"] == 0
    assert c["balance"] == 18000.0
    assert c["status"] == "draft"
    assert [ln["amount"] for ln in c["lines"]] == [12500.0, 5500.0]


def test_contract_no_must_be_unique(store):
    store.create_contract("HT-1", "F1")
    with pytest.raises(ValueError):
        store.create_contract("HT-1", "F2")


def test_contract_requires_no_and_factory(store):
    with pytest.raises(ValueError):
        store.create_contract("", "F1")
    with pytest.raises(ValueError):
        store.create_contract("HT-2", "")


def test_payment_log_drives_paid_and_balance(store):
    cid = store.create_contract("HT-1", "F1", lines=_LINES)
    store.add_payment(cid, "2026-07-20", 10000, method="电汇",
                      recorded_by="james")
    store.add_payment(cid, "2026-08-01", 5000)

    c = store.get_contract(cid)
    assert c["paid_total"] == 15000
    assert c["balance"] == 3000.0
    assert len(c["payments"]) == 2

    # Negative amount = refund; zero rejected.
    store.add_payment(cid, "2026-08-05", -2000, note="refund")
    assert store.get_contract(cid)["paid_total"] == 13000
    with pytest.raises(ValueError):
        store.add_payment(cid, "2026-08-06", 0)


def test_replace_lines_updates_agreed(store):
    cid = store.create_contract("HT-1", "F1", lines=_LINES)
    store.replace_lines(cid, [{"po_number": "PO9", "style": "S9",
                               "qty": 100, "unit_price": 5.0}])
    c = store.get_contract(cid)
    assert c["agreed_total"] == 500.0
    assert len(c["lines"]) == 1


def test_update_status_validated_and_unknown_field_rejected(store):
    cid = store.create_contract("HT-1", "F1")
    store.update_contract(cid, status="signed", updated_by="james")
    assert store.get_contract(cid)["status"] == "signed"
    with pytest.raises(ValueError):
        store.update_contract(cid, status="bogus")
    with pytest.raises(ValueError):
        store.update_contract(cid, contract_no="NEW")   # not an allowed field


def test_delete_contract_cascades(store):
    cid = store.create_contract("HT-1", "F1", lines=_LINES)
    store.add_payment(cid, "2026-07-20", 100)
    store.delete_contract(cid)
    assert store.get_contract(cid) is None
    assert store.list_contracts() == []


def test_list_contracts_filters(store):
    store.create_contract("HT-1", "F1")
    c2 = store.create_contract("HT-2", "F2")
    store.update_contract(c2, status="signed")
    assert len(store.list_contracts()) == 2
    assert len(store.list_contracts(factory="F1")) == 1
    assert len(store.list_contracts(status="signed")) == 1
    assert store.list_factories() == ["F1", "F2"]


def test_next_contract_no_sequence(store):
    from datetime import datetime
    year = datetime.now().strftime("%Y")

    assert store.next_contract_no() == f"CMPT-{year}-001"

    store.create_contract(f"CMPT-{year}-001", "F1")
    store.create_contract(f"CMPT-{year}-002", "F1")
    assert store.next_contract_no() == f"CMPT-{year}-003"

    # Gaps are not refilled; other formats and other years are ignored.
    store.create_contract(f"CMPT-{year}-010", "F1")
    store.create_contract("HT-CUSTOM-99", "F1")
    store.create_contract("CMPT-1999-500", "F1")
    assert store.next_contract_no() == f"CMPT-{year}-011"


# ── RMB capital amounts ─────────────────────────────────────────────────────

@pytest.mark.parametrize("amount,expected", [
    (0,            "零元整"),
    (1,            "壹元整"),
    (10,           "壹拾元整"),
    (110,          "壹佰壹拾元整"),
    (1005,         "壹仟零伍元整"),
    (10001,        "壹万零壹元整"),
    (100000001,    "壹亿零壹元整"),
    (12345.67,     "壹万贰仟叁佰肆拾伍元陆角柒分"),
    (0.05,         "伍分"),
    (5.05,         "伍元零伍分"),
    (5.50,         "伍元伍角整"),
    (80808.08,     "捌万零捌佰零捌元零捌分"),
    (-12.30,       "负壹拾贰元叁角整"),
    (1000000,      "壹佰万元整"),
    (100200300,    "壹亿零贰拾万零叁佰元整"),
    (18000,        "壹万捌仟元整"),
])
def test_rmb_capital(amount, expected):
    assert rmb_capital(amount) == expected


# ── Document generation ─────────────────────────────────────────────────────

def _make_template() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "加工合同 Processing Contract {{contract_no}}"
    ws["A2"] = "甲方 Party A: {{company}}    乙方 Party B: {{factory}}"
    ws["A3"] = "签订日期 Date: {{contract_date}}"
    ws["A5"] = "{{line.no}}"
    ws["B5"] = "{{line.po}}"
    ws["C5"] = "{{line.style}}"
    ws["D5"] = "{{line.qty}}"
    ws["E5"] = "{{line.unit_price}}"
    ws["F5"] = "{{line.amount}}"
    ws["A7"] = "合计 Total: {{total_qty}} 件, {{currency}} {{total_amount}}"
    ws["A8"] = "大写: {{total_amount_cn}}"
    ws["A10"] = "甲方签章：＿＿＿＿＿＿    乙方签章：＿＿＿＿＿＿"
    ws["A11"] = "{{unknown_token}}"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _contract_dict(store):
    cid = store.create_contract(
        "HT-2026-001", "Factory Alpha", company="Sky East",
        contract_date="2026-07-19", lines=_LINES,
    )
    return store.get_contract(cid)


def test_generation_fills_header_and_replicates_lines(store):
    doc = generate_cmpt_contract_xlsx(_make_template(), _contract_dict(store))
    ws = openpyxl.load_workbook(io.BytesIO(doc)).active

    assert ws["A1"].value == "加工合同 Processing Contract HT-2026-001"
    assert "Sky East" in ws["A2"].value and "Factory Alpha" in ws["A2"].value
    assert ws["A3"].value == "签订日期 Date: 2026-07-19"

    # Two lines -> prototype row 5 + one inserted row 6.
    assert ws["B5"].value == "PO1" and ws["C5"].value == "STY1"
    assert ws["D5"].value == "1,000" and ws["F5"].value == "12,500.00"
    assert ws["B6"].value == "PO2" and ws["D6"].value == "500"
    assert ws["A5"].value == "1" and ws["A6"].value == "2"

    # Rows below shifted down by 1 (7->8, 8->9, 10->11).
    assert ws["A8"].value == "合计 Total: 1,500 件, RMB 18,000.00"
    assert ws["A9"].value == "大写: 壹万捌仟元整"
    assert "签章" in ws["A11"].value
    # Unknown tokens stay visible (typo detection), now at row 12.
    assert ws["A12"].value == "{{unknown_token}}"


def test_generation_single_line_no_row_insert(store):
    cid = store.create_contract("HT-2", "F1", lines=[_LINES[0]])
    doc = generate_cmpt_contract_xlsx(_make_template(), store.get_contract(cid))
    ws = openpyxl.load_workbook(io.BytesIO(doc)).active
    assert ws["B5"].value == "PO1"
    # No shift: totals stay at row 7.
    assert ws["A7"].value.startswith("合计")


def test_generation_rejects_unreadable_template(store):
    with pytest.raises(ValueError):
        generate_cmpt_contract_xlsx(b"not an xlsx", _contract_dict(store))
