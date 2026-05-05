"""Regression tests for the wash-label Excel generator."""
from __future__ import annotations

import io
from dataclasses import dataclass

import pandas as pd
import pytest

from po_extractor.ui_helpers.wash_label import write_wash_label_excel


@dataclass
class _FabricPart:
    seq: int
    body_part: str
    hhn_no: str
    composition: str | None = None


def _open(buf: bytes):
    from openpyxl import load_workbook
    return load_workbook(io.BytesIO(buf))


def test_returns_xlsx_bytes_with_header_row():
    df = pd.DataFrame([{"style": "S1", "fabric_item_no": "HHN-1", "composition_en": "100% Cotton"}])
    out = write_wash_label_excel(df, image_cache={})
    assert isinstance(out, bytes) and out[:2] == b"PK"
    wb = _open(out)
    ws = wb.active
    assert ws.title == "Wash Labels"
    assert [ws.cell(row=1, column=c).value for c in range(1, 7)] == [
        "Style", "Photo", "Seq", "Body Part", "Fabric Code", "Composition",
    ]


def test_fallback_uses_fabric_item_no_when_no_parts():
    df = pd.DataFrame([
        {"style": "S1", "fabric_item_no": "HHN-1", "composition_en": "100% Cotton"},
    ])
    wb = _open(write_wash_label_excel(df, image_cache={}))
    ws = wb.active
    assert ws.cell(row=2, column=1).value == "S1"
    assert ws.cell(row=2, column=3).value == "1"  # seq
    assert ws.cell(row=2, column=5).value == "HHN-1"
    assert ws.cell(row=2, column=6).value == "100% Cotton"


def test_fabric_parts_take_precedence_over_fallback():
    df = pd.DataFrame([{"style": "S1", "fabric_item_no": "FALLBACK-HHN"}])
    parts = {"S1": [
        _FabricPart(1, "Shell", "HHN-A", "100% Wool"),
        _FabricPart(2, "Lining", "HHN-B", "100% Polyester"),
    ]}
    wb = _open(write_wash_label_excel(df, image_cache={}, fabric_parts_by_style=parts))
    ws = wb.active
    # row 2: first part, with style label
    assert ws.cell(row=2, column=1).value == "S1"
    assert ws.cell(row=2, column=4).value == "Shell"
    assert ws.cell(row=2, column=5).value == "HHN-A"
    # row 3: second part, style label suppressed (openpyxl stores "" as None)
    assert ws.cell(row=3, column=1).value in ("", None)
    assert ws.cell(row=3, column=5).value == "HHN-B"


def test_nan_composition_renders_blank():
    df = pd.DataFrame([{"style": "S1", "fabric_item_no": "HHN-1", "composition_en": "nan"}])
    wb = _open(write_wash_label_excel(df, image_cache={}))
    ws = wb.active
    assert ws.cell(row=2, column=6).value in ("", None)


def test_part_without_hhn_is_skipped():
    df = pd.DataFrame([{"style": "S1"}])
    parts = {"S1": [
        _FabricPart(1, "Shell", "", None),
        _FabricPart(2, "Lining", "HHN-X", "100% Cotton"),
    ]}
    wb = _open(write_wash_label_excel(df, image_cache={}, fabric_parts_by_style=parts))
    ws = wb.active
    # only one data row should be written (the one with hhn)
    assert ws.cell(row=2, column=5).value == "HHN-X"
    assert ws.cell(row=3, column=5).value is None


def test_multiple_styles_preserve_input_order():
    df = pd.DataFrame([
        {"style": "ALPHA", "fabric_item_no": "A"},
        {"style": "BETA",  "fabric_item_no": "B"},
        {"style": "ALPHA", "fabric_item_no": "A"},  # duplicate, should be ignored
    ])
    wb = _open(write_wash_label_excel(df, image_cache={}))
    ws = wb.active
    assert ws.cell(row=2, column=1).value == "ALPHA"
    assert ws.cell(row=3, column=1).value == "BETA"


def test_freeze_panes_set_at_a2():
    df = pd.DataFrame([{"style": "S1"}])
    wb = _open(write_wash_label_excel(df, image_cache={}))
    assert wb.active.freeze_panes == "A2"
