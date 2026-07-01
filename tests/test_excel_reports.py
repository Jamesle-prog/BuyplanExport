"""Regression tests for color-plan and PO-summary report generators."""
import io

import openpyxl
import pandas as pd

from po_extractor.ui_helpers.excel_reports import (
    SIZE_ORDER, generate_color_plan_excel, generate_po_summary_excel,
)


# ── Color plan ────────────────────────────────────────────────────────────────

def test_color_plan_empty_returns_empty_bytes():
    assert generate_color_plan_excel(pd.DataFrame()) == b""
    assert generate_color_plan_excel(None) == b""


def _size_rows():
    return pd.DataFrame([
        {"PO Number": "PO1", "Style": "S1", "Color": "Red", "Size": "S",  "Units": 5},
        {"PO Number": "PO1", "Style": "S1", "Color": "Red", "Size": "M",  "Units": 7},
        {"PO Number": "PO1", "Style": "S1", "Color": "Red", "Size": "L",  "Units": 3},
        {"PO Number": "PO1", "Style": "S1", "Color": "Blue", "Size": "M", "Units": 2},
    ])


def test_color_plan_pivots_sizes_to_columns():
    data = generate_color_plan_excel(_size_rows())
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["Color Plan"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    # Required identity cols
    assert headers[:3] == ["PO Number", "Style", "Color"]
    # Sizes follow standard order
    for sz in ("S", "M", "L"):
        assert sz in headers
    # Total at end
    assert headers[-1] == "Total"


def test_color_plan_total_column_correct():
    data = generate_color_plan_excel(_size_rows())
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["Color Plan"]
    # Find Total column index
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    total_col = headers.index("Total") + 1
    # Sum of values for "Red" row (S=5, M=7, L=3) → 15
    red_row = None
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 3).value == "Red":
            red_row = r
            break
    assert red_row is not None
    assert ws.cell(red_row, total_col).value == 15


def test_color_plan_size_columns_in_standard_order():
    df = pd.DataFrame([
        {"PO Number": "PO1", "Style": "S1", "Color": "R", "Size": "XL", "Units": 1},
        {"PO Number": "PO1", "Style": "S1", "Color": "R", "Size": "XS", "Units": 1},
        {"PO Number": "PO1", "Style": "S1", "Color": "R", "Size": "M",  "Units": 1},
    ])
    data = generate_color_plan_excel(df)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["Color Plan"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    sizes_present = [h for h in headers if h in SIZE_ORDER]
    # Order matches SIZE_ORDER
    assert sizes_present == ["XS", "M", "XL"]


# ── PO summary ────────────────────────────────────────────────────────────────
#
# generate_po_summary_excel() emits a fixed 3-section "PO Summary" sheet, not
# a simple one-row-per-input-column table:
#   Row 1              : title ("PO Summary - <vendor>")
#   Rows 2-13          : a 12-row Vendor/Factory/.../HTS key-value metadata block
#   Row 15             : the detail table header (row 14 is a spacer)
#   Row 16+            : one data row per PO (or per size-broken-down colour)
# The detail columns are a FIXED set (PO Number, Style, Color Code, Color
# Name, FOB, Cost P/U, <sizes>, Total Units, Extended Cost) written
# unconditionally with blank/zero fallbacks for missing input data — the
# function no longer derives columns dynamically from whatever happens to be
# in df_pos. label_for is accepted for backward compatibility but the current
# implementation hardcodes English labels everywhere and never calls it.

_DETAIL_HDR_ROW = 15   # title (1) + 12 metadata rows (2-13) + spacer (14)


def _pos_df():
    return pd.DataFrame([
        {"company": "G3", "po_number": "PO1", "style": "STY1",
         "factory": "F1", "country_of_origin": "CN",
         "xport_date": "2026-01-01", "issue_date": "2026-01-01",
         "version": 1, "division_name": "DKNY",
         "total_units": 100, "source_format": "infor", "extracted_at": "2026-01-02"},
    ])


def test_po_summary_uses_fallback_labels_when_no_resolver():
    """Title row uses the vendor fallback chain, and the detail table header
    (fixed English labels) appears at the expected row."""
    data = generate_po_summary_excel(_pos_df())
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["PO Summary"]
    assert ws.cell(1, 1).value == "PO Summary - G3"
    detail_headers = [ws.cell(_DETAIL_HDR_ROW, c).value for c in range(1, ws.max_column + 1)]
    assert "PO Number" in detail_headers
    assert "Style" in detail_headers
    assert "Total Units" in detail_headers


def test_po_summary_label_for_param_is_currently_a_no_op():
    """label_for is still accepted (backward-compatible signature) but the
    current implementation never calls it -- passing a custom resolver must
    not change the output or raise.
    """
    def my_label(db_col, fallback):
        return {"company": "公司", "po_number": "采购单号"}.get(db_col, fallback)

    default_data = generate_po_summary_excel(_pos_df())
    custom_data  = generate_po_summary_excel(_pos_df(), label_for=my_label)
    assert default_data == custom_data


def test_po_summary_always_emits_full_fixed_detail_columns():
    """Even a minimal 2-column input still gets the complete fixed detail
    table (missing fields render blank/zero, columns aren't pruned)."""
    df = pd.DataFrame([{"company": "G3", "po_number": "PO1"}])
    data = generate_po_summary_excel(df)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["PO Summary"]
    detail_headers = [ws.cell(_DETAIL_HDR_ROW, c).value for c in range(1, ws.max_column + 1)]
    assert detail_headers[:6] == [
        "PO Number", "Style", "Color Code", "Color Name", "FOB", "Cost P/U",
    ]
    assert detail_headers[-2:] == ["Total Units", "Extended Cost"]


def test_po_summary_data_row_values_preserved():
    data = generate_po_summary_excel(_pos_df())
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["PO Summary"]
    data_row = _DETAIL_HDR_ROW + 1
    assert ws.cell(data_row, 1).value == "PO1"    # PO Number
    assert ws.cell(data_row, 2).value == "STY1"   # Style
