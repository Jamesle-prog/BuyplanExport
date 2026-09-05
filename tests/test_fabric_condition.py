"""Tests for the actual fabric condition module (面料情况).

The source is a hand-kept shop-floor log, not a structured export — headers
repeat ambiguously, and the "numeric" columns are full of ranges,
approximations and notes instead of clean numbers. The fixtures below
reproduce those specifically, since they're exactly what would break a
parser that assumed clean, single-format data.
"""
from __future__ import annotations

import io

import openpyxl
import pytest

from po_extractor.parsers.fabric_condition import (
    FabricConditionParseError, parse_fabric_condition,
)
from po_extractor.store.fabric_condition_store import FabricConditionStore


def _sheet(ws, rows: list[list]) -> None:
    """Write the sheet's real two-row header, then *rows* of data."""
    ws["J1"] = "面料缩率"
    ws.merge_cells("J1:K1")
    ws["L1"] = "纸板缩率"
    ws.merge_cells("L1:M1")
    head = ["序号", "日期", "面料编号", "款号", "部位", "颜色",
            "有效门幅 (cm)", "毛门幅(cm)", "克重(g)",
            "径向 (length)", "纬向", "径向", "纬向",
            "超裁%", "裁剪最高层数", "裁剪最大长度", "裁剪方向及要求",
            "采购净单耗(cm) -大身", "采购净单耗(cm) -滚条",
            "排版净单耗(cm) -大身", "排版净单耗(cm) -滚条",
            "剩余面料(匹）", "剩余面料(kg）", "剩余面料(m）"]
    for c, v in enumerate(head, start=1):
        ws.cell(2, c, v)
    for i, row in enumerate(rows, start=3):
        for c, v in enumerate(row, start=1):
            if v is not None:
                ws.cell(i, c, v)


def _book(rows: list[list]) -> io.BytesIO:
    wb = openpyxl.Workbook()
    wb.active.title = "1.面料情况"
    _sheet(wb.active, rows)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _row(seq=None, date=None, fabric_no="", style="", body_part="",
        color="", width_net=None, width_gross=None, weight=None,
        shrink_fw=None, shrink_ffw=None, shrink_pw=None, shrink_pfw=None,
        remaining_m=None) -> list:
    return [seq, date, fabric_no, style, body_part, color, width_net,
           width_gross, weight, shrink_fw, shrink_ffw, shrink_pw,
           shrink_pfw, None, None, None, None, None, None, None, None,
           None, None, remaining_m]


@pytest.fixture
def store(tmp_path):
    FabricConditionStore._forget_schema()
    return FabricConditionStore(str(tmp_path / "po_history.db"))


# ── Parser: the two ambiguous 径向/纬向 pairs ────────────────────────────────

def test_the_two_shrinkage_pairs_are_told_apart_by_their_row1_group():
    """径向/纬向 read identically on their own row — only the merged group
    heading above them (面料缩率 vs 纸板缩率) says which is which."""
    rows = [_row(style="S1", color="Black",
                shrink_fw=0.01, shrink_ffw=0.02,
                shrink_pw=0.03, shrink_pfw=0.04)]
    rec = parse_fabric_condition(_book(rows))["records"][0]
    assert (rec["shrink_fabric_warp"], rec["shrink_fabric_weft"]) == ("0.01", "0.02")
    assert (rec["shrink_pattern_warp"], rec["shrink_pattern_weft"]) == ("0.03", "0.04")


def test_the_length_suffix_on_one_heading_does_not_break_the_match():
    """The fabric-shrinkage 径向 column is headed '径向 (length)' — an extra
    annotation the pattern-shrinkage column's plain '径向' doesn't have."""
    rows = [_row(style="S1", color="Black", shrink_fw=0.05)]
    rec = parse_fabric_condition(_book(rows))["records"][0]
    assert rec["shrink_fabric_warp"] == "0.05"


# ── Parser: everything is kept as text, verbatim ─────────────────────────────

def test_a_range_is_kept_whole_not_truncated_to_one_number():
    """'176-178' as a float would silently become 176 or fail outright —
    either way the second reading is gone."""
    rows = [_row(style="S1", color="Black", width_net="176-178",
                weight="184-175")]
    rec = parse_fabric_condition(_book(rows))["records"][0]
    assert rec["width_net"] == "176-178"
    assert rec["weight_gsm"] == "184-175"


@pytest.mark.parametrize("raw", ["同上", "无", "/", "100层左右", "负0.6%",
                                 "二次测试待定"])
def test_a_word_instead_of_a_number_survives_untouched(raw):
    """'同上' (ditto), '无' (none), '负0.6%' (a typo'd minus sign) are never
    resolved or coerced here — that would be a guess this parser has no
    business making."""
    rows = [_row(style="S1", color="Black", shrink_fw=raw)]
    rec = parse_fabric_condition(_book(rows))["records"][0]
    assert rec["shrink_fabric_warp"] == raw


def test_style_keeps_a_leading_zero_when_the_cell_is_already_text():
    rows = [_row(style="0063", color="Blue")]
    assert parse_fabric_condition(_book(rows))["records"][0]["style"] == "0063"


def test_a_bare_int_style_is_still_text_not_a_number():
    """Some rows type the style as a literal number (6122); it must come
    back as a string like every other style value, not an int."""
    rows = [_row(style=6122, color="Black")]
    rec = parse_fabric_condition(_book(rows))["records"][0]
    assert rec["style"] == "6122" and isinstance(rec["style"], str)


# ── Parser: the date column ─────────────────────────────────────────────────

def test_a_clean_excel_serial_becomes_an_iso_date():
    """The column's cells are General-formatted, not date-formatted — a bare
    int here is a date that lost its formatting, not a different value."""
    rows = [_row(date=44872, style="S1", color="Black")]
    assert parse_fabric_condition(_book(rows))["records"][0]["test_date"] \
        == "2022-11-07"


def test_a_malformed_date_is_kept_as_typed():
    """A real example from the sheet: '2023/215' — not a parseable date, and
    not silently dropped either."""
    rows = [_row(date="2023/215", style="S1", color="Black")]
    assert parse_fabric_condition(_book(rows))["records"][0]["test_date"] \
        == "2023/215"


def test_a_blank_date_is_a_blank_string_not_none():
    rows = [_row(style="S1", color="Black")]
    assert parse_fabric_condition(_book(rows))["records"][0]["test_date"] == ""


# ── Parser: which rows count as records ─────────────────────────────────────

def test_a_fully_blank_row_is_not_a_record():
    rows = [_row(style="S1", color="Black"), _row(), _row(style="S2", color="Blue")]
    recs = parse_fabric_condition(_book(rows))["records"]
    assert [r["style"] for r in recs] == ["S1", "S2"]


def test_seq_is_optional_and_can_repeat():
    """The tail of the real sheet stops assigning 序号 at all, and the one
    place it IS reused is two colourways of the same delivery."""
    rows = [_row(seq=7, style="S1", color="Blue"),
           _row(seq=7, style="S1", color="Black"),
           _row(style="S2", color="Red")]        # no seq at all
    recs = parse_fabric_condition(_book(rows))["records"]
    assert [r["seq"] for r in recs] == [7, 7, None]


def test_a_row_needs_at_least_one_identifying_field():
    """A row with only, say, a shrink reading and nothing identifying it is
    noise, not a record — matches how the settlement parser draws this line."""
    rows = [[None, None, None, None, None, None, None, None, None,
            0.01] + [None] * 14]                  # a lone shrink value
    assert parse_fabric_condition(_book(rows))["records"] == []


def test_a_workbook_with_no_header_row_is_rejected():
    wb = openpyxl.Workbook()
    wb.active["A1"] = "not a fabric condition sheet"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    with pytest.raises(FabricConditionParseError):
        parse_fabric_condition(buf)


def test_falls_back_to_the_first_sheet_when_the_named_one_is_missing():
    wb = openpyxl.Workbook()
    wb.active.title = "renamed"
    _sheet(wb.active, [_row(style="S1", color="Black")])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    assert parse_fabric_condition(buf)["records"][0]["style"] == "S1"


# ── Store ───────────────────────────────────────────────────────────────────

def test_import_replaces_the_whole_table(store):
    """Unlike settlement (several client-years in one file), this is one
    running log from one sheet — a re-import has nothing to preserve."""
    p1 = parse_fabric_condition(_book([_row(style="S1", color="Black")]))
    store.import_parsed(p1)
    assert store.count() == 1

    p2 = parse_fabric_condition(_book([_row(style="S2", color="Blue"),
                                       _row(style="S3", color="Red")]))
    store.import_parsed(p2)
    assert store.count() == 2
    assert list(store.list_records()["style"]) == ["S2", "S3"]


def test_filters_and_search(store):
    rows = [_row(style="S1", color="Black"), _row(style="S2", color="Blue")]
    store.import_parsed(parse_fabric_condition(_book(rows)))
    assert store.distinct("style") == ["S1", "S2"]
    assert len(store.list_records(styles=["S1"])) == 1
    assert list(store.list_records(search="Blue")["style"]) == ["S2"]
    assert store.list_records(search="nothing-matches").empty


def test_import_history_records_the_upload(store):
    store.import_parsed(
        parse_fabric_condition(_book([_row(style="S1", color="Black")])),
        source_file="面料情况.xlsx", file_bytes=b"x", imported_by="angel")
    row = store.list_imports().iloc[0]
    assert (row["source_file"], row["imported_by"], row["n_records"]) \
        == ("面料情况.xlsx", "angel", 1)


def test_an_empty_store_lists_cleanly(store):
    assert store.count() == 0
    assert store.list_records().empty
    assert store.distinct("style") == []
