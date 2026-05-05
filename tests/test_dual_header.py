"""Regression tests for the dual-header Excel writer."""
from __future__ import annotations

import io

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from po_extractor.ui_helpers.dual_header import (
    DUAL_HEADER_STATIC, get_dual_header, write_dual_header_excel,
)


class _FakeWriter:
    """Minimal stand-in for a pandas ExcelWriter — exposes .book."""
    def __init__(self):
        self.book = Workbook()
        # Drop the default sheet so create_sheet is the only one
        self.book.remove(self.book.active)


def _label_for(db_col, fallback):
    return f"LBL[{db_col}]"


def test_dual_header_static_shape():
    assert all(len(t) == 3 for t in DUAL_HEADER_STATIC)
    db_cols = [t[0] for t in DUAL_HEADER_STATIC]
    # spot-check the canonical Sky East fields
    for required in ("pc_no", "style", "color_name", "xs", "xxl", "total_qty"):
        assert required in db_cols


def test_get_dual_header_uses_label_for_callable():
    rows = get_dual_header(_label_for)
    style_row = next(r for r in rows if r[0] == "style")
    assert style_row[1] == "Main Supplier Config SKU"
    assert style_row[2] == "LBL[style]"


def test_write_dual_header_writes_three_rows():
    df = pd.DataFrame([
        {"style": "S1", "color_name": "Red", "xs": 10, "total_qty": 10},
    ])
    w = _FakeWriter()
    write_dual_header_excel(df, "Sheet1", w, label_for=_label_for)
    ws = w.book["Sheet1"]
    # Row 1 = client names
    row1 = [ws.cell(row=1, column=c).value for c in range(1, 5)]
    assert "Main Supplier Config SKU" in row1
    # Row 2 = standard labels via label_for
    row2 = [ws.cell(row=2, column=c).value for c in range(1, 5)]
    assert any(v and v.startswith("LBL[") for v in row2)
    # Row 3 = data
    row3 = [ws.cell(row=3, column=c).value for c in range(1, 5)]
    assert "S1" in row3 and "Red" in row3 and 10 in row3


def test_write_dual_header_only_includes_present_columns():
    df = pd.DataFrame([{"style": "S1", "color_name": "Red"}])
    w = _FakeWriter()
    write_dual_header_excel(df, "Sheet1", w, label_for=_label_for)
    ws = w.book["Sheet1"]
    # Should have 2 columns only (style + color_name), no xs/total_qty
    assert ws.cell(row=1, column=3).value is None


def test_write_dual_header_falls_back_when_label_for_none():
    df = pd.DataFrame([{"style": "S1"}])
    w = _FakeWriter()
    write_dual_header_excel(df, "Sheet1", w, label_for=None)
    ws = w.book["Sheet1"]
    # row2 falls back to row1 client name
    assert ws.cell(row=2, column=1).value == ws.cell(row=1, column=1).value


def test_write_dual_header_unknown_columns_use_passthrough():
    df = pd.DataFrame([{"foo": 1, "bar": 2}])
    w = _FakeWriter()
    write_dual_header_excel(df, "Sheet1", w, label_for=_label_for)
    ws = w.book["Sheet1"]
    # No DUAL_HEADER_STATIC fields present, so columns become foo/bar
    headers = sorted([ws.cell(row=1, column=c).value for c in range(1, 3)])
    assert headers == ["bar", "foo"]


def test_write_dual_header_no_photo_column_without_image_cache():
    df = pd.DataFrame([{"style": "S1", "picture_id": "P1"}])
    w = _FakeWriter()
    write_dual_header_excel(df, "Sheet1", w, image_cache=None, label_for=_label_for)
    ws = w.book["Sheet1"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, 5)]
    assert "Sample_Pic" not in headers


def test_write_dual_header_inserts_photo_column_after_style():
    df = pd.DataFrame([{"style": "S1", "color_name": "Red", "picture_id": "P1"}])
    w = _FakeWriter()
    write_dual_header_excel(df, "Sheet1", w, image_cache={"P1": b""}, label_for=_label_for)
    ws = w.book["Sheet1"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, 5)]
    # Sample_Pic should come right after style
    style_idx = headers.index("Main Supplier Config SKU")
    assert headers[style_idx + 1] == "Sample_Pic"
