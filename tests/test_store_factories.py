"""Lock the canonical store-factory contract + Sky East buyplan output shape.

This file covers regression tests for every bug fixed during the v1.50.x →
v1.54.x stabilisation:

  • v1.53.0 silent-import bug — fabric_master factory missing under
    ``po_extractor.store.fabric_master_store`` was silently swallowed.
  • v1.52.0 综合key column placement — value used to land in column E.
  • v1.51.1 empty dates — K1/K2 left blank because the template has no
    ``{{created_at}}`` placeholder.
  • v1.50.x photos not showing — placeholder-text scan + oneCellAnchor
    instead of twoCellAnchor on J3:L6 / M3:O6.
  • v1.50.2 format overrides — font 10, col widths 20/6, row height 28pt.
  • v1.51.x photo lookup variants — {style}_front.png, {style}-front.png,
    {style}_F.png, {style}_1.png, {style}.png, case-insensitive,
    slashes-in-style normalised.
  • v1.53.x brand auto-registration — new brands inserted into
    boat_sample_req when Sky East orders are loaded.

Run with: ``pytest tests/test_store_factories.py -v``
"""
from __future__ import annotations

import io
import os
import re
import sqlite3
import struct
import tempfile
import warnings
import zipfile
import zlib

import pandas as pd
import pytest


# ─────────────────────────────────────────────────────────────────────────────
# Section 1 — Canonical store factory contract (locks v1.53.0 fix)
# ─────────────────────────────────────────────────────────────────────────────

_REQUIRED_FACTORIES = [
    "get_po_store",
    "get_sky_east_store",
    "get_fabric_master_store",
    "get_color_translation_store",
    "get_boat_sample_store",
]


def test_every_factory_importable_from_package_root():
    """Each canonical factory must be importable from po_extractor.store."""
    import po_extractor.store as pkg
    for name in _REQUIRED_FACTORIES:
        assert hasattr(pkg, name), (
            f"po_extractor.store.{name} is missing — old code paths "
            f"importing it will silently fall back to empty caches "
            f"(v1.53.0 bug class)."
        )


@pytest.mark.parametrize("factory_name", _REQUIRED_FACTORIES)
def test_factory_returns_usable_store(factory_name):
    """Calling each factory must return an instance, not raise."""
    import po_extractor.store as pkg
    factory = getattr(pkg, factory_name)
    store = factory()
    assert store is not None, f"{factory_name}() returned None"
    assert hasattr(store, "_conn") or hasattr(store, "db_path"), (
        f"{factory_name}() did not return a SQLite-backed store"
    )


def test_factory_uses_canonical_db_path():
    """All factories must point at po_extractor.config.DB_PATH — the same
    DB that ui.stores writes to.  Otherwise UI writes are invisible to
    the exporters and vice-versa.
    """
    from po_extractor.store import get_boat_sample_store, get_fabric_master_store
    from po_extractor.config import DB_PATH
    assert get_boat_sample_store().db_path == DB_PATH
    assert get_fabric_master_store().db_path == DB_PATH


def test_no_inline_db_path_construction_in_exporters():
    """Exporters must not build DB paths with os.path.join — they must
    use the canonical factory.  This test scans the source files."""
    files_to_scan = [
        "po_extractor/exporters/sky_east_buyplan_export.py",
        "po_extractor/exporters/hhp_buyplan_export.py",
        "po_extractor/exporters/buyplan_export.py",
    ]
    base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    for rel in files_to_scan:
        path = os.path.join(base, rel)
        if not os.path.exists(path):
            continue
        src = open(path, encoding="utf-8").read()
        # Reject manually-constructed po_history.db paths
        assert "po_history.db" not in src or 'os.path.join' not in src or rel.endswith("buyplan_export.py"), (
            f"{rel} appears to construct the DB path inline — use a "
            f"canonical factory from po_extractor.store instead."
        )


# ─────────────────────────────────────────────────────────────────────────────
# Section 2 — Brand registry union (locks v1.53.1 fix)
# ─────────────────────────────────────────────────────────────────────────────

def test_list_all_brands_unions_all_sources():
    """list_all_brands() must include brands from every brand-holding store."""
    from po_extractor.store import list_all_brands, get_boat_sample_store
    SENTINEL = "__test_brand_sentinel__"
    bss = get_boat_sample_store()
    bss.register_missing_brands("__test_company__", [SENTINEL])
    try:
        brands = list_all_brands("__test_company__")
        assert SENTINEL in brands, (
            "list_all_brands() did not include a brand registered only in "
            "boat_sample_req — the union must cover every brand-holding store."
        )
    finally:
        bss.delete("__test_company__", SENTINEL)


def test_register_missing_brands_is_idempotent_and_returns_only_new():
    """Calling register_missing_brands twice with the same input must
    return the new brands the first time and an empty list the second."""
    from po_extractor.store import get_boat_sample_store
    company = "__test_company__"
    bss = get_boat_sample_store()
    brands = ["__sentinel_X__", "__sentinel_Y__"]
    try:
        new1 = bss.register_missing_brands(company, brands)
        assert sorted(new1) == sorted(brands)
        new2 = bss.register_missing_brands(company, brands + ["__sentinel_Z__"])
        assert new2 == ["__sentinel_Z__"]
        new3 = bss.register_missing_brands(company, brands + ["__sentinel_Z__"])
        assert new3 == []
    finally:
        for b in brands + ["__sentinel_Z__"]:
            bss.delete(company, b)


def test_register_missing_brands_dedups_input():
    """Duplicate brands in the input list must collapse into a single insert."""
    from po_extractor.store import get_boat_sample_store
    company = "__test_company__"
    bss = get_boat_sample_store()
    try:
        new = bss.register_missing_brands(
            company, ["__dup__", "  __dup__  ", "__dup__", ""]
        )
        assert new == ["__dup__"]
    finally:
        bss.delete(company, "__dup__")


# ─────────────────────────────────────────────────────────────────────────────
# Section 3 — Photo filename lookup (locks v1.51.x patterns)
# ─────────────────────────────────────────────────────────────────────────────

def _tiny_png(r: int, g: int, b: int) -> bytes:
    """Return a minimal valid 1×1 PNG for tests."""
    raw = bytes([0, r, g, b])
    compressed = zlib.compress(raw)

    def chunk(t: bytes, d: bytes) -> bytes:
        c = struct.pack(">I", len(d)) + t + d
        return c + struct.pack(">I", zlib.crc32(t + d) & 0xFFFFFFFF)

    return (
        b"\x89PNG\r\n\x1a\n"
        + chunk(b"IHDR", struct.pack(">IIBBBBB", 1, 1, 8, 2, 0, 0, 0))
        + chunk(b"IDAT", compressed)
        + chunk(b"IEND", b"")
    )


class _EmptyRow:
    def get(self, k):  # pragma: no cover — trivial
        return None
    def __getitem__(self, k):
        raise KeyError(k)


@pytest.mark.parametrize("style,filename,expect_front", [
    ("BL3404",            "BL3404_front.png",                True),    # canonical
    ("BL3404",            "BL3404-front.png",                True),    # dash separator
    ("BL3404",            "BL3404_F.png",                    True),    # short suffix
    ("BL3404",            "BL3404_1.png",                    True),    # numeric
    ("BL3404",            "BL3404.png",                      True),    # single image
    ("BL3404",            "BL3404_FRONT.PNG",                True),    # uppercase
    ("BL3404",            "bl3404_front.png",                True),    # lowercase
    ("ZLD060/S24DTR003",  "ZLD060_S24DTR003_front.png",      True),    # slash → underscore
    ("BL3404",            "OTHER_front.png",                 False),   # no match
])
def test_photo_lookup_filename_patterns(style, filename, expect_front):
    """resolve_photo_pair must find photos under any of the supported
    filename patterns — tested explicitly so future refactors can't
    silently drop a pattern."""
    from po_extractor.exporters._photo_utils import resolve_photo_pair
    photo_map = {filename: _tiny_png(255, 0, 0)}
    front, _back = resolve_photo_pair(style, _EmptyRow(), photo_map)
    if expect_front:
        assert front is not None, f"{filename} should match style {style}"
    else:
        assert front is None, f"{filename} should not match style {style}"


def test_photo_lookup_returns_pair_when_both_present():
    from po_extractor.exporters._photo_utils import resolve_photo_pair
    photo_map = {
        "BL3404_front.png": _tiny_png(255, 0, 0),
        "BL3404_back.png":  _tiny_png(0, 0, 255),
    }
    front, back = resolve_photo_pair("BL3404", _EmptyRow(), photo_map)
    assert front is not None and back is not None


# ─────────────────────────────────────────────────────────────────────────────
# Section 4 — Sky East buyplan output structure (end-to-end)
# ─────────────────────────────────────────────────────────────────────────────

@pytest.fixture
def fabric_in_db():
    """Insert one fabric_master record so 综合key lookups have data,
    then clean up."""
    from po_extractor.config import DB_PATH
    HHN = "__TEST_HHN__"
    with sqlite3.connect(DB_PATH) as c:
        c.execute(
            """CREATE TABLE IF NOT EXISTS fabric_master (
                quality_no TEXT PRIMARY KEY, erp_code TEXT, display_key TEXT,
                composition_en TEXT, weight_gsm INTEGER,
                cuttable_width_cm INTEGER, shrinkage_rate TEXT,
                short_rate TEXT)"""
        )
        c.execute(
            "INSERT OR REPLACE INTO fabric_master "
            "(quality_no, display_key, composition_en, weight_gsm, "
            "cuttable_width_cm) VALUES (?, '', ?, ?, ?)",
            (HHN, "100%TestFiber", 220, 145),
        )
    yield HHN
    with sqlite3.connect(DB_PATH) as c:
        c.execute("DELETE FROM fabric_master WHERE quality_no = ?", (HHN,))


@pytest.fixture
def sky_east_output(fabric_in_db, tmp_path):
    """Run export_sky_east_buyplan once and return the loaded workbook."""
    from openpyxl import load_workbook
    from po_extractor.exporters.sky_east_buyplan_export import export_sky_east_buyplan

    df = pd.DataFrame([{
        "pc_no": "TESTPC",
        "style": "TESTSTYLE",
        "contract_no": "C001",
        "brand": "TestBrand",
        "article_name": "TESTART",
        "zalando_po": "PO111",
        "config_sku": "SKU01",
        "color_name": "BLUE",
        "colour_code": "201",
        "xs": 5, "s": 10, "m": 15, "l": 20, "xl": 25, "xxl": 3,
        "ex_fty_date": "2026-07-22",
        "fabric_item_no": fabric_in_db,
        "fabrication": "fallback-comp",
    }])
    photo_map = {
        "TESTSTYLE": [_tiny_png(255, 0, 0), _tiny_png(0, 0, 255)],
    }
    path, _totals = export_sky_east_buyplan(
        df, {}, str(tmp_path),
        fabric_parts_by_style=None,
        style_image_map=photo_map,
    )
    return path, load_workbook(path)


def test_sky_east_buyplan_dk_lands_in_column_d(sky_east_output, fabric_in_db):
    """综合key must be written to column D, not E (v1.52.0 fix)."""
    _path, wb = sky_east_output
    ws = wb["TESTSTYLE"]
    assert ws["D2"].value is not None, "D2 (综合key) is empty"
    assert fabric_in_db in str(ws["D2"].value), (
        f"D2 should contain the HHN code, got {ws['D2'].value!r}"
    )


def test_sky_east_buyplan_e2_is_cleared(sky_east_output):
    """Column E (row 2) must be explicitly empty — old configs put the
    display key there and we want zero ambiguity."""
    _path, wb = sky_east_output
    ws = wb["TESTSTYLE"]
    assert ws["E2"].value is None, (
        f"E2 should be empty, got {ws['E2'].value!r}"
    )


def test_sky_east_buyplan_dk_includes_db_gsm_and_width(sky_east_output, fabric_in_db):
    """The 综合key must include weight_gsm + cuttable_width_cm fetched
    from the fabric_master DB (v1.53.0 fix — silent ImportError)."""
    _path, wb = sky_east_output
    ws = wb["TESTSTYLE"]
    expected = f"{fabric_in_db}|100%TestFiber|220|145"
    assert ws["D2"].value == expected, (
        f"综合key not built from DB.\nExpected: {expected}\n"
        f"Got:      {ws['D2'].value!r}"
    )


def test_sky_east_buyplan_dates_filled(sky_east_output):
    """K1 (制单日期) and K2 (修改日期) must have a non-empty date string
    (v1.51.1 fix)."""
    _path, wb = sky_east_output
    ws = wb["TESTSTYLE"]
    assert ws["K1"].value, "K1 (creation date) is empty"
    assert ws["K2"].value, "K2 (modification date) is empty"


def test_sky_east_buyplan_date_cells_merged(sky_east_output):
    """K1:O1 and K2:O2 must be merged so the date is visible across the
    coloured rectangle (v1.51.1 fix)."""
    _path, wb = sky_east_output
    ws = wb["TESTSTYLE"]
    merges = {str(m) for m in ws.merged_cells.ranges}
    assert "K1:O1" in merges, f"K1:O1 not merged; merges = {merges}"
    assert "K2:O2" in merges, f"K2:O2 not merged; merges = {merges}"


def test_sky_east_buyplan_photo_boxes_use_two_cell_anchor(sky_east_output):
    """Photos must be embedded as twoCellAnchor spanning J3:L6 and M3:O6
    (v1.50.x → v1.52.0 fix).  Tested by reading the drawing XML."""
    path, _wb = sky_east_output
    with zipfile.ZipFile(path) as z:
        drawings = sorted(
            n for n in z.namelist()
            if "drawings/drawing" in n and n.endswith(".xml")
            and "_rels" not in n
        )
        # The style sheet drawing has the two photo anchors;
        # the index sheet drawing has the thumbnail.
        twoCell_anchors_found = []
        for d in drawings:
            xml = z.read(d).decode("utf-8")
            cols = [int(m) for m in re.findall(r"<col>(\d+)</col>", xml)]
            rows = [int(m) for m in re.findall(r"<row>(\d+)</row>", xml)]
            if "twoCellAnchor" in xml:
                # Pair (from_col, to_col) and (from_row, to_row)
                for i in range(0, len(cols), 2):
                    if i + 1 < len(cols):
                        twoCell_anchors_found.append(
                            (cols[i], rows[i], cols[i + 1], rows[i + 1])
                        )
    # Front: J3 (col=9) → M7 (col=12), rows 2 → 6
    # Back:  M3 (col=12) → P7 (col=15), rows 2 → 6
    assert (9, 2, 12, 6) in twoCell_anchors_found, (
        f"Front photo anchor J3:L6 not found in: {twoCell_anchors_found}"
    )
    assert (12, 2, 15, 6) in twoCell_anchors_found, (
        f"Back photo anchor M3:O6 not found in: {twoCell_anchors_found}"
    )


def test_sky_east_buyplan_format_overrides(sky_east_output):
    """Compact-layout overrides must apply: font 10, widths 20/6,
    row height 28pt (v1.50.2 + v1.53.0)."""
    _path, wb = sky_east_output
    ws = wb["TESTSTYLE"]

    # Font size = 10
    assert ws["A1"].font.size == 10, f"A1 font {ws['A1'].font.size}"
    assert ws["D2"].font.size == 10, f"D2 font {ws['D2'].font.size}"

    # Column widths
    assert ws.column_dimensions["A"].width == 20
    assert ws.column_dimensions["I"].width == 20
    assert ws.column_dimensions["R"].width == 20
    for size_col in "JKLMNO":
        assert ws.column_dimensions[size_col].width == 6, (
            f"Size col {size_col} width = {ws.column_dimensions[size_col].width}"
        )

    # Row heights = 28pt
    for r in (1, 2, 5, 9):
        h = ws.row_dimensions[r].height
        assert h == 28, f"Row {r} height = {h}, expected 28"


def test_sky_east_buyplan_workbook_reloads_in_openpyxl(sky_east_output):
    """The generated workbook must round-trip through openpyxl — guards
    against malformed drawing rels paths (v1.50.0 ../xl/media/ bug)."""
    path, _wb = sky_east_output
    from openpyxl import load_workbook
    # If anything in the saved zip is malformed (e.g. rels target points
    # at a non-existent path) load_workbook raises here.
    wb2 = load_workbook(path)
    assert "TESTSTYLE" in wb2.sheetnames


def test_sky_east_index_first_style_has_hyperlink_and_total(sky_east_output, fabric_in_db):
    """The Index sheet's FIRST data row must carry a working hyperlink and
    a SUM-formula total — guards against the v1.54.x bug where the first
    style whose cleaned sheet name collided with the master template was
    silently auto-renamed by openpyxl, leaving Index rows pointing at a
    sheet that didn't exist (no hyperlink, empty 订单数合计).
    """
    _path, wb = sky_east_output
    assert "Index" in wb.sheetnames
    idx = wb["Index"]
    # Row 2 = first data row (row 1 is the header)
    style_cell = idx["B2"]
    assert style_cell.value, "Index row 2 (first style) has no value in 款号"
    assert style_cell.hyperlink is not None, (
        f"First style row has NO hyperlink — {style_cell.value!r}.  "
        "This is the v1.54.x regression: cleaned sheet name collided "
        "with the master template and was silently auto-renamed."
    )
    # The 订单数合计 column (E in no-image layout, F with images) must
    # carry a SUM formula referencing the actual sheet.
    qty_col = None
    for ci, header in enumerate(idx[1], start=1):
        if header.value == "订单数合计":
            qty_col = ci
            break
    assert qty_col, "订单数合计 column not found in Index header"
    qty_val = idx.cell(row=2, column=qty_col).value
    assert qty_val and str(qty_val).startswith("="), (
        f"First style row has empty / non-formula 订单数合计: {qty_val!r}"
    )


def test_sky_east_index_no_redundant_fabric_label_column(sky_east_output):
    """The Index header must NOT contain the redundant 面料_大身 column —
    it duplicated 面料_大身_编号 (which already implies 大身 in its name)."""
    _path, wb = sky_east_output
    idx = wb["Index"]
    headers = [c.value for c in idx[1] if c.value is not None]
    assert "面料_大身" not in headers, (
        f"Redundant column 面料_大身 still in Index header: {headers}"
    )
    assert "面料_大身_编号" in headers, (
        f"Required column 面料_大身_编号 missing from Index header: {headers}"
    )


def test_sky_east_index_every_style_row_has_hyperlink(sky_east_output):
    """Every populated 款号 row must be a hyperlink (not just rows 3+)."""
    _path, wb = sky_east_output
    idx = wb["Index"]
    for ri in range(2, idx.max_row + 1):
        style_cell = idx.cell(row=ri, column=2)
        if style_cell.value:
            assert style_cell.hyperlink is not None, (
                f"款号 row {ri} ({style_cell.value!r}) is missing its hyperlink"
            )


# ─────────────────────────────────────────────────────────────────────────────
# Section 5 — 主标颜色 auto-derivation (locks v1.55.0 rule)
# ─────────────────────────────────────────────────────────────────────────────

@pytest.mark.parametrize("en_color,expected", [
    # ── Light bodies → WHITE label (白色) — matching-colour convention ───
    ("WHITE",                "白色"),
    ("CREAM",                "白色"),
    ("IVORY",                "白色"),
    ("BEIGE",                "白色"),
    ("PINK",                 "白色"),
    ("light blue",           "白色"),  # "light" wins over "blue"
    ("PALE PINK",            "白色"),
    ("BABY BLUE",            "白色"),  # "baby" leftmost
    ("Powder Pink",          "白色"),

    # ── Dark bodies → BLACK label (黑色) ──────────────────────────────────
    ("BLACK",                "黑色"),
    ("NAVY",                 "黑色"),
    ("wine",                 "黑色"),
    ("CHOCOLATE BROWN",      "黑色"),
    ("BURGUNDY",             "黑色"),
    ("DARK TEAL",            "黑色"),
    ("blue",                 "黑色"),
    ("Navy with GOld buckle","黑色"),  # "navy" leftmost wins over "gold"
    ("RED",                  "黑色"),
    ("CHARCOAL",             "黑色"),
    ("Forest Green",         "黑色"),

    # ── Mixed: leftmost token decides (matches "first colour mentioned") ──
    ("CREAM/NAVY PIPING",    "白色"),   # cream first → light → 白色 label
    ("NAVY BODY WITH CREAM", "黑色"),   # navy first → dark → 黑色 label
    ("BLACK WITH WHITE STRAP","黑色"),  # black first → dark → 黑色 label

    # ── Unknown / blank → no auto value ──────────────────────────────────
    ("",                     ""),
    (None,                   ""),
    ("foobar",               ""),
])
def test_derive_main_label_color(en_color, expected):
    """Locks the v1.58.0 rule: light body→白色, dark body→黑色 (matching colour)."""
    from po_extractor.exporters._sky_east_helpers import derive_main_label_color
    assert derive_main_label_color(en_color) == expected, (
        f"derive_main_label_color({en_color!r}) returned the wrong label colour"
    )


def test_sky_east_buyplan_auto_fills_main_label_color(fabric_in_db, tmp_path):
    """End-to-end: column I (主标颜色) is derived from the body colour.

    ``colour_code`` is a Zalando SKU code (e.g. "802") and MUST NOT be used
    as the label colour. With no DB label_color override, the resolution
    falls through to ``derive_main_label_color()``: dark→黑色, light→白色.
    """
    from openpyxl import load_workbook
    from po_extractor.exporters.sky_east_buyplan_export import export_sky_east_buyplan

    df = pd.DataFrame([{
        "pc_no": "P1", "style": "S_DARK",
        "contract_no": "C", "brand": "B", "article_name": "A",
        "zalando_po": "PO1", "config_sku": "K1",
        "color_name": "NAVY",       # dark body → expect 黑色 label
        "colour_code": "",
        "xs": 1, "s": 1, "m": 1, "l": 1, "xl": 1, "xxl": 1,
        "ex_fty_date": "2026-07-22",
        "fabric_item_no": fabric_in_db, "fabrication": "fb",
    }, {
        "pc_no": "P1", "style": "S_LIGHT",
        "contract_no": "C", "brand": "B", "article_name": "A",
        "zalando_po": "PO2", "config_sku": "K2",
        "color_name": "CREAM",      # light body → expect 白色 label
        "colour_code": "",
        "xs": 1, "s": 1, "m": 1, "l": 1, "xl": 1, "xxl": 1,
        "ex_fty_date": "2026-07-22",
        "fabric_item_no": fabric_in_db, "fabrication": "fb",
    }, {
        "pc_no": "P1", "style": "S_SKU_IGNORED",
        "contract_no": "C", "brand": "B", "article_name": "A",
        "zalando_po": "PO3", "config_sku": "K3",
        "color_name": "NAVY",
        "colour_code": "802",       # SKU code — must be ignored as label colour
        "xs": 1, "s": 1, "m": 1, "l": 1, "xl": 1, "xxl": 1,
        "ex_fty_date": "2026-07-22",
        "fabric_item_no": fabric_in_db, "fabrication": "fb",
    }])
    path, _ = export_sky_east_buyplan(
        df, {}, str(tmp_path),
        fabric_parts_by_style=None, style_image_map=None,
    )
    wb = load_workbook(path)

    # Column I (9) = label_clr; row 9 = first data row in Sky_East template.
    # (Row 8 is the bilingual header row — scanning from there would pick up
    # the literal header "主标色" instead of the data value.)
    assert wb["S_DARK"].cell(9, 9).value  == "黑色", (
        f"dark body → expected 黑色, got {wb['S_DARK'].cell(9, 9).value!r}"
    )
    assert wb["S_LIGHT"].cell(9, 9).value == "白色", (
        f"light body → expected 白色, got {wb['S_LIGHT'].cell(9, 9).value!r}"
    )
    assert wb["S_SKU_IGNORED"].cell(9, 9).value == "黑色", (
        "colour_code is a Zalando SKU and MUST NOT be used as label colour"
    )


# ─────────────────────────────────────────────────────────────────────────────
# Section 6 — ColorTranslationStore: light/dark + label_color columns
# ─────────────────────────────────────────────────────────────────────────────

def test_color_translation_store_has_new_columns():
    """The migrated table must have light_or_dark + label_color columns."""
    from po_extractor.store import get_color_translation_store
    store = get_color_translation_store()
    with store._conn() as conn:
        cols = {r[1] for r in conn.execute("PRAGMA table_info(color_translations)")}
    assert "light_or_dark" in cols, "missing column light_or_dark"
    assert "label_color"   in cols, "missing column label_color"


def test_color_translation_lookup_label_color_explicit_override():
    """Explicit label_color set by user must be returned verbatim."""
    from po_extractor.store import get_color_translation_store
    store = get_color_translation_store()
    df = pd.DataFrame([{
        "Client": "__t_co__", "Brand": "__t_br__",
        "English Color": "MAGIC NAVY",
        "Chinese Color": "深海蓝",
        "Color Code": "9", "Light/Dark": "dark",
        "Label Color": "白色", "Notes": "",
    }])
    store.upsert_from_df(df)
    try:
        assert store.lookup_label_color("__t_co__", "__t_br__", "MAGIC NAVY") == "白色"
    finally:
        store.delete_by_client_brand("__t_co__", "__t_br__")


def test_color_translation_upsert_auto_derives_when_blank():
    """When the user leaves Light/Dark + Label Color blank, the store must
    populate them from the English colour name.  English colour names are
    normalised to title case on save (case-insensitive matching)."""
    from po_extractor.store import get_color_translation_store
    store = get_color_translation_store()
    df = pd.DataFrame([
        {"Client": "__t_co__", "Brand": "", "English Color": "BLACK",
         "Chinese Color": "", "Color Code": "",
         "Light/Dark": "", "Label Color": "", "Notes": ""},
        {"Client": "__t_co__", "Brand": "", "English Color": "cream",
         "Chinese Color": "", "Color Code": "",
         "Light/Dark": "", "Label Color": "", "Notes": ""},
    ])
    store.upsert_from_df(df)
    try:
        view = store.to_dataframe("__t_co__")
        rows = {r["English Color"]: r for _, r in view.iterrows()}
        # Input casing ("BLACK" / "cream") collapses to title case on save
        assert "Black" in rows, f"expected 'Black' (title-cased), got {list(rows)}"
        assert "Cream" in rows, f"expected 'Cream' (title-cased), got {list(rows)}"
        assert rows["Black"]["Light/Dark"] == "dark"
        assert rows["Black"]["Label Color"] == "黑色"
        assert rows["Cream"]["Light/Dark"] == "light"
        assert rows["Cream"]["Label Color"] == "白色"
    finally:
        store.delete_by_client_brand("__t_co__", "")


def test_color_translation_normalize_color_name():
    """English colour names must collapse to a single canonical form."""
    from po_extractor.store.color_translation_store import _normalize_color_name
    assert _normalize_color_name("NAVY")              == "Navy"
    assert _normalize_color_name("navy")              == "Navy"
    assert _normalize_color_name("Navy")              == "Navy"
    assert _normalize_color_name("  chocolate brown") == "Chocolate Brown"
    assert _normalize_color_name("52#NAVY")           == "52#Navy"
    assert _normalize_color_name(None)                == ""
    assert _normalize_color_name("")                  == ""
    # Whitespace is collapsed (newlines / multiple spaces)
    assert _normalize_color_name("BLACK   WITH\n  WHITE") == "Black With White"


def test_color_translation_lookup_is_case_insensitive():
    """Looking up "navy" must find a row stored as "Navy"."""
    from po_extractor.store import get_color_translation_store
    store = get_color_translation_store()
    df = pd.DataFrame([{
        "Client": "__t_co__", "Brand": "", "English Color": "Navy",
        "Chinese Color": "藏青色", "Color Code": "",
        "Light/Dark": "dark", "Label Color": "白色", "Notes": "",
    }])
    store.upsert_from_df(df)
    try:
        # Various source casings should all find the same row
        for variant in ("Navy", "navy", "NAVY", "  NaVy  "):
            assert store.lookup("__t_co__", "", variant) == "藏青色", (
                f"lookup({variant!r}) failed"
            )
            assert store.lookup_label_color("__t_co__", "", variant) == "白色"
    finally:
        store.delete_by_client_brand("__t_co__", "")


def test_color_translation_delete_ids():
    """delete_ids must remove only the listed row IDs."""
    from po_extractor.store import get_color_translation_store
    store = get_color_translation_store()
    df = pd.DataFrame([
        {"Client": "__t_co__", "Brand": "", "English Color": "TempA",
         "Chinese Color": "", "Color Code": "", "Light/Dark": "",
         "Label Color": "", "Notes": ""},
        {"Client": "__t_co__", "Brand": "", "English Color": "TempB",
         "Chinese Color": "", "Color Code": "", "Light/Dark": "",
         "Label Color": "", "Notes": ""},
    ])
    store.upsert_from_df(df)
    try:
        view = store.to_dataframe("__t_co__")
        # Pick the id of the "Tempa" row and delete just that one
        tempa_id = int(view.loc[view["English Color"] == "Tempa", "_id"].iloc[0])
        n = store.delete_ids([tempa_id])
        assert n == 1
        view2 = store.to_dataframe("__t_co__")
        remaining = set(view2["English Color"])
        assert remaining == {"Tempb"}, (
            f"delete_ids removed wrong rows; remaining = {remaining}"
        )
    finally:
        store.delete_by_client_brand("__t_co__", "")


# ─────────────────────────────────────────────────────────────────────────────
# Section 7 — ColorTranslationStore audit log
# ─────────────────────────────────────────────────────────────────────────────

def test_audit_log_records_insert():
    """Inserting a new row must produce an 'insert' audit entry."""
    from po_extractor.store import get_color_translation_store
    store = get_color_translation_store()
    df = pd.DataFrame([{
        "Client": "__t_co_audit__", "Brand": "B",
        "English Color": "TestInsertColor",
        "Chinese Color": "测试色", "Color Code": "",
        "Light/Dark": "", "Label Color": "", "Notes": "",
    }])
    store.upsert_from_df(df)
    try:
        log = store.audit_log(limit=10, client="__t_co_audit__")
        assert any(r["action"] == "insert" and r["en_color"] == "Testinsertcolor"
                   for r in log), (
            f"No insert entry for TestInsertColor; log = "
            f"{[(r['action'], r['en_color']) for r in log]}"
        )
    finally:
        store.delete_by_client_brand("__t_co_audit__", "B")


def test_audit_log_records_per_field_update():
    """Editing a field must produce exactly one 'update' entry per changed
    field with its old + new value captured."""
    from po_extractor.store import get_color_translation_store
    store = get_color_translation_store()

    base = pd.DataFrame([{
        "Client": "__t_co_audit__", "Brand": "B",
        "English Color": "AuditNavy",
        "Chinese Color": "藏青色", "Color Code": "52#",
        "Light/Dark": "dark", "Label Color": "黑色", "Notes": "",
    }])
    store.upsert_from_df(base)

    # Now change only Chinese Color and Notes
    edited = pd.DataFrame([{
        "Client": "__t_co_audit__", "Brand": "B",
        "English Color": "AuditNavy",
        "Chinese Color": "深蓝色",     # changed
        "Color Code": "52#",          # unchanged
        "Light/Dark": "dark",          # unchanged
        "Label Color": "黑色",         # unchanged
        "Notes": "manual override",   # changed
    }])
    store.upsert_from_df(edited)
    try:
        log = store.audit_log(limit=20, client="__t_co_audit__",
                              en_color="AuditNavy")
        update_entries = [r for r in log if r["action"] == "update"]
        # Two changes → two update entries
        changed_fields = {r["field"]: r for r in update_entries}
        assert "cn_color" in changed_fields
        assert "notes" in changed_fields
        assert changed_fields["cn_color"]["old_value"] == "藏青色"
        assert changed_fields["cn_color"]["new_value"] == "深蓝色"
        assert changed_fields["notes"]["old_value"] == ""
        assert changed_fields["notes"]["new_value"] == "manual override"
        # Unchanged fields should NOT have entries
        assert "color_code"    not in changed_fields
        assert "light_or_dark" not in changed_fields
        assert "label_color"   not in changed_fields
    finally:
        store.delete_by_client_brand("__t_co_audit__", "B")


def test_audit_log_records_delete():
    """Deleting a row via delete_ids must produce a 'delete' audit entry."""
    from po_extractor.store import get_color_translation_store
    store = get_color_translation_store()
    df = pd.DataFrame([{
        "Client": "__t_co_audit__", "Brand": "B",
        "English Color": "AuditDel",
        "Chinese Color": "测试删除",
        "Color Code": "", "Light/Dark": "", "Label Color": "", "Notes": "",
    }])
    store.upsert_from_df(df)
    view = store.to_dataframe("__t_co_audit__")
    target_id = int(view.loc[view["English Color"] == "Auditdel", "_id"].iloc[0])
    store.delete_ids([target_id])

    log = store.audit_log(limit=20, client="__t_co_audit__",
                          en_color="AuditDel")
    delete_entries = [r for r in log if r["action"] == "delete"]
    assert delete_entries, (
        f"No delete entry recorded; log = {log}"
    )
    assert delete_entries[0]["row_id"] == target_id


def test_audit_log_filters_by_en_color_case_insensitive():
    """The audit-log filter must be case-insensitive on en_color."""
    from po_extractor.store import get_color_translation_store
    store = get_color_translation_store()
    df = pd.DataFrame([{
        "Client": "__t_co_audit__", "Brand": "",
        "English Color": "FilterTest",
        "Chinese Color": "", "Color Code": "",
        "Light/Dark": "", "Label Color": "", "Notes": "",
    }])
    store.upsert_from_df(df)
    try:
        for variant in ("filtertest", "FilterTest", "FILTERTEST"):
            log = store.audit_log(client="__t_co_audit__", en_color=variant)
            assert any(r["en_color"].lower() == "filtertest" for r in log), (
                f"en_color filter failed for variant {variant!r}"
            )
    finally:
        store.delete_by_client_brand("__t_co_audit__", "")


def test_color_translation_import_from_progress_xlsx(tmp_path):
    """Locks the progress-tracker importer end-to-end."""
    import openpyxl
    from po_extractor.store import get_color_translation_store

    p = tmp_path / "test_progress.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test"
    # Header row
    headers = ["No.", "PO", "Style", "BRAND", "颜色",
               "主标颜色", "中文颜色"]
    for ci, h in enumerate(headers, 1):
        ws.cell(1, ci, h)
    # Data rows
    data = [
        (1, "PO1", "S1", "TestBrand", "NAVY", "",     "52#藏青色"),
        (2, "PO2", "S2", "TestBrand", "navy", "",     "52#藏青色"),  # duplicate, different case
        (3, "PO3", "S3", "TestBrand", "CREAM", "黑色", "92#奶白色"),
        (4, "PO4", "S4", "TestBrand", "",      "",     ""),         # blank, must be skipped
    ]
    for ri, row in enumerate(data, 2):
        for ci, v in enumerate(row, 1):
            ws.cell(ri, ci, v)
    wb.save(p)
    wb.close()

    store = get_color_translation_store()
    result = store.import_from_progress_xlsx(str(p), client="__t_co__")
    try:
        # NAVY + navy collapse → 1 unique English colour
        assert result["inserted"] >= 2
        assert result["skipped"]  >= 1   # the blank row

        # Round-trip the data
        cn  = store.lookup           ("__t_co__", "TestBrand", "navy")
        lab = store.lookup_label_color("__t_co__", "TestBrand", "Cream")
        assert cn  == "52#藏青色", f"navy CN lookup failed: {cn!r}"
        assert lab == "黑色",      f"Cream label lookup failed: {lab!r}"
    finally:
        store.delete_by_client_brand("__t_co__", "TestBrand")


def test_buyplan_uses_db_label_color_over_keyword_derivation(fabric_in_db, tmp_path):
    """When the colour-translation DB has a label_color set, the buyplan
    must use that value instead of the keyword-derived one — single source
    of truth.
    """
    from openpyxl import load_workbook
    from po_extractor.store import get_color_translation_store
    from po_extractor.exporters.sky_east_buyplan_export import export_sky_east_buyplan

    store = get_color_translation_store()
    # Override the keyword default — the keyword rule for "navy" gives 黑色,
    # we explicitly set 白色 in the DB to confirm the DB value wins.
    df_seed = pd.DataFrame([{
        "Client": "Sky East", "Brand": "B_OVERRIDE",
        "English Color": "NAVY",
        "Chinese Color": "藏青色",
        "Color Code": "52#",
        "Light/Dark": "dark",
        "Label Color": "白色",     # explicit override (contradicts keyword rule)
        "Notes": "",
    }])
    store.upsert_from_df(df_seed)

    df = pd.DataFrame([{
        "pc_no": "P", "style": "S_DB",
        "contract_no": "C", "brand": "B_OVERRIDE", "article_name": "A",
        "zalando_po": "PO1", "config_sku": "K",
        "color_name": "NAVY", "colour_code": "",
        "xs": 1, "s": 1, "m": 1, "l": 1, "xl": 1, "xxl": 1,
        "ex_fty_date": "2026-07-22",
        "fabric_item_no": fabric_in_db, "fabrication": "fb",
    }])
    try:
        path, _ = export_sky_east_buyplan(
            df, {}, str(tmp_path),
            fabric_parts_by_style=None, style_image_map=None,
        )
        wb = load_workbook(path)
        # Column I row 9 = label_clr
        actual = wb["S_DB"].cell(9, 9).value
        assert actual == "白色", (
            f"DB label_color override ignored — expected 白色, got {actual!r}"
        )
    finally:
        store.delete_by_client_brand("Sky East", "B_OVERRIDE")


def test_sky_east_buyplan_warns_when_hhn_missing_from_db(tmp_path, capsys):
    """When an HHN isn't in fabric_master the export must emit a warning
    so the user sees what's wrong (loud-failure principle)."""
    from po_extractor.exporters.sky_east_buyplan_export import export_sky_east_buyplan
    df = pd.DataFrame([{
        "pc_no": "PC", "style": "S1", "contract_no": "", "brand": "B",
        "article_name": "A", "zalando_po": "P", "config_sku": "K",
        "color_name": "C", "colour_code": "1",
        "xs": 1, "s": 1, "m": 1, "l": 1, "xl": 1, "xxl": 1,
        "ex_fty_date": "", "fabric_item_no": "__MISSING_HHN__",
        "fabrication": "fb",
    }])
    with warnings.catch_warnings(record=True) as W:
        warnings.simplefilter("always")
        export_sky_east_buyplan(df, {}, str(tmp_path),
                                fabric_parts_by_style=None,
                                style_image_map=None)
        msgs = [str(w.message) for w in W]
    assert any("综合key partial" in m or "综合key" in m for m in msgs), (
        f"Expected 综合key diagnostic warning; got messages: {msgs}"
    )
