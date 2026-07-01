"""Tests for set_internal_hyperlink() — the WPS-safe way to link one sheet
cell to a cell on another sheet in the same workbook.

Regression: a plain-string ``cell.hyperlink = f"#'{sheet}'!A1"`` writes an
*external* relationship (``TargetMode="External"``) whose literal Target is
the ``"#..."`` string. Excel special-cases a ``"#"``-prefixed external
target and follows it as an internal jump, but that's an Excel-only
leniency, not part of the OOXML spec — WPS (and other readers) take the
target literally, so the link does nothing there.
"""
from __future__ import annotations

from openpyxl import Workbook

from po_extractor.exporters._excel_helpers import set_internal_hyperlink


def test_set_internal_hyperlink_uses_location_not_target():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Index"
    ws2 = wb.create_sheet("1_DR5124")

    cell = ws1.cell(1, 1, value="DR5124")
    set_internal_hyperlink(cell, "1_DR5124")

    assert cell.hyperlink is not None
    assert cell.hyperlink.target is None
    assert cell.hyperlink.location == "'1_DR5124'!A1"


def test_set_internal_hyperlink_writes_no_external_relationship():
    """The whole point: no relationship means no chance of a broken/foreign
    Target string reaching the saved file — verify by round-tripping through
    a real save/load and checking the sheet has zero hyperlink relationships.
    """
    import io
    from openpyxl import load_workbook

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Index"
    wb.create_sheet("Style1")
    set_internal_hyperlink(ws1.cell(1, 1, value="Style1"), "Style1")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    wb2 = load_workbook(buf)
    ws1_reloaded = wb2["Index"]
    cell = ws1_reloaded.cell(1, 1)
    assert cell.hyperlink is not None
    assert cell.hyperlink.target is None
    assert cell.hyperlink.location == "'Style1'!A1"


def test_set_internal_hyperlink_custom_anchor():
    wb = Workbook()
    ws1 = wb.active
    wb.create_sheet("Detail")
    cell = ws1.cell(1, 1)
    set_internal_hyperlink(cell, "Detail", anchor="B5")
    assert cell.hyperlink.location == "'Detail'!B5"


def test_set_internal_hyperlink_preserves_existing_cell_value():
    """Setting the hyperlink must not clobber a value already written to the
    cell (the display text stays whatever the caller set it to)."""
    wb = Workbook()
    ws1 = wb.active
    wb.create_sheet("Style1")
    cell = ws1.cell(1, 1, value="DR5124")
    set_internal_hyperlink(cell, "Style1")
    assert cell.value == "DR5124"
