"""Tests for fabric_master versioning: snapshot history, incremental diff log,
version-scoped enrichment, and pruning.

Locks the v2.75.0 feature (current + 3 previous snapshots stay queryable;
older snapshot DATA is pruned but ``fabric_versions``/``fabric_version_diff``
never are; ``get_batch_enrichment`` can be pinned to a historical version)
plus the v2.77.x refinement: an import only mints a new version when the
fabric DATA actually differs from the latest version -- a byte-identical
re-upload (even under a new filename) is a no-op version-wise.
"""
from __future__ import annotations

import io

import openpyxl
import pytest

from po_extractor.store.fabric_master_store import FabricMasterStore


# The importer positionally fallback-maps EVERY field via _COL_MAP_FALLBACK
# (setdefault), so a fixture that packs its 4 fields into columns 1-4 leaks
# the weight/width values into the text fields whose canonical columns those
# are (supplier_no=3, supplier=4). Write each field at its canonical column
# instead, leaving every other column empty (-> None -> no cross-mapping).
_FIXTURE_COLS = {          # field -> 1-based canonical column (_COL_MAP_FALLBACK)
    "quality_no": 1, "composition_en": 5, "weight_gsm": 10, "cuttable_width_cm": 11,
}
_FIXTURE_HEADERS = {
    "quality_no": "公司面料编号", "composition_en": "面料成分(英文)",
    "weight_gsm": "克重(gsm)", "cuttable_width_cm": "有效门幅(cm)",
}


def _make_fabric_xlsx(rows: list[dict]) -> bytes:
    """Build a minimal 面料统计表.xlsx ('all' sheet) from row dicts keyed by
    quality_no/composition_en/weight_gsm/cuttable_width_cm."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "all"
    for field, col in _FIXTURE_COLS.items():
        ws.cell(row=1, column=col, value=_FIXTURE_HEADERS[field])
    for ri, r in enumerate(rows, start=2):
        for field, col in _FIXTURE_COLS.items():
            ws.cell(row=ri, column=col, value=r.get(field))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _import(store, tmp_path, name: str, rows: list[dict]) -> dict:
    path = tmp_path / name
    path.write_bytes(_make_fabric_xlsx(rows))
    return store.import_from_xlsx(str(path), source_file_name=name)


@pytest.fixture
def store(tmp_path):
    return FabricMasterStore(str(tmp_path / "fabric_master.db"))


def test_first_import_creates_version_1_all_added(store, tmp_path):
    result = _import(store, tmp_path, "v1.xlsx", [
        {"quality_no": "A", "composition_en": "100%Cotton", "weight_gsm": 200, "cuttable_width_cm": 140},
        {"quality_no": "B", "composition_en": "100%Poly", "weight_gsm": 150, "cuttable_width_cm": 150},
    ])
    assert result["version_id"] == 1
    assert result["inserted"] == 2 and result["updated"] == 0

    versions = store.list_versions()
    assert len(versions) == 1 and versions[0]["version_id"] == 1
    assert versions[0]["row_count"] == 2

    summary = store.get_diff_summary(1)
    assert summary == {"added": 2, "removed": 0, "changed": 0}
    diff = store.get_version_diff(1)
    assert {d["quality_no"] for d in diff} == {"A", "B"}
    assert all(d["change_type"] == "added" for d in diff)


def test_second_import_diffs_changed_field(store, tmp_path):
    _import(store, tmp_path, "v1.xlsx", [
        {"quality_no": "A", "composition_en": "100%Cotton", "weight_gsm": 200, "cuttable_width_cm": 140},
    ])
    result = _import(store, tmp_path, "v2.xlsx", [
        {"quality_no": "A", "composition_en": "100%Cotton", "weight_gsm": 220, "cuttable_width_cm": 140},
    ])
    assert result["version_id"] == 2
    summary = store.get_diff_summary(2)
    assert summary == {"added": 0, "removed": 0, "changed": 1}

    diff = store.get_version_diff(2)
    weight_diff = next(d for d in diff if d["field"] == "weight_gsm")
    assert weight_diff["quality_no"] == "A"
    assert weight_diff["old_value"] == "200.0"
    assert weight_diff["new_value"] == "220.0"
    # composition_en unchanged -- must NOT appear in the diff
    assert not any(d["field"] == "composition_en" for d in diff)


def test_removed_row_only_detected_after_delete(store, tmp_path):
    """Update/Add mode never deletes, so a row absent from a later upload but
    never explicitly deleted must NOT show as 'removed' -- it's still live."""
    _import(store, tmp_path, "v1.xlsx", [
        {"quality_no": "A", "composition_en": "100%Cotton", "weight_gsm": 200, "cuttable_width_cm": 140},
        {"quality_no": "B", "composition_en": "100%Poly", "weight_gsm": 150, "cuttable_width_cm": 150},
    ])
    # Simulate "Update/Add": B is simply omitted from the new file, but never
    # deleted. Nothing actually differs, so (post-v2.77.x) this is a no-op
    # version-wise -- the point stands: B is NOT flagged "removed".
    result = _import(store, tmp_path, "v2_update.xlsx", [
        {"quality_no": "A", "composition_en": "100%Cotton", "weight_gsm": 200, "cuttable_width_cm": 140},
    ])
    assert result["unchanged"] is True
    assert result["version_id"] == 1        # no new version minted
    assert store.count() == 2   # B is still there, untouched

    # Now simulate "Clear & Reimport": delete_all() then a fresh, smaller file.
    store.delete_all()
    result2 = _import(store, tmp_path, "v3_clear.xlsx", [
        {"quality_no": "A", "composition_en": "100%Cotton", "weight_gsm": 200, "cuttable_width_cm": 140},
        {"quality_no": "C", "composition_en": "100%Wool", "weight_gsm": 300, "cuttable_width_cm": 145},
    ])
    summary = store.get_diff_summary(result2["version_id"])
    assert summary == {"added": 1, "removed": 1, "changed": 0}
    diff = store.get_version_diff(result2["version_id"])
    assert {d["quality_no"]: d["change_type"] for d in diff} == {"B": "removed", "C": "added"}


def test_version_numbering_increments_sequentially(store, tmp_path):
    ids = []
    for i in range(4):
        r = _import(store, tmp_path, f"v{i}.xlsx", [
            {"quality_no": f"Q{i}", "composition_en": "Cotton", "weight_gsm": 100 + i, "cuttable_width_cm": 140},
        ])
        ids.append(r["version_id"])
    assert ids == [1, 2, 3, 4]
    assert [v["version_id"] for v in store.list_versions()] == [4, 3, 2, 1]


def test_pruning_keeps_current_plus_3_previous(store, tmp_path):
    """5 uploads: only versions 2-5 keep snapshot DATA; version 1's data is
    pruned, but its metadata + diff rows must still exist (never pruned)."""
    for i in range(1, 6):
        _import(store, tmp_path, f"v{i}.xlsx", [
            {"quality_no": f"Q{i}", "composition_en": "Cotton", "weight_gsm": 100, "cuttable_width_cm": 140},
        ])

    with store._conn() as conn:
        snap_versions = {r[0] for r in conn.execute(
            "SELECT DISTINCT version_id FROM fabric_master_snapshot").fetchall()}
    assert snap_versions == {2, 3, 4, 5}, (
        f"expected only the current + 3 previous versions to retain snapshot "
        f"data, got {snap_versions}"
    )

    # Metadata and diff log survive regardless of pruning.
    assert {v["version_id"] for v in store.list_versions()} == {1, 2, 3, 4, 5}
    assert store.get_diff_summary(1) == {"added": 1, "removed": 0, "changed": 0}

    # get_batch_enrichment against a pruned version now returns nothing.
    assert store.get_batch_enrichment(["Q1"], version_id=1) == {}
    # ...but a retained version still resolves.
    assert store.get_batch_enrichment(["Q2"], version_id=2) != {}


def test_identical_reimport_does_not_bump_version(store, tmp_path):
    """Re-uploading a byte-identical fabric list (even under a different
    filename) must NOT mint a new version -- no duplicate snapshot, no empty
    diff entry, version stays where it was."""
    rows = [
        {"quality_no": "A", "composition_en": "100%Cotton", "weight_gsm": 200, "cuttable_width_cm": 140},
        {"quality_no": "B", "composition_en": "100%Poly", "weight_gsm": 150, "cuttable_width_cm": 150},
    ]
    first = _import(store, tmp_path, "v1.xlsx", rows)
    assert first["version_id"] == 1
    assert first.get("unchanged") is False

    again = _import(store, tmp_path, "same_data_new_name.xlsx", rows)
    assert again["version_id"] == 1, "identical re-upload must not bump the version"
    assert again["unchanged"] is True
    # The upsert itself still ran (rows refreshed in place) -- only the
    # VERSION machinery was skipped.
    assert again["updated"] == 2

    assert [v["version_id"] for v in store.list_versions()] == [1]
    with store._conn() as conn:
        n_snap_versions = conn.execute(
            "SELECT COUNT(DISTINCT version_id) FROM fabric_master_snapshot"
        ).fetchone()[0]
    assert n_snap_versions == 1


def test_identical_reimport_then_real_change_still_bumps(store, tmp_path):
    rows = [{"quality_no": "A", "composition_en": "100%Cotton",
             "weight_gsm": 200, "cuttable_width_cm": 140}]
    _import(store, tmp_path, "v1.xlsx", rows)
    _import(store, tmp_path, "v1_again.xlsx", rows)          # no-op, still v1

    changed = [{"quality_no": "A", "composition_en": "100%Cotton",
                "weight_gsm": 220, "cuttable_width_cm": 140}]
    result = _import(store, tmp_path, "v2.xlsx", changed)
    assert result["version_id"] == 2
    assert result["unchanged"] is False
    assert store.get_diff_summary(2) == {"added": 0, "removed": 0, "changed": 1}


def test_identical_reimport_does_not_consume_retention_window(store, tmp_path):
    """The real cost of the old behaviour: each no-op re-upload pushed a
    genuinely different old snapshot out of the current+3 retention window.
    After 4 real versions, any number of identical re-uploads must leave all
    4 snapshots (including the oldest) intact and queryable."""
    for i in range(1, 5):   # 4 real versions, each with different data
        _import(store, tmp_path, f"v{i}.xlsx", [
            {"quality_no": "A", "composition_en": "100%Cotton",
             "weight_gsm": 100 + i, "cuttable_width_cm": 140},
        ])

    last_rows = [{"quality_no": "A", "composition_en": "100%Cotton",
                  "weight_gsm": 104, "cuttable_width_cm": 140}]
    for j in range(3):      # 3 no-op re-uploads of the latest data
        r = _import(store, tmp_path, f"noop{j}.xlsx", last_rows)
        assert r["unchanged"] is True

    # Version 1's snapshot must still be inside the retention window.
    assert store.get_batch_enrichment(["A"], version_id=1)["A"]["weight_gsm"] == 101
    assert [v["version_id"] for v in store.list_versions()] == [4, 3, 2, 1]


def test_numeric_precision_noise_beyond_2dp_is_not_a_change(store, tmp_path):
    """Fabric-table numbers are only meaningful to 2 decimal places -- a
    re-upload whose numeric values differ only beyond 2 dp (Excel float
    artifacts, 66.666667 vs 66.67 on file) must count as unchanged and must
    NOT mint a new version."""
    _import(store, tmp_path, "v1.xlsx", [
        {"quality_no": "A", "composition_en": "100%Cotton",
         "weight_gsm": 66.666666, "cuttable_width_cm": 140.004},
    ])
    result = _import(store, tmp_path, "v1_rounded.xlsx", [
        {"quality_no": "A", "composition_en": "100%Cotton",
         "weight_gsm": 66.67, "cuttable_width_cm": 140.0},
    ])
    assert result["unchanged"] is True
    assert result["version_id"] == 1
    assert [v["version_id"] for v in store.list_versions()] == [1]


def test_numeric_change_at_2dp_still_bumps_and_records_rounded_values(store, tmp_path):
    _import(store, tmp_path, "v1.xlsx", [
        {"quality_no": "A", "composition_en": "100%Cotton",
         "weight_gsm": 66.66, "cuttable_width_cm": 140},
    ])
    result = _import(store, tmp_path, "v2.xlsx", [
        {"quality_no": "A", "composition_en": "100%Cotton",
         "weight_gsm": 66.68, "cuttable_width_cm": 140},
    ])
    assert result["unchanged"] is False
    assert result["version_id"] == 2

    diff = store.get_version_diff(2)
    weight_diff = next(d for d in diff if d["field"] == "weight_gsm")
    assert weight_diff["old_value"] == "66.66"
    assert weight_diff["new_value"] == "66.68"


def test_get_batch_enrichment_no_version_matches_live_table(store, tmp_path):
    """Regression guard: get_batch_enrichment(fabric_nos) with no version_id
    must behave exactly as before versioning existed -- reads live fabric_master."""
    _import(store, tmp_path, "v1.xlsx", [
        {"quality_no": "A", "composition_en": "100%Cotton", "weight_gsm": 200, "cuttable_width_cm": 140},
    ])
    _import(store, tmp_path, "v2.xlsx", [
        {"quality_no": "A", "composition_en": "100%Cotton", "weight_gsm": 999, "cuttable_width_cm": 140},
    ])
    live = store.get_batch_enrichment(["A"])
    assert live["A"]["weight_gsm"] == 999   # reflects the LATEST live state


def test_get_batch_enrichment_specific_version_reads_snapshot(store, tmp_path):
    _import(store, tmp_path, "v1.xlsx", [
        {"quality_no": "A", "composition_en": "100%Cotton", "weight_gsm": 200, "cuttable_width_cm": 140},
    ])
    _import(store, tmp_path, "v2.xlsx", [
        {"quality_no": "A", "composition_en": "100%Cotton", "weight_gsm": 999, "cuttable_width_cm": 140},
    ])
    v1_data = store.get_batch_enrichment(["A"], version_id=1)
    assert v1_data["A"]["weight_gsm"] == 200   # the OLDER value, from the v1 snapshot

    v2_data = store.get_batch_enrichment(["A"], version_id=2)
    assert v2_data["A"]["weight_gsm"] == 999


def test_uploaded_by_defaults_to_system_outside_streamlit(store, tmp_path):
    result = _import(store, tmp_path, "v1.xlsx", [
        {"quality_no": "A", "composition_en": "Cotton", "weight_gsm": 100, "cuttable_width_cm": 140},
    ])
    versions = store.list_versions()
    assert versions[0]["uploaded_by"] == "system"
    assert versions[0]["version_id"] == result["version_id"]


# ─────────────────────────────────────────────────────────────────────────────
# migrate_from_db: src_conn must not leak on a mid-read exception
# ─────────────────────────────────────────────────────────────────────────────

def test_migrate_from_db_closes_src_conn_on_exception(tmp_path, monkeypatch):
    """If reading the source DB raises between connect() and the explicit
    .close() call (e.g. a transient OperationalError), migrate_from_db must
    still close src_conn via a finally -- previously the explicit
    ``src_conn.close()`` calls were only reached on the success and the two
    early-return ("no table" / "empty table") paths, so any other exception
    leaked the connection.
    """
    import sqlite3
    from po_extractor.store.fabric_master_store import FabricMasterStore

    src_path = str(tmp_path / "src.db")
    FabricMasterStore(src_path)  # creates the fabric_master schema
    with sqlite3.connect(src_path) as c:
        c.execute(
            "INSERT INTO fabric_master (quality_no, composition_en) VALUES (?, ?)",
            ("Q1", "100%Cotton"),
        )

    closed = {"called": False}

    class _SpyConnection(sqlite3.Connection):
        def execute(self, sql, *args, **kwargs):
            if sql.strip() == "SELECT * FROM fabric_master":
                raise sqlite3.OperationalError("simulated read failure")
            return super().execute(sql, *args, **kwargs)

        def close(self):
            closed["called"] = True
            return super().close()

    real_connect = sqlite3.connect

    def _connect_with_spy(path, *a, **kw):
        if path == src_path:
            kw["factory"] = _SpyConnection
        return real_connect(path, *a, **kw)

    monkeypatch.setattr(sqlite3, "connect", _connect_with_spy)

    dst_path = str(tmp_path / "dst.db")
    result = FabricMasterStore.migrate_from_db(src_path, dst_path)

    assert "Cannot read source DB" in result["message"]
    assert closed["called"], (
        "src_conn.close() must run even when reading the source DB raises "
        "between connect() and the explicit close() call"
    )
