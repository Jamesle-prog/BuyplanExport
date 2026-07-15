"""Unit tests for the pieces extracted out of ``export_sky_east_buyplan`` so
they can be tested without generating a whole workbook.

Covers ``_FabricMasterCache.display_key`` (was a nested closure over three
locals) and the batch pre-fetch helpers.
"""
from __future__ import annotations

import pandas as pd

from openpyxl import Workbook

from po_extractor.exporters.sky_east_buyplan_export import (
    _FabricMasterCache, _prefetch_boat_sample_cache, _prefetch_fabric_master_cache,
    _RowContext, _fill_one_style_row, _COLOR_NOT_FOUND, _present_order_sizes,
)


# ── _present_order_sizes (drives dynamic 核料 size columns) ──────────────────

def test_present_order_sizes_returns_only_nonzero_in_canonical_order():
    df = pd.DataFrame([
        {"xs": 0, "s": 5, "m": 0, "l": 2, "xl": 0, "xxl": 0},
        {"xs": 1, "s": 0, "m": 0, "l": 0, "xl": 0, "xxl": 0},
    ])
    assert _present_order_sizes(df) == ["xs", "s", "l"]


def test_present_order_sizes_falls_back_to_full_set_when_no_size_data():
    assert _present_order_sizes(pd.DataFrame([{"style": "S1"}])) == \
        ["xs", "s", "m", "l", "xl", "xxl"]


def test_present_order_sizes_all_zero_falls_back_to_full_set():
    df = pd.DataFrame([{"xs": 0, "s": 0, "m": 0, "l": 0, "xl": 0, "xxl": 0}])
    assert _present_order_sizes(df) == ["xs", "s", "m", "l", "xl", "xxl"]


def _basic_col() -> dict:
    keys = ["contract", "style", "brand", "article", "po", "config", "color_en",
            "color_cn", "label_clr", "xs", "s", "m", "l", "xl", "xxl",
            "total", "ex_fty"]
    return {k: i + 1 for i, k in enumerate(keys)}


def _ctx(**overrides) -> _RowContext:
    base = dict(
        col=_basic_col(), cn_lookup={}, cn_code_lookup={}, cn_by_pc_lookup=None,
        brand_by_pc=None,
        label_lookup={}, ai_enhance=False, ai_api_key="", ai_model="deepseek-chat",
        bsr_cache={}, boat_sample_col=18, se_store=None, color_source="db",
        sty_norm_cache={}, label_warnings=[], label_missing=[],
    )
    base.update(overrides)
    return _RowContext(**base)


# ── _FabricMasterCache.display_key ──────────────────────────────────────────

def test_display_key_uses_db_display_key_verbatim():
    c = _FabricMasterCache({"HHN-A": {"display_key": "Q1|Cotton|200|150"}})
    assert c.display_key("HHN-A") == "Q1|Cotton|200|150"
    assert c.misses == set()


def test_display_key_rebuilds_from_db_columns_when_no_display_key():
    c = _FabricMasterCache({"HHN-A": {
        "quality_no": "Q9", "composition_en": "Wool",
        "weight_gsm": 220, "cuttable_width_cm": 145,
    }})
    assert c.display_key("HHN-A") == "Q9|Wool|220|145"


def test_display_key_db_wins_but_caller_fills_null_db_fields():
    # composition present in DB, gsm/width NULL → caller fallbacks fill them.
    c = _FabricMasterCache({"HHN-A": {
        "quality_no": "Q9", "composition_en": "Wool",
        "weight_gsm": None, "cuttable_width_cm": None,
    }})
    out = c.display_key("HHN-A", fallback_gsm=180, fallback_width=140)
    assert out == "Q9|Wool|180|140"


def test_display_key_case_and_whitespace_insensitive():
    c = _FabricMasterCache({"HHN-A": {"display_key": "K"}})
    assert c.display_key("  hhn-a ") == "K"


def test_display_key_miss_records_and_builds_partial_from_fallbacks():
    c = _FabricMasterCache({})
    out = c.display_key("HHN-Z", fallback_composition="Poly")
    assert out == "HHN-Z|Poly||"
    assert c.misses == {"HHN-Z"}


def test_display_key_miss_with_no_fallbacks_returns_bare_hhn():
    c = _FabricMasterCache({})
    assert c.display_key("HHN-Z") == "HHN-Z"
    assert c.misses == {"HHN-Z"}


def test_display_key_empty_hhn_returns_empty_and_not_a_miss():
    c = _FabricMasterCache({})
    assert c.display_key("") == ""
    assert c.misses == set()


# ── pre-fetch helpers ───────────────────────────────────────────────────────

def test_prefetch_boat_sample_cache_empty_when_no_brand_column():
    df = pd.DataFrame([{"style": "S1"}])
    assert _prefetch_boat_sample_cache(df) == {}


def test_prefetch_fabric_master_cache_empty_df_yields_empty_cache():
    df = pd.DataFrame([{"style": "S1"}])   # no fabric_item_no, no parts
    fm = _prefetch_fabric_master_cache(None, df)
    assert isinstance(fm, _FabricMasterCache)
    assert fm.by_hhn == {}
    assert fm.misses == set()


# ── _fill_one_style_row (extracted per-row writer, testable on a blank sheet) ─

def test_fill_one_style_row_writes_cells_and_returns_overview():
    ws = Workbook().active
    col = _basic_col()
    grp = pd.DataFrame([
        {"style": "DR1", "brand": "Anna Field", "contract_no": "C1",
         "article_name": "A", "zalando_po": "PO1", "config_sku": "S1",
         "color_name": "dark blue", "pc_no": "PC1", "ex_fty_date": "2026-08-11",
         "xs": 1, "s": 2, "m": 0, "l": 0, "xl": 0, "xxl": 0},
        {"style": "DR1", "brand": "Anna Field", "contract_no": "C1",
         "article_name": "A", "zalando_po": "PO1", "config_sku": "S1",
         "color_name": "dark blue", "pc_no": "PC1", "ex_fty_date": "2026-08-11",
         "xs": 3, "s": 1, "m": 0, "l": 0, "xl": 0, "xxl": 0},
    ])
    total, ov = _fill_one_style_row(
        ws, 2, grp.iloc[0], grp, "DR1", "1_DR1", None, [], _ctx(col=col),
    )
    assert total == 7                                   # xs 1+3=4, s 2+1=3
    assert ov["style"] == "DR1"
    assert ov["color_en"] == "Dark Blue"                # bracket-strip + title
    assert ov["total"] == 7
    assert ov["label_color"] == ""                      # no label on file → blank, not derived
    assert ws.cell(2, col["style"]).value == "DR1"
    assert ws.cell(2, col["color_en"]).value == "Dark Blue"
    assert ws.cell(2, col["total"]).value == 7


def test_fill_one_style_row_uses_base_style_when_row_style_blank():
    ws = Workbook().active
    grp = pd.DataFrame([{
        "style": "", "brand": "B", "color_name": "Red",
        "xs": 1, "s": 0, "m": 0, "l": 0, "xl": 0, "xxl": 0,
    }])
    _, ov = _fill_one_style_row(
        ws, 2, grp.iloc[0], grp, "BASE9", "1_BASE9", None, [], _ctx(),
    )
    assert ov["style"] == "BASE9"


def test_fill_one_style_row_logs_miss_and_attaches_comment():
    calls = []

    class _Store:
        def log_color_miss(self, **kw):
            calls.append(kw)

    ws = Workbook().active
    col = _basic_col()
    grp = pd.DataFrame([{
        "style": "DR1", "brand": "Anna Field", "contract_no": "C1",
        "article_name": "A", "zalando_po": "PO1", "config_sku": "S1",
        "color_name": "Mauve", "pc_no": "PC1", "ex_fty_date": "",
        "xs": 1, "s": 0, "m": 0, "l": 0, "xl": 0, "xxl": 0,
    }])
    # cn_by_pc_lookup={} → progress source selected, no match → miss.
    ctx = _ctx(col=col, cn_by_pc_lookup={}, se_store=_Store(), color_source="progress")
    _, ov = _fill_one_style_row(ws, 2, grp.iloc[0], grp, "DR1", "1_DR1", None, [], ctx)

    assert ov["color_cn"] == _COLOR_NOT_FOUND
    assert len(calls) == 1
    assert calls[0]["style"] == "DR1" and calls[0]["source"] == "progress"
    assert ws.cell(2, col["color_cn"]).comment is not None   # diagnostic comment


def test_fill_one_style_row_injects_boat_sample_from_cache():
    ws = Workbook().active
    col = _basic_col()
    grp = pd.DataFrame([{
        "style": "DR1", "brand": "Anna Field", "color_name": "Red",
        "xs": 1, "s": 0, "m": 0, "l": 0, "xl": 0, "xxl": 0,
    }])
    ctx = _ctx(col=col, bsr_cache={"Anna Field": "confirm boat sample"}, boat_sample_col=18)
    _fill_one_style_row(ws, 2, grp.iloc[0], grp, "DR1", "1_DR1", None, [], ctx)
    assert ws.cell(2, 18).value == "confirm boat sample"


# ── 主标颜色: 大货进度表 authoritative, derived heuristic only cross-checks ────

def _navy_row():
    return pd.DataFrame([{
        "style": "DR1", "brand": "B", "pc_no": "PC1", "color_name": "Navy",
        "xs": 1, "s": 0, "m": 0, "l": 0, "xl": 0, "xxl": 0,
    }])


def test_label_uses_progress_value_and_flags_mismatch_with_derived():
    """大货进度表 says 白色 for a "Navy" body colour; the derived heuristic says
    黑色.  大货进度表's value is kept (not overridden), a cell comment is
    attached, and the disagreement is collected for the end-of-run warning.
    """
    from po_extractor.lookups.progress_lookup import PCColorMatch
    from po_extractor.store.color_translation_store import _normalize_color_name as _nz

    ws = Workbook().active
    col = _basic_col()
    cn_by_pc = {("PC1", "DR1", _nz("Navy")): PCColorMatch("藏青", "52#", "白色")}
    warnings: list = []
    ctx = _ctx(col=col, cn_by_pc_lookup=cn_by_pc, color_source="progress",
               label_warnings=warnings)
    _, ov = _fill_one_style_row(ws, 2, _navy_row().iloc[0], _navy_row(),
                                "DR1", "1_DR1", None, [], ctx)
    assert ov["label_color"] == "白色"                     # 大货进度表 kept, not derived 黑色
    assert ws.cell(2, col["label_clr"]).comment is not None
    assert warnings == [("DR1", "Navy", "白色", "黑色")]


def test_label_no_warning_when_progress_agrees_with_derived():
    from po_extractor.lookups.progress_lookup import PCColorMatch
    from po_extractor.store.color_translation_store import _normalize_color_name as _nz

    ws = Workbook().active
    col = _basic_col()
    cn_by_pc = {("PC1", "DR1", _nz("Navy")): PCColorMatch("藏青", "52#", "黑色")}
    warnings: list = []
    ctx = _ctx(col=col, cn_by_pc_lookup=cn_by_pc, color_source="progress",
               label_warnings=warnings)
    _, ov = _fill_one_style_row(ws, 2, _navy_row().iloc[0], _navy_row(),
                                "DR1", "1_DR1", None, [], ctx)
    assert ov["label_color"] == "黑色"
    assert ws.cell(2, col["label_clr"]).comment is None
    assert warnings == []


def test_label_missing_leaves_cell_blank_and_flags_it():
    # No 大货进度表 / DB value → the cell is genuinely blank (never derived),
    # gets a "missing" comment, and is collected in label_missing — not treated
    # as a mismatch (there was nothing authoritative to disagree with).
    ws = Workbook().active
    col = _basic_col()
    mismatches: list = []
    missing: list = []
    ctx = _ctx(col=col, cn_by_pc_lookup=None,
               label_warnings=mismatches, label_missing=missing)
    _, ov = _fill_one_style_row(ws, 2, _navy_row().iloc[0], _navy_row(),
                                "DR1", "1_DR1", None, [], ctx)
    assert ov["label_color"] == ""                        # blank, not derived 黑色
    assert ws.cell(2, col["label_clr"]).value in ("", None)
    assert ws.cell(2, col["label_clr"]).comment is not None   # "missing" comment
    assert mismatches == []
    assert missing == [("DR1", "Navy")]
