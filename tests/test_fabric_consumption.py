"""Tests for fabric consumption (单耗/排版): kg↔cm reconcile, template I/O, store."""
from __future__ import annotations

import io

import openpyxl

from po_extractor.ui_helpers.fabric_consumption import (
    reconcile_consumption, consumption_template_bytes, parse_consumption_upload,
    CONSUMPTION_COLUMNS,
)


# ── reconcile ─────────────────────────────────────────────────────────────────

def test_derive_kg_from_cm_uses_gross_width():
    # 毛门幅 = 有效门幅 150 + 5 = 155 → 165*155*200/1e7 = 0.5115
    kg, cm, warn = reconcile_consumption(None, 165, 150, 200)
    assert kg == 0.5115 and cm == 165 and warn == ""


def test_derive_cm_from_kg_uses_gross_width():
    kg, cm, warn = reconcile_consumption(0.5115, None, 150, 200)
    assert kg == 0.5115 and cm == 165.0 and warn == ""


def test_both_given_consistent_no_warning():
    kg, cm, warn = reconcile_consumption(0.5115, 165, 150, 200)
    assert warn == "" and kg == 0.5115 and cm == 165


def test_both_given_inconsistent_warns():
    kg, cm, warn = reconcile_consumption(0.90, 165, 150, 200)   # kg way off
    assert "不一致" in warn and "毛门幅" in warn
    assert kg == 0.90 and cm == 165          # values kept, not silently changed


def test_cannot_derive_without_width_or_gsm():
    kg, cm, warn = reconcile_consumption(None, 165, None, 200)  # no width
    assert kg is None and cm == 165 and "无法" in warn


def test_blank_row_no_warning():
    kg, cm, warn = reconcile_consumption(None, None, 150, 200)
    assert kg is None and cm is None and warn == ""


# ── template round-trip ───────────────────────────────────────────────────────

def test_template_has_all_columns_and_example():
    wb = openpyxl.load_workbook(io.BytesIO(consumption_template_bytes()))
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    assert headers == [h for h, _ in CONSUMPTION_COLUMNS]
    assert ws.cell(2, 1).value == "S5DTN67A"      # example style


def test_parse_upload_reconciles_and_keys_by_style():
    # build a filled template: cm given, kg blank → derived
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([h for h, _ in CONSUMPTION_COLUMNS])
    ws.append(["ST1", None, 165, 82, 24, 150, 200])
    ws.append(["", 1, 1, 1, 1, 1, 1])             # no style → skipped
    buf = io.BytesIO(); wb.save(buf)
    records, warns = parse_consumption_upload(buf.getvalue())
    assert len(records) == 1
    r = records[0]
    assert r["style"] == "ST1" and r["cons_kg"] == 0.5115 and r["cons_cm"] == 165


def test_parse_upload_missing_style_column():
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["单耗(kg)", "单耗(cm)"])
    buf = io.BytesIO(); wb.save(buf)
    records, warns = parse_consumption_upload(buf.getvalue())
    assert records == [] and any("款号" in w for w in warns)


# ── store round-trip ──────────────────────────────────────────────────────────

def test_store_save_load_roundtrip(tmp_path):
    from po_extractor.store.po_store import POStore
    store = POStore(str(tmp_path / "t.db"))
    n = store.save_fabric_consumption([
        {"style": "ST1", "cons_kg": 0.495, "cons_cm": 165, "util": 82,
         "marker_pcs": 24, "width_cm": 150, "gsm": 200},
    ])
    assert n == 1
    got = store.load_fabric_consumption(["ST1", "NOPE"])
    assert "NOPE" not in got
    assert got["ST1"]["cons_cm"] == 165 and got["ST1"]["gsm"] == 200
    # upsert (not duplicate) on same style
    store.save_fabric_consumption([{"style": "ST1", "cons_cm": 170}])
    assert store.load_fabric_consumption(["ST1"])["ST1"]["cons_cm"] == 170
    assert len(store.load_all_fabric_consumption()) == 1
    # clear wipes the table
    store.save_fabric_consumption([{"style": "ST2", "cons_cm": 1}])
    assert len(store.load_all_fabric_consumption()) == 2
    store.clear_fabric_consumption()
    assert store.load_all_fabric_consumption() == []
