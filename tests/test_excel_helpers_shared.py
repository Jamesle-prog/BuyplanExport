"""Shared Excel helpers consolidated in v2.125.2 (Phase 3 of the refactor):
border / fill / header / total / data styling, write_cell, template
placeholder + clear-area helpers, sheet-name de-duplication, file names."""
import openpyxl
import pytest

from po_extractor.exporters._excel_helpers import (
    clean_sheet_name, clear_data_area, replace_placeholders, safe_filename,
    solid_fill, style_data, style_header, style_total, thin_border,
    unique_sheet_name, write_cell,
)


# ── sheet names ─────────────────────────────────────────────────────────────

def test_unique_sheet_name_returns_base_when_free():
    assert unique_sheet_name("Style A", set()) == "Style A"


def test_unique_sheet_name_suffixes_in_order():
    taken = {"S", "S_2", "S_3"}
    assert unique_sheet_name("S", taken) == "S_4"


def test_unique_sheet_name_never_exceeds_31_chars():
    base = "A" * 31
    taken = {base} | {f"{base[:31 - len(f'_{i}')]}_{i}" for i in range(2, 150)}
    name = unique_sheet_name(base, taken)
    assert len(name) <= 31 and name.endswith("_150") and name not in taken


def test_unique_sheet_name_accepts_workbook_sheetnames():
    wb = openpyxl.Workbook()
    wb.active.title = "PO1"
    assert unique_sheet_name("PO1", wb.sheetnames) == "PO1_2"


def test_clean_then_unique_round_trip():
    used = set()
    a = unique_sheet_name(clean_sheet_name("A/B", fallback="PO"), used); used.add(a)
    b = unique_sheet_name(clean_sheet_name("A/B", fallback="PO"), used); used.add(b)
    assert (a, b) == ("A_B", "A_B_2")


# ── file names ──────────────────────────────────────────────────────────────

@pytest.mark.parametrize("raw, expected", [
    ("Sky East", "Sky_East"),
    ('a<b>:"c/d\\e|f?g*h', "a_b_c_d_e_f_g_h"),
    ("  ", "unknown"),
    (None, "unknown"),
    ("__x__", "x"),
])
def test_safe_filename(raw, expected):
    assert safe_filename(raw) == expected


def test_safe_filename_custom_fallback():
    assert safe_filename("", fallback="output") == "output"


# ── styling ─────────────────────────────────────────────────────────────────

def test_thin_border_is_cached_per_colour():
    assert thin_border("FF000000") is thin_border("FF000000")
    assert thin_border("FF000000") is not thin_border("FFAAAAAA")
    b = thin_border("FFAAAAAA")
    assert b.left.style == "thin" and b.left.color.rgb == "FFAAAAAA"
    assert b.right.style == b.top.style == b.bottom.style == "thin"


def test_solid_fill():
    f = solid_fill("FFFFFF00")
    assert f.fill_type == "solid" and f.start_color.rgb == "FFFFFF00"


def test_style_header_total_data():
    ws = openpyxl.Workbook().active
    style_header(ws["A1"], "H", fill="000000", font_color="FFFFFF", border_color="000000")
    assert ws["A1"].value == "H" and ws["A1"].font.bold and ws["A1"].font.color.rgb == "00FFFFFF"
    assert ws["A1"].fill.start_color.rgb == "00000000"
    assert ws["A1"].alignment.horizontal == "center" and ws["A1"].border.left.style == "thin"

    style_total(ws["A2"], 1234)
    assert ws["A2"].number_format == "#,##0" and ws["A2"].fill.start_color.rgb == "FFFFFF00"
    style_total(ws["A3"], "Total")
    assert ws["A3"].number_format == "General"

    style_data(ws["A4"], 5)
    assert ws["A4"].number_format == "#,##0" and ws["A4"].alignment.horizontal == "center"
    style_data(ws["A5"], "x")
    assert ws["A5"].number_format == "General" and ws["A5"].border.bottom.style == "thin"


def test_write_cell_defaults_and_flags():
    ws = openpyxl.Workbook().active
    c = write_cell(ws, 1, 1, "plain")
    assert c.font.name == "Arial" and c.font.size == 10 and not c.font.bold
    assert c.font.color.rgb == "FF000000" and c.alignment.horizontal == "left"
    assert c.alignment.wrap_text is True and c.fill.fill_type is None
    assert c.border.left.style is None                     # no border unless given

    c = write_cell(ws, 2, 1, 12.5, bold=True, bg="FF1F3864", white=True,
                   center=True, wrap=False, num="0.00", border=thin_border())
    assert c.font.bold and c.font.color.rgb == "FFFFFFFF"
    assert c.fill.fgColor.rgb == "FF1F3864" and c.alignment.horizontal == "center"
    assert c.alignment.wrap_text is False and c.number_format == "0.00"
    assert c.border.left.style == "thin"


def test_write_cell_partial_binding_overridable():
    from functools import partial
    ws = openpyxl.Workbook().active
    cell = partial(write_cell, ws, center=True, border=thin_border())
    assert cell(1, 1, "a").alignment.horizontal == "center"
    assert cell(1, 2, "b", center=False).alignment.horizontal == "left"


# ── template sheets ─────────────────────────────────────────────────────────

def test_replace_placeholders_only_touches_strings():
    ws = openpyxl.Workbook().active
    ws["A1"] = "Factory: {{factory}} / {{missing}}"
    ws["A2"] = 42
    ws["A3"] = "{{n}}"
    replace_placeholders(ws, {"factory": "XZY", "missing": None, "n": 7})
    assert ws["A1"].value == "Factory: XZY / "
    assert ws["A2"].value == 42
    assert ws["A3"].value == "7"


def test_clear_data_area_unmerges_and_blanks_from_start_row():
    ws = openpyxl.Workbook().active
    for r in range(1, 6):
        ws.cell(r, 1, f"r{r}"); ws.cell(r, 2, r)
    ws.merge_cells("A1:B1")
    ws.merge_cells("A4:B4")
    clear_data_area(ws, 3)
    assert ws["A1"].value == "r1" and ws["A2"].value == "r2"
    assert all(ws.cell(r, c).value is None for r in (3, 4, 5) for c in (1, 2))
    assert [str(m) for m in ws.merged_cells.ranges] == ["A1:B1"]
