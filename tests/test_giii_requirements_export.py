"""Tests for the upload-time PO requirements document (resolver + Excel)."""
from __future__ import annotations

import io

import openpyxl

from po_extractor.models.po_data import POData, POMetadata
from po_extractor.exporters.giii_requirements_export import export_giii_requirements
from po_extractor.ui_helpers.giii_requirements import resolve_po_requirements


class _Cprs:
    """Fake CPRS /evaluate/po — decodes a raw PO and returns {decoded, evaluation}."""
    def __init__(self):
        self.calls = 0

    def evaluate_po(self, raw):
        self.calls += 1
        brand = str(raw.get("brand", "")).upper()
        if not any(k in brand for k in ("DKNY", "KARL", "CALVIN")):
            return {"decoded": {}, "evaluation": {"results": []}}   # not decoded
        acct = "MACYS" if "MACY" in str(raw.get("account", "")).upper() else ""
        return {"decoded": {
                    "clientId": "a1", "clientName": brand, "channel": "WHOLESALE",
                    "accountCode": acct, "warehouseCode": raw.get("warehouseCode", ""),
                    "warehouseInfo": {"region": "US", "rfid_default": False,
                                      "msrp_required_default": False},
                    "warnings": []},
                "evaluation": {"results": [
                    {"domain": "label", "subtype": "care_label", "status": "confirmed",
                     "resultJson": {"standard": "Care label per FTC", "source": "Manual p3"}},
                    {"domain": "carton", "subtype": "red_carton_sticker",
                     "status": "pending_input", "resultJson": {"waiting_for": "dim_code"}},
                    {"domain": "packaging", "subtype": "polybag",
                     "status": "not_applicable", "resultJson": {}},
                ]}}

    def manual_image(self, image_id):
        return None


def _po(po="PO1", division="DKNY Sportswear", dest="UC", customer="MY MACY'S"):
    # buyer is the G-III vendor entity (never a CPRS account); the retail
    # account is the PO's customer — that's what the resolver sends.
    return POData(metadata=POMetadata(po_number=po, style="ST1",
                                      division_name=division, destination_code=dest,
                                      buyer="G-III APPAREL GROUP LTD",
                                      customer=customer,
                                      country_of_origin="China"))


# ── resolver ─────────────────────────────────────────────────────────────────

def test_resolver_builds_contexts_from_decoded():
    cprs = _Cprs()
    contexts, warns = resolve_po_requirements(cprs, [_po("PO1"), _po("PO2")])
    assert len(contexts) == 2
    assert cprs.calls == 2                       # one /evaluate/po per PO
    assert contexts[0]["warehouse"] == "UC"      # from CPRS decoded.warehouseCode
    assert contexts[0]["account"] == "MACYS"     # from CPRS decoded.accountCode
    assert contexts[0]["region"] == "US"
    assert warns == []


def test_resolver_unknown_brand_warns_and_skips():
    contexts, warns = resolve_po_requirements(_Cprs(), [_po(division="ACME")])
    assert contexts == []
    assert any("not decoded" in w for w in warns)


def test_resolver_no_cprs():
    contexts, warns = resolve_po_requirements(None, [_po()])
    assert contexts == [] and any("not configured" in w for w in warns)


def test_resolver_account_from_customer_and_wrh_stripped():
    """The account is the PO's customer (retail account), not the G-III buyer;
    and the 'WRH' warehouse prefix is stripped before it reaches CPRS."""
    from po_extractor.models.po_data import POData, POMetadata
    cprs = _Cprs()
    po = POData(metadata=POMetadata(
        po_number="DW843124UC", style="ST", division_name="DW",
        destination_code="WRHUC", buyer="G-III APPAREL GROUP LTD",
        customer="MY MACY'S", country_of_origin="China"))
    contexts, warns = resolve_po_requirements(cprs, [po])
    assert len(contexts) == 1
    assert contexts[0]["account"] == "MACYS"    # from customer, NOT the G-III buyer
    assert contexts[0]["warehouse"] == "UC"     # 'WRH' prefix stripped


def test_resolver_no_customer_sends_no_account():
    """A G-III-direct PO (no customer) sends no account — the buyer is never a
    CPRS account, so we don't create a false 'not matched' signal."""
    from po_extractor.models.po_data import POData, POMetadata
    cprs = _Cprs()
    po = POData(metadata=POMetadata(
        po_number="DW843120DN", style="ST", division_name="DW",
        destination_code="WRHDN", buyer="G-III LEATHER FASHIONS",
        customer=None, country_of_origin="China"))
    contexts, _ = resolve_po_requirements(cprs, [po])
    assert len(contexts) == 1
    assert contexts[0]["account"] == ""         # no customer → no account


def test_resolver_cprs_down_one_reason_no_evaluate():
    """CPRS outage (health() down) → ONE actionable warning, evaluate_po never
    called, no per-PO noise."""
    class _Outage(_Cprs):
        def health(self):
            return False, "Could not reach CPRS: Connection refused"
        def evaluate_po(self, raw):
            raise AssertionError("evaluate_po called despite CPRS being down")
    contexts, warns = resolve_po_requirements(_Outage(), [_po("PO1"), _po("PO2")])
    assert contexts == []
    assert len(warns) == 1 and "not reachable" in warns[0]
    assert "Connection refused" in warns[0]


# ── exporter ─────────────────────────────────────────────────────────────────

def _sample_results():
    return [
        {"domain": "label", "subtype": "care_label", "status": "confirmed",
         "resultJson": {"standard": "Care label per FTC", "source": "Manual p3"}},
        {"domain": "carton", "subtype": "red_carton_sticker",
         "status": "pending_input", "resultJson": {"waiting_for": "dim_code"}},
        {"domain": "packaging", "subtype": "polybag", "status": "not_applicable",
         "resultJson": {}},
    ]


def _ctx(po="PO1"):
    return {"po_number": po, "style": "ST1", "brand": "DKNY Sportswear",
            "warehouse": "UC", "account": "MACYS", "channel": "WHOLESALE",
            "results": _sample_results()}


def test_export_summary_and_per_po_sheets():
    data = export_giii_requirements([_ctx("PO1"), _ctx("PO2")],
                                    warnings=["something to know"])
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert wb.sheetnames[0] == "PO Index"          # KL layout leads with PO Index
    assert "Summary 汇总" in wb.sheetnames
    assert "PO1" in wb.sheetnames and "PO2" in wb.sheetnames

    ws = wb["Summary 汇总"]
    # header + one row per PO; counts: 1 confirmed, 1 pending, 0 conflict, 1 N/A
    assert ws.cell(2, 1).value == "PO1"
    assert ws.cell(2, 7).value == 1 and ws.cell(2, 8).value == 1
    assert ws.cell(2, 10).value == 1
    flat = [c.value for row in ws.iter_rows() for c in row]
    assert "something to know" in flat

    s = wb["PO1"]
    rows = [[c.value for c in row] for row in s.iter_rows()]
    flat = [c for r in rows for c in r]
    assert "care_label" in flat
    assert any(v and "待定" in str(v) for v in flat)      # pending marker
    assert "Care label per FTC" in " ".join(str(v) for v in flat if v)
    # N/A rows sort last: last data row is the polybag one
    assert "polybag" in [str(v) for v in rows[-1]]


def test_missing_mandatory_context_counts_as_pending():
    """A PO whose results all need context must not show 0/0/0/0 in Summary."""
    ctx = {"po_number": "PO9", "style": "S", "brand": "B", "warehouse": "",
           "account": "", "channel": "WHOLESALE",
           "results": [{"domain": "label", "subtype": "x",
                        "status": "missing_mandatory_context", "resultJson": {}}]}
    wb = openpyxl.load_workbook(io.BytesIO(export_giii_requirements([ctx])))
    ws = wb["Summary 汇总"]
    assert ws.cell(2, 8).value == 1        # folded into 待定 Pending
    flat = [c.value for row in wb["PO9"].iter_rows() for c in row]
    assert any(v and "缺少信息" in str(v) for v in flat)


def test_export_duplicate_po_numbers_get_unique_sheets():
    data = export_giii_requirements([_ctx("PO1"), _ctx("PO1")])
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert "PO1" in wb.sheetnames and "PO1_2" in wb.sheetnames     # unique per-PO


def test_export_sanitizes_bad_sheet_names():
    ctx = _ctx("PO/1:*?")
    data = export_giii_requirements([ctx])
    wb = openpyxl.load_workbook(io.BytesIO(data))
    # no crash; the fixed KL-style sheets are present alongside the sanitized PO
    assert "PO Index" in wb.sheetnames and "Requirement Matrix" in wb.sheetnames


# ── KL-style illustrated sheets ───────────────────────────────────────────────

def _ctx_kl(po, style, account, region, units, results):
    return {"po_number": po, "style": style, "brand": "Karl Lagerfeld",
            "warehouse": "NJ", "account": account, "channel": "WHOLESALE",
            "units": units, "article": "KL-9727", "region": region,
            "destination": "US — Dayton, NJ", "packing": "Flat pack", "msrp": "59",
            "coo": "China", "source_file": f"{po}.pdf", "is_prepack": False,
            "results": results}


def test_full_kl_layout_sheet_order():
    r = [{"domain": "label", "subtype": "main_label", "status": "confirmed",
          "resultJson": {"standard": "Main label", "source": "KL Manual"}}]
    data = export_giii_requirements([_ctx_kl("P1", "ST1", "", "US", 80, r)])
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert wb.sheetnames[:7] == [
        "PO Index", "Summary 汇总", "款号对比 By Style", "Requirements + Pictures",
        "Requirement Matrix", "Pre-pack", "Actions & Confirm"]


def test_po_index_sheet_lists_pos_and_total():
    r = [{"domain": "label", "subtype": "x", "status": "confirmed", "resultJson": {}}]
    data = export_giii_requirements([
        _ctx_kl("P1", "ST1", "", "US", 80, r),
        _ctx_kl("P2", "ST1", "AMRG", "US", 3740, r)])
    ws = openpyxl.load_workbook(io.BytesIO(data))["PO Index"]
    hdr = [ws.cell(2, c).value for c in range(1, 11)]
    assert hdr[:4] == ["PO Number", "款号 Style", "品名 Article", "数量 Units"]
    vals = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
    assert "P1" in vals and "P2" in vals and "3740" in vals
    assert "TOTAL" in vals and "3820" in vals            # 80 + 3740


def test_requirement_matrix_columns_by_destination():
    common = {"domain": "label", "subtype": "main_label", "status": "confirmed",
              "resultJson": {"standard": "Main label"}}
    only_amrg = {"domain": "carton", "subtype": "red_sticker",
                 "status": "confirmed", "resultJson": {}}
    data = export_giii_requirements([
        _ctx_kl("P1", "ST1", "", "US", 80, [dict(common)]),
        _ctx_kl("P2", "ST1", "AMRG", "US", 100, [dict(common), dict(only_amrg)])])
    ws = openpyxl.load_workbook(io.BytesIO(data))["Requirement Matrix"]
    hdr = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
    assert hdr[0] == "Domain / Subtype" and "STOCK (US)" in hdr and "AMRG (US)" in hdr
    rows = [[c.value for c in row] for row in ws.iter_rows(min_row=3)]
    red = next(r for r in rows if r[0] == "carton/red_sticker")
    si, ai = hdr.index("STOCK (US)"), hdr.index("AMRG (US)")
    assert red[ai] == "✓" and red[si] == "—"       # only AMRG has the red sticker


def test_actions_sheet_flags_conflicts():
    r = [{"domain": "hangtag", "subtype": "main_hangtag", "status": "conflict",
          "resultJson": {"standard": "two rules disagree"}}]
    data = export_giii_requirements([_ctx_kl("P1", "ST1", "", "US", 80, r)])
    ws = openpyxl.load_workbook(io.BytesIO(data))["Actions & Confirm"]
    vals = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
    assert any("Conflict" in v for v in vals) and "Confirm" in vals


# ── by-style comparison sheet ─────────────────────────────────────────────────

def _ctx_style(po, style, results):
    return {"po_number": po, "style": style, "brand": "DKNY Sportswear",
            "warehouse": "UC", "account": "MACYS", "channel": "WHOLESALE",
            "results": results}


def test_by_style_sheet_combines_shared_and_splits_differing():
    shared = {"domain": "label", "subtype": "care_label", "status": "confirmed",
              "resultJson": {"standard": "Care label per FTC"}}
    reqA = {"domain": "carton", "subtype": "carton_mark", "status": "confirmed",
            "resultJson": {"standard": "Mark ABC"}}
    reqB = {"domain": "carton", "subtype": "carton_mark", "status": "confirmed",
            "resultJson": {"standard": "Mark XYZ"}}
    data = export_giii_requirements([
        _ctx_style("PO1", "STYLE_A", [dict(shared), dict(reqA)]),
        _ctx_style("PO2", "STYLE_B", [dict(shared), dict(reqB)]),
    ])
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert "款号对比 By Style" in wb.sheetnames
    s = wb["款号对比 By Style"]
    rows = [[c.value for c in row] for row in s.iter_rows(min_row=3)]
    # care_label is identical across both styles → one row, "全部 All (2)"
    care = [r for r in rows if r[1] == "care_label"]
    assert len(care) == 1 and "全部 All (2)" in str(care[0][4])
    # carton_mark differs → two rows, each scoped to its own style
    marks = [r for r in rows if r[1] == "carton_mark"]
    assert len(marks) == 2
    style_cells = {str(r[4]) for r in marks}
    assert style_cells == {"STYLE_A", "STYLE_B"}
    assert any("Mark ABC" in str(r[3]) for r in marks)
    assert any("Mark XYZ" in str(r[3]) for r in marks)


# ── images ────────────────────────────────────────────────────────────────────

def _png(color="red") -> bytes:
    from PIL import Image
    b = io.BytesIO()
    Image.new("RGB", (30, 20), color).save(b, format="PNG")
    return b.getvalue()


def test_all_pictures_embedded_on_po_sheet():
    res = {"domain": "carton", "subtype": "red_carton_sticker", "status": "confirmed",
           "resultJson": {}, "_images": [_png("red"), _png("blue")]}
    ctx = _ctx_style("PO1", "ST1", [res])
    wb = openpyxl.load_workbook(io.BytesIO(export_giii_requirements([ctx])))
    s = wb["PO1"]
    assert s.cell(2, 6).value == "图示 Image"        # image column header
    assert len(s._images) == 2                       # BOTH pictures embedded


def test_bad_image_bytes_do_not_break_export():
    res = {"domain": "carton", "subtype": "x", "status": "confirmed",
           "resultJson": {}, "_images": [b"not-a-png"]}
    wb = openpyxl.load_workbook(io.BytesIO(
        export_giii_requirements([_ctx_style("PO1", "ST1", [res])])))
    assert wb["PO1"]._images == []                   # skipped, no crash


# ── resolver attaches all image bytes ─────────────────────────────────────────

class _CprsImg(_Cprs):
    def __init__(self):
        super().__init__()
        self.fetches = []

    def evaluate_po(self, raw):
        self.calls += 1
        return {"decoded": {"clientId": "a1", "warehouseCode": raw.get("warehouseCode", ""),
                            "accountCode": "", "channel": "WHOLESALE",
                            "warehouseInfo": {}, "warnings": []},
                "evaluation": {"results": [
                    {"domain": "carton", "subtype": "red_carton_sticker",
                     "status": "confirmed", "resultJson": {},
                     "images": [{"id": "img-1"}, {"id": "img-2"}]},
                ]}}

    def manual_image(self, image_id):
        self.fetches.append(image_id)
        return f"BYTES:{image_id}".encode()


def test_resolver_attaches_all_image_bytes_deduped():
    cprs = _CprsImg()
    contexts, _ = resolve_po_requirements(cprs, [_po("PO1"), _po("PO2")])
    imgs = contexts[0]["results"][0]["_images"]
    assert imgs == [b"BYTES:img-1", b"BYTES:img-2"]   # ALL images fetched
    # two POs share one order context → each image fetched once, not per-PO
    assert sorted(cprs.fetches) == ["img-1", "img-2"]
