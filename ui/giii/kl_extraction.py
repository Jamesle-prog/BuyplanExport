"""KL PO section for the GIII Upload tab.

Accepts PDF files for KL-format purchase orders (G-III / Ross Stores).
Layout is nearly identical to MSG/CSKHHA but with:
  - Direct PDF upload (no .msg wrapper)
  - MSRP DETAILS on page 3: "001 G5DTN93C JVS MSRP $69.00"
  - FOB line has two prices; take tariff-adjusted (2nd value before WW//)
  - CPO encoded as doubled "CCUUSSTT PPOO::"
  - HANGER appears at end of description line (not doubled)
"""
from __future__ import annotations

import io
import re

import streamlit as st

from ui.giii._shared import (
    _XLSX_MIME, _undouble, _SIZE_CODES, _FIRST_RE, _CONT_RE, files_signature,
    FAX_SIZE_ORDER, XL_NAVY, XL_WHITE, XL_YELLOW, XL_GREY, XL_LTBLUE, XL_GREEN,
    drop_stale_results, iter_pdf_payloads, make_excel_style_kit,
    persist_fax_pos,
)
from ui.i18n import t
from ui.session_keys import SK
from ui.shared import _th

# (parser helpers are imported from _shared)


def _parse_kl_pdf(pdf_bytes: bytes) -> dict:
    """Parse a KL-format PO PDF. Returns a PO dict."""
    import pdfplumber

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        all_text  = '\n'.join(p.extract_text() or '' for p in pdf.pages)
        text_lines = all_text.split('\n')

    def grep(pat: str):
        for l in text_lines:
            m = re.search(pat, l)
            if m:
                return m
        return None

    # ── Standard header fields ────────────────────────────────────────────────
    m = grep(r'PO NUMBER\s+(\S+)')
    po_number  = _undouble(m.group(1)) if m else '?'
    m = grep(r'S T Y L E #\s+(\S+)')
    style   = _undouble(m.group(1)) if m else '?'
    m = grep(r'PO DATE\s+([\d/]+)')
    po_date = _undouble(m.group(1)) if m else '?'
    m = grep(r'P R T\s+([\d/]+)')
    ship    = _undouble(m.group(1)) if m else '?'

    etd_m = grep(r'([\d/]+)\s+EETTDD')
    etd   = _undouble(etd_m.group(1)) if etd_m else '?'

    vline  = text_lines[3] if len(text_lines) > 3 else ''
    vendor = _undouble(re.sub(r'\s*S T Y L E #.*', '', vline).strip())

    fact_m  = grep(r'F A C T O R Y\s+(.+?)\s+INCO')
    factory = _undouble(fact_m.group(1)).strip() if fact_m else '?'

    # ── FOB price: KL has "FFOOBB::$$4.17// $$3.75 WW//TTAARRIIFFFF"
    #    Take the tariff-adjusted price (2nd $$ value, before WW//).
    fob = '?'
    for l in text_lines:
        if re.search(r'FFOOBB::\s*UUNNCCOONNFFIIRRMMEEDD', l):
            fob = 'UNCONFIRMED'
            break
        m = re.search(r'FFOOBB::.*?\$\$([\d.]+)\s+WW', l)
        if m:
            fob = _undouble(m.group(1))
            break
        if re.search(r'FFOOBB', l):
            m2 = re.search(r'\$\$([\d.]+)', l)
            if m2:
                fob = _undouble(m2.group(1))
                break

    # ── Description: strip leading "DESCRIPTION " and trailing "HANGER" ──────
    desc_m = grep(r'DESCRIPTION\s+(.+)')
    if desc_m:
        raw_desc = desc_m.group(1)
        description = re.sub(r'\s+HANGER\s*$', '', _undouble(raw_desc)).strip()
    else:
        description = '?'

    # ── Customer name ─────────────────────────────────────────────────────────
    cust_m        = grep(r'PO NUMBER\s+\S+\s+(.+)')
    customer_name = _undouble(cust_m.group(1).strip()) if cust_m else '?'

    # ── Ship To ───────────────────────────────────────────────────────────────
    ship_street   = _undouble(text_lines[5]).strip() if len(text_lines) > 5 else ''
    ship_type_raw = re.sub(r'\s*C N T R Y.*', '', text_lines[6] if len(text_lines) > 6 else '').strip()
    ship_type     = _undouble(ship_type_raw)
    city_m        = grep(r'PO DATE\s+[\d/]+\s+(.+?)\s+PPRRIICCEE')
    ship_city     = _undouble(city_m.group(1).strip()) if city_m else ''
    ship_to       = ' / '.join(p for p in [customer_name, ship_street, ship_type, ship_city] if p and p != '?')

    # ── Hanger info: look for line with HHAANNGGEERR (doubled) or HANGER ────────
    hanger_line = next(
        (l for l in text_lines if 'HHAANNGGEERR' in l or
         ('DESCRIPTION' in l and 'HANGER' in l.upper())),
        None,
    )
    hanger_info = _undouble(hanger_line).strip() if hanger_line else '?'

    # ── Pack ratio: scan all lines undoubled for "(d-d-d-d)" pattern ─────────
    pack_ratio = '?'
    for raw_l in text_lines:
        pr_m = re.search(r'\(([\d]+-[\d-]+)\)', _undouble(raw_l))
        if pr_m:
            pack_ratio = pr_m.group(1)
            break

    # ── HTS# (page 2, normal font; doubled variant carries "..") ─────────────
    # Only undouble a fax-doubled match — unconditional undoubling corrupted
    # normal-font codes with repeated digits: 6110.20.2079 → 610.20.2079.
    hts_m   = grep(r'(\d{4,8}\.+\d{2,4}\.+\d{4,8})')
    if hts_m:
        hts_raw = hts_m.group(1)
        hts_num = _undouble(hts_raw) if '..' in hts_raw else hts_raw
    else:
        hts_num = '?'

    # ── CPO: KL uses "CUST PO:" (doubled: "CCUUSSTT PPOO::") ─────────────────
    cpo_m = grep(r'CCUUSSTT\s+PPOO::\s*(\S+)')
    if cpo_m:
        cpo = _undouble(cpo_m.group(1))
    else:
        cpo_fb = grep(r'C+P+O+:+\s*(\S+)')
        cpo    = _undouble(cpo_fb.group(1)) if cpo_fb else '?'

    # ── MSRP: page 3 line "001 G5DTN93C JVS MSRP $69.00" (normal font) ───────
    msrp_m = grep(r'MSRP\s+\$([\d.]+)')
    if msrp_m:
        msrp = msrp_m.group(1)
    else:
        msrp_d = grep(r'MMSSRRPP::\s*\$\$?([\d.]+)')
        msrp   = _undouble(msrp_d.group(1)) if msrp_d else '?'

    # ── Table rows ────────────────────────────────────────────────────────────
    line_items: list[dict] = []
    in_table   = False
    dash_count = 0
    cur_item: dict | None = None

    for l in text_lines:
        if re.match(r'-{30,}', l):
            dash_count += 1
            if dash_count == 4:
                in_table = True
            continue
        if not in_table:
            continue
        if re.match(r'TOTAL\s+', l) or 'FLAT PACK' in l or 'PACK TTL' in l:
            continue

        m = _FIRST_RE.match(l)
        if m:
            cur_item = {
                'ln':    m.group(1),
                'style': m.group(2),
                'color': m.group(3),
                'sizes': [(m.group(4), int(m.group(5)), m.group(6), float(m.group(7)))],
            }
            line_items.append(cur_item)
            continue

        m2 = _CONT_RE.match(l)
        if m2 and cur_item:
            cur_item['sizes'].append((m2.group(1), int(m2.group(2)), m2.group(3), None))

    return dict(
        po_number=po_number, style=style, po_date=po_date, ship_date=ship,
        etd=etd, vendor=vendor, factory=factory, fob_price=fob,
        description=description, line_items=line_items,
        customer_name=customer_name, ship_to=ship_to,
        hanger_info=hanger_info, pack_ratio=pack_ratio,
        hts_num=hts_num, cpo=cpo, msrp=msrp,
    )


def _parse_kl_pdfs(pdf_files) -> list[dict]:
    """Accept Streamlit UploadedFiles — bare .pdf or .msg-wrapped (the
    auto-router can hand this section KL POs that arrived as fax emails) —
    and return parsed PO dicts."""
    results: list[dict] = []
    for name, pdf_data, _subject in iter_pdf_payloads(pdf_files):
        try:
            po = _parse_kl_pdf(pdf_data)
        except Exception as exc:
            st.warning(f"{t('Parse error in')} {name}: {exc}")
            continue
        if po['po_number'] == '?':
            # derive from filename e.g. LSKHHN015R-G5DTN93C 5.13.pdf
            po['po_number'] = name.split('-')[0] if '-' in name else name
        po['source_file'] = name
        results.append(po)
    return results


# ---------------------------------------------------------------------------
# Excel builder
# ---------------------------------------------------------------------------

_SIZE_ORDER = FAX_SIZE_ORDER

_NAVY    = XL_NAVY
_WHITE   = XL_WHITE
_YELLOW  = XL_YELLOW
_ORANGE  = 'FFF4B942'
_LT_BLUE = XL_LTBLUE
_GREY    = XL_GREY
_GREEN   = XL_GREEN


@st.cache_data(max_entries=8, show_spinner=False)
# Pure: results in -> xlsx bytes out. Uncached, the whole workbook was
# rebuilt on every rerun of this page -- behind a visible "Building
# Excel..." spinner, so every click on the page paid for it whether or
# not anyone wanted the download. Same treatment the summary tab's
# builders already have.
def _build_kl_excel(results: list[dict]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    all_sizes  = {s[0] for po in results for li in po['line_items'] for s in li['sizes']}
    sizes_cols = [s for s in _SIZE_ORDER if s in all_sizes]

    _kit = make_excel_style_kit()
    _border, _fill, _align = _kit.border, _kit.fill, _kit.align
    _style, _hdr, _autofit = _kit.style, _kit.hdr, _kit.autofit

    wb = Workbook()

    # ── Sheet 1: PO Detail ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title        = 'PO Detail'
    ws1.freeze_panes = 'A2'
    ws1.row_dimensions[1].height = 28

    D_HDRS = [
        'PO Number', 'Style', 'Color', 'Size', 'Units',
        'UPC', 'Unit Price (FOB)', 'MSRP', 'Line Total ($)',
        'ETD', 'PO Date', 'Ship Date',
        'Customer Name', 'Ship To', 'Hanger Info', 'Pack Ratio',
        'HTS#', 'CPO',
        'Description', 'Factory', 'Vendor',
    ]
    for ci, h in enumerate(D_HDRS, 1):
        _hdr(ws1.cell(1, ci), h)

    # MSRP column index
    msrp_col = D_HDRS.index('MSRP') + 1

    dr = 2
    for po in results:
        is_unc = po['fob_price'] == 'UNCONFIRMED'
        try:
            msrp_val = float(po['msrp']) if po['msrp'] not in ('?', '') else None
        except ValueError:
            msrp_val = None

        for li in po['line_items']:
            first_price = li['sizes'][0][3]
            for sz, units, upc, price in li['sizes']:
                unit_price = price if price is not None else first_price
                bg = _LT_BLUE if dr % 2 == 0 else None

                vals = [
                    po['po_number'], li['style'], li['color'], sz, units,
                    upc, unit_price, msrp_val, None,   # placeholder for Line Total
                    po['etd'], po['po_date'], po['ship_date'],
                    po['customer_name'], po['ship_to'], po['hanger_info'], po['pack_ratio'],
                    po['hts_num'], po['cpo'],
                    po['description'], po['factory'], po['vendor'],
                ]
                for ci, v in enumerate(vals, 1):
                    c = ws1.cell(dr, ci, v)
                    _style(c, bg=bg, align='right' if ci in (5, 7, 8, 9) else 'left')

                # MSRP formatting (green highlight)
                mc = ws1.cell(dr, msrp_col)
                _style(mc, bg=_GREEN if msrp_val else bg, align='right', num_fmt='$#,##0.00')

                # Line Total formula  (column 9 = H in 1-indexed = I actually)
                lt_col_idx = D_HDRS.index('Line Total ($)') + 1
                e_col = get_column_letter(D_HDRS.index('Units') + 1)
                g_col = get_column_letter(D_HDRS.index('Unit Price (FOB)') + 1)
                lt = ws1.cell(dr, lt_col_idx)
                lt.value          = f'={e_col}{dr}*{g_col}{dr}'
                lt.number_format  = '$#,##0.00'
                lt.fill           = _fill(_ORANGE if is_unc else bg)
                lt.border         = _border()
                lt.alignment      = _align('right')
                lt.font           = Font(name='Arial', bold=is_unc, size=10)

                # Unit price formatting
                pc = ws1.cell(dr, D_HDRS.index('Unit Price (FOB)') + 1)
                if is_unc and unit_price is None:
                    pc.value = 'UNCONFIRMED'
                    _style(pc, bold=True, bg=_ORANGE, align='center')
                else:
                    _style(pc, bg=bg, align='right', num_fmt='$#,##0.00')

                _style(ws1.cell(dr, D_HDRS.index('Units') + 1),
                       bg=bg, align='right', num_fmt='#,##0')

                dr += 1

    # Grand total row
    gt = dr
    ws1.cell(gt, 1, 'GRAND TOTAL')
    _style(ws1.cell(gt, 1), bold=True, bg=_YELLOW)
    for ci in range(2, len(D_HDRS) + 1):
        _style(ws1.cell(gt, ci), bold=True, bg=_YELLOW)

    u_col  = get_column_letter(D_HDRS.index('Units') + 1)
    lt_col = get_column_letter(D_HDRS.index('Line Total ($)') + 1)

    c_units = ws1.cell(gt, D_HDRS.index('Units') + 1)
    c_units.value         = f'=SUM({u_col}2:{u_col}{gt-1})'
    c_units.number_format = '#,##0'
    c_units.fill          = _fill(_YELLOW)
    c_units.border        = _border()
    c_units.alignment     = _align('right')
    c_units.font          = Font(name='Arial', bold=True, size=10)

    c_total = ws1.cell(gt, D_HDRS.index('Line Total ($)') + 1)
    c_total.value         = f'=SUM({lt_col}2:{lt_col}{gt-1})'
    c_total.number_format = '$#,##0.00'
    c_total.fill          = _fill(_YELLOW)
    c_total.border        = _border()
    c_total.alignment     = _align('right')
    c_total.font          = Font(name='Arial', bold=True, size=10)

    _autofit(ws1)

    # ── Sheet 2: Summary ────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Summary')
    ws2.freeze_panes          = 'A2'
    ws2.row_dimensions[1].height = 28

    S_HDRS = (['PO Number', 'Style', 'Description', 'Color', 'ETD', 'FOB Price', 'MSRP']
              + sizes_cols + ['Total Units'])
    for ci, h in enumerate(S_HDRS, 1):
        _hdr(ws2.cell(1, ci), h)

    po_row   = 2
    style_rows: dict[str, list[int]] = {}

    for po in results:
        is_unc = po['fob_price'] == 'UNCONFIRMED'
        try:
            msrp_val = float(po['msrp']) if po['msrp'] not in ('?', '') else ''
        except ValueError:
            msrp_val = ''
        if is_unc:
            fob_val = po['fob_price']
        elif po['fob_price'] != '?':
            try:
                fob_val = float(po['fob_price'])
            except ValueError:
                fob_val = po['fob_price']       # e.g. "12.50." — keep raw string
        else:
            fob_val = ''

        for li in po['line_items']:
            bg        = _LT_BLUE if (po_row - 2) % 2 == 0 else None
            sizes_map = {s[0]: s[1] for s in li['sizes']}
            row_total = sum(s[1] for s in li['sizes'])

            vals = (
                [po['po_number'], li['style'], po['description'], li['color'],
                 po['etd'], fob_val, msrp_val]
                + [sizes_map.get(sz, '') for sz in sizes_cols]
                + [row_total]
            )
            for ci, v in enumerate(vals, 1):
                c = ws2.cell(po_row, ci, v)
                is_price  = ci == 6
                is_msrp   = ci == 7
                is_size   = 8 <= ci <= 7 + len(sizes_cols)
                is_totcol = ci == len(S_HDRS)
                aln  = 'right' if (is_price or is_msrp or is_size or is_totcol) else 'left'
                c_bg = _ORANGE if (is_price and is_unc) else (_GREEN if (is_msrp and msrp_val != '') else bg)
                fmt  = None
                if is_size or is_totcol:           fmt = '#,##0'
                if is_price and not is_unc and fob_val != '': fmt = '$#,##0.00'
                if is_msrp and msrp_val != '':     fmt = '$#,##0.00'
                _style(c, bold=(is_price and is_unc), bg=c_bg, align=aln, num_fmt=fmt)

            style_rows.setdefault(li['style'], []).append(po_row)
            po_row += 1

    last_po_row = po_row - 1

    # Blank + style rollup header
    po_row += 1
    rl_hdr = po_row
    for ci, h in enumerate(['Style', 'Description', 'FOB Price', 'MSRP', 'Total Units', 'PO Count'], 1):
        _hdr(ws2.cell(rl_hdr, ci), h)
    po_row += 1

    seen_styles: dict[str, dict] = {}
    for po in results:
        try:
            msrp_val = float(po['msrp']) if po['msrp'] not in ('?', '') else ''
        except ValueError:
            msrp_val = ''
        for li in po['line_items']:
            k = li['style']
            if k not in seen_styles:
                seen_styles[k] = {'desc': po['description'], 'fob': po['fob_price'],
                                  'msrp': msrp_val, 'pos': set()}
            seen_styles[k]['pos'].add(po['po_number'])

    for style_key, info in seen_styles.items():
        rows    = style_rows.get(style_key, [])
        tu_col  = get_column_letter(len(S_HDRS))
        formula = f'=SUM({",".join(f"{tu_col}{r}" for r in rows)})' if rows else 0
        is_unc  = info['fob'] == 'UNCONFIRMED'
        if is_unc:
            fob_val = info['fob']
        elif info['fob'] != '?':
            try:
                fob_val = float(info['fob'])
            except ValueError:
                fob_val = info['fob']           # e.g. "12.50." — keep raw string
        else:
            fob_val = ''
        vals    = [style_key, info['desc'], fob_val, info['msrp'], formula, len(info['pos'])]

        for ci, v in enumerate(vals, 1):
            c        = ws2.cell(po_row, ci, v)
            is_fob   = ci == 3
            is_msrp  = ci == 4
            fmt = None
            if ci == 5:                             fmt = '#,##0'
            if is_fob and not is_unc and fob_val != '': fmt = '$#,##0.00'
            if is_msrp and info['msrp'] != '':      fmt = '$#,##0.00'
            c_bg = _ORANGE if (is_fob and is_unc) else (_GREEN if (is_msrp and info['msrp'] != '') else _GREY)
            _style(c, bold=True, bg=c_bg, align='right' if ci >= 3 else 'left', num_fmt=fmt)

        po_row += 1

    last_rl_row = po_row - 1

    # Grand total (summary)
    gt2 = po_row
    ws2.cell(gt2, 1, 'GRAND TOTAL')
    _style(ws2.cell(gt2, 1), bold=True, bg=_YELLOW)
    for ci in [2, 3, 4]:
        _style(ws2.cell(gt2, ci, ''), bold=True, bg=_YELLOW)

    tu_col = get_column_letter(len(S_HDRS))
    c_gu = ws2.cell(gt2, 5)
    c_gu.value         = f'=SUM({tu_col}2:{tu_col}{last_po_row})'
    c_gu.number_format = '#,##0'
    c_gu.fill          = _fill(_YELLOW)
    c_gu.border        = _border()
    c_gu.alignment     = _align('right')
    c_gu.font          = Font(name='Arial', bold=True, size=10)

    c_gp = ws2.cell(gt2, 6)
    c_gp.value         = f'=SUM(F{rl_hdr+1}:F{last_rl_row})'
    c_gp.number_format = '#,##0'
    c_gp.fill          = _fill(_YELLOW)
    c_gp.border        = _border()
    c_gp.alignment     = _align('right')
    c_gp.font          = Font(name='Arial', bold=True, size=10)

    _autofit(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Streamlit section
# ---------------------------------------------------------------------------

def show_kl_upload_section(files=None) -> None:
    """Render the KL PO section inside the GIII Upload tab.

    ``files`` — pre-routed UploadedFiles from the combined "Other PO types"
    uploader (auto-detection); ``None`` renders this section's own uploader.
    """

    # No divider/header here — this renders inside a labeled expander on the
    # GIII New Contracts tab, which already names the section.
    st.caption(t(
        "Upload KL-format purchase order PDFs directly. "
        "The system parses PO fields, MSRP details, HTS codes and produces "
        "a formatted Excel workbook ready for download."
    ))

    if files is None:
        uploaded_pdfs = st.file_uploader(
            t("Upload KL PO PDF files"),
            type=["pdf"],
            accept_multiple_files=True,
            label_visibility="collapsed",
            key="kl_uploader",
        )
        if not uploaded_pdfs:
            st.info(t("Upload one or more KL PO PDF files to get started."))
            return
    else:
        uploaded_pdfs = files
        if not uploaded_pdfs:
            return

    st.caption(f"{len(uploaded_pdfs)} " + t("file(s) selected"))

    sig = files_signature(uploaded_pdfs)

    if SK.GIII_KL_RESULTS not in st.session_state:
        st.session_state[SK.GIII_KL_RESULTS] = None

    if st.button(t("▶  Extract KL POs"), type="primary", width="stretch", key="run_kl"):
        st.session_state[SK.GIII_KL_RESULTS] = None
        with st.spinner(t("Parsing KL PO PDFs…")):
            results = _parse_kl_pdfs(uploaded_pdfs)

        if not results:
            st.error(t("No POs could be parsed from the uploaded files."))
            return

        st.session_state[SK.GIII_KL_RESULTS] = results
        st.session_state[SK.GIII_KL_SIG]     = sig
        persist_fax_pos(results, "kl")

    results = drop_stale_results(SK.GIII_KL_RESULTS, SK.GIII_KL_SIG, sig)
    if not results:
        return

    # ── Summary table ────────────────────────────────────────────────────────
    all_sizes  = {s[0] for po in results for li in po['line_items'] for s in li['sizes']}
    sizes_cols = [s for s in _SIZE_ORDER if s in all_sizes]

    rows = []
    total_col = _th('Total')
    for po in results:
        for li in po['line_items']:
            sz_map    = {s[0]: s[1] for s in li['sizes']}
            row_total = sum(s[1] for s in li['sizes'])
            fob_disp  = po['fob_price'] if po['fob_price'] in ('?', 'UNCONFIRMED') else f"${po['fob_price']}"
            msrp_disp = f"${po['msrp']}" if po['msrp'] not in ('?', '') else po['msrp']
            row = {
                _th('PO Number'):     po['po_number'],
                _th('Style'):         li['style'],
                _th('Color'):         li['color'],
                _th('ETD'):           po['etd'],
                _th('FOB'):           fob_disp,
                _th('MSRP'):          msrp_disp,
                _th('HTS#'):          po['hts_num'],
                _th('CPO'):           po['cpo'],
                _th('Customer Name'): po['customer_name'],
                _th('Ship To'):       po['ship_to'],
                _th('Hanger Info'):   po['hanger_info'],
                _th('Pack Ratio'):    po['pack_ratio'],
            }
            for sz in sizes_cols:
                row[sz] = sz_map.get(sz, '')
            row[total_col] = row_total
            rows.append(row)

    import pandas as pd
    df = pd.DataFrame(rows)
    st.dataframe(df, hide_index=True, width="stretch")

    grand_total = sum(r[total_col] for r in rows)
    st.caption(
        f"**{len(results)} {t('PO(s)')}** · **{grand_total:,} {t('total units')}** · "
        f"{len(sizes_cols)} {t('size(s)')}: {', '.join(sizes_cols)}"
    )

    # ── Metadata expander ────────────────────────────────────────────────────
    with st.expander(t("PO Metadata"), expanded=False):
        meta_rows = [{
            _th('PO Number'):     po['po_number'],
            _th('Style'):         po['style'],
            _th('Description'):   po['description'],
            _th('Customer Name'): po['customer_name'],
            _th('Ship To'):       po['ship_to'],
            _th('PO Date'):       po['po_date'],
            _th('Ship Date'):     po['ship_date'],
            _th('ETD'):           po['etd'],
            _th('FOB Price'):     po['fob_price'],
            _th('MSRP'):          po['msrp'],
            _th('HTS#'):          po['hts_num'],
            _th('CPO'):           po['cpo'],
            _th('Hanger Info'):   po['hanger_info'],
            _th('Pack Ratio'):    po['pack_ratio'],
            _th('Factory'):       po['factory'],
            _th('Vendor'):        po['vendor'],
        } for po in results]
        st.dataframe(pd.DataFrame(meta_rows), hide_index=True, width="stretch")

    # ── Download ─────────────────────────────────────────────────────────────
    with st.spinner(t("Building Excel…")):
        xlsx_bytes = _build_kl_excel(results)

    first_po = results[0]['po_number'] if results else 'KL_POs'
    prefix   = re.sub(r'\d+R$', '', first_po)
    fname    = f"{prefix}_KL_POs.xlsx" if prefix else "KL_POs.xlsx"

    st.download_button(
        label=t("⬇️ Download Excel"),
        data=xlsx_bytes,
        file_name=fname,
        mime=_XLSX_MIME,
        type="primary",
        width="stretch",
        key="kl_dl_xlsx",
    )
