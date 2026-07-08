"""MSG / Vendor Fax PO section for the GIII Upload tab.

Accepts Outlook .msg email files, extracts the embedded PDF attachment,
parses PO fields (PO number, style, color, sizes, units, UPC, FOB price),
and produces a downloadable Excel workbook with a PO Detail sheet and a
Summary sheet.

The fax-copy PDFs use a doubled-character font rendering for headers
(e.g. "CCSSKKHHHHAA000011RR" → "CSKHHA001R").  Table rows (line items)
are rendered normally.
"""
from __future__ import annotations

import io
import os
import re
import tempfile

import streamlit as st

from ui.giii._shared import _XLSX_MIME, _undouble, _SIZE_CODES, _FIRST_RE, _CONT_RE, files_signature
from ui.i18n import t
from ui.session_keys import SK
from ui.shared import _th

# (parser helpers are imported from _shared)


def _parse_pdf_bytes(pdf_bytes: bytes) -> dict:
    """Extract PO fields from raw PDF bytes. Returns a PO dict."""
    import pdfplumber

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        all_text  = '\n'.join(p.extract_text() or '' for p in pdf.pages)
        text_lines = all_text.split('\n')

    def grep(pat: str):
        for l in text_lines:
            m = re.search(pat, l)
            if m:
                return m
        return None

    m = grep(r'PO NUMBER\s+(\S+)')
    po_num  = _undouble(m.group(1)) if m else '?'
    m = grep(r'S T Y L E #\s+(\S+)')
    style   = _undouble(m.group(1)) if m else '?'
    m = grep(r'PO DATE\s+([\d/]+)')
    po_date = _undouble(m.group(1)) if m else '?'
    m = grep(r'P R T\s+([\d/]+)')
    ship    = _undouble(m.group(1)) if m else '?'
    etd_m   = grep(r'([\d/]+)\s+EETTDD')
    etd     = _undouble(etd_m.group(1)) if etd_m else '?'

    vline  = text_lines[3] if len(text_lines) > 3 else ''
    vendor = _undouble(re.sub(r'\s*S T Y L E #.*', '', vline).strip())

    fact_m  = grep(r'F A C T O R Y\s+(.+?)\s+INCO')
    factory = _undouble(fact_m.group(1)).strip() if fact_m else '?'

    fob = '?'
    for l in text_lines:
        if re.search(r'FFOOBB::\s*UUNNCCOONNFFIIRRMMEEDD', l):
            fob = 'UNCONFIRMED'
            break
        if re.search(r'FFOOBB', l):
            m = re.search(r'\$\$([\d.]+)', l)
            if m:
                fob = _undouble(m.group(1))
                break

    desc_m      = grep(r'DESCRIPTION\s+(.+)')
    description = _undouble(desc_m.group(1)).strip() if desc_m else '?'

    # Customer name: text after PO NUMBER + PO# on line 4
    cust_m        = grep(r'PO NUMBER\s+\S+\s+(.+)')
    customer_name = _undouble(cust_m.group(1).strip()) if cust_m else '?'

    # Ship To: assembled from lines 5-7
    ship_street   = _undouble(text_lines[5]).strip() if len(text_lines) > 5 else ''
    ship_type_raw = re.sub(r'\s*C N T R Y.*', '', text_lines[6] if len(text_lines) > 6 else '').strip()
    ship_type     = _undouble(ship_type_raw)
    city_m        = grep(r'PO DATE\s+[\d/]+\s+(.+?)\s+PPRRIICCEE')
    ship_city     = _undouble(city_m.group(1).strip()) if city_m else ''
    ship_to       = ' / '.join(p for p in [customer_name, ship_street, ship_type, ship_city] if p and p != '?')

    # Hanger info: line containing "HHAANNGGEERR"
    hanger_line  = next((l for l in text_lines if 'HHAANNGGEERR' in l), None)
    hanger_info  = _undouble(hanger_line).strip() if hanger_line else '?'

    # Pack ratio: e.g. "1-2-2-1" from "(1-2-2-1)" in hanger line
    pack_ratio = '?'
    if hanger_info != '?':
        pr_m = re.search(r'\(([\d-]+)\)', hanger_info)
        if pr_m:
            pack_ratio = pr_m.group(1)

    # HTS#: tariff code like 6106.20.2010 (may be doubled in header, normal in
    # data row).  Only undouble when the match is actually fax-doubled — the
    # doubled form always carries doubled dots ("..").  Unconditional
    # undoubling corrupted normal-font codes with legitimately repeated
    # digits: 6110.20.2079 → 610.20.2079.
    hts_m   = grep(r'(\d{4,8}\.+\d{2,4}\.+\d{4,8})')
    if hts_m:
        hts_raw = hts_m.group(1)
        hts_num = _undouble(hts_raw) if '..' in hts_raw else hts_raw
    else:
        hts_num = '?'

    # CPO: e.g. "CPO: TBD" or doubled "CCPPOO:: TTBBDD"
    cpo_m = grep(r'C+P+O+:+\s*(\S+)')
    cpo   = _undouble(cpo_m.group(1)) if cpo_m else '?'

    # Parse table rows (NOT doubled)
    line_items: list[dict] = []
    in_table   = False
    dash_count = 0
    cur_item: dict | None = None

    for l in text_lines:
        if re.match(r'-{30,}', l):
            dash_count += 1
            if dash_count == 4:
                in_table = True
            continue
        if not in_table:
            continue
        if re.match(r'TOTAL\s+', l) or 'FLAT PACK' in l or 'PACK TTL' in l:
            continue

        m = _FIRST_RE.match(l)
        if m:
            cur_item = {
                'ln':    m.group(1),
                'style': m.group(2),
                'color': m.group(3),
                'sizes': [(m.group(4), int(m.group(5)), m.group(6), float(m.group(7)))],
            }
            line_items.append(cur_item)
            continue

        m2 = _CONT_RE.match(l)
        if m2 and cur_item:
            cur_item['sizes'].append((m2.group(1), int(m2.group(2)), m2.group(3), None))

    return dict(
        po_number=po_num, style=style, po_date=po_date, ship_date=ship,
        etd=etd, vendor=vendor, factory=factory, fob_price=fob,
        description=description, line_items=line_items,
        customer_name=customer_name, ship_to=ship_to,
        hanger_info=hanger_info, pack_ratio=pack_ratio,
        hts_num=hts_num, cpo=cpo, msrp='?',
    )


def _extract_and_parse_msgs(msg_files) -> list[dict]:
    """Accept a list of Streamlit UploadedFile (.msg), return parsed PO dicts."""
    import extract_msg

    results: list[dict] = []
    with tempfile.TemporaryDirectory() as tmp:
        for uf in msg_files:
            msg_path = os.path.join(tmp, uf.name)
            with open(msg_path, 'wb') as f:
                f.write(uf.getbuffer())

            try:
                msg = extract_msg.openMsg(msg_path)
            except Exception as exc:
                st.warning(f"{t('Could not open')} {uf.name}: {exc}")
                continue

            pdf_data: bytes | None = None
            for att in msg.attachments:
                fname = (att.longFilename or att.shortFilename or '').lower()
                if fname.endswith('.pdf'):
                    pdf_data = att.data
                    break

            if pdf_data is None:
                st.warning(f"{t('No PDF attachment in')} {uf.name} — {t('skipped.')}")
                continue

            try:
                po = _parse_pdf_bytes(pdf_data)
            except Exception as exc:
                st.warning(f"{t('Parse error in')} {uf.name}: {exc}")
                continue

            # Derive PO number from email subject as fallback
            if po['po_number'] == '?':
                subj = msg.subject or ''
                nums = [p for p in subj.split() if re.match(r'CSK[A-Z0-9]+', p)]
                if nums:
                    po['po_number'] = nums[0]

            po['source_file'] = uf.name
            results.append(po)

    return results


# ---------------------------------------------------------------------------
# Excel builder (in-memory)
# ---------------------------------------------------------------------------

_SIZE_ORDER = ['XS', 'S', 'M', 'L', 'XL', 'XXL', '1X', '2X', '3X']

# Colours
_NAVY    = 'FF1F3864'
_WHITE   = 'FFFFFFFF'
_YELLOW  = 'FFFFF2CC'
_ORANGE  = 'FFF4B942'
_LT_BLUE = 'FFDEEAF1'
_GREY    = 'FFD9D9D9'
_BLACK   = 'FF000000'


def _build_excel_bytes(results: list[dict]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    all_sizes  = {s[0] for po in results for li in po['line_items'] for s in li['sizes']}
    sizes_cols = [s for s in _SIZE_ORDER if s in all_sizes]

    def _side():
        return Side(style='thin', color='FF000000')

    def _border():
        s = _side()
        return Border(left=s, right=s, top=s, bottom=s)

    def _hdr_font():
        return Font(name='Arial', bold=True, color=_WHITE, size=10)

    def _hdr_fill():
        return PatternFill('solid', fgColor=_NAVY)

    def _hdr_align():
        return Alignment(horizontal='center', vertical='center', wrap_text=True)

    def _body_font(bold=False):
        return Font(name='Arial', bold=bold, size=10)

    def _fill(colour):
        return PatternFill('solid', fgColor=colour) if colour else PatternFill()

    def _align(h='left'):
        return Alignment(horizontal=h, vertical='center')

    def _style(cell, bold=False, bg=None, align='left', num_fmt=None):
        cell.font      = _body_font(bold)
        cell.fill      = _fill(bg)
        cell.alignment = _align(align)
        cell.border    = _border()
        if num_fmt:
            cell.number_format = num_fmt

    def _hdr(cell, value):
        cell.value     = value
        cell.font      = _hdr_font()
        cell.fill      = _hdr_fill()
        cell.alignment = _hdr_align()
        cell.border    = _border()

    def _autofit(ws, mn=8, mx=50):
        for col in ws.columns:
            ltr = get_column_letter(col[0].column)
            w   = max((len(str(c.value or '')) for c in col), default=0)
            ws.column_dimensions[ltr].width = min(max(w + 2, mn), mx)

    wb = Workbook()

    # ── Sheet 1: PO Detail ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title          = 'PO Detail'
    ws1.freeze_panes   = 'A2'
    ws1.row_dimensions[1].height = 28

    D_HDRS = [
        'PO Number', 'Style', 'Color', 'Size', 'Units',
        'UPC', 'Unit Price (FOB)', 'Line Total ($)',
        'ETD', 'PO Date', 'Ship Date',
        'Customer Name', 'Ship To', 'Hanger Info', 'Pack Ratio',
        'HTS#', 'CPO',
        'Description', 'Factory', 'Vendor',
    ]
    for ci, h in enumerate(D_HDRS, 1):
        _hdr(ws1.cell(1, ci), h)

    dr = 2  # detail row counter
    for po in results:
        fob_num: float | None = None
        if po['fob_price'] not in ('?', 'UNCONFIRMED'):
            try:
                fob_num = float(po['fob_price'])
            except ValueError:
                pass

        is_unc = po['fob_price'] == 'UNCONFIRMED'

        for li in po['line_items']:
            first_price = li['sizes'][0][3]
            for sz, units, upc, price in li['sizes']:
                unit_price = price if price is not None else first_price
                bg = _LT_BLUE if dr % 2 == 0 else None

                vals = [
                    po['po_number'], li['style'], li['color'], sz, units,
                    upc, unit_price, None,          # placeholder for formula
                    po['etd'], po['po_date'], po['ship_date'],
                    po['customer_name'], po['ship_to'], po['hanger_info'], po['pack_ratio'],
                    po['hts_num'], po['cpo'],
                    po['description'], po['factory'], po['vendor'],
                ]
                for ci, v in enumerate(vals, 1):
                    c = ws1.cell(dr, ci, v)
                    _style(c, bg=bg, align='right' if ci in (5, 7, 8) else 'left')

                # Line Total formula
                lt = ws1.cell(dr, 8)
                lt.value          = f'=E{dr}*G{dr}'
                lt.number_format  = '$#,##0.00'
                lt.fill           = _fill(_ORANGE if is_unc else bg)
                lt.border         = _border()
                lt.alignment      = _align('right')
                lt.font           = _body_font(is_unc)

                # Unit price formatting
                pc = ws1.cell(dr, 7)
                if is_unc and unit_price is None:
                    pc.value = 'UNCONFIRMED'
                    _style(pc, bold=True, bg=_ORANGE, align='center')
                else:
                    _style(pc, bg=bg, align='right', num_fmt='$#,##0.00')

                _style(ws1.cell(dr, 5), bg=bg, align='right', num_fmt='#,##0')

                dr += 1

    # Grand total row
    gt = dr
    ws1.cell(gt, 1, 'GRAND TOTAL')
    _style(ws1.cell(gt, 1), bold=True, bg=_YELLOW)
    for ci in range(2, len(D_HDRS) + 1):
        _style(ws1.cell(gt, ci), bold=True, bg=_YELLOW)

    e_col = get_column_letter(5)
    h_col = get_column_letter(8)
    c_units = ws1.cell(gt, 5)
    c_units.value          = f'=SUM({e_col}2:{e_col}{gt-1})'
    c_units.number_format  = '#,##0'
    c_units.fill           = _fill(_YELLOW)
    c_units.border         = _border()
    c_units.alignment      = _align('right')
    c_units.font           = _body_font(True)

    c_total = ws1.cell(gt, 8)
    c_total.value          = f'=SUM({h_col}2:{h_col}{gt-1})'
    c_total.number_format  = '$#,##0.00'
    c_total.fill           = _fill(_YELLOW)
    c_total.border         = _border()
    c_total.alignment      = _align('right')
    c_total.font           = _body_font(True)

    _autofit(ws1)

    # ── Sheet 2: Summary ────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Summary')
    ws2.freeze_panes          = 'A2'
    ws2.row_dimensions[1].height = 28

    S_HDRS = ['PO Number', 'Style', 'Description', 'Color', 'ETD', 'FOB Price'] + sizes_cols + ['Total Units']
    for ci, h in enumerate(S_HDRS, 1):
        _hdr(ws2.cell(1, ci), h)

    # PO rows
    po_row = 2
    style_rows: dict[str, list[int]] = {}   # style → [row, ...]

    for po in results:
        is_unc = po['fob_price'] == 'UNCONFIRMED'
        for li in po['line_items']:
            bg = _LT_BLUE if (po_row - 2) % 2 == 0 else None
            sizes_map  = {s[0]: s[1] for s in li['sizes']}
            row_total  = sum(s[1] for s in li['sizes'])
            if is_unc:
                fob_val = po['fob_price']
            elif po['fob_price'] != '?':
                try:
                    fob_val = float(po['fob_price'])
                except ValueError:
                    fob_val = po['fob_price']   # e.g. "12.50." — keep raw string
            else:
                fob_val = ''

            vals = (
                [po['po_number'], li['style'], po['description'], li['color'], po['etd'], fob_val]
                + [sizes_map.get(sz, '') for sz in sizes_cols]
                + [row_total]
            )
            for ci, v in enumerate(vals, 1):
                c = ws2.cell(po_row, ci, v)
                is_price  = ci == 6
                is_size   = 7 <= ci <= 6 + len(sizes_cols)
                is_totcol = ci == len(S_HDRS)
                aln = 'right' if (is_price or is_size or is_totcol) else 'left'
                c_bg = _ORANGE if (is_price and is_unc) else bg
                fmt  = None
                if is_size or is_totcol: fmt = '#,##0'
                if is_price and not is_unc and fob_val != '': fmt = '$#,##0.00'
                _style(c, bold=(is_price and is_unc), bg=c_bg, align=aln, num_fmt=fmt)

            style_rows.setdefault(li['style'], []).append(po_row)
            po_row += 1

    last_po_row = po_row - 1

    # Blank separator + style rollup header
    po_row += 1
    rl_hdr = po_row
    for ci, h in enumerate(['Style', 'Description', 'FOB Price', 'Total Units', 'PO Count'], 1):
        _hdr(ws2.cell(rl_hdr, ci), h)
    po_row += 1

    # Style rollup rows
    seen_styles: dict[str, dict] = {}
    for po in results:
        for li in po['line_items']:
            k = li['style']
            if k not in seen_styles:
                seen_styles[k] = {'desc': po['description'], 'fob': po['fob_price'], 'pos': set()}
            seen_styles[k]['pos'].add(po['po_number'])

    for style_key, info in seen_styles.items():
        rows = style_rows.get(style_key, [])
        tu_col  = get_column_letter(len(S_HDRS))
        formula = f'=SUM({",".join(f"{tu_col}{r}" for r in rows)})' if rows else 0
        is_unc  = info['fob'] == 'UNCONFIRMED'
        if is_unc:
            fob_val = info['fob']
        elif info['fob'] != '?':
            try:
                fob_val = float(info['fob'])
            except ValueError:
                fob_val = info['fob']           # e.g. "12.50." — keep raw string
        else:
            fob_val = ''
        vals    = [style_key, info['desc'], fob_val, formula, len(info['pos'])]

        for ci, v in enumerate(vals, 1):
            c = ws2.cell(po_row, ci, v)
            is_price = ci == 3
            fmt = None
            if ci == 4:            fmt = '#,##0'
            if is_price and not is_unc and fob_val != '': fmt = '$#,##0.00'
            _style(c, bold=True, bg=_GREY, align='right' if ci >= 3 else 'left', num_fmt=fmt)
            if is_price and is_unc:
                _style(c, bold=True, bg=_ORANGE, align='center')

        po_row += 1

    last_rl_row = po_row - 1

    # Grand total (summary sheet)
    gt2 = po_row
    ws2.cell(gt2, 1, 'GRAND TOTAL')
    _style(ws2.cell(gt2, 1), bold=True, bg=_YELLOW)
    for ci in [2, 3]:
        _style(ws2.cell(gt2, ci, ''), bold=True, bg=_YELLOW)

    tu_col = get_column_letter(len(S_HDRS))
    c_gu = ws2.cell(gt2, 4)
    c_gu.value          = f'=SUM({tu_col}2:{tu_col}{last_po_row})'
    c_gu.number_format  = '#,##0'
    c_gu.fill           = _fill(_YELLOW)
    c_gu.border         = _border()
    c_gu.alignment      = _align('right')
    c_gu.font           = _body_font(True)

    c_gp = ws2.cell(gt2, 5)
    c_gp.value          = f'=SUM(E{rl_hdr+1}:E{last_rl_row})'
    c_gp.number_format  = '#,##0'
    c_gp.fill           = _fill(_YELLOW)
    c_gp.border         = _border()
    c_gp.alignment      = _align('right')
    c_gp.font           = _body_font(True)

    _autofit(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Streamlit section
# ---------------------------------------------------------------------------

def show_msg_upload_section() -> None:
    """Render the MSG / Vendor Fax PO upload section inside the GIII Upload tab."""

    # No divider/header here — this renders inside a labeled expander on the
    # GIII New Contracts tab, which already names the section.
    st.caption(t(
        "Upload Outlook **.msg** emails (vendor fax copies from AS400). "
        "The system extracts the embedded PDF, parses PO fields, and produces "
        "a formatted Excel workbook ready for download."
    ))

    uploaded_msgs = st.file_uploader(
        "Upload .msg files",
        type=["msg"],
        accept_multiple_files=True,
        label_visibility="collapsed",
        key="msg_uploader",
    )

    if not uploaded_msgs:
        st.info(t("Upload one or more .msg vendor fax emails to get started."))
        return

    st.caption(f"{len(uploaded_msgs)} " + t("file(s) selected"))

    sig = files_signature(uploaded_msgs)

    if SK.GIII_MSG_RESULTS not in st.session_state:
        st.session_state[SK.GIII_MSG_RESULTS] = None

    if st.button(t("▶  Extract MSG POs"), type="primary", use_container_width=True, key="run_msg"):
        st.session_state[SK.GIII_MSG_RESULTS] = None
        with st.spinner(t("Extracting PDFs and parsing POs…")):
            try:
                import extract_msg  # noqa: F401 — verify library present
            except ImportError:
                st.error(t(
                    "**extract-msg** library not installed. "
                    "Run `pip install extract-msg` and restart the app."
                ))
                return
            results = _extract_and_parse_msgs(uploaded_msgs)

        if not results:
            st.error(t("No POs could be parsed from the uploaded files."))
            return

        st.session_state[SK.GIII_MSG_RESULTS] = results
        st.session_state[SK.GIII_MSG_SIG]     = sig

    results = st.session_state.get(SK.GIII_MSG_RESULTS)
    if results and st.session_state.get(SK.GIII_MSG_SIG) != sig:
        # Upload selection changed since extraction — drop the stale results
        # so a swapped file set can't display/download the previous batch.
        st.session_state[SK.GIII_MSG_RESULTS] = None
        results = None
    if not results:
        return

    # ── Summary table ────────────────────────────────────────────────────────
    all_sizes  = {s[0] for po in results for li in po['line_items'] for s in li['sizes']}
    sizes_cols = [s for s in _SIZE_ORDER if s in all_sizes]

    rows = []
    total_col = _th('Total')
    for po in results:
        for li in po['line_items']:
            sz_map     = {s[0]: s[1] for s in li['sizes']}
            row_total  = sum(s[1] for s in li['sizes'])
            fob_disp   = po['fob_price'] if po['fob_price'] in ('?', 'UNCONFIRMED') else f"${po['fob_price']}"
            row = {
                _th('PO Number'):     po['po_number'],
                _th('Style'):         li['style'],
                _th('Color'):         li['color'],
                _th('ETD'):           po['etd'],
                _th('FOB'):           fob_disp,
                _th('HTS#'):          po['hts_num'],
                _th('CPO'):           po['cpo'],
                _th('Customer Name'): po['customer_name'],
                _th('Ship To'):       po['ship_to'],
                _th('Hanger Info'):   po['hanger_info'],
                _th('Pack Ratio'):    po['pack_ratio'],
            }
            for sz in sizes_cols:
                row[sz] = sz_map.get(sz, '')
            row[total_col] = row_total
            rows.append(row)

    import pandas as pd
    df = pd.DataFrame(rows)
    st.dataframe(df, hide_index=True, use_container_width=True)

    grand_total = sum(r[total_col] for r in rows)
    st.caption(
        f"**{len(results)} {t('PO(s)')}** · **{grand_total:,} {t('total units')}** · "
        f"{len(sizes_cols)} {t('size(s)')}: {', '.join(sizes_cols)}"
    )

    # ── Metadata expander ────────────────────────────────────────────────────
    with st.expander(t("PO Metadata"), expanded=False):
        meta_rows = [{
            _th('PO Number'):     po['po_number'],
            _th('Style'):         po['style'],
            _th('Description'):   po['description'],
            _th('Customer Name'): po['customer_name'],
            _th('Ship To'):       po['ship_to'],
            _th('PO Date'):       po['po_date'],
            _th('Ship Date'):     po['ship_date'],
            _th('ETD'):           po['etd'],
            _th('FOB Price'):     po['fob_price'],
            _th('HTS#'):          po['hts_num'],
            _th('CPO'):           po['cpo'],
            _th('Hanger Info'):   po['hanger_info'],
            _th('Pack Ratio'):    po['pack_ratio'],
            _th('Factory'):       po['factory'],
            _th('Vendor'):        po['vendor'],
        } for po in results]
        st.dataframe(pd.DataFrame(meta_rows), hide_index=True, use_container_width=True)

    # ── Download ─────────────────────────────────────────────────────────────
    with st.spinner(t("Building Excel…")):
        xlsx_bytes = _build_excel_bytes(results)

    # Derive a filename from the first PO number
    first_po = results[0]['po_number'] if results else 'MSG_POs'
    prefix   = re.sub(r'\d+R$', '', first_po)          # e.g. "CSKHHA" from "CSKHHA001R"
    fname    = f"{prefix}_POs.xlsx" if prefix else "MSG_POs.xlsx"

    st.download_button(
        label=t("⬇ Download Excel"),
        data=xlsx_bytes,
        file_name=fname,
        mime=_XLSX_MIME,
        type="primary",
        use_container_width=True,
        key="msg_dl_xlsx",
    )
