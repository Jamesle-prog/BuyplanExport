"""InforNexus PO section for the GIII Upload tab.

Accepts structured InforNexus-format PDF files (clean text, no doubled chars).
Each PDF is a G-III purchase order printed from the Infor Nexus portal.

Key layout differences from KL/MSG fax format:
  - PO# on labelled line "Order Number Issue Date Version"
  - Line items: one summary row + Size/UPC/Qty sub-rows
  - MSRP explicit: "Freight Pay Method Collect MSRP MSRP $69.00"
  - HTS: "HTS Classification 6106.20.2010"
  - ETD / CUST PO in Special Instructions (plain text)
  - No doubled characters
"""
from __future__ import annotations

import io
import re

import streamlit as st

from ui.giii._shared import _XLSX_MIME

_SIZE_ORDER = ['XXS', 'XS', 'S', 'M', 'L', 'XL', 'XXL', '1X', '2X', '3X', 'OSFM']


# ---------------------------------------------------------------------------
# Parser
# ---------------------------------------------------------------------------

def _parse_infornexus_pdf(pdf_bytes: bytes) -> dict:
    """Parse one InforNexus PO PDF. Returns a dict matching the KL/MSG schema."""
    import pdfplumber

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        all_text   = '\n'.join(p.extract_text() or '' for p in pdf.pages)
        text_lines = all_text.split('\n')

    def grep(pat: str):
        for l in text_lines:
            m = re.search(pat, l)
            if m:
                return m
        return None

    # ── PO Number + Issue Date ────────────────────────────────────────────────
    # Line format: "LSKHHN009R 2026-05-12 2026-05-14T00:36:35Z"
    po_m    = grep(r'^(LS\w+R)\s+(\d{4}-\d{2}-\d{2})')
    po_num  = po_m.group(1)            if po_m else '?'
    po_date = po_m.group(2)            if po_m else '?'

    # ── Style / Color / Units from line-item row ──────────────────────────────
    # "1 G5DTN93A JVS JET BLACK - EA EA 8400 3.75 USD ..."
    li_m = grep(r'^(\d+)\s+(\S+)\s+(\S+)\s+(.+?)\s+-\s+EA\s+EA\s+([\d,]+)\s+([\d.]+)')
    ln_num      = li_m.group(1)                        if li_m else '1'
    style       = li_m.group(2)                        if li_m else '?'
    color_code  = li_m.group(3)                        if li_m else '?'
    color_name  = li_m.group(4).strip()                if li_m else '?'
    color       = f"{color_code}/{color_name}"         if li_m else '?'
    fob_price   = li_m.group(6)                        if li_m else '?'

    # ── Ship Date (factory ship date) ────────────────────────────────────────
    sd_m      = grep(r'Factory Ship by Date\s+(\S+)')
    ship_date = sd_m.group(1) if sd_m else '?'

    # ── ETD / CUST PO ────────────────────────────────────────────────────────
    etd_m = grep(r'([\d/]+)\s+ETD')
    etd   = etd_m.group(1) if etd_m else '?'

    cpo_m = grep(r'CUST PO:\s+(\S+)')
    cpo   = re.sub(r'[,;]$', '', cpo_m.group(1)) if cpo_m else '?'

    # ── MSRP ─────────────────────────────────────────────────────────────────
    # Preferred: "MSRP MSRP $69.00" on the detail line
    msrp_m = grep(r'MSRP\s+MSRP\s+\$([\d.]+)')
    if not msrp_m:
        # Fallback: Special Instructions "MSRP: $69"
        msrp_m = grep(r'MSRP:\s+\$([\d.]+)')
    msrp = msrp_m.group(1) if msrp_m else '?'

    # ── HTS ──────────────────────────────────────────────────────────────────
    hts_m   = grep(r'HTS Classification\s+([\d.]+)')
    hts_num = hts_m.group(1) if hts_m else '?'

    # ── Description ──────────────────────────────────────────────────────────
    desc_m      = grep(r'Style Description\s+(.+)')
    description = desc_m.group(1).strip() if desc_m else '?'

    # ── Vendor / Factory (line 5) ────────────────────────────────────────────
    # "G-III APPAREL GROUP LTD  <VENDOR>  <FACTORY CITY>"
    line5    = text_lines[5] if len(text_lines) > 5 else ''
    l5_clean = re.sub(r'^G-III APPAREL GROUP LTD\s*', '', line5).strip()
    # Try to split at known factory city marker
    fact_city_m = re.search(r'(CHANGZHOU\s+\S.*|NINGBO\s+\S.*|SHANGHAI\s+\S.*)', l5_clean)
    if fact_city_m:
        vendor       = l5_clean[:fact_city_m.start()].strip()
        factory_name = fact_city_m.group(1).strip()
    else:
        vendor       = l5_clean
        factory_name = '?'
    # Factory code from label line (e.g. "R HHN R 01423")
    fc_m    = grep(r'\b(\d{5})\b')
    factory = f"{fc_m.group(1)} — {factory_name}" if fc_m else factory_name

    # ── Customer Name ─────────────────────────────────────────────────────────
    cust_m        = grep(r'Customer\s+(.+?)\s+Packaging')
    customer_name = cust_m.group(1).strip() if cust_m else '?'

    # ── Ship To — clean lines appearing after garbled column block ────────────
    ship_to = '?'
    for i, l in enumerate(text_lines):
        if re.match(r'^ROSS STORES', l) and i + 2 < len(text_lines):
            street   = text_lines[i + 1].strip()
            dist_raw = text_lines[i + 2] if i + 2 < len(text_lines) else ''
            city_raw = text_lines[i + 3] if i + 3 < len(text_lines) else ''
            dist     = re.sub(r'^[A-Z]\s+', '', dist_raw).strip()
            city     = re.sub(r'^(?:[A-Z]\s+){1,3}', '', city_raw).strip()
            ship_to  = ' / '.join(p for p in [l.strip(), street, dist, city] if p)
            break
    if ship_to == '?':
        ship_to = customer_name   # fallback

    # ── Hanger Info ───────────────────────────────────────────────────────────
    # "LSKHHN009R KL SEE SPECIAL INSTRUCTIONS 0.75% China"
    hanger_m = None
    if po_num != '?':
        hanger_m = grep(re.escape(po_num) + r'\s+\S+\s+(.+?)\s+[\d.]+%')
    hanger_info = hanger_m.group(1).strip() if hanger_m else '?'

    # ── Pack Ratio ────────────────────────────────────────────────────────────
    pack_ratio = '?'
    for l in text_lines:
        pr_m = re.search(r'\(([\d]+-[\d-]+)\)', l)
        if pr_m:
            pack_ratio = pr_m.group(1)
            break

    # ── Sizes / UPCs / Quantities (structured sub-rows) ───────────────────────
    size_m = grep(r'^Size:\s+(.+)')
    upc_m  = grep(r'^UPC:\s+(.+)')
    qty_m  = grep(r'^Qty:\s+(.+)')

    sizes_list: list[tuple] = []
    if size_m and upc_m and qty_m:
        szs  = [s for s in size_m.group(1).split() if s != '-']
        upcs = [u for u in upc_m.group(1).split()  if u != '-']
        qtys = [q.replace(',', '') for q in qty_m.group(1).split() if q != '-']
        price = float(fob_price) if fob_price != '?' else None
        for sz, upc, qty in zip(szs, upcs, qtys):
            sizes_list.append((sz, int(qty), upc, price))

    line_items = [{
        'ln':    ln_num,
        'style': style,
        'color': color,
        'sizes': sizes_list,
    }] if sizes_list else []

    return dict(
        po_number=po_num, style=style, po_date=po_date, ship_date=ship_date,
        etd=etd, vendor=vendor, factory=factory, fob_price=fob_price,
        description=description, line_items=line_items,
        customer_name=customer_name, ship_to=ship_to,
        hanger_info=hanger_info, pack_ratio=pack_ratio,
        hts_num=hts_num, cpo=cpo, msrp=msrp,
    )


def _parse_infornexus_pdfs(pdf_files) -> list[dict]:
    results: list[dict] = []
    for uf in pdf_files:
        try:
            po = _parse_infornexus_pdf(uf.read())
        except Exception as exc:
            st.warning(f"Parse error in {uf.name}: {exc}")
            continue
        if po['po_number'] == '?':
            po['po_number'] = re.sub(r'[Oo]', '0', uf.name.split('.')[0])
        po['source_file'] = uf.name
        results.append(po)
    return results


# ---------------------------------------------------------------------------
# Comparison Excel builder
# ---------------------------------------------------------------------------

_NAVY   = 'FF1F3864'
_WHITE  = 'FFFFFFFF'
_YELLOW = 'FFFFF2CC'
_GREEN  = 'FFE2EFDA'
_RED    = 'FFFFC7CE'
_GREY   = 'FFD9D9D9'
_LT_BLU = 'FFDEEAF1'
_ORANGE = 'FFF4B942'

_CMP_FIELDS = [
    ('Style',          'style'),
    ('Color',          '_color'),
    ('Customer Name',  'customer_name'),
    ('Ship To',        'ship_to'),
    ('FOB ($)',        'fob_price'),
    ('MSRP ($)',       'msrp'),
    ('Total Units',    '_units'),
    ('S',              '_sz_S'),
    ('M',              '_sz_M'),
    ('L',              '_sz_L'),
    ('XL',             '_sz_XL'),
    ('ETD',            'etd'),
    ('Ship Date',      'ship_date'),
    ('PO Date',        'po_date'),
    ('HTS#',           'hts_num'),
    ('CPO',            'cpo'),
    ('Description',    'description'),
    ('Pack Ratio',     'pack_ratio'),
    ('Hanger',         'hanger_info'),
    ('Factory',        'factory'),
    ('Vendor',         'vendor'),
]


def _po_flat(po: dict) -> dict:
    """Flatten a PO dict for comparison (add computed size/unit fields)."""
    sizes_map: dict[str, int] = {}
    total = 0
    color = '?'
    for li in po['line_items']:
        color = li['color']
        for sz, qty, upc, _ in li['sizes']:
            sizes_map[sz] = sizes_map.get(sz, 0) + qty
            total += qty
    d = dict(po)
    d['_color'] = color
    d['_units'] = str(total)
    for s in ['S', 'M', 'L', 'XL', 'XXS', 'XS', 'XXL']:
        d[f'_sz_{s}'] = str(sizes_map.get(s, '')) if sizes_map.get(s) else ''
    return d


def build_comparison_excel(kl_results: list[dict], in_results: list[dict]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    def _side():
        return Side(style='thin', color='FF000000')

    def _border():
        s = _side()
        return Border(left=s, right=s, top=s, bottom=s)

    def _hdr(cell, value, bg=_NAVY):
        cell.value     = value
        cell.font      = Font(name='Arial', bold=True, color=_WHITE, size=10)
        cell.fill      = PatternFill('solid', fgColor=bg)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = _border()

    def _fill(colour):
        return PatternFill('solid', fgColor=colour) if colour else PatternFill()

    def _style(cell, bold=False, bg=None, align='left', num_fmt=None):
        cell.font      = Font(name='Arial', bold=bold, size=10)
        cell.fill      = _fill(bg)
        cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
        cell.border    = _border()
        if num_fmt:
            cell.number_format = num_fmt

    def _autofit(ws, mn=8, mx=60):
        for col in ws.columns:
            ltr = get_column_letter(col[0].column)
            w   = max((len(str(c.value or '')) for c in col), default=0)
            ws.column_dimensions[ltr].width = min(max(w + 2, mn), mx)

    # Index by PO number
    kl_by_po = {po['po_number']: _po_flat(po) for po in kl_results}
    in_by_po = {po['po_number']: _po_flat(po) for po in in_results}
    all_pos  = sorted(set(kl_by_po) | set(in_by_po))

    wb = Workbook()

    # ── Sheet 1: Side-by-side comparison ────────────────────────────────────
    ws = wb.active
    ws.title          = 'Comparison'
    ws.freeze_panes   = 'C2'
    ws.row_dimensions[1].height = 30

    # Header row
    _hdr(ws.cell(1, 1), 'PO Number')
    _hdr(ws.cell(1, 2), 'Field')
    _hdr(ws.cell(1, 3), 'KL (Fax)',       bg='FF2E75B6')
    _hdr(ws.cell(1, 4), 'InforNexus',     bg='FF375623')
    _hdr(ws.cell(1, 5), 'Match?')

    row = 2
    for po_num in all_pos:
        kl = kl_by_po.get(po_num)
        nx = in_by_po.get(po_num)

        # Group header for the PO
        for ci in range(1, 6):
            ws.cell(row, ci).fill   = _fill(_GREY)
            ws.cell(row, ci).border = _border()
            ws.cell(row, ci).font   = Font(name='Arial', bold=True, size=10)
        ws.cell(row, 1, po_num)
        row += 1

        for label, key in _CMP_FIELDS:
            kl_val = kl.get(key, '—') if kl else '—'
            nx_val = nx.get(key, '—') if nx else '—'
            kl_str = str(kl_val) if kl_val is not None else '—'
            nx_str = str(nx_val) if nx_val is not None else '—'

            match = kl_str.strip().lower() == nx_str.strip().lower()
            match_txt = '✓' if match else '✗'
            match_bg  = _GREEN if match else _RED

            _style(ws.cell(row, 1), bg=None, align='left')   # PO# blank (grouped above)
            _style(ws.cell(row, 2, label), align='left')
            _style(ws.cell(row, 3, kl_str), bg=_LT_BLU, align='left')
            _style(ws.cell(row, 4, nx_str), bg='FFEDFFE0', align='left')
            _style(ws.cell(row, 5, match_txt), bold=True, bg=match_bg, align='center')

            row += 1

        # Blank spacer
        row += 1

    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 32
    ws.column_dimensions['D'].width = 32
    ws.column_dimensions['E'].width = 8

    # ── Sheet 2: InforNexus Detail (size-level rows) ─────────────────────────
    ws2 = wb.create_sheet('InforNexus Detail')
    ws2.freeze_panes = 'A2'
    ws2.row_dimensions[1].height = 28

    all_sizes  = {s[0] for po in in_results for li in po['line_items'] for s in li['sizes']}
    sizes_cols = [s for s in _SIZE_ORDER if s in all_sizes]

    D2_HDRS = [
        'PO Number', 'Style', 'Color', 'Size', 'Units', 'UPC',
        'Unit Price (FOB)', 'MSRP',
        'ETD', 'PO Date', 'Ship Date',
        'Customer Name', 'Ship To', 'Hanger Info', 'Pack Ratio',
        'HTS#', 'CPO', 'Description', 'Factory', 'Vendor',
    ]
    for ci, h in enumerate(D2_HDRS, 1):
        _hdr(ws2.cell(1, ci), h)

    dr = 2
    for po in in_results:
        try:
            msrp_val = float(po['msrp'])
        except (ValueError, TypeError):
            msrp_val = po['msrp']
        for li in po['line_items']:
            first_price = li['sizes'][0][3] if li['sizes'] else None
            for sz, units, upc, price in li['sizes']:
                unit_price = price if price is not None else first_price
                bg = _LT_BLU if dr % 2 == 0 else None
                vals = [
                    po['po_number'], li['style'], li['color'], sz, units, upc,
                    unit_price, msrp_val,
                    po['etd'], po['po_date'], po['ship_date'],
                    po['customer_name'], po['ship_to'], po['hanger_info'], po['pack_ratio'],
                    po['hts_num'], po['cpo'], po['description'], po['factory'], po['vendor'],
                ]
                for ci, v in enumerate(vals, 1):
                    c = ws2.cell(dr, ci, v)
                    _style(c, bg=bg, align='right' if ci in (5, 7, 8) else 'left',
                           num_fmt='$#,##0.00' if ci in (7, 8) else (
                               '#,##0' if ci == 5 else None))
                dr += 1

    _autofit(ws2)

    # ── Sheet 3: KL Detail ────────────────────────────────────────────────────
    ws3 = wb.create_sheet('KL Detail')
    ws3.freeze_panes = 'A2'
    ws3.row_dimensions[1].height = 28
    for ci, h in enumerate(D2_HDRS, 1):
        _hdr(ws3.cell(1, ci), h)

    dr = 2
    for po in kl_results:
        try:
            msrp_val = float(po.get('msrp', '?'))
        except (ValueError, TypeError):
            msrp_val = po.get('msrp', '?')
        for li in po['line_items']:
            first_price = li['sizes'][0][3] if li['sizes'] else None
            for sz, units, upc, price in li['sizes']:
                unit_price = price if price is not None else first_price
                bg = _LT_BLU if dr % 2 == 0 else None
                vals = [
                    po['po_number'], li['style'], li['color'], sz, units, upc,
                    unit_price, msrp_val,
                    po['etd'], po['po_date'], po['ship_date'],
                    po['customer_name'], po['ship_to'], po['hanger_info'], po['pack_ratio'],
                    po['hts_num'], po['cpo'], po['description'], po['factory'], po['vendor'],
                ]
                for ci, v in enumerate(vals, 1):
                    c = ws3.cell(dr, ci, v)
                    _style(c, bg=bg, align='right' if ci in (5, 7, 8) else 'left',
                           num_fmt='$#,##0.00' if ci in (7, 8) else (
                               '#,##0' if ci == 5 else None))
                dr += 1

    _autofit(ws3)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Streamlit section
# ---------------------------------------------------------------------------

def show_infornexus_upload_section() -> None:
    """InforNexus PO upload + optional comparison against KL PDFs."""

    st.divider()
    st.markdown("#### 🗂 InforNexus POs")
    st.caption(
        "Upload InforNexus-format PO PDFs. "
        "Optionally also upload the matching KL fax PDFs to generate a side-by-side comparison."
    )

    col_in, col_kl = st.columns(2)
    with col_in:
        st.markdown("**InforNexus PDFs**")
        in_files = st.file_uploader(
            "InforNexus PDFs", type=["pdf"], accept_multiple_files=True,
            label_visibility="collapsed", key="in_uploader",
        )
        if in_files:
            st.caption(f"{len(in_files)} file(s)")

    with col_kl:
        st.markdown("**KL Fax PDFs** *(for comparison)*")
        kl_files = st.file_uploader(
            "KL PDFs", type=["pdf"], accept_multiple_files=True,
            label_visibility="collapsed", key="in_kl_uploader",
        )
        if kl_files:
            st.caption(f"{len(kl_files)} file(s)")

    if not in_files:
        st.info("Upload InforNexus PDFs to get started.")
        return

    if "in_results" not in st.session_state:
        st.session_state.in_results = None
    if "in_kl_results" not in st.session_state:
        st.session_state.in_kl_results = None

    if st.button("▶  Extract & Compare", type="primary", use_container_width=True, key="run_in"):
        st.session_state.in_results    = None
        st.session_state.in_kl_results = None

        with st.spinner("Parsing InforNexus PDFs…"):
            in_results = _parse_infornexus_pdfs(in_files)

        kl_results: list[dict] = []
        if kl_files:
            from ui.giii.kl_extraction import _parse_kl_pdf
            with st.spinner("Parsing KL fax PDFs…"):
                for uf in kl_files:
                    try:
                        po = _parse_kl_pdf(uf.read())
                    except Exception as exc:
                        st.warning(f"KL parse error {uf.name}: {exc}")
                        continue
                    if po['po_number'] == '?':
                        po['po_number'] = uf.name.split('-')[0]
                    po['source_file'] = uf.name
                    kl_results.append(po)

        if not in_results:
            st.error("No POs could be parsed.")
            return

        st.session_state.in_results    = in_results
        st.session_state.in_kl_results = kl_results

    in_results = st.session_state.get("in_results")
    if not in_results:
        return

    kl_results = st.session_state.get("in_kl_results") or []

    # ── Summary table ────────────────────────────────────────────────────────
    rows = []
    for po in in_results:
        total = sum(s[1] for li in po['line_items'] for s in li['sizes'])
        rows.append({
            'PO Number': po['po_number'],
            'Style':     po['style'],
            'Color':     po['line_items'][0]['color'] if po['line_items'] else '?',
            'ETD':       po['etd'],
            'FOB':       f"${po['fob_price']}",
            'MSRP':      f"${po['msrp']}" if po['msrp'] != '?' else '?',
            'HTS#':      po['hts_num'],
            'CPO':       po['cpo'],
            'Units':     total,
            'Pack':      po['pack_ratio'],
        })

    import pandas as pd
    st.dataframe(pd.DataFrame(rows), hide_index=True, use_container_width=True)
    total_units = sum(r['Units'] for r in rows)
    st.caption(f"**{len(in_results)} PO(s)** · **{total_units:,} total units**")

    # ── Download ─────────────────────────────────────────────────────────────
    with st.spinner("Building Excel…"):
        xlsx = build_comparison_excel(kl_results, in_results)

    label = "⬇ Download Comparison Excel" if kl_results else "⬇ Download InforNexus Excel"
    first_po = in_results[0]['po_number'] if in_results else 'IN_POs'
    prefix   = re.sub(r'\d+R$', '', first_po)
    fname    = f"{prefix}_comparison.xlsx" if kl_results else f"{prefix}_IN_POs.xlsx"

    st.download_button(
        label=label, data=xlsx, file_name=fname,
        mime=_XLSX_MIME, type="primary", use_container_width=True, key="in_dl",
    )
