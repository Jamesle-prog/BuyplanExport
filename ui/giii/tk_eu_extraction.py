"""TK EU (Kostroma / TJX UK) PO section for the GIII Upload tab.

Accepts Outlook .msg vendor fax emails for TK EU purchase orders.
Layout differences from CSKHHA/Ross MSG format:
  - Brand header: "K O S T R O M A , LTD" (not G-III)
  - PO number suffix: U (not R)
  - Customer: TJX UK PROCESSING CENTRE / C/O APL LOGISTICS / LONDON, UK
  - FOB: "NOT CONFIRMED" in header — actual unit price extracted from table row
  - No MSRP, no HTS details page, no CPO
  - No pack ratio
  - Hanger line mixed-case doubling: "HHaannggeerr,..."
  - Ship To address: C/O APL LOGISTICS line 5, city after PO DATE in line 7
"""
from __future__ import annotations

import io
import re

import streamlit as st

from ui.giii._shared import (
    _XLSX_MIME, _undouble, _SIZE_CODES, _FIRST_RE, _CONT_RE, files_signature,
    FAX_SIZE_ORDER, XL_NAVY, XL_WHITE, XL_YELLOW, XL_GREY, XL_LTBLUE, XL_GREEN,
    drop_stale_results, iter_pdf_payloads, make_excel_style_kit,
    persist_fax_pos,
)
from ui.i18n import t
from ui.session_keys import SK
from ui.shared import _th

# (parser helpers are imported from _shared)


def _parse_tk_eu_pdf(pdf_bytes: bytes) -> dict:
    """Parse a TK EU Kostroma PO PDF. Returns a PO dict."""
    import pdfplumber

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        all_text   = '\n'.join(p.extract_text() or '' for p in pdf.pages)
        text_lines = all_text.split('\n')

    def grep(pat: str):
        for l in text_lines:
            m = re.search(pat, l)
            if m:
                return m
        return None

    # ── Standard header fields ────────────────────────────────────────────────
    m = grep(r'PO NUMBER\s+(\S+)')
    po_number  = _undouble(m.group(1)) if m else '?'
    m = grep(r'S T Y L E #\s+(\S+)')
    style   = _undouble(m.group(1)) if m else '?'
    m = grep(r'PO DATE\s+([\d/]+)')
    po_date = _undouble(m.group(1)) if m else '?'
    m = grep(r'P R T\s+([\d/]+)')
    ship    = _undouble(m.group(1)) if m else '?'

    etd_m = grep(r'([\d/]+)\s+EETTDD')
    etd   = _undouble(etd_m.group(1)) if etd_m else '?'

    # Vendor: line 3, before "S T Y L E #"
    vline  = text_lines[3] if len(text_lines) > 3 else ''
    vendor = _undouble(re.sub(r'\s*S T Y L E #.*', '', vline).strip())

    fact_m  = grep(r'F A C T O R Y\s+(.+?)\s+INCO')
    factory = _undouble(fact_m.group(1)).strip() if fact_m else '?'

    # ── Description ───────────────────────────────────────────────────────────
    desc_m      = grep(r'DESCRIPTION\s+(.+)')
    description = _undouble(desc_m.group(1)).strip() if desc_m else '?'

    # ── Customer name: after PO NUMBER + PO# ─────────────────────────────────
    cust_m        = grep(r'PO NUMBER\s+\S+\s+(.+)')
    customer_name = _undouble(cust_m.group(1).strip()) if cust_m else '?'

    # ── Ship To: TJX UK / C/O APL LOGISTICS / LONDON, UK E152GW ─────────────
    # Line 5: C/O APL LOGISTICS (doubled)
    # Line 7: "PO DATE 5/14/26 LONDON,UK E152GW PRICE TYPE..."
    ship_forwarding = _undouble(text_lines[5]).strip() if len(text_lines) > 5 else ''
    city_m = grep(r'PO DATE\s+[\d/]+\s+(.+?)\s+PPRRIICCEE')
    ship_city = _undouble(city_m.group(1).strip()) if city_m else ''
    ship_to = ' / '.join(p for p in [customer_name, ship_forwarding, ship_city] if p and p != '?')

    # ── FOB: header says "NOT CONFIRMED" — extract actual price from table row
    fob = 'NOT CONFIRMED'

    # ── MSRP / HTS / CPO — not present in TK EU format ───────────────────────
    msrp    = '?'
    hts_num = '?'
    cpo     = '?'

    # ── Hanger: "HHaannggeerr,..." mixed-case doubling ─────────────────────────
    hanger_line = next(
        (l for l in text_lines if re.search(r'[Hh]{2}[Aa]{2}[Nn]{2}[Gg]{2}[Ee]{2}[Rr]{2}', l)),
        None,
    )
    hanger_info = _undouble(hanger_line).strip() if hanger_line else '?'

    # ── Pack ratio: not used in TK EU ────────────────────────────────────────
    pack_ratio = '?'

    # ── Table rows ─────────────────────────────────────────────────────────
    line_items: list[dict] = []
    in_table   = False
    dash_count = 0
    cur_item: dict | None = None

    for l in text_lines:
        if re.match(r'-{30,}', l):
            dash_count += 1
            if dash_count == 4:
                in_table = True
            continue
        if not in_table:
            continue
        if re.match(r'TOTAL\s+', l) or 'FLAT PACK' in l or 'PACK TTL' in l:
            continue

        m = _FIRST_RE.match(l)
        if m:
            row_fob = float(m.group(7))
            # Use first table-row price as the confirmed FOB
            if fob == 'NOT CONFIRMED':
                fob = str(row_fob)
            cur_item = {
                'ln':    m.group(1),
                'style': m.group(2),
                'color': m.group(3),
                'sizes': [(m.group(4), int(m.group(5)), m.group(6), row_fob)],
            }
            line_items.append(cur_item)
            continue

        m2 = _CONT_RE.match(l)
        if m2 and cur_item:
            cur_item['sizes'].append((m2.group(1), int(m2.group(2)), m2.group(3), None))

    return dict(
        po_number=po_number, style=style, po_date=po_date, ship_date=ship,
        etd=etd, vendor=vendor, factory=factory, fob_price=fob,
        description=description, line_items=line_items,
        customer_name=customer_name, ship_to=ship_to,
        hanger_info=hanger_info, pack_ratio=pack_ratio,
        hts_num=hts_num, cpo=cpo, msrp=msrp,
    )


def _extract_and_parse_tk_eu(msg_files) -> list[dict]:
    """Accept Streamlit UploadedFiles — .msg (embedded-PDF fax emails) or the
    same fax PDFs uploaded directly — and return parsed TK EU PO dicts."""
    results: list[dict] = []
    for name, pdf_data, _subject in iter_pdf_payloads(msg_files):
        try:
            po = _parse_tk_eu_pdf(pdf_data)
        except Exception as exc:
            st.warning(t("Parse error in {name}: {exc}").format(name=name, exc=exc))
            continue

        # Derive PO number from filename if not found
        if po['po_number'] == '?':
            m = re.search(r'(DU\w+U)', name)
            if m:
                po['po_number'] = m.group(1)

        po['source_file'] = name
        results.append(po)

    return results


# ---------------------------------------------------------------------------
# Excel builder
# ---------------------------------------------------------------------------

_SIZE_ORDER = FAX_SIZE_ORDER

_NAVY    = XL_NAVY
_WHITE   = XL_WHITE
_YELLOW  = XL_YELLOW
_LT_BLUE = XL_LTBLUE
_GREY    = XL_GREY
_GREEN   = XL_GREEN
_TEAL    = 'FF1F6B75'  # TK EU accent colour


def _build_tk_eu_excel(results: list[dict]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    all_sizes  = {s[0] for po in results for li in po['line_items'] for s in li['sizes']}
    sizes_cols = [s for s in _SIZE_ORDER if s in all_sizes]

    _kit = make_excel_style_kit(hdr_bg=_TEAL)
    _border, _fill, _align = _kit.border, _kit.fill, _kit.align
    _style, _hdr, _autofit = _kit.style, _kit.hdr, _kit.autofit

    wb = Workbook()

    # ── Sheet 1: PO Detail ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title        = 'PO Detail'
    ws1.freeze_panes = 'A2'
    ws1.row_dimensions[1].height = 28

    D_HDRS = [
        'PO Number', 'Style', 'Color', 'Size', 'Units',
        'UPC', 'Unit Price (FOB)', 'Line Total ($)',
        'ETD', 'PO Date', 'Ship Date',
        'Customer Name', 'Ship To', 'Hanger Info',
        'Description', 'Factory', 'Vendor',
    ]
    for ci, h in enumerate(D_HDRS, 1):
        _hdr(ws1.cell(1, ci), h)

    dr = 2
    for po in results:
        for li in po['line_items']:
            first_price = li['sizes'][0][3]
            for sz, units, upc, price in li['sizes']:
                unit_price = price if price is not None else first_price
                bg = _LT_BLUE if dr % 2 == 0 else None

                vals = [
                    po['po_number'], li['style'], li['color'], sz, units,
                    upc, unit_price, None,
                    po['etd'], po['po_date'], po['ship_date'],
                    po['customer_name'], po['ship_to'], po['hanger_info'],
                    po['description'], po['factory'], po['vendor'],
                ]
                for ci, v in enumerate(vals, 1):
                    c = ws1.cell(dr, ci, v)
                    _style(c, bg=bg, align='right' if ci in (5, 7, 8) else 'left')

                # Line Total formula
                lt = ws1.cell(dr, 8)
                lt.value         = f'=E{dr}*G{dr}'
                lt.number_format = '$#,##0.00'
                lt.fill          = _fill(bg)
                lt.border        = _border()
                lt.alignment     = _align('right')
                lt.font          = Font(name='Arial', size=10)

                _style(ws1.cell(dr, 7), bg=bg, align='right', num_fmt='$#,##0.00')
                _style(ws1.cell(dr, 5), bg=bg, align='right', num_fmt='#,##0')
                dr += 1

    # Grand total
    gt = dr
    ws1.cell(gt, 1, 'GRAND TOTAL')
    _style(ws1.cell(gt, 1), bold=True, bg=_YELLOW)
    for ci in range(2, len(D_HDRS) + 1):
        _style(ws1.cell(gt, ci), bold=True, bg=_YELLOW)

    c_units = ws1.cell(gt, 5)
    c_units.value = f'=SUM(E2:E{gt-1})'
    c_units.number_format = '#,##0'
    c_units.fill = _fill(_YELLOW); c_units.border = _border()
    c_units.alignment = _align('right')
    c_units.font = Font(name='Arial', bold=True, size=10)

    c_total = ws1.cell(gt, 8)
    c_total.value = f'=SUM(H2:H{gt-1})'
    c_total.number_format = '$#,##0.00'
    c_total.fill = _fill(_YELLOW); c_total.border = _border()
    c_total.alignment = _align('right')
    c_total.font = Font(name='Arial', bold=True, size=10)

    _autofit(ws1)

    # ── Sheet 2: Summary ────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Summary')
    ws2.freeze_panes = 'A2'
    ws2.row_dimensions[1].height = 28

    S_HDRS = ['PO Number', 'Style', 'Description', 'Color',
              'ETD', 'FOB Price'] + sizes_cols + ['Total Units']
    for ci, h in enumerate(S_HDRS, 1):
        _hdr(ws2.cell(1, ci), h)

    po_row = 2
    style_rows: dict[str, list[int]] = {}

    for po in results:
        try:    fob_val = float(po['fob_price'])
        except: fob_val = po['fob_price']

        for li in po['line_items']:
            bg        = _LT_BLUE if (po_row - 2) % 2 == 0 else None
            sizes_map = {s[0]: s[1] for s in li['sizes']}
            row_total = sum(s[1] for s in li['sizes'])

            vals = (
                [po['po_number'], li['style'], po['description'], li['color'],
                 po['etd'], fob_val]
                + [sizes_map.get(sz, '') for sz in sizes_cols]
                + [row_total]
            )
            for ci, v in enumerate(vals, 1):
                c = ws2.cell(po_row, ci, v)
                is_fob   = ci == 6
                is_size  = 7 <= ci <= 6 + len(sizes_cols)
                is_total = ci == len(S_HDRS)
                fmt = None
                if is_size or is_total: fmt = '#,##0'
                if is_fob and isinstance(fob_val, float): fmt = '$#,##0.00'
                aln = 'right' if (is_fob or is_size or is_total) else 'left'
                _style(c, bg=bg, align=aln, num_fmt=fmt)

            style_rows.setdefault(li['style'], []).append(po_row)
            po_row += 1

    last_po_row = po_row - 1
    po_row += 1  # blank separator

    # Style rollup header
    rl_hdr = po_row
    for ci, h in enumerate(['Style', 'Description', 'FOB Price', 'Total Units', 'PO Count'], 1):
        _hdr(ws2.cell(rl_hdr, ci), h)
    po_row += 1

    seen_styles: dict[str, dict] = {}
    for po in results:
        for li in po['line_items']:
            k = li['style']
            if k not in seen_styles:
                seen_styles[k] = {'desc': po['description'], 'fob': po['fob_price'], 'pos': set()}
            seen_styles[k]['pos'].add(po['po_number'])

    for style_key, info in seen_styles.items():
        rows   = style_rows.get(style_key, [])
        tu_col = get_column_letter(len(S_HDRS))
        formula = f'=SUM({",".join(f"{tu_col}{r}" for r in rows)})' if rows else 0
        try:    fob_val = float(info['fob'])
        except: fob_val = info['fob']

        for ci, v in enumerate([style_key, info['desc'], fob_val, formula, len(info['pos'])], 1):
            c = ws2.cell(po_row, ci, v)
            is_fob = ci == 3
            fmt = None
            if ci == 4: fmt = '#,##0'
            if is_fob and isinstance(fob_val, float): fmt = '$#,##0.00'
            _style(c, bold=True, bg=_GREY,
                   align='right' if ci >= 3 else 'left', num_fmt=fmt)
        po_row += 1

    last_rl_row = po_row - 1

    # Grand total (summary)
    gt2 = po_row
    ws2.cell(gt2, 1, 'GRAND TOTAL')
    _style(ws2.cell(gt2, 1), bold=True, bg=_YELLOW)
    for ci in [2, 3]:
        _style(ws2.cell(gt2, ci, ''), bold=True, bg=_YELLOW)

    tu_col = get_column_letter(len(S_HDRS))
    c_gu = ws2.cell(gt2, 4)
    c_gu.value = f'=SUM({tu_col}2:{tu_col}{last_po_row})'
    c_gu.number_format = '#,##0'
    c_gu.fill = _fill(_YELLOW); c_gu.border = _border()
    c_gu.alignment = _align('right'); c_gu.font = Font(name='Arial', bold=True, size=10)

    c_gp = ws2.cell(gt2, 5)
    c_gp.value = f'=SUM(E{rl_hdr+1}:E{last_rl_row})'
    c_gp.number_format = '#,##0'
    c_gp.fill = _fill(_YELLOW); c_gp.border = _border()
    c_gp.alignment = _align('right'); c_gp.font = Font(name='Arial', bold=True, size=10)

    _autofit(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Streamlit section
# ---------------------------------------------------------------------------

def show_tk_eu_upload_section(files=None) -> None:
    """Render the TK EU PO section inside the GIII Upload tab.

    ``files`` — pre-routed UploadedFiles from the combined "Other PO types"
    uploader (auto-detection); ``None`` renders this section's own uploader.
    """

    # No divider/header here — this renders inside a labeled expander on the
    # GIII New Contracts tab, which already names the section.
    st.caption(t(
        "Upload Outlook **.msg** vendor fax emails for TK EU / Kostroma "
        "purchase orders (TJX UK) — or the fax **PDF** files directly. The "
        "system extracts the PDF, parses PO fields, and produces a formatted Excel."
    ))

    if files is None:
        uploaded_msgs = st.file_uploader(
            t("Upload TK EU .msg or fax PDF files"),
            type=["msg", "pdf"],
            accept_multiple_files=True,
            label_visibility="collapsed",
            key="tk_eu_uploader",
        )
        if not uploaded_msgs:
            st.info(t("Upload TK EU .msg vendor fax emails — or the fax PDFs directly — to get started."))
            return
    else:
        uploaded_msgs = files
        if not uploaded_msgs:
            return

    sig = files_signature(uploaded_msgs)

    st.caption(f"{len(uploaded_msgs)} {t('file(s) selected')}")

    if SK.GIII_TKEU_RESULTS not in st.session_state:
        st.session_state[SK.GIII_TKEU_RESULTS] = None

    if st.button(f"▶  {t('Extract TK EU POs')}", type="primary",
                 width="stretch", key="run_tk_eu"):
        st.session_state[SK.GIII_TKEU_RESULTS] = None
        with st.spinner(t("Extracting PDFs and parsing POs…")):
            if any(uf.name.lower().endswith(".msg") for uf in uploaded_msgs):
                try:
                    import extract_msg  # noqa
                except ImportError:
                    st.error(t("**extract-msg** library not installed. Run `pip install extract-msg`."))
                    return
            results = _extract_and_parse_tk_eu(uploaded_msgs)

        if not results:
            st.error(t("No POs could be parsed."))
            return
        st.session_state[SK.GIII_TKEU_RESULTS] = results
        st.session_state[SK.GIII_TKEU_SIG]     = sig
        persist_fax_pos(results, "tk_eu")

    # Drop stale results when the uploaded file set changed since extraction.
    results = drop_stale_results(SK.GIII_TKEU_RESULTS, SK.GIII_TKEU_SIG, sig)
    if not results:
        return

    # ── Summary table ────────────────────────────────────────────────────────
    all_sizes  = {s[0] for po in results for li in po['line_items'] for s in li['sizes']}
    sizes_cols = [s for s in _SIZE_ORDER if s in all_sizes]

    rows = []
    grand_total = 0
    for po in results:
        for li in po['line_items']:
            sz_map    = {s[0]: s[1] for s in li['sizes']}
            row_total = sum(s[1] for s in li['sizes'])
            fob_disp  = (po['fob_price']
                         if po['fob_price'] in ('?', 'UNCONFIRMED', 'NOT CONFIRMED')
                         else f"${po['fob_price']}")
            row = {
                _th('PO Number'):     po['po_number'],
                _th('Style'):         li['style'],
                _th('Color'):         li['color'],
                _th('ETD'):           po['etd'],
                _th('FOB'):           fob_disp,
                _th('Customer Name'): po['customer_name'],
                _th('Ship To'):       po['ship_to'],
                _th('Hanger Info'):   po['hanger_info'],
            }
            for sz in sizes_cols:
                row[sz] = sz_map.get(sz, '')
            row[_th('Total')] = row_total
            grand_total += row_total
            rows.append(row)

    import pandas as pd
    st.dataframe(pd.DataFrame(rows), hide_index=True, width="stretch")

    st.caption(f"**{len(results)} {t('PO(s)')}** · **{grand_total:,} {t('total units')}**")

    with st.expander(t("PO Metadata"), expanded=False):
        meta_rows = [{
            _th('PO Number'):     po['po_number'],
            _th('Style'):         po['style'],
            _th('Description'):   po['description'],
            _th('Customer Name'): po['customer_name'],
            _th('Ship To'):       po['ship_to'],
            _th('PO Date'):       po['po_date'],
            _th('Ship Date'):     po['ship_date'],
            _th('ETD'):           po['etd'],
            _th('FOB Price'):     po['fob_price'],
            _th('Hanger Info'):   po['hanger_info'],
            _th('Factory'):       po['factory'],
            _th('Vendor'):        po['vendor'],
        } for po in results]
        st.dataframe(pd.DataFrame(meta_rows), hide_index=True, width="stretch")

    with st.spinner(t("Building Excel…")):
        xlsx_bytes = _build_tk_eu_excel(results)

    first_po = results[0]['po_number'] if results else 'TK_EU_POs'
    prefix   = re.sub(r'\d+U$', '', first_po)
    fname    = f"{prefix}_TK_EU_POs.xlsx" if prefix else "TK_EU_POs.xlsx"

    st.download_button(
        label=f"⬇️ {t('Download Excel')}",
        data=xlsx_bytes, file_name=fname,
        mime=_XLSX_MIME, type="primary",
        width="stretch", key="tk_eu_dl",
    )
