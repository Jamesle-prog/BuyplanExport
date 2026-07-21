"""GIII fabric-mapping reference management."""
from __future__ import annotations
import io
import os
import tempfile
import streamlit as st
import pandas as pd
from po_extractor.ui_helpers import (
    detect_fabric_mapping_columns as _detect_fabric_mapping_columns_impl,
    parse_fabric_mapping_rows as _parse_fabric_mapping_rows,
)
from ui.i18n import t
from ui.session_keys import SK
from ui.stores import get_store, get_fabric_master_store
from ui.giii._shared import _BODY_PART_LIST, _XLSX_MIME


def _generate_fabric_mapping_template() -> bytes:
    """
    Shared style-fabric mapping template (GIII and Sky East).

    Layout — one bilingual header row + two example rows:
      A: Style No. / 款式号
      B: Fabric 1 Body Part / 面料1部位  (dropdown)   C: Fabric 1 Code / 面料1编号
      D: Fabric 2 Body Part / 面料2部位  (dropdown)   E: Fabric 2 Code / 面料2编号
      F: Fabric 3 Body Part / 面料3部位  (dropdown)   G: Fabric 3 Code / 面料3编号
      H: Fabric 4 Body Part / 面料4部位  (dropdown)   I: Fabric 4 Code / 面料4编号
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Style Fabric Mapping"

    headers = [
        "Style No.\n款式号",
        "Fabric 1 Body Part\n面料1部位",  "Fabric 1 Code\n面料1编号",
        "Fabric 2 Body Part\n面料2部位",  "Fabric 2 Code\n面料2编号",
        "Fabric 3 Body Part\n面料3部位",  "Fabric 3 Code\n面料3编号",
        "Fabric 4 Body Part\n面料4部位",  "Fabric 4 Code\n面料4编号",
    ]
    col_widths = [20, 24, 18, 24, 18, 24, 18, 24, 18]

    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)

    for ci, (hdr, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=ci, value=hdr)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = width

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    # Example rows  (Body Part first, then Code — within each fabric slot)
    examples = [
        ["S25DDR2036", "Main Body / 大身",   "HHN-JA-01715", "Lining / 里布",        "HHN-MS-01794", "",                   "",             "", ""],
        ["S25JKT1042", "Upper Body / 上身",  "HHN-JA-02301", "Pocket Mesh / 网眼布", "HHN-PO-00891", "Sleeve / 袖子",       "HHN-JA-00712", "", ""],
    ]
    ex_font = Font(color="808080", italic=True, size=10)
    for ri, row in enumerate(examples, start=2):
        for ci, val in enumerate(row, start=1):
            ws.cell(row=ri, column=ci, value=val).font = ex_font

    # Instruction row
    note = ws.cell(
        row=4, column=1,
        value="↑ 请替换示例行 / Replace example rows.  "
              "Body Part 部位 — select from dropdown or leave blank if single fabric.  "
              "Fabric Code = HHN编号 (e.g. HHN-JA-01715).  "
              "Composition 成分 is looked up automatically from the Fabric DB.",
    )
    note.font = Font(color="C00000", size=9, italic=True)
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=9)

    # Dropdown validation for all four Body Part columns (B=2, D=4, F=6, H=8)
    # Excel dropdown formula: quoted comma-separated list (max ~255 chars)
    dv_formula = '"' + ",".join(_BODY_PART_LIST) + '"'
    for col_letter in ("B", "D", "F", "H"):
        dv = DataValidation(
            type="list",
            formula1=dv_formula,
            allow_blank=True,
            showDropDown=False,   # False = show the arrow
            showErrorMessage=True,
            errorTitle="Invalid entry",
            error="Please select from the dropdown list or leave blank.",
        )
        dv.sqref = f"{col_letter}2:{col_letter}1000"
        ws.add_data_validation(dv)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# Sky East uses the same shared template
_generate_se_fabric_mapping_template = _generate_fabric_mapping_template


def _detect_fabric_mapping_columns(header_row: tuple) -> dict:
    return _detect_fabric_mapping_columns_impl(header_row)


def _parse_fabric_mapping_file(path: str) -> dict:
    """
    Parse a filled-in style-fabric mapping file.

    Layout is detected dynamically from the header row (Row 1).
    Recognised column names (case-insensitive, bilingual):
      Style:         "Style No.", "Style", "款式号", …
      Fabric N Body Part: "Fabric N Body Part", "面料N部位", …
      Fabric N Code:      "Fabric N Code", "面料N编号", …

    Falls back to the standard template layout (A=style, B/C=slot1, D/E=slot2, …)
    when no headers are recognised.

    Returns
    -------
    dict  style → list[FabricPart]   (parts with at least an hhn_no)
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return _parse_fabric_mapping_rows(all_rows)


def _show_giii_reference_section():
    """Style-Fabric mapping template download + upload for GIII."""
    store = get_store()

    st.markdown("#### 📋 Style-Fabric Mapping (GIII)")
    st.caption(
        "Map each GIII style to up to 4 HHN fabric codes. "
        "Composition is looked up automatically from the fabric database."
    )

    mapping_file = st.file_uploader(
        "Upload filled-in mapping (.xlsx)",
        type=["xlsx", "xls"],
        key="giii_mapping_uploader",
        help="Fill in the template and upload here. "
             "Fabric composition will be looked up automatically.",
    )
    st.caption(
        "💡 Need a blank template? Get it from **Admin → 📄 Templates → "
        "Style-Fabric Mapping Template**."
    )

    st.markdown("---")

    if mapping_file:
        col_imp, col_dry = st.columns(2)
        with col_imp:
            if st.button("▶ Import Mapping", type="primary",
                         use_container_width=True, key="giii_import_btn"):
                _run_giii_mapping_import(mapping_file)
        with col_dry:
            if st.button("🔍 Preview (dry run)", use_container_width=True,
                         key="giii_preview_btn"):
                _run_giii_mapping_import(mapping_file, dry_run=True)

    if st.session_state.get(SK.GIII_MAPPING):
        r = st.session_state.giii_mapping_result
        if r.get("dry_run"):
            st.info(
                f"**Dry-run preview** — {r['styles']} style(s), "
                f"{r['parts']} fabric part(s) found, "
                f"{r['enriched']} would be enriched from the HHN cache. "
                "Press **Import Mapping** to commit."
            )
        else:
            st.success(
                f"✅ Imported {r['styles']} style(s), "
                f"{r['parts']} fabric part(s) saved "
                f"({r['enriched']} enriched with composition from cache)."
            )
        if r.get("bad_format_hhns"):
            st.warning(
                f"⚠️ {len(r['bad_format_hhns'])} HHN code(s) do not match the expected "
                "format (e.g. HHN-AB-12345): "
                + ", ".join(r["bad_format_hhns"][:10])
                + (" …" if len(r["bad_format_hhns"]) > 10 else "")
            )
        if r.get("missing_hhns"):
            st.warning(
                f"⚠️ {len(r['missing_hhns'])} HHN code(s) not found in the HHN cache "
                "(composition left blank): "
                + ", ".join(r["missing_hhns"][:10])
                + (" …" if len(r["missing_hhns"]) > 10 else "")
            )
        if r.get("not_in_master"):
            st.error(
                f"🔴 {len(r['not_in_master'])} HHN code(s) also absent from the "
                "fabric master database — import may be incomplete: "
                + ", ".join(r["not_in_master"][:10])
                + (" …" if len(r["not_in_master"]) > 10 else "")
            )

    # Stored GIII fabric parts (read-only browse)
    df_giii = store.load_fabric_parts(source="giii")
    if not df_giii.empty:
        n_styles = df_giii["style"].nunique()
        with st.expander(f"Stored GIII fabric parts ({n_styles} styles)", expanded=False):
            st.dataframe(
                df_giii.rename(columns={
                    "style": "Style", "seq": "Seq", "body_part": "Body Part",
                    "hhn_no": "HHN No.", "composition": "Composition",
                    "weight_gsm": "Weight (g/m²)", "width_cm": "Width (cm)",
                    "updated_at": "Updated",
                }).drop(columns=["id", "source"], errors="ignore"),
                use_container_width=True, hide_index=True,
            )
            if st.button("🗑 Clear GIII fabric parts", key="giii_clear_parts"):
                store.delete_fabric_parts("giii")
                st.success("GIII fabric parts cleared.")
                st.rerun()

    st.markdown("---")
    _show_giii_consumption_section(store)


def _show_giii_consumption_section(store):
    """单耗/排版 (fabric consumption + marker) template download/upload — feeds
    the trailing columns of the buy plan. kg↔cm reconciled on import."""
    from po_extractor.ui_helpers.fabric_consumption import (
        consumption_template_bytes, parse_consumption_upload, CONSUMPTION_COLUMNS,
    )

    st.markdown("#### 🧵 " + t("Fabric Consumption / Marker (单耗 · 排版)"))
    st.caption(t(
        "Per-style 单耗 (fabric consumption) and 排版 (marker) data — appended "
        "as the last columns of the buy plan. Provide 单耗 in kg OR cm plus the "
        "effective width and fabric weight, and the other is calculated; give "
        "both and they're consistency-checked."
    ))

    existing = store.load_all_fabric_consumption()
    col_dl, col_up = st.columns(2)
    with col_dl:
        st.download_button(
            "⬇️ " + t("Download template / current data (.xlsx)"),
            data=consumption_template_bytes(existing or None),
            file_name="fabric_consumption_template.xlsx",
            mime=_XLSX_MIME, use_container_width=True, key="cons_tpl_dl",
        )
    with col_up:
        cons_file = st.file_uploader(
            t("Upload filled 单耗/排版 template (.xlsx)"),
            type=["xlsx"], key="cons_uploader",
        )

    if cons_file and st.button("▶ " + t("Import 单耗/排版"), type="primary",
                               use_container_width=True, key="cons_import_btn"):
        try:
            records, warns = parse_consumption_upload(cons_file.getvalue())
            for w in warns:
                st.warning(f"⚠️ {w}")
            if records:
                n = store.save_fabric_consumption(records)
                st.success(f"✅ {t('Imported')} {n} {t('style(s).')}")
                st.rerun()
            else:
                st.error(t("No valid rows found — check the 款号 column."))
        except Exception as exc:
            st.error(t("Import failed:") + f" {exc}")

    # ── Searchable stored-consumption panel ───────────────────────────────────
    st.markdown("##### 🔎 " + t("Stored consumption data"))
    if not existing:
        st.info(t("No consumption data yet — download the template above, fill it "
                  "in, and upload."))
        return

    _hdr = {k: h for h, k in CONSUMPTION_COLUMNS}
    df_cons = pd.DataFrame(existing).rename(columns=_hdr)

    q = st.text_input(t("Search by 款号 / style"), key="cons_search",
                      placeholder=t("Type part of a style number…")).strip()
    view = df_cons
    if q:
        view = df_cons[df_cons["款号"].astype(str).str.contains(q, case=False, na=False)]

    st.caption(f"**{len(view)}** / {len(df_cons)} {t('styles')}")
    st.dataframe(view, use_container_width=True, hide_index=True)

    st.download_button(
        "⬇️ " + t("Download shown rows (.xlsx)"),
        data=_consumption_view_bytes(view, _hdr),
        file_name="fabric_consumption_view.xlsx",
        mime=_XLSX_MIME, key="cons_view_dl",
    )
    if st.button("🗑 " + t("Clear all consumption data"), key="cons_clear"):
        store.clear_fabric_consumption()
        st.success(t("Consumption data cleared."))
        st.rerun()


def _consumption_view_bytes(df, hdr_map) -> bytes:
    """The currently-filtered consumption table as .xlsx (Chinese headers)."""
    import io as _io
    buf = _io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xw:
        df.to_excel(xw, index=False, sheet_name="单耗排版")
    return buf.getvalue()


def _run_giii_mapping_import(mapping_file, dry_run: bool = False,
                             source: str = "giii"):
    """Parse the uploaded mapping file and (optionally) save to the DB.

    Parameters
    ----------
    mapping_file : Streamlit UploadedFile object
    dry_run      : when True, preview only — do not write to the database
    source       : client identifier stored in ``style_fabric_parts.source``
                   (``"giii"`` or ``"sky_east"``); controls both the DB write
                   and the session-state key used to surface results.
    """
    import shutil as _shutil
    tmpdir = tempfile.mkdtemp()
    try:
        path = os.path.join(tmpdir, mapping_file.name)
        with open(path, "wb") as f:
            f.write(mapping_file.getbuffer())

        try:
            style_parts = _parse_fabric_mapping_file(path)
        except Exception as exc:
            st.error(f"Could not parse mapping file: {exc}")
            return

        store = get_store()
        fabric_store = get_fabric_master_store()
        total_parts = sum(len(v) for v in style_parts.values())
        enriched = 0
        missing_hhns: list[str] = []        # not in HHN cache
        not_in_master: list[str] = []       # not in fabric_master either
        bad_format_hhns: list[str] = []     # HHN format doesn't match pattern

        import re as _re
        _HHN_RE = _re.compile(r'^[A-Za-z]{2,5}-[A-Za-z]{1,5}-?\d{4,8}$')

        # Enrich each FabricPart with composition from the HHN cache
        from po_extractor.models.fabric_part import FabricPart
        for style, parts in style_parts.items():
            for p in parts:
                if p.hhn_no:
                    # Format check
                    if not _HHN_RE.match(p.hhn_no.strip()) and p.hhn_no not in bad_format_hhns:
                        bad_format_hhns.append(p.hhn_no)

                    detail = store.get_fabric_by_hhn(p.hhn_no)
                    if detail:
                        p.composition = detail.get("composition", "")
                        p.weight_gsm  = detail.get("weight_gsm", 0) or 0
                        p.width_cm    = detail.get("width_cm",   0) or 0
                        if p.composition:
                            enriched += 1
                    else:
                        if p.hhn_no not in missing_hhns:
                            missing_hhns.append(p.hhn_no)
                        # Also cross-check fabric_master directly
                        master_rec = fabric_store.get_by_quality_no(p.hhn_no.strip())
                        if not master_rec and p.hhn_no not in not_in_master:
                            not_in_master.append(p.hhn_no)

        result = {
            "dry_run":        dry_run,
            "styles":         len(style_parts),
            "parts":          total_parts,
            "enriched":       enriched,
            "missing_hhns":   missing_hhns,
            "not_in_master":  not_in_master,
            "bad_format_hhns": bad_format_hhns,
        }

        if not dry_run:
            n = store.save_fabric_parts_batch(source, style_parts)
            result["saved"] = n

        # Store under a source-specific key so GIII and Sky East results don't clash
        st.session_state[f"{source}_mapping_result"] = result
    finally:
        _shutil.rmtree(tmpdir, ignore_errors=True)


def _compute_giii_missing_df(df: "pd.DataFrame") -> "pd.DataFrame":
    """Return the rows of *df* that are missing factory or export date.

    *df* is a ``list_pos`` frame already scoped to the caller's assigned
    companies (show_smart_upload_tab fetches it once per rerun and shares it
    across sub-tabs — an unassigned non-admin gets an empty frame). Without
    that scoping the tab leaked — and let a non-admin edit — every
    company's POs.
    """
    if df.empty:
        return pd.DataFrame()
    mask = (
        df["factory"].fillna("").str.strip().eq("") |
        df["xport_date"].fillna("").str.strip().eq("")
    )
    return df[mask].copy()
