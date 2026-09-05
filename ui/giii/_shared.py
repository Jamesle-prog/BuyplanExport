"""Shared constants and helpers for the GIII tab sub-modules."""
from __future__ import annotations
import re
import streamlit as st
from po_extractor.ui_helpers import (
    live_label_for,
    load_live_schema, schema_seed_rows,
    enrich_cn_color as _enrich_cn_color_impl,
)
from po_extractor.utils.normalize import normalize_header as _normalize_header
from ui.shared import XLSX_MIME, CSV_MIME, ZIP_MIME, ProgressTracker
from ui.stores import get_color_translation_store

# ---------------------------------------------------------------------------
# MIME aliases
# ---------------------------------------------------------------------------
_XLSX_MIME = XLSX_MIME
_CSV_MIME  = CSV_MIME
_ZIP_MIME  = ZIP_MIME

# ---------------------------------------------------------------------------
# ProgressTracker alias
# ---------------------------------------------------------------------------
_ProgressTracker = ProgressTracker

# ---------------------------------------------------------------------------
# Live schema / label helpers
# ---------------------------------------------------------------------------
from po_extractor.config import SCHEMA_PATH as _SCHEMA_PATH, CACHE_TTL_SECONDS


@st.cache_data(ttl=CACHE_TTL_SECONDS)
def _cached_schema() -> list[dict]:
    rows = load_live_schema(_SCHEMA_PATH)
    return rows if rows else schema_seed_rows()


def live_label(db_col: str, fallback: str | None = None) -> str:
    return live_label_for(_cached_schema(), db_col, fallback)


# ---------------------------------------------------------------------------
# Color enrichment wrapper
# ---------------------------------------------------------------------------

def _enrich_cn_color(df_size, df_meta):
    lookup = get_color_translation_store().build_lookup_dict()
    return _enrich_cn_color_impl(df_size, df_meta, lookup)


# ---------------------------------------------------------------------------
# Module-level aliases
# ---------------------------------------------------------------------------

_norm_mapping_header = _normalize_header

# ---------------------------------------------------------------------------
# Smart upload confidence badges
# ---------------------------------------------------------------------------
_CONF_BADGE = {"high": "🟢", "medium": "🟡", "low": "🔴"}

# ---------------------------------------------------------------------------
# Fax-copy (AS400 doubled-font) parser helpers
# shared by msg_extraction, kl_extraction, tk_eu_extraction
# ---------------------------------------------------------------------------

def _undouble(s: str) -> str:
    """Collapse doubled characters produced by the AS400 fax-copy font."""
    return re.sub(r'(.)\1', r'\1', s)


def files_signature(files) -> tuple:
    """Order-insensitive identity of an uploader's file set (name + size).

    Stored next to extraction results so a changed upload selection
    invalidates them — previously the table (and its download button) kept
    serving the previous batch after the files were swapped.
    """
    return tuple(sorted((f.name, getattr(f, "size", None)) for f in (files or [])))


# Canonical fax-PO size column order (msg / kl / tk_eu builders).
# infornexus_extraction keeps its own superset — its output column order is
# part of that workbook's layout and must not silently change.
FAX_SIZE_ORDER = ['XS', 'S', 'M', 'L', 'XL', 'XXL', '1X', '2X', '3X']

# Shared workbook palette (every GIII extraction Excel builder uses these).
XL_NAVY   = 'FF1F3864'
XL_WHITE  = 'FFFFFFFF'
XL_YELLOW = 'FFFFF2CC'
XL_GREY   = 'FFD9D9D9'
XL_LTBLUE = 'FFDEEAF1'
XL_GREEN  = 'FFE2EFDA'


def make_excel_style_kit(hdr_bg: str = XL_NAVY, autofit_max: int = 50,
                         wrap: bool = False):
    """Standard cell styling for the GIII extraction Excel builders.

    Returns a namespace of the helpers every builder previously nested
    inline (border/fill/align/style/hdr/autofit) — Arial 10, thin black
    borders, white-on-``hdr_bg`` wrapped headers. ``hdr_bg`` sets the
    default header colour (TK EU uses its teal); ``hdr()`` still accepts a
    per-call ``bg`` override; ``wrap`` makes ``style()`` wrap text too (the
    InforNexus comparison workbook).
    """
    from types import SimpleNamespace
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    def side():
        return Side(style='thin', color='FF000000')

    def border():
        s = side()
        return Border(left=s, right=s, top=s, bottom=s)

    def fill(colour):
        return PatternFill('solid', fgColor=colour) if colour else PatternFill()

    def align(h='left'):
        return Alignment(horizontal=h, vertical='center')

    def style(cell, bold=False, bg=None, align='left', num_fmt=None):
        cell.font      = Font(name='Arial', bold=bold, size=10)
        cell.fill      = fill(bg)
        cell.alignment = Alignment(horizontal=align, vertical='center',
                                   wrap_text=wrap)
        cell.border    = border()
        if num_fmt:
            cell.number_format = num_fmt

    def hdr(cell, value, bg=None):
        cell.value     = value
        cell.font      = Font(name='Arial', bold=True, color=XL_WHITE, size=10)
        cell.fill      = fill(bg or hdr_bg)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = border()

    def autofit(ws, mn=8, mx=None):
        mx = autofit_max if mx is None else mx
        for col in ws.columns:
            ltr = get_column_letter(col[0].column)
            w   = max((len(str(c.value or '')) for c in col), default=0)
            ws.column_dimensions[ltr].width = min(max(w + 2, mn), mx)

    return SimpleNamespace(side=side, border=border, fill=fill, align=align,
                           style=style, hdr=hdr, autofit=autofit)


def pdf_text_lines(pdf_bytes: bytes):
    """``(lines, grep)`` for a fax-copy PDF: every page's text joined and
    split into lines, plus ``grep(pattern)`` — the first ``re.search`` match
    across those lines, or None. The opening block every fax extractor
    (MSG / KL / TK EU / InforNexus) used to repeat."""
    import io
    import pdfplumber

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        lines = '\n'.join(p.extract_text() or '' for p in pdf.pages).split('\n')

    def grep(pat: str):
        for l in lines:
            m = re.search(pat, l)
            if m:
                return m
        return None

    return lines, grep


def iter_pdf_payloads(files):
    """Yield ``(name, pdf_bytes, email_subject)`` for each uploaded file.

    Accepts the mixed uploads the fax sections take: bare fax ``.pdf`` files
    pass straight through (subject ``''``); ``.msg`` Outlook emails are
    unpacked to their first PDF attachment, with the email subject carried
    along for PO-number fallbacks. Unreadable emails and emails without a
    PDF attachment are warned about and skipped — shared by the MSG and
    TK EU sections, which previously each had their own copy of this loop.
    """
    import os
    import tempfile

    with tempfile.TemporaryDirectory() as tmp:
        for uf in files:
            if uf.name.lower().endswith('.pdf'):
                yield uf.name, bytes(uf.getbuffer()), ''
                continue

            import extract_msg
            msg_path = os.path.join(tmp, os.path.basename(uf.name))
            with open(msg_path, 'wb') as f:
                f.write(uf.getbuffer())
            try:
                msg = extract_msg.openMsg(msg_path)
            except Exception as exc:
                st.warning(f"Could not open {uf.name}: {exc}")
                continue

            pdf_data = None
            for att in msg.attachments:
                if (att.longFilename or att.shortFilename or '').lower().endswith('.pdf'):
                    pdf_data = att.data
                    break
            if pdf_data is None:
                st.warning(f"No PDF attachment in {uf.name} — skipped.")
                continue

            yield uf.name, pdf_data, (msg.subject or '')


def mask_ai_creds() -> tuple[str | None, str]:
    """Return (api_key, model) for AI-assisted price masking.

    The api_key is the configured DeepSeek key only when the admin has enabled
    "AI-assisted price masking" (Admin → Settings); otherwise None, so masking
    uses the built-in pattern/keyword rules alone. Shared by every price-mask
    call site (GIII PDF, GIII/Zalando Excel, Sky East).
    """
    from po_extractor.store.app_settings_store import (
        KEY_MASK_USE_AI, KEY_DEEPSEEK_API_KEY, KEY_DEEPSEEK_MODEL,
    )
    from ui.stores import get_app_settings_store
    s = get_app_settings_store()
    model = s.get(KEY_DEEPSEEK_MODEL, "deepseek-chat")
    if s.get(KEY_MASK_USE_AI, "false") == "true":
        return (s.get(KEY_DEEPSEEK_API_KEY, "") or None), model
    return None, model


def make_en_to_cn_translator():
    """Return a cached English→Chinese translator callable (DeepSeek), or None
    if no API key is configured. Used to render CPRS requirement wording (which
    comes back in English) as Chinese on the GIII buy plan. Text that already
    contains Chinese is passed through untouched."""
    from po_extractor.store.app_settings_store import (
        KEY_DEEPSEEK_API_KEY, KEY_DEEPSEEK_MODEL,
    )
    from ui.stores import get_app_settings_store
    s = get_app_settings_store()
    key = s.get(KEY_DEEPSEEK_API_KEY, "")
    if not key:
        return None
    model = s.get(KEY_DEEPSEEK_MODEL, "deepseek-chat")
    cache: dict[str, str] = {}

    def translate(text: str) -> str:
        text = (text or "").strip()
        if not text or any("一" <= c <= "鿿" for c in text):
            return text
        if text in cache:
            return cache[text]
        try:
            from openai import OpenAI
            from po_extractor.utils.deepseek_client import chat_kwargs, max_tokens_for
            client = OpenAI(api_key=key, base_url="https://api.deepseek.com")
            resp = client.chat.completions.create(
                model=model,
                messages=[{"role": "system", "content":
                           "Translate the user's English garment packaging / "
                           "carton requirement text into concise Simplified "
                           "Chinese. Return only the translation."},
                          {"role": "user", "content": text[:2000]}],
                max_tokens=max_tokens_for(model, 256),
                **chat_kwargs(model),
            )
            out = (resp.choices[0].message.content or "").strip() or text
        except Exception:
            out = text
        cache[text] = out
        return out

    return translate


def drop_stale_results(results_key: str, sig_key: str, sig: tuple):
    """Return current results for *results_key*, dropping them first if the
    uploader's file set changed since extraction (signature mismatch) — a
    swapped selection must not display/download the previous batch."""
    results = st.session_state.get(results_key)
    if results and st.session_state.get(sig_key) != sig:
        st.session_state[results_key] = None
        results = None
    return results


_SIZE_CODES = r'(?:XXS|XS|XXL|XL|[123]XL|[123]X|OSFM|OSM|OSF|OS|S|M|L)'
_FIRST_RE   = re.compile(
    rf'^(\d{{3}})\s+(\S+)\s+(.+?)\s+({_SIZE_CODES})\s+(\d+)\s+(\d{{12,13}})\s+([\d.]+)'
)
_CONT_RE    = re.compile(rf'^({_SIZE_CODES})\s+(\d+)\s+(\d{{12,13}})')

# ---------------------------------------------------------------------------
# Body part list (used for fabric mapping template)
# ---------------------------------------------------------------------------
_BODY_PART_LIST = [
    "Main Body / 大身",
    "Upper Body / 上身",
    "Lower Body / 下身",
    "Lining / 里布",
    "Sleeve / 袖子",
    "Collar / 领子",
    "Cuff / 袖口",
    "Hood / 帽子",
    "Pocket / 口袋布",
    "Pocket Lining / 口袋里布",
    "Pocket Mesh / 网眼布",
    "Waistband / 腰头",
    "Front Panel / 前片",
    "Back Panel / 后片",
    "Facing / 贴边",
    "Interlining / 衬布",
    "Piping / 嵌条",
    "Trim / 辅料",
]
