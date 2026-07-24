"""Cross-company Order Summary tab."""
from __future__ import annotations

import io

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from auth.companies import COMPANY_GIII, COMPANY_SKY_EAST
from po_extractor.ui_helpers.combined_summary import (
    STANDARD_KEYS, STANDARD_LABELS, load_standard_orders,
)
from ui.i18n import t
from ui.session_keys import SK
from ui.shared import _th, _tr, guard_multiselect_state, XLSX_MIME
from ui.stores import get_store, get_sky_east_store


# ---------------------------------------------------------------------------
# PO Tracker helpers
# ---------------------------------------------------------------------------

_TRACKER_COLS = {
    "po_number":          "PO Number",
    "issue_date":         "Issue Date",
    "version":            "Version",
    "buyer":              "Buyer",
    "seller":             "Seller",
    "factory":            "Factory",
    "ship_to":            "Ship To",
    "customer":           "End Customer",
    "incoterm":           "Incoterm",
    "origin_port":        "Origin Port",
    "payment_terms":      "Payment Terms",
    "discount":           "Discount",
    "approval_status":    "Approved",
    "country_of_origin":  "COO",
    "division_code":      "Div Code",
    "division_name":      "Division",
    "season":             "Season",
    "issued_by":          "Issued By",
    "style":              "Style",
    "fabric":             "Fabric",
    "style_description":  "Description",
    "style_group":        "Style Group",
    "unit_cost":          "Unit Cost",
    "total_units":        "Total Qty",
    "line_extended_cost": "Extended Cost",
    "xport_date":         "Ex-Fty Date",
    "factory_ship_date":  "Factory Ship-by",
    "company":            "Company",
    "source_format":      "Format",
}

_DEFAULT_COLS = [
    "po_number", "issue_date", "buyer", "seller", "factory",
    "customer", "incoterm", "country_of_origin", "division_name",
    "season", "style", "style_description", "unit_cost",
    "total_units", "line_extended_cost", "xport_date",
]


@st.cache_data(max_entries=8, show_spinner=False)
def _build_tracker_excel(df: pd.DataFrame) -> bytes:
    """Pure function (DataFrame in → xlsx bytes out) — safe to cache so the
    styled workbook is not rebuilt on every fragment rerun."""
    wb = Workbook()
    ws = wb.active
    ws.title = "PO Tracker"

    NAVY, BLUE, LBLUE, WHITE = "1F3864", "2C5F8A", "D9E6F2", "FFFFFF"

    def _bdr():
        s = Side(style="thin", color="CCCCCC")
        return Border(left=s, right=s, top=s, bottom=s)

    # Title row  (cell var named `tc` — `t` is the i18n translator import)
    ws.merge_cells(f"A1:{get_column_letter(len(df.columns))}1")
    tc = ws["A1"]
    tc.value = "PO Tracker — Commercial Summary"
    tc.font = Font(name="Arial", bold=True, size=13, color=WHITE)
    tc.fill = PatternFill("solid", fgColor=NAVY)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # Header row
    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(row=2, column=ci, value=col)
        c.font = Font(name="Arial", bold=True, size=10, color=WHITE)
        c.fill = PatternFill("solid", fgColor=BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _bdr()
    ws.row_dimensions[2].height = 28

    # Data rows
    for ri, row in enumerate(df.itertuples(index=False), 3):
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = Font(name="Arial", size=10)
            c.fill = PatternFill("solid", fgColor=LBLUE if ri % 2 == 0 else WHITE)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = _bdr()

    # Column widths
    width_map = {
        "PO Number": 16, "Issue Date": 13, "Version": 13, "Buyer": 24, "Seller": 28,
        "Factory": 32, "Ship To": 20, "End Customer": 22, "Incoterm": 14,
        "Origin Port": 13, "Payment Terms": 16, "Discount": 10, "Approved": 10,
        "COO": 10, "Div Code": 10, "Division": 18, "Season": 9, "Issued By": 18,
        "Style": 12, "Description": 24, "Style Group": 14, "Unit Cost": 11,
        "Total Qty": 11, "Extended Cost": 16, "Ex-Fty Date": 13,
        "Factory Ship-by": 15, "Company": 16, "Format": 14,
    }
    for ci, col in enumerate(df.columns, 1):
        ws.column_dimensions[get_column_letter(ci)].width = width_map.get(col, 14)

    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@st.cache_data(max_entries=8, show_spinner=False)
def _build_all_orders_excel(df: pd.DataFrame) -> bytes:
    """Pure function (DataFrame in → xlsx bytes out) — cached like
    ``_build_tracker_excel`` so the export isn't rebuilt on every rerun."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as wr:
        df.to_excel(wr, sheet_name="All Orders", index=False)
    return buf.getvalue()


@st.cache_data(max_entries=8, show_spinner=False)
def _build_overview_excel(summary_df: pd.DataFrame,
                          giii_df: pd.DataFrame | None,
                          se_df: pd.DataFrame | None) -> bytes:
    """Pure function (DataFrames in → xlsx bytes out) — cached like
    ``_build_tracker_excel``.  ``giii_df`` / ``se_df`` are the pre-renamed
    detail sheets, or None to omit that sheet."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as wr:
        summary_df.to_excel(wr, sheet_name="Summary", index=False)
        if giii_df is not None:
            giii_df.to_excel(wr, sheet_name="GIII POs", index=False)
        if se_df is not None:
            se_df.to_excel(wr, sheet_name="Sky East Items", index=False)
    return buf.getvalue()


def _show_po_tracker(df: pd.DataFrame) -> None:
    st.subheader(f"📋 {t('PO Tracker')}")
    st.caption(t("One row per PO — all commercial fields for comparison and tracking."))

    if df.empty:
        st.info(t("No POs stored yet. Upload PDFs via the GIII tab to populate."))
        return

    # ── Filters ──────────────────────────────────────────────────────────────
    fc1, fc2, fc3 = st.columns(3)
    with fc1:
        seasons = sorted(df["season"].dropna().unique().tolist()) if "season" in df.columns else []
        guard_multiselect_state("tracker_season", seasons)
        sel_season = st.multiselect(t("Season"), seasons, key="tracker_season")
    with fc2:
        divs = sorted(df["division_name"].dropna().unique().tolist()) if "division_name" in df.columns else []
        guard_multiselect_state("tracker_div", divs)
        sel_div = st.multiselect(t("Division"), divs, key="tracker_div")
    with fc3:
        buyers = sorted(df["buyer"].dropna().unique().tolist()) if "buyer" in df.columns else []
        guard_multiselect_state("tracker_buyer", buyers)
        sel_buyer = st.multiselect(t("Buyer"), buyers, key="tracker_buyer")

    if sel_season:
        df = df[df["season"].isin(sel_season)]
    if sel_div:
        df = df[df["division_name"].isin(sel_div)]
    if sel_buyer:
        df = df[df["buyer"].isin(sel_buyer)]

    # ── Column selector ───────────────────────────────────────────────────────
    avail = [k for k in _TRACKER_COLS if k in df.columns]
    default = [k for k in _DEFAULT_COLS if k in avail]
    # Seed-once + guard (never key= AND default= together — CLAUDE.md).
    if "tracker_cols" not in st.session_state:
        st.session_state["tracker_cols"] = default
    else:
        guard_multiselect_state("tracker_cols", avail)
    picked = st.multiselect(
        t("Columns"),
        options=avail,
        format_func=lambda k: _TRACKER_COLS.get(k, k),
        key="tracker_cols",
    )
    show_cols = picked or default
    display_df = df[show_cols].rename(columns=_TRACKER_COLS)

    st.caption(f"{len(display_df):,} {t('PO(s)')}")
    # Translate headers only for on-screen display; the Excel export keeps the
    # English headers (its width_map and cache key stay language-independent).
    st.dataframe(display_df.rename(columns={c: _th(c) for c in display_df.columns}),
                 use_container_width=True, hide_index=True)

    # ── Download ──────────────────────────────────────────────────────────────
    xlsx = _build_tracker_excel(display_df)
    st.download_button(
        "⬇️ " + t("Download PO Tracker (.xlsx)"),
        data=xlsx,
        file_name="po_tracker.xlsx",
        mime=XLSX_MIME,
        key="tracker_dl",
    )


def show_summary_tab(user_cos: list[str], admin_mode: bool) -> None:
    """Cross-company order summary, filtered by user permissions."""
    # Same convention as the Tracking tab: a non-admin with no assigned
    # companies sees nothing.  Without this gate the empty list fell through
    # the stores' falsy check (`if companies:`) to an UNFILTERED query —
    # i.e. a freshly created user saw every company's commercial data.
    if not admin_mode and not user_cos:
        st.info(t(
            "No companies assigned to your account. "
            "Contact an administrator to be granted access."
        ))
        return

    # ── Shared data — fetched ONCE per rerun ─────────────────────────────────
    # st.tabs renders every sub-tab body on each rerun, so per-sub-tab store
    # reads used to run 2–4x per interaction.  The two company-filter shapes
    # are kept distinct on purpose: Tracker/All Orders treat an admin as
    # unrestricted, Overview scopes an admin to their assigned companies.
    tracker_cos  = user_cos if user_cos and not admin_mode else None
    overview_cos = user_cos if user_cos else None
    unrestricted = admin_mode or not user_cos
    can_see_giii    = unrestricted or bool(user_cos)
    can_see_zalando = unrestricted or COMPANY_SKY_EAST in [c.strip() for c in user_cos]

    po_store = get_store()
    se_store = get_sky_east_store()
    tracker_df = po_store.list_pos(companies=tracker_cos)
    overview_giii_df = (tracker_df if overview_cos == tracker_cos
                        else po_store.list_pos(companies=overview_cos))
    se_items = se_store.list_items() if can_see_zalando else pd.DataFrame()

    std_all_df = load_standard_orders(
        po_store, se_store,
        companies=tracker_cos,
        include_giii=can_see_giii,
        include_sky_east=can_see_zalando,
        sky_east_company=COMPANY_SKY_EAST,
        giii_df=tracker_df,
        se_items=se_items if can_see_zalando else None,
    )
    std_overview_df = (std_all_df if overview_cos == tracker_cos
                       else load_standard_orders(
                           po_store, se_store,
                           companies=overview_cos,
                           include_giii=can_see_giii,
                           include_sky_east=can_see_zalando,
                           sky_east_company=COMPANY_SKY_EAST,
                           giii_df=overview_giii_df,
                           se_items=se_items if can_see_zalando else None,
                       ))

    tab_overview, tab_tracker, tab_all = st.tabs([
        f"📊 {t('Overview')}", f"📋 {t('PO Tracker')}", f"🧾 {t('All Orders')}",
    ])

    with tab_tracker:
        _show_po_tracker(tracker_df)

    with tab_overview:
        _show_overview(user_cos, admin_mode, overview_giii_df, se_items,
                       std_overview_df)

    with tab_all:
        _show_all_orders(std_all_df)


def _show_all_orders(df: pd.DataFrame) -> None:
    """One combined table of every client's orders in the standard column set.

    *df* is the standard-shape frame loaded once in ``show_summary_tab``.
    """
    st.subheader(f"🧾 {t('All Orders')}")
    st.caption(t(
        "Every client's orders in one table with standardized columns — "
        "GIII POs and Sky East contract items side by side."
    ))

    if df.empty:
        st.info(t("No order data available for your permitted companies."))
        return

    # ── Filters ──────────────────────────────────────────────────────────────
    fc1, fc2 = st.columns([1, 2])
    with fc1:
        companies = sorted(x for x in df["company"].dropna().unique().tolist() if x)
        guard_multiselect_state(SK.SUM_ALL_CLIENTS, companies)
        sel_cos = st.multiselect(t("Client"), companies, key=SK.SUM_ALL_CLIENTS)
    with fc2:
        query = st.text_input(
            t("Search (PO / Contract / Style / Color)"),
            key=SK.SUM_ALL_SEARCH, placeholder="PO123 / PC001 / ...",
        ).strip()

    if sel_cos:
        df = df[df["company"].isin(sel_cos)]
    if query:
        q = query.lower()
        mask = (
            df["po_number"].astype(str).str.lower().str.contains(q, na=False)
            | df["contract_no"].astype(str).str.lower().str.contains(q, na=False)
            | df["style"].astype(str).str.lower().str.contains(q, na=False)
            | df["color"].astype(str).str.lower().str.contains(q, na=False)
        )
        df = df[mask]

    # ── Column picker (standard keys, English labels translated on screen) ──
    if SK.SUM_ALL_COLS not in st.session_state:
        st.session_state[SK.SUM_ALL_COLS] = STANDARD_KEYS
    else:
        guard_multiselect_state(SK.SUM_ALL_COLS, STANDARD_KEYS)
    picked = st.multiselect(
        t("Columns"),
        options=STANDARD_KEYS,
        format_func=lambda k: STANDARD_LABELS.get(k, k),
        key=SK.SUM_ALL_COLS,
    )
    show = picked or STANDARD_KEYS

    display = df[show].rename(columns={k: _th(STANDARD_LABELS[k]) for k in show})
    st.caption(f"{len(display):,} {t('row(s)')} · "
               f"{int(df['units'].sum()):,} {t('units')}")
    st.dataframe(display, use_container_width=True, hide_index=True)

    # ── Download (English standard headers, language-independent) ───────────
    xlsx = _build_all_orders_excel(df[show].rename(columns=STANDARD_LABELS))
    st.download_button(
        "⬇️ " + t("Download All Orders (.xlsx)"),
        data=xlsx,
        file_name="all_orders.xlsx",
        mime=XLSX_MIME,
        key="sum_all_dl",
    )


def _show_overview(user_cos: list[str], admin_mode: bool,
                   giii_df: pd.DataFrame, se_items: pd.DataFrame,
                   std_df: pd.DataFrame) -> None:
    """Original cross-company overview content.

    *giii_df* / *se_items* / *std_df* are loaded once in ``show_summary_tab``
    (Overview-scoped company filtering) and passed in.
    """
    st.subheader(f"📊 {t('Order Summary')}")
    st.caption(t("Aggregated view of all orders across clients, filtered to your permitted companies."))

    # Permission helpers
    # admin or empty user_cos = unrestricted
    unrestricted = admin_mode or not user_cos

    can_see_giii    = unrestricted or bool(user_cos)          # any company assigned includes GIII-type data
    can_see_zalando = unrestricted or COMPANY_SKY_EAST in [c.strip() for c in user_cos]

    # Copies: the expanders below add virtual columns in place, and the
    # frames are shared with the other sub-tabs.
    giii_df = giii_df.copy() if can_see_giii else pd.DataFrame()
    se_df = (se_items.copy()
             if can_see_zalando and not se_items.empty else pd.DataFrame())

    # Unified summary rows (one row per Company) come from the standard
    # cross-pipeline shape (std_df) — no per-pipeline branching here.
    # Columns: Company | POs | Styles | Units | Latest Ex-Fty | Factory | COO | Source

    def _mode_nonblank(series: pd.Series) -> str:
        vals = series.replace("", pd.NA).dropna()
        return vals.mode()[0] if not vals.empty else ""

    summary_rows = []
    for company, grp in std_df.groupby("company", dropna=False):
        is_sky_east = (grp["source"] == "sky_east").all()
        summary_rows.append({
            "Company":       company or "—",
            "Source":        COMPANY_SKY_EAST if is_sky_east else COMPANY_GIII,
            "POs":           grp["po_number"].nunique(),
            "Styles":        grp["style"].nunique(),
            "Units":         int(grp["units"].sum()),
            "Factory":       _mode_nonblank(grp["factory"]),
            "COO":           _mode_nonblank(grp["country_of_origin"]),
            "Latest Ex-Fty": grp["ex_fty_date"].max(),
        })

    # Top metrics
    total_pos    = sum(r["POs"]    for r in summary_rows)
    total_styles = sum(r["Styles"] for r in summary_rows)
    total_units  = sum(r["Units"]  for r in summary_rows)
    total_cos    = len(summary_rows)

    mc = st.columns(4)
    mc[0].metric(_th("Companies"),    f"{total_cos:,}")
    mc[1].metric(_th("Total POs"),    f"{total_pos:,}")
    mc[2].metric(_th("Total Styles"), f"{total_styles:,}")
    mc[3].metric(_th("Total Units"),  f"{total_units:,}")

    st.divider()

    # Unified summary table
    if summary_rows:
        summary_df = pd.DataFrame(summary_rows,
                                  columns=["Company", "Source", "POs", "Styles", "Units",
                                           "Factory", "COO", "Latest Ex-Fty"])
        _sum_rename = {c: _th(c) for c in summary_df.columns}
        st.dataframe(summary_df.rename(columns=_sum_rename),
                     use_container_width=True, hide_index=True)
    else:
        st.info(t("No order data available for your permitted companies."))

    st.divider()

    # ── Detail expanders with user-configurable column selection ─────────────
    show_cols: list[str] = []
    labels: dict = {}
    se_show: list[str] = []
    se_labels: dict = {}

    if can_see_giii and not giii_df.empty:
        with st.expander("🔍 " + t("GIII — full PO list")):
            # Virtual "pc_no" column — mirrors po_number for users whose
            # external workflow refers to it as "PC No." (Purchase Contract No.).
            if "po_number" in giii_df.columns and "pc_no" not in giii_df.columns:
                giii_df["pc_no"] = giii_df["po_number"]

            # All selectable columns (key → human label)
            giii_field_labels: dict = {
                "po_number":         "PO No.",
                "pc_no":             "PC No.",
                "company":           "Company",
                "style":              "Style",
                "factory":           "Factory",
                "country_of_origin": "COO",
                "xport_date":        "Ex-Fty",
                "issue_date":        "Issue Date",
                "version":           "Version",
                "division_code":     "Division Code",
                "division_name":     "Division Name",
                "total_units":       "Units",
            }
            giii_avail = [k for k in giii_field_labels if k in giii_df.columns]
            giii_default = [k for k in
                            ["po_number", "pc_no", "company", "style", "factory",
                             "country_of_origin", "xport_date", "total_units"]
                            if k in giii_avail]

            # Seed-once + guard (never key= AND default= together — CLAUDE.md).
            if "sum_giii_cols" not in st.session_state:
                st.session_state["sum_giii_cols"] = giii_default
            else:
                guard_multiselect_state("sum_giii_cols", giii_avail)
            picked = st.multiselect(
                t("Columns to display"),
                options=giii_avail,
                format_func=lambda k: giii_field_labels.get(k, k),
                key="sum_giii_cols",
            )
            show_cols = picked or giii_default
            labels = _tr({k: giii_field_labels[k] for k in show_cols})
            st.dataframe(giii_df[show_cols].rename(columns=labels),
                         use_container_width=True, hide_index=True)

    if can_see_zalando and not se_df.empty:
        with st.expander("🔍 " + t("Sky East — full item list")):
            # Virtual "company" column — Sky East is the source by definition.
            if "company" not in se_df.columns:
                se_df["company"] = COMPANY_SKY_EAST

            se_field_labels: dict = {
                "pc_no":       "PC No.",
                "zalando_po":  "PO No.",
                "company":     "Company",
                "style":       "Style",
                "brand":       "Brand",
                "color_name":  "Color",
                "config_sku":  "Config SKU",
                "fabric_item_no": "Fabric Code",
                "total_qty":   "Units",
                "ex_fty_date": "Ex-Fty",
                "return_label": "Return Label",
            }
            se_avail = [k for k in se_field_labels if k in se_df.columns]
            se_default = [k for k in
                          ["pc_no", "zalando_po", "company", "style", "brand",
                           "color_name", "total_qty", "ex_fty_date", "return_label"]
                          if k in se_avail]

            # Seed-once + guard (never key= AND default= together — CLAUDE.md).
            if "sum_se_cols" not in st.session_state:
                st.session_state["sum_se_cols"] = se_default
            else:
                guard_multiselect_state("sum_se_cols", se_avail)
            se_picked = st.multiselect(
                t("Columns to display"),
                options=se_avail,
                format_func=lambda k: se_field_labels.get(k, k),
                key="sum_se_cols",
            )
            se_show = se_picked or se_default
            se_labels = _tr({k: se_field_labels[k] for k in se_show})
            st.dataframe(se_df[se_show].rename(columns=se_labels),
                         use_container_width=True, hide_index=True)

    # Download
    if summary_rows:
        giii_export = (giii_df[show_cols].rename(columns=labels)
                       if can_see_giii and not giii_df.empty and show_cols
                       else None)
        se_export = (se_df[se_show].rename(columns=se_labels)
                     if can_see_zalando and not se_df.empty and se_show
                     else None)
        xlsx = _build_overview_excel(summary_df, giii_export, se_export)
        st.download_button("⬇️ " + t("Download Full Summary"), xlsx,
                           "order_summary.xlsx", key="sum_dl_all")

    if not can_see_giii and not can_see_zalando:
        st.warning(t("You don't have permission to view any company's orders. Contact an admin."))
