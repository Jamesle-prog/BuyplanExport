"""Admin: Output column-mapping (schema) editor view."""
from __future__ import annotations

import io
import os
from typing import Callable

import pandas as pd
import streamlit as st

from ui.i18n import t
from po_extractor.ui_helpers import load_live_schema, save_live_schema
from ui.shared import XLSX_MIME as _XLSX_MIME


def show_schema_editor(
    schema_path: str,
    on_schema_change: Callable[[], None] | None = None,
) -> None:
    """Admin UI: view and edit the output column mapping table.

    Parameters
    ----------
    schema_path
        Filesystem path of the live schema JSON.
    on_schema_change
        Optional callback fired after save/reset (typically clears the
        Streamlit cache holding the schema).
    """
    st.subheader(f"📋 {t('Output Column Mapping')}")
    st.caption(t(
        "This table controls every column heading in all export files (Excel downloads, reports). "
        "Edit the **Standard Label** to rename a column across all outputs instantly. "
        "Client alias columns show what that client calls the same field in their input files — "
        "useful for the dual-header row in Sky East Excel downloads. "
        "Delete rows to hide fields from outputs. (The internal DB Field is "
        "fixed because it must match the code — new fields are added in code.)"
    ))

    rows = load_live_schema(schema_path)
    df = pd.DataFrame(
        rows,
        columns=["db_col", "label", "sky_east", "infor", "legacy", "required", "notes"],
    )

    edited = st.data_editor(
        df,
        column_config={
            "db_col": st.column_config.TextColumn(
                t("DB Field (internal)"),
                help=t("Internal database / DataFrame column name. "
                       "Do not rename — this must match the code."),
                disabled=True, width="medium",
            ),
            "label": st.column_config.TextColumn(
                t("Standard Label ✏️"),
                help=t("Our company's standard output heading used in ALL export files."),
                width="medium",
            ),
            "sky_east": st.column_config.TextColumn(
                t("Sky East (client alias)"),
                help=t("How Sky East calls this field in their own files. "
                       "Shown in row 1 of dual-header Excel downloads."),
                width="medium",
            ),
            "infor": st.column_config.TextColumn(
                t("Infor Nexus (client alias)"),
                help=t("Column name as it appears in Infor Nexus PO files."),
                width="medium",
            ),
            "legacy": st.column_config.TextColumn(
                t("Legacy GIII (client alias)"),
                help=t("Column name in legacy GIII Excel PO files."),
                width="medium",
            ),
            "required": st.column_config.CheckboxColumn(
                t("Required"),
                help=t("Always included in standard outputs."),
                width="small",
            ),
            "notes": st.column_config.TextColumn(
                t("Notes"),
                help=t("Free-text explanation for this field."),
                width="large",
            ),
        },
        num_rows="dynamic",
        width="stretch",
        hide_index=True,
        key="schema_editor_tbl",
    )

    col_save, col_reset, col_dl = st.columns([1, 1, 2])

    with col_save:
        if st.button(f"💾 {t('Save Changes')}", type="primary", key="schema_save_btn"):
            records = edited.dropna(subset=["db_col"]).to_dict("records")
            records = [r for r in records if str(r.get("db_col", "")).strip()]
            save_live_schema(schema_path, records)
            if on_schema_change:
                on_schema_change()
            st.success(t("Saved. All exports will use the updated labels."))
            st.rerun()

    with col_reset:
        if st.button(f"↩ {t('Reset to Defaults')}", key="schema_reset_btn"):
            if os.path.exists(schema_path):
                os.remove(schema_path)
            if on_schema_change:
                on_schema_change()
            st.success(t("Reset to built-in defaults."))
            st.rerun()

    with col_dl:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Column Mapping"
        headers = ["DB Field", "Standard Label", "Sky East Alias",
                   "Infor Nexus Alias", "Legacy GIII Alias", "Required", "Notes"]
        for ci, h in enumerate(headers, 1):
            ws.cell(row=1, column=ci, value=h)
        for ri, r in enumerate(rows, 2):
            ws.cell(row=ri, column=1, value=r.get("db_col", ""))
            ws.cell(row=ri, column=2, value=r.get("label", ""))
            ws.cell(row=ri, column=3, value=r.get("sky_east", ""))
            ws.cell(row=ri, column=4, value=r.get("infor", ""))
            ws.cell(row=ri, column=5, value=r.get("legacy", ""))
            ws.cell(row=ri, column=6, value="Yes" if r.get("required") else "No")
            ws.cell(row=ri, column=7, value=r.get("notes", ""))
        buf = io.BytesIO()
        wb.save(buf)
        st.download_button(
            f"⬇ {t('Download mapping as Excel')}",
            data=buf.getvalue(),
            file_name="Column_Mapping.xlsx",
            mime=_XLSX_MIME,
            key="schema_dl_btn",
        )
