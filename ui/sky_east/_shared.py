"""Sky East tab — shared schema helpers and Excel-writer wrappers."""
from __future__ import annotations
import streamlit as st
from ui.session_keys import SK, COLOR_SOURCE_DB, COLOR_SOURCE_PROGRESS
from ui.i18n import t as _t
from po_extractor.ui_helpers import (
    live_label_for,
    load_live_schema, schema_seed_rows,
    parse_fabric_mapping_rows as _parse_fabric_mapping_rows,
    write_wash_label_excel as _write_wash_label_excel_impl,
    write_dual_header_excel as _write_dual_header_excel_impl,
    get_dual_header as _get_dual_header_impl,
)
from po_extractor.ui_helpers.fabric_mapping_template import generate_fabric_mapping_template

# Sky East uses the same shared template as GIII
_generate_se_fabric_mapping_template = generate_fabric_mapping_template

# Schema cache (module-local; re-uses same TTL pattern as app.py)
from po_extractor.config import SCHEMA_PATH as _SCHEMA_PATH, CACHE_TTL_SECONDS


@st.cache_data(ttl=CACHE_TTL_SECONDS)
def _cached_schema() -> list[dict]:
    rows = load_live_schema(_SCHEMA_PATH)
    return rows if rows else schema_seed_rows()


def live_label(db_col: str, fallback: str | None = None) -> str:
    return live_label_for(_cached_schema(), db_col, fallback)


# ---------------------------------------------------------------------------
# Local wrappers for ui_helpers Excel writers
# ---------------------------------------------------------------------------

def _get_dual_header() -> list[tuple[str, str, str]]:
    return _get_dual_header_impl(live_label)


def _write_wash_label_excel(
    df_enriched: "pd.DataFrame",
    image_cache: dict,
    fabric_parts_by_style: dict | None = None,
    styles: list[str] | None = None,
) -> bytes:
    return _write_wash_label_excel_impl(df_enriched, image_cache, fabric_parts_by_style,
                                        styles=styles)


def _write_dual_header_excel(
    df_enriched: "pd.DataFrame",
    sheet_name: str,
    writer,
    image_cache: dict | None = None,
) -> None:
    _write_dual_header_excel_impl(
        df_enriched, sheet_name, writer, image_cache, label_for=live_label
    )


# ---------------------------------------------------------------------------
# Fabric mapping file parser (local -- pure openpyxl)
# ---------------------------------------------------------------------------

def _parse_fabric_mapping_file(path: str) -> dict:
    """Parse a filled-in style-fabric mapping file; returns {style: [FabricPart]}."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return _parse_fabric_mapping_rows(all_rows)


def _parse_fabric_mapping_bytes(raw_bytes: bytes) -> dict:
    """Parse a style-fabric mapping from raw bytes (e.g. UploadedFile.getvalue()).

    Accepts the same file format as ``_parse_fabric_mapping_file`` but takes
    bytes instead of a path, so no temp file is needed for Streamlit uploads.
    Returns ``{style: [FabricPart, ...]}``.
    """
    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return _parse_fabric_mapping_rows(all_rows)


# ---------------------------------------------------------------------------
# Chinese color-source radio — shared between New Contracts and Buy Plan tabs
# ---------------------------------------------------------------------------

# Stable ordering for radio options (translation-independent).
COLOR_SOURCE_KEYS = [COLOR_SOURCE_DB, COLOR_SOURCE_PROGRESS]

# English keys used for _t() lookup.
_COLOR_SOURCE_EN: dict[str, str] = {
    COLOR_SOURCE_DB:       "Internal Database",
    COLOR_SOURCE_PROGRESS: "大货进度表 (HHN Contract File)",
}

# Per-option lookup-key explainer shown beneath the radio.
_COLOR_SOURCE_CAPTIONS: dict[str, str] = {
    COLOR_SOURCE_DB: (
        "🔑 **Lookup keys:** Client · Brand · English color name  "
        "→ **Returns:** 中文颜色, 中文颜色代码, 主标颜色  "
        "*(source: 🎨 Colors tab — color_translation table)*"
    ),
    COLOR_SOURCE_PROGRESS: (
        "🔑 **Lookup keys (priority order):**  \n"
        "1. PC No (所在PO / 客人PC No) · 款式 · 颜色  \n"
        "2. PO# · 款式 · 颜色  \n"
        "3. 款式 · 颜色  \n"
        "4. PC No · 款式 · 颜色代码  \n"
        "5. 款式 · 颜色代码  \n"
        "6. PC No · 款式  \n"
        "7. 款式 only (last resort)  \n"
        "→ **Returns:** 中文颜色, 中文颜色代码, 主标颜色  "
        "*(source: HHN Contract No. file — 大货进度表)*"
    ),
}


def show_color_source_radio(widget_key: str = "se_color_src_radio") -> None:
    """Render the Chinese-colour-source radio + per-option lookup-key caption.

    Reads the admin-configured default on the first render (when
    ``SK.SE_COLOR_SOURCE`` is None).  Writes the user's choice back into
    ``SK.SE_COLOR_SOURCE`` so both the New Contracts and Buy Plan sections
    share the same state regardless of which one the user interacts with.
    Pass a unique *widget_key* when rendering in more than one place.
    """
    cur = st.session_state.get(SK.SE_COLOR_SOURCE)
    if cur is None:
        from ui.stores import get_app_settings_store
        cur = get_app_settings_store().get("default_color_source", COLOR_SOURCE_PROGRESS)
        st.session_state[SK.SE_COLOR_SOURCE] = cur

    labels = [_t(_COLOR_SOURCE_EN[k]) for k in COLOR_SOURCE_KEYS]
    chosen = st.radio(
        _t("Chinese color mapping source") + " (中文颜色 / 中文颜色代码)",
        labels,
        index=COLOR_SOURCE_KEYS.index(cur) if cur in COLOR_SOURCE_KEYS else 0,
        horizontal=True,
        key=widget_key,
    )
    new = COLOR_SOURCE_KEYS[labels.index(chosen)]
    if new != cur:
        st.session_state[SK.SE_COLOR_SOURCE] = new
    st.caption(_COLOR_SOURCE_CAPTIONS[new])
