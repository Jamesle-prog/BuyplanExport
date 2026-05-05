import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.styles import numbers
from datetime import datetime
import fitz  # PyMuPDF
import re
import logging


# Set up logging
def setup_logging(log_file_path):
    logging.basicConfig(filename=log_file_path, level=logging.ERROR, format='%(asctime)s %(message)s')


# Define size information
SIZE_INFO = 'PS|PM|PL|PXL|XXS|XS|S|M|L|XL|1X|2X|3X'
SIZE_ORDER = SIZE_INFO.split('|')
SIZE_PATTERN = rf'({SIZE_INFO})'
PO_NUMBER_PATTERN = r'PO NUMBER\s+(\w+)'
STYLE_PATTERN = r'STYLE#\s+(.+)'
COLOR_PATTERN = r'(\w+(?:/\w+)+|\w+/\w+(?:\s\w+)?)'
UNITS_PATTERN = r'(\d+)\s+(\d{12})'
FULL_PATTERN = rf'{COLOR_PATTERN}\s+{SIZE_PATTERN}\s+{UNITS_PATTERN}'
LN_START = "LN#"
VENDOR_PATTERN = r'VENDOR\s+(\w+)'
ISSUED_BY_PATTERN = r'ISSUED BY\s+([a-zA-Z0-9.]+)'
PO_DATE_PATTERN = r'PO DATE\s+(\d{1,2}/\d{1,2}/\d{2,4})'
VEND_CNTRY_PATTERN = r'VEND CNTRY\s+(\w+(?:\s*-\s*\w+)?)'
FACTORY_PATTERN = r'FACTORY\s+(\d+)\s*-\s*([A-Z]+(?:\s[A-Z]+)*)(?=\s{2,}|$)'
HANGER_PATTERN = r'HANGER'
CNTRY_OF_ORIGIN_PATTERN = r'CNTRY OF ORIGIN\s+(\w+)'


def read_csv_files(data_path, metadata_path):
    """Read the CSV files."""

    def validate_file_path(path):
        if not os.path.exists(path):
            raise FileNotFoundError(f"File not found: {path}")

    validate_file_path(data_path)
    validate_file_path(metadata_path)

    data = pd.read_csv(data_path)
    metadata = pd.read_csv(metadata_path)
    return data, metadata


def transform_data(data):
    """Transform the data by pivoting and reordering."""
    style_data = {}
    unique_styles = data['Style'].unique()

    for style in unique_styles:
        style_specific_data = data[data['Style'] == style]
        pivot_data = style_specific_data.pivot_table(
            index=['PO Number', 'Style', 'Color'], columns='Size',
            values='Units', aggfunc='sum', fill_value=0
        )
        pivot_data = pivot_data.loc[:, (pivot_data != 0).any(axis=0)]
        pivot_data = pivot_data.reindex(columns=[size for size in SIZE_ORDER if size in pivot_data.columns])
        pivot_data['Total'] = pivot_data.sum(axis=1)

        # Add total row for each size
        total_row = pivot_data.sum(axis=0)
        total_row.name = ('Total', style, '')
        pivot_data = pd.concat([pivot_data, pd.DataFrame(total_row).T])

        style_data[style] = pivot_data

    return style_data


def get_output_file_path(output_dir, base_name, extension='.xlsx'):
    """Generate an output file path with versioning."""
    version = 1
    output_file_path = os.path.join(output_dir, base_name + extension)
    while os.path.isfile(output_file_path):
        version += 1
        output_file_path = os.path.join(output_dir, f"{base_name}_v{version}{extension}")
    return output_file_path


def adjust_column_widths(worksheet):
    """Adjust column widths to fit the content."""
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = adjusted_width


def format_table(worksheet, int_tbl_starting_row=5):
    """Format the table with black boxes, highlight total rows and columns, and add borders."""
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    total_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    total_font = Font(color="000000", bold=True)
    border_style = Side(border_style="thin", color="000000")
    border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

    max_row = worksheet.max_row

    # Find the last column of row 5
    last_col = worksheet.max_column
    for cell in worksheet[int_tbl_starting_row]:
        if cell.value:
            last_col = cell.column

    for row in worksheet.iter_rows(min_row=int_tbl_starting_row, max_row=max_row, min_col=1, max_col=last_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if cell.row == int_tbl_starting_row:  # Header row
                cell.fill = header_fill
                cell.font = header_font
            if cell.row == max_row:  # Total row
                cell.fill = total_fill
                cell.font = total_font
            if cell.row < max_row and cell.column == last_col:  # Total column
                cell.fill = total_fill
                cell.font = total_font
            if cell.row > int_tbl_starting_row and cell.row < max_row and cell.column > 3:  # Data rows
                cell.number_format = '#,##0'


def write_to_excel(style_data, metadata, output_file_path):
    """Write the transformed data to an Excel file with additional information."""
    with pd.ExcelWriter(output_file_path, engine='openpyxl') as excel_writer:
        for style, df in style_data.items():
            sheet_name = style[:31]
            df.to_excel(excel_writer, sheet_name=sheet_name, startrow=4)
            style_metadata = metadata[metadata['Style'] == style]
            factory_info = style_metadata['Factory'].values[0] if not style_metadata.empty else "N/A"
            worksheet = excel_writer.sheets[sheet_name]
            worksheet.cell(row=1, column=1, value='工厂信息：')
            worksheet.cell(row=1, column=2, value=factory_info)
            worksheet.cell(row=2, column=1, value='款号：')
            worksheet.cell(row=2, column=2, value=style)
            worksheet.cell(row=3, column=1, value='面料信息：')

            # Add file creation date to K1
            worksheet.cell(row=1, column=10, value='创建时间：')
            creation_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            worksheet.cell(row=1, column=11, value=creation_date)

            # Add and format the headers
            headers = ['PO Number', 'Style', 'Color'] + df.columns.tolist()
            for col_num, header in enumerate(headers, 1):
                cell = worksheet.cell(row=5, column=col_num, value=header)
                cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(left=Side(border_style="thin", color="000000"),
                                     right=Side(border_style="thin", color="000000"),
                                     top=Side(border_style="thin", color="000000"),
                                     bottom=Side(border_style="thin", color="000000"))

            adjust_column_widths(worksheet)
            format_table(worksheet)

    print(f'File saved to {output_file_path}')


def extract_metadata(text, file_path):
    """Extract metadata from the combined text."""

    def search_pattern(pattern, text, group_index=1):
        match = re.search(pattern, text)
        return match.group(group_index).strip() if match else None

    po_number = search_pattern(PO_NUMBER_PATTERN, text)
    style = search_pattern(STYLE_PATTERN, text)
    vendor = search_pattern(VENDOR_PATTERN, text)
    issued_by = search_pattern(ISSUED_BY_PATTERN, text)
    po_date = search_pattern(PO_DATE_PATTERN, text)
    vendor_country = search_pattern(VEND_CNTRY_PATTERN, text)
    factory_number = search_pattern(FACTORY_PATTERN, text, 1)
    factory_name = search_pattern(FACTORY_PATTERN, text, 2)
    factory = f"{factory_number} - {factory_name}" if factory_number and factory_name else None
    country_of_origin = search_pattern(CNTRY_OF_ORIGIN_PATTERN, text)
    hanger = None

    hanger_match = re.search(HANGER_PATTERN, text)
    if hanger_match:
        hanger_start = hanger_match.end()
        hanger_line = text[hanger_start:].strip().split('\n')[0]
        hanger = hanger_line.strip()

    metadata = {
        'PO Number': po_number,
        'Style': style,
        'Vendor': vendor,
        'Issued By': issued_by,
        'PO Date': po_date,
        'Vendor Country': vendor_country,
        'Factory': factory,
        'Country of Origin': country_of_origin,
        'File Path': file_path,
        'File Name': os.path.basename(file_path),
        'Hanger': hanger
    }

    return metadata


def extract_data_by_size_color(lines, metadata):
    """Extract data by size and color from the lines."""
    data_by_size_color = []
    data_summary = []
    start_searching = False
    current_color = None

    for line in lines:
        if line.startswith(LN_START):
            start_searching = True
            continue

        if start_searching:
            match = re.search(FULL_PATTERN, line)
            if match:
                color = match.group(1)
                size = match.group(2)
                units = int(match.group(3))
                upc = match.group(4)
                data_by_size_color.append([metadata['PO Number'], metadata['Style'], color, size, units, upc])
                current_color = color
            else:
                if "TTL" in line:
                    hanger = line.split("TTL")[0].strip()
                    data_summary.append([metadata['PO Number'], metadata['Style'], current_color, hanger])

                size_match = re.search(SIZE_PATTERN, line)
                if size_match:
                    size = size_match.group(1)
                    units_match = re.search(UNITS_PATTERN, line)
                    if units_match:
                        units = int(units_match.group(1))
                        upc = units_match.group(2)
                        data_by_size_color.append(
                            [metadata['PO Number'], metadata['Style'], current_color, size, units, upc])
    return data_by_size_color, data_summary


def extract_data_from_pdf(pdf_path):
    """Extract data from the given PDF file."""
    try:
        pdf_document = fitz.open(pdf_path)
    except Exception as e:
        logging.error(f"Error opening PDF file {pdf_path}: {e}")
        return [], [], {}

    combined_text = ""
    for page_number in range(len(pdf_document)):
        try:
            page = pdf_document.load_page(page_number)
            text = page.get_text()
            combined_text += text + "\n"
        except Exception as e:
            logging.error(f"Error reading page {page_number} of {pdf_path}: {e}")
            continue

    metadata = extract_metadata(combined_text, pdf_path)
    lines = combined_text.splitlines()
    data_by_size_color, data_summary = extract_data_by_size_color(lines, metadata)

    return data_by_size_color, data_summary, metadata


def scan_folders_for_pdfs(folder_paths):
    """Scan folders for PDF files."""
    pdf_files = []
    for folder_path in folder_paths:
        for root, dirs, files in os.walk(folder_path):
            for file in files:
                if file.lower().endswith('.pdf'):
                    pdf_files.append(os.path.join(root, file))
    return pdf_files


def get_next_versioned_filename(output_dir, base_name, extension):
    """Get the next versioned filename."""
    version = 1
    output_file_path = os.path.join(output_dir, base_name + extension)
    while os.path.isfile(output_file_path):
        version += 1
        output_file_path = os.path.join(output_dir, f"{base_name}_v{version}{extension}")
    return output_file_path


def main(data=None, metadata=None, output_dir='D:\\Users\\Desktop\\data',
         base_output_file_name='transformed_data_by_style_filtered_with_totals_and_metadata',
         data_file_path='D:\\Users\\Desktop\\data\\all_extracted_data_by_size_color_v1.csv',
         metadata_file_path='D:\\Users\\Desktop\\data\\all_extracted_metadata_v1.csv'):
    """Main function to process data and write to Excel.

    Args:
        data (pd.DataFrame, optional): DataFrame containing the data. Defaults to None.
        metadata (pd.DataFrame, optional): DataFrame containing the metadata. Defaults to None.
        output_dir (str, optional): Directory for the output Excel file. Defaults to 'D:\\Users\\Desktop\\data'.
        base_output_file_name (str, optional): Base name for the output Excel file. Defaults to 'transformed_data_by_style_filtered_with_totals_and_metadata'.
        data_file_path (str, optional): Path to the data CSV file. Defaults to 'D:\\Users\\Desktop\\data\\all_extracted_data_by_size_color_v1.csv'.
        metadata_file_path (str, optional): Path to the metadata CSV file. Defaults to 'D:\\Users\\Desktop\\data\\all_extracted_metadata_v1.csv'.
    """

    if data is None or metadata is None:
        data, metadata = read_csv_files(data_file_path, metadata_file_path)

    # Transform the data
    style_data = transform_data(data)

    # Generate the output file path
    output_file_path = get_output_file_path(output_dir, base_output_file_name)

    # Write the transformed data to an Excel file
    write_to_excel(style_data, metadata, output_file_path)


if __name__ == "__main__":
    # Parameters
    folder_paths = [

        "E:\\新庄源订单资料\\2024\\2.大货\\12.AMS\\Fall 24\\PO"
    ]

    #"D:/Users/Desktop/data",
    #"D:/公司网盘/新庄源订单资料/2024/2.大货/1.CBCJ/Q3/3-44532/1.PO_BuyPlan",
    #"D:\\公司网盘\\新庄源订单资料\\2024\\2.大货\\3.DKNY\\DKNY SPORTSWEAR\\Fall 24\\P4HHEXTF P4HHCXTF\\1.PO_BuyPlan",
    #"D:\\公司网盘\\新庄源订单资料\\2024\\2.大货\\3.DKNY\\DKNY SPORTSWEAR\\Fall 24\\P4HHDXTG P4HHCXTG\\1.PO_BuyPlan",


    log_file_path = "E:\\新庄源订单资料\\2024\\2.大货\\12.AMS\\Fall 24\\PO\\pdf_extraction_errors.log"
    output_dir = "E:\\新庄源订单资料\\2024\\2.大货\\12.AMS\\Fall 24\\PO"
    base_filename_by_size_color = "all_extracted_data_by_size_color"
    base_filename_summary = "all_extracted_data_summary"
    base_filename_metadata = "all_extracted_metadata"
    base_output_file_name = 'transformed_data_by_style_filtered_with_totals_and_metadata'
    extension = ".csv"

    # Setup logging
    setup_logging(log_file_path)

    # Scan the folders for PDF files
    pdf_files = scan_folders_for_pdfs(folder_paths)

    # Initialize an empty list to store all data
    all_data_by_size_color = []
    all_data_summary = []
    all_metadata = []

    # Process each PDF file and append data to the list
    for pdf_file in pdf_files:
        extracted_data_by_size_color, extracted_data_summary, metadata = extract_data_from_pdf(pdf_file)
        if extracted_data_by_size_color:
            all_data_by_size_color.extend(extracted_data_by_size_color)
        if extracted_data_summary:
            all_data_summary.extend(extracted_data_summary)
        if metadata:
            all_metadata.append(metadata)

    # Convert the extracted data into pandas DataFrames
    df_by_size_color = pd.DataFrame(all_data_by_size_color,
                                    columns=['PO Number', 'Style', 'Color', 'Size', 'Units', 'UPC'])
    df_summary = pd.DataFrame(all_data_summary, columns=['PO Number', 'Style', 'Color', 'Hanger'])
    df_metadata = pd.DataFrame(all_metadata)

    # Get the next versioned filenames
    output_csv_path_by_size_color = get_next_versioned_filename(output_dir, base_filename_by_size_color, extension)
    output_csv_path_summary = get_next_versioned_filename(output_dir, base_filename_summary, extension)
    output_csv_path_metadata = get_next_versioned_filename(output_dir, base_filename_metadata, extension)

    # Save the DataFrames to versioned CSV files
    df_by_size_color.to_csv(output_csv_path_by_size_color, index=False)
    df_summary.to_csv(output_csv_path_summary, index=False)
    df_metadata.to_csv(output_csv_path_metadata, index=False)

    print(f"Data by Size and Color saved to: {output_csv_path_by_size_color}")
    print(f"Summary Data saved to: {output_csv_path_summary}")
    print(f"Metadata saved to: {output_csv_path_metadata}")

    # Call the main function with the DataFrames and parameters
    main(data=df_by_size_color, metadata=df_metadata, output_dir=output_dir,
         base_output_file_name=base_output_file_name)
