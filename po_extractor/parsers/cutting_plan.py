"""Parser for externally-generated cutting plans (Optitex Cut Plan export).

The cut plan itself is produced *outside* this app by the marker/nesting
software (Optitex — ``CpAlgo.dll``): marker efficiency, marker length and ply
counts all come from the real ``.mrk`` markers, so nothing here recomputes
them.  This module only *reads* the workbook the operator hands over, so the
plan can be linked to the PO(s) it covers and re-emitted in one standard
layout (see ``po_extractor.exporters.cutting_plan_export``).

Workbook layout (one sheet, named after the run timestamp)::

    Date / Time / Order name / Style file N / Style name N /
    Cut plan operator / Client / Output folder path      <- key-value header
    Order demands        <- colours x (style, size) matrix, + Sum row
    ... then, repeated once per material ...
    Marker Definition    <- material, width, spreading, ply limits
    Marker Ratio         <- one row per marker: file + pieces of each size
    Spreading Plies      <- per marker: plies, pieces, fabric length, efficiency
    Solution             <- achieved qty per colour/size + fabric length
    Total Efficiency / Total Cost / Total cut length / Total Tables /
    Average Length
    Total Tables         <- grand total across all materials (last row)

Nothing about the row offsets is fixed: block anchors are found by their
label text and the metric columns by their header text, because the size
columns (and therefore everything to their right) shift with the number of
styles in the plan.
"""
from __future__ import annotations

from typing import Any

# Header labels in the key/value block at the top of the sheet.
_H_DATE      = "date"
_H_TIME      = "time"
_H_ORDER     = "order name"
_H_OPERATOR  = "cut plan operator"
_H_CLIENT    = "client"
_H_FOLDER    = "output folder path"

# Block anchors (matched case-insensitively against column A/B).
_B_DEMANDS   = "order demands"
_B_MARKERDEF = "marker definition"
_B_RATIO     = "marker ratio"
_B_SPREAD    = "spreading plies"
_B_SOLUTION  = "solution"

# Metric column headers inside the Spreading Plies / Solution blocks.
_M_FABRIC_LEN   = "fabric length,m"
_M_EFFICIENCY   = "efficiency,%"
_M_CUT_LEN      = "cut length,m"
_M_MARKER_LEN   = "marker length,cm"
_M_COST         = "material cost,cny"
_M_TOTAL_QTY    = "total quantity"

_T_EFFICIENCY   = "total efficiency"
_T_COST         = "total cost"
_T_CUT_LENGTH   = "total cut length"
_T_TABLES       = "total tables"
_T_AVG_LENGTH   = "average length"


class CuttingPlanParseError(ValueError):
    """Raised when a workbook doesn't look like a cut plan export at all."""


# ---------------------------------------------------------------------------
# Small grid helpers
# ---------------------------------------------------------------------------

def _norm(v: Any) -> str:
    """Normalised text of a cell — trimmed, lower-cased, whitespace collapsed."""
    if v is None:
        return ""
    return " ".join(str(v).split()).strip().lower()


def _txt(v: Any) -> str:
    """Display text of a cell — trimmed, original case, '' for None."""
    if v is None:
        return ""
    return str(v).strip()


def _num(v: Any) -> float | None:
    """Float value of a cell, or None when it isn't numeric."""
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _int(v: Any) -> int | None:
    f = _num(v)
    return None if f is None else int(round(f))


def _cell(grid: list[list[Any]], r: int, c: int) -> Any:
    """grid[r][c] with bounds tolerance — out-of-range reads give None."""
    if r < 0 or r >= len(grid):
        return None
    row = grid[r]
    if c < 0 or c >= len(row):
        return None
    return row[c]


def _row_is_blank(grid: list[list[Any]], r: int) -> bool:
    if r < 0 or r >= len(grid):
        return True
    return all(v is None or str(v).strip() == "" for v in grid[r])


def _find_label(grid: list[list[Any]], label: str, *,
                start: int = 0, end: int | None = None,
                max_col: int = 3) -> int:
    """Return the first row at/after *start* whose first ``max_col`` cells hold
    *label* (case-insensitive), or -1.  Restricting the column window keeps a
    style name that happens to read "Solution" from anchoring a block."""
    end = len(grid) if end is None else min(end, len(grid))
    target = _norm(label)
    for r in range(max(start, 0), end):
        for c in range(min(max_col, len(grid[r]))):
            if _norm(grid[r][c]) == target:
                return r
    return -1


def _find_all_labels(grid: list[list[Any]], label: str, *,
                     max_col: int = 3) -> list[int]:
    rows: list[int] = []
    r = _find_label(grid, label, max_col=max_col)
    while r >= 0:
        rows.append(r)
        r = _find_label(grid, label, start=r + 1, max_col=max_col)
    return rows


def _metric_cols(grid: list[list[Any]], r: int) -> dict[str, int]:
    """Map metric header text → column index for the header row *r*.

    Metric columns sit to the right of the size columns, so their position
    depends on how many styles the plan covers — always resolve by name.
    """
    out: dict[str, int] = {}
    for c, v in enumerate(grid[r] if r < len(grid) else []):
        key = _norm(v)
        if key in (_M_FABRIC_LEN, _M_EFFICIENCY, _M_CUT_LEN, _M_MARKER_LEN,
                   _M_COST, _M_TOTAL_QTY):
            out[key] = c
    return out


def _size_columns(grid: list[list[Any]], style_row: int, size_row: int,
                  default_style: str = "",
                  first_col: int = 0) -> list[tuple[int, str, str]]:
    """Resolve the (column, style, size) triples of a size matrix.

    *style_row* carries a style name above the first column of each group
    (the export merges it across the group), *size_row* carries one size per
    column.  A style name therefore owns every column up to the next name.

    A single-style plan has no band at all — *style_row* is then ``-1`` and
    every column belongs to *default_style*, which the caller takes from the
    header block (``Style name``).

    *first_col* is the column the ``Sizes`` heading sits in; nothing left of
    it is a size.  With a band, the block's own labels (``Colors``, ``Plies
    number``) sit on the band row and the size row is empty under them — but
    a single-style block has no band, so those labels share the size row, and
    without this bound ``Plies number`` was read as a size and its ply count
    as a garment quantity.
    """
    out: list[tuple[int, str, str]] = []
    current_style = default_style
    width = max(len(_row_at(grid, style_row)), len(_row_at(grid, size_row)))
    for c in range(max(first_col, 0), width):
        name = _txt(_cell(grid, style_row, c))
        if name:
            current_style = name
        size = _txt(_cell(grid, size_row, c))
        if size:
            out.append((c, current_style, size))
    return out


def _row_at(grid: list[list[Any]], r: int) -> list[Any]:
    return grid[r] if 0 <= r < len(grid) else []


# Text that can sit in a block's label column on a *header* row — anything
# else there is the first data row.  Used to find where the header ends.
_HEADER_LABELS = frozenset({
    "colors", "color", "plies number", "quantity", "sizes", "size",
    "marker", "file name",
})


def _first_data_row(grid: list[list[Any]], sizes_row: int, label_col: int,
                    *, window: int = 3) -> int:
    """Row where the block's data starts, or -1 when it can't be told.

    Found by the label column: every header row leaves it empty or carries a
    fixed heading, and the first row that puts anything else there is the
    first colour/marker. Reading the geometry off this rather than off a fixed
    offset is what makes a single-style plan work — see :class:`_Matrix`.
    """
    for r in range(sizes_row + 1, min(sizes_row + 1 + window, len(grid))):
        if _row_is_blank(grid, r):
            break
        label = _norm(_cell(grid, r, label_col))
        if label and label not in _HEADER_LABELS:
            return r
    return -1


class _Matrix:
    """The header geometry shared by every colour x (style, size) block.

    Each block is anchored by the cell reading ``Sizes``.  Below it come the
    size labels, and — **only when the plan covers more than one style** — a
    band of style names between the two, merged across each style's group of
    sizes.  Then the data rows.

    That band is not always there, and assuming it was is what made a
    single-style plan wrong in a way nothing flagged: every size label was
    read as a style name and the first data row as the sizes, which left the
    plan with no style totals at all (Plan qty 0) and its quantities filed
    under size names, so Cut qty came out as the sum of two mis-keyed buckets
    rather than the units cut.  The rows are therefore located from where the
    data actually starts, not from a fixed offset.

    Only the *label* columns move around — ``Colors`` sits on the Sizes row in
    the Order demands / Solution blocks but on the style row in Spreading
    Plies — so both rows are searched for them.  Metric headers (``Fabric
    Length,m`` …) always live on the Sizes row.
    """

    __slots__ = ("sizes_row", "style_row", "size_row", "cols", "metrics",
                 "label_col", "plies_col", "kind_col", "first_data_row")

    def __init__(self, grid: list[list[Any]], sizes_row: int,
                 default_style: str = ""):
        self.sizes_row = sizes_row
        # Provisional geometry — the label columns are read off these, and
        # they are then corrected below once the data row is known.
        self.style_row = sizes_row + 1
        self.size_row = sizes_row + 2
        self.first_data_row = sizes_row + 3
        self.label_col = self._col(grid, "colors", default=0)

        found = _first_data_row(grid, sizes_row, self.label_col)
        if found > sizes_row + 1:
            self.first_data_row = found
            self.size_row = found - 1
            # No room for a band → one style, and its name is in the header
            # block rather than over the sizes.
            self.style_row = found - 2 if found - 2 > sizes_row else -1

        self.cols = _size_columns(grid, self.style_row, self.size_row,
                                  default_style,
                                  self._col(grid, "sizes", default=0))
        self.metrics = _metric_cols(grid, sizes_row)
        self.plies_col = self._col(grid, "plies number", default=-1)
        # Present only when the plan couldn't hit the ordered quantity exactly:
        # the Solution block then splits each colour into 'Order' and 'Real'
        # sub-rows under a 'Quantity' column.
        self.kind_col = self._col(grid, "quantity", default=-1)

    def _col(self, grid: list[list[Any]], label: str, *, default: int) -> int:
        # The size row is searched too: with no band it is where ``Colors``
        # and ``Plies number`` end up, and skipping it read every marker's
        # ply count as zero.  ``_row_at`` tolerates the -1 of an absent band.
        target = _norm(label)
        for r in (self.sizes_row, self.style_row, self.size_row):
            for c, v in enumerate(_row_at(grid, r)):
                if _norm(v) == target:
                    return c
        return default

    def metric(self, key: str) -> int:
        return self.metrics.get(key, -1)


def _find_matrix(grid: list[list[Any]], start: int, end: int,
                 *, within: int = 4,
                 default_style: str = "") -> _Matrix | None:
    """Locate the matrix header within *within* rows of *start*."""
    row = _find_label(grid, "sizes", start=start,
                      end=min(start + within, end), max_col=6)
    if row < 0:
        return None
    m = _Matrix(grid, row, default_style)
    return m if m.cols else None


def _groups_from_cols(cols: list[tuple[int, str, str]]
                      ) -> list[tuple[str, list[str]]]:
    """Column order of a matrix as ``[(style, [size, ...]), ...]``.

    The header row lists every style/size the block uses, left to right, so
    this is the block's own ordering — the only reliable one.  Individual data
    rows can't supply it: a marker holds only some of the sizes.
    """
    groups: dict[str, list[str]] = {}
    for _c, style, size in cols:
        groups.setdefault(style, [])
        if size not in groups[style]:
            groups[style].append(size)
    return list(groups.items())


def _block_groups(grid: list[list[Any]], start: int, end: int,
                  default_style: str = "") -> list[tuple[str, list[str]]]:
    """Size groups of the block starting at *start*, or [] when not found."""
    m = _find_matrix(grid, start, end, default_style=default_style)
    return _groups_from_cols(m.cols) if m else []


def _matrix_rows(grid: list[list[Any]], first_row: int, label_col: int,
                 cols: list[tuple[int, str, str]], *,
                 stop_labels: tuple[str, ...] = ("sum",),
                 max_rows: int = 200) -> list[tuple[str, int]]:
    """Yield (label, row_index) for the data rows of a matrix block.

    Walks down from *first_row* until a blank row, a stop label ("Sum"), or a
    row whose label cell is empty.
    """
    out: list[tuple[str, int]] = []
    r = first_row
    limit = min(first_row + max_rows, len(grid))
    while r < limit:
        if _row_is_blank(grid, r):
            break
        label = _txt(_cell(grid, r, label_col))
        if _norm(label) in stop_labels:
            break
        if not label:
            # No label but real numbers to the right → keep scanning a row;
            # a fully empty label column means the block has ended.
            if not any(_num(_cell(grid, r, c)) is not None for c, _, _ in cols):
                break
        else:
            out.append((label, r))
        r += 1
    return out


# ---------------------------------------------------------------------------
# Block parsers
# ---------------------------------------------------------------------------

def _style_slot(key: str, prefix: str) -> int:
    """Which style a ``Style file N`` / ``Style name N`` header line belongs to.

    A plan covering one style drops the number and writes plain ``Style file``
    / ``Style name``. Numbering those by arrival order put the file and the
    name in two *different* slots, so a single-style plan came out as two
    styles — one with only a filename, one with only a name.
    """
    return _int(key.replace(prefix, "").strip()) or 1


def _parse_header(grid: list[list[Any]]) -> dict[str, Any]:
    """Key/value header block — labels in col A (or B), values one col right."""
    head: dict[str, Any] = {
        "plan_date": "", "plan_time": "", "order_name": "",
        "operator": "", "client": "", "output_folder": "",
        "styles": [],
    }
    styles: dict[int, dict[str, str]] = {}
    scan_to = min(len(grid), 40)
    for r in range(scan_to):
        for c in range(min(2, len(grid[r]))):
            key = _norm(_cell(grid, r, c))
            if not key:
                continue
            val = _txt(_cell(grid, r, c + 1))
            if key == _H_DATE:
                head["plan_date"] = val
            elif key == _H_TIME:
                head["plan_time"] = val
            elif key == _H_ORDER:
                head["order_name"] = val
            elif key == _H_OPERATOR:
                head["operator"] = val
            elif key == _H_CLIENT:
                head["client"] = val
            elif key == _H_FOLDER:
                head["output_folder"] = val
            elif key.startswith("style name"):
                styles.setdefault(_style_slot(key, "style name"), {})["name"] = val
            elif key.startswith("style file"):
                styles.setdefault(_style_slot(key, "style file"), {})["file"] = val
    # "Time" is written as a real time value; normalise to HH:MM.
    raw_time = _cell_time(grid)
    if raw_time:
        head["plan_time"] = raw_time
    head["styles"] = [
        {"name": styles[i].get("name", ""), "file": styles[i].get("file", "")}
        for i in sorted(styles)
        if styles[i].get("name") or styles[i].get("file")
    ]
    return head


def _cell_time(grid: list[list[Any]]) -> str:
    """Find the value next to a 'Time' label and render it as HH:MM."""
    for r in range(min(len(grid), 20)):
        for c in range(min(len(grid[r]), 8)):
            if _norm(grid[r][c]) == _H_TIME:
                v = _cell(grid, r, c + 1)
                if v is None:
                    return ""
                if hasattr(v, "strftime"):
                    return v.strftime("%H:%M")
                return _txt(v)
    return ""


def _parse_demands(grid: list[list[Any]],
                   default_style: str = "") -> list[dict[str, Any]]:
    """Order demands block → [{style, color, size, qty}]."""
    anchor = _find_label(grid, _B_DEMANDS, max_col=4)
    if anchor < 0:
        return []
    m = _find_matrix(grid, anchor + 1, len(grid), within=6,
                     default_style=default_style)
    if m is None:
        return []
    out: list[dict[str, Any]] = []
    for color, r in _matrix_rows(grid, m.first_data_row, m.label_col, m.cols):
        for c, style, size in m.cols:
            qty = _int(_cell(grid, r, c)) or 0
            if qty:
                out.append({"style": style, "color": color,
                            "size": size, "qty": qty})
    return out


def _parse_marker_definition(grid: list[list[Any]], anchor: int) -> dict[str, Any]:
    """Marker Definition block → material attributes.

    The header spans two merged rows, so the data row is not always
    ``anchor + 2``; scan forward for the first row with a material name.
    """
    hdr = anchor + 1
    keys = {_norm(v): c for c, v in enumerate(_row_at(grid, hdr)) if _norm(v)}

    def col(name: str) -> int | None:
        return keys.get(_norm(name))

    data = -1
    for r in range(hdr + 1, min(hdr + 6, len(grid))):
        if _txt(_cell(grid, r, 0)) and _norm(_cell(grid, r, 0)) != "material":
            data = r
            break
    if data < 0:
        return {}

    def val(name: str) -> Any:
        c = col(name)
        return None if c is None else _cell(grid, data, c)

    return {
        "material":     _txt(_cell(grid, data, 0)),
        "n_markers":    _int(val("N Markers")),
        "spreading":    _txt(val("Spreading")),
        "width_cm":     _num(val("Width, cm")),
        "length_cm":    _num(val("Length, cm")),
        "min_plies":    _int(val("Min Plies")),
        "max_plies":    _int(val("Max Plies")),
        "waste_limits": _txt(val("Waste Limits, cm")),
    }


def _parse_marker_ratio(grid: list[list[Any]], anchor: int, end: int,
                        default_style: str = "") -> list[dict[str, Any]]:
    """Marker Ratio block → one entry per marker with its size ratio."""
    m = _find_matrix(grid, anchor, end, default_style=default_style)
    if m is None:
        return []
    out: list[dict[str, Any]] = []
    r = m.first_data_row
    while r < end and not _row_is_blank(grid, r):
        marker_no = _int(_cell(grid, r, 0))
        if marker_no is None:
            break
        ratio = []
        for c, style, size in m.cols:
            qty = _int(_cell(grid, r, c)) or 0
            if qty:
                ratio.append({"style": style, "size": size, "qty": qty})
        out.append({
            "marker_no": marker_no,
            "file_name": _txt(_cell(grid, r, 1)),
            "ratio": ratio,
        })
        r += 1
    return out


def _parse_spreading(grid: list[list[Any]], anchor: int, end: int,
                     default_style: str = "") -> list[dict[str, Any]]:
    """Spreading Plies block → one entry per marker with its ply rows."""
    out: list[dict[str, Any]] = []
    r = anchor + 1
    while r < end:
        if _norm(_cell(grid, r, 0)).startswith("marker "):
            out.append(_parse_spread_marker(grid, r, end, default_style))
            r += 1
            continue
        r += 1
    return [m for m in out if m]


def _parse_spread_marker(grid: list[list[Any]], row: int, end: int,
                         default_style: str = "") -> dict[str, Any]:
    marker_no = _int(_norm(_cell(grid, row, 0)).replace("marker", "").strip())
    # The marker row *is* the matrix header here — 'Sizes' and the metric
    # headers share it with the marker number and file name.
    m = _find_matrix(grid, row, end, within=1, default_style=default_style)
    if m is None:
        return {}
    rows: list[dict[str, Any]] = []
    for color, r in _matrix_rows(grid, m.first_data_row, m.label_col, m.cols):
        pieces = []
        for c, style, size in m.cols:
            qty = _int(_cell(grid, r, c)) or 0
            if qty:
                pieces.append({"style": style, "size": size, "qty": qty})
        rows.append({
            "color": color,
            "plies": _int(_cell(grid, r, m.plies_col)) or 0,
            "pieces": pieces,
            "fabric_length_m":  _num(_cell(grid, r, m.metric(_M_FABRIC_LEN))),
            "efficiency_pct":   _num(_cell(grid, r, m.metric(_M_EFFICIENCY))),
            "cut_length_m":     _num(_cell(grid, r, m.metric(_M_CUT_LEN))),
            "marker_length_cm": _num(_cell(grid, r, m.metric(_M_MARKER_LEN))),
            "cost":             _num(_cell(grid, r, m.metric(_M_COST))),
        })
    return {
        "marker_no": marker_no,
        "file_name": _txt(_cell(grid, row, 1)),
        "rows": rows,
    }


def _parse_solution(grid: list[list[Any]], anchor: int, end: int,
                    default_style: str = "") -> list[dict[str, Any]]:
    """Solution block → achieved qty per colour, with fabric length."""
    m = _find_matrix(grid, anchor + 1, end, default_style=default_style)
    if m is None:
        return []
    out: list[dict[str, Any]] = []
    color = ""
    r = m.first_data_row
    while r < min(end, len(grid)) and not _row_is_blank(grid, r):
        label = _txt(_cell(grid, r, m.label_col))
        if _norm(label) == "sum":
            break
        if label:
            color = label
        # 'Order' / 'Real' sub-rows repeat the colour only on the first row.
        kind = _norm(_cell(grid, r, m.kind_col)) if m.kind_col >= 0 else ""
        cells = []
        for c, style, size in m.cols:
            qty = _int(_cell(grid, r, c)) or 0
            if qty:
                cells.append({"style": style, "size": size, "qty": qty})
        if not cells and not color:
            break
        out.append({
            "color": color,
            "kind": kind,
            "cells": cells,
            "total_qty": _int(_cell(grid, r, m.metric(_M_TOTAL_QTY))) or 0,
            "fabric_length_m": _num(_cell(grid, r, m.metric(_M_FABRIC_LEN))),
        })
        r += 1
    return out


def achieved_rows(solution: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """The solution rows that describe what will actually be cut.

    When the plan splits into 'Order' / 'Real' sub-rows, 'Real' is the cut
    quantity (it can exceed the order — a marker ratio rarely divides the
    demand exactly).  Plans that hit the demand exactly have neither label
    and every row counts.
    """
    real = [s for s in solution if s.get("kind") == "real"]
    if real:
        return real
    return [s for s in solution if s.get("kind") != "order"]


def _parse_totals(grid: list[list[Any]], start: int, end: int) -> dict[str, Any]:
    """The five per-material total lines between *start* and *end*.

    First occurrence wins: the grand-total "Total Tables" line at the very
    bottom of the sheet falls inside the last material's row range and would
    otherwise overwrite that material's own table count.
    """
    out: dict[str, Any] = {}
    labels = {
        _T_EFFICIENCY: ("total_efficiency_pct", _num),
        _T_COST:       ("total_cost", _num),
        _T_CUT_LENGTH: ("total_cut_length_m", _num),
        _T_TABLES:     ("total_tables", _int),
        _T_AVG_LENGTH: ("average_length_m", _num),
    }
    for r in range(start, min(end, len(grid))):
        key = _norm(_cell(grid, r, 0))
        if key in labels:
            field, conv = labels[key]
            if field not in out:
                out[field] = conv(_cell(grid, r, 1))
    return out


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def parse_cut_plan_grid(grid: list[list[Any]]) -> dict[str, Any]:
    """Parse an already-materialised sheet grid.  See :func:`parse_cut_plan`."""
    if not grid:
        raise CuttingPlanParseError("The worksheet is empty.")

    plan: dict[str, Any] = _parse_header(grid)
    plan["format"] = "optitex"
    # A plan covering one style names it only in the header block — there is
    # no style band over the size columns to read it from.
    one_style = (plan["styles"][0].get("name", "")
                 if len(plan["styles"]) == 1 else "")
    plan["demands"] = _parse_demands(grid, one_style)

    def_rows = _find_all_labels(grid, _B_MARKERDEF, max_col=2)
    if not def_rows and not plan["demands"]:
        raise CuttingPlanParseError(
            "This file doesn't look like a cut plan export — no 'Order demands' "
            "or 'Marker Definition' block was found."
        )

    materials: list[dict[str, Any]] = []
    for i, anchor in enumerate(def_rows):
        end = def_rows[i + 1] if i + 1 < len(def_rows) else len(grid)
        mat = _parse_marker_definition(grid, anchor)
        ratio_at = _find_label(grid, _B_RATIO, start=anchor, end=end, max_col=2)
        spread_at = _find_label(grid, _B_SPREAD, start=anchor, end=end, max_col=2)
        sol_at = _find_label(grid, _B_SOLUTION, start=anchor, end=end, max_col=2)
        mat["markers"] = (
            _parse_marker_ratio(grid, ratio_at + 1,
                                spread_at if spread_at > 0 else end,
                                one_style)
            if ratio_at >= 0 else []
        )
        mat["spreads"] = (
            _parse_spreading(grid, spread_at,
                             sol_at if sol_at > 0 else end, one_style)
            if spread_at >= 0 else []
        )
        mat["solution"] = (
            _parse_solution(grid, sol_at, end, one_style) if sol_at >= 0 else []
        )
        # The material's own column order, kept so it can be re-emitted
        # faithfully — a lining is often cut for only some of the styles or
        # sizes, so it can't be derived from the order-demands matrix.
        mat["groups"] = (
            (_block_groups(grid, ratio_at + 1, end, one_style)
             if ratio_at >= 0 else [])
            or (_block_groups(grid, sol_at + 1, end, one_style)
                if sol_at >= 0 else [])
            or (_block_groups(grid, spread_at + 1, end, one_style)
                if spread_at >= 0 else [])
        )
        mat.update(_parse_totals(grid, sol_at if sol_at >= 0 else anchor, end))
        real = achieved_rows(mat["solution"])
        mat["cut_qty"] = sum(int(s.get("total_qty") or 0) for s in real)
        mat["fabric_length_m"] = sum(
            float(s.get("fabric_length_m") or 0) for s in real
        ) or None
        materials.append(mat)
    plan["materials"] = materials

    # The very last "Total Tables" line sits outside every material block and
    # is the grand total; per-material totals were already consumed above.
    table_rows = _find_all_labels(grid, _T_TABLES, max_col=1)
    grand = None
    if table_rows:
        last = table_rows[-1]
        if not def_rows or last > (def_rows[-1] if def_rows else -1):
            in_material = any(
                m.get("total_tables") is not None for m in materials
            )
            # Only treat it as the grand total when it isn't the single
            # material's own total line (one-material plans repeat neither).
            if len(table_rows) > len(materials) or not in_material:
                grand = _int(_cell(grid, last, 1))
    if grand is None:
        grand = sum(int(m.get("total_tables") or 0) for m in materials) or None
    plan["total_tables"] = grand

    # Per-style demand totals: the export repeats the same garment count under
    # every style group (a garment cut from two style files is still one
    # garment), so the order size is the per-style total, not their sum.
    per_style: dict[str, int] = {}
    for d in plan["demands"]:
        per_style[d["style"]] = per_style.get(d["style"], 0) + int(d["qty"])
    plan["style_totals"] = per_style
    plan["total_qty"] = max(per_style.values()) if per_style else 0
    plan["colors"] = sorted({d["color"] for d in plan["demands"]})
    if not plan["styles"]:
        plan["styles"] = [{"name": s, "file": ""} for s in sorted(per_style)]
    return plan


def parse_cut_plan(source: Any, *, sheet: str | None = None) -> dict[str, Any]:
    """Parse a cut plan workbook into a structured dict.

    *source* is anything ``openpyxl.load_workbook`` accepts — a path, or a
    file-like object (e.g. a Streamlit ``UploadedFile`` / ``BytesIO``).
    *sheet* selects a worksheet by name; the first sheet is used by default
    (the export names its only sheet after the run timestamp).
    """
    from openpyxl import load_workbook

    try:
        wb = load_workbook(source, data_only=True, read_only=True)
    except Exception as exc:                       # noqa: BLE001 - user file
        raise CuttingPlanParseError(
            f"Could not open the workbook: {exc}"
        ) from exc
    try:
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.worksheets[0]
        grid = [list(row) for row in ws.iter_rows(values_only=True)]
        sheet_name = ws.title
    finally:
        wb.close()

    plan = parse_cut_plan_grid(grid)
    plan["sheet_name"] = sheet_name
    return plan
