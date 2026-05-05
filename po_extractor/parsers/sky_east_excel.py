"""Parser for Sky East purchase contract Excel files.

File layout:
  Row 3:  PC NO. (col B=2, value col E=5)
  Row 4:  Date
  Row 5:  Buyer (col E)
  Row 7:  Seller (col E)
  Row 9:  Currency (col E)
  Row 10: Payment Method (col E)
  Row 13: Trade Term (col E)
  Row 16: Column headers
  Row 17+: Data rows (brand sub-section headers + item rows + totals row)
"""
import hashlib
import os
import re
from datetime import datetime
from pathlib import Path
import openpyxl

from ..models.sky_east_data import SkyEastContract, SkyEastItem

PARSER_VERSION = "1.0.0"

# Regex to extract DISPIMG picture ID from Excel formula string
_DISPIMG_RE = re.compile(r'DISPIMG\("(ID_[A-F0-9]+)"', re.IGNORECASE)

# Expected size column names in order
_SIZE_LABELS = ["XS", "S", "M", "L", "XL", "2XL"]


def _file_hash(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _clean(value) -> str | None:
    """Return stripped string or None for blank/None values."""
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _to_int(value) -> int:
    try:
        return int(value) if value is not None else 0
    except (ValueError, TypeError):
        return 0


def _to_float(value) -> float:
    try:
        return float(value) if value is not None else 0.0
    except (ValueError, TypeError):
        return 0.0


def _format_date(value) -> str | None:
    """Convert a cell value to ISO date string, handling both datetime and text."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    return s if s else None


def _extract_picture_id(cell_value) -> str | None:
    """Extract the DISPIMG ID string from an Excel formula value."""
    if cell_value is None:
        return None
    m = _DISPIMG_RE.search(str(cell_value))
    return m.group(1) if m else None


def _find_header_row(ws, search_from: int = 14, search_to: int = 20) -> int | None:
    """Find the row containing 'Style No.' or 'PO NUMBER' in any cell."""
    for row_num in range(search_from, search_to + 1):
        for col_num in range(1, 30):
            val = ws.cell(row=row_num, column=col_num).value
            if val and str(val).strip() in ("Style No.", "PO NUMBER"):
                return row_num
    return None


def _map_size_columns(ws, header_row: int) -> dict:
    """Return {size_label: col_num} for size columns found in the header row."""
    mapping = {}
    for col_num in range(1, 30):
        val = ws.cell(row=header_row, column=col_num).value
        if val:
            label = str(val).strip()
            if label in _SIZE_LABELS:
                mapping[label] = col_num
    return mapping


def _is_brand_section_row(row_values: dict) -> str | None:
    """
    Return the brand name if this row is a brand sub-section header.
    Criteria: col 1 has a non-numeric string, col 4 (Style No.) is empty.
    """
    item_val = row_values.get(1)
    style_val = row_values.get(4)
    if item_val is None or style_val is not None:
        return None
    s = str(item_val).strip()
    if s and not s.replace(".", "").isdigit():
        return s
    return None


def _is_totals_row(row_values: dict) -> bool:
    """Return True if this is a totals/summary row (no style, 'Total' text, or only numeric sum)."""
    style_val = row_values.get(4)
    if style_val is not None:
        return False
    # Check for 'Total' keyword anywhere in the row
    for v in row_values.values():
        if v is not None and str(v).strip().lower() == "total":
            return True
    # Row with no item number and no style is also a totals/summary row
    item_val = row_values.get(1)
    if item_val is None:
        return True
    return False


def parse_sky_east(file_path: str) -> SkyEastContract:
    """
    Parse a Sky East purchase contract Excel file and return a SkyEastContract.

    Supports .xlsx and .xlsm files. Uses data_only=True so formula results
    are read where available (requires the file was last saved with calculated values).
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    try:
        ws = wb.active
        contract = _parse_worksheet(ws, path)
    finally:
        wb.close()

    return contract


def _parse_worksheet(ws, path: Path) -> SkyEastContract:
    # ------------------------------------------------------------------ #
    # Header block (rows 3-15)                                             #
    # ------------------------------------------------------------------ #
    def cell(row, col):
        return ws.cell(row=row, column=col).value

    pc_no_raw = _clean(cell(3, 5))
    pc_date = _format_date(cell(4, 5))
    buyer = _clean(cell(5, 5))
    seller = _clean(cell(7, 5))
    currency = _clean(cell(9, 5))
    payment_terms = _clean(cell(10, 5))
    trade_term = _clean(cell(13, 5))

    if not pc_no_raw:
        raise ValueError(f"Could not find PC NO. in row 3 col E of {path.name}")

    # ------------------------------------------------------------------ #
    # Find header row dynamically                                          #
    # ------------------------------------------------------------------ #
    header_row = _find_header_row(ws, search_from=14, search_to=20)
    if header_row is None:
        # Fall back to row 16 as per spec
        header_row = 16

    size_col_map = _map_size_columns(ws, header_row)

    # ------------------------------------------------------------------ #
    # Data rows                                                            #
    # ------------------------------------------------------------------ #
    items: list[SkyEastItem] = []
    current_brand = ""

    # Determine data start row
    data_start = header_row + 1

    # Iterate rows; stop at a large buffer past last non-empty row
    # Use a row limit to avoid the max_row=1048529 issue with openpyxl read_only
    MAX_EMPTY_ROWS = 10
    empty_streak = 0

    for row_num in range(data_start, data_start + 5000):
        # Read up to column 30
        row_values = {}
        any_value = False
        for col_num in range(1, 30):
            v = ws.cell(row=row_num, column=col_num).value
            if v is not None:
                row_values[col_num] = v
                any_value = True

        if not any_value:
            empty_streak += 1
            if empty_streak >= MAX_EMPTY_ROWS:
                break
            continue
        empty_streak = 0

        # Brand sub-section header?
        brand_name = _is_brand_section_row(row_values)
        if brand_name is not None:
            current_brand = brand_name
            continue

        # Totals / summary row?
        if _is_totals_row(row_values):
            continue

        # Must have a style value to be a data row
        style = _clean(row_values.get(4))
        if not style:
            continue

        # ---- Extract fields ---- #
        zalando_po = _clean(row_values.get(5)) or ""
        config_sku = _clean(row_values.get(6)) or ""
        article_name = _clean(row_values.get(7)) or ""
        picture_id = _extract_picture_id(row_values.get(8))
        fabric_item_no = _clean(row_values.get(9)) or ""
        fabrication = _clean(row_values.get(10)) or ""
        brand = _clean(row_values.get(11)) or current_brand
        color_name_raw = row_values.get(12)
        color_name = _clean(color_name_raw) or ""
        # Normalise newlines in color name
        color_name = " ".join(color_name.split())
        colour_code_raw = row_values.get(13)
        colour_code = str(colour_code_raw).strip() if colour_code_raw is not None else ""
        # BUG-37 fix: when col 14 is a datetime, _clean(str(dt)) yields
        # "2025-01-01 00:00:00" (with time).  Use _format_date for ISO output
        # consistent with pc_date / ex_fty_date.
        launch_date = _format_date(row_values.get(14)) or ""
        total_qty = _to_int(row_values.get(21))
        fob_usd = _to_float(row_values.get(22))
        total_cost_usd = _to_float(row_values.get(23))
        ex_fty_date = _format_date(row_values.get(24))

        # Sizes
        sizes: dict[str, int] = {}
        for size_label in _SIZE_LABELS:
            col = size_col_map.get(size_label)
            if col is not None:
                sizes[size_label] = _to_int(row_values.get(col))

        item = SkyEastItem(
            pc_no=pc_no_raw,
            zalando_po=zalando_po,
            style=style,
            config_sku=config_sku,
            article_name=article_name,
            brand=brand,
            color_name=color_name,
            colour_code=colour_code,
            launch_date=launch_date,
            fabric_item_no=fabric_item_no,
            fabrication=fabrication,
            sizes=sizes,
            total_qty=total_qty,
            fob_usd=fob_usd,
            total_cost_usd=total_cost_usd,
            ex_fty_date=ex_fty_date,
            picture_id=picture_id,
        )
        items.append(item)

    # ------------------------------------------------------------------ #
    # Build contract                                                        #
    # ------------------------------------------------------------------ #
    contract = SkyEastContract(
        pc_no=pc_no_raw,
        pc_date=pc_date,
        buyer=buyer,
        seller=seller,
        currency=currency,
        payment_terms=payment_terms,
        trade_term=trade_term,
        items=items,
        source_file=path.name,
        file_path=str(path.resolve()),
        extracted_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        parser_version=PARSER_VERSION,
        parse_confidence=90,
        source_file_hash=_file_hash(str(path)),
        processed_by="sky_east_excel",
    )
    return contract
