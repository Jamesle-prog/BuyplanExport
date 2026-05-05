"""Excel client PO parser — reads the two-row header mapping format.

Sheet layout (1.1.PO_Client):
  Row 1  — client's original column headings (varies per client)
  Row 2  — internal standardised column names (fixed schema)
  Row 3+ — data rows

The parser maps client headers → internal names via row 2, then normalises
each data row into SizeRow + POMetadata objects compatible with the rest of
the pipeline.
"""
from __future__ import annotations

import os
import openpyxl

from ..models import POData, POMetadata, SizeRow

# ---------------------------------------------------------------------------
# Internal column schema (row 2 values in 1.1.PO_Client)
# ---------------------------------------------------------------------------

# Columns the parser REQUIRES to produce useful output.
REQUIRED_INTERNAL = {
    "Purchase Order Number",
    "Main Supplier Config SKU",
}

# Size columns (order matters for the buy plan)
SIZE_COLUMNS = ["XS", "S", "M", "L", "XL", "XXL"]

# Map internal-header → SizeRow/POMetadata field
_FIELD_MAP = {
    "Purchase Order Number":            "po_number",
    "Main Supplier Config SKU":         "style",
    "Article Name":                     "style_description",
    "Config SKU":                       "config_sku",
    "Brand":                            "brand",
    "Main Supplier Color Description":  "color",
    "颜色":                              "color_cn",
    "辅助色":                            "secondary_color",
    "合同号":                            "contract_no",
    "入厂时间":                          "factory_entry_date",
    "质量要求":                          "quality_req",
    "备注1":                             "note1",
    "备注2":                             "note2",
    "备注3":                             "note3",
    "备注4":                             "note4",
    "面料_面料":                         "fabric1",
    "面料_面料1":                        "fabric2",
    "面料_面料2":                        "fabric3",
    "面料_面料3":                        "fabric4",
    "面料_面料_编号":                    "fabric1_code",
    "面料_面料1_编号":                   "fabric2_code",
    "面料_面料2_编号":                   "fabric3_code",
    "面料_面料3_编号":                   "fabric4_code",
    "面料_面料_部位":                    "fabric1_body_part",
    "面料_面料1_部位":                   "fabric2_body_part",
    "面料_面料2_部位":                   "fabric3_body_part",
    "面料_面料3_部位":                   "fabric4_body_part",
    "TOTAL QUANTITY":                   "total_qty",
    "Photo1":                           "photo1",
    "Photo2":                           "photo2",
}


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def parse_client_excel(path: str, sheet_name: str = "1.1.PO_Client") -> POData:
    """Parse a client Excel file using the two-row mapping header format."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
        )

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 3:
        raise ValueError("Sheet must have at least 3 rows (2 header rows + 1 data row).")

    # Row 1 = client headers (index 0), Row 2 = internal headers (index 1)
    client_headers = [str(v).strip() if v is not None else "" for v in rows[0]]
    internal_headers = [str(v).strip() if v is not None else "" for v in rows[1]]

    # Build column-index maps
    internal_col: dict[str, int] = {}   # internal_name → col_index
    for idx, hdr in enumerate(internal_headers):
        if hdr:
            internal_col[hdr] = idx

    # Validate required columns are present
    missing = REQUIRED_INTERNAL - set(internal_col.keys())
    if missing:
        raise ValueError(
            f"Missing required internal columns in row 2: {sorted(missing)}\n"
            f"Row 2 contains: {[h for h in internal_headers if h]}"
        )

    # Collect size columns that are actually present
    present_sizes = [s for s in SIZE_COLUMNS if s in internal_col]

    size_rows: list[SizeRow] = []
    metadata_by_po: dict[str, POMetadata] = {}

    for row_idx, row in enumerate(rows[2:], start=3):
        fields = _extract_fields(row, internal_col)

        po = str(fields.get("po_number") or "").strip()
        style = str(fields.get("style") or "").strip()
        color = str(fields.get("color") or fields.get("color_cn") or "").strip()

        if not po or not style:
            continue  # skip blank / summary rows

        # Build / update metadata for this PO
        if po not in metadata_by_po:
            meta = POMetadata(
                po_number=po,
                style=style,
                style_description=fields.get("style_description"),
                company=fields.get("brand"),
                source_format="client_excel",
                file_name=os.path.basename(path),
                file_path=path,
            )
            # Store extra HHP-specific fields as dynamic attributes.
            # BUG-16 fix: the old code used `setattr(...) if hasattr(meta, attr)`
            # which is always False for these fields (not declared on POMetadata),
            # silently dropping all fabric/note data.  Use plain setattr instead.
            for attr in ("fabric1", "fabric2", "fabric3", "fabric4",
                         "fabric1_code", "fabric2_code", "fabric3_code", "fabric4_code",
                         "fabric1_body_part", "fabric2_body_part",
                         "fabric3_body_part", "fabric4_body_part",
                         "note1", "note2", "note3", "note4",
                         "factory_entry_date", "quality_req",
                         "contract_no", "secondary_color", "photo1", "photo2"):
                val = fields.get(attr)
                if val is not None:
                    setattr(meta, attr, val)
            metadata_by_po[po] = meta

        # Emit one SizeRow per size that has non-zero quantity
        for size in present_sizes:
            qty_raw = row[internal_col[size]] if internal_col[size] < len(row) else None
            qty = _to_int(qty_raw)
            if qty:
                upc = ""  # not present in this format
                size_rows.append(SizeRow(
                    po_number=po,
                    style=style,
                    color=color,
                    size=size,
                    units=qty,
                    upc=upc,
                ))

    wb.close()

    # Use the first metadata record as the "primary" for this file
    primary_meta = next(iter(metadata_by_po.values())) if metadata_by_po else POMetadata(
        source_format="client_excel",
        file_name=os.path.basename(path),
        file_path=path,
    )

    return POData(metadata=primary_meta, size_rows=size_rows)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _extract_fields(row: tuple, col_map: dict[str, int]) -> dict:
    """Extract all known internal fields from a data row."""
    out = {}
    for internal_name, field_name in _FIELD_MAP.items():
        idx = col_map.get(internal_name)
        if idx is not None and idx < len(row):
            val = row[idx]
            if val is not None and str(val).strip():
                out[field_name] = str(val).strip()
    return out


def _to_int(val) -> int:
    if val is None:
        return 0
    try:
        return int(float(str(val).replace(",", "")))
    except (ValueError, TypeError):
        return 0


def parse_client_excel_to_df(path: str, sheet_name: str = "1.1.PO_Client") -> "pd.DataFrame":
    """Return a rich DataFrame (one row per source data row) preserving all group-level
    fields — fabric info, notes, photos, contract no., etc.

    Columns returned use the *internal* names from row 2 of the mapping sheet.
    An extra ``_source_file`` column records the originating filename.
    """
    import pandas as pd

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {path}. Available: {wb.sheetnames}")

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 3:
        raise ValueError("Sheet must have at least 3 rows (2 header rows + 1 data row).")

    internal_headers = [str(v).strip() if v is not None else "" for v in rows[1]]

    # Build column-index → internal-name map (skip blanks)
    col_names = []
    for h in internal_headers:
        if h:
            col_names.append(h)
        else:
            col_names.append(f"__unnamed_{len(col_names)}")

    records = []
    for raw_row in rows[2:]:
        # Pad or trim to match header width
        padded = list(raw_row) + [None] * max(0, len(col_names) - len(raw_row))
        record = {col_names[i]: padded[i] for i in range(len(col_names))}

        # Skip completely empty rows
        po = str(record.get("Purchase Order Number") or "").strip()
        style = str(record.get("Main Supplier Config SKU") or "").strip()
        if not po and not style:
            continue

        record["_source_file"] = os.path.basename(path)
        records.append(record)

    df = pd.DataFrame(records)
    # Coerce size columns to numeric
    for sz in SIZE_COLUMNS:
        if sz in df.columns:
            df[sz] = pd.to_numeric(df[sz].astype(str).str.replace(",", "", regex=False), errors="coerce").fillna(0).astype(int)
    if "TOTAL QUANTITY" in df.columns:
        df["TOTAL QUANTITY"] = pd.to_numeric(df["TOTAL QUANTITY"].astype(str).str.replace(",", "", regex=False), errors="coerce").fillna(0).astype(int)

    return df


def get_client_header_row(path: str, sheet_name: str = "1.1.PO_Client") -> list[str]:
    """Return row 1 (client's original headers) from the mapping sheet."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    wb.close()
    return [str(v).strip() if v else "" for v in row]


def extract_fabric_parts_from_row(row: dict) -> list:
    """
    Build a list of FabricPart objects from a DataFrame row or field dict
    produced by parse_client_excel_to_df / _extract_fields.

    Reads up to 4 fabric slots using the internal column names:
        fabricN_code  → hhn_no
        fabricN_body_part → body_part
        fabricN        → composition   (may be blank if only code is known)

    Returns a list of non-empty FabricPart objects (those with at least an
    hhn_no or composition value), ordered by fabric slot number.
    """
    from ..models.fabric_part import FabricPart

    parts = []
    for seq in range(1, 5):
        suffix = "" if seq == 1 else str(seq - 1)   # fabric1→"", fabric2→"1", ...
        hhn    = str(row.get(f"fabric{seq}_code")   or "").strip()
        part   = str(row.get(f"fabric{seq}_body_part") or "").strip()
        comp   = str(row.get(f"fabric{seq}")        or "").strip()
        if hhn or comp:
            parts.append(FabricPart(
                seq=seq,
                body_part=part,
                hhn_no=hhn,
                composition=comp,
            ))
    return parts


def get_internal_header_row(path: str, sheet_name: str = "1.1.PO_Client") -> list[str]:
    """Return row 2 (internal standardised headers) from the mapping sheet."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    wb.close()
    return [str(v).strip() if v else "" for v in (rows[0] if rows else [])]
