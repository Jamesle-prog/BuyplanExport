"""Parser for Sky East Purchase Contract Excel files.

All structural positions are determined dynamically:

  Contract sheet  : first worksheet whose header block contains "Sky East"
                    or a "PC NO." label; falls back to worksheets[0].
  Header row      : first row that contains both a style-like and a PO-like
                    column header (scanned up to row 40).
  Header columns  : every header token is matched against _COL_ALIASES; size
                    columns are recognised from _KNOWN_SIZES and mapped to the
                    6 DB size keys via _SIZE_TO_DB.
  Contract fields : label → value scan in rows 1-25, cols 1-6; value is the
                    first non-empty cell to the right of the label cell.

Returns SkyEastContract (with SkyEastItem list) matching models/sky_east_data.py.

Same PC No., same (style + color + zalando_po) across files → quantities are
aggregated by the store.  Amendment detection (size breakdown change) is also
handled in the store.
"""
from __future__ import annotations

import hashlib
import os
import re
from datetime import datetime
import openpyxl

from ..models.sky_east_data import SkyEastContract, SkyEastItem
from ..models.fabric_part import FabricPart
from ..utils.image_extractor import extract_dispimg_positions

PARSER_VERSION = "1.3"

# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

_DISPIMG_RE = re.compile(r'DISPIMG\("(ID_[0-9A-Fa-f]+)"', re.IGNORECASE)

# ── Size recognition ──────────────────────────────────────────────────────────
# Any header token whose stripped, upper-cased value is in this set is treated
# as a size column.  Add new variants here; no other code needs to change.
_KNOWN_SIZES: frozenset[str] = frozenset({
    # Letter sizes
    "XXS", "XS", "S", "SM", "M", "MED", "L", "LG", "XL", "XXL", "2XL", "3XL",
    "4XL", "XXXL", "XXXXL",
    # Plus / extended
    "1X", "2X", "3X", "4X", "5X", "6X",
    # Numeric (women's / children's)
    "0", "2", "4", "6", "8", "10", "12", "14", "16", "18", "20", "22", "24",
    "2T", "3T", "4T", "5T", "6T",
    # Other
    "OS", "ONE SIZE", "ONE-SIZE", "FREE SIZE", "FREESIZE",
})

# Map every recognised size (upper-cased) → DB column key.
# Unmapped sizes still contribute to total_qty in the store.
SIZE_TO_DB: dict[str, str] = {
    "XXS": "xs",
    "XS": "xs", "X-SMALL": "xs",
    "S": "s",  "SM": "s",  "SMALL": "s",
    "M": "m",  "MED": "m", "MEDIUM": "m",
    "L": "l",  "LG": "l",  "LARGE": "l",
    "XL": "xl", "X-LARGE": "xl", "XLARGE": "xl", "1X": "xl",
    "XXL": "xxl", "2XL": "xxl", "3XL": "xxl", "4XL": "xxl",
    "XXXL": "xxl", "XXXXL": "xxl",
    "2X": "xxl", "3X": "xxl", "4X": "xxl", "5X": "xxl", "6X": "xxl",
}

# Legacy constants kept for external callers that import them
SIZE_KEYS  = ("XS", "S", "M", "L", "XL", "2XL")
_SIZE_COLS = ("xs",  "s", "m", "l", "xl", "xxl")   # DB column names


def _is_size_header(text: str) -> bool:
    """Return True if *text* looks like a garment size column header."""
    return text.strip().upper() in _KNOWN_SIZES


def _v(cell_value) -> str:
    if cell_value is None:
        return ""
    if isinstance(cell_value, datetime):
        return cell_value.strftime("%Y-%m-%d")
    return str(cell_value).strip()


def _int(v) -> int:
    try:
        return int(float(str(v))) if v not in (None, "") else 0
    except (ValueError, TypeError):
        return 0


def _float(v) -> float:
    try:
        return float(str(v)) if v not in (None, "") else 0.0
    except (ValueError, TypeError):
        return 0.0


def _dispimg_id(val) -> str:
    if not val:
        return ""
    m = _DISPIMG_RE.search(str(val))
    return m.group(1) if m else ""


# ---------------------------------------------------------------------------
# Contract header detection (label → value scan)
# ---------------------------------------------------------------------------

# Each entry: (field_name, set_of_label_aliases)
# Labels appear somewhere in the header block (rows 1-20), values are read
# from the same row — first non-empty cell to the right of the label cell.
_HEADER_LABEL_ALIASES: list[tuple[str, set[str]]] = [
    ("pc_no", {
        "pc no.", "pc no", "p.c. no.", "pc number", "pc#",
        "合同编号", "合同号", "contract no", "contract no.", "contract number",
        "contract#", "po contract no", "purchase contract no",
    }),
    ("pc_date", {
        "date", "contract date", "pc date", "order date",
        "日期", "合同日期", "签约日期",
    }),
    ("party_a", {
        "party a", "buyer", "party a (buyer)", "甲方", "买方",
        "issued to", "sold to", "bill to", "customer", "client",
    }),
    ("party_b", {
        "party b", "seller", "party b (seller)", "乙方", "卖方",
        "vendor", "supplier", "factory", "manufacturer",
    }),
    ("currency", {
        "currency", "currency:", "curr.", "货币", "结算货币",
    }),
    ("payment_terms", {
        "payment", "payment method", "payment terms", "payment condition",
        "terms of payment", "付款方式", "付款条款",
    }),
    ("trade_term", {
        "trade term", "trade terms", "incoterm", "incoterms",
        "delivery term", "delivery terms", "贸易条款", "交货条款",
    }),
]


def _norm_label(s: str) -> str:
    """Lower-case, strip, remove trailing punctuation for label matching."""
    return str(s).strip().lower().rstrip(".: ：")


def _parse_contract_header(ws) -> dict[str, str]:
    """Scan rows 1-20 searching for label text; return {field: value_str}.

    Strategy per row:
      1. Scan cols A-D for a cell whose normalised text matches a known label.
      2. Once found, read the value from col E (the standard value column in
         Sky East contracts).  If col E is blank, walk cols F-H for the first
         non-empty cell.
      3. Fall back to the hardcoded row/col positions if nothing is found
         (keeps backward-compatibility with files that have no label text).
    """
    # Build normalised alias → field lookup
    alias_to_field: dict[str, str] = {}
    for field, aliases in _HEADER_LABEL_ALIASES:
        for alias in aliases:
            alias_to_field[_norm_label(alias)] = field

    found: dict[str, str] = {}

    for r in range(1, 21):
        for c in range(1, 5):   # cols A-D are label columns
            cell_val = ws.cell(row=r, column=c).value
            if cell_val is None:
                continue
            norm = _norm_label(str(cell_val))
            field = alias_to_field.get(norm)
            if field and field not in found:
                # Value is in the next non-empty cell on this row (col c+1 … H).
                # Standard layout puts it in col E (5); some files use col D (4).
                val = ""
                for vc in range(c + 1, 9):
                    val = _v(ws.cell(row=r, column=vc).value)
                    if val:
                        break
                found[field] = val
                break   # move to next row once a field is found on this row

    # Hard-coded fallbacks for any field not found by label search.
    # Try col 4 first (newer layout), then col 5 (classic layout).
    _FALLBACK_ROWS = {
        "pc_no": (3,), "pc_date": (4,), "party_a": (5,),
        "party_b": (7,), "currency": (9,),
        "payment_terms": (10,), "trade_term": (13,),
    }
    for field, (row,) in _FALLBACK_ROWS.items():
        if field not in found:
            for col in (4, 5, 6):
                val = _v(ws.cell(row=row, column=col).value)
                if val:
                    found[field] = val
                    break

    return found


# ---------------------------------------------------------------------------
# Column detection
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Contract sheet detection
# ---------------------------------------------------------------------------

_CONTRACT_SIGNALS = frozenset({
    "sky east international trading limited",
    "sky east",
    "pc no.",
    "pc no",
    "purchase contract",
    "purchase order contract",
})


def looks_like_sky_east_contract(wb) -> bool:
    """Cheap Sky East signature check for the universal file detector.

    Scans the first 15 rows × 9 columns of each sheet for the same contract
    keywords ``_find_contract_sheet`` scores on.  Uses ``iter_rows`` (not
    ``ws.cell``) so it also works on read-only workbooks, which is how the
    detector opens files.
    """
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=15, max_col=9, values_only=True):
            for val in row:
                raw = str(val or "").strip().lower()
                if raw and any(sig in raw for sig in _CONTRACT_SIGNALS):
                    return True
    return False


def _find_contract_sheet(wb):
    """Return the worksheet most likely to be the Sky East purchase contract.

    Scans the first 15 rows × 10 columns of each sheet for keywords that
    only appear in contract sheets (Sky East company name, PC NO. label, etc.).
    Falls back to ``worksheets[0]`` if nothing matches.
    """
    best_ws   = wb.worksheets[0]
    best_hits = 0

    for ws in wb.worksheets:
        hits = 0
        for r in range(1, 16):
            for c in range(1, 10):
                raw = str(ws.cell(r, c).value or "").strip().lower()
                if not raw:
                    continue
                if any(sig in raw for sig in _CONTRACT_SIGNALS):
                    hits += 1
                # Direct PC No. pattern match (e.g. HHPPC038, LSKHHN005R …)
                if re.match(r'^[A-Z]{2,8}PC\d{3}', raw, re.IGNORECASE):
                    hits += 2
        if hits > best_hits:
            best_hits = hits
            best_ws   = ws

    return best_ws


# ---------------------------------------------------------------------------
# Column detection
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Column aliases
# ---------------------------------------------------------------------------
# Non-size column aliases — sizes are detected dynamically via _is_size_header.
# Add synonyms freely; _find_header_row and _map_columns both use this table.
# ---------------------------------------------------------------------------
_COL_ALIASES: dict[str, set[str]] = {
    "item": {
        "item", "item no", "item no.", "no", "no.", "seq", "seq.",
        "line", "line no", "line no.", "line item",
    },
    "return_label": {
        "需要挂 return label", "return label", "return label ",
        "return label?", "need return label",
    },
    "style_no": {
        "style no.", "style no", "style", "style#", "style number",
        "style code", "design no", "design no.", "design number",
        "supplier style", "supplier style no", "supplier style number",
        "supplier article number", "supplier article no",
        "main supplier config sku", "vendor style",
        "item style", "sku style",
    },
    "po_number": {
        "po number", "po no.", "po no", "po#", "po_number",
        "purchase order number", "purchase order no",
        "purchase order no.", "purchase order#",
        "order number", "order no", "order no.", "order#",
        "zalando po", "zalando po number", "buyer po", "buyer po number",
        "client po", "retailer po",
    },
    "config_sku": {
        "config_sku", "config sku", "config sku.", "sku",
        "buyer sku", "buyer article sku", "retailer sku",
        "article sku", "product sku",
    },
    "article_name": {
        "supplier article name", "article name", "article description",
        "product name", "product description", "garment description",
        "description", "item description", "style description",
        "style name", "item name",
    },
    "picture": {
        "piicture", "picture", "image", "photo", "photo id",
        "picture id", "image id", "style image",
    },
    "fabric_no": {
        "fabric item number", "fabric item no", "fabric item no.",
        "fabric no", "fabric no.", "fabric number", "fabric code",
        "fabric reference", "fabric ref", "hhn no",
        "material no", "material number", "material code",
    },
    "fabrication": {
        "fabrication", "fabrication ", "composition",
        "fabric composition", "material composition",
        "fabric content", "content",
    },
    "brand": {
        "brand", "brand name", "brand/label", "label",
        "buyer brand", "division", "sub-brand",
    },
    "color_name": {
        "color name", "colour name", "color", "colour",
        "color description", "colour description",
        "colorway", "colourway", "color/colour",
        "supplier color name", "supplier colour name",
        "color description en", "shade", "shade name",
    },
    "color_code": {
        "colour code", "color code", "color code.", "colour code.",
        "color ref", "colour ref", "color reference",
        "colour reference", "color id", "colour id",
        "supplier color code", "supplier colour code",
        "color#", "colour#",
    },
    "launch_date": {
        "launch date", "handover window from",
        "delivery date", "ship date", "ship window",
        "delivery window", "delivery window from",
        "required delivery date", "required ship date",
        "target delivery", "target ship date",
        "in-store date", "in store date",
    },
    "total_qty": {
        "total quantity", "qty", "total qty",
        "quantity ordered", "quantity", "order qty",
        "total pcs", "pcs", "total pieces", "total units",
        "units", "total order qty", "ordered qty",
    },
    "fob_usd": {
        "fob\nusd", "fob usd", " fob \nusd", "fob",
        "fob price", "fob unit price", "unit price usd",
        "unit price", "unit fob", "price usd", "price (usd)",
        "net price", "selling price usd",
    },
    "total_cost": {
        "total cost\nusd", "total cost usd", "total cost",
        "total amount", "total amount usd", "amount usd",
        "total value", "total value usd",
        "extended cost", "total extended",
    },
    "ex_fty": {
        "ex-fty", "ex fty", "exfty", "离厂期",
        "ex factory", "ex-factory", "ex factory date",
        "ex-factory date", "ex fac", "ex fac date",
        "ship date", "ship ex factory", "departure date",
        "date ex factory",
    },
}

# Minimal positional fallbacks — only applied to fields not found by header scan.
# Prefer dynamic detection; these only protect against completely headerless files.
_DEFAULTS: dict[str, int] = {
    "item": 1, "style_no": 4, "po_number": 5,
    "config_sku": 6, "article_name": 7, "picture": 8,
    "fabric_no": 9, "fabrication": 10, "brand": 11,
    "color_name": 12, "color_code": 13, "launch_date": 14,
    "total_qty": 21, "fob_usd": 22, "total_cost": 23, "ex_fty": 24,
}


def _find_header_row(ws) -> int:
    """Return the 1-based row index of the column-header row.

    Strategy: score every row up to row 40 by counting how many cells match
    a known column alias (from _COL_ALIASES) or a known size token.  The row
    with the highest score is the header row.  Ties are broken by taking the
    earlier row.  Falls back to row 16 when nothing scores above 0.

    Auto-derives its signal set from _COL_ALIASES so adding new aliases there
    automatically improves header detection — no separate list to maintain.
    """
    # Build the full signal set from _COL_ALIASES
    _signals: frozenset[str] = frozenset(
        alias for aliases in _COL_ALIASES.values() for alias in aliases
    )

    max_col = min((ws.max_column or 0) + 1, 50)
    best_row, best_score = 16, 0

    for r in range(1, 41):
        score = 0
        for c in range(1, max_col):
            v = _v(ws.cell(row=r, column=c).value).strip().lower()
            if not v:
                continue
            if v in _signals or _is_size_header(v):
                score += 1
        if score > best_score:
            best_score = score
            best_row = r

    return best_row


def _map_columns(ws, hrow: int) -> dict[str, int]:
    """Build a ``{field_name: column_index}`` map from the header row.

    Three passes over the header cells:

    Pass 1 — Exact alias match
        Each normalised cell value is looked up directly in ``_COL_ALIASES``.

    Pass 2 — Size columns
        Anything not matched in pass 1 is tested with ``_is_size_header``
        and stored under its canonical upper-cased name (``"XS"``, ``"2XL"`` …).

    Pass 3 — Partial / substring match (fallback for near-miss headers)
        For every field still unmapped, check whether any alias string is a
        substring of the cell text or vice-versa (minimum alias length: 5
        characters to avoid false positives on short words like "no", "qty").

    Positional defaults from ``_DEFAULTS`` are applied last, only for fields
    that remained undetected after all three passes.
    """
    col_map: dict[str, int] = {}
    max_col = min((ws.max_column or 0) + 1, 50)

    # Collect (col, norm) pairs for all non-empty header cells
    header_cells: list[tuple[int, str]] = []
    for c in range(1, max_col):
        raw  = _v(ws.cell(row=hrow, column=c).value)
        norm = raw.lower().strip()
        if not norm:
            continue
        header_cells.append((c, norm))

        # Pass 1: exact alias match
        matched = False
        for key, aliases in _COL_ALIASES.items():
            if norm in aliases and key not in col_map:
                col_map[key] = c
                matched = True
                break

        # Pass 2: size columns
        if not matched and _is_size_header(norm):
            canonical = norm.strip().upper()
            if canonical not in col_map:
                col_map[canonical] = c

    # Pass 3: partial / substring matching for fields still unmapped
    claimed_cols = set(col_map.values())
    for c, norm in header_cells:
        if c in claimed_cols:
            continue
        for key, aliases in _COL_ALIASES.items():
            if key in col_map:
                continue
            for alias in aliases:
                # Guard: alias must be long enough to avoid noise
                if len(alias) < 5:
                    continue
                if alias in norm or norm in alias:
                    col_map[key] = c
                    claimed_cols.add(c)
                    break

    # Apply positional defaults only for fields still missing
    for k, v in _DEFAULTS.items():
        col_map.setdefault(k, v)

    return col_map


# ---------------------------------------------------------------------------
# Row classification
# ---------------------------------------------------------------------------

def _is_brand_divider(item_val, style_val) -> tuple[bool, str]:
    """Return (True, brand_name) for brand section header rows."""
    if item_val is None or item_val == "":
        return False, ""
    if isinstance(item_val, (int, float)):
        return False, ""
    s = str(item_val).strip()
    if re.match(r'^\d+$', s):
        return False, ""
    if re.match(r'^(SAY |TOTAL)', s, re.IGNORECASE):
        return False, ""
    if style_val and str(style_val).strip():
        return False, ""        # has a style number → it's a data row
    return True, s


def _is_footer(item_val) -> bool:
    if item_val is None:
        return False
    s = str(item_val).strip()
    return bool(re.match(r'^(SAY |TOTAL)', s, re.IGNORECASE))


# ---------------------------------------------------------------------------
# Multi-fabric extraction
# ---------------------------------------------------------------------------

def _extract_fabric_parts(fabric_cell) -> list[FabricPart]:
    """
    Parse a (possibly multi-line) fabric-number cell into an ordered list
    of FabricPart objects.

    Examples
    --------
    "HHN-JA-01715"
        → [FabricPart(seq=1, body_part="", hhn_no="HHN-JA-01715")]

    "大身HHN-JA-01715\\n网布HHN-MS-01794"
        → [FabricPart(seq=1, body_part="大身", hhn_no="HHN-JA-01715"),
           FabricPart(seq=2, body_part="网布", hhn_no="HHN-MS-01794")]

    "大身：HHN-JA-01715，300克\\n口袋：HHN-MS-01794"
        → same as above (tolerates Chinese punctuation)
    """
    raw = _v(fabric_cell)
    if not raw:
        return []

    parts: list[FabricPart] = []
    seq = 1
    for line in raw.split("\n"):
        line = line.strip()
        if not line:
            continue
        matches = list(re.finditer(r'(HHN-\S+)', line))
        if not matches:
            continue
        for m in matches:
            hhn = m.group(1).strip("，,：: ")
            prefix = line[:m.start()].strip()
            prefix = re.sub(r'[：:,，\s]+$', '', prefix)
            parts.append(FabricPart(seq=seq, body_part=prefix, hhn_no=hhn))
            seq += 1
    return parts


def _fabric_item_no(parts: list[FabricPart]) -> str:
    """Primary HHN number for backward-compat fabric_item_no field."""
    return parts[0].hhn_no if parts else ""


def _fabrication_display(parts: list[FabricPart], fabrication_cell) -> str:
    """Human-readable multi-fabric string for the fabrication display field."""
    fab_raw = _v(fabrication_cell)
    if not parts:
        return fab_raw
    segments = []
    for p in parts:
        seg = p.hhn_no
        if p.body_part:
            seg = f"{p.body_part}: {seg}"
        segments.append(seg)
    result = " | ".join(segments)
    if fab_raw:
        result += f" — {fab_raw}"
    return result


# ---------------------------------------------------------------------------
# Public parse function
# ---------------------------------------------------------------------------

def parse(path: str, processed_by: str = "") -> SkyEastContract:
    """Parse a Sky East Excel file and return a SkyEastContract.

    Sheet selection, header-row position, column mapping, and size columns are
    all determined dynamically from the file content — no hardcoded positions.

    Each SkyEastItem has:
      • fabric_item_no  — primary HHN (backward compat)
      • fabrication     — human-readable display string
      • fabric_parts    — list[FabricPart] with full structured data
      • sizes           — dict keyed by actual size labels found in the file
                          (e.g. {"XS": 49, "S": 143, "1X": 60})

    processed_by  : username of the person who triggered the upload (optional)
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = _find_contract_sheet(wb)     # ← dynamic sheet detection

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    fname   = os.path.basename(path)

    # ── Source file hash ──────────────────────────────────────────────────────
    with open(path, "rb") as fh:
        file_hash = hashlib.md5(fh.read()).hexdigest()

    # ── Contract header — dynamic label search with position fallbacks ────────
    # .get with "" fallback: _parse_contract_header only inserts keys whose
    # label (or hardcoded fallback cell) was found — a variant contract that
    # omits an optional row (e.g. Trade Term) must not KeyError the whole file.
    hdr      = _parse_contract_header(ws)
    pc_no    = hdr.get("pc_no", "")
    pc_date  = hdr.get("pc_date", "")
    party_a  = hdr.get("party_a", "")
    party_b  = hdr.get("party_b", "")
    currency = hdr.get("currency", "")
    payment  = hdr.get("payment_terms", "")
    trade    = hdr.get("trade_term", "")

    # ── Column mapping ────────────────────────────────────────────────────────
    hrow = _find_header_row(ws)
    col  = _map_columns(ws, hrow)

    # Pre-extract DISPIMG positions directly from XML (reliable even with data_only=True).
    # Use the *found* contract sheet's index — hardcoding 0 read image positions
    # from sheet 1 even when the contract lives on a later sheet, attaching
    # wrong/no photos by (row, col) collision.
    dispimg_pos = extract_dispimg_positions(
        path, sheet_index=wb.sheetnames.index(ws.title)
    )

    def cv(row, key):
        c = col.get(key)
        return ws.cell(row=row, column=c).value if c else None

    # ── Parse data rows ───────────────────────────────────────────────────────
    items: list[SkyEastItem] = []
    skipped_zero_qty: list[dict] = []
    current_brand = ""

    for r in range(hrow + 1, ws.max_row + 1):
        item_raw  = cv(r, "item")
        style_raw = cv(r, "style_no")

        if _is_footer(item_raw):
            break

        is_div, brand_name = _is_brand_divider(item_raw, style_raw)
        if is_div:
            current_brand = brand_name
            continue

        # Accept rows that either have an integer item number OR a valid style number.
        style_no = _v(style_raw).strip("\n ")
        if item_raw is None or item_raw == "":
            if not style_no:
                continue
        else:
            try:
                int(float(str(item_raw)))
            except (ValueError, TypeError):
                if not style_no:
                    continue

        if not style_no:
            continue

        # Brand: prefer per-row brand col, fall back to last seen divider
        brand = _v(cv(r, "brand")) or current_brand

        # ── Dynamic sizes ─────────────────────────────────────────────────────
        # All keys in col_map that pass _is_size_header are size columns.
        sizes: dict[str, int] = {}
        for key, cidx in col.items():
            if _is_size_header(key):
                qty = _int(ws.cell(row=r, column=cidx).value)
                if qty:   # only store non-zero for compactness
                    sizes[key.upper()] = qty

        total_qty = _int(cv(r, "total_qty"))
        if total_qty == 0:
            total_qty = sum(sizes.values())

        # Zero-unit rows are typically a cancelled/blanked-out order line
        # (strikethrough style, PO, price all wiped but the row left in
        # place) — importing them creates a phantom item with no real
        # quantity. Drop the row and record it so it's surfaced in the
        # upload log rather than silently disappearing.
        if total_qty == 0:
            skipped_zero_qty.append({
                "row": r, "style": style_no,
                "po": _v(cv(r, "po_number")),
            })
            continue

        # ── Multi-fabric extraction ───────────────────────────────────────────
        fabric_parts = _extract_fabric_parts(cv(r, "fabric_no"))

        item = SkyEastItem(
            pc_no          = pc_no,
            zalando_po     = _v(cv(r, "po_number")),
            style          = style_no,
            config_sku     = _v(cv(r, "config_sku")),
            article_name   = _v(cv(r, "article_name")),
            brand          = brand,
            color_name     = _v(cv(r, "color_name")).replace("\n", " ").strip(),
            colour_code    = _v(cv(r, "color_code")),
            launch_date    = _v(cv(r, "launch_date")),
            fabric_item_no = _fabric_item_no(fabric_parts),
            fabrication    = _fabrication_display(fabric_parts, cv(r, "fabrication")),
            contract_no    = "",
            sizes          = sizes,
            total_qty      = total_qty,
            fob_usd        = _float(cv(r, "fob_usd")),
            total_cost_usd = _float(cv(r, "total_cost")),
            ex_fty_date    = _v(cv(r, "ex_fty")) or None,
            picture_id     = (dispimg_pos.get((r, col.get("picture", 8)))
                              or _dispimg_id(cv(r, "picture"))),
            fabric_parts   = fabric_parts,
        )
        items.append(item)

    wb.close()

    # Confidence + validation status — same formula and scale the GIII PDF
    # parsers use (100 minus 10 per missing key field; "exception" when the
    # file yielded no line items at all), so downstream consumers can treat
    # both pipelines' quality grades interchangeably.
    _key_fields = [pc_no, pc_date, party_a, party_b, currency, payment]
    confidence = 100 - 10 * sum(1 for f in _key_fields if not (f or "").strip())
    if not items:
        validation_status = "exception"
    elif confidence >= 70:
        validation_status = "valid"
    else:
        validation_status = "warning"

    return SkyEastContract(
        pc_no            = pc_no,
        pc_date          = pc_date,
        buyer            = party_a,
        seller           = party_b,
        currency         = currency,
        payment_terms    = payment,
        trade_term       = trade,
        items            = items,
        source_file      = fname,
        file_path        = path,
        extracted_at     = now_str,
        parser_version   = PARSER_VERSION,
        parse_confidence = confidence,
        validation_status = validation_status,
        source_file_hash = file_hash,
        processed_by     = processed_by,
        skipped_zero_qty = skipped_zero_qty,
    )
