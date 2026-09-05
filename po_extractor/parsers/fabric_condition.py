"""Parser for the 面料情况 (actual fabric condition) sheet.

This is a hand-kept shop-floor log, not a structured export: one row per
fabric delivery/test — measured width and weight against nominal, shrinkage
test results, cutting requirements (max plies, max length, direction), net
consumption, and remaining stock.

Everything is stored as text, verbatim
------------------------------------------
Every field here except the sheet's own row number (``序号``) is kept as
plain text, with no numeric coercion at all. The sheet is full of things a
number can't hold without losing information:

  * ranges instead of a single reading —  ``"176-178"``, ``"148-175"``
  * approximations — ``"100层左右"`` ("about 100 layers")
  * a word instead of a value — ``"无"`` (none), ``"/"``, ``"同上"`` (ditto,
    referring to the row above — never resolved here; that would be a
    judgement call this parser has no business making)
  * a typo'd sign — ``"负0.6%"`` (Chinese "negative" instead of a minus sign)

Coercing any of these to a float would silently discard the part that didn't
fit. The one exception is ``日期``: its cells carry a plain Excel serial
number with **General** formatting (not a date format — openpyxl would have
handed back a real ``datetime`` if it were), so a clean integer is converted
to an ISO date; anything that isn't a clean integer is kept as typed.

Two ambiguous headings
-----------------------
The sheet repeats ``径向`` / ``纬向`` twice — once under the merged heading
``面料缩率`` (fabric shrinkage) and once under ``纸板缩率`` (pattern-paper
shrinkage). Row 1's merge tells which is which; the columns are otherwise
identical text.
"""
from __future__ import annotations

from typing import Any

import openpyxl
from openpyxl.utils.datetime import from_excel
from ..utils.normalize import cell_text, norm_header_key, to_int


_norm = norm_header_key


def _txt(v) -> str:
    """Verbatim display text — the whole point of this parser."""
    return cell_text(v, int_floats=True)


def _int(v) -> int | None:
    return to_int(v, strict=True)

SHEET_NAME = "1.面料情况"


class FabricConditionParseError(ValueError):
    """Raised when the workbook carries no recognisable header row."""


# ── cell helpers ────────────────────────────────────────────────────────────


def _date_text(v: Any) -> str:
    """ISO date when *v* is a clean Excel-serial integer; raw text otherwise.

    The column's own cell format is General (checked against the source
    workbook), so a real date typed here would already come back as a
    ``datetime`` — a bare int is a serial number that lost its date
    formatting, not a different kind of value.
    """
    n = _int(v)
    if n is not None and 1 <= n <= 2_958_465:      # openpyxl's valid range
        try:
            return from_excel(n).strftime("%Y-%m-%d")
        except (ValueError, OverflowError):
            pass
    return _txt(v)


# ── header mapping ──────────────────────────────────────────────────────────

_FIELDS: dict[str, tuple[str, ...]] = {
    "seq":            ("序号",),
    "test_date":      ("日期",),
    "fabric_no":      ("面料编号",),
    "style":          ("款号",),
    "body_part":      ("部位",),
    "color":          ("颜色",),
    "width_net":      ("有效门幅(cm)", "有效门幅"),
    "width_gross":    ("毛门幅(cm)", "毛门幅"),
    "weight_gsm":     ("克重(g)", "克重"),
    "over_cut_pct":   ("超裁%",),
    "max_plies":      ("裁剪最高层数",),
    "max_cut_length": ("裁剪最大长度",),
    "cut_direction":  ("裁剪方向及要求",),
    "consumption_purchase_body":    ("采购净单耗(cm)-大身",),
    "consumption_purchase_binding": ("采购净单耗(cm)-滚条",),
    "consumption_layout_body":      ("排版净单耗(cm)-大身",),
    "consumption_layout_binding":   ("排版净单耗(cm)-滚条",),
    "remaining_rolls": ("剩余面料(匹)",),
    "remaining_kg":     ("剩余面料(kg)",),
    "remaining_m":      ("剩余面料(m)",),
}
_HEAD_TO_FIELD = {_norm(h): f for f, heads in _FIELDS.items() for h in heads}

_GROUP_FABRIC  = "面料缩率"   # fabric shrinkage
_GROUP_PATTERN = "纸板缩率"   # pattern-paper shrinkage

# Display order, matching the sheet's own column order.
FIELD_ORDER = (
    "seq", "test_date", "fabric_no", "style", "body_part", "color",
    "width_net", "width_gross", "weight_gsm",
    "shrink_fabric_warp", "shrink_fabric_weft",
    "shrink_pattern_warp", "shrink_pattern_weft",
    "over_cut_pct", "max_plies", "max_cut_length", "cut_direction",
    "consumption_purchase_body", "consumption_purchase_binding",
    "consumption_layout_body", "consumption_layout_binding",
    "remaining_rolls", "remaining_kg", "remaining_m",
)

# A row counts as a record when it names at least one of these.
_KEY_FIELDS = ("fabric_no", "style", "color")


def _row1_groups(ws) -> dict[int, str]:
    """``{column: group label}`` from row-1 merged headings.

    A heading's text lives only on the merge's top-left cell; this maps it
    across every column the merge spans, so a plain per-column lookup works
    for the rest of the parser.
    """
    groups: dict[int, str] = {}
    for rng in ws.merged_cells.ranges:
        if rng.min_row != 1 or rng.max_row != 1:
            continue
        label = _txt(ws.cell(1, rng.min_col).value)
        if label:
            for c in range(rng.min_col, rng.max_col + 1):
                groups[c] = label
    return groups


def _find_header_row(ws) -> int:
    for r in range(1, min(ws.max_row, 10) + 1):
        heads = {_norm(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)}
        if _norm("序号") in heads and _norm("颜色") in heads:
            return r
    return -1


def _map_columns(ws, header_row: int) -> dict[str, int]:
    """``{field: column}``, resolving the two ambiguous 径向/纬向 pairs via
    the row-1 group each sits under rather than by text (both pairs read
    identically on their own row)."""
    groups = _row1_groups(ws)
    cols: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        head = _norm(ws.cell(header_row, c).value)
        if not head:
            continue
        group = groups.get(c, "")
        if group == _GROUP_FABRIC and head.startswith(_norm("径向")):
            cols.setdefault("shrink_fabric_warp", c)
        elif group == _GROUP_FABRIC and head.startswith(_norm("纬向")):
            cols.setdefault("shrink_fabric_weft", c)
        elif group == _GROUP_PATTERN and head.startswith(_norm("径向")):
            cols.setdefault("shrink_pattern_warp", c)
        elif group == _GROUP_PATTERN and head.startswith(_norm("纬向")):
            cols.setdefault("shrink_pattern_weft", c)
        elif head in _HEAD_TO_FIELD:
            cols.setdefault(_HEAD_TO_FIELD[head], c)
    return cols


# ── main entry point ────────────────────────────────────────────────────────

def parse_fabric_condition(source: Any) -> dict[str, Any]:
    """Parse the 面料情况 sheet. Returns ``{"records": [...], "n_cols": int}``."""
    wb = openpyxl.load_workbook(source, data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.worksheets[0]

    header_row = _find_header_row(ws)
    if header_row < 0:
        raise FabricConditionParseError(
            "No fabric-condition header row found — expected 序号 and 颜色 "
            "among the column headings."
        )
    cols = _map_columns(ws, header_row)

    def cell(r: int, field: str):
        c = cols.get(field)
        return ws.cell(r, c).value if c else None

    records: list[dict[str, Any]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        rec = {"row_no": r, "seq": _int(cell(r, "seq")),
              "test_date": _date_text(cell(r, "test_date"))}
        for field in FIELD_ORDER:
            if field in ("seq", "test_date"):
                continue
            rec[field] = _txt(cell(r, field))
        if not any(rec.get(k) for k in _KEY_FIELDS):
            continue
        records.append(rec)

    return {"records": records, "n_cols": len(cols)}
