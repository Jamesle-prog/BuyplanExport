"""Parser for the settlement statistics workbook (结算统计表).

One row per invoice line: what was contracted, what shipped, what was invoiced
and received, and what went out again in payments and fees.  The workbook is
maintained by hand, so nothing here relies on a column sitting in a fixed
place — every field is resolved by its heading.

Layout
------
Each bulk sheet carries a **two-row** header.  The upper row holds the plain
columns plus two group headings, ``支付`` and ``费用``; the lower row names the
columns inside those groups.  A payment's date column is simply headed ``日期``
and only means anything by position, so it is taken as the column immediately
right of the payment it belongs to.

Sheets differ in column *order* between clients and years — ``Zalando2026``
puts 辅助列/款号/PO# where ``Zalando2025`` puts PO#/辅助列/款号, and the
non-Zalando sheets have neither 翻单还是新款 nor 溢短装.  Heading lookup makes
that a non-issue.

Colour is data here, not decoration
-----------------------------------
Two facts are recorded only as cell fills, and the sheet says so in its own
legend rows:

* a red ``INVOICE NO.`` cell means the line is at **discount risk**
  (``红底的是有折扣风险的``);
* the fill on ``合同号`` says whether the contract has been **sent to the
  factory** or **signed back** (``合同已寄工厂`` / ``合同已回签``).

The legend cells are read for their own colours rather than the colours being
hardcoded, so a workbook that re-themes still comes out right.
"""
from __future__ import annotations

import re
from typing import Any

import openpyxl
from ..utils.normalize import cell_text, norm_header_key, to_float
from ._sheet import cell_getter, find_header_row, header_index


_norm = norm_header_key


def _txt(v) -> str:
    return cell_text(v, dates=True, int_floats=True)


# Ex-factory is sometimes a revision, not a date (``2025/8/28->9/4``): a
# real date prints ISO, anything else is kept verbatim.
_date = _txt


def _num(v) -> float | None:
    # amounts may carry a thousands separator or a currency mark; '3.25%' → 0.0325
    return to_float(v, strip_currency=True, percent="ratio")

# Sheets that are recomputed rather than imported: they are views of the bulk
# sheets (overdue receivables, discount-risk lines), and importing them would
# create a second copy that goes stale the moment a payment lands.
DERIVED_SHEETS = ("到期未付款明细", "折扣风险明细")
SAMPLE_SHEET = "样品"

# Anything else with no recognisable header is skipped as a scratch sheet.


class SettlementParseError(ValueError):
    """Raised when a workbook carries no settlement sheet at all."""


# ── cell helpers ────────────────────────────────────────────────────────────


# ── header mapping ──────────────────────────────────────────────────────────

# heading → field.  Several spellings per field: the sheets disagree on
# 采购单价 vs 加工费 for the same column, and SC2026 marks its currency in the
# heading itself (``FOB（£）``).
_FIELDS: dict[str, tuple[str, ...]] = {
    "invoice_no":      ("invoiceno.", "invoiceno", "invoice"),
    "po_number":       ("po#", "po号", "ponumber"),
    "style":           ("款号",),
    "contract_no":     ("合同号",),
    "contract_qty":    ("合同数量",),
    "repeat_or_new":   ("翻单还是新款",),
    "factory":         ("工厂",),
    "unit_cost":       ("采购单价", "加工费"),
    "fob":             ("fob", "fob($)", "fob(£)", "fob(¥)"),
    "contract_amount": ("合同金额", "合同金额($)", "合同金额(£)"),
    "ex_factory":      ("离厂时间", "离厂日期"),
    "shipped_qty":     ("出货数", "出货数量"),
    "over_short_pct":  ("溢短装",),
    "invoice_amount":  ("发票金额(报关金额)", "发票金额", "报关金额"),
    "received":        ("实际收汇",),
    "received_date":   ("日期",),
    "helper":          ("辅助列",),
}
_HEAD_TO_FIELD = {h: f for f, heads in _FIELDS.items() for h in heads}

# Group headings on the upper row, and the columns inside them on the lower.
_GROUP_PAY = "支付"
_GROUP_FEE = "费用"
_PAY_FIELDS = {"面料款": "pay_fabric", "辅料款": "pay_trim", "加工费": "pay_cmt"}
_FEE_FIELDS = {
    "港杂费(含运费)": "fee_port", "港杂费": "fee_port",
    "其他一": "fee_other1",
    "其他二(税金)": "fee_other2", "其他二": "fee_other2",
}
_DATE_HEAD = "日期"

# A row counts as a settlement line when it names at least one of these; a
# sheet's stray totals and notes carry none of them.
_KEY_FIELDS = ("invoice_no", "po_number", "style", "contract_no")


def _find_header_row(ws) -> int:
    """Row of the upper header, or -1.  Identified by the headings on it."""
    return find_header_row(
        ws, lambda heads: "合同号" in heads and bool(heads & {"款号", "invoiceno.", "po#"}),
        max_rows=15, norm=_norm)


def _map_columns(ws, header_row: int) -> dict[str, int]:
    """``{field: column}`` for one sheet.

    The two rows are read together: a plain heading on the upper row maps
    straight to its field, while ``支付`` / ``费用`` hand off to the lower row.
    Each payment's ``日期`` is the column to its right — the heading alone
    cannot say which payment it dates.
    """
    cols: dict[str, int] = {}
    sub = header_row + 1
    for c in range(1, ws.max_column + 1):
        upper = _norm(ws.cell(header_row, c).value)
        lower = _norm(ws.cell(sub, c).value)
        field = _HEAD_TO_FIELD.get(upper)
        if field and field not in cols:
            cols[field] = c
        if lower in _PAY_FIELDS:
            cols.setdefault(_PAY_FIELDS[lower], c)
            if _norm(ws.cell(sub, c + 1).value) == _DATE_HEAD:
                cols.setdefault(f"{_PAY_FIELDS[lower]}_date", c + 1)
        elif lower in _FEE_FIELDS:
            cols.setdefault(_FEE_FIELDS[lower], c)
    return cols


def _currency(ws, header_row: int) -> str:
    """Currency marked in the FOB / 合同金额 heading; USD when unmarked."""
    for c in range(1, ws.max_column + 1):
        head = _norm(ws.cell(header_row, c).value)
        if head.startswith(("fob", "合同金额")):
            if "£" in head:
                return "GBP"
            if "¥" in head or "rmb" in head:
                return "CNY"
    return "USD"


# ── colour flags ────────────────────────────────────────────────────────────

def _fill_key(cell) -> str:
    """Comparable key for a cell's fill, or '' when it has none."""
    f = getattr(cell, "fill", None)
    if f is None or f.patternType is None:
        return ""
    fg = f.fgColor
    rgb = getattr(fg, "rgb", None)
    if isinstance(rgb, str) and rgb not in ("00000000",):
        return rgb.upper()
    theme = getattr(fg, "theme", None)
    if theme is not None and not isinstance(theme, type(None)):
        return f"theme{theme}:{getattr(fg, 'tint', 0) or 0:.3f}"
    return ""


def _legend(ws, header_row: int) -> dict[str, str]:
    """Fill key → meaning, read from the sheet's own legend rows.

    The legend sits above the header (``合同已寄工厂`` / ``合同已回签`` /
    ``红底的是有折扣风险的``).  Reading the colours off it rather than
    hardcoding them means a re-themed workbook still parses.
    """
    out: dict[str, str] = {}
    for r in range(1, header_row):
        for c in range(1, min(ws.max_column, 8) + 1):
            text = _txt(ws.cell(r, c).value)
            if not text:
                continue
            key = _fill_key(ws.cell(r, c))
            if not key:
                continue
            if "寄工厂" in text:
                out[key] = "sent"
            elif "回签" in text:
                out[key] = "signed"
            elif "折扣" in text:
                out[key] = "discount_risk"
    return out


# ── sheet identity ──────────────────────────────────────────────────────────

_YEAR_RE = re.compile(r"(20\d{2})")


def sheet_identity(name: str) -> tuple[str, str]:
    """``(client, year)`` inferred from a sheet name.

    The workbook files one sheet per client per year and says so nowhere else
    — ``Zalando2026``, ``SC2025``, ``2025其它 客人大货``.
    """
    raw = str(name or "").strip()
    year_m = _YEAR_RE.search(raw)
    year = year_m.group(1) if year_m else ""
    client = _YEAR_RE.sub("", raw).strip()
    client = client.replace("客人大货", "").replace("大货", "").strip()
    return (client or "—", year)


# ── main entry points ───────────────────────────────────────────────────────

def parse_settlement(source: Any) -> dict[str, Any]:
    """Parse a settlement workbook.

    Returns ``{"rows": [...], "samples": [...], "sheets": [...],
    "skipped": [...]}``.  Every row carries the sheet it came from, so a
    re-import can replace one client's year without touching the rest.
    """
    wb = openpyxl.load_workbook(source, data_only=True)
    rows: list[dict[str, Any]] = []
    samples: list[dict[str, Any]] = []
    sheets: list[dict[str, Any]] = []
    skipped: list[str] = []

    for ws in wb.worksheets:
        title = str(ws.title).strip()
        if any(title.startswith(d) for d in DERIVED_SHEETS):
            skipped.append(f"{title} (recomputed, not imported)")
            continue
        if title.startswith(SAMPLE_SHEET):
            got = _parse_samples(ws)
            samples.extend(got)
            sheets.append({"sheet": title, "kind": "samples", "rows": len(got)})
            continue
        header_row = _find_header_row(ws)
        if header_row < 0:
            skipped.append(f"{title} (no settlement header)")
            continue
        got = _parse_sheet(ws, header_row, title)
        rows.extend(got)
        client, year = sheet_identity(title)
        sheets.append({"sheet": title, "kind": "settlement", "rows": len(got),
                       "client": client, "year": year,
                       "currency": _currency(ws, header_row)})

    if not rows and not samples:
        raise SettlementParseError(
            "No settlement sheet was found — expected a header row with "
            "合同号 and 款号 / PO# / INVOICE NO."
        )
    return {"rows": rows, "samples": samples, "sheets": sheets,
            "skipped": skipped}


def _parse_sheet(ws, header_row: int, title: str) -> list[dict[str, Any]]:
    cols = _map_columns(ws, header_row)
    legend = _legend(ws, header_row)
    client, year = sheet_identity(title)
    currency = _currency(ws, header_row)
    inv_col = cols.get("invoice_no", 1)
    con_col = cols.get("contract_no", 0)

    cell = cell_getter(ws, cols)

    out: list[dict[str, Any]] = []
    for r in range(header_row + 2, ws.max_row + 1):
        rec = {
            "sheet": title, "client": client, "year": year,
            "currency": currency, "row_no": r,
            "invoice_no":   _txt(cell(r, "invoice_no")),
            "po_number":    _txt(cell(r, "po_number")),
            "style":        _txt(cell(r, "style")),
            "contract_no":  _txt(cell(r, "contract_no")),
            "factory":      _txt(cell(r, "factory")),
            "repeat_or_new": _txt(cell(r, "repeat_or_new")),
            "contract_qty": _num(cell(r, "contract_qty")),
            "unit_cost":    _num(cell(r, "unit_cost")),
            "fob":          _num(cell(r, "fob")),
            "contract_amount": _num(cell(r, "contract_amount")),
            "ex_factory":   _date(cell(r, "ex_factory")),
            "shipped_qty":  _num(cell(r, "shipped_qty")),
            "over_short_pct": _num(cell(r, "over_short_pct")),
            "invoice_amount": _num(cell(r, "invoice_amount")),
            "received":     _num(cell(r, "received")),
            "received_date": _date(cell(r, "received_date")),
            "pay_fabric":   _num(cell(r, "pay_fabric")),
            "pay_fabric_date": _date(cell(r, "pay_fabric_date")),
            "pay_trim":     _num(cell(r, "pay_trim")),
            "pay_trim_date": _date(cell(r, "pay_trim_date")),
            "pay_cmt":      _num(cell(r, "pay_cmt")),
            "pay_cmt_date": _date(cell(r, "pay_cmt_date")),
            "fee_port":     _num(cell(r, "fee_port")),
            "fee_other1":   _num(cell(r, "fee_other1")),
            "fee_other2":   _num(cell(r, "fee_other2")),
        }
        if not any(rec[k] for k in _KEY_FIELDS):
            continue
        rec["discount_risk"] = int(
            legend.get(_fill_key(ws.cell(r, inv_col))) == "discount_risk")
        rec["contract_status"] = (
            legend.get(_fill_key(ws.cell(r, con_col)), "") if con_col else "")
        if rec["contract_status"] == "discount_risk":
            rec["contract_status"] = ""
        out.append(rec)
    return out


def _parse_samples(ws) -> list[dict[str, Any]]:
    """样品 sheet → per-style sample cost breakdown."""
    header_row = find_header_row(
        ws, lambda heads: "款号" in heads and "工厂" in heads, max_rows=12, norm=_norm)
    if header_row < 0:
        return []
    val = cell_getter(ws, header_index(ws, header_row, _norm))

    out: list[dict[str, Any]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        style = _txt(val(r, "款号"))
        if not style:
            continue
        out.append({
            "style":   style,
            "client":  _txt(val(r, "客人")),
            "factory": _txt(val(r, "工厂")),
            "fabric":  _num(val(r, "面料")),
            "lining":  _num(val(r, "里布")),
            "rib":     _num(val(r, "螺纹")),
            "trim":    _num(val(r, "辅料")),
            "total":   _num(val(r, "合计")),
            "row_no":  r,
        })
    return out


# ── derived figures ─────────────────────────────────────────────────────────

def outstanding(rec: dict[str, Any]) -> float:
    """Invoiced but not yet received, in the row's own currency.

    Falls back to the contract amount when nothing has been invoiced yet —
    the money is still owed, and a row that has shipped but not been invoiced
    would otherwise read as settled.
    """
    billed = rec.get("invoice_amount") or rec.get("contract_amount") or 0
    return round(float(billed) - float(rec.get("received") or 0), 2)


def cost_total(rec: dict[str, Any]) -> float:
    """Everything paid out against the line: the three payments plus fees."""
    keys = ("pay_fabric", "pay_trim", "pay_cmt",
            "fee_port", "fee_other1", "fee_other2")
    return round(sum(float(rec.get(k) or 0) for k in keys), 2)


def margin(rec: dict[str, Any]) -> float | None:
    """Money in less money out, or None when nothing has been paid out yet.

    Returning 0 for an un-costed line would read as break-even, which is a
    different claim from "not known yet".
    """
    out = cost_total(rec)
    if not out:
        return None
    earned = rec.get("received") or rec.get("invoice_amount") \
        or rec.get("contract_amount") or 0
    return round(float(earned) - out, 2)
