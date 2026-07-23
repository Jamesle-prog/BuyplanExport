"""Buyer DSP file → ``dspTrims[]`` rows for the CPRS export endpoints.

CPRS ≥1.6.16 accepts caller-provided trim facts (``dspTrims[]``) and turns
the doc-suite trim list DSP-first: DSP rows become the A section with
order-quantity formulas, CPRS rule rows become the B section marked
"以 DSP 为准" where they differ. CPRS deliberately does NOT read mailboxes
or files — extracting the DSP attachment and structuring it is the
caller's job. This module is that job.

Buyer DSP sheets vary in layout, so the header row is FOUND (first row
containing at least a style-ish and a material-ish header, fuzzy-matched
against bilingual synonyms) rather than assumed, and each recognised column
maps to one ``dspTrims`` field:

    style / materialName / materialCode / supplier / placement /
    qtyPerPc / color / appliesToOrders[]

Everything is passed through verbatim — no invention: a blank qty stays 0
(the API renders "按TP"), unknown columns are ignored, and rows without a
material name are reported, not guessed at.
"""
from __future__ import annotations

import io
import re

import openpyxl

# Field → lowercase header synonyms (English + Chinese, punctuation-insensitive).
_HEADER_SYNONYMS: dict[str, tuple[str, ...]] = {
    "style":        ("style", "style no", "style#", "styleno", "款号", "款式",
                     "style number"),
    "materialName": ("material name", "material", "trim", "trim name", "item",
                     "item name", "description", "name", "辅料", "辅料名称",
                     "品名", "描述", "名称"),
    "materialCode": ("material code", "item code", "code", "ref", "ref no",
                     "reference", "article", "article no", "料号", "编号",
                     "货号"),
    "supplier":     ("supplier", "vendor", "mill", "nominated supplier",
                     "供应商", "供货商"),
    "placement":    ("placement", "position", "location", "部位", "位置"),
    "qtyPerPc":     ("qty per pc", "qty/pc", "qty per piece", "qty", "usage",
                     "consumption", "用量", "单件用量", "单耗"),
    "color":        ("color", "colour", "颜色", "色号"),
    "appliesToOrders": ("po", "pos", "po no", "po number", "orders", "order",
                        "适用订单", "订单号", "po list"),
}


def _norm_header(v) -> str:
    s = str(v or "").strip().lower()
    return re.sub(r"[\s_./#:：（）()-]+", " ", s).strip()


def _match_field(header: str) -> str | None:
    h = _norm_header(header)
    if not h:
        return None
    for field, names in _HEADER_SYNONYMS.items():
        if h in names:
            return field
    # Tolerant second pass: a header that CONTAINS a synonym ("trim material
    # name (english)" → materialName). Longest synonym wins to keep
    # "material code" from matching the bare "material" of materialName.
    best: tuple[int, str] | None = None
    for field, names in _HEADER_SYNONYMS.items():
        for n in names:
            if n in h and (best is None or len(n) > best[0]):
                best = (len(n), field)
    return best[1] if best else None


def parse_dsp_trims(content: bytes) -> dict:
    """Parse a buyer DSP file — PDF (the usual format) or Excel — into CPRS
    ``dspTrims[]`` rows.

    Returns ``{"trims": [...], "issues": [str], "sheet": str}`` (``sheet``
    is the worksheet name for Excel, a page summary for PDF). Raises
    ``ValueError`` only for file-level problems (unreadable file, no
    recognisable trim table) — cell-level junk becomes an issue line, never
    an exception.
    """
    if content[:5] == b"%PDF-":
        return _parse_pdf(content)
    return _parse_xlsx(content)


def _parse_xlsx(content: bytes) -> dict:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True,
                                    read_only=True)
    except Exception as exc:
        raise ValueError(f"Not a readable Excel file: {exc}") from exc

    try:
        best = None   # (score, sheet_name, header_row_idx, {col_idx: field})
        for ws in wb.worksheets:
            for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30,
                                                     values_only=True), 1):
                cols: dict[int, str] = {}
                for c_idx, cell in enumerate(row or ()):
                    f = _match_field(cell)
                    if f and f not in cols.values():
                        cols[c_idx] = f
                fields = set(cols.values())
                # A usable header names a material plus a way to attach the
                # row to orders (style or explicit PO list).
                if "materialName" in fields and (
                        "style" in fields or "appliesToOrders" in fields):
                    score = len(fields)
                    if best is None or score > best[0]:
                        best = (score, ws.title, r_idx, cols)
            if best and best[0] >= 5:
                break   # good enough — don't scan further sheets

        if best is None:
            raise ValueError(
                "No sheet has a recognisable DSP header row (need at least a "
                "material/辅料 column plus a style/款号 or PO column).")

        _score, sheet_name, header_row, cols = best
        ws = wb[sheet_name]
        trims: list[dict] = []
        issues: list[str] = []
        for r_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1,
                                                 values_only=True),
                                    header_row + 1):
            vals = {f: row[c] if c < len(row) else None
                    for c, f in cols.items()}
            name = str(vals.get("materialName") or "").strip()
            style = str(vals.get("style") or "").strip()
            orders_raw = str(vals.get("appliesToOrders") or "").strip()
            if not any((name, style, orders_raw)):
                continue                      # blank spacer row
            if not name:
                issues.append(f"row {r_idx}: no material name — skipped")
                continue

            qty_raw = vals.get("qtyPerPc")
            qty: float = 0
            if qty_raw not in (None, ""):
                try:
                    qty = float(str(qty_raw).replace(",", "").strip())
                except (TypeError, ValueError):
                    issues.append(
                        f"row {r_idx} ({name}): qty {qty_raw!r} is not a "
                        f"number — sent as 0 (按TP)")
                    qty = 0

            trim = {
                "style":        style,
                "materialName": name,
                "materialCode": str(vals.get("materialCode") or "").strip(),
                "supplier":     str(vals.get("supplier") or "").strip(),
                "placement":    str(vals.get("placement") or "").strip(),
                "qtyPerPc":     qty,
                "color":        str(vals.get("color") or "").strip(),
            }
            orders = [o for o in re.split(r"[,;/\s]+", orders_raw) if o]
            if orders:
                trim["appliesToOrders"] = orders
            trims.append(trim)

        return {"trims": trims, "issues": issues, "sheet": sheet_name}
    finally:
        wb.close()


# ── PDF path (the usual buyer format) ───────────────────────────────────────

# A style number stated in page text ("STYLE: DU5105" / "款号 DU5105") — the
# style often lives in the PDF's page header rather than a table column.
_PAGE_STYLE_RE = re.compile(
    r"(?:style\s*(?:no\.?|number|#)?|款号|款式)\s*[:：#]?\s*"
    r"([A-Z][A-Z0-9]{2,}(?:[-/][A-Z0-9]+)*)",
    re.IGNORECASE)


def _page_style(text: str) -> str:
    m = _PAGE_STYLE_RE.search(text or "")
    return (m.group(1).strip().upper() if m else "")


def trims_from_pdf_pages(pages: list[dict]) -> dict:
    """Pure core of the PDF path — *pages* is
    ``[{"tables": [[row_cells…]…], "text": str}, …]`` (what pdfplumber
    extracts). Returns the same shape as :func:`parse_dsp_trims`.

    Header rows are matched with the same bilingual synonyms as the Excel
    path. A table with no header row of its own continues the previous
    table's columns (multi-page tables repeat or omit headers page by
    page). When a page has no style column, a style stated in the page text
    ("STYLE: DU5105") applies to that page's rows — extraction of a stated
    fact, not invention.
    """
    trims: list[dict] = []
    issues: list[str] = []
    carry_cols: dict[int, str] | None = None
    pages_used = 0

    for p_idx, page in enumerate(pages, 1):
        default_style = _page_style(page.get("text") or "")
        page_hit = False
        for table in (page.get("tables") or []):
            rows = [r or [] for r in (table or []) if r]
            if not rows:
                continue
            # Find the header row inside the first few rows of this table.
            cols: dict[int, str] | None = None
            start = 0
            for i, row in enumerate(rows[:4]):
                cand: dict[int, str] = {}
                for c_idx, cell in enumerate(row):
                    f = _match_field(cell)
                    if f and f not in cand.values():
                        cand[c_idx] = f
                if "materialName" in cand.values():
                    cols, start = cand, i + 1
                    break
            if cols is None:
                if carry_cols is None:
                    continue            # decorative table before any header
                cols, start = carry_cols, 0   # continuation page
            else:
                carry_cols = cols
            page_hit = True

            for r_off, row in enumerate(rows[start:], start + 1):
                vals = {f: (row[c] if c < len(row) else None)
                        for c, f in cols.items()}
                name = str(vals.get("materialName") or "").strip()
                style = str(vals.get("style") or "").strip() or default_style
                orders_raw = str(vals.get("appliesToOrders") or "").strip()
                if not any((name, style, orders_raw)):
                    continue
                if not name:
                    continue            # spacer/subtotal line in the grid
                # A repeated header row inside a continuation table.
                if _match_field(name) == "materialName":
                    continue

                qty_raw = vals.get("qtyPerPc")
                qty: float = 0
                if qty_raw not in (None, ""):
                    try:
                        qty = float(str(qty_raw).replace(",", "").strip())
                    except (TypeError, ValueError):
                        issues.append(
                            f"page {p_idx} row {r_off} ({name}): qty "
                            f"{qty_raw!r} is not a number — sent as 0 (按TP)")
                        qty = 0

                trim = {
                    "style":        style,
                    "materialName": name,
                    "materialCode": str(vals.get("materialCode") or "").strip(),
                    "supplier":     str(vals.get("supplier") or "").strip(),
                    "placement":    str(vals.get("placement") or "").strip(),
                    "qtyPerPc":     qty,
                    "color":        str(vals.get("color") or "").strip(),
                }
                orders = [o for o in re.split(r"[,;/\s]+", orders_raw) if o]
                if orders:
                    trim["appliesToOrders"] = orders
                trims.append(trim)
        if page_hit:
            pages_used += 1

    if not trims:
        raise ValueError(
            "No recognisable trim table found in the PDF (need a table with "
            "a material/辅料 column). If the DSP is a scanned image rather "
            "than a text PDF, export it to Excel first.")
    return {"trims": trims, "issues": issues,
            "sheet": f"PDF · {pages_used} page(s)"}


def _parse_pdf(content: bytes) -> dict:
    """pdfplumber glue around :func:`trims_from_pdf_pages`."""
    try:
        import pdfplumber
    except Exception as exc:                     # pragma: no cover
        raise ValueError(f"PDF support unavailable: {exc}") from exc
    try:
        pages: list[dict] = []
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            for page in pdf.pages:
                try:
                    tables = page.extract_tables() or []
                except Exception:
                    tables = []
                try:
                    text = page.extract_text() or ""
                except Exception:
                    text = ""
                pages.append({"tables": tables, "text": text})
    except ValueError:
        raise
    except Exception as exc:
        raise ValueError(f"Not a readable PDF: {exc}") from exc
    return trims_from_pdf_pages(pages)


def trims_for_request(trims: list[dict], rq: dict) -> list[dict]:
    """The subset of *trims* belonging to one export request (order context).

    Routing is data plumbing, not a business rule: a trim belongs to a
    context when its explicit ``appliesToOrders`` intersects the context's
    PO register, or (no explicit list) when its style appears there. Trims
    without style/orders match every context — they're context-free facts
    the caller chose to send.
    """
    pos = rq.get("pos") or []
    ctx_orders = {str(p.get("order") or "").strip() for p in pos} - {""}
    ctx_styles = {str(p.get("style") or "").strip() for p in pos} - {""}
    out = []
    for tr in trims:
        orders = {str(o).strip() for o in (tr.get("appliesToOrders") or [])}
        if orders:
            if orders & ctx_orders:
                out.append(tr)
        elif tr.get("style"):
            if str(tr["style"]).strip() in ctx_styles:
                out.append(tr)
        else:
            out.append(tr)
    return out
