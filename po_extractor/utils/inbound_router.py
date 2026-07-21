"""Identify what an emailed spreadsheet actually is, and summarise it.

An inbound attachment could be any of the four files the system already
knows how to read. Rather than making the sender label it (they won't),
each candidate parser is tried in turn — cheapest and most specific first —
and the first confident match wins.

Detection is *structural* (sheet names, header cells), never filename-based:
factories rename files constantly, and a filename is not evidence.

Nothing here writes to the database. It returns a description of what the
file is and what it would change, so the UI can show a review before a
human applies it — the same shape as the fabric upload review gate.
"""
from __future__ import annotations

import io

import openpyxl

# Kind constants — stable strings stored on the inbox row.
KIND_PROGRESS_FORM = "progress_form"     # 进度回报表 (factory quantity + milestones)
KIND_BUYPLAN       = "buyplan_index"     # a returned buy plan's Index tab
KIND_HHN_PROGRESS  = "hhn_progress"      # 大货进度表
KIND_FABRIC_LIST   = "fabric_list"       # 面料统计表
KIND_UNKNOWN       = "unknown"

KIND_LABELS: dict[str, str] = {
    KIND_PROGRESS_FORM: "进度回报表 Factory progress form",
    KIND_BUYPLAN:       "采购计划 Returned buy plan",
    KIND_HHN_PROGRESS:  "大货进度表 HHN progress",
    KIND_FABRIC_LIST:   "面料统计表 Fabric list",
    KIND_UNKNOWN:       "Unrecognised spreadsheet",
}


def _sheet_names(content: bytes) -> list[str]:
    """Sheet names only — a cheap read that avoids parsing whole workbooks
    just to classify them."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()
    except Exception:
        return []


def _first_row_texts(content: bytes, sheet: str, limit: int = 30) -> set[str]:
    """Header-ish cell texts from the first rows of *sheet*."""
    out: set[str] = set()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        try:
            if sheet not in wb.sheetnames:
                return out
            ws = wb[sheet]
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= limit:
                    break
                for cell in row:
                    if cell is not None:
                        out.add(str(cell).strip())
        finally:
            wb.close()
    except Exception:
        pass
    return out


def detect_kind(content: bytes) -> str:
    """Classify an emailed workbook. Returns one of the KIND_* constants."""
    if not content:
        return KIND_UNKNOWN
    names = _sheet_names(content)
    if not names:
        return KIND_UNKNOWN
    lowered = {n.strip().lower() for n in names}

    # Our own generated forms are unambiguous — they carry known sheet names.
    if any("进度回报" in n or "里程碑" in n for n in names):
        return KIND_PROGRESS_FORM
    if "index" in lowered:
        return KIND_BUYPLAN
    if "all" in lowered:                       # 面料统计表's data sheet
        return KIND_FABRIC_LIST

    # 大货进度表 has no fixed sheet name; recognise it by its header cells.
    for sheet in names[:3]:
        texts = _first_row_texts(content, sheet)
        if any("客人PC" in x or "所在PO" in x for x in texts) and any(
                "款式" in x or "款号" in x for x in texts):
            return KIND_HHN_PROGRESS
    return KIND_UNKNOWN


def summarise(content: bytes, kind: str) -> dict:
    """Parse *content* far enough to describe what importing it would do.

    Returns ``{"ok": bool, "summary": str, "detail": list[str],
    "error": str}`` — never raises, because this runs while rendering an
    inbox listing and one bad attachment must not break the page.
    """
    try:
        if kind == KIND_PROGRESS_FORM:
            from ..exporters.factory_progress_form import parse_progress_report_xlsx
            parsed = parse_progress_report_xlsx(content)
            n_rep = len(parsed.get("reports") or [])
            n_ms = len(parsed.get("milestones") or [])
            return {
                "ok": bool(n_rep or n_ms),
                "summary": f"{n_rep} quantity report(s), {n_ms} milestone update(s)",
                "detail": (parsed.get("issues") or [])[:10],
                "error": "",
            }

        if kind == KIND_BUYPLAN:
            from ..exporters.factory_progress_form import parse_buyplan_index_tracking
            parsed = parse_buyplan_index_tracking(content)
            rows = parsed.get("rows") or []
            n_dates = sum(len(r.get("planned") or {}) for r in rows)
            return {
                "ok": bool(rows),
                "summary": f"{len(rows)} row(s) with tracking data, {n_dates} date(s)",
                "detail": [f"{r['pc_no']} · {r['style']}" for r in rows[:10]],
                "error": "",
            }

        if kind == KIND_HHN_PROGRESS:
            from ..lookups.progress_lookup import parse_progress_rows
            records = parse_progress_rows(content)
            return {
                "ok": bool(records),
                "summary": f"{len(records)} progress record(s)",
                "detail": [
                    f"{r.get('pc_no', '')} · {r.get('style_display') or r.get('style', '')}"
                    for r in records[:10]
                ],
                "error": "",
            }

        if kind == KIND_FABRIC_LIST:
            names = _sheet_names(content)
            return {
                "ok": "all" in {n.lower() for n in names},
                "summary": "Fabric list — imports through the existing approval queue",
                "detail": [f"sheets: {', '.join(names[:5])}"],
                "error": "",
            }

        return {"ok": False, "summary": "Not a file this system recognises",
                "detail": [], "error": ""}
    except Exception as exc:                     # never break the listing
        return {"ok": False, "summary": "Could not be read",
                "detail": [], "error": str(exc)[:300]}
