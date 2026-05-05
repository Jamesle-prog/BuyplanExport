"""Generate the 1.1.PO_Client mapping template sheet.

The sheet has two header rows:
  Row 1 — client's original column headings (filled in by the user per client)
  Row 2 — internal standardised column names (fixed, do not change)
  Row 3+ — paste client data here

Usage:
    python -m po_extractor.utils.client_template --output my_template.xlsx
    python -m po_extractor.utils.client_template --output my_template.xlsx --client Zalando
"""
from __future__ import annotations

import argparse
import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side, PatternFill,
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Internal column schema — in display order
# Each tuple: (internal_header, example_client_header, group, notes)
# ---------------------------------------------------------------------------
SCHEMA = [
    # ── Core PO / style identifiers ──────────────────────────────────────────
    ("Purchase Order Number",           "PO Number",            "PO",       "Required"),
    ("Main Supplier Config SKU",        "Style #",              "PO",       "Required — groups rows into one buy-plan sheet"),
    ("Config SKU",                      "Config SKU",           "PO",       ""),
    ("Article Name",                    "Description",          "PO",       ""),
    ("Brand",                           "Brand",                "PO",       ""),
    ("Main Supplier Color Description", "Color Description",    "PO",       ""),
    # ── Size quantities ───────────────────────────────────────────────────────
    ("XS",                              "XS",                   "Size",     ""),
    ("S",                               "S",                    "Size",     ""),
    ("M",                               "M",                    "Size",     ""),
    ("L",                               "L",                    "Size",     ""),
    ("XL",                              "XL",                   "Size",     ""),
    ("XXL",                             "XXL",                  "Size",     ""),
    ("TOTAL QUANTITY",                  "Total Qty",            "Size",     ""),
    # ── Internal / Chinese fields ─────────────────────────────────────────────
    ("颜色",                             "Color (CN)",           "内部",     "Chinese color name"),
    ("辅助色",                           "Secondary Color",      "内部",     ""),
    ("合同号",                           "Contract No.",         "内部",     ""),
    ("入厂时间",                         "Factory Entry Date",   "内部",     ""),
    ("质量要求",                         "Quality Requirements", "内部",     ""),
    ("备注1",                            "Note 1",               "内部",     ""),
    ("备注2",                            "Note 2",               "内部",     ""),
    ("备注3",                            "Note 3",               "内部",     ""),
    ("备注4",                            "Note 4",               "内部",     ""),
    # ── Fabric ────────────────────────────────────────────────────────────────
    ("面料_面料",                        "Fabric 1",             "面料",     ""),
    ("面料_面料_编号",                   "Fabric 1 Code",        "面料",     "Used to group Template_P files"),
    ("面料_面料_部位",                   "Fabric 1 Body Part",   "面料",     ""),
    ("面料_面料1",                       "Fabric 2",             "面料",     ""),
    ("面料_面料1_编号",                  "Fabric 2 Code",        "面料",     ""),
    ("面料_面料1_部位",                  "Fabric 2 Body Part",   "面料",     ""),
    ("面料_面料2",                       "Fabric 3",             "面料",     ""),
    ("面料_面料2_编号",                  "Fabric 3 Code",        "面料",     ""),
    ("面料_面料2_部位",                  "Fabric 3 Body Part",   "面料",     ""),
    # ── Photos ────────────────────────────────────────────────────────────────
    ("Photo1",                          "Photo Path 1",         "Photo",    "Full file path"),
    ("Photo2",                          "Photo Path 2",         "Photo",    "Full file path"),
]

# Known client header aliases — extend as new clients are on-boarded
CLIENT_ALIASES: dict[str, dict[str, str]] = {
    "Zalando": {
        "Purchase Order Number":            "Purchase Order Number",
        "Main Supplier Config SKU":         "Main Supplier Config SKU",
        "Config SKU":                       "Config SKU",
        "Article Name":                     "Article Name",
        "Brand":                            "Brand",
        "Main Supplier Color Description":  "Main Supplier Color Description",
        "XS":   "XS",
        "S":    "S",
        "M":    "M",
        "L":    "L",
        "XL":   "XL",
        "XXL":  "XXL",
        "TOTAL QUANTITY":                   "Qty Ordered aCanc & aClosing (PO)",
    },
    # Add more clients here:
    # "ClientName": { internal_header: client_header, ... }
}

# Colour palette
_FILL_CLIENT   = PatternFill(start_color="FFDCE6F1", end_color="FFDCE6F1", fill_type="solid")   # light blue
_FILL_INTERNAL = PatternFill(start_color="FFD9EAD3", end_color="FFD9EAD3", fill_type="solid")   # light green
_FILL_GROUP: dict[str, PatternFill] = {
    "PO":    PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid"),
    "Size":  PatternFill(start_color="FFFFE0E0", end_color="FFFFE0E0", fill_type="solid"),
    "内部":  PatternFill(start_color="FFEDDDF4", end_color="FFEDDDF4", fill_type="solid"),
    "面料":  PatternFill(start_color="FFFDECDC", end_color="FFFDECDC", fill_type="solid"),
    "Photo": PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid"),
}
_FONT_BOLD  = Font(bold=True)
_FONT_NOTE  = Font(italic=True, color="FF666666", size=9)
_ALIGN_C    = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _thin_border():
    s = Side(border_style="thin", color="FFBBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)


# ---------------------------------------------------------------------------
# Public helpers
# ---------------------------------------------------------------------------

def build_mapping_sheet(
    ws,
    client: str | None = None,
    *,
    include_notes_row: bool = True,
) -> None:
    """Write the two header rows (+ optional notes row) to an openpyxl worksheet."""

    aliases = CLIENT_ALIASES.get(client or "", {})
    num_cols = len(SCHEMA)

    # ── Row 1: client headers ─────────────────────────────────────────────────
    for col_idx, (internal, example, group, _) in enumerate(SCHEMA, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = aliases.get(internal) or example
        cell.fill = _FILL_CLIENT
        cell.font = _FONT_BOLD
        cell.alignment = _ALIGN_C
        cell.border = _thin_border()

    # ── Row 2: internal headers ───────────────────────────────────────────────
    for col_idx, (internal, example, group, _) in enumerate(SCHEMA, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = internal
        cell.fill = _FILL_GROUP.get(group, _FILL_INTERNAL)
        cell.font = _FONT_BOLD
        cell.alignment = _ALIGN_C
        cell.border = _thin_border()

    # ── Row 3: column notes (optional, italic) ────────────────────────────────
    if include_notes_row:
        for col_idx, (internal, example, group, note) in enumerate(SCHEMA, start=1):
            cell = ws.cell(row=3, column=col_idx)
            cell.value = note or ""
            cell.font = _FONT_NOTE
            cell.alignment = _ALIGN_C

    # ── Column widths ─────────────────────────────────────────────────────────
    SIZE_COLS = {"XS", "S", "M", "L", "XL", "XXL", "TOTAL QUANTITY"}
    for col_idx, (internal, *_) in enumerate(SCHEMA, start=1):
        ltr = get_column_letter(col_idx)
        if internal in SIZE_COLS:
            ws.column_dimensions[ltr].width = 9
        elif internal.startswith("面料"):
            ws.column_dimensions[ltr].width = 18
        elif internal.startswith("Photo"):
            ws.column_dimensions[ltr].width = 35
        else:
            ws.column_dimensions[ltr].width = 22

    # ── Freeze panes below the two header rows ────────────────────────────────
    ws.freeze_panes = "A3" if not include_notes_row else "A4"

    # ── Legend in a far-right column ─────────────────────────────────────────
    legend_col = num_cols + 2
    ltr = get_column_letter(legend_col)
    ws.column_dimensions[ltr].width = 28
    ws.cell(row=1, column=legend_col, value="Row 1 = client's column headers").font = _FONT_NOTE
    ws.cell(row=2, column=legend_col, value="Row 2 = internal names (do NOT edit)").font = Font(bold=True, size=9)
    if include_notes_row:
        ws.cell(row=3, column=legend_col, value="Row 3 = notes (delete before use)").font = _FONT_NOTE
    ws.cell(row=4, column=legend_col, value="Paste client data from row 3 / 4 down").font = _FONT_NOTE


def create_template(output_path: str, client: str | None = None) -> str:
    """Create a blank mapping template Excel and return the saved path."""
    wb = Workbook()
    ws = wb.active
    ws.title = "1.1.PO_Client"
    ws.sheet_view.showGridLines = True

    build_mapping_sheet(ws, client=client)

    wb.save(output_path)
    return output_path


def inject_mapping_sheet(
    workbook_path: str,
    client: str | None = None,
    sheet_name: str = "1.1.PO_Client",
) -> None:
    """Add (or replace) the mapping sheet in an existing workbook."""
    wb = load_workbook(workbook_path)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name, 0)
    build_mapping_sheet(ws, client=client)
    wb.save(workbook_path)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _cli():
    ap = argparse.ArgumentParser(description="Generate 1.1.PO_Client mapping template")
    ap.add_argument("--output", "-o", required=True, help="Output .xlsx path")
    ap.add_argument("--client", "-c", default=None,
                    choices=list(CLIENT_ALIASES.keys()) + [None],
                    help="Pre-fill row 1 with known client headers")
    args = ap.parse_args()
    path = create_template(args.output, client=args.client)
    print(f"Template saved: {path}")


if __name__ == "__main__":
    _cli()
