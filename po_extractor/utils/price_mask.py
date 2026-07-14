"""Mask prices in PDF and Excel files.

PDF: uses PyMuPDF redaction — scans every page for tokens that look like
prices (digits.digits) and covers them with white-filled redaction rectangles.

Excel: uses openpyxl — detects columns whose headers contain price-related
keywords and replaces numeric cell values with "***".

Optional AI assist (DeepSeek): when an API key is supplied, the model is asked
to identify prices/price-columns from context and its findings are UNIONED with
the pattern/keyword detection — never replacing it. So AI can catch
context-dependent prices the pattern misses (a whole-dollar FOB, an odd format),
while the regex still guarantees the obvious ones even if AI is off or errors.

Output is saved to output_dir/masked/<original_filename>.
"""
import os
import re

# A PDF token counts as a maskable amount when it has exactly two fraction
# digits, optionally led by a currency symbol / comma thousands separators (the
# integer part may be absent, so a leading-dot form like ".75%" in prose is
# caught too) and optionally FOLLOWED BY '%' — so the cost figures (unit cost,
# extended cost) AND the discount / tariff percentages are all covered wherever
# they appear. The two-decimal requirement keeps it conservative — it won't grab
# bare integers (quantities, UPCs, PO numbers).
#   matches: 4.17 · 69.00 · 1,234.00 · $4.17 · €1,000.00 · 0.75% · .75% · 000.00
_PRICE_RE = re.compile(r'^[$£€¥]?[\d,]*\.\d{2}%?$')

# Header keywords that identify CONFIDENTIAL "price" columns in Excel sheets.
# Any column whose header (first 25 rows) contains one of these has its NUMERIC
# cells masked — text cells (names, headers) are left intact, so broad keywords
# are safe. Covers the cost/FOB headers this system's own exports use plus
# currency-symbol headers like "Line Total ($)". Retail prices (MSRP/SRP/RRP)
# are PUBLIC and handled by _RETAIL_EXCLUDE below — they are never masked.
_PRICE_KEYWORDS = (
    "fob", "cost", "price", "usd", "amount", "total cost", "unit price",
    "wholesale", "extended", "line total",
    "$", "€", "£", "¥",
)

# Retail prices are PUBLIC (printed on the hangtag) — never confidential, so
# they must not be masked even when the header also contains a generic price
# word (e.g. "Suggested Retail Price" contains "price"). A column whose header
# matches any of these is excluded from masking, overriding keyword/AI matches.
_RETAIL_EXCLUDE = ("msrp", "srp", "rrp", "retail")


# ---------------------------------------------------------------------------
# AI-assisted detection (DeepSeek) — optional, augments the pattern/keywords
# ---------------------------------------------------------------------------

def _norm_price_token(s) -> str:
    """Strip currency symbols / thousands commas / spaces for token matching."""
    return str(s).translate({ord(ch): None for ch in "$£€¥, "}).strip()


def _price_float(s):
    """Numeric value of a price token ('$1,234.00' → 1234.0), or None when it
    isn't a plain number. Lets '59' and '59.00' compare equal, so a known retail
    price (MSRP) matches its PDF token regardless of trailing-zero formatting."""
    try:
        return float(_norm_price_token(s))
    except (ValueError, TypeError):
        return None


# The PO labels its public retail price ("MSRP: $54.00", "SRP $59"). Extract it
# straight from the file's own text so it is kept visible even when the parser
# didn't capture an MSRP (only the KL path does) — parser-independent.
_RETAIL_LABEL_RE = re.compile(r'(?:MSRP|SRP|RRP)\b\s*[:=]?\s*\$?\s*([\d,]*\.?\d+)',
                              re.IGNORECASE)


def _retail_values_from_text(text) -> set:
    """Numeric values of any MSRP/SRP/RRP-labelled retail prices in *text*."""
    out: set = set()
    for m in _RETAIL_LABEL_RE.finditer(str(text or "")):
        pf = _price_float(m.group(1))
        if pf is not None:
            out.add(pf)
    return out


_AI_PRICE_SYSTEM = (
    "You find CONFIDENTIAL monetary costs in purchase-order text — the amounts a "
    "vendor must not reveal downstream. Return strict JSON "
    '{"prices": [<exact substrings>]} listing every token that is a cost, FOB, '
    "wholesale price, unit cost, or extended/line total, written EXACTLY as it "
    "appears (keep currency symbols and separators). Do NOT include the RETAIL "
    "price — MSRP, SRP, RRP, or suggested/recommended retail — which is PUBLIC "
    "and must stay visible. Also exclude quantities, UPC/EAN barcodes, PO/order/"
    "style numbers, dates, sizes, percentages, and measurements. If there are "
    'none, return {"prices": []}.'
)

_AI_COLUMN_SYSTEM = (
    "You are given column headers from a purchase-order spreadsheet. Return "
    'strict JSON {"price_headers": [<headers>]} listing exactly the headers '
    "that denote a CONFIDENTIAL monetary cost column (FOB, cost, unit cost, "
    "wholesale, extended/line total, amount). Do NOT include RETAIL price "
    "columns (MSRP, SRP, RRP, suggested retail) — those are public. Exclude "
    "quantity, size, UPC, style, PO, and date columns. Copy each header string "
    "exactly as given."
)


def _deepseek_json(system: str, user: str, api_key: str, model: str) -> dict:
    """One DeepSeek JSON call; returns {} on any failure (never raises)."""
    if not (api_key and user.strip()):
        return {}
    try:
        import json
        from openai import OpenAI
        from .deepseek_client import chat_kwargs, max_tokens_for
        client = OpenAI(api_key=api_key, base_url="https://api.deepseek.com")
        resp = client.chat.completions.create(
            model=model,
            messages=[{"role": "system", "content": system},
                      {"role": "user", "content": user[:12000]}],
            max_tokens=max_tokens_for(model, 1024),
            response_format={"type": "json_object"},
            **chat_kwargs(model),
        )
        return json.loads(resp.choices[0].message.content or "{}")
    except Exception:
        return {}   # graceful: AI failure never breaks masking (regex is the floor)


def detect_prices_ai(text: str, api_key: str, model: str = "deepseek-chat") -> set[str]:
    """Return the set of normalized price tokens the model found in *text*."""
    data = _deepseek_json(_AI_PRICE_SYSTEM, text, api_key, model)
    prices = data.get("prices", []) if isinstance(data, dict) else []
    return {_norm_price_token(p) for p in prices if str(p).strip()} - {""}


def detect_price_headers_ai(headers: list[str], api_key: str,
                            model: str = "deepseek-chat") -> set[str]:
    """Return the subset of *headers* the model classifies as price columns."""
    if not headers:
        return set()
    data = _deepseek_json(_AI_COLUMN_SYSTEM, "\n".join(headers), api_key, model)
    found = data.get("price_headers", []) if isinstance(data, dict) else []
    return {str(h).strip() for h in found if str(h).strip()}


# ---------------------------------------------------------------------------
# PDF masking
# ---------------------------------------------------------------------------

def _pdf_text(pdf_path: str) -> str:
    """Full text of a PDF (kept separate so the AI-detection pass can run
    without holding a PyMuPDF document open across threads)."""
    import fitz
    doc = fitz.open(pdf_path)
    try:
        return "\n".join(page.get_text() for page in doc)
    finally:
        doc.close()


def mask_prices(pdf_path: str, output_dir: str,
                api_key: str | None = None, model: str = "deepseek-chat",
                ai_prices: set[str] | None = None,
                keep: "set | list | None" = None) -> str:
    """Write a price-redacted copy of a PDF; return the output path.

    When *api_key* is given, DeepSeek-detected prices are unioned with the
    pattern matches (AI augments, never replaces the regex). Pass *ai_prices* to
    supply an already-computed AI price set (e.g. from a parallelised batch) and
    skip the per-file model call.

    *keep* is a set of retail prices (e.g. the PO's MSRP) that must NEVER be
    redacted — retail prices are public. A token whose numeric value equals a
    keep value is left visible even if the regex/AI would otherwise mask it.
    """
    import fitz  # PyMuPDF — loaded lazily so Excel-only callers don't need it

    masked_dir = os.path.join(output_dir, "masked")
    os.makedirs(masked_dir, exist_ok=True)
    out_path = os.path.join(masked_dir, os.path.basename(pdf_path))

    keep_floats = {v for v in (_price_float(k) for k in (keep or ())) if v is not None}

    doc = fitz.open(pdf_path)
    try:
        full_text = "\n".join(page.get_text() for page in doc)
        # Keep any retail price the PO labels (MSRP/SRP/RRP) — public, and works
        # even when the parser captured no MSRP (so the caller passed no keep).
        keep_floats |= _retail_values_from_text(full_text)

        ai_norm: set[str] = set(ai_prices) if ai_prices is not None else set()
        if api_key and ai_prices is None:
            ai_norm = detect_prices_ai(full_text, api_key, model)

        for page in doc:
            for word in page.get_text("words"):
                token = word[4]
                pf = _price_float(token)
                if pf is not None and pf in keep_floats:
                    continue        # retail price (MSRP/…) — public, never mask
                if _PRICE_RE.match(token) or (ai_norm and _norm_price_token(token) in ai_norm):
                    rect = fitz.Rect(word[:4])
                    page.add_redact_annot(rect, fill=(1, 1, 1))
            page.apply_redactions()
        doc.save(out_path)
    finally:
        doc.close()

    return out_path


def mask_prices_batch(pdf_paths: list[str], output_dir: str,
                      errors: list[str] | None = None,
                      api_key: str | None = None,
                      model: str = "deepseek-chat",
                      on_progress=None,
                      keep: "set | list | None" = None) -> list[str]:
    """Mask prices in a list of PDFs; return output paths.

    Failures are appended to *errors* (if given) so UI callers can surface
    them — a console print is invisible in Streamlit and the file is just
    missing from the result. *api_key* enables the AI assist (see mask_prices).

    The AI price detection is a slow network call per file, so with several
    files it runs CONCURRENTLY (the real bottleneck), while all PyMuPDF work
    stays serial — fitz is not thread-safe. *on_progress(done, total)* is called
    as detection completes so the UI can show live per-file progress.
    """
    total = len(pdf_paths)

    def _tick(done: int) -> None:
        if on_progress:
            try:
                on_progress(done, total)
            except Exception:
                pass

    # Concurrent AI detection (network-bound) → {path: price set}; fitz work
    # below stays single-threaded.
    ai_by_path: dict[str, set] = {}
    if api_key and total > 1:
        texts: dict[str, str] = {}
        for path in pdf_paths:
            try:
                texts[path] = _pdf_text(path)
            except Exception:
                texts[path] = ""
        from concurrent.futures import ThreadPoolExecutor, as_completed
        done = 0
        with ThreadPoolExecutor(max_workers=min(6, total)) as ex:
            futs = {ex.submit(detect_prices_ai, texts.get(p, ""), api_key, model): p
                    for p in pdf_paths}
            for fut in as_completed(futs):
                p = futs[fut]
                try:
                    ai_by_path[p] = fut.result()
                except Exception:
                    ai_by_path[p] = set()
                done += 1
                _tick(done)

    results = []
    for i, path in enumerate(pdf_paths, 1):
        try:
            out = mask_prices(path, output_dir, api_key=api_key, model=model,
                              ai_prices=ai_by_path.get(path), keep=keep)
            results.append(out)
            print(f"masked: {out}")
        except Exception as e:
            msg = f"{os.path.basename(path)}: {e}"
            if errors is not None:
                errors.append(msg)
            print(f"  mask FAILED: {msg}")
        if not (api_key and total > 1):
            _tick(i)     # serial path: report as each file is redacted
    return results


# ---------------------------------------------------------------------------
# Excel masking
# ---------------------------------------------------------------------------

def mask_prices_excel(xlsx_path: str, output_dir: str,
                      price_keywords: tuple = _PRICE_KEYWORDS,
                      api_key: str | None = None,
                      model: str = "deepseek-chat") -> str:
    """Write a price-redacted copy of an xlsx; return the output path.

    Detection strategy
    ------------------
    * Scans the first 25 rows of each sheet for cells whose text contains
      any of *price_keywords* (case-insensitive).
    * When *api_key* is given, DeepSeek classifies the sheet's header strings
      and any it calls price columns are unioned in (AI augments the keywords).
    * All columns that matched are treated as price columns.
    * In those columns, every numeric cell value (int / float) is replaced
      with the literal string ``"***"``.  String cells (headers) are left
      untouched so column headers remain readable.
    """
    import openpyxl

    ext = os.path.splitext(xlsx_path)[1].lower()
    if ext == ".xls":
        # openpyxl cannot read the legacy binary format; failing loudly here
        # beats the old behavior (an exception swallowed into a console
        # print, with the file silently missing from the masked zip).
        raise ValueError(
            "legacy .xls files cannot be price-masked — re-save as .xlsx/.xlsm"
        )

    masked_dir = os.path.join(output_dir, "masked")
    os.makedirs(masked_dir, exist_ok=True)
    out_path = os.path.join(masked_dir, os.path.basename(xlsx_path))

    # keep_vba: without it a .xlsm is silently re-packaged as a plain xlsx
    # under the .xlsm name — Excel then refuses to open the output.
    wb = openpyxl.load_workbook(xlsx_path, keep_vba=(ext == ".xlsm"))
    for ws in wb.worksheets:
        if ws.max_row is None or ws.max_column is None:
            continue

        # ── Detect price columns from first 25 rows ───────────────────────────
        price_cols: set[int] = set()
        retail_cols: set[int] = set()   # public retail (MSRP/SRP/…) — never mask
        header_by_col: dict[int, str] = {}
        scan_rows = min(25, ws.max_row)
        for r in range(1, scan_rows + 1):
            for c in range(1, ws.max_column + 1):
                val = ws.cell(row=r, column=c).value
                if val is None:
                    continue
                text = str(val)
                val_lower = text.lower().replace("\n", " ")
                if any(rk in val_lower for rk in _RETAIL_EXCLUDE):
                    retail_cols.add(c)
                if any(kw in val_lower for kw in price_keywords):
                    price_cols.add(c)
                # Remember a header-ish string per column for the AI pass —
                # a non-numeric cell in the scan band.
                if c not in header_by_col:
                    try:
                        float(text.replace(",", "").strip())
                    except (ValueError, TypeError):
                        header_by_col[c] = text.strip().replace("\n", " ")

        # ── AI assist: union in columns the model calls prices ────────────────
        if api_key and header_by_col:
            ai_headers = detect_price_headers_ai(
                list(header_by_col.values()), api_key, model)
            if ai_headers:
                for col, hdr in header_by_col.items():
                    if hdr in ai_headers:
                        price_cols.add(col)

        # Retail-price columns are public — exclusion wins over any match above.
        price_cols -= retail_cols

        if not price_cols:
            continue

        # ── Redact numeric values in price columns ────────────────────────────
        for r in range(1, ws.max_row + 1):
            for c in price_cols:
                cell = ws.cell(row=r, column=c)
                if cell.value is None:
                    continue
                if cell.data_type == "f":
                    # A price formula (=Dn*En, cross-sheet ref, …) would parse
                    # as a string and survive — but Excel recalculates it on
                    # open, leaking the price from a file believed redacted.
                    cell.value = "***"
                elif isinstance(cell.value, (int, float)):
                    cell.value = "***"
                else:
                    # String value in a price column: parse as a number after
                    # stripping currency symbols / thousands commas / spaces,
                    # so "$4.17", "1,234.00", "€1000" also mask. Header text and
                    # genuinely non-numeric strings fail the parse and stay.
                    cleaned = str(cell.value).translate(
                        {ord(ch): None for ch in "$£€¥, "}
                    ).strip()
                    try:
                        float(cleaned)
                        cell.value = "***"
                    except (ValueError, TypeError):
                        pass

    wb.save(out_path)
    wb.close()
    return out_path


def mask_prices_excel_batch(xlsx_paths: list[str], output_dir: str,
                            errors: list[str] | None = None,
                            api_key: str | None = None,
                            model: str = "deepseek-chat") -> list[str]:
    """Mask prices in a list of Excel files; return output paths.

    Failures are appended to *errors* (if given) so UI callers can surface
    them — see mask_prices_batch. *api_key* enables the AI column assist.
    """
    results = []
    for path in xlsx_paths:
        try:
            out = mask_prices_excel(path, output_dir, api_key=api_key, model=model)
            results.append(out)
            print(f"masked: {out}")
        except Exception as e:
            msg = f"{os.path.basename(path)}: {e}"
            if errors is not None:
                errors.append(msg)
            print(f"  mask FAILED: {msg}")
    return results
