"""SQLite-backed store for the fabric master database (面料统计表).

Imports from the '面料统计表.xlsx' 'all' sheet and exposes lookup by
公司面料编号 (Quality No.).

Schema DDL, header-alias tables, column-map fallback, and low-level row
helpers are kept in _fabric_master_schema.py to separate static definitions
from the store class logic.
"""
from __future__ import annotations

import json
import sqlite3
from datetime import datetime
from pathlib import Path

import openpyxl

from .base_store import BaseSQLiteStore, current_actor
from ._fabric_master_schema import (
    _SCHEMA, _NUMERIC_FIELDS,
    _build_col_map, _v, _num, _make_display_key,
)
from ._fabric_version_schema import (
    _VERSION_SCHEMA, _ALL_FABRIC_COLUMNS, _DIFF_FIELDS, _SNAPSHOT_KEEP_VERSIONS,
    _HIGH_RISK_REMOVED_ROWS, _HIGH_RISK_TOUCHED_PCT,
)


_current_actor = current_actor


class FabricMasterStore(BaseSQLiteStore):
    """Read/write access to the fabric_master SQLite table."""

    # Summary columns used by search / list_all / list_page (no expensive fields)
    _SUMMARY_COLS = (
        "quality_no", "erp_code", "supplier", "composition_en",
        "weight_gsm", "cuttable_width_cm", "dyeing_process",
        "shrinkage_rate", "short_rate", "notes_cn", "display_key",
    )

    def __init__(self, db_path: str):
        self._init_db(db_path)

    # ── Internal helpers ──────────────────────────────────────────────────────

    def _setup_schema(self, conn) -> None:
        conn.executescript(_SCHEMA)
        conn.executescript(_VERSION_SCHEMA)
        self._migrate_swap_widths(conn)
        self._migrate_add_spot_price_cols(conn)
        # Peer-review columns added after versioning first shipped
        self._ensure_columns(conn, "fabric_versions",
                             (("approved_by", "TEXT"), ("review_comment", "TEXT")))
        self._ensure_columns(conn, "fabric_pending_import",
                             (("fields_json", "TEXT NOT NULL DEFAULT '[]'"),
                              ("col_map_json", "TEXT NOT NULL DEFAULT '{}'")))

    @staticmethod
    def _migrate_add_spot_price_cols(conn: sqlite3.Connection) -> None:
        """Add is_in_stock / spot_price_kg / spot_price_m columns if missing.

        BUG-11 fix: the NULL-clearing UPDATE previously fired unconditionally
        whenever spot_price_kg was absent from the schema — which is always true
        on a fresh install, wiping any fabric data already imported in the same
        session.  Now we use user_version=2 as a sentinel so the destructive
        UPDATE only runs once on databases that genuinely pre-date this migration.
        New installations jump straight to version 2 without touching data.
        """
        version = conn.execute("PRAGMA user_version").fetchone()[0]
        if version >= 2:
            return  # migration already applied, nothing to do

        existing_cols = {
            row[1] for row in conn.execute("PRAGMA table_info(fabric_master)").fetchall()
        }

        for col_name, col_type in [
            ("is_in_stock",  "TEXT"),
            ("spot_price_kg", "REAL"),
            ("spot_price_m",  "REAL"),
        ]:
            if col_name not in existing_cols:
                FabricMasterStore._add_column_if_missing(
                    conn, "fabric_master", col_name, col_type
                )

        # Only NULL out stale data when upgrading a pre-existing DB that lacked
        # the spot_price_kg column.  On a fresh install existing_cols already
        # contains the column (added by the DDL in _SCHEMA), so this branch is
        # skipped and no data is touched.
        if "spot_price_kg" not in existing_cols:
            conn.execute(
                """UPDATE fabric_master
                   SET cost_per_kg    = NULL,
                       cost_per_m     = NULL,
                       quote_date     = NULL,
                       shrinkage_rate = NULL,
                       short_rate     = NULL,
                       notes_cn       = NULL,
                       notes_en       = NULL,
                       quote_history  = NULL"""
            )

        conn.execute("PRAGMA user_version = 2")

    @staticmethod
    def _migrate_swap_widths(conn: sqlite3.Connection) -> None:
        """One-time migration: swap full_width_cm ↔ cuttable_width_cm if reversed."""
        version = conn.execute("PRAGMA user_version").fetchone()[0]
        if version >= 1:
            return

        needs_swap = conn.execute(
            """SELECT COUNT(*) FROM fabric_master
               WHERE cuttable_width_cm IS NOT NULL
                 AND full_width_cm IS NOT NULL
                 AND cuttable_width_cm > full_width_cm"""
        ).fetchone()[0]

        if needs_swap:
            # BUG-34 fix: the previous UPDATE had no WHERE clause, so it swapped
            # cuttable ↔ full for EVERY row — corrupting rows that were already
            # correct.  Restrict the swap to rows that actually need it (the
            # same predicate used to detect the issue).
            conn.execute(
                """UPDATE fabric_master
                   SET cuttable_width_cm = full_width_cm,
                       full_width_cm     = cuttable_width_cm
                   WHERE cuttable_width_cm IS NOT NULL
                     AND full_width_cm IS NOT NULL
                     AND cuttable_width_cm > full_width_cm"""
            )
            # Rebuild display_key only for the rows whose width just changed
            rows = conn.execute(
                """SELECT quality_no, composition_en, weight_gsm, cuttable_width_cm
                   FROM fabric_master
                   WHERE cuttable_width_cm IS NOT NULL"""
            ).fetchall()
            for row in rows:
                qno   = row[0]
                comp  = row[1] or ""
                gsm   = str(int(row[2])) if row[2] else ""
                width = str(int(row[3])) if row[3] else ""
                key   = f"{qno}|{comp}|{gsm}|{width}"
                conn.execute(
                    "UPDATE fabric_master SET display_key=? WHERE quality_no=?",
                    (key, qno),
                )

        conn.execute("PRAGMA user_version = 1")

    # ── Import ────────────────────────────────────────────────────────────────

    def import_from_xlsx(self, xlsx_path: str,
                         source_file_name: str | None = None,
                         clear_first: bool = False) -> dict:
        """Import all rows from the 'all' sheet of 面料统计表.xlsx.

        *clear_first*=True wipes fabric_master inside the SAME transaction
        as the import ("Clear & Reimport") -- a failed import rolls the
        delete back too, instead of leaving an emptied table behind.

        Returns a summary dict:
            {"inserted": int, "updated": int, "skipped": int, "total": int,
             "cleared": int, "col_map": dict, "unmatched_headers": list,
             "version_id": int, "unchanged": bool}
        """
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        # try/finally: a mid-import exception used to skip wb.close(), and on
        # Windows the open zip handle kept the uploaded file locked until GC —
        # an immediate retry of the same file failed with a share violation.
        try:
            return self._import_from_open_workbook(wb, xlsx_path, source_file_name,
                                                   clear_first=clear_first)
        finally:
            wb.close()

    @staticmethod
    def _capture_old_state(conn) -> tuple[int | None, dict]:
        """(latest version_id or None, {quality_no: diff-field dict}) — the
        baseline every subsequent diff compares against. Reads the latest
        version's own snapshot when one exists; falls back to a live
        fabric_master read on a never-versioned DB (legacy data). MUST be
        called BEFORE mutating fabric_master — the live fallback would
        otherwise capture the post-mutation state as the baseline.
        """
        v_prev_row = conn.execute(
            "SELECT MAX(version_id) AS v FROM fabric_versions"
        ).fetchone()
        v_prev = v_prev_row["v"] if v_prev_row and v_prev_row["v"] is not None else None

        if v_prev is not None:
            old_rows = conn.execute(
                f"SELECT quality_no, {', '.join(_DIFF_FIELDS)} "
                f"FROM fabric_master_snapshot WHERE version_id=?",
                (v_prev,),
            ).fetchall()
        else:
            old_rows = conn.execute(
                f"SELECT quality_no, {', '.join(_DIFF_FIELDS)} FROM fabric_master"
            ).fetchall()
        return v_prev, {r["quality_no"]: dict(r) for r in old_rows}

    def _mint_version(self, conn, v_prev: int | None, old_state: dict, *,
                      actor: str, source_file: str, stamp: str,
                      inserted: int = 0, updated: int = 0,
                      skipped: int = 0, approved_by: str | None = None,
                      review_comment: str | None = None) -> tuple[int | None, bool]:
        """Version the CURRENT live fabric_master state against *old_state*.

        Shared by every mutation path (xlsx import, manual row deletion,
        approved pending upload) so no data change can slip past the version
        history. Returns ``(version_id, unchanged)``: when nothing differs
        from the latest version, no new version is minted and the latest id
        is returned with ``unchanged=True``. The first-ever version is always
        minted (the baseline snapshot future diffs compare against).
        """
        new_rows = conn.execute(
            f"SELECT {', '.join(_ALL_FABRIC_COLUMNS)} FROM fabric_master"
        ).fetchall()
        new_state = {r["quality_no"]: dict(r) for r in new_rows}
        diff_rows = self._compute_version_diff(old_state, new_state)

        if v_prev is not None and not diff_rows:
            return v_prev, True

        new_version_id = 1 if v_prev is None else v_prev + 1

        conn.execute(
            """INSERT INTO fabric_versions
                  (version_id, created_at, uploaded_by, source_file,
                   row_count, inserted, updated, skipped,
                   approved_by, review_comment)
               VALUES (?,?,?,?,?,?,?,?,?,?)""",
            (new_version_id, stamp, actor, source_file,
             len(new_state), inserted, updated, skipped,
             approved_by, review_comment),
        )

        if new_state:
            cols = ", ".join(_ALL_FABRIC_COLUMNS)
            ph = ", ".join("?" * len(_ALL_FABRIC_COLUMNS))
            conn.executemany(
                f"INSERT INTO fabric_master_snapshot (version_id, {cols}) "
                f"VALUES (?, {ph})",
                [
                    (new_version_id,) + tuple(rec.get(c) for c in _ALL_FABRIC_COLUMNS)
                    for rec in new_state.values()
                ],
            )

        if diff_rows:
            conn.executemany(
                """INSERT INTO fabric_version_diff
                      (version_id, quality_no, change_type, field, old_value, new_value, diffed_at)
                   VALUES (?,?,?,?,?,?,?)""",
                [(new_version_id, *row, stamp) for row in diff_rows],
            )

        # Prune snapshot DATA older than the retention window --
        # fabric_versions / fabric_version_diff rows are never pruned
        # (permanent audit trail).
        conn.execute(
            "DELETE FROM fabric_master_snapshot WHERE version_id <= ?",
            (new_version_id - _SNAPSHOT_KEEP_VERSIONS,),
        )
        return new_version_id, False

    def _import_from_open_workbook(self, wb, xlsx_path: str,
                                   source_file_name: str | None,
                                   clear_first: bool = False) -> dict:
        if "all" not in wb.sheetnames:
            raise ValueError("Sheet 'all' not found in workbook.")

        ws = wb["all"]
        imported_at = datetime.utcnow().isoformat()
        source_file = source_file_name or Path(xlsx_path).name

        records, field_to_col, unmatched, skipped = self._parse_workbook_records(
            ws, imported_at, source_file,
        )
        cleared = 0
        actor = _current_actor()

        with self._conn() as conn:
            # Baseline captured before any mutation (see _capture_old_state) —
            # including before clear_first's wipe, which also closes the old
            # "first import happens to be Clear & Reimport" gap: the legacy
            # live baseline is read while it still exists.
            v_prev, old_state = self._capture_old_state(conn)

            if clear_first:
                cleared = conn.execute(
                    "SELECT COUNT(*) FROM fabric_master").fetchone()[0]
                conn.execute("DELETE FROM fabric_master")

            inserted, updated = self._apply_records(conn, records)

            # ── Versioning: diff FIRST, and only mint a new version when the
            # data actually changed (see _mint_version). A re-upload of the
            # same data under a new filename/timestamp counts as unchanged --
            # the diff compares _DIFF_FIELDS only (imported_at/source_file/
            # display_key excluded), numeric fields at 2 dp.
            version_id, unchanged = self._mint_version(
                conn, v_prev, old_state,
                actor=actor, source_file=source_file, stamp=imported_at,
                inserted=inserted, updated=updated, skipped=skipped,
            )

        return {
            "inserted": inserted,
            "updated": updated,
            "skipped": skipped,
            "cleared": cleared,
            "total": inserted + updated,
            "col_map": {f: c for f, c in field_to_col.items()},
            "unmatched_headers": unmatched,
            "version_id": version_id,
            "unchanged": unchanged,
        }

    @staticmethod
    def _parse_workbook_records(ws, imported_at: str,
                                source_file: str) -> tuple[list[dict], dict, list, int]:
        """Parse the 'all' sheet into record dicts (one per distinct
        quality_no, last occurrence wins). Returns
        ``(records, field_to_col, unmatched_headers, skipped)`` where
        *skipped* counts source rows dropped for a missing quality_no.
        Every record carries the SAME key set: the file's mapped fields plus
        display_key/imported_at/source_file.
        """
        field_to_col, unmatched = _build_col_map(ws)
        col_field_pairs = [(col, field) for field, col in field_to_col.items()]

        skipped = 0
        by_qno: dict[str, dict] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            record: dict = {}
            for col_idx, field in col_field_pairs:
                raw = row[col_idx - 1] if col_idx - 1 < len(row) else None
                if field in _NUMERIC_FIELDS:
                    record[field] = _num(raw)
                elif field == "quote_date":
                    if raw is None:
                        record[field] = None
                    elif hasattr(raw, "isoformat"):
                        record[field] = raw.date().isoformat()
                    else:
                        record[field] = _v(raw) or None
                else:
                    record[field] = _v(raw) or None

            quality_no = record.get("quality_no")
            if not quality_no:
                skipped += 1
                continue

            record["display_key"] = _make_display_key(
                quality_no,
                record.get("composition_en") or "",
                record.get("weight_gsm"),
                record.get("cuttable_width_cm"),
            )
            record["imported_at"] = imported_at
            record["source_file"] = source_file
            by_qno[quality_no] = record

        return list(by_qno.values()), field_to_col, unmatched, skipped

    @staticmethod
    def _apply_records(conn, records: list[dict]) -> tuple[int, int]:
        """Upsert *records* into the live fabric_master table (only each
        record's own keys are written -- fields the source file didn't map
        are left untouched on existing rows). Returns (inserted, updated)."""
        inserted = updated = 0
        for record in records:
            quality_no = record["quality_no"]
            existing = conn.execute(
                "SELECT quality_no FROM fabric_master WHERE quality_no=?",
                (quality_no,)
            ).fetchone()

            fields = list(record.keys())
            if existing:
                set_clause = ", ".join(f"{f}=?" for f in fields if f != "quality_no")
                vals = [record[f] for f in fields if f != "quality_no"] + [quality_no]
                conn.execute(
                    f"UPDATE fabric_master SET {set_clause} WHERE quality_no=?",
                    vals,
                )
                updated += 1
            else:
                placeholders = ", ".join("?" * len(fields))
                conn.execute(
                    f"INSERT INTO fabric_master ({', '.join(fields)}) VALUES ({placeholders})",
                    [record[f] for f in fields],
                )
                inserted += 1
        return inserted, updated

    @staticmethod
    def _diff_value_norm(field: str, value) -> str:
        """Canonical string form of *value* for version-diff comparison.

        Numeric fabric fields are capped at 2 decimal places before
        comparing: the fabric tables' numbers are only meaningful to 2 dp,
        so precision noise beyond that (Excel float artifacts, a source file
        carrying 66.666667 where 66.67 is already on file, int-vs-float
        representation like 200 vs 200.0) must not register as a "changed"
        field -- and hence must not mint a new version by itself.
        """
        if value is None:
            return ""
        if field in _NUMERIC_FIELDS:
            try:
                return str(round(float(value), 2))
            except (ValueError, TypeError):
                pass   # non-numeric text in a numeric column: string-compare
        return str(value)

    @classmethod
    def _compute_version_diff(cls, old_state: dict, new_state: dict) -> list[tuple]:
        """Return one (quality_no, change_type, field, old_value, new_value)
        tuple per added/removed quality_no, and one per changed field for a
        quality_no present in both -- mirrors ColorTranslationStore._audit_diff's
        field-level diff shape. Empty list == the two states are identical
        across _DIFF_FIELDS (i.e. no version bump warranted). Numeric fields
        compare (and are recorded) at 2 dp -- see _diff_value_norm."""
        rows: list[tuple] = []
        for qno in set(old_state) | set(new_state):
            old_row = old_state.get(qno)
            new_row = new_state.get(qno)
            if old_row is None:
                rows.append((qno, "added", None, None, None))
            elif new_row is None:
                rows.append((qno, "removed", None, None, None))
            else:
                for f in _DIFF_FIELDS:
                    ov_s = cls._diff_value_norm(f, old_row.get(f))
                    nv_s = cls._diff_value_norm(f, new_row.get(f))
                    if ov_s != nv_s:
                        rows.append((qno, "changed", f, ov_s, nv_s))
        return rows

    # ── Lookups ───────────────────────────────────────────────────────────────

    def get_by_quality_no(self, quality_no: str) -> dict | None:
        """Return full record dict or None. Matches quality_no or erp_code."""
        key = quality_no.strip()
        with self._conn() as conn:
            row = conn.execute(
                "SELECT * FROM fabric_master WHERE quality_no=? OR erp_code=?",
                (key, key),
            ).fetchone()
        return dict(row) if row else None

    def get_key_info(self, fabric_no: str) -> dict | None:
        """Return 6 key fields for display, or None. Matches quality_no or erp_code."""
        key = fabric_no.strip()
        with self._conn() as conn:
            row = conn.execute(
                """SELECT composition_en, weight_gsm, cuttable_width_cm,
                          shrinkage_rate, short_rate, notes_cn
                   FROM fabric_master WHERE quality_no=? OR erp_code=?""",
                (key, key),
            ).fetchone()
        return dict(row) if row else None

    def get_display_key(self, fabric_no: str) -> str:
        """Return 'quality_no|composition_en|weight_gsm|cuttable_width_cm' or empty str."""
        key = fabric_no.strip()
        with self._conn() as conn:
            row = conn.execute(
                "SELECT display_key FROM fabric_master WHERE quality_no=? OR erp_code=?",
                (key, key),
            ).fetchone()
        return row["display_key"] if row else ""

    def get_batch_enrichment(self, fabric_nos: list, version_id: int | None = None) -> dict:
        """Return {fabric_no: record} for all requested fabric numbers in one pass.

        *version_id*=None (default) reads the live ``fabric_master`` table --
        unchanged behavior for every existing caller. A specific *version_id*
        reads the archived ``fabric_master_snapshot`` for that version instead
        (only the current + 3 most recent versions have snapshot data
        available; an older, pruned version_id simply returns no matches).
        """
        table = "fabric_master" if version_id is None else "fabric_master_snapshot"
        extra_where = "" if version_id is None else " AND version_id=?"
        _SQL = f"""SELECT quality_no, erp_code, display_key,
                         composition_en, weight_gsm, cuttable_width_cm,
                         shrinkage_rate, short_rate
                  FROM {table} WHERE {{col}} IN ({{ph}}){extra_where}"""
        keys = list({str(f).strip() for f in fabric_nos if f})
        if not keys:
            return {}

        version_args = [] if version_id is None else [version_id]
        result: dict = {}
        ph = ",".join("?" * len(keys))
        with self._conn() as conn:
            rows = conn.execute(
                _SQL.format(col="quality_no", ph=ph), keys + version_args
            ).fetchall()
            matched_by_qno = set()
            for row in rows:
                d = dict(row)
                result[d["quality_no"]] = d
                matched_by_qno.add(d["quality_no"])

            unmatched = [k for k in keys if k not in matched_by_qno]
            if unmatched:
                ph2 = ",".join("?" * len(unmatched))
                for row in conn.execute(
                    _SQL.format(col="erp_code", ph=ph2), unmatched + version_args
                ).fetchall():
                    d = dict(row)
                    result[d["erp_code"]] = d

        return result

    # ── Versioning: read/query API ───────────────────────────────────────────

    def list_versions(self, limit: int = 20) -> list[dict]:
        """Return version metadata, newest first -- for populating a version picker."""
        with self._conn() as conn:
            rows = conn.execute(
                """SELECT version_id, created_at, uploaded_by, source_file,
                          row_count, inserted, updated, skipped,
                          approved_by, review_comment
                   FROM fabric_versions ORDER BY version_id DESC LIMIT ?""",
                (limit,),
            ).fetchall()
        return [dict(r) for r in rows]

    def get_diff_summary(self, version_id: int) -> dict:
        """Return {'added': n, 'removed': n, 'changed': n} distinct-quality_no
        counts for one version's diff (a 'changed' quality_no may have several
        field rows, counted once)."""
        summary = {"added": 0, "removed": 0, "changed": 0}
        with self._conn() as conn:
            rows = conn.execute(
                """SELECT change_type, COUNT(DISTINCT quality_no) AS n
                   FROM fabric_version_diff WHERE version_id=?
                   GROUP BY change_type""",
                (version_id,),
            ).fetchall()
        for r in rows:
            summary[r["change_type"]] = r["n"]
        return summary

    def get_version_diff(self, version_id: int) -> list[dict]:
        """Return every diff row for one version, for the detail view."""
        with self._conn() as conn:
            rows = conn.execute(
                """SELECT quality_no, change_type, field, old_value, new_value, diffed_at
                   FROM fabric_version_diff WHERE version_id=?
                   ORDER BY quality_no, id""",
                (version_id,),
            ).fetchall()
        return [dict(r) for r in rows]

    # ── Peer review: staged uploads (propose → review → approve/reject) ──────
    #
    # An UPLOAD never touches fabric_master directly: it is parsed, validated,
    # diffed against the latest version, and staged in fabric_pending_import/
    # fabric_pending_rows until a reviewer approves (two-person rule: the
    # proposer cannot approve their own upload without a justification
    # comment, and the UI additionally restricts approval to admins).
    # Single-record deletes stay immediate (already versioned) by design.

    def get_pending(self) -> dict | None:
        """Return the current pending proposal's metadata (or None).

        ``warnings`` is the parsed list; ``fields`` the mapped-field list.
        At most one proposal can be pending at a time.
        """
        with self._conn() as conn:
            row = conn.execute(
                "SELECT * FROM fabric_pending_import WHERE status='pending' "
                "ORDER BY id DESC LIMIT 1"
            ).fetchone()
        if row is None:
            return None
        d = dict(row)
        try:
            d["warnings"] = json.loads(d.get("warnings_json") or "[]")
        except Exception:
            d["warnings"] = []
        try:
            d["fields"] = json.loads(d.get("fields_json") or "[]")
        except Exception:
            d["fields"] = []
        return d

    def _prospective_diff(self, conn, old_state: dict,
                          records: list[dict], clear_first: bool) -> list[tuple]:
        """Diff the state *records* WOULD produce (without writing anything)
        against *old_state* -- same tuples as _compute_version_diff."""
        if clear_first:
            base: dict[str, dict] = {}
        else:
            live_rows = conn.execute(
                f"SELECT quality_no, {', '.join(_DIFF_FIELDS)} FROM fabric_master"
            ).fetchall()
            base = {r["quality_no"]: dict(r) for r in live_rows}

        for rec in records:
            entry = dict(base.get(rec["quality_no"])
                         or {f: None for f in _DIFF_FIELDS})
            for f in _DIFF_FIELDS:
                if f in rec:
                    entry[f] = rec[f]
            base[rec["quality_no"]] = entry
        return self._compute_version_diff(old_state, base)

    @staticmethod
    def _quality_warnings(records: list[dict], cap: int = 50) -> list[str]:
        """Data-quality warnings for the review panel (composition sums,
        unknown fibers, out-of-range numerics) -- advisory, never blocking."""
        try:
            from ..utils.composition_check import validate_all
            from ..utils.fabric_validators import validate_all_records
            comp = validate_all([
                {"quality_no": r.get("quality_no"),
                 "composition_en": r.get("composition_en")}
                for r in records
            ])
            fields = validate_all_records(records)
            msgs = ([f"{i.quality_no}: {i.detail}" for i in comp]
                    + [f"{i.quality_no}: {i.detail}" for i in fields])
            return msgs[:cap]
        except Exception:
            return []

    def propose_import(self, xlsx_path: str,
                       source_file_name: str | None = None,
                       proposed_by: str | None = None,
                       clear_first: bool = False) -> dict:
        """Stage an upload for peer review. Nothing touches fabric_master.

        Returns one of:
          {"blocked_by_pending": id, "pending_proposed_by": str}
          {"unchanged": True, "version_id": latest}      -- no-op file
          {"pending_id": id, "diff_added"/"diff_removed"/"diff_changed": n,
           "warnings": [...], "high_risk": bool, "row_count": n, ...}
        """
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        try:
            if "all" not in wb.sheetnames:
                raise ValueError("Sheet 'all' not found in workbook.")
            ws = wb["all"]
            stamp = datetime.utcnow().isoformat()
            source_file = source_file_name or Path(xlsx_path).name
            records, field_to_col, unmatched, skipped = \
                self._parse_workbook_records(ws, stamp, source_file)
        finally:
            wb.close()

        proposed_by = proposed_by or _current_actor()

        with self._conn() as conn:
            existing = conn.execute(
                "SELECT id, proposed_by FROM fabric_pending_import "
                "WHERE status='pending' LIMIT 1"
            ).fetchone()
            if existing:
                return {"blocked_by_pending": existing["id"],
                        "pending_proposed_by": existing["proposed_by"]}

            v_prev, old_state = self._capture_old_state(conn)
            diff_rows = self._prospective_diff(conn, old_state, records, clear_first)
            if v_prev is not None and not diff_rows:
                return {"unchanged": True, "version_id": v_prev,
                        "row_count": len(records)}

            added   = len({r[0] for r in diff_rows if r[1] == "added"})
            removed = len({r[0] for r in diff_rows if r[1] == "removed"})
            changed = len({r[0] for r in diff_rows if r[1] == "changed"})
            touched = added + removed + changed
            high_risk = (removed > _HIGH_RISK_REMOVED_ROWS
                         or touched > _HIGH_RISK_TOUCHED_PCT * max(len(old_state), 1))
            warnings = self._quality_warnings(records)
            fields = list(records[0].keys()) if records else []

            cur = conn.execute(
                """INSERT INTO fabric_pending_import
                      (created_at, proposed_by, source_file, clear_first,
                       status, row_count, skipped,
                       diff_added, diff_removed, diff_changed,
                       warnings_json, high_risk, fields_json, col_map_json)
                   VALUES (?,?,?,?,'pending',?,?,?,?,?,?,?,?,?)""",
                (stamp, proposed_by, source_file, int(clear_first),
                 len(records), skipped, added, removed, changed,
                 json.dumps(warnings, ensure_ascii=False), int(high_risk),
                 json.dumps(fields), json.dumps(field_to_col)),
            )
            pending_id = cur.lastrowid
            self._stage_pending_rows(conn, pending_id, records)

        return {"pending_id": pending_id, "row_count": len(records),
                "skipped": skipped, "diff_added": added,
                "diff_removed": removed, "diff_changed": changed,
                "warnings": warnings, "high_risk": high_risk,
                "unmatched_headers": unmatched,
                "col_map": dict(field_to_col)}

    @staticmethod
    def _stage_pending_rows(conn, pending_id: int, records: list[dict]) -> None:
        if not records:
            return
        cols = ", ".join(_ALL_FABRIC_COLUMNS)
        ph = ", ".join("?" * len(_ALL_FABRIC_COLUMNS))
        conn.executemany(
            f"INSERT INTO fabric_pending_rows (pending_id, {cols}) VALUES (?, {ph})",
            [(pending_id,) + tuple(rec.get(c) for c in _ALL_FABRIC_COLUMNS)
             for rec in records],
        )

    def _load_pending_records(self, conn, pending_id: int,
                              fields: list[str]) -> list[dict]:
        """Reconstruct the proposal's records, restricted to the fields the
        source file actually mapped -- so approving preserves the same
        only-touch-mapped-fields semantics as a direct import."""
        keep = [f for f in _ALL_FABRIC_COLUMNS if f in set(fields) | {"quality_no"}]
        rows = conn.execute(
            f"SELECT {', '.join(keep)} FROM fabric_pending_rows WHERE pending_id=?",
            (pending_id,),
        ).fetchall()
        return [dict(r) for r in rows]

    def get_pending_diff(self, pending_id: int) -> list[dict]:
        """Prospective field-level diff for the review panel (recomputed
        against the CURRENT live/latest state, so it stays truthful even if
        records were manually deleted after the proposal was made)."""
        with self._conn() as conn:
            meta = conn.execute(
                "SELECT clear_first, fields_json FROM fabric_pending_import WHERE id=?",
                (pending_id,),
            ).fetchone()
            if meta is None:
                return []
            fields = json.loads(meta["fields_json"] or "[]")
            records = self._load_pending_records(conn, pending_id, fields)
            _v_prev, old_state = self._capture_old_state(conn)
            diff_rows = self._prospective_diff(
                conn, old_state, records, bool(meta["clear_first"]))
        return [
            {"quality_no": q, "change_type": ct, "field": f,
             "old_value": ov, "new_value": nv}
            for (q, ct, f, ov, nv) in diff_rows
        ]

    def approve_pending(self, pending_id: int, reviewed_by: str,
                        comment: str = "") -> dict:
        """Apply a pending proposal to the live table and mint a version.

        Two-person rule: the proposer may not approve their own upload
        unless a non-empty justification *comment* is given (the UI
        additionally restricts this self-approval path to admins).
        """
        reviewed_by = (reviewed_by or "").strip() or "system"
        now = datetime.utcnow().isoformat()

        with self._conn() as conn:
            meta = conn.execute(
                "SELECT * FROM fabric_pending_import WHERE id=?", (pending_id,)
            ).fetchone()
            if meta is None or meta["status"] != "pending":
                raise ValueError(f"No pending proposal with id={pending_id}.")
            if reviewed_by == meta["proposed_by"] and not comment.strip():
                raise ValueError(
                    "Self-approval requires a justification comment "
                    "(two-person rule)."
                )

            fields = json.loads(meta["fields_json"] or "[]")
            records = self._load_pending_records(conn, pending_id, fields)

            v_prev, old_state = self._capture_old_state(conn)
            if meta["clear_first"]:
                conn.execute("DELETE FROM fabric_master")
            inserted, updated = self._apply_records(conn, records)

            version_id, unchanged = self._mint_version(
                conn, v_prev, old_state,
                actor=meta["proposed_by"], source_file=meta["source_file"] or "",
                stamp=now, inserted=inserted, updated=updated,
                skipped=meta["skipped"],
                approved_by=reviewed_by, review_comment=comment.strip() or None,
            )

            conn.execute(
                """UPDATE fabric_pending_import
                   SET status='approved', reviewed_by=?, reviewed_at=?,
                       review_comment=?, version_id=?
                   WHERE id=?""",
                (reviewed_by, now, comment.strip() or None, version_id, pending_id),
            )
            conn.execute(
                "DELETE FROM fabric_pending_rows WHERE pending_id=?", (pending_id,)
            )

        return {"version_id": version_id, "unchanged": unchanged,
                "inserted": inserted, "updated": updated}

    def reject_pending(self, pending_id: int, reviewed_by: str,
                       comment: str) -> None:
        """Discard a pending proposal. A rejection reason is mandatory."""
        if not (comment or "").strip():
            raise ValueError("A rejection reason is required.")
        now = datetime.utcnow().isoformat()
        with self._conn() as conn:
            meta = conn.execute(
                "SELECT status FROM fabric_pending_import WHERE id=?", (pending_id,)
            ).fetchone()
            if meta is None or meta["status"] != "pending":
                raise ValueError(f"No pending proposal with id={pending_id}.")
            conn.execute(
                """UPDATE fabric_pending_import
                   SET status='rejected', reviewed_by=?, reviewed_at=?, review_comment=?
                   WHERE id=?""",
                ((reviewed_by or "").strip() or "system", now, comment.strip(), pending_id),
            )
            conn.execute(
                "DELETE FROM fabric_pending_rows WHERE pending_id=?", (pending_id,)
            )

    def cancel_pending(self, pending_id: int, by: str) -> None:
        """Withdraw a pending proposal (proposer changing their mind — the
        UI restricts this to the proposer or an admin). No comment needed."""
        now = datetime.utcnow().isoformat()
        with self._conn() as conn:
            meta = conn.execute(
                "SELECT status FROM fabric_pending_import WHERE id=?", (pending_id,)
            ).fetchone()
            if meta is None or meta["status"] != "pending":
                raise ValueError(f"No pending proposal with id={pending_id}.")
            conn.execute(
                """UPDATE fabric_pending_import
                   SET status='cancelled', reviewed_by=?, reviewed_at=?
                   WHERE id=?""",
                ((by or "").strip() or "system", now, pending_id),
            )
            conn.execute(
                "DELETE FROM fabric_pending_rows WHERE pending_id=?", (pending_id,)
            )

    def propose_restore(self, version_id: int,
                        proposed_by: str | None = None) -> dict:
        """Stage a rollback: restore version *version_id*'s snapshot as a new
        pending proposal (full replacement). Goes through the same review
        gate as any upload. Only versions still inside the snapshot
        retention window can be restored."""
        proposed_by = proposed_by or _current_actor()
        stamp = datetime.utcnow().isoformat()
        source_file = f"restore v{version_id}"

        with self._conn() as conn:
            existing = conn.execute(
                "SELECT id, proposed_by FROM fabric_pending_import "
                "WHERE status='pending' LIMIT 1"
            ).fetchone()
            if existing:
                return {"blocked_by_pending": existing["id"],
                        "pending_proposed_by": existing["proposed_by"]}

            snap_rows = conn.execute(
                f"SELECT {', '.join(_ALL_FABRIC_COLUMNS)} "
                f"FROM fabric_master_snapshot WHERE version_id=?",
                (version_id,),
            ).fetchall()
            if not snap_rows:
                raise ValueError(
                    f"Version {version_id} has no snapshot data (pruned or unknown)."
                )
            records = [dict(r) for r in snap_rows]

            v_prev, old_state = self._capture_old_state(conn)
            diff_rows = self._prospective_diff(conn, old_state, records,
                                               clear_first=True)
            if v_prev is not None and not diff_rows:
                return {"unchanged": True, "version_id": v_prev}

            added   = len({r[0] for r in diff_rows if r[1] == "added"})
            removed = len({r[0] for r in diff_rows if r[1] == "removed"})
            changed = len({r[0] for r in diff_rows if r[1] == "changed"})
            touched = added + removed + changed
            high_risk = (removed > _HIGH_RISK_REMOVED_ROWS
                         or touched > _HIGH_RISK_TOUCHED_PCT * max(len(old_state), 1))

            cur = conn.execute(
                """INSERT INTO fabric_pending_import
                      (created_at, proposed_by, source_file, clear_first,
                       status, row_count, skipped,
                       diff_added, diff_removed, diff_changed,
                       warnings_json, high_risk, fields_json, col_map_json)
                   VALUES (?,?,?,1,'pending',?,0,?,?,?,'[]',?,?, '{}')""",
                (stamp, proposed_by, source_file,
                 len(records), added, removed, changed, int(high_risk),
                 json.dumps(list(_ALL_FABRIC_COLUMNS))),
            )
            pending_id = cur.lastrowid
            self._stage_pending_rows(conn, pending_id, records)

        return {"pending_id": pending_id, "row_count": len(records),
                "diff_added": added, "diff_removed": removed,
                "diff_changed": changed, "high_risk": high_risk,
                "warnings": []}

    def search(self, query: str, limit: int = 200) -> list[dict]:
        """Search by quality_no, composition_en, supplier, or structure_en."""
        q = f"%{query.strip()}%"
        cols = ", ".join(self._SUMMARY_COLS)
        with self._conn() as conn:
            rows = conn.execute(
                f"""SELECT {cols}
                   FROM fabric_master
                   WHERE quality_no LIKE ? OR composition_en LIKE ?
                      OR supplier LIKE ? OR structure_en LIKE ?
                   ORDER BY quality_no
                   LIMIT ?""",
                (q, q, q, q, limit),
            ).fetchall()
        return [dict(r) for r in rows]

    def list_all(self, limit: int = 5000) -> list[dict]:
        """Return all records (summary columns only)."""
        cols = ", ".join(self._SUMMARY_COLS) + ", imported_at, source_file"
        with self._conn() as conn:
            rows = conn.execute(
                f"SELECT {cols} FROM fabric_master ORDER BY quality_no LIMIT ?",
                (limit,),
            ).fetchall()
        return [dict(r) for r in rows]

    def list_page(self, offset: int = 0, limit: int = 200) -> list[dict]:
        """Return one page of records (summary columns only), ordered by quality_no."""
        cols = ", ".join(self._SUMMARY_COLS) + ", imported_at, source_file"
        with self._conn() as conn:
            rows = conn.execute(
                f"SELECT {cols} FROM fabric_master ORDER BY quality_no LIMIT ? OFFSET ?",
                (limit, offset),
            ).fetchall()
        return [dict(r) for r in rows]

    def list_all_compositions(self) -> list[dict]:
        """Return {quality_no, composition_en} for every record — used for validation."""
        with self._conn() as conn:
            rows = conn.execute(
                "SELECT quality_no, composition_en FROM fabric_master ORDER BY quality_no"
            ).fetchall()
        return [dict(r) for r in rows]

    def delete_all(self) -> int:
        """Delete every row from fabric_master. Returns the number of rows removed.

        Version-TRANSPARENT by design: intended only as a step inside a
        larger versioned operation (e.g. ``import_from_xlsx(clear_first=True)``
        does the equivalent wipe inside the import's own transaction, so the
        subsequent version diff records the removals). Do not expose this
        directly to users -- a standalone wipe would bypass version history.
        """
        with self._conn() as conn:
            n = conn.execute("SELECT COUNT(*) FROM fabric_master").fetchone()[0]
            conn.execute("DELETE FROM fabric_master")
        return n

    def delete_by_quality_nos(self, quality_nos: list[str]) -> int:
        """Delete rows whose quality_no is in *quality_nos*. Returns rows deleted.

        VERSIONED: a manual deletion is a data change like any other, so it
        mints a version (source labelled "manual delete") with "removed" diff
        rows -- otherwise the deletion would be invisible in the history, the
        live table would silently drift from the latest snapshot, and a later
        re-upload containing the deleted rows would be reported "unchanged"
        while actually restoring them.
        """
        if not quality_nos:
            return 0
        keys = [str(q).strip() for q in quality_nos if q]
        ph   = ",".join("?" * len(keys))
        with self._conn() as conn:
            v_prev, old_state = self._capture_old_state(conn)
            cur = conn.execute(
                f"DELETE FROM fabric_master WHERE quality_no IN ({ph})", keys
            )
            n = cur.rowcount
            if n:
                self._mint_version(
                    conn, v_prev, old_state,
                    actor=_current_actor(), source_file="manual delete",
                    stamp=datetime.utcnow().isoformat(),
                )
        return n

    def list_all_quality_nos(self) -> list[str]:
        """Return all quality_no values for cross-system HHN orphan checks."""
        with self._conn() as conn:
            rows = conn.execute(
                "SELECT quality_no FROM fabric_master WHERE quality_no IS NOT NULL"
            ).fetchall()
        return [r[0] for r in rows]

    def list_all_for_validation(self) -> list[dict]:
        """Return all records with the fields needed for field-range validation."""
        with self._conn() as conn:
            rows = conn.execute(
                """SELECT quality_no, weight_gsm, cuttable_width_cm, full_width_cm,
                          shrinkage_rate, short_rate
                   FROM fabric_master ORDER BY quality_no"""
            ).fetchall()
        return [dict(r) for r in rows]

    def count(self) -> int:
        with self._conn() as conn:
            return conn.execute("SELECT COUNT(*) FROM fabric_master").fetchone()[0]

    def last_import_info(self) -> dict | None:
        """Return imported_at and source_file of the most recent import."""
        with self._conn() as conn:
            row = conn.execute(
                "SELECT imported_at, source_file FROM fabric_master "
                "ORDER BY imported_at DESC LIMIT 1"
            ).fetchone()
        return dict(row) if row else None

    # ── Cross-DB migration ─────────────────────────────────────────────────────

    @classmethod
    def migrate_from_db(cls, src_db_path: str, dst_db_path: str) -> dict:
        """Copy the fabric_master table from one SQLite database to another.

        Designed for the one-time migration from the legacy shared ``po_history.db``
        into the dedicated ``fabric_master.db``.  Reads rows into Python memory
        from the source then bulk-inserts into the destination — avoids ATTACH
        transaction-locking issues that occur in WAL mode.

        Existing rows in *dst_db_path* with matching ``quality_no`` are replaced.

        Returns::
            {
                "rows_read":      int,   # rows read from source
                "net_added":      int,   # net new rows in destination
                                          # (= dst_count_after - dst_count_before)
                "already_in_dst": int,   # rows in dst before migration
                "message":        str,   # human-readable summary
            }
        """
        import sqlite3 as _sqlite3

        # ── Read from source ────────────────────────────────────────────────
        src_conn = None
        try:
            src_conn = _sqlite3.connect(src_db_path)
            src_conn.row_factory = _sqlite3.Row
            has_table = src_conn.execute(
                "SELECT name FROM sqlite_master "
                "WHERE type='table' AND name='fabric_master'"
            ).fetchone()
            if not has_table:
                return {"rows_read": 0, "net_added": 0, "already_in_dst": 0,
                        "message": "Source DB has no fabric_master table."}

            src_rows = src_conn.execute("SELECT * FROM fabric_master").fetchall()
            if not src_rows:
                return {"rows_read": 0, "net_added": 0, "already_in_dst": 0,
                        "message": "Source fabric_master is empty — nothing to migrate."}

            # Column names from the source
            src_cols = [desc[0] for desc in src_conn.execute(
                "SELECT * FROM fabric_master LIMIT 0"
            ).description]
        except Exception as exc:
            return {"rows_read": 0, "net_added": 0, "already_in_dst": 0,
                    "message": f"Cannot read source DB: {exc}"}
        finally:
            if src_conn is not None:
                src_conn.close()

        # ── Ensure destination schema, then write ───────────────────────────
        dst_store = cls(dst_db_path)
        already_in_dst = dst_store.count()

        # Only insert columns present in the destination schema.
        with dst_store._conn() as dst_conn:
            dst_cols = {
                row[1]
                for row in dst_conn.execute(
                    "PRAGMA table_info(fabric_master)"
                ).fetchall()
            }
            common = [c for c in src_cols if c in dst_cols]
            # Quote column identifiers so reserved-word or whitespace-bearing
            # column names round-trip safely.  Doubling embedded `"` per the
            # SQL-92 escaping rule keeps the SQL well-formed even with
            # adversarial schemas.
            col_list = ", ".join(f'"{c.replace(chr(34), chr(34)*2)}"' for c in common)
            ph       = ", ".join("?" * len(common))
            rows_to_insert = [
                tuple(row[c] for c in common) for row in src_rows
            ]
            dst_conn.executemany(
                f"INSERT OR REPLACE INTO fabric_master ({col_list}) VALUES ({ph})",
                rows_to_insert,
            )

        net_added = dst_store.count() - already_in_dst
        return {
            "rows_read":      len(src_rows),
            "net_added":      net_added,
            "already_in_dst": already_in_dst,
            "message": (
                f"Read {len(src_rows)} rows from source DB; "
                f"{net_added} added to destination "
                f"({already_in_dst} were already present)."
            ),
        }
