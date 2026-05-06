"""SQLite-backed store for the color-name translation table.

Stores English ↔ Chinese color name pairs, grouped by client + brand.

Schema
------
color_translations(
    id          INTEGER PRIMARY KEY,
    client      TEXT NOT NULL,          -- e.g. "GIII", "Sky East"
    brand       TEXT NOT NULL DEFAULT '',  -- e.g. "Karl Lagerfeld", "Anna Field"
    en_color    TEXT NOT NULL,
    cn_color    TEXT,
    color_code  TEXT,
    notes       TEXT,
    updated_at  TEXT,
    UNIQUE(client, brand, en_color)
)
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd

from .base_store import BaseSQLiteStore
from ._color_translation_schema import (
    _SCHEMA, _MIGRATION_ADD_BRAND, _MIGRATION_ADD_LABEL_COLS, _AUDIT_SCHEMA,
    _COL_CLIENT, _COL_BRAND, _COL_EN_COLOR, _COL_CN_COLOR, _COL_COLOR_CODE,
    _COL_LIGHT_DARK, _COL_LABEL_CLR, _COL_NOTES,
    _v, _match_col,
)
from auth.companies import COMPANY_GIII, COMPANY_SKY_EAST


# Tracked columns whose changes are logged into the audit table.
_AUDIT_FIELDS = ("cn_color", "color_code", "light_or_dark",
                 "label_color", "notes")


def _current_actor() -> str:
    """Best-effort: return the logged-in Streamlit user, else ``'system'``."""
    try:
        import streamlit as st
        from ui.session_keys import SK
        return str(st.session_state.get(SK.USERNAME) or "system").strip() or "system"
    except Exception:
        return "system"


def _audit_log(conn, action: str, row_id: int | None, client: str, brand: str,
               en_color: str, field: str,
               old_value: str | None, new_value: str | None,
               actor: str | None = None) -> None:
    """Insert one audit-log row.  Called inside an active connection so the
    audit insert and the data change share the same transaction.
    """
    conn.execute(
        """INSERT INTO color_translations_audit
              (action, row_id, client, brand, en_color, field,
               old_value, new_value, changed_at, changed_by)
           VALUES (?,?,?,?,?,?,?,?,?,?)""",
        (action, row_id, client or "", brand or "", en_color or "",
         field or "*",
         None if old_value is None else str(old_value),
         None if new_value is None else str(new_value),
         datetime.utcnow().isoformat(),
         actor or _current_actor()),
    )


def _audit_diff(conn, row_id: int, client: str, brand: str, en_color: str,
                old_row: dict, new_row: dict, actor: str | None = None) -> None:
    """Compare *old_row* and *new_row* on the audited fields and emit one
    audit entry per changed field."""
    for f in _AUDIT_FIELDS:
        ov = (old_row.get(f) or "")
        nv = (new_row.get(f) or "")
        if ov != nv:
            _audit_log(conn, "update", row_id, client, brand, en_color,
                       f, ov, nv, actor)


def _normalize_color_name(s: str | None) -> str:
    """Canonical English-colour-name form: case-insensitive, title-case,
    whitespace-collapsed.  Used everywhere so that "NAVY", "navy" and
    "Navy" map to the same row in the table.

    Examples
    --------
    >>> _normalize_color_name("NAVY")              → "Navy"
    >>> _normalize_color_name("navy")              → "Navy"
    >>> _normalize_color_name("  chocolate brown") → "Chocolate Brown"
    >>> _normalize_color_name("52#NAVY")           → "52#Navy"
    >>> _normalize_color_name(None)                → ""
    """
    if not s:
        return ""
    return " ".join(str(s).split()).lower().title()


def _derive_shade_and_label(en_color: str) -> tuple[str, str]:
    """Auto-classify *en_color* using the same keyword sets as the
    buyplan exporter.  Returns ``(shade, label)`` where shade is
    'light'/'dark'/'' and label is '黑色'/'白色'/''.
    """
    if not en_color:
        return "", ""
    from po_extractor.exporters._sky_east_helpers import (
        derive_main_label_color,
        _LIGHT_BODY_KEYWORDS, _DARK_BODY_KEYWORDS,
    )
    t = str(en_color).lower()
    light_pos = min((t.find(k) for k in _LIGHT_BODY_KEYWORDS if k in t), default=-1)
    dark_pos  = min((t.find(k) for k in _DARK_BODY_KEYWORDS  if k in t), default=-1)
    if light_pos == -1 and dark_pos == -1:
        return "", ""
    if dark_pos == -1:
        return "light", derive_main_label_color(en_color)
    if light_pos == -1:
        return "dark",  derive_main_label_color(en_color)
    return ("dark", derive_main_label_color(en_color)) if dark_pos < light_pos else (
        "light", derive_main_label_color(en_color)
    )


class ColorTranslationStore(BaseSQLiteStore):
    def __init__(self, db_path: str | Path):
        self.db_path = str(db_path)
        Path(db_path).parent.mkdir(parents=True, exist_ok=True)
        self._ensure_schema()

    # ── Internal ─────────────────────────────────────────────────────────────

    def _ensure_schema(self) -> None:
        with self._conn() as conn:
            tbl = conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name='color_translations'"
            ).fetchone()
            if tbl:
                cols = {r[1] for r in conn.execute("PRAGMA table_info(color_translations)")}
                if "brand" not in cols:
                    conn.executescript(_MIGRATION_ADD_BRAND)
                    cols = {r[1] for r in conn.execute("PRAGMA table_info(color_translations)")}
                # v1 → v2: add light_or_dark + label_color columns
                if "light_or_dark" not in cols or "label_color" not in cols:
                    conn.executescript(_MIGRATION_ADD_LABEL_COLS)
                    self._backfill_light_dark_and_label(conn)
            else:
                conn.executescript(_SCHEMA)
            # v2 → v3: ensure the audit-log table exists (idempotent — uses
            # CREATE TABLE IF NOT EXISTS so safe to run on every startup).
            conn.executescript(_AUDIT_SCHEMA)

    @staticmethod
    def _backfill_light_dark_and_label(conn) -> None:
        """Populate light_or_dark + label_color for every existing row using
        the same auto-derivation rule as the buyplan exporter.

        Light body → label_color = 黑色;  Dark body → label_color = 白色.
        """
        from po_extractor.exporters._sky_east_helpers import (
            derive_main_label_color,
            _LIGHT_BODY_KEYWORDS,
            _DARK_BODY_KEYWORDS,
        )

        def classify_shade(en: str) -> str:
            if not en:
                return ""
            t = en.lower()
            light_pos = min((t.find(k) for k in _LIGHT_BODY_KEYWORDS if k in t), default=-1)
            dark_pos  = min((t.find(k) for k in _DARK_BODY_KEYWORDS  if k in t), default=-1)
            if light_pos == -1 and dark_pos == -1:
                return ""
            if dark_pos == -1:
                return "light"
            if light_pos == -1:
                return "dark"
            return "dark" if dark_pos < light_pos else "light"

        rows = conn.execute(
            "SELECT id, en_color, light_or_dark, label_color "
            "FROM color_translations"
        ).fetchall()
        for r in rows:
            en = str(r["en_color"] or "")
            shade = (r["light_or_dark"] or "").strip()
            label = (r["label_color"] or "").strip()
            if shade and label:
                continue
            new_shade = shade or classify_shade(en)
            new_label = label or derive_main_label_color(en)
            if new_shade != shade or new_label != label:
                conn.execute(
                    "UPDATE color_translations SET light_or_dark=?, label_color=? "
                    "WHERE id=?",
                    (new_shade, new_label, r["id"])
                )

    # ── Import from Excel ─────────────────────────────────────────────────────

    def import_from_xlsx(self, xlsx_path: str,
                         default_client: str = "",
                         default_brand: str = "") -> dict:
        """Import rows from an Excel file.

        Required column: *en_color* (or alias).
        Optional: *client*, *brand*, *cn_color*, *color_code*, *notes*.

        Returns {"inserted", "updated", "skipped", "total"}.
        """
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            return {"inserted": 0, "updated": 0, "skipped": 0, "total": 0}

        headers = [_v(c) for c in rows[0]]
        col_client = col_brand = col_en = col_cn = col_code = col_notes = None
        col_shade = col_label = None
        for idx, h in enumerate(headers):
            if _match_col(h, _COL_CLIENT):        col_client = idx
            elif _match_col(h, _COL_BRAND):       col_brand  = idx
            elif _match_col(h, _COL_EN_COLOR):    col_en     = idx
            elif _match_col(h, _COL_CN_COLOR):    col_cn     = idx
            elif _match_col(h, _COL_COLOR_CODE):  col_code   = idx
            elif _match_col(h, _COL_LIGHT_DARK):  col_shade  = idx
            elif _match_col(h, _COL_LABEL_CLR):   col_label  = idx
            elif _match_col(h, _COL_NOTES):       col_notes  = idx

        if col_en is None:
            raise ValueError(
                f"Excel must have an 'en_color' column. Found: {headers}"
            )

        now = datetime.utcnow().isoformat()
        inserted = updated = skipped = 0

        with self._conn() as conn:
            for row in rows[1:]:
                def cell(i):
                    return _v(row[i]) if i is not None and i < len(row) else ""

                client   = cell(col_client) or default_client or "Unknown"
                brand    = cell(col_brand)  or default_brand  or ""
                en_color = _normalize_color_name(cell(col_en))
                if not en_color:
                    skipped += 1
                    continue
                cn_color   = cell(col_cn)
                color_code = cell(col_code)
                notes      = cell(col_notes)
                shade      = cell(col_shade).lower()
                label      = cell(col_label)
                # Auto-derive missing shade/label from en_color so imported
                # files don't have to fill them in manually.
                if not shade or not label:
                    derived_shade, derived_label = _derive_shade_and_label(en_color)
                    shade = shade or derived_shade
                    label = label or derived_label

                existing = conn.execute(
                    "SELECT id, cn_color, color_code, light_or_dark, "
                    "label_color, notes FROM color_translations "
                    "WHERE client=? AND brand=? AND en_color=?",
                    (client, brand, en_color)
                ).fetchone()

                if existing:
                    conn.execute(
                        """UPDATE color_translations
                           SET cn_color=?, color_code=?, light_or_dark=?, label_color=?,
                               notes=?, updated_at=?
                           WHERE client=? AND brand=? AND en_color=?""",
                        (cn_color, color_code, shade, label, notes, now,
                         client, brand, en_color)
                    )
                    updated += 1
                    _audit_diff(conn, existing["id"], client, brand, en_color,
                                {f: existing[f] for f in _AUDIT_FIELDS},
                                {"cn_color": cn_color, "color_code": color_code,
                                 "light_or_dark": shade, "label_color": label,
                                 "notes": notes})
                else:
                    conn.execute(
                        """INSERT INTO color_translations
                           (client, brand, en_color, cn_color, color_code,
                            light_or_dark, label_color, notes, updated_at)
                           VALUES (?,?,?,?,?,?,?,?,?)""",
                        (client, brand, en_color, cn_color, color_code,
                         shade, label, notes, now)
                    )
                    inserted += 1
                    new_id = conn.execute(
                        "SELECT id FROM color_translations "
                        "WHERE client=? AND brand=? AND en_color=?",
                        (client, brand, en_color)
                    ).fetchone()[0]
                    _audit_log(conn, "insert", new_id, client, brand, en_color,
                               "*", None,
                               f"cn={cn_color!r}, code={color_code!r}, "
                               f"shade={shade!r}, label={label!r} (xlsx import)")

        return {"inserted": inserted, "updated": updated,
                "skipped": skipped, "total": inserted + updated}

    # ── Load from existing PO / Sky East data ─────────────────────────────────

    def load_from_po_data(self, skip_existing: bool = True) -> dict:
        """Scan po_size_rows + po_metadata and sky_east_items in the same DB
        and insert any (client, brand, color) combinations not already present.

        GIII:      client = company,  brand = division_name
        Sky East:  client = COMPANY_SKY_EAST, brand = item brand (e.g. "Anna Field")
        """
        now = datetime.utcnow().isoformat()
        inserted = skipped = giii_n = se_n = 0

        with self._conn() as conn:
            # ── GIII: po_size_rows + po_metadata ─────────────────────────────
            giii_rows = conn.execute(
                """SELECT DISTINCT
                       COALESCE(m.company, ?) AS client,
                       COALESCE(m.division_name, '') AS brand,
                       r.color AS en_color
                   FROM po_size_rows r
                   JOIN po_metadata m ON r.po_number = m.po_number
                   WHERE r.color IS NOT NULL AND trim(r.color) != ''
                   ORDER BY client, brand, r.color""",
                (COMPANY_GIII,),
            ).fetchall()

            for row in giii_rows:
                # BUG-36 defensive: SQLite may return non-str types (int, float)
                # if the source column was numeric.  Coerce to str before strip.
                client   = str(row["client"] or "").strip()
                brand    = str(row["brand"] or "").strip()
                en_color = str(row["en_color"] or "").strip()
                exists = conn.execute(
                    "SELECT id FROM color_translations "
                    "WHERE client=? AND brand=? AND en_color=?",
                    (client, brand, en_color)
                ).fetchone()
                if exists and skip_existing:
                    skipped += 1
                    continue
                if not exists:
                    conn.execute(
                        """INSERT OR IGNORE INTO color_translations
                           (client, brand, en_color, cn_color, color_code, notes, updated_at)
                           VALUES (?,?,?,?,?,?,?)""",
                        (client, brand, en_color, "", "", "", now)
                    )
                    inserted += 1
                    giii_n += 1
                else:
                    skipped += 1

            # ── Sky East: sky_east_items ──────────────────────────────────────
            se_rows = conn.execute(
                """SELECT DISTINCT
                       ? AS client,
                       COALESCE(brand, '') AS brand,
                       color_name AS en_color,
                       COALESCE(colour_code, '') AS color_code
                   FROM sky_east_items
                   WHERE color_name IS NOT NULL AND trim(color_name) != ''
                   ORDER BY brand, color_name""",
                (COMPANY_SKY_EAST,),
            ).fetchall()

            for row in se_rows:
                client     = COMPANY_SKY_EAST
                # BUG-36 defensive: same coercion as GIII branch above
                brand      = str(row["brand"] or "").strip()
                en_color   = str(row["en_color"] or "").strip()
                color_code = str(row["color_code"] or "").strip()
                if not en_color:
                    skipped += 1
                    continue
                exists = conn.execute(
                    "SELECT id FROM color_translations "
                    "WHERE client=? AND brand=? AND en_color=?",
                    (client, brand, en_color)
                ).fetchone()
                if exists and skip_existing:
                    skipped += 1
                    continue
                if not exists:
                    conn.execute(
                        """INSERT OR IGNORE INTO color_translations
                           (client, brand, en_color, cn_color, color_code, notes, updated_at)
                           VALUES (?,?,?,?,?,?,?)""",
                        (client, brand, en_color, "", color_code, "", now)
                    )
                    inserted += 1
                    se_n += 1
                else:
                    if color_code:
                        conn.execute(
                            """UPDATE color_translations SET color_code=?, updated_at=?
                               WHERE client=? AND brand=? AND en_color=?
                               AND (color_code IS NULL OR color_code='')""",
                            (color_code, now, client, brand, en_color)
                        )
                    skipped += 1

        return {
            "inserted": inserted, "skipped": skipped,
            "sources": {"giii": giii_n, "sky_east": se_n},
        }

    # ── Upsert from DataFrame (data_editor save) ──────────────────────────────

    def upsert_from_df(self, df: pd.DataFrame) -> int:
        now = datetime.utcnow().isoformat()
        actor = _current_actor()
        count = 0
        with self._conn() as conn:
            for _, row in df.iterrows():
                client   = _v(row.get("Client"))
                brand    = _v(row.get("Brand"))
                en_color = _normalize_color_name(_v(row.get("English Color")))
                if not client or not en_color:
                    continue
                cn_color   = _v(row.get("Chinese Color"))
                color_code = _v(row.get("中文颜色代码") or row.get("Color Code"))
                shade      = _v(row.get("Light/Dark")).lower()
                label      = _v(row.get("Label Color"))
                notes      = _v(row.get("Notes"))
                if not shade or not label:
                    derived_shade, derived_label = _derive_shade_and_label(en_color)
                    shade = shade or derived_shade
                    label = label or derived_label

                # Snapshot the old row (if any) so we can diff for the audit log
                old = conn.execute(
                    """SELECT id, cn_color, color_code, light_or_dark,
                              label_color, notes
                       FROM color_translations
                       WHERE client=? AND brand=? AND en_color=?""",
                    (client, brand, en_color)
                ).fetchone()

                new_vals = {
                    "cn_color": cn_color, "color_code": color_code,
                    "light_or_dark": shade, "label_color": label,
                    "notes": notes,
                }
                conn.execute(
                    """INSERT INTO color_translations
                       (client, brand, en_color, cn_color, color_code,
                        light_or_dark, label_color, notes, updated_at)
                       VALUES (?,?,?,?,?,?,?,?,?)
                       ON CONFLICT(client, brand, en_color) DO UPDATE SET
                           cn_color=excluded.cn_color,
                           color_code=excluded.color_code,
                           light_or_dark=excluded.light_or_dark,
                           label_color=excluded.label_color,
                           notes=excluded.notes,
                           updated_at=excluded.updated_at""",
                    (client, brand, en_color, cn_color, color_code,
                     shade, label, notes, now)
                )
                count += 1

                if old is None:
                    # New row inserted
                    new_id = conn.execute(
                        "SELECT id FROM color_translations "
                        "WHERE client=? AND brand=? AND en_color=?",
                        (client, brand, en_color)
                    ).fetchone()[0]
                    _audit_log(conn, "insert", new_id, client, brand, en_color,
                               "*", None,
                               f"cn={cn_color!r}, code={color_code!r}, "
                               f"shade={shade!r}, label={label!r}",
                               actor)
                else:
                    _audit_diff(conn, old["id"], client, brand, en_color,
                                {f: old[f] for f in _AUDIT_FIELDS},
                                new_vals, actor)
        return count

    def delete_by_client_brand(self, client: str, brand: str = "") -> int:
        actor = _current_actor()
        with self._conn() as conn:
            # Snapshot rows about to be deleted for the audit log
            if brand:
                rows = conn.execute(
                    "SELECT id, client, brand, en_color, cn_color, color_code, "
                    "light_or_dark, label_color, notes "
                    "FROM color_translations WHERE client=? AND brand=?",
                    (client, brand),
                ).fetchall()
            else:
                rows = conn.execute(
                    "SELECT id, client, brand, en_color, cn_color, color_code, "
                    "light_or_dark, label_color, notes "
                    "FROM color_translations WHERE client=?",
                    (client,),
                ).fetchall()
            for r in rows:
                _audit_log(
                    conn, "delete", r["id"], r["client"], r["brand"],
                    r["en_color"], "*",
                    f"cn={r['cn_color']!r}, code={r['color_code']!r}, "
                    f"shade={r['light_or_dark']!r}, label={r['label_color']!r}",
                    None, actor,
                )
            if brand:
                cur = conn.execute(
                    "DELETE FROM color_translations WHERE client=? AND brand=?",
                    (client, brand)
                )
            else:
                cur = conn.execute(
                    "DELETE FROM color_translations WHERE client=?", (client,)
                )
        return cur.rowcount

    def delete_ids(self, ids: list[int]) -> int:
        """Delete the rows whose primary-key ``id`` is in *ids*.  Returns
        the number of rows actually deleted."""
        clean = [int(i) for i in ids if i is not None and str(i).strip() != ""]
        if not clean:
            return 0
        ph = ",".join("?" * len(clean))
        actor = _current_actor()
        with self._conn() as conn:
            rows = conn.execute(
                f"SELECT id, client, brand, en_color, cn_color, color_code, "
                f"light_or_dark, label_color, notes "
                f"FROM color_translations WHERE id IN ({ph})", clean,
            ).fetchall()
            for r in rows:
                _audit_log(
                    conn, "delete", r["id"], r["client"], r["brand"],
                    r["en_color"], "*",
                    f"cn={r['cn_color']!r}, code={r['color_code']!r}, "
                    f"shade={r['light_or_dark']!r}, label={r['label_color']!r}",
                    None, actor,
                )
            cur = conn.execute(
                f"DELETE FROM color_translations WHERE id IN ({ph})", clean
            )
        return cur.rowcount

    # ── Specialised importer for the Zalando 大货进度表 progress tracker ──

    def import_from_progress_xlsx(
        self,
        xlsx_path: str,
        client: str = COMPANY_SKY_EAST,
    ) -> dict:
        """Read the columns 颜色 / 主标颜色 / 中文颜色 (+ BRAND when present)
        from a 大货进度表 workbook and upsert into ``color_translations``.

        Headers are detected case-insensitively in the first 5 rows of every
        sheet.  Colour names are normalised to title case so "NAVY", "navy"
        and "Navy" all collapse onto the same row.

        Returns ``{"inserted", "updated", "skipped", "total_rows", "sheets"}``.
        """
        import openpyxl as _ox
        wb = _ox.load_workbook(xlsx_path, read_only=True, data_only=True)

        BRAND_HEADERS    = {"brand", "客户品牌", "品牌"}
        EN_HEADERS       = {"颜色", "英文颜色", "color", "color (en)", "en color",
                            "english color", "colour", "colour (en)"}
        MAIN_HEADERS     = {"主标颜色", "main label color", "label color"}
        CN_HEADERS       = {"中文颜色", "颜色(中文)", "颜色（中文）", "color (cn)",
                            "colour (cn)", "cn color"}
        CODE_HEADERS     = {"中文颜色代码", "color code", "colour code", "colorcode"}

        def _h(v):
            return str(v).strip().lower() if v else ""

        inserted = updated = skipped = 0
        seen_keys: set = set()
        sheets_with_data = 0

        for sn in wb.sheetnames:
            try:
                ws = wb[sn]
            except Exception:
                continue
            cols = {}
            header_row = None
            for ri, row in enumerate(ws.iter_rows(max_row=5, values_only=True), 1):
                for ci, v in enumerate(row, 1):
                    h = _h(v)
                    if not h:
                        continue
                    if h in BRAND_HEADERS:    cols.setdefault("brand", ci)
                    elif h in EN_HEADERS:     cols.setdefault("en",    ci)
                    elif h in MAIN_HEADERS:   cols.setdefault("main",  ci)
                    elif h in CN_HEADERS:     cols.setdefault("cn",    ci)
                    elif h in CODE_HEADERS:   cols.setdefault("code",  ci)
                if "en" in cols and ("cn" in cols or "main" in cols or "code" in cols):
                    header_row = ri
                    break
            if not header_row:
                continue
            sheets_with_data += 1
            now = datetime.utcnow().isoformat()

            with self._conn() as conn:
                for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                    if cols["en"] > len(row):
                        continue
                    en_raw = row[cols["en"] - 1]
                    if en_raw is None or str(en_raw).strip() == "":
                        skipped += 1
                        continue

                    en_color = _normalize_color_name(en_raw)

                    def _cell(key):
                        ci = cols.get(key)
                        if ci and ci <= len(row) and row[ci - 1] is not None:
                            return str(row[ci - 1]).strip()
                        return ""

                    cn_color   = _cell("cn")
                    label      = _cell("main")
                    brand      = _cell("brand")
                    color_code = _cell("code")

                    key = (client, brand, en_color)
                    if key in seen_keys:
                        # Already processed in this run — skip duplicates
                        # within the same workbook
                        continue
                    seen_keys.add(key)

                    # Auto-derive shade + label when not in source.  We use
                    # the keyword classifier just like the buyplan exporter
                    # so the table stays consistent.
                    derived_shade, derived_label = _derive_shade_and_label(en_color)
                    shade_final = derived_shade
                    label_final = label or derived_label

                    existing = conn.execute(
                        "SELECT id, cn_color, color_code, light_or_dark, "
                        "label_color, notes FROM color_translations "
                        "WHERE client=? AND brand=? AND en_color=?",
                        key,
                    ).fetchone()
                    if existing:
                        # Preserve existing manual values when source is blank
                        old_cn    = existing["cn_color"]    or ""
                        old_code  = existing["color_code"]  or ""
                        old_label = existing["label_color"] or ""
                        new_cn    = cn_color    or old_cn
                        new_code  = color_code  or old_code
                        new_label = label_final or old_label
                        conn.execute(
                            """UPDATE color_translations SET
                                  cn_color=?, color_code=?, light_or_dark=?,
                                  label_color=?, updated_at=?
                               WHERE id=?""",
                            (new_cn, new_code, shade_final, new_label, now,
                             existing["id"]),
                        )
                        updated += 1
                        _audit_diff(
                            conn, existing["id"], client, brand, en_color,
                            {f: existing[f] for f in _AUDIT_FIELDS},
                            {"cn_color": new_cn,
                             "color_code": new_code,
                             "light_or_dark": shade_final,
                             "label_color": new_label,
                             "notes": existing["notes"] or ""},
                        )
                    else:
                        conn.execute(
                            """INSERT INTO color_translations
                                  (client, brand, en_color, cn_color,
                                   color_code, light_or_dark, label_color,
                                   notes, updated_at)
                               VALUES (?,?,?,?,?,?,?,?,?)""",
                            (client, brand, en_color, cn_color, color_code,
                             shade_final, label_final, "", now),
                        )
                        inserted += 1
                        new_id = conn.execute(
                            "SELECT id FROM color_translations "
                            "WHERE client=? AND brand=? AND en_color=?",
                            key,
                        ).fetchone()[0]
                        _audit_log(
                            conn, "insert", new_id, client, brand, en_color,
                            "*", None,
                            f"cn={cn_color!r}, code={color_code!r}, "
                            f"shade={shade_final!r}, label={label_final!r} "
                            f"(progress-xlsx import)",
                        )

        wb.close()
        return {
            "inserted":   inserted,
            "updated":    updated,
            "skipped":    skipped,
            "total_rows": inserted + updated,
            "sheets":     sheets_with_data,
        }

    # ── Queries ───────────────────────────────────────────────────────────────

    def get_all(self) -> list[dict]:
        with self._conn() as conn:
            rows = conn.execute(
                "SELECT * FROM color_translations ORDER BY client, brand, en_color"
            ).fetchall()
        return [dict(r) for r in rows]

    def get_by_client(self, client: str, brand: str = "") -> list[dict]:
        with self._conn() as conn:
            if brand:
                rows = conn.execute(
                    "SELECT * FROM color_translations WHERE client=? AND brand=? "
                    "ORDER BY en_color",
                    (client, brand)
                ).fetchall()
            else:
                rows = conn.execute(
                    "SELECT * FROM color_translations WHERE client=? ORDER BY brand, en_color",
                    (client,)
                ).fetchall()
        return [dict(r) for r in rows]

    def lookup(self, client: str, brand: str, en_color: str) -> str:
        """Return cn_color string for (client, brand, en_color), or ''.

        Look-up is case-insensitive — the en_color is normalised to title
        case before matching, and the underlying column is also compared
        with COLLATE NOCASE for safety with legacy rows that haven't been
        re-normalised yet.
        """
        en_color = _normalize_color_name(en_color)
        with self._conn() as conn:
            row = conn.execute(
                "SELECT cn_color FROM color_translations "
                "WHERE client=? AND brand=? AND en_color = ? COLLATE NOCASE",
                (client, brand, en_color)
            ).fetchone()
            if not row and brand:
                row = conn.execute(
                    "SELECT cn_color FROM color_translations "
                    "WHERE client=? AND brand='' AND en_color = ? COLLATE NOCASE",
                    (client, en_color)
                ).fetchone()
        return (row[0] or "") if row else ""

    def build_lookup_dict(self) -> dict[tuple, str]:
        """Return {(client, brand, normalised_en_color): cn_color} for fast
        batch lookups — keys are normalised to title case so callers can
        pass the raw English color regardless of source casing.
        """
        with self._conn() as conn:
            rows = conn.execute(
                "SELECT client, brand, en_color, cn_color FROM color_translations"
            ).fetchall()
        return {
            (r["client"], r["brand"], _normalize_color_name(r["en_color"])):
                r["cn_color"] or ""
            for r in rows
        }

    def build_label_lookup_dict(self) -> dict[tuple, str]:
        """Return {(client, brand, normalised_en_color): label_color}.
        Same case-insensitive keying as :py:meth:`build_lookup_dict`.
        """
        with self._conn() as conn:
            rows = conn.execute(
                "SELECT client, brand, en_color, label_color FROM color_translations"
            ).fetchall()
        return {
            (r["client"], r["brand"], _normalize_color_name(r["en_color"])):
                r["label_color"] or ""
            for r in rows
        }

    def build_cn_code_lookup_dict(self) -> dict[tuple, str]:
        """Return {(client, brand, normalised_en_color): color_code} (中文颜色代码).
        Same case-insensitive keying as :py:meth:`build_lookup_dict`.
        """
        with self._conn() as conn:
            rows = conn.execute(
                "SELECT client, brand, en_color, color_code FROM color_translations"
            ).fetchall()
        return {
            (r["client"], r["brand"], _normalize_color_name(r["en_color"])):
                r["color_code"] or ""
            for r in rows
        }

    def lookup_label_color(self, client: str, brand: str, en_color: str) -> str:
        """Return label_color for (client, brand, en_color), or '' if unset.

        Case-insensitive on en_color.  Falls back to brand-agnostic entry
        the same way ``lookup`` does.
        """
        en_color = _normalize_color_name(en_color)
        with self._conn() as conn:
            row = conn.execute(
                "SELECT label_color FROM color_translations "
                "WHERE client=? AND brand=? AND en_color = ? COLLATE NOCASE",
                (client, brand, en_color),
            ).fetchone()
            if not row and brand:
                row = conn.execute(
                    "SELECT label_color FROM color_translations "
                    "WHERE client=? AND brand='' AND en_color = ? COLLATE NOCASE",
                    (client, en_color),
                ).fetchone()
        return (row[0] or "") if row else ""

    def list_clients(self) -> list[str]:
        with self._conn() as conn:
            rows = conn.execute(
                "SELECT DISTINCT client FROM color_translations ORDER BY client"
            ).fetchall()
        return [r[0] for r in rows]

    def list_brands(self, client: str = "") -> list[str]:
        with self._conn() as conn:
            if client:
                rows = conn.execute(
                    "SELECT DISTINCT brand FROM color_translations "
                    "WHERE client=? ORDER BY brand", (client,)
                ).fetchall()
            else:
                rows = conn.execute(
                    "SELECT DISTINCT brand FROM color_translations ORDER BY brand"
                ).fetchall()
        return [r[0] for r in rows if r[0]]

    # ── Audit log ─────────────────────────────────────────────────────────────

    def audit_log(
        self,
        limit: int = 200,
        client: str = "",
        brand: str = "",
        en_color: str = "",
    ) -> list[dict]:
        """Return the most recent audit-log entries (newest first).

        Optional filters: client, brand, en_color (case-insensitive on
        en_color via the same normalisation used everywhere else).
        """
        sql = ("SELECT id, action, row_id, client, brand, en_color, field, "
               "old_value, new_value, changed_at, changed_by "
               "FROM color_translations_audit WHERE 1=1")
        params: list = []
        if client:
            sql += " AND client = ?"
            params.append(client)
        if brand:
            sql += " AND brand = ?"
            params.append(brand)
        if en_color:
            sql += " AND en_color = ? COLLATE NOCASE"
            params.append(_normalize_color_name(en_color))
        sql += " ORDER BY id DESC LIMIT ?"
        params.append(int(limit))
        with self._conn() as conn:
            rows = conn.execute(sql, params).fetchall()
        return [dict(r) for r in rows]

    def audit_log_count(self) -> int:
        with self._conn() as conn:
            return conn.execute(
                "SELECT COUNT(*) FROM color_translations_audit"
            ).fetchone()[0]

    def clear_audit_log(self) -> int:
        """Remove all audit-log rows.  Returns the number of rows deleted."""
        with self._conn() as conn:
            cur = conn.execute("DELETE FROM color_translations_audit")
        return cur.rowcount

    def count(self) -> int:
        with self._conn() as conn:
            return conn.execute("SELECT COUNT(*) FROM color_translations").fetchone()[0]

    # ── Export ────────────────────────────────────────────────────────────────

    def to_dataframe(self, client: str = "", brand: str = "") -> pd.DataFrame:
        cols = ["Client", "Brand", "English Color", "Chinese Color",
                "中文颜色代码", "Light/Dark", "Label Color", "Notes"]
        rows = self.get_by_client(client, brand) if client else self.get_all()
        if not rows:
            return pd.DataFrame(columns=cols)
        return pd.DataFrame([{
            "Client":        r["client"],
            "Brand":         r.get("brand") or "",
            "English Color": r["en_color"],
            "Chinese Color": r.get("cn_color") or "",
            "中文颜色代码":    r.get("color_code") or "",
            "Light/Dark":    r.get("light_or_dark") or "",
            "Label Color":   r.get("label_color") or "",
            "Notes":         r.get("notes") or "",
            "_id":           r["id"],
        } for r in rows])
