"""HHP / Zalando buy-plan exporter — one sheet per style.

Architecture
------------
This exporter **uses the BuyPlan_Template.xlsx** as the source of all
formatting (column widths, row heights, merged cells, borders, fonts,
photo boxes, bilingual headers).  It does **not** recreate any of that
in code.  Per-style sheets are produced by ``wb.copy_worksheet`` and only
the *data* cells are populated.

Template lookup order (same convention as ``buyplan_export.py``):
  1. ``data/buyplan_templates/zalando.xlsx``  (client-specific)
  2. ``data/buyplan_templates/default.xlsx``  (shared fallback)

Template structure (all positions 1-based, the template is the contract):
  L1, L2  — creation / modification date
  B2:B5   — fabric position labels (Main Body / Other 1-3)
  D2:D5   — fabric details (code|composition|weight|width)
  J3:L6   — front photo box (merged, filled by inject_style_photos)
  M3:O6   — back  photo box (merged, filled by inject_style_photos)
  Row 7-8 — bilingual column headers (English / Chinese) — preserved
  Row 9+  — data rows  (one per PO × ConfigSKU × ColorDesc group)
            Cols A-S map to 合同号 / 款号 / Brand / Article Name /
            PO No / Config SKU / 英文颜色 / 主标颜色 / 中文颜色 / 中文颜色代码 /
            K-P sizes / 船样要求 / TTL Total / X-FTY date.
  R5      — grand total cell (referenced by the Index sheet).

Index sheet
-----------
A separate "Index" sheet is created at workbook position 0.  Each row is
one style with a hyperlink to its sheet plus production-planning columns.
"""
from __future__ import annotations

import warnings
from copy import copy as _copy
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..utils.file_utils import versioned_path
from ..config import EXCEL_PALETTE as _P
from ._excel_helpers import clean_sheet_name, stable_unique, cell_value, apply_print_settings
from ._image_inject import inject_style_photos
from ._photo_utils import resolve_photo_pair


# ── Layout constants (must match the template) ───────────────────────────────
# Template column layout (1-based):
#   A=合同号  B=款号  C=Brand  D=Article Name  E=PO No.  F=Config SKU
#   G=英文颜色  H=主标颜色  I=中文颜色  J=中文颜色代码
#   K-P=sizes (XS-XXL)  Q=船样要求  R=TTL Total  S=X-FTY
SIZES = ["XS", "S", "M", "L", "XL", "XXL"]
DATA_START_ROW   = 9       # first PO data row
SIZE_COL_START   = 11      # column K
TOTAL_FORMULA_COL = 18     # column R
FACTORY_DATE_COL  = 19     # column S
GRAND_TOTAL_CELL  = "R5"
MAX_DATA_COL      = 19     # column S (last column we write to)

# ── Template path ─────────────────────────────────────────────────────────────
_TEMPLATES_DIR = Path(__file__).parent.parent.parent / "data" / "buyplan_templates"


def _resolve_template_path() -> Path | None:
    """Return the BuyPlan_Template.xlsx path, or None if not installed."""
    for stem in ("zalando", "default"):
        p = _TEMPLATES_DIR / f"{stem}.xlsx"
        if p.exists():
            return p
    return None


# ── Index-sheet planning headers ─────────────────────────────────────────────
_PLANNING_HEADERS = [
    "供应商(计划确认时间)", "工厂(计划确认时间)", "版样(计划确认时间)",
    "初版(计划打样时间)", "全版(计划打样时间)",
    "剪样(计划打样时间)", "剪样(计划打样时间2)", "剪样号",
    "单位(计划打样时间)", "出货(计划打样时间)", "出货号",
]

_HDRFILL = PatternFill(start_color=_P["hdr_blue"], end_color=_P["hdr_blue"],
                       fill_type="solid")


def _thin() -> Border:
    s = Side(border_style="thin", color=_P["border_grey"])
    return Border(left=s, right=s, top=s, bottom=s)


# ─────────────────────────────────────────────────────────────────────────────
# Public API
# ─────────────────────────────────────────────────────────────────────────────

def export_hhp_buyplan(
    df: pd.DataFrame,
    output_dir: str,
    *,
    photo_map: dict | None = None,
) -> str:
    """Build the HHP / Zalando buy-plan workbook and return the saved path.

    Parameters
    ----------
    df         : combined DataFrame from ``combine_excel_files``.  Must have
                 a ``Main Supplier Config SKU`` column (one sheet per value).
    output_dir : directory to save the workbook.
    photo_map  : optional dict — see :func:`resolve_photo_pair` for the
                 supported shapes (style → (front, back) tuple, or
                 filename → bytes).
    """
    photo_map = photo_map or {}
    path = versioned_path(output_dir, "Zalando_BuyPlan", ".xlsx")

    template_path = _resolve_template_path()
    if template_path is None:
        warnings.warn(
            "BuyPlan template not found in data/buyplan_templates/. "
            "Falling back to a plain workbook (no formatting)."
        )
        wb = Workbook()
        wb.remove(wb.active)
        template_ws = None
    else:
        wb = load_workbook(str(template_path))
        template_ws = wb.worksheets[0]

    # Index sheet — inserted at the front so it's the first tab the user sees
    ws_index = wb.create_sheet("Index", 0)
    _write_index_headers(ws_index)

    styles_in_order = stable_unique(
        df["Main Supplier Config SKU"].dropna().astype(str).str.strip()
        if "Main Supplier Config SKU" in df.columns else pd.Series(dtype=str)
    )

    # Pre-fetch fabric_master enrichment for all HHN codes used in this df —
    # one DB round-trip instead of N per style.  Used to fill weight (克重)
    # and width (有效门幅) into the column-D 综合 key.
    fm_cache = _build_fabric_master_cache(df)

    # Track sheet_name → style for the post-save photo injection
    sheet_style_map: dict[str, str] = {}

    index_row = 2
    for style in styles_in_order:
        sub = df[df["Main Supplier Config SKU"].astype(str).str.strip() == style].copy()
        if sub.empty:
            continue

        sheet_name = _unique_sheet_name(wb, clean_sheet_name(style))

        if template_ws is not None:
            ws = wb.copy_worksheet(template_ws)
            ws.title = sheet_name
            _fill_template_sheet(ws, style, sub, fm_cache)
        else:
            ws = wb.create_sheet(sheet_name)
            _fill_blank_sheet(ws, style, sub)

        sheet_style_map[sheet_name] = style
        _fill_index_row(ws_index, index_row, sheet_name, style, sub)
        index_row += 1

    # Remove the master template sheet now that all per-style copies exist
    if template_ws is not None and template_ws in wb.worksheets:
        wb.remove(template_ws)

    _set_index_widths(ws_index)
    apply_print_settings(wb)
    wb.save(path)

    # Inject photos (front in J3:L6, back in M3:O6) via zip-level patch.
    # openpyxl copy_worksheet drops drawings, so we do this after save.
    sheet_photo_map = _build_photo_map(df, sheet_style_map, photo_map)
    if sheet_photo_map:
        inject_style_photos(path, sheet_photo_map)

    return path


# ─────────────────────────────────────────────────────────────────────────────
# Per-style sheet filling
# ─────────────────────────────────────────────────────────────────────────────

def _fill_template_sheet(ws, style: str, sub: pd.DataFrame,
                         fm_cache: dict | None = None) -> None:
    """Populate the *data* of one sheet that was copied from the template.

    All formatting is inherited from the template — we only set values.
    For data rows past the template's row 9, we clone row 9's cell styles
    so additional rows match.

    *fm_cache* is the workbook-level fabric_master enrichment dict; pass
    None to skip DB lookup (the 综合 key will then only show columns that
    exist on the input row).
    """
    first_row = sub.iloc[0]

    # Date cells — merge K1:O1 / K2:O2 so the long datetime string is visible
    # even after the user's compact-layout override sets J–O width to 6.
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for rng in ("K1:O1", "K2:O2"):
        if rng not in (str(m) for m in ws.merged_cells.ranges):
            try:
                ws.merge_cells(rng)
            except Exception:
                pass
    ws["K1"] = now   # top-left of K1:O1
    ws["K2"] = now   # top-left of K2:O2

    # Fabric block (rows 2-5)
    _fill_fabric_block(ws, first_row, fm_cache)

    # ── Data rows ────────────────────────────────────────────────────────────
    group_cols = [c for c in ["Purchase Order Number", "Config SKU",
                              "Main Supplier Color Description"]
                  if c in sub.columns]
    groups = list(sub.groupby(group_cols, sort=False, dropna=False)) if group_cols \
             else [((), sub)]

    # Capture row 9's styling once — used to clone style onto rows 10+
    template_styles = _capture_row_styles(ws, DATA_START_ROW)

    last_row = DATA_START_ROW
    for g_idx, (_key, grp) in enumerate(groups):
        out_row = DATA_START_ROW + g_idx
        if out_row > DATA_START_ROW:
            _apply_row_styles(ws, out_row, template_styles)
        _write_data_row(ws, out_row, style, grp)
        last_row = out_row

    # Grand total — referenced from the Index sheet
    if groups:
        ws[GRAND_TOTAL_CELL] = (
            f"=SUM({get_column_letter(TOTAL_FORMULA_COL)}{DATA_START_ROW}"
            f":{get_column_letter(TOTAL_FORMULA_COL)}{last_row})"
        )

    # ── Compact-layout overrides (user-requested in v1.50.2) ───────────────
    # • All text font size = 10
    # • Column widths: J–O (sizes) = 6, others = 20
    # • Row heights = 20 for every row in use
    _apply_compact_layout(ws, last_row=last_row)


def _apply_compact_layout(ws, *, last_row: int) -> None:
    """Force the per-style sheet into a compact uniform layout.

    Overrides whatever the template defined:
      • font size 10 on every populated cell (preserves bold / colour)
      • column widths: A–I and P–R = 20, J–O = 6
      • row heights for rows 1..last_row+1 = 20 pt
    """
    from openpyxl.styles import Font

    # Font size = 10 on every cell with a value or an explicit style
    for row in ws.iter_rows():
        for cell in row:
            f = cell.font
            cell.font = Font(
                name=f.name, size=10, bold=f.bold, italic=f.italic,
                vertAlign=f.vertAlign, underline=f.underline, strike=f.strike,
                color=f.color,
            )

    # Column widths
    for col_idx in range(1, FACTORY_DATE_COL + 1):           # A..R
        letter = get_column_letter(col_idx)
        if SIZE_COL_START <= col_idx < SIZE_COL_START + len(SIZES):
            ws.column_dimensions[letter].width = 6           # J–O
        else:
            ws.column_dimensions[letter].width = 20          # all others

    # Row heights = 20 for every row up to the last data row (+ a buffer)
    for r in range(1, max(last_row + 2, 14)):
        ws.row_dimensions[r].height = 20


def _fill_fabric_block(ws, first_row, fm_cache: dict | None = None) -> None:
    """Write the four fabric rows into the template's B (position) and
    D (combined ``code|composition|gsm|width`` key) columns.  Empty rows
    are left untouched.

    *fm_cache* is the fabric_master enrichment dict (HHN → record);
    pass None for a no-DB fallback.
    """
    fm_cache = fm_cache or {}
    fabric_specs = [
        (2, "",   "面料_面料"),         # main fabric (no numeric suffix)
        (3, "1",  "面料_面料1"),
        (4, "2",  "面料_面料2"),
        (5, "3",  "面料_面料3"),
    ]
    for row, suffix, full_key in fabric_specs:
        position = cell_value(first_row, f"面料_面料{suffix}_部位")
        if position:
            ws.cell(row=row, column=2).value = position

        details = _fabric_details(first_row, suffix, full_key, fm_cache)
        if details:
            # Always column D = 4.  Explicitly clear E so any prior
            # overflow / sample value the template may have had cannot
            # mislead the user into thinking the value sits in E.
            ws.cell(row=row, column=4).value = details
            ws.cell(row=row, column=5).value = None


def _fabric_details(first_row, suffix: str, full_key: str,
                    fm_cache: dict) -> str | None:
    """Return ``code|composition|gsm|width`` for one fabric row.

    Resolution order:
      1. Pre-combined ``面料_面料{suffix}`` value if it already contains
         pipes — used as-is unless we can enrich it with DB-supplied
         weight / width when those are missing.
      2. Look the HHN code up in the fabric_master DB (composition,
         weight_gsm, cuttable_width_cm) and assemble the key.
      3. Fall back to whatever individual columns exist on *first_row*.
    """
    code = cell_value(first_row, f"面料_面料{suffix}_编号") or ""
    pre  = cell_value(first_row, full_key)

    # Pull from fabric_master DB if we have a code and a cache entry
    rec = fm_cache.get(code) if code else None
    db_comp  = (rec or {}).get("composition_en") or ""
    db_gsm   = (rec or {}).get("weight_gsm")
    db_width = (rec or {}).get("cuttable_width_cm")

    # Helper to format gsm / width as plain integer strings (or "")
    def _num(v):
        if v in (None, "", 0):
            return ""
        try:
            return str(int(float(v)))
        except (ValueError, TypeError):
            return ""

    gsm_s   = _num(db_gsm)
    width_s = _num(db_width)

    # Case 1: pre-combined value already pipe-separated
    if pre and "|" in pre:
        parts = pre.split("|")
        # Pad to 4 fields
        while len(parts) < 4:
            parts.append("")
        # Fill blanks from DB if available
        if not parts[0].strip() and code:
            parts[0] = code
        if not parts[2].strip() and gsm_s:
            parts[2] = gsm_s
        if not parts[3].strip() and width_s:
            parts[3] = width_s
        return "|".join(parts[:4])

    # Case 2: assemble from DB
    comp_s = (cell_value(first_row, f"面料_面料{suffix}_成分") or db_comp or "")
    if any((code, comp_s, gsm_s, width_s)):
        return f"{code}|{comp_s}|{gsm_s}|{width_s}"
    return None


def _write_data_row(ws, out_row: int, style: str, grp: pd.DataFrame) -> None:
    """Write one data row.  Styles are assumed to be already applied."""
    g_first = grp.iloc[0]

    ws.cell(out_row, 1).value  = cell_value(g_first, "合同号")
    ws.cell(out_row, 2).value  = style
    ws.cell(out_row, 3).value  = cell_value(g_first, "Brand")
    ws.cell(out_row, 4).value  = cell_value(g_first, "Article Name")
    ws.cell(out_row, 5).value  = cell_value(g_first, "Purchase Order Number")
    ws.cell(out_row, 6).value  = cell_value(g_first, "Config SKU")
    ws.cell(out_row, 7).value  = cell_value(g_first, "Main Supplier Color Description")  # 英文颜色

    ws.cell(out_row, 8).value  = cell_value(g_first, "主标颜色")                           # H: label color

    # I: 中文颜色 — combine as  #中文颜色代码|中文颜色  when both are present.
    # Fall back to the source '颜色' field (pre-filled in some Zalando sheets).
    _cn_name = cell_value(g_first, "中文颜色") or cell_value(g_first, "颜色") or ""
    _cn_code = cell_value(g_first, "中文颜色代码") or ""
    ws.cell(out_row, 9).value  = f"#{_cn_code}|{_cn_name}" if (_cn_code and _cn_name) \
                                 else (_cn_name or None)

    ws.cell(out_row, 10).value = _cn_code or None                                        # J: 中文颜色代码 (standalone)

    # Sizes J-O
    for s_idx, sz in enumerate(SIZES):
        qty = int(grp[sz].sum()) if sz in grp.columns else 0
        ws.cell(out_row, SIZE_COL_START + s_idx).value = qty if qty else None

    # Row total formula in Q
    ws.cell(out_row, TOTAL_FORMULA_COL).value = (
        f"=SUM({get_column_letter(SIZE_COL_START)}{out_row}"
        f":{get_column_letter(SIZE_COL_START + len(SIZES) - 1)}{out_row})"
    )

    # X-FTY date in R
    ws.cell(out_row, FACTORY_DATE_COL).value = cell_value(g_first, "入厂时间")


# ─────────────────────────────────────────────────────────────────────────────
# Style cloning helpers (for data rows past row 9)
# ─────────────────────────────────────────────────────────────────────────────

def _capture_row_styles(ws, row: int, max_col: int = MAX_DATA_COL) -> list[dict]:
    """Snapshot per-cell styling of *row* in *ws* for later replay."""
    out: list[dict] = []
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        out.append({
            "font":          _copy(cell.font),
            "fill":          _copy(cell.fill),
            "alignment":     _copy(cell.alignment),
            "border":        _copy(cell.border),
            "number_format": cell.number_format,
        })
    return out


def _apply_row_styles(ws, dst_row: int, styles: list[dict]) -> None:
    """Apply previously captured *styles* to *dst_row*."""
    for col_idx, st in enumerate(styles, start=1):
        c = ws.cell(row=dst_row, column=col_idx)
        c.font          = st["font"]
        c.fill          = st["fill"]
        c.alignment     = st["alignment"]
        c.border        = st["border"]
        c.number_format = st["number_format"]


# ─────────────────────────────────────────────────────────────────────────────
# Photo plumbing
# ─────────────────────────────────────────────────────────────────────────────

def _build_fabric_master_cache(df: pd.DataFrame) -> dict:
    """One DB round-trip to fetch enrichment (composition / gsm / width) for
    every HHN code referenced anywhere in *df*.  Returns ``{hhn: record}``.

    Returns an empty dict if the fabric_master DB is unavailable.
    """
    code_cols = [c for c in (
        "面料_面料_编号", "面料_面料1_编号", "面料_面料2_编号", "面料_面料3_编号"
    ) if c in df.columns]
    if not code_cols:
        return {}

    hhns: set[str] = set()
    for col in code_cols:
        for v in df[col].dropna().astype(str):
            v = v.strip()
            if v and v.lower() != "nan":
                hhns.add(v)
    if not hhns:
        return {}

    try:
        from ..store import get_fabric_master_store
        return get_fabric_master_store().get_batch_enrichment(list(hhns)) or {}
    except Exception as exc:
        warnings.warn(f"[buyplan] fabric_master lookup failed: {exc!r}")
        return {}


def _to_bytes(photo) -> bytes | None:
    """Normalise a resolved photo (bytes or path) to raw bytes, or None."""
    if isinstance(photo, (bytes, bytearray)):
        return bytes(photo)
    if isinstance(photo, str):
        try:
            with open(photo, "rb") as fh:
                return fh.read()
        except Exception:
            return None
    return None


def _build_photo_map(
    df: pd.DataFrame,
    sheet_style_map: dict[str, str],
    photo_map: dict,
) -> dict[str, dict[str, bytes | None]]:
    """Build ``{sheet_title: {'front': bytes|None, 'back': bytes|None}}``.

    Path strings returned by :func:`resolve_photo_pair` are read from disk
    here so the post-save zip patcher gets raw bytes.

    Emits a one-line summary via ``warnings.warn`` showing how many photos
    were resolved (visible in the Streamlit terminal output) — helpful when
    "the picture is not showing" needs diagnosing.
    """
    out: dict[str, dict[str, bytes | None]] = {}
    style_col = "Main Supplier Config SKU"

    misses: list[str] = []
    n_front = n_back = 0

    for sheet_name, style in sheet_style_map.items():
        sub = df[df[style_col].astype(str).str.strip() == style] \
              if style_col in df.columns else pd.DataFrame()
        if sub.empty:
            continue
        front_raw, back_raw = resolve_photo_pair(style, sub.iloc[0], photo_map)
        front = _to_bytes(front_raw)
        back  = _to_bytes(back_raw)
        out[sheet_name] = {"front": front, "back": back}
        if front is not None:
            n_front += 1
        if back is not None:
            n_back += 1
        if front is None and back is None:
            misses.append(style)

    warnings.warn(
        f"[buyplan photos] photo_map keys={len(photo_map)} | "
        f"styles={len(sheet_style_map)} | "
        f"front-found={n_front} back-found={n_back} | "
        f"missed={misses[:5]}"
        + (f" (+{len(misses) - 5} more)" if len(misses) > 5 else "")
    )
    return out


# ─────────────────────────────────────────────────────────────────────────────
# Index sheet
# ─────────────────────────────────────────────────────────────────────────────

def _write_index_headers(ws) -> None:
    headers = ["No.", "款号", "客户品号", "面料_面料", "面料_面料_编号",
               "总数量合计", "入厂时间"] + _PLANNING_HEADERS
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color=_P["white"])
        c.fill = _HDRFILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _thin()
    ws.freeze_panes = "A2"


def _fill_index_row(ws_index, index_row: int, sheet_name: str,
                    style: str, sub: pd.DataFrame) -> None:
    first_row = sub.iloc[0]
    ws_index.cell(index_row, 1).value = index_row - 1

    link_cell = ws_index.cell(index_row, 2)
    link_cell.value = style
    link_cell.hyperlink = f"#'{sheet_name}'!A1"
    link_cell.font = Font(color="FF0000FF", underline="single")

    ws_index.cell(index_row, 3).value = cell_value(first_row, "Brand")
    ws_index.cell(index_row, 4).value = cell_value(first_row, "面料_面料")
    ws_index.cell(index_row, 5).value = cell_value(first_row, "面料_面料_编号")
    ws_index.cell(index_row, 6).value = f"='{sheet_name}'!{GRAND_TOTAL_CELL}"
    ws_index.cell(index_row, 6).font  = Font(bold=True)
    ws_index.cell(index_row, 7).value = cell_value(first_row, "入厂时间")

    for col in range(1, 8 + len(_PLANNING_HEADERS)):
        c = ws_index.cell(index_row, col)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _thin()


def _set_index_widths(ws_index) -> None:
    ws_index.column_dimensions["A"].width = 5
    ws_index.column_dimensions["B"].width = 28
    for col_ltr in ("C", "D", "E"):
        ws_index.column_dimensions[col_ltr].width = 22
    ws_index.column_dimensions["F"].width = 14
    ws_index.column_dimensions["G"].width = 18
    for i, _ in enumerate(_PLANNING_HEADERS, start=8):
        ws_index.column_dimensions[get_column_letter(i)].width = 20
    ws_index.freeze_panes = "B2"


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _unique_sheet_name(wb, base: str) -> str:
    existing = {s.title for s in wb.worksheets}
    if base not in existing:
        return base
    stem = base[:28]
    sfx = 2
    while f"{stem}_{sfx}" in existing:
        sfx += 1
    return f"{stem}_{sfx}"


# ─────────────────────────────────────────────────────────────────────────────
# Fallback for missing template — minimal layout matching column scheme
# ─────────────────────────────────────────────────────────────────────────────

def _fill_blank_sheet(ws, style: str, sub: pd.DataFrame) -> None:
    """Write a no-frills version when the template is missing.  Provides
    correct column layout but no merges / styling / photos.
    """
    headers = ["合同号", "款号", "Brand", "Article Name", "PO No.",
               "Config SKU", "英文颜色", "主标颜色", "中文颜色", "中文颜色代码"] + SIZES + \
              ["", "Total", "X-FTY"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=DATA_START_ROW - 1, column=col, value=h)

    group_cols = [c for c in ["Purchase Order Number", "Config SKU",
                              "Main Supplier Color Description"]
                  if c in sub.columns]
    groups = list(sub.groupby(group_cols, sort=False, dropna=False)) if group_cols \
             else [((), sub)]
    last_row = DATA_START_ROW
    for g_idx, (_k, grp) in enumerate(groups):
        out_row = DATA_START_ROW + g_idx
        _write_data_row(ws, out_row, style, grp)
        last_row = out_row
    if groups:
        ws[GRAND_TOTAL_CELL] = (
            f"=SUM({get_column_letter(TOTAL_FORMULA_COL)}{DATA_START_ROW}"
            f":{get_column_letter(TOTAL_FORMULA_COL)}{last_row})"
        )
