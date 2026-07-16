"""Private helpers for Sky East buy-plan exporters (constants + utility functions).

Re-exported via ``from ._sky_east_helpers import *`` in sky_east_buyplan_export.py.
Names prefixed with `_` are excluded from `import *` by default, so we publish
them explicitly through ``__all__`` (regression fix: NameError: '_SE_TEMPLATE'
is not defined when generating buy plans).
"""
from __future__ import annotations

import re
from pathlib import Path

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from auth.companies import COMPANY_SKY_EAST
from ._excel_helpers import set_internal_hyperlink

__all__ = [
    # Template paths
    "_DATA_DIR", "_TEMPLATES_DIR", "_SE_TEMPLATE", "_SE_TEMPLATE_P",
    # Size lists
    "_SIZES_LC", "_SIZES_UC",
    # Fallback column positions
    "_COL_CONTRACT", "_COL_STYLE", "_COL_BRAND", "_COL_ARTICLE", "_COL_PO",
    "_COL_CONFIG", "_COL_COLOR_EN", "_COL_COLOR_CN", "_COL_LABEL_CLR",
    "_COL_XS", "_COL_S", "_COL_M", "_COL_L", "_COL_XL", "_COL_XXL",
    "_COL_BOAT_SAMPLE", "_COL_TOTAL", "_COL_EXFTY", "_COL_RETURN_LABEL",
    # Fabric header fallback column positions
    "_COL_COMPOSITION", "_COL_DISPLAY_KEY",
    "_DATA_ROW_FB",
    # Header alias dictionaries
    "_NUK_COLOR_ALIASES", "_NUK_SIZE_ALIASES",
    # Helper functions
    "_norm", "_thin",
    "_apply_config_overrides", "_clean_sheet_name",
    "_clear_data_area", "_cn_color", "_create_index_sheet", "_create_overview_sheet",
    "_fabric_version_note",
    "_detect_buyplan_layout", "_detect_fabric_rows", "_detect_nukuryou_layout",
    "_embed_style_photos", "_prep_image_for_embed",
    "_replace_placeholders", "_set_sheet_column_widths", "_strip_color_brackets",
    "_style_data", "_style_total",
    "derive_main_label_color",
    "_COLOR_NOT_FOUND", "_color_miss_comment_text",
]

# ---------------------------------------------------------------------------
# Colour resolution — explicit "not found" marker
# ---------------------------------------------------------------------------
# Distinct from a soft blank/"NA" (matched, but that field was empty in the
# source row) — this means the selected colour source has no entry at all
# for the (style, colour) key.  Shared by both buyplan exporters and the
# Overview sheet builder below, which attaches a diagnostic comment to any
# cell showing this value.
_COLOR_NOT_FOUND = "未找到"


def _color_miss_comment_text(client_po_color: str, progress_colors: list[str] | None) -> str:
    """Build the diagnostic comment text for a 未找到 colour cell.

    Always shows the client's PO colour exactly as the order file had it
    (before bracket-stripping / splitting). When *progress_colors* is a
    (possibly empty) list — meaning 大货进度表 is the selected source — a
    second line lists the colour name(s) 大货进度表 actually has on file for
    that PC No./Style, so a reviewer can spot a naming mismatch (e.g. client
    says "Dark Blue", 大货进度表 says "Navy") without opening the source file.
    ``None`` (the internal-DB-source case) omits that second line entirely.
    """
    lines = [f"Client's PO colour: \"{client_po_color}\""]
    if progress_colors is not None:
        if progress_colors:
            lines.append(f"大货进度表 colour(s) on file: {', '.join(progress_colors)}")
        else:
            lines.append("大货进度表 has no colour recorded for this PC No./Style")
    return "\n".join(lines)

# ---------------------------------------------------------------------------
# Template paths
# ---------------------------------------------------------------------------
_DATA_DIR      = Path(__file__).parent.parent.parent / "data"
_TEMPLATES_DIR = _DATA_DIR / "buyplan_templates"
_SE_TEMPLATE   = _TEMPLATES_DIR / "Sky_East.xlsx"
_SE_TEMPLATE_P = _TEMPLATES_DIR / "Sky_East_P.xlsx"

# Sky East size columns
_SIZES_LC = ["xs", "s", "m", "l", "xl", "xxl"]
_SIZES_UC = ["XS", "S", "M", "L", "XL", "XXL"]

# ---------------------------------------------------------------------------
# Fallback column positions (1-based) — used when header detection fails.
# These reflect the canonical Sky_East.xlsx template layout.
# ---------------------------------------------------------------------------
_COL_CONTRACT  =  1   # A  合同号
_COL_STYLE     =  2   # B  款号
_COL_BRAND     =  3   # C  Brand
_COL_ARTICLE   =  4   # D  Article Name
_COL_PO        =  5   # E  PO Number
_COL_CONFIG    =  6   # F  Config SKU
_COL_COLOR_EN  =  7   # G  ColorDesc
_COL_COLOR_CN  =  8   # H  颜色
_COL_LABEL_CLR =  9   # I  主标色
_COL_XS        = 10   # J  XS
_COL_S         = 11   # K  S
_COL_M         = 12   # L  M
_COL_L         = 13   # M  L
_COL_XL        = 14   # N  XL
_COL_XXL       = 15   # O  XXL
_COL_BOAT_SAMPLE = 16 # P  船样要求 (populated from BoatSampleStore when available)
_COL_TOTAL     = 17   # Q  Total
_COL_EXFTY     = 18   # R  离厂时间
_COL_RETURN_LABEL = 19  # S  Return Label (Yes/No/NA) — new column, one past ex_fty
_DATA_ROW_FB   =  8   # Fallback data start row

# Fabric-header section fallback column positions (rows 2-5).
# Column D in the template's row 1 has the literal header
# "面料编号|成分|克重|有效门幅" — i.e. column D is the 综合标识Key column.
# The legacy code split the value across D (composition) and E (display key);
# we now write the *full* combined key to D and leave E empty.
_COL_COMPOSITION =  4   # D  (kept for backward-compat, same as DISPLAY_KEY now)
_COL_DISPLAY_KEY =  4   # D  综合标识Key (quality_no|composition_en|gsm|width)

# Header aliases: logical field → set of recognised header strings (lowercased).
_BUY_PLAN_COL_ALIASES: dict[str, set[str]] = {
    "contract":  {"合同号", "contract no", "contract no.", "contract"},
    "style":     {"款号", "style", "style no", "style no."},
    "brand":     {"brand", "客户品牌", "客户品牌:"},
    "article":   {"article name", "supplier article name", "article"},
    "po":        {"po number", "po no.", "po no", "po#", "po"},
    "config":    {"config sku", "config_sku", "configsku"},
    "color_en":  {"colordesc", "color name", "color (en)", "颜色（英文）", "colour name"},
    "color_cn":  {"颜色", "color (cn)", "颜色（中文）", "colour"},
    "label_clr": {"主标色", "colour code", "color code", "label color"},
    "xs":        {"xs"},
    "s":         {"s"},
    "m":         {"m"},
    "l":         {"l"},
    "xl":        {"xl"},
    "xxl":       {"xxl", "2xl", "2xl"},
    # 船样 = "shipping sample" (a pre-shipment sample), not "boat sample" — the
    # old English aliases are kept for backward compatibility with existing
    # templates that already use that wording.
    "boat_sample": {"船样要求", "shipping sample", "shipping sample req",
                    "boat sample", "boat sample req", "船样"},
    "total":     {"total", "订单数合计", "qty", "total qty", "数量合计"},
    "ex_fty":    {"离厂时间", "ex-fty", "ex fty", "exfty", "ex_fty"},
    "return_label": {"return label", "退货标签", "退货唛头", "需要挂 return label"},
}

# For nukuryou (Sky_East_P.xlsx) — same size aliases plus color column
_NUK_COLOR_ALIASES  = {"color name", "colour name", "color", "颜色", "颜色（英文）"}
_NUK_SIZE_ALIASES   = {"xs", "s", "m", "l", "xl", "xxl", "2xl"}


# ---------------------------------------------------------------------------
# Template column / row detection
# ---------------------------------------------------------------------------

def _norm(s) -> str:
    """Lowercase and strip for header matching."""
    return str(s).strip().lower() if s is not None else ""


def _detect_buyplan_layout(ws) -> tuple[dict[str, int], int]:
    """Scan the template worksheet and return (col_map, data_start_row).

    Scans rows 1-25 for the header row — identified as the first row that
    contains both a style-like header and at least one size label.
    Returns a dict mapping logical field names to 1-based column numbers.
    Falls back to the module-level _COL_* constants for any field not found.
    """
    # Build normalised alias → logical_field lookup
    alias_to_field: dict[str, str] = {}
    for field, aliases in _BUY_PLAN_COL_ALIASES.items():
        for alias in aliases:
            alias_to_field.setdefault(_norm(alias), field)

    header_row_num = None
    col_map: dict[str, int] = {}

    for r in range(1, 26):
        row_map: dict[str, int] = {}
        for cell in ws[r]:
            if cell.value is None:
                continue
            norm = _norm(cell.value)
            field = alias_to_field.get(norm)
            if field and field not in row_map:
                row_map[field] = cell.column

        # Accept this row as the header if it has ≥3 recognised fields
        # AND at least one size column
        size_fields = {"xs", "s", "m", "l", "xl", "xxl"}
        if len(row_map) >= 3 and row_map.keys() & size_fields:
            header_row_num = r
            col_map = row_map
            break

    # Fill fallbacks for any missing field
    fallbacks = {
        "contract": _COL_CONTRACT, "style":    _COL_STYLE,
        "brand":    _COL_BRAND,    "article":  _COL_ARTICLE,
        "po":       _COL_PO,       "config":   _COL_CONFIG,
        "color_en": _COL_COLOR_EN, "color_cn": _COL_COLOR_CN,
        "label_clr":_COL_LABEL_CLR,
        "xs": _COL_XS, "s": _COL_S, "m": _COL_M, "l": _COL_L,
        "xl": _COL_XL, "xxl": _COL_XXL,
        "boat_sample": _COL_BOAT_SAMPLE,
        "total": _COL_TOTAL, "ex_fty": _COL_EXFTY,
        "return_label": _COL_RETURN_LABEL,
    }
    for field, col in fallbacks.items():
        col_map.setdefault(field, col)

    data_start = (header_row_num + 1) if header_row_num else _DATA_ROW_FB
    return col_map, data_start


def _detect_nukuryou_layout(ws) -> tuple[int, dict[str, int], int]:
    """Scan Sky_East_P template and return (color_col, size_col_map, data_start_row).

    Scans rows 1-10 for a row that contains recognised size labels.
    color_col is the first column in that row that is NOT a size label
    (defaults to column 1 if nothing else is found).
    """
    size_col_map: dict[str, int] = {}   # "xs"/"s"/… → col_num
    color_col = 1
    size_header_row = None

    for r in range(1, 11):
        row_sizes: dict[str, int] = {}
        non_size_cols: list[int] = []
        color_found = False
        for cell in ws[r]:
            if cell.value is None:
                continue
            norm = _norm(cell.value)
            if norm in _NUK_SIZE_ALIASES:
                sz = "xxl" if norm in ("2xl", "xxl") else norm
                row_sizes.setdefault(sz, cell.column)
            elif norm in _NUK_COLOR_ALIASES:
                color_col = cell.column
                color_found = True
            else:
                non_size_cols.append(cell.column)

        if len(row_sizes) >= 3:   # found the size header row
            size_header_row = r
            size_col_map = row_sizes
            if not color_found and non_size_cols:
                # No 颜色/Color header in this row — fall back to the
                # leftmost non-size column.
                color_col = min(non_size_cols)
            break

    # Fallbacks: canonical B-G layout
    if not size_col_map:
        size_col_map = {"xs": 2, "s": 3, "m": 4, "l": 5, "xl": 6, "xxl": 7}
        color_col = 1

    data_start = (size_header_row + 1) if size_header_row else 3
    return color_col, size_col_map, data_start


def _apply_config_overrides(
    pipeline_id: str,
    col: dict[str, int],
    data_row: int,
    fabric_rows: list[tuple[int, int, int, int]],
) -> tuple[dict[str, int], int, list[tuple[int, int, int, int]]]:
    """Apply per-pipeline JSON overrides to the auto-detected layout.

    Reads the JSON config registered under *pipeline_id* (see
    ``template_config``) and returns adjusted ``(col, data_row, fabric_rows)``.
    Missing/blank keys leave auto-detection in place.
    """
    try:
        from . import template_config as _tc
    except Exception:
        return col, data_row, fabric_rows

    cfg = _tc.load_config(pipeline_id)

    # Logical-field aliases used in column_map (lowercase).  We map both the
    # user-facing names ("Style", "PO Number") and the internal short codes
    # ("style", "po") to the same exporter slot.
    field_aliases = {
        "contract":   "contract",  "contract no": "contract",  "合同号": "contract",
        "style":      "style",     "款号": "style",
        "brand":      "brand",
        "article":    "article",   "article name": "article",
        "po":         "po",        "po number": "po",  "po no": "po",
        "config":     "config",    "config sku": "config",  "configsku": "config",
        "color_en":   "color_en",  "color":     "color_en",  "colordesc": "color_en",
        "color_cn":   "color_cn",  "color (cn)": "color_cn",  "颜色": "color_cn",
        "label_clr":  "label_clr", "主标色": "label_clr",
        "total":      "total",     "qty": "total",  "数量合计": "total",
        "ex_fty":     "ex_fty",    "xfactory_date": "ex_fty",  "ex-fty": "ex_fty",
        "xs": "xs", "s": "s", "m": "m", "l": "l", "xl": "xl", "xxl": "xxl", "2xl": "xxl",
    }

    new_col = dict(col)
    for raw_key, raw_val in (cfg.get("column_map") or {}).items():
        slot = field_aliases.get(str(raw_key).strip().lower())
        if not slot:
            continue
        try:
            new_col[slot] = _tc.column_letter_to_int(raw_val)
        except Exception:
            pass

    for size, raw_val in (cfg.get("size_column_map") or {}).items():
        slot = field_aliases.get(str(size).strip().lower())
        if not slot:
            continue
        try:
            new_col[slot] = _tc.column_letter_to_int(raw_val)
        except Exception:
            pass

    for meta, raw_val in (cfg.get("meta_column_map") or {}).items():
        slot = field_aliases.get(str(meta).strip().lower())
        if not slot:
            continue
        try:
            new_col[slot] = _tc.column_letter_to_int(raw_val)
        except Exception:
            pass

    new_data_row = data_row
    if cfg.get("data_start_row"):
        try:
            new_data_row = int(cfg["data_start_row"])
        except (TypeError, ValueError):
            pass
    elif cfg.get("header_row"):
        # header_row is one above the data block, matching the GIII convention
        try:
            new_data_row = int(cfg["header_row"]) + 1
        except (TypeError, ValueError):
            pass

    new_fabric_rows = list(fabric_rows)
    cfg_slots = cfg.get("fabric_slots") or []
    if cfg_slots:
        new_fabric_rows = []
        for s in cfg_slots:
            try:
                r  = int(s.get("row") or 0)
                bc = _tc.column_letter_to_int(s.get("body_part") or s.get("body") or "B")
                hc = _tc.column_letter_to_int(s.get("hhn")       or "C")
                cc = _tc.column_letter_to_int(s.get("composition") or s.get("comp") or "D")
                kc = _tc.column_letter_to_int(s.get("key")       or "E")
                if r > 0:
                    new_fabric_rows.append((r, bc, hc, cc, kc))
            except Exception:
                continue

    return new_col, new_data_row, new_fabric_rows


def _detect_fabric_rows(ws, max_row: int = 7) -> list[tuple[int, int, int, int, int]]:
    """Find fabric header rows above the data area.

    Scans rows 1 to max_row and returns a list of
    (row, body_part_col, hhn_col, composition_col, display_key_col) for each row
    that looks like a fabric-slot row (contains an HHN-like value or a cell
    labelled with a body-part keyword).

    The five positions map to the canonical Sky_East.xlsx layout:
        B  = 大身 / body part label
        C  = HHN code (公司面料编号)
        D  = 面料成分 (fabric composition text)
        E  = 综合标识Key (quality_no|composition_en|gsm|width)

    Falls back to the canonical template positions (rows 2-5, cols B/C/D/E)
    when no fabric rows are auto-detected.
    """
    # Look for rows that have an HHN code pattern or a fabric keyword
    import re as _re
    _HHN_LIKE = _re.compile(r'[A-Za-z]{2,5}-[A-Za-z]{1,5}-?\d{4,8}')
    _BODY_KW   = {"大身", "身", "lining", "body", "shell", "inner", "outer",
                  "网布", "针织", "梭织"}

    found: list[tuple[int, int, int, int, int]] = []

    for r in range(1, max_row + 1):
        cols_with_value: dict[int, str] = {}
        for cell in ws[r]:
            if cell.value is not None:
                cols_with_value[cell.column] = _norm(str(cell.value))

        # Heuristic: row has a body-part keyword or an HHN-like value
        has_hhn  = any(_HHN_LIKE.search(str(ws.cell(r, c).value or ""))
                       for c in cols_with_value)
        has_body = any(v in _BODY_KW for v in cols_with_value.values())
        if not (has_hhn or has_body):
            continue

        # Best-guess column assignments from sorted non-empty cols.
        # The template has four fabric columns: body / HHN / composition / display_key.
        sorted_cols = sorted(cols_with_value.keys())
        body_col = sorted_cols[0] if len(sorted_cols) >= 1 else 2
        hhn_col  = sorted_cols[1] if len(sorted_cols) >= 2 else 3
        comp_col = sorted_cols[2] if len(sorted_cols) >= 3 else _COL_COMPOSITION
        dk_col   = sorted_cols[3] if len(sorted_cols) >= 4 else _COL_DISPLAY_KEY
        found.append((r, body_col, hhn_col, comp_col, dk_col))

    # Fallback: canonical rows 2-5, cols B(2) / C(3) / D(4) / E(5)
    if not found:
        found = [(r, 2, 3, _COL_COMPOSITION, _COL_DISPLAY_KEY) for r in range(2, 6)]

    return found


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

def _thin() -> Border:
    s = Side(border_style="thin", color="FF000000")
    return Border(left=s, right=s, top=s, bottom=s)


_DATA_ROW_HEIGHT_BASE = 15.75   # Excel default line height in points
_DATA_ROW_HEIGHT_MIN  = 18      # minimum row height even for single-line rows

# Per-column caps: (min_width, max_width) in Excel column-width units.
# Text columns wrap when content exceeds max_width.
# Size/numeric columns stay compact and never wrap.
_COL_CAPS: dict[str, tuple[float, float]] = {
    # field        min   max
    "contract": ( 8,   22),
    "style":    ( 8,   22),
    "brand":    ( 8,   22),
    "article":  (10,   32),   # long names wrap inside 32-char column
    "po":       ( 8,   18),
    "config":   ( 8,   22),
    "color_en": ( 8,   22),
    "color_cn": ( 8,   20),
    "label_clr":( 6,   14),
    "xs":       ( 5,    7),   # size cols — never wrap
    "s":        ( 5,    7),
    "m":        ( 5,    7),
    "l":        ( 5,    7),
    "xl":       ( 5,    7),
    "xxl":      ( 5,    8),
    "total":    ( 6,   10),
    "ex_fty":   ( 8,   16),
}
_BOAT_SAMPLE_MIN  =  8
_BOAT_SAMPLE_MAX  = 24
# Fields that should never wrap (size / numeric columns)
_NO_WRAP_FIELDS = {"xs", "s", "m", "l", "xl", "xxl", "total", "label_clr"}


def _cell_display_len(value) -> float:
    """Approximate display width of a cell value in Excel column-width units.

    CJK / full-width characters each count as ~2 units; ASCII as 1.
    A small fudge factor (×1.05) accounts for font kerning.
    """
    if value is None:
        return 0
    s = str(value)
    n = sum(2 if ord(c) > 0x2E7F else 1 for c in s)
    return n * 1.05


def _set_sheet_column_widths(ws, col: dict[str, int],
                              data_start_row: int = 6) -> None:
    """Dynamically size columns and set row heights for a buy-plan style sheet.

    Algorithm
    ---------
    1. Scan every cell → compute max content display width per column.
    2. Clamp to per-field (min, max) caps.  Text columns whose content exceeds
       the cap get ``wrap_text=True``; numeric/size columns never wrap.
    3. For each data row (≥ *data_start_row*) calculate how many wrapped lines
       each cell needs given the final column width, then set the row height to
       ``max_lines × line_height``.
    """
    import math
    from openpyxl.utils import get_column_letter as _gcl

    # Reverse map: col_num → field name
    col_to_field: dict[int, str] = {v: k for k, v in col.items()}

    # ── Pass 1: find max content width per column ─────────────────────────────
    max_content: dict[int, float] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            w = _cell_display_len(cell.value)
            if w > max_content.get(cell.column, 0):
                max_content[cell.column] = w

    # ── Pass 2: compute final column widths ───────────────────────────────────
    final_widths: dict[int, float] = {}
    for cn, raw_w in max_content.items():
        field = col_to_field.get(cn)
        if field and field in _COL_CAPS:
            lo, hi = _COL_CAPS[field]
        elif cn == 16:          # col P 船样要求
            lo, hi = _BOAT_SAMPLE_MIN, _BOAT_SAMPLE_MAX
        else:
            lo, hi = 8, 40
        final_widths[cn] = max(lo, min(hi, raw_w + 2))
        ws.column_dimensions[_gcl(cn)].width = final_widths[cn]

    # ── Pass 3: compute per-row heights based on wrapping ────────────────────
    # Alignment wrapText was already set by _style_data; here we only adjust
    # row heights so wrapped text is fully visible.
    for row in ws.iter_rows(min_row=data_start_row):
        max_lines = 1
        for cell in row:
            if cell.value is None:
                continue
            field  = col_to_field.get(cell.column)
            col_w  = final_widths.get(cell.column, 10)
            disp   = _cell_display_len(cell.value)
            # Only count wrapping for text (non-numeric) cells
            if field not in _NO_WRAP_FIELDS and disp > col_w:
                lines = math.ceil(disp / max(col_w, 1))
                if lines > max_lines:
                    max_lines = lines

        row_height = max(_DATA_ROW_HEIGHT_MIN,
                         max_lines * _DATA_ROW_HEIGHT_BASE + 2)
        ws.row_dimensions[row[0].row].height = row_height


def _style_data(cell, value) -> None:
    cell.value = value
    # Numbers: centre-aligned, no wrap.
    # Text: left-aligned with wrap so long values are fully visible.
    if isinstance(value, (int, float)):
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrapText=False)
        cell.number_format = "#,##0"
    else:
        cell.alignment = Alignment(horizontal="left", vertical="center",
                                   wrapText=True, indent=1)
    cell.border = _thin()


def _style_total(cell, value) -> None:
    cell.value = value
    cell.fill  = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
    cell.font  = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _thin()
    if isinstance(value, (int, float)):
        cell.number_format = "#,##0"


def _replace_placeholders(ws, values: dict) -> None:
    """Substitute {{key}} in every string cell."""
    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            v = cell.value
            for key, val in values.items():
                v = v.replace(f"{{{{{key}}}}}", str(val or ""))
            cell.value = v


def _clear_data_area(ws, start_row: int) -> None:
    """Unmerge and clear all cells from start_row downward."""
    to_unmerge = [str(r) for r in ws.merged_cells.ranges if r.min_row >= start_row]
    for r in to_unmerge:
        ws.unmerge_cells(r)
    for row in ws.iter_rows(min_row=start_row):
        for cell in row:
            cell.value = None


def _clean_sheet_name(name: str) -> str:
    """Backward-compat shim — delegates to the shared helper.

    Original behaviour: strip Excel-illegal characters (``/ \\ [ ] * ? :``
    plus apostrophe — BUG-41 mitigation) and clip to 31 chars.
    """
    from ._excel_helpers import clean_sheet_name
    return clean_sheet_name(name, fallback="")


_PAREN_GAP_RE    = re.compile(r'\)\s*\(')
_PAREN_CHARS_RE  = re.compile(r'[()\[\]{}]')


def _strip_color_brackets(s: str) -> str:
    """Strip decorative wrapping brackets/parentheses from a colour string.

    Some Sky East order files store the colour name fully wrapped in
    parentheses, e.g. ``"(dark blue)"``, or two colours concatenated as
    ``"(dark blue)(white)"``. Left as-is these:
      • display with ugly brackets in the buy plan ("(Dark Blue)"), and
      • never exact-match a lookup key from 大货进度表 or the internal
        colour DB (both store plain names like "Dark Blue"), so the
        Chinese colour name / colour code silently come back empty even
        when the DB has a matching entry.

    Adjacent ")(" boundaries (concatenated multi-colour values) are joined
    with " / " first so a two-tone label stays readable —
    ``"(dark blue)(white)"`` → ``"dark blue / white"`` — before every
    remaining bracket character is dropped. Colours with no brackets at
    all (the common case) pass through unchanged.
    """
    if not s or ("(" not in s and "[" not in s and "{" not in s):
        return s
    s = _PAREN_GAP_RE.sub(') / (', s)
    s = _PAREN_CHARS_RE.sub('', s)
    return " ".join(s.split())


def _cn_color(cn_lookup: dict, brand: str, color_en: str) -> str:
    """Look up Chinese color — primary (brand-specific) key only, no fallback.

    The keys in ``cn_lookup`` are normalised to title-case English by
    :meth:`ColorTranslationStore.build_lookup_dict`, so we apply the same
    normalisation here before the lookup.

    No brand-agnostic fallback: returning a Chinese name from a different
    brand's mapping would be misleading when the exact brand entry is absent.
    """
    from po_extractor.store.color_translation_store import _normalize_color_name
    norm = _normalize_color_name(color_en)
    return cn_lookup.get((COMPANY_SKY_EAST, brand, norm), "")


# ── 主标颜色 (main-label colour) auto-derivation ─────────────────────────────
# Rule (matching-colour convention observed in the user's source data,
# locked in v1.58.0 after consistency analysis of 大货进度表--Angel 2026.xlsx):
#   • light body colour → white label  (白色)
#   • dark  body colour → black label  (黑色)
# These keyword sets are matched as case-insensitive substrings against the
# English body-colour text.  When a colour token from BOTH lists is found
# (e.g. "BLACK WITH CREAM PIPING"), the *first* matching token wins so the
# label colour reflects the dominant body colour.

_LIGHT_BODY_KEYWORDS = {
    # Whites and off-whites
    "white", "cream", "ivory", "beige", "ecru", "snow", "pearl", "champagne",
    "oatmeal", "eggshell", "natural", "nude", "bone",
    # Light yellows / pastels
    "lemon", "yellow", "mustard", "butter", "vanilla", "honey",
    # Pinks / corals
    "pink", "blush", "rose", "peach", "salmon", "apricot", "coral",
    "fushia", "fuchsia",
    # Light blues / greens / purples
    "sky", "powder", "aqua", "mint", "lavender", "lilac", "sage", "pista",
    # Light neutrals / metallics
    "silver", "tan", "stone", "khaki", "taupe", "sand",
    # Generic modifiers
    "light", "pale", "pastel", "soft", "baby",
}

_DARK_BODY_KEYWORDS = {
    # Blacks / browns
    "black", "noir", "onyx", "charcoal", "graphite", "anthracite", "pewter",
    "brown", "chocolate", "coffee", "espresso", "mocha", "mahogany", "cognac",
    "rust", "burnt", "camel",
    # Reds / wines
    "red", "wine", "burgundy", "bordeaux", "maroon", "crimson", "oxblood",
    "scarlet", "ruby", "garnet",
    # Blues / greens (generally darker than the listed light variants)
    "navy", "indigo", "midnight", "cobalt", "royal", "denim", "blue",
    "teal", "emerald", "forest", "hunter", "olive", "moss",
    # Purples
    "purple", "violet", "plum", "eggplant", "aubergine", "magenta",
    # Misc dark / greys / golds
    "dark", "deep", "rich", "gold",
    # Catchy garment descriptors
    "chocolate brown",
}


def derive_main_label_color(en_color: str | None) -> str:
    """Return ``白色`` or ``黑色`` based on whether *en_color* describes a
    light or dark body colour, or ``""`` when the colour can't be classified.

    Rule (matching-colour convention observed in production data):
      • Light body → 白色 label (white label on light fabric)
      • Dark  body → 黑色 label (black label on dark fabric)
      • When both light and dark tokens appear, the leftmost token wins
        (the first colour mentioned is usually the dominant one — e.g.
        "Navy with cream piping" → navy → 黑色 label).
    """
    if not en_color:
        return ""
    text = str(en_color).lower()

    first_light_pos = min(
        (text.find(k) for k in _LIGHT_BODY_KEYWORDS if k in text),
        default=-1,
    )
    first_dark_pos = min(
        (text.find(k) for k in _DARK_BODY_KEYWORDS if k in text),
        default=-1,
    )

    if first_dark_pos == -1 and first_light_pos == -1:
        return ""           # unknown colour — leave blank for manual entry
    if first_dark_pos == -1:
        return "白色"        # only light tokens found
    if first_light_pos == -1:
        return "黑色"        # only dark tokens found
    # Both present — leftmost wins
    return "黑色" if first_dark_pos < first_light_pos else "白色"


# ---------------------------------------------------------------------------
# Index sheet helper (VBA CreateIndexSheet)
# ---------------------------------------------------------------------------

def _create_index_sheet(wb, df_items, total_anchor: str = "Q5",
                        style_image_map: dict | None = None,
                        sheet_meta_list: list | None = None,
                        fabric_version_note: str | None = None) -> None:
    """Create Index sheet as the first sheet — mirrors VBA CreateIndexSheet.

    Parameters
    ----------
    wb               : target openpyxl Workbook
    df_items         : items DataFrame (must have 'style' column); used only
                       when *sheet_meta_list* is not provided.
    total_anchor     : cell address of the style-total formula (e.g. "Q5")
    style_image_map  : optional ``{style_name: [front_bytes, back_bytes|None]}``
                       — when provided, a "图片" column is inserted after "款号"
                       and each style row gets the front-view thumbnail.
    sheet_meta_list  : optional list of dicts, one per generated sheet::

                           {
                               "style":       str,   # style number
                               "sheet_name":  str,   # exact Excel sheet title
                               "brand":       str,
                               "body_part":   str,   # fabric body part label
                               "hhn_no":      str,   # HHN fabric code
                               "ex_fty_date": str,
                               "pc_no":       str,   # 客人PC NO — optional
                               "return_label": str,  # "Yes"/"No"/"NA" — optional,
                                                      # representative (first) value
                                                      # for this sheet's colours/POs
                           }

                       When supplied this is used directly (one Index row per
                       entry) instead of aggregating *df_items* by style.
                       Pass this when each style may have several sheets (one
                       per fabric part).
    """
    import io as _io
    from openpyxl.styles import Alignment, Font, PatternFill

    idx_ws = wb.create_sheet("Index", 0)

    has_images = bool(style_image_map)

    # ── Header row ────────────────────────────────────────────────────────────
    # Removed redundant "面料_大身" column — its value is always the body-part
    # label "Main Body / 大身" and the next column already says 面料_大身_编号
    # (i.e. the body-part identity is encoded in the header itself).
    _base_headers = [
        "No.", "款号", "客人PC NO", "客户品牌", "面料_大身_编号",
        "订单数合计", "离厂时间", "Return Label", "生产工厂", "工厂交期",
        "面料（计划）到厂时间", "辅料（计划）到厂时间", "样衣（计划）确认时间",
        "大货版（计划）完成时间", "全码版（计划）完成时间",
        "裁剪计划（计划）完成时间", "裁剪（计划）完成时间",
        "裁剪数", "车位（计划）完成时间", "后道（计划）完成时间", "出货数",
    ]
    if has_images:
        # Insert "图片" between "款号" and "客人PC NO"
        headers = _base_headers[:2] + ["图片"] + _base_headers[2:]
    else:
        headers = _base_headers

    for ci, h in enumerate(headers, 1):
        cell = idx_ws.cell(1, ci, value=h)
        cell.fill = PatternFill(start_color="FF000000", end_color="FF000000", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    idx_ws.row_dimensions[1].height = 36   # header row: tall enough for wrapped CJK text

    # ── Column index map (1-based) — shifts right by 1 when image col present ─
    # Layout (no image col / with image col):
    #   No.    | 款号   | [图片] | 客人PC NO | 客户品牌 | 面料_大身_编号 | 订单数合计 | 离厂时间
    #     1        2       (3)      3/4          4/5        5/6            6/7        7/8
    _off = 1 if has_images else 0
    _C_NO    = 1
    _C_STYLE = 2
    _C_IMG   = 3 if has_images else None
    _C_PCNO  = 3 + _off
    _C_BRAND = 4 + _off
    _C_FABNO = 5 + _off
    _C_QTY   = 6 + _off
    _C_EXFTY = 7 + _off
    _C_RETLBL = 8 + _off

    from openpyxl.utils import get_column_letter as _gcl
    _IMG_PX  = 160  # thumbnail size in pixels (larger = sharper rendering)
    _ROW_PT  = 122  # row height in points (~160 px × 0.75 pt-per-px + margin)

    # ── Build row data ────────────────────────────────────────────────────────
    if sheet_meta_list is not None:
        # One Index row per (style, fabric) sheet — caller pre-built this list.
        # Use display_key (综合标识Key) for the fabric column when available;
        # fall back to the bare HHN code so older callers still work.
        rows_iter = [
            (
                meta["style"],
                meta["sheet_name"],
                meta.get("brand",       ""),
                meta.get("body_part",   ""),
                meta.get("display_key") or meta.get("hhn_no", ""),
                meta.get("ex_fty_date", ""),
                meta.get("pc_no",       ""),
                meta.get("return_label", "NA"),
            )
            for meta in sheet_meta_list
        ]
    else:
        # Legacy path: one row per unique style, aggregated from df_items.
        _agg_kwargs = dict(
            brand          = ("brand",          "first"),
            fabrication    = ("fabrication",    "first"),
            fabric_item_no = ("fabric_item_no", "first"),
            ex_fty_date    = ("ex_fty_date",    "first"),
        )
        if "pc_no" in df_items.columns:
            _agg_kwargs["pc_no"] = ("pc_no", "first")
        if "return_label" in df_items.columns:
            _agg_kwargs["return_label"] = ("return_label", "first")
        agg = (
            df_items.groupby("style", sort=False)
            .agg(**_agg_kwargs)
            .reset_index()
        )
        rows_iter = [
            (
                str(row.style          or ""),
                _clean_sheet_name(str(row.style or "")),
                str(row.brand          or ""),
                str(row.fabrication    or ""),
                str(row.fabric_item_no or ""),
                str(row.ex_fty_date    or ""),
                str(getattr(row, "pc_no", "") or ""),
                str(getattr(row, "return_label", "") or "").strip() or "NA",
            )
            for row in agg.itertuples(index=False)
        ]

    for ri, (style_name, sheet_name, brand, body_part, fab_key, ex_fty_date, pc_no, return_label) in \
            enumerate(rows_iter, start=2):

        idx_ws.cell(ri, _C_NO).value = ri - 1
        cell_style = idx_ws.cell(ri, _C_STYLE, value=style_name)
        if sheet_name in wb.sheetnames:
            set_internal_hyperlink(cell_style, sheet_name)
            cell_style.style = "Hyperlink"
        idx_ws.cell(ri, _C_PCNO).value = pc_no

        # ── Style picture thumbnail (front image only for Index) ──────────
        if has_images and _C_IMG:
            _imgs = (style_image_map or {}).get(style_name) or []
            img_bytes = _imgs[0] if _imgs else None
            if img_bytes:
                try:
                    from openpyxl.drawing.image import Image as _XLImage
                    _prepped = _prep_image_for_embed(img_bytes, _IMG_PX)
                    xl_img = _XLImage(_io.BytesIO(_prepped))
                    xl_img.height = _IMG_PX
                    xl_img.width  = _IMG_PX
                    idx_ws.add_image(xl_img, f"{_gcl(_C_IMG)}{ri}")
                    idx_ws.row_dimensions[ri].height = _ROW_PT
                except Exception:
                    pass  # non-fatal — skip broken images silently

        idx_ws.cell(ri, _C_BRAND).value = brand
        idx_ws.cell(ri, _C_FABNO).value = fab_key
        if sheet_name in wb.sheetnames:
            idx_ws.cell(ri, _C_QTY).value = f"='{sheet_name}'!{total_anchor}"
        idx_ws.cell(ri, _C_EXFTY).value = ex_fty_date
        idx_ws.cell(ri, _C_RETLBL).value = return_label

    # ── Align data cells ─────────────────────────────────────────────────────
    # No., Qty, Return Label: centre; hyperlinked style: left; all others: left.
    for rn in range(2, idx_ws.max_row + 1):
        for cn in range(1, len(headers) + 1):
            c = idx_ws.cell(rn, cn)
            if cn == _C_NO or cn == _C_QTY or cn == _C_RETLBL:
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # ── Column widths ─────────────────────────────────────────────────────────
    # Explicit widths for the fixed columns; fall back to header-length for the
    # many schedule/planning columns that rarely have long data.
    _idx_fixed_widths: dict[int, float] = {
        _C_NO:    6,
        _C_STYLE: 20,   # 款号  style names can be 15-18 chars
        _C_PCNO:  14,   # 客人PC NO
        _C_BRAND: 16,   # 客户品牌
        _C_FABNO: 44,   # 面料_大身_编号  综合标识Key e.g. "HHN-JA-01715|100%POLYESTER|280|150"
        _C_QTY:   12,   # 订单数合计
        _C_EXFTY: 14,   # 离厂时间
        _C_RETLBL: 14,  # Return Label
    }
    if has_images and _C_IMG:
        _idx_fixed_widths[_C_IMG] = 26   # 图片 thumbnail — 160 px image ≈ 26 col units

    for ci, h in enumerate(headers, 1):
        if ci in _idx_fixed_widths:
            w = _idx_fixed_widths[ci]
        else:
            # CJK characters are roughly double-width visually
            h_str = str(h)
            cjk_count = sum(1 for ch in h_str if ord(ch) > 0x2E7F)
            w = max(14, len(h_str) + cjk_count + 4)
        idx_ws.column_dimensions[idx_ws.cell(1, ci).column_letter].width = w

    # ── Fabric-list version note (single cell, discreet footer) ─────────────
    # Written AFTER the alignment/width passes so it keeps its own subtle
    # styling. One blank row below the last data row, column B (under 款号).
    if fabric_version_note:
        note_cell = idx_ws.cell(idx_ws.max_row + 2, 2, value=fabric_version_note)
        note_cell.font = Font(size=9, italic=True, color="FF808080")
        note_cell.alignment = Alignment(horizontal="left", vertical="center")

    # Freeze the header row so it stays visible while scrolling
    idx_ws.freeze_panes = "A2"


def _create_overview_sheet(wb, overview_rows: list[dict], n_fabric_slots: int = 0) -> None:
    """Item-level cross-check sheet — one row per (style, PO, colour) item
    across the whole workbook, inserted right after the Index sheet.

    Mirrors the Contract History item-browser preview table (Style, Photo,
    Color, Brand, PO No., Config SKU, Article Name, Color Code, sizes,
    Ex-Fty, Fabric N / 综合标识 Key N) plus the Chinese colour name alongside
    the English one and the plain colour code, so every value written to a
    per-style sheet can be spot-checked against one flat table without
    hopping between tabs.

    Parameters
    ----------
    wb             : target openpyxl Workbook — the Index sheet must already
                     exist (this sheet is inserted at position 1, right after it).
    overview_rows  : one dict per item row::

                         {
                             "contract_no": str, "pc_no": str, "style": str,
                             "sheet_name": str,
                             "color_en": str, "color_cn": str, "color_code": str,
                             "client_po_color": str,
                             "progress_colors": list[str] | None,
                             "label_color": str,
                             "brand": str, "po": str, "config_sku": str,
                             "article_name": str,
                             "xs": int, "s": int, "m": int, "l": int,
                             "xl": int, "xxl": int, "total": int, "ex_fty": str,
                             "return_label": str,   # "Yes" / "No" / "NA"
                             "photo": bytes | None,
                             "fabrics": [(label, display_key), ...],
                         }

                     ``client_po_color`` is the colour text exactly as the
                     client's order file had it, before bracket-stripping or
                     multi-colour splitting. ``progress_colors`` is the list of
                     colour name(s) 大货进度表 actually has on file for that PC
                     No./Style (``None`` when the internal DB, not 大货进度表, is
                     the selected source). When ``color_cn``/``color_code`` is
                     :data:`_COLOR_NOT_FOUND`, both are combined into a cell
                     comment so a reviewer can spot a naming mismatch (client
                     says "Dark Blue", 大货进度表 says "Navy") without opening
                     the source order file.

    n_fabric_slots : number of Fabric / 综合标识 Key column pairs to render —
                     matches the template's fabric-header row count, so every
                     sheet's fabric combination fits regardless of how many
                     fabrics any single style/combo actually uses.
    """
    import io as _io
    from openpyxl.comments import Comment as _Comment
    from openpyxl.utils import get_column_letter as _gcl

    has_photos = any(r.get("photo") for r in overview_rows)

    _base_headers = [
        "No.", "Contract No.", "客人PC NO", "Style", "Color (EN)", "Color (CN)",
        "Color Code", "主标颜色",
        "Brand", "PO No.", "Config SKU", "Article Name",
        "XS", "S", "M", "L", "XL", "2XL", "Total Qty", "Ex-Fty", "Return Label",
    ]
    headers = (_base_headers[:3] + ["Photo"] + _base_headers[3:]
              if has_photos else list(_base_headers))
    for i in range(1, n_fabric_slots + 1):
        headers += [f"Fabric {i}", f"综合标识 Key {i}"]

    ov_ws = wb.create_sheet("Overview", 1)   # right after Index

    for ci, h in enumerate(headers, 1):
        cell = ov_ws.cell(1, ci, value=h)
        cell.fill = PatternFill(start_color="FF1F3864", end_color="FF1F3864", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ov_ws.row_dimensions[1].height = 30

    # ── Column index map (1-based) — shifts right by 1 when Photo is present ─
    _off        = 1 if has_photos else 0
    _C_NO       = 1
    _C_CONTRACT = 2
    _C_PCNO     = 3
    _C_IMG      = 4 if has_photos else None
    _C_STYLE    = 4 + _off
    _C_COLOR_EN = 5 + _off
    _C_COLOR_CN = 6 + _off
    _C_CODE     = 7 + _off
    _C_LABEL    = 8 + _off
    _C_BRAND    = 9 + _off
    _C_PO       = 10 + _off
    _C_CONFIG   = 11 + _off
    _C_ARTICLE  = 12 + _off
    _C_XS       = 13 + _off
    _C_S        = 14 + _off
    _C_M        = 15 + _off
    _C_L        = 16 + _off
    _C_XL       = 17 + _off
    _C_XXL      = 18 + _off
    _C_TOTAL    = 19 + _off
    _C_EXFTY    = 20 + _off
    _C_RETLBL   = 21 + _off
    _FAB_START  = 22 + _off

    _IMG_PX = 60   # small thumbnail — this sheet is one row per item, potentially
    _ROW_PT = 46   # far more rows than the style-level Index sheet

    for ri, row in enumerate(overview_rows, start=2):
        ov_ws.cell(ri, _C_NO,       value=ri - 1)
        ov_ws.cell(ri, _C_CONTRACT, value=row.get("contract_no", ""))
        ov_ws.cell(ri, _C_PCNO,     value=row.get("pc_no", ""))
        style_cell = ov_ws.cell(ri, _C_STYLE, value=row.get("style", ""))
        sheet_name = row.get("sheet_name", "")
        if sheet_name and sheet_name in wb.sheetnames:
            set_internal_hyperlink(style_cell, sheet_name)
            style_cell.style = "Hyperlink"
        ov_ws.cell(ri, _C_COLOR_EN, value=row.get("color_en", ""))
        color_cn_cell = ov_ws.cell(ri, _C_COLOR_CN, value=row.get("color_cn", ""))
        color_code_cell = ov_ws.cell(ri, _C_CODE, value=row.get("color_code", ""))
        if row.get("color_cn") == _COLOR_NOT_FOUND:
            _client_color = row.get("client_po_color") or row.get("color_en", "")
            _note_text = _color_miss_comment_text(_client_color, row.get("progress_colors"))
            color_cn_cell.comment   = _Comment(_note_text, "PO Extractor")
            color_code_cell.comment = _Comment(_note_text, "PO Extractor")
        ov_ws.cell(ri, _C_LABEL,    value=row.get("label_color", ""))
        ov_ws.cell(ri, _C_BRAND,    value=row.get("brand", ""))
        ov_ws.cell(ri, _C_PO,       value=row.get("po", ""))
        ov_ws.cell(ri, _C_CONFIG,   value=row.get("config_sku", ""))
        ov_ws.cell(ri, _C_ARTICLE,  value=row.get("article_name", ""))
        ov_ws.cell(ri, _C_XS,  value=row.get("xs", 0))
        ov_ws.cell(ri, _C_S,   value=row.get("s", 0))
        ov_ws.cell(ri, _C_M,   value=row.get("m", 0))
        ov_ws.cell(ri, _C_L,   value=row.get("l", 0))
        ov_ws.cell(ri, _C_XL,  value=row.get("xl", 0))
        ov_ws.cell(ri, _C_XXL, value=row.get("xxl", 0))
        ov_ws.cell(ri, _C_TOTAL, value=row.get("total", 0))
        ov_ws.cell(ri, _C_EXFTY, value=row.get("ex_fty", ""))
        ov_ws.cell(ri, _C_RETLBL, value=row.get("return_label", "NA"))

        if has_photos and _C_IMG:
            img_bytes = row.get("photo")
            if img_bytes:
                try:
                    from openpyxl.drawing.image import Image as _XLImage
                    _prepped = _prep_image_for_embed(img_bytes, _IMG_PX)
                    xl_img = _XLImage(_io.BytesIO(_prepped))
                    xl_img.height = _IMG_PX
                    xl_img.width  = _IMG_PX
                    ov_ws.add_image(xl_img, f"{_gcl(_C_IMG)}{ri}")
                except Exception:
                    pass  # non-fatal — skip broken images silently

        fabrics = row.get("fabrics") or []
        for i in range(n_fabric_slots):
            label, key = fabrics[i] if i < len(fabrics) else ("", "")
            ov_ws.cell(ri, _FAB_START + i * 2,     value=label)
            ov_ws.cell(ri, _FAB_START + i * 2 + 1, value=key)

        ov_ws.row_dimensions[ri].height = _ROW_PT if has_photos else 18

    # ── Alignment ─────────────────────────────────────────────────────────
    _center_cols = {_C_NO, _C_XS, _C_S, _C_M, _C_L, _C_XL, _C_XXL, _C_TOTAL, _C_RETLBL}
    for rn in range(2, ov_ws.max_row + 1):
        for cn in range(1, len(headers) + 1):
            c = ov_ws.cell(rn, cn)
            c.alignment = (
                Alignment(horizontal="center", vertical="center") if cn in _center_cols
                else Alignment(horizontal="left", vertical="center", indent=1)
            )

    # ── Column widths ─────────────────────────────────────────────────────
    _fixed_widths: dict[int, float] = {
        _C_NO: 6, _C_CONTRACT: 16, _C_PCNO: 14, _C_STYLE: 20, _C_COLOR_EN: 14, _C_COLOR_CN: 12,
        _C_CODE: 10, _C_LABEL: 10, _C_BRAND: 14, _C_PO: 16, _C_CONFIG: 16, _C_ARTICLE: 22,
        _C_XS: 6, _C_S: 6, _C_M: 6, _C_L: 6, _C_XL: 6, _C_XXL: 6,
        _C_TOTAL: 10, _C_EXFTY: 12, _C_RETLBL: 12,
    }
    if has_photos and _C_IMG:
        _fixed_widths[_C_IMG] = 10
    for i in range(n_fabric_slots):
        _fixed_widths[_FAB_START + i * 2]     = 24
        _fixed_widths[_FAB_START + i * 2 + 1] = 40

    for ci in range(1, len(headers) + 1):
        ov_ws.column_dimensions[_gcl(ci)].width = _fixed_widths.get(ci, 14)

    ov_ws.freeze_panes = "A2"


def _fabric_version_note(fabric_version_id: int | None) -> str | None:
    """One-line provenance note for the Index sheet: which fabric-list
    version enriched this buy plan. Deliberately terse -- buy plans are sent
    to clients/factories, so internal detail (uploader, internal filename,
    change counts) is omitted; the version id + date is enough for our own
    team to look up the full record in the Fabric DB tab's Version History.

    ``fabric_version_id=None`` means "Latest" (the live table / newest
    version); a specific id means the export was pinned to that archived
    version. Returns None (no note) when no version info is available --
    best-effort, never breaks the export.
    """
    try:
        from ..store import get_fabric_master_store
        versions = get_fabric_master_store().list_versions(limit=10)
        if fabric_version_id is None:
            meta = versions[0] if versions else None
            suffix = ""
        else:
            meta = next((v for v in versions
                         if v["version_id"] == fabric_version_id), None)
            suffix = " · 指定历史版本 pinned"
        if meta is None:
            return None
        when = str(meta["created_at"])[:10]
        return f"面料表版本 Fabric list version: v{meta['version_id']} ({when}){suffix}"
    except Exception:
        return None


def _prep_image_for_embed(img_bytes: bytes, display_px: int) -> bytes:
    """Resample *img_bytes* to fit within 2× *display_px* for Excel embedding.

    Uses thumbnail() for fast progressive downsampling of large images, then
    saves as JPEG (much faster encode/decode than PNG for photos).
    Falls back to the original bytes if Pillow is unavailable or decode fails.
    """
    target = display_px * 2          # 2× for HiDPI rendering
    try:
        from PIL import Image as _PILImage
        import io as _io
        img = _PILImage.open(_io.BytesIO(img_bytes))
        # Convert RGBA → RGB for JPEG output; keep RGB as-is
        if img.mode == "RGBA":
            bg = _PILImage.new("RGB", img.size, (255, 255, 255))
            bg.paste(img, mask=img.split()[3])
            img = bg
        elif img.mode != "RGB":
            img = img.convert("RGB")
        w, h = img.size
        if max(w, h) > target:
            # thumbnail() uses progressive halving — much faster than resize()
            # on large source images (15 MB PNGs, etc.)
            img.thumbnail((target, target), _PILImage.BILINEAR)
        elif max(w, h) < target:
            scale = target / max(w, h)
            new_size = (max(1, int(w * scale)), max(1, int(h * scale)))
            img = img.resize(new_size, _PILImage.BILINEAR)
        buf = _io.BytesIO()
        img.save(buf, format="JPEG", quality=85)
        return buf.getvalue()
    except Exception:
        return img_bytes              # PIL unavailable or broken image — use as-is


def _embed_style_photos(ws, front_bytes, back_bytes) -> None:
    """Embed front photo into J3:L6 and back photo into M3:O6 of *ws*.

    Uses ``twoCellAnchor editAs="twoCell"`` so the image fills (and is
    therefore centred in) the merged photo box exactly, regardless of
    image DPI / aspect ratio / column widths.

    ``front_bytes`` and ``back_bytes`` may be ``None``; missing photos
    are skipped silently and any old placeholder text is cleaned up.

    The legacy behaviour scanned for Photo1/Photo2 placeholder strings
    in the cell values and anchored a fixed-size image at that position.
    That left photos floating over arbitrary cells and never centred them
    inside the merged photo boxes.  This implementation always targets
    the canonical Sky_East.xlsx layout (J3:L6 / M3:O6).
    """
    import io as _io
    from openpyxl.drawing.image import Image as _XLImage
    from openpyxl.drawing.spreadsheet_drawing import (
        AnchorMarker as _AnchorMarker,
        TwoCellAnchor as _TwoCellAnchor,
    )

    # ── Clean up any legacy Photo1/Photo2 placeholder text first ──────────
    def _pnorm(v):
        return str(v).strip().lower().replace(" ", "").replace("_", "") \
                                    .replace("（", "(").replace("）", ")")
    _ALL_ALIASES = {"photo1", "{{photo1}}", "photo(front)", "photo(正面)",
                    "photo2", "{{photo2}}", "photo(back)",  "photo(背面)"}
    for row in ws.iter_rows(min_row=1, max_row=10, min_col=1, max_col=20):
        for cell in row:
            if isinstance(cell.value, str) and _pnorm(cell.value) in _ALL_ALIASES:
                cell.value = None

    # ── Target regions (1-based, inclusive) ───────────────────────────────
    # Front: J3:L6 → cols 10-12, rows 3-6
    # Back:  M3:O6 → cols 13-15, rows 3-6
    _SLOTS = (
        (front_bytes, 10, 3, 12, 6),
        (back_bytes,  13, 3, 15, 6),
    )

    for img_bytes, fc, fr, tc, tr in _SLOTS:
        if not img_bytes:
            continue
        try:
            _prepped = _prep_image_for_embed(img_bytes, 400)
            xl_img = _XLImage(_io.BytesIO(_prepped))
            # twoCellAnchor: image stretches to fill the cell range fc,fr → tc,tr
            xl_img.anchor = _TwoCellAnchor(
                editAs="twoCell",
                _from=_AnchorMarker(col=fc - 1, colOff=0, row=fr - 1, rowOff=0),
                to=_AnchorMarker(col=tc,        colOff=0, row=tr,    rowOff=0),
            )
            ws.add_image(xl_img)
        except Exception as _exc:
            # Loud failure so silent bugs don't bite us again
            import warnings as _w
            _w.warn(f"[sky_east photo embed] failed at "
                    f"({fc},{fr})-({tc},{tr}): {_exc!r}")
