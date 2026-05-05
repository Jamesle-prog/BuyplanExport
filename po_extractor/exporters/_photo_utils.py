"""Shared photo handling for buy-plan exporters.

Consolidates three previously duplicated concerns:

  • Photo region constants (J3:L6 front, M3:O6 back)
  • Photo embedding via openpyxl twoCellAnchor (used when building from
    scratch, e.g. hhp_buyplan_export)
  • Photo lookup helpers (disk-based, dict-based with fallbacks)

Region representation
---------------------
Internally everything is 1-based (matching openpyxl's worksheet API).
Two helpers are available for the post-save zip patcher in
``_image_inject.py`` which needs 0-based indices for OOXML drawing XML.
"""
from __future__ import annotations

import io
import os
from dataclasses import dataclass

from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor


# ---------------------------------------------------------------------------
# Region constants — single source of truth
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class PhotoRegion:
    """A rectangular cell region for a photo (1-based, inclusive)."""
    from_col: int
    from_row: int
    to_col:   int
    to_row:   int

    @property
    def merge_range(self) -> str:
        """Return Excel range like ``J3:L6`` for ``ws.merge_cells``."""
        from openpyxl.utils import get_column_letter
        return (
            f"{get_column_letter(self.from_col)}{self.from_row}"
            f":{get_column_letter(self.to_col)}{self.to_row}"
        )

    def to_zero_based(self) -> dict:
        """Return 0-based dict for the post-save XML patcher in
        ``_image_inject.py``.

        Note: in OOXML drawingML, the *to* marker is exclusive, so we add 1
        to to_col / to_row.
        """
        return {
            "from_col": self.from_col - 1,
            "from_row": self.from_row - 1,
            "to_col":   self.to_col,
            "to_row":   self.to_row,
        }


# Default merged-box positions used by the Zalando / HHP buy-plan layout
PHOTO_REGION_FRONT = PhotoRegion(from_col=10, from_row=3, to_col=12, to_row=6)
PHOTO_REGION_BACK  = PhotoRegion(from_col=13, from_row=3, to_col=15, to_row=6)


# ---------------------------------------------------------------------------
# Embedding (direct openpyxl — for exporters that build from scratch)
# ---------------------------------------------------------------------------

def embed_photo(ws, photo: bytes | str, region: PhotoRegion) -> bool:
    """Embed *photo* so it fills (and is therefore centred in) *region*.

    Uses ``twoCellAnchor editAs="twoCell"`` — the picture stretches to
    exactly match the merged box bounds, perfectly centred regardless of
    image DPI, aspect ratio, or column / row sizes.

    *photo* may be ``bytes`` or a file path.  Returns True on success,
    False (silently) if the photo is missing or invalid.
    """
    try:
        if isinstance(photo, (bytes, bytearray)):
            img = XLImage(io.BytesIO(photo))
        elif isinstance(photo, str) and os.path.isfile(photo):
            img = XLImage(photo)
        else:
            return False

        # AnchorMarker uses 0-based indices.  The "to" marker is the cell
        # *after* the region, so the image's bottom-right snaps to the
        # bottom-right of the last cell included.
        img.anchor = TwoCellAnchor(
            editAs="twoCell",
            _from=AnchorMarker(col=region.from_col - 1, colOff=0,
                               row=region.from_row - 1, rowOff=0),
            to=AnchorMarker(col=region.to_col,   colOff=0,
                            row=region.to_row,   rowOff=0),
        )
        ws.add_image(img)
        return True
    except Exception:
        return False


# ---------------------------------------------------------------------------
# Lookup helpers
# ---------------------------------------------------------------------------

# Recognised image extensions in priority order
_PHOTO_EXTS = (".png", ".jpg", ".jpeg")


def load_photo_from_disk(images_dir: str, style: str, side: str) -> bytes | None:
    """Look up ``{style}_{side}.{ext}`` on disk and return raw bytes.

    Tries .png / .jpg / .jpeg in that order.  Style names may contain
    Excel-illegal characters (``/``) — they are normalised the same way
    as sheet names (slash → underscore) for the file lookup.

    Returns None if no matching file is found / readable.
    """
    if not images_dir or not style or not side:
        return None

    # Normalise style name for filename use (slash etc. → underscore)
    safe_style = style
    for ch in r"/\[]*?:'":
        safe_style = safe_style.replace(ch, "_")

    for ext in _PHOTO_EXTS:
        path = os.path.join(images_dir, f"{safe_style}_{side}{ext}")
        if os.path.isfile(path):
            try:
                with open(path, "rb") as fh:
                    return fh.read()
            except Exception:
                pass
    return None


_PHOTO_EXTS_LOWER = (".png", ".jpg", ".jpeg", ".bmp", ".gif", ".tiff")

# Recognised "front / back" suffix variants (lowercase, with leading separator)
_FRONT_TOKENS = ("_front", "-front", " front", "_f", "-f", "_1", "-1")
_BACK_TOKENS  = ("_back",  "-back",  " back",  "_b", "-b", "_2", "-2")


def _safe_style(style: str) -> str:
    """Replace Excel-illegal chars in a style name for filename matching."""
    out = style
    for ch in r"/\[]*?:'":
        out = out.replace(ch, "_")
    return out


def _classify_filename(stem: str, safe_style_lc: str) -> str | None:
    """Return ``"front"``, ``"back"``, ``"single"``, or None if no match.

    *stem* is the filename without extension, lowercased.
    *safe_style_lc* is the sanitised style name, lowercased.
    """
    if not stem.startswith(safe_style_lc):
        # also allow safe_style appearing anywhere in stem (last-resort match)
        if safe_style_lc not in stem:
            return None
    suffix = stem[len(safe_style_lc):] if stem.startswith(safe_style_lc) else ""
    suffix_lc = suffix.lower()
    for tok in _BACK_TOKENS:
        if suffix_lc.startswith(tok):
            return "back"
    for tok in _FRONT_TOKENS:
        if suffix_lc.startswith(tok):
            return "front"
    if suffix_lc == "":          # exact match → treat as single (front)
        return "single"
    return None


def resolve_photo_pair(
    style: str,
    first_row,
    photo_map: dict,
) -> tuple[bytes | str | None, bytes | str | None]:
    """Resolve (front, back) photos for *style* using a multi-strategy lookup.

    Lookup order — first hit wins for each side:

      1. **Style-keyed** in *photo_map*
         ``photo_map[style]`` is bytes (front only) or a (front, back) tuple.

      2. **Filename pattern** in *photo_map* — handles the common case where
         ``photo_map`` is built by :func:`load_photo_map_from_dir`
         (keys are filenames like ``{style}_front.png``).
         Style names containing ``/`` are normalised to ``_`` first so a
         style ``ZLD060/S24DTR003`` matches the file ``ZLD060_S24DTR003_*``.

      3. **Photo1 / Photo2 columns** on *first_row* — legacy:
            • basename match against *photo_map*
            • style-prefix match against *photo_map* keys
            • absolute disk path

    Returned values are bytes, a path string, or None.
    """
    from ._excel_helpers import cell_value

    # ── 1. Direct style-keyed ──────────────────────────────────────────────
    entry = photo_map.get(style)
    if isinstance(entry, (list, tuple)) and len(entry) >= 2:
        return entry[0], entry[1]
    if isinstance(entry, (bytes, bytearray)):
        return entry, None

    # ── 2. Filename pattern in photo_map (folder-loaded photo_map) ────────
    safe = _safe_style(style)
    safe_lc = safe.lower()

    front_bytes: bytes | None = None
    back_bytes:  bytes | None = None

    # Walk photo_map once, classify each filename, take the first hit per side.
    for fname, blob in photo_map.items():
        if not isinstance(fname, str) or not isinstance(blob, (bytes, bytearray)):
            continue
        fname_lc = fname.lower()
        # Strip extension for classification
        stem_lc = fname_lc
        for ext in _PHOTO_EXTS_LOWER:
            if stem_lc.endswith(ext):
                stem_lc = stem_lc[: -len(ext)]
                break
        else:
            continue   # skip non-image filenames

        kind = _classify_filename(stem_lc, safe_lc)
        if kind == "back" and back_bytes is None:
            back_bytes = blob
        elif kind in ("front", "single") and front_bytes is None:
            front_bytes = blob

        if front_bytes is not None and back_bytes is not None:
            break

    if front_bytes is not None or back_bytes is not None:
        return front_bytes, back_bytes

    # ── 3. Photo1 / Photo2 column lookup (legacy) ─────────────────────────
    p1_path = cell_value(first_row, "Photo1")
    p2_path = cell_value(first_row, "Photo2")

    def _try(path_or_name: str | None):
        if not path_or_name:
            return None
        fname = os.path.basename(path_or_name)
        # 3a. Filename-keyed in photo_map
        if fname in photo_map:
            return photo_map[fname]
        # 3b. Style-prefix match (uses safe_style so slashes don't break it)
        prefix = safe[:8] if safe else ""
        if prefix:
            for k, v in photo_map.items():
                if isinstance(k, str) and k.startswith(prefix) \
                        and isinstance(v, (bytes, bytearray)):
                    return v
        # 3c. Disk path
        if os.path.isfile(path_or_name):
            return path_or_name
        return None

    return _try(p1_path), _try(p2_path)
