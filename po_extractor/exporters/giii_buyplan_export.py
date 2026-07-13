"""GIII buy-plan exporter — the A–W bilingual reference layout.

NOTE (v2.40+): no UI path uses this exporter any more — the buy plan the app
generates is :mod:`giii_production_plan` (the 生产计划单 format the factory
uses). This module is kept as the documented A–W reference implementation
(docs/GIII_BuyPlan_Field_Mapping.md) with its tests; retire it if the A–W
layout is never needed again.

Generates the A–W bilingual layout of the canonical GIII buy plan (see
docs/GIII_BuyPlan_Field_Mapping.md) from scratch, rather than filling a static
template — the reference file is a produced document, and generating it makes
every cell testable and the layout robust to size-count changes.

Sources per the spec:
* PO / 进度表 / color-translation → the caller assembles these into ``rows``.
* CPRS knowledge base → this exporter resolves the requirement-driven columns
  (red sticker P, carton mark Q, prepack ratio T, pcs/box U, MSRP V, RFID W)
  and the warehouse code E, via an optional :class:`CprsClient`. When no client
  is supplied (or a lookup fails) those cells are left blank — the buy plan
  still generates.
"""
from __future__ import annotations

import io
from dataclasses import dataclass, field

# Fixed left columns (before the dynamic size block).
_LEFT = [
    ("合同号", "Contract No."),   # A
    ("款号", "Style"),            # B
    ("PO号", "PO No."),           # C
    ("CPO#", "CPO#"),             # D
    ("仓库代码", "Warehouse"),    # E
    ("买家", "Buyer"),            # F
    ("颜色(英文)", "Color EN"),   # G
    ("颜色(中文)", "Color CN"),   # H
]

# Fixed right columns (after the dynamic size block).
_RIGHT = [
    ("总数量", "Total"),          # N-ish
    ("离厂时间", "X-FTY"),        # O
    ("红色箱贴纸", "Red Sticker"),  # P
    ("主箱唛", "Carton Mark"),    # Q
    ("包装方式", "Packing"),      # R
    ("是否预包", "Prepack"),      # S
    ("预包比例", "Prepack Ratio"),  # T
    ("每箱件数", "PCs/Box"),      # U
    ("MSRP", "MSRP"),             # V
    ("RFID", "RFID"),             # W
]

_NAVY = "FF1F3864"
_WHITE = "FFFFFFFF"
_GREY = "FFD9D9D9"
_YELLOW = "FFFFF2CC"


@dataclass
class BuyPlanRow:
    contract_no: str = ""
    style: str = ""
    po_number: str = ""
    cpo: str = ""
    ship_to: str = ""          # resolved to warehouse code via CPRS
    warehouse_code: str = ""   # if already known, skips CPRS resolution
    buyer: str = ""
    color_en: str = ""
    color_cn: str = ""
    sizes: dict = field(default_factory=dict)   # {"XS": 9, "S": 12, ...}
    ex_fty: str = ""
    packing_method: str = ""   # R — PO-sourced
    is_prepack: bool | None = None  # S — PO-sourced (PPK marker)

    @property
    def total(self) -> int:
        return sum(int(v or 0) for v in self.sizes.values())


@dataclass
class BuyPlanHeader:
    manufacturer: str = "江苏新万新服饰有限公司"
    supplier: str = ""
    fabric: str = ""
    description: str = ""
    brand: str = ""            # drives CPRS clientId
    date: str = ""
    updated: str = ""
    updated_2nd: str = ""


def _col(df, *names):
    """Return the first present column name from *names* (case-insensitive)."""
    lower = {c.lower(): c for c in df.columns}
    for n in names:
        if n.lower() in lower:
            return lower[n.lower()]
    return None


def assemble_buyplan_rows(df_size, df_meta, *, contract_by_po=None,
                          contract_by_style=None, color_lookup=None) -> list[BuyPlanRow]:
    """Assemble BuyPlanRows (one per PO × Color) from stored GIII frames.

    Pure — dataframes/dicts in, rows out (no store/CPRS). *df_size* has
    po_number/style/color/size/units; *df_meta* one row per PO. Contract maps
    (from the 大货进度表) and the EN→CN color lookup fill A and H.
    """
    from ..lookups.progress_lookup import _norm_key

    contract_by_po = contract_by_po or {}
    contract_by_style = contract_by_style or {}
    color_lookup = color_lookup or {}

    if df_size is None or df_size.empty:
        return []

    po_c   = _col(df_size, "po_number", "PO Number")
    style_c = _col(df_size, "style", "Style")
    color_c = _col(df_size, "color", "Color")
    size_c = _col(df_size, "size", "Size")
    units_c = _col(df_size, "units", "Units")

    # Metadata lookup by PO number.
    meta_by_po: dict[str, dict] = {}
    if df_meta is not None and not df_meta.empty:
        mpo = _col(df_meta, "po_number", "PO Number")
        for rec in df_meta.to_dict("records"):
            meta_by_po[str(rec.get(mpo, ""))] = rec

    def m(po, *names, default=""):
        rec = meta_by_po.get(str(po), {})
        for n in names:
            for k in rec:
                if k.lower() == n.lower() and rec[k] not in (None, ""):
                    return rec[k]
        return default

    rows: list[BuyPlanRow] = []
    grp_keys = [k for k in (po_c, color_c) if k]
    for (key, sub) in df_size.groupby(grp_keys, dropna=False):
        po = key[0] if isinstance(key, tuple) else key
        color = key[1] if isinstance(key, tuple) and len(key) > 1 else ""
        style = str(sub.iloc[0][style_c]) if style_c else ""
        sizes = {}
        for _, sr in sub.iterrows():
            sz = str(sr[size_c]).strip() if size_c else ""
            if sz:
                sizes[sz] = sizes.get(sz, 0) + int(sr[units_c] or 0) if units_c else 0

        contract = (contract_by_po.get(_norm_key(po))
                    or contract_by_style.get(_norm_key(style)) or "")
        color_en = str(color or "")
        packaging = str(m(po, "packaging"))
        rows.append(BuyPlanRow(
            contract_no=contract, style=style, po_number=str(po),
            cpo=str(m(po, "cpo")), ship_to=str(m(po, "ship_to")),
            buyer=str(m(po, "buyer", "customer")),
            color_en=color_en,
            color_cn=str(color_lookup.get(color_en, "")
                         or color_lookup.get(color_en.strip().upper(), "")),
            sizes=sizes, ex_fty=str(m(po, "xport_date", "factory_ship_date")),
            packing_method=packaging,
            is_prepack=("PPK" in packaging.upper() or "PREPACK" in packaging.upper())
                       if packaging else None,
        ))
    return rows


def _present_sizes(rows: list[BuyPlanRow]) -> list[str]:
    """Union of size keys across rows, in a stable garment order."""
    order = ["XXS", "XS", "S", "M", "L", "XL", "XXL", "1X", "2X", "3X",
             "OSFM", "OS"]
    seen = {s for r in rows for s in r.sizes}
    ordered = [s for s in order if s in seen]
    return ordered + [s for s in sorted(seen) if s not in order]


def _yn(v) -> str:
    return "" if v is None else ("Y" if v else "N")


def export_giii_buyplan(header: BuyPlanHeader, rows: list[BuyPlanRow],
                        cprs=None, manual: dict | None = None,
                        translate=None, requirements: dict | None = None) -> bytes:
    """Build the GIII buy-plan workbook and return the .xlsx bytes.

    Requirement resolution lives in
    :mod:`po_extractor.ui_helpers.giii_requirements` — pass its output as
    *requirements* ({id(row): RowRequirements}); or pass *cprs* (+ *manual* /
    *translate*) and it is resolved here for convenience.
    """
    from dataclasses import asdict
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from ..ui_helpers.giii_requirements import resolve_requirements

    sizes = _present_sizes(rows)
    n_left, n_size, n_right = len(_LEFT), len(sizes), len(_RIGHT)
    n_cols = n_left + n_size + n_right

    if requirements is None and cprs is not None:
        requirements, _ = resolve_requirements(cprs, header.brand, rows,
                                               manual=manual, translate=translate)
    cprs_fields = {rid: asdict(req) for rid, req in (requirements or {}).items()}

    wb = Workbook()
    ws = wb.active
    ws.title = (rows[0].style if rows else "BuyPlan")[:31] or "BuyPlan"

    thin = Side(style="thin", color="FF000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def cell(r, c, v, *, bold=False, bg=None, white=False, center=True, num=None):
        cl = ws.cell(r, c, v)
        cl.font = Font(name="Arial", size=10, bold=bold,
                       color=_WHITE if white else "FF000000")
        if bg:
            cl.fill = PatternFill("solid", fgColor=bg)
        cl.alignment = Alignment(horizontal="center" if center else "left",
                                 vertical="center", wrap_text=True)
        cl.border = border
        if num:
            cl.number_format = num
        return cl

    last_col = get_column_letter(n_cols)

    # ── Banner + title ───────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{last_col}1")
    cell(1, 1, header.manufacturer, bold=True, bg=_NAVY, white=True)
    ws.merge_cells(f"A2:{last_col}2")
    cell(2, 1, "生产计划单（buy plan）", bold=True, bg=_GREY)

    # ── Header meta (rows 4–6) ───────────────────────────────────────────────
    meta = [("供应商名称：", header.supplier, "日期：", header.date),
            ("面料/FIBER:", header.fabric, "更新日期：", header.updated),
            ("品名/Description:", header.description, "2ND更新日期：", header.updated_2nd)]
    for i, (la, va, lb, vb) in enumerate(meta):
        r = 4 + i
        cell(r, 1, la, bold=True, center=False)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=max(2, n_cols - 3))
        cell(r, 2, va, center=False)
        cell(r, n_cols - 1, lb, bold=True)
        cell(r, n_cols, vb)

    # ── Column headers (rows 8–9) ────────────────────────────────────────────
    hr1, hr2 = 8, 9
    ci = 1
    for zh, en in _LEFT:
        cell(hr1, ci, zh, bold=True, bg=_NAVY, white=True)
        cell(hr2, ci, en, bold=True, bg=_NAVY, white=True)
        ci += 1
    # size block: merged "尺码搭配" over row 8, size letters on row 9
    if n_size:
        ws.merge_cells(start_row=hr1, start_column=ci, end_row=hr1, end_column=ci + n_size - 1)
        cell(hr1, ci, "尺码搭配 / SIZE BREAK DOWN", bold=True, bg=_NAVY, white=True)
        for s in sizes:
            cell(hr2, ci, s, bold=True, bg=_NAVY, white=True)
            ci += 1
    for zh, en in _RIGHT:
        cell(hr1, ci, zh, bold=True, bg=_NAVY, white=True)
        cell(hr2, ci, en, bold=True, bg=_NAVY, white=True)
        ci += 1

    # ── Data rows ────────────────────────────────────────────────────────────
    def _embed(img_bytes: bytes | None, col: int, row_no: int) -> bool:
        """Anchor requirement artwork (red sticker / carton mark) into a cell,
        DISPIMG-style like the reference workbook. Never fails the export —
        a bad image just leaves the text value in place."""
        if not img_bytes:
            return False
        try:
            import io as _io
            from openpyxl.drawing.image import Image as XLImage
            img = XLImage(_io.BytesIO(img_bytes))
            scale = min(1.0, 76.0 / img.height if img.height else 1.0,
                        76.0 / img.width if img.width else 1.0)
            img.height, img.width = int(img.height * scale), int(img.width * scale)
            ws.add_image(img, f"{get_column_letter(col)}{row_no}")
            ws.row_dimensions[row_no].height = max(
                ws.row_dimensions[row_no].height or 0, 60)
            return True
        except Exception:
            return False

    red_col = n_left + n_size + 3   # 红色箱贴纸 (P-equivalent)
    mark_col = red_col + 1          # 主箱唛 (Q-equivalent)

    r = hr2 + 1
    for row in rows:
        cf = cprs_fields.get(id(row), {})
        left_vals = [row.contract_no, row.style, row.po_number, row.cpo,
                     cf.get("warehouse", row.warehouse_code), row.buyer,
                     row.color_en, row.color_cn]
        right_vals = [
            row.total, row.ex_fty,
            cf.get("red_sticker", ""), cf.get("carton_mark", ""),
            row.packing_method,
            _yn(row.is_prepack), cf.get("prepack_ratio", ""),
            cf.get("pcs_box", ""), cf.get("msrp", ""), cf.get("rfid", ""),
        ]
        ci = 1
        for v in left_vals:
            cell(r, ci, v, center=False); ci += 1
        for s in sizes:
            cell(r, ci, int(row.sizes.get(s, 0) or 0)); ci += 1
        for v in right_vals:
            cell(r, ci, v); ci += 1
        # Requirement artwork on top of (not instead of) the text values.
        _embed(cf.get("red_img"), red_col, r)
        _embed(cf.get("mark_img"), mark_col, r)
        r += 1

    # ── TTL + per-color subtotals ────────────────────────────────────────────
    total_col = n_left + n_size + 1   # 总数量 column index
    grand = sum(row.total for row in rows)
    cell(r, 1, "TTL", bold=True, bg=_YELLOW, center=False)
    cell(r, total_col, grand, bold=True, bg=_YELLOW)
    r += 1

    by_color: dict[str, dict] = {}
    for row in rows:
        key = row.color_cn or row.color_en or "—"
        acc = by_color.setdefault(key, {s: 0 for s in sizes})
        for s in sizes:
            acc[s] += int(row.sizes.get(s, 0) or 0)
    for color, acc in by_color.items():
        cell(r, n_left, color, bold=True, bg=_GREY, center=False)
        ci = n_left + 1
        for s in sizes:
            cell(r, ci, acc[s], bold=True, bg=_GREY); ci += 1
        cell(r, total_col, sum(acc.values()), bold=True, bg=_GREY)
        r += 1

    # ── Column widths + freeze ───────────────────────────────────────────────
    for c in range(1, n_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 11
    ws.column_dimensions["A"].width = 15
    ws.freeze_panes = "A10"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
