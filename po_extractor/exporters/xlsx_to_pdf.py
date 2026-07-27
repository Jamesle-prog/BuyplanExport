"""Render a worksheet to PDF with every column on one page.

Written for the cutting plan, whose sheets are far wider than they are tall:
the whole point of the printout is seeing every size column of every style
side by side, so the page must never split columns across sheets of paper.

Layout rules, matching Excel's "fit all columns on one page" print setting
(``_excel_helpers.apply_print_settings``):

  * **Columns always fit the page width.**  Column widths, row heights and
    the font are scaled by one factor, exactly as Excel's fit-to-width does,
    so proportions are preserved.
  * **Rows flow onto further pages** — height is not squeezed.
  * **Margins are minimal** (2 mm by default) so the widest possible sheet
    still gets a readable font.

Rendering goes through PyMuPDF, which is already a dependency, rather than
driving Excel over COM: the app runs headless on a server that may not have
Office installed, and COM automation from a Streamlit worker thread is
fragile.  The trade-off is that this is a focused renderer — values, fonts,
fills, borders, merges and alignment — not a full re-implementation of
Excel's display engine.
"""
from __future__ import annotations

import datetime as _dt
import io
import re
from typing import Any

import fitz
from openpyxl.utils import get_column_letter

# Page sizes in points (72 dpi), portrait orientation.
PAGE_SIZES: dict[str, tuple[float, float]] = {
    "A4":     (595.28, 841.89),
    "A3":     (841.89, 1190.55),
    "A5":     (419.53, 595.28),
    "Letter": (612.0, 792.0),
    "Legal":  (612.0, 1008.0),
}

# 2 mm — "minimal" edges: small enough to hand the width to the content,
# still inside the non-printable border of a typical office printer.
MINIMAL_MARGIN_PT = 5.67

# Excel's column-width unit is "characters of the default font"; the usual
# conversion is px = width * 7 + 5 at 96 dpi, and points = px * 72/96.
_PX_PER_CHAR = 7.0
_COL_PADDING_PX = 5.0
_PX_TO_PT = 0.75

_DEFAULT_COL_WIDTH = 8.43     # Excel's own default, in width units
_DEFAULT_ROW_HEIGHT = 15.0    # points

# Widest a single column may get from autofit (~55 characters).  Past this a
# lone long value would starve every other column of the page.
_MAX_AUTOFIT_COL_PT = (55.0 * _PX_PER_CHAR + _COL_PADDING_PX) * _PX_TO_PT

# Base-14 fonts cover Latin only; CJK needs one of MuPDF's built-in CJK
# fonts, which also carry Latin glyphs (so mixed strings render correctly).
_FONT_LATIN = "helv"
_FONT_LATIN_BOLD = "hebo"
_FONT_CJK = "china-s"

_CJK_RE = re.compile(
    r"[⺀-鿿豈-﫿＀-￯　-〿]"
)

_BLACK = (0.0, 0.0, 0.0)
_GRID = (0.62, 0.62, 0.62)


class PdfRenderError(RuntimeError):
    """Raised when a workbook can't be rendered at all."""


# ---------------------------------------------------------------------------
# Value formatting
# ---------------------------------------------------------------------------

def _decimals_from_format(fmt: str) -> int | None:
    """Decimal places implied by a number format, or None for General."""
    if not fmt or fmt.lower() in ("general", "@"):
        return None
    m = re.search(r"\.(0+)", fmt)
    if m:
        return len(m.group(1))
    if re.search(r"\b0\b|^0$|#,##0(?!\.)", fmt):
        return 0
    return None


def _format_value(value: Any, number_format: str = "General") -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, _dt.datetime):
        return value.strftime("%Y-%m-%d %H:%M" if (value.hour or value.minute)
                              else "%Y-%m-%d")
    if isinstance(value, _dt.date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, _dt.time):
        return value.strftime("%H:%M")

    fmt = number_format or "General"
    if isinstance(value, (int, float)):
        pct = "%" in fmt
        num = float(value) * (100.0 if pct else 1.0)
        dec = _decimals_from_format(fmt)
        if dec is None:
            # General: show enough precision to be useful, without a tail of
            # floating-point noise.
            text = f"{num:.6f}".rstrip("0").rstrip(".")
            if text in ("", "-"):
                text = "0"
        else:
            text = f"{num:,.{dec}f}" if "," in fmt else f"{num:.{dec}f}"
        return text + ("%" if pct else "")
    return str(value)


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

def _rgb(color) -> tuple[float, float, float] | None:
    """openpyxl colour → (r, g, b) floats, or None when not a plain RGB."""
    if color is None:
        return None
    rgb = getattr(color, "rgb", None)
    if not isinstance(rgb, str) or len(rgb) < 6:
        return None
    hexpart = rgb[-6:]
    try:
        r, g, b = (int(hexpart[i:i + 2], 16) / 255.0 for i in (0, 2, 4))
    except ValueError:
        return None
    return (r, g, b)


def _fill_color(cell) -> tuple[float, float, float] | None:
    fill = getattr(cell, "fill", None)
    if fill is None or getattr(fill, "patternType", None) != "solid":
        return None
    color = _rgb(getattr(fill, "fgColor", None))
    # A white fill is the default look; skipping it keeps the PDF smaller.
    if color is None or color == (1.0, 1.0, 1.0):
        return None
    return color


def _has_cjk(text: str) -> bool:
    return bool(_CJK_RE.search(text))


def _font_for(text: str, bold: bool) -> str:
    if _has_cjk(text):
        # MuPDF's built-in CJK fonts have no bold variant; the regular face
        # is used for both rather than dropping the glyphs.
        return _FONT_CJK
    return _FONT_LATIN_BOLD if bold else _FONT_LATIN


def _runs(text: str, bold: bool) -> list[tuple[str, str]]:
    """Split *text* into (run, fontname) pairs by script.

    A path like ``D:\\LJ\\zalando\\S24DTR003-裤子.PDS`` is mostly Latin with
    two Chinese characters.  Setting the whole string in the CJK face renders
    every Latin letter full-width — legible but stretched to twice the space
    it needs.  Splitting keeps Latin in Helvetica and hands only the CJK runs
    to the CJK face.
    """
    if not text:
        return []
    out: list[tuple[str, str]] = []
    buf, buf_cjk = [], _has_cjk(text[0])
    for ch in text:
        is_cjk = _has_cjk(ch)
        if is_cjk != buf_cjk and buf:
            out.append(("".join(buf), _font_for("中" if buf_cjk else "a", bold)))
            buf, buf_cjk = [], is_cjk
        buf.append(ch)
    if buf:
        out.append(("".join(buf), _font_for("中" if buf_cjk else "a", bold)))
    return out


def _run_width(run: str, fontname: str, fontsize: float) -> float:
    try:
        return fitz.get_text_length(run, fontname=fontname, fontsize=fontsize)
    except Exception:                                   # noqa: BLE001
        # Conservative estimate — CJK glyphs are full-width.
        return sum(fontsize * (1.0 if _has_cjk(ch) else 0.5) for ch in run)


def _measure(text: str, bold: bool, fontsize: float) -> float:
    return sum(_run_width(run, fn, fontsize) for run, fn in _runs(text, bold))


# ---------------------------------------------------------------------------
# Geometry
# ---------------------------------------------------------------------------

def _used_extent(ws) -> tuple[int, int]:
    """(last_row, last_col) that actually hold a value.

    ``ws.max_row`` / ``max_column`` count cells that were merely styled or
    touched, which on a generated sheet can stretch far past the data and
    would otherwise squeeze the scale factor to nothing.
    """
    last_row = last_col = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and str(cell.value).strip() != "":
                if cell.row > last_row:
                    last_row = cell.row
                if cell.column > last_col:
                    last_col = cell.column
    return last_row, last_col


def _column_widths(ws, last_col: int) -> list[float]:
    default = getattr(ws.sheet_format, "defaultColWidth", None) or _DEFAULT_COL_WIDTH
    widths = []
    for idx in range(1, last_col + 1):
        dim = ws.column_dimensions.get(get_column_letter(idx))
        width = getattr(dim, "width", None) if dim is not None else None
        if not width or width <= 0:
            width = default
        widths.append((width * _PX_PER_CHAR + _COL_PADDING_PX) * _PX_TO_PT)
    return widths


def _autofit_widths(ws, last_row: int, last_col: int, widths: list[float],
                    covered: set, anchors: dict, base_font_size: float,
                    ) -> list[float]:
    """Size each column to the width its own content needs.

    The stored column widths are deliberately *not* honoured here.  They are
    tuned for reading on screen, where a column can be narrower than its
    contents because Excel spills text over its neighbours and the user can
    always widen it — the Optitex export ships a 12-character column holding
    50-character marker paths.  Printing has no such escape: whatever doesn't
    fit is simply lost.  Measuring the content instead costs some scale
    (a smaller font) and buys a page with nothing truncated, which is the
    point of the printout.

    A single outlier value can't wreck the sheet: no column is allowed past
    ``_MAX_AUTOFIT_COL_PT``, and anything longer falls back to spilling and
    clipping as it would in Excel.

    Cells that anchor a merge across *columns* are skipped: their text spans
    several columns, so it must not force any single one wide.  A merge that
    only spans rows still lives in one column and is measured normally —
    missing that is how ``Material Cost,CNY`` ends up as ``Mat…``.
    """
    pad = 2.0 * _PX_PER_CHAR * _PX_TO_PT      # ~2 characters of breathing room
    floor = 2.0 * _PX_PER_CHAR * _PX_TO_PT
    needed = [0.0] * last_col
    for r in range(1, last_row + 1):
        for c in range(1, last_col + 1):
            if (r, c) in covered:
                continue
            if anchors.get((r, c), (1, 1))[1] > 1:
                continue
            cell = ws.cell(row=r, column=c)
            if cell.value is None:
                continue
            text = _format_value(cell.value, cell.number_format)
            if not text:
                continue
            width = _measure(text, bool(getattr(
                getattr(cell, "font", None), "bold", False)), base_font_size)
            if width > needed[c - 1]:
                needed[c - 1] = width
    return [
        min(_MAX_AUTOFIT_COL_PT, max(floor, need + pad)) if need else floor
        for need in needed
    ]


def _row_heights(ws, last_row: int) -> list[float]:
    default = (getattr(ws.sheet_format, "defaultRowHeight", None)
               or _DEFAULT_ROW_HEIGHT)
    heights = []
    for idx in range(1, last_row + 1):
        dim = ws.row_dimensions.get(idx)
        height = getattr(dim, "height", None) if dim is not None else None
        heights.append(float(height) if height else float(default))
    return heights


def _merge_map(ws, last_row: int, last_col: int
               ) -> tuple[dict[tuple[int, int], tuple[int, int]], set]:
    """(anchor → (rowspan, colspan), covered non-anchor cells)."""
    anchors: dict[tuple[int, int], tuple[int, int]] = {}
    covered: set[tuple[int, int]] = set()
    for rng in ws.merged_cells.ranges:
        if rng.min_row > last_row or rng.min_col > last_col:
            continue
        anchors[(rng.min_row, rng.min_col)] = (
            min(rng.max_row, last_row) - rng.min_row + 1,
            min(rng.max_col, last_col) - rng.min_col + 1,
        )
        for r in range(rng.min_row, min(rng.max_row, last_row) + 1):
            for c in range(rng.min_col, min(rng.max_col, last_col) + 1):
                if (r, c) != (rng.min_row, rng.min_col):
                    covered.add((r, c))
    return anchors, covered


def _paginate(heights: list[float], avail_h: float) -> list[tuple[int, int]]:
    """Split row indices (0-based) into pages that fit *avail_h*."""
    pages: list[tuple[int, int]] = []
    start = 0
    used = 0.0
    for i, h in enumerate(heights):
        if used + h > avail_h and i > start:
            pages.append((start, i))
            start, used = i, 0.0
        used += h
    if start < len(heights):
        pages.append((start, len(heights)))
    return pages or [(0, 0)]


# ---------------------------------------------------------------------------
# Rendering
# ---------------------------------------------------------------------------

def _draw_text(page, rect: fitz.Rect, text: str, *, bold: bool,
               align: str, fontsize: float, min_fontsize: float,
               color: tuple[float, float, float],
               overflow_to: float | None = None) -> None:
    """Draw one cell's text, clipped to *rect* (or to *overflow_to*).

    ``overflow_to`` is an absolute x the text may run to when the cells to
    the right are empty — Excel's spill behaviour, which is what keeps a
    marker's file path readable instead of clipped at the column edge.
    """
    if not text:
        return
    pad = min(2.0, rect.width * 0.08)
    limit = overflow_to if (overflow_to and align == "left") else rect.x1
    avail = max(limit - rect.x0 - 2 * pad, 1.0)

    size = fontsize
    width = _measure(text, bold, size)
    if width > avail:
        # Shrink this cell's text before truncating — losing a digit off a
        # quantity is worse than a slightly smaller font.
        size = max(min_fontsize, size * avail / width)
        width = _measure(text, bold, size)
    if width > avail and len(text) > 1:
        while len(text) > 1 and width > avail:
            text = text[:-1]
            width = _measure(text + "…", bold, size)
        text += "…"
        width = _measure(text, bold, size)

    if align == "center":
        x = rect.x0 + (rect.width - width) / 2
    elif align == "right":
        x = rect.x1 - pad - width
    else:
        x = rect.x0 + pad
    baseline = rect.y0 + (rect.height + size * 0.62) / 2

    for run, fontname in _runs(text, bold):
        try:
            page.insert_text((x, baseline), run, fontname=fontname,
                             fontsize=size, color=color)
        except Exception:                               # noqa: BLE001
            # A glyph this face can't encode — retry with the CJK font,
            # which covers the widest range, rather than dropping the run.
            try:
                page.insert_text((x, baseline), run, fontname=_FONT_CJK,
                                 fontsize=size, color=color)
            except Exception:                           # noqa: BLE001
                pass
        x += _run_width(run, fontname, size)


def _draw_borders(page, rect: fitz.Rect, border, scale: float) -> None:
    if border is None:
        return
    width = max(0.3, 0.5 * scale)
    for side, (p0, p1) in (
        ("left",   ((rect.x0, rect.y0), (rect.x0, rect.y1))),
        ("right",  ((rect.x1, rect.y0), (rect.x1, rect.y1))),
        ("top",    ((rect.x0, rect.y0), (rect.x1, rect.y0))),
        ("bottom", ((rect.x0, rect.y1), (rect.x1, rect.y1))),
    ):
        edge = getattr(border, side, None)
        if edge is None or not getattr(edge, "style", None):
            continue
        color = _rgb(getattr(edge, "color", None)) or _GRID
        page.draw_line(fitz.Point(*p0), fitz.Point(*p1),
                       color=color, width=width)


def _spill_limit(ws, row: int, last_used_col: int, last_col: int,
                 covered: set, x_at: list[float]) -> float | None:
    """How far right a cell's text may run before it must be clipped.

    Excel lets text spill over the cells to its right for as long as they are
    empty; the moment one holds a value the text is cut off.  Reproducing that
    is what keeps a marker's ``.mrk`` path readable on the printout instead of
    stopping at ``D:\\LJ\\zalando\\0…``.
    """
    col = last_used_col + 1
    while col <= last_col:
        if (row, col) in covered:
            break
        value = ws.cell(row=row, column=col).value
        if value is not None and str(value).strip() != "":
            break
        col += 1
    if col == last_used_col + 1:
        return None
    return x_at[min(col - 1, len(x_at) - 1)]


def _alignment(cell) -> str:
    horiz = getattr(getattr(cell, "alignment", None), "horizontal", None)
    if horiz in ("center", "centerContinuous"):
        return "center"
    if horiz == "right":
        return "right"
    if horiz in ("left", "justify", "distributed", "fill"):
        return "left"
    # Excel's implicit default: numbers right, everything else left.
    if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
        return "right"
    return "left"


def plan_layout(ws, *, avail_w: float, base_font_size: float = 9.0,
                min_font_size: float = 4.0, autofit: bool = True) -> dict:
    """Work out the page geometry for *ws* without drawing anything.

    Returns ``{last_row, last_col, widths, heights, scale, font_size,
    anchors, covered}`` with widths/heights already scaled.  Exposed so the
    fit can be asserted in tests and inspected without producing a PDF.
    """
    last_row, last_col = _used_extent(ws)
    if not last_row or not last_col:
        return {"last_row": 0, "last_col": 0, "widths": [], "heights": [],
                "scale": 1.0, "font_size": base_font_size,
                "anchors": {}, "covered": set()}

    anchors, covered = _merge_map(ws, last_row, last_col)
    widths = _column_widths(ws, last_col)
    if autofit:
        widths = _autofit_widths(ws, last_row, last_col, widths, covered,
                                 anchors, base_font_size)
    heights = _row_heights(ws, last_row)

    total_w = sum(widths) or 1.0
    # One scale for widths, heights and the font — Excel's fit-to-width
    # behaviour.  Never scale up: a narrow sheet keeps its natural size.
    scale = min(1.0, avail_w / total_w)
    return {
        "last_row": last_row, "last_col": last_col,
        "widths": [w * scale for w in widths],
        "heights": [h * scale for h in heights],
        "scale": scale,
        "font_size": max(min_font_size, base_font_size * scale),
        "anchors": anchors, "covered": covered,
    }


def _render_sheet(doc, ws, *, page_w: float, page_h: float, margin: float,
                  base_font_size: float, min_font_size: float,
                  autofit: bool = True) -> int:
    avail_w = page_w - 2 * margin
    avail_h = page_h - 2 * margin
    layout = plan_layout(ws, avail_w=avail_w, base_font_size=base_font_size,
                         min_font_size=min_font_size, autofit=autofit)
    last_row, last_col = layout["last_row"], layout["last_col"]
    if not last_row or not last_col:
        page = doc.new_page(width=page_w, height=page_h)
        _draw_text(page, fitz.Rect(margin, margin, page_w - margin, margin + 20),
                   f"{ws.title} — (empty)", bold=True, align="left",
                   fontsize=base_font_size, min_fontsize=min_font_size,
                   color=_BLACK)
        return 1

    widths, heights = layout["widths"], layout["heights"]
    scale, font_size = layout["scale"], layout["font_size"]
    anchors, covered = layout["anchors"], layout["covered"]

    x_at = [margin]
    for w in widths:
        x_at.append(x_at[-1] + w)

    pages = _paginate(heights, avail_h)

    for first, stop in pages:
        page = doc.new_page(width=page_w, height=page_h)
        y_at = [margin]
        for h in heights[first:stop]:
            y_at.append(y_at[-1] + h)

        def rect_for(row0: int, col0: int, rowspan: int = 1,
                     colspan: int = 1) -> fitz.Rect:
            y0 = y_at[row0 - first]
            y1 = y_at[min(row0 - first + rowspan, len(y_at) - 1)]
            return fitz.Rect(x_at[col0], y0,
                             x_at[min(col0 + colspan, len(x_at) - 1)], y1)

        # Fills first, then text and borders on top.
        for r in range(first, stop):
            for c in range(last_col):
                if (r + 1, c + 1) in covered:
                    continue
                cell = ws.cell(row=r + 1, column=c + 1)
                color = _fill_color(cell)
                if color is None:
                    continue
                span = anchors.get((r + 1, c + 1), (1, 1))
                page.draw_rect(rect_for(r, c, *span), color=None,
                               fill=color, width=0)

        for r in range(first, stop):
            for c in range(last_col):
                if (r + 1, c + 1) in covered:
                    continue
                cell = ws.cell(row=r + 1, column=c + 1)
                span = anchors.get((r + 1, c + 1), (1, 1))
                rect = rect_for(r, c, *span)
                _draw_borders(page, rect, getattr(cell, "border", None), scale)
                text = _format_value(cell.value, cell.number_format)
                if not text:
                    continue
                font = getattr(cell, "font", None)
                _draw_text(
                    page, rect, text,
                    bold=bool(getattr(font, "bold", False)),
                    align=_alignment(cell),
                    fontsize=font_size, min_fontsize=min_font_size,
                    color=_rgb(getattr(font, "color", None)) or _BLACK,
                    overflow_to=_spill_limit(ws, r + 1, c + 1 + span[1] - 1,
                                             last_col, covered, x_at),
                )
    return len(pages)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def workbook_to_pdf(source: Any, *, sheets: list[str] | None = None,
                    orientation: str = "landscape",
                    page_size: str = "A4",
                    margin_pt: float = MINIMAL_MARGIN_PT,
                    base_font_size: float = 9.0,
                    min_font_size: float = 4.0,
                    autofit: bool = True) -> bytes:
    """Render a workbook to PDF bytes, every column on one page.

    *source* is anything ``openpyxl.load_workbook`` accepts — a path or a
    file-like object.  *sheets* selects worksheets by name (all by default);
    each starts on a new page.  *margin_pt* is the edge on all four sides,
    defaulting to a minimal 2 mm.  *autofit* narrows columns to their content
    first (never past the sheet's own width) so the fit-to-width scale — and
    therefore the font — stays as large as possible.
    """
    from openpyxl import load_workbook

    try:
        wb = load_workbook(source, data_only=True)
    except Exception as exc:                            # noqa: BLE001
        raise PdfRenderError(f"Could not open the workbook: {exc}") from exc

    portrait = PAGE_SIZES.get(page_size) or PAGE_SIZES["A4"]
    page_w, page_h = (max(portrait), min(portrait)) \
        if orientation == "landscape" else (min(portrait), max(portrait))
    margin = max(0.0, float(margin_pt))

    doc = fitz.open()
    try:
        targets = [ws for ws in wb.worksheets
                   if sheets is None or ws.title in sheets]
        if not targets:
            raise PdfRenderError("No matching worksheet to render.")
        for ws in targets:
            _render_sheet(doc, ws, page_w=page_w, page_h=page_h,
                          margin=margin, base_font_size=base_font_size,
                          min_font_size=min_font_size, autofit=autofit)
        if doc.page_count == 0:
            raise PdfRenderError("Nothing was rendered.")
        return doc.tobytes()
    finally:
        doc.close()
        wb.close()


def xlsx_bytes_to_pdf(data: bytes, **kwargs) -> bytes:
    """Convenience wrapper: PDF bytes from in-memory .xlsx bytes."""
    return workbook_to_pdf(io.BytesIO(data), **kwargs)
