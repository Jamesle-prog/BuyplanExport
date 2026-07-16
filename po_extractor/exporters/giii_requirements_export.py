"""GIII PO requirements document — Excel export of CPRS requirement sets.

Generated automatically at upload time (when CPRS is configured): a Summary
sheet (one row per PO's order context, with per-status counts) plus one sheet
per PO listing every resolved requirement — domain, subtype, status, and the
requirement text extracted from the CPRS result.

Input is the ``contexts`` list from
:func:`po_extractor.ui_helpers.giii_requirements.resolve_po_requirements`.
"""
from __future__ import annotations

import io
import re

_NAVY = "FF1F3864"
_WHITE = "FFFFFFFF"
_GREY = "FFD9D9D9"
_YELLOW = "FFFFF2CC"
_RED = "FFFFC7CE"
_GREEN = "FFE2EFDA"

_STATUS_CN = {
    "confirmed": "必须 Required",
    "pending_input": "待定 Pending",
    "conflict": "冲突 Conflict",
    "not_applicable": "不适用 N/A",
    "missing_mandatory_context": "缺少信息 Missing context",
}

_STATUS_BG = {
    "confirmed": _GREEN,
    "pending_input": _YELLOW,
    "conflict": _RED,
}

_STATUS_RANK = {"confirmed": 0, "pending_input": 1, "conflict": 2,
                "missing_mandatory_context": 3, "not_applicable": 4}

# resultJson keys most likely to carry the human-readable spec, in order.
_TEXT_KEYS = ("current_wording", "standard", "description", "prompt",
              "applies_to", "reference", "change_summary")


def _req_text(result: dict) -> str:
    """Extract a readable requirement text from a CPRS result."""
    rj = result.get("resultJson") or {}
    parts = [str(rj[k]) for k in _TEXT_KEYS if rj.get(k)]
    if not parts:
        # compact leftover scalars (skip booleans/ids/nested structures)
        parts = [f"{k}: {v}" for k, v in rj.items()
                 if isinstance(v, (str, int, float)) and not isinstance(v, bool)
                 and k not in ("source", "scope_level", "updated")][:4]
    return " | ".join(parts)[:500]


def _sheet_name(base: str, used: set[str]) -> str:
    name = re.sub(r"[\\/*?:\[\]]", "_", base or "PO")[:28] or "PO"
    candidate, i = name, 2
    while candidate in used:
        candidate = f"{name}_{i}"
        i += 1
    used.add(candidate)
    return candidate


def _thumb(raw: bytes, max_h: int = 70):
    """Scale image bytes to a thumbnail; return (PNG BytesIO, w, h) or None on
    bad bytes (a broken image must never fail the document)."""
    try:
        from PIL import Image as PImage
        im = PImage.open(io.BytesIO(raw))
        im.load()
        if im.mode not in ("RGB", "RGBA"):
            im = im.convert("RGB")
        w, h = im.size
        if h > max_h:
            w = max(1, int(w * max_h / h))
            h = max_h
            im = im.resize((w, h))
        out = io.BytesIO()
        im.save(out, format="PNG")
        out.seek(0)
        return out, w, h
    except Exception:
        return None


def _embed_thumbs(ws, row: int, col: int, images, thumb_h: int = 70) -> None:
    """Embed each image side-by-side starting at (row, col); size the row to
    fit and widen the columns used. No-op when there are no images."""
    if not images:
        return
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter
    x, max_h = col, 0
    for raw in images:
        t = _thumb(raw, max_h=thumb_h)
        if not t:
            continue
        buf, w, h = t
        img = XLImage(buf)
        img.width, img.height = w, h
        letter = get_column_letter(x)
        ws.add_image(img, f"{letter}{row}")
        cur = ws.column_dimensions[letter].width or 0
        ws.column_dimensions[letter].width = max(cur, w / 7.0 + 2)
        x += 1
        max_h = max(max_h, h)
    if max_h:
        cur = ws.row_dimensions[row].height or 15
        ws.row_dimensions[row].height = max(cur, max_h * 0.78)


def export_giii_requirements(contexts: list[dict],
                             warnings: list[str] | None = None) -> bytes:
    """Build the requirements workbook; return .xlsx bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    thin = Side(style="thin", color="FF000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()

    def cell(ws, r, c, v, *, bold=False, bg=None, white=False, wrap=True,
             center=False):
        cl = ws.cell(r, c, v)
        cl.font = Font(name="Arial", size=10, bold=bold,
                       color=_WHITE if white else "FF000000")
        if bg:
            cl.fill = PatternFill("solid", fgColor=bg)
        cl.alignment = Alignment(horizontal="center" if center else "left",
                                 vertical="center", wrap_text=wrap)
        cl.border = border
        return cl

    # ── Summary sheet ────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary 汇总"
    headers = ["PO", "Style 款号", "Brand 品牌", "Warehouse 仓库",
               "Account 客户", "Channel 渠道",
               "必须 Required", "待定 Pending", "冲突 Conflict", "不适用 N/A"]
    for c, h in enumerate(headers, 1):
        cell(ws, 1, c, h, bold=True, bg=_NAVY, white=True, center=True)
    widths = [16, 14, 16, 11, 13, 12, 12, 12, 11, 11]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + c)].width = w

    for r, ctx in enumerate(contexts, 2):
        counts = {"confirmed": 0, "pending_input": 0, "conflict": 0,
                  "not_applicable": 0, "missing_mandatory_context": 0}
        for res in ctx["results"]:
            s = res.get("status", "")
            if s in counts:
                counts[s] += 1
        # missing_mandatory_context needs operator attention just like
        # pending_input — fold it into the 待定 column so a PO whose results
        # all need context doesn't show 0/0/0/0 (indistinguishable from
        # "no requirements").
        pending = counts["pending_input"] + counts["missing_mandatory_context"]
        vals = [ctx["po_number"], ctx["style"], ctx["brand"], ctx["warehouse"],
                ctx["account"], ctx["channel"], counts["confirmed"],
                pending, counts["conflict"], counts["not_applicable"]]
        for c, v in enumerate(vals, 1):
            cell(ws, r, c, v, center=(c >= 7),
                 bg=_RED if (c == 9 and v) else (_YELLOW if (c == 8 and v) else None))
    ws.freeze_panes = "A2"

    if warnings:
        wr = len(contexts) + 3
        cell(ws, wr, 1, "⚠ Warnings", bold=True, bg=_YELLOW)
        for i, w in enumerate(warnings, 1):
            cell(ws, wr + i, 1, w, wrap=True)
            ws.merge_cells(start_row=wr + i, start_column=1,
                           end_row=wr + i, end_column=len(headers))

    # ── By-style comparison sheet ────────────────────────────────────────────
    # Combines identical requirements (one row, "全部 All") and breaks out the
    # ones that differ per style, so you see at a glance where styles diverge.
    _write_by_style_sheet(wb, contexts, cell)

    # ── KL-style illustrated sheets (PO Index leads; then the illustrated /
    #    matrix / pre-pack / actions sheets) ───────────────────────────────────
    _write_po_index_sheet(wb, contexts, cell, index=0)
    _write_requirements_pictures_sheet(wb, contexts, cell, index=3)
    _write_requirement_matrix_sheet(wb, contexts, cell, index=4)
    _write_prepack_sheet(wb, contexts, cell, index=5)
    _write_actions_sheet(wb, contexts, warnings, cell, index=6)

    # ── One sheet per PO context ─────────────────────────────────────────────
    used: set[str] = {"Summary 汇总", "款号对比 By Style", "PO Index",
                      "Requirements + Pictures", "Requirement Matrix",
                      "Pre-pack", "Actions & Confirm"}
    for ctx in contexts:
        s = wb.create_sheet(_sheet_name(ctx["po_number"], used))
        cell(s, 1, 1, f"PO {ctx['po_number']} · {ctx['style']} · {ctx['brand']} · "
                      f"{ctx['warehouse']} {ctx['account']} ({ctx['channel']})",
             bold=True, bg=_NAVY, white=True)
        s.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)

        hdrs = ["Domain 类别", "Item 项目", "Status 状态", "Requirement 要求",
                "Source 来源", "图示 Image"]
        for c, h in enumerate(hdrs, 1):
            cell(s, 2, c, h, bold=True, bg=_NAVY, white=True, center=True)
        for c, w in zip(range(1, 7), (14, 22, 20, 70, 34, 16)):
            s.column_dimensions[chr(64 + c)].width = w

        # applicable first, N/A last; stable by domain within each group
        results = sorted(ctx["results"],
                         key=lambda x: (x.get("status") == "not_applicable",
                                        x.get("domain", ""), x.get("subtype", "")))
        r = 3
        for res in results:
            status = res.get("status", "")
            rj = res.get("resultJson") or {}
            cell(s, r, 1, res.get("domain", ""))
            cell(s, r, 2, res.get("subtype", ""))
            cell(s, r, 3, _STATUS_CN.get(status, status),
                 bg=_STATUS_BG.get(status), center=True)
            cell(s, r, 4, _req_text(res))
            cell(s, r, 5, str(rj.get("source", "")))
            cell(s, r, 6, "")                       # image cell (border); art below
            _embed_thumbs(s, r, 6, res.get("_images"))
            r += 1
        s.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_by_style_sheet(wb, contexts: list[dict], cell) -> None:
    """Second tab: requirements pivoted by style. Identical requirements across
    every style collapse to ONE row ("全部 All"); requirements that differ show
    one row per distinct value with the styles that carry it, highlighted."""
    from collections import OrderedDict

    styles_order: list[str] = []
    for ctx in contexts:
        st = ctx.get("style") or "—"
        if st not in styles_order:
            styles_order.append(st)
    all_styles = set(styles_order)
    n = len(all_styles)

    # (domain, subtype) -> {(status, value): set(styles)}
    groups: "OrderedDict[tuple, OrderedDict]" = OrderedDict()
    for ctx in contexts:
        st = ctx.get("style") or "—"
        for res in ctx.get("results", []):
            key = (res.get("domain", ""), res.get("subtype", ""))
            vk = (res.get("status", ""), _req_text(res))
            groups.setdefault(key, OrderedDict()).setdefault(vk, set()).add(st)

    s = wb.create_sheet("款号对比 By Style", 1)
    cell(s, 1, 1, f"款号对比 By-Style Requirements · {n} 款 styles: "
                  + "、".join(styles_order), bold=True, bg=_NAVY, white=True)
    s.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    for c, h in enumerate(["类别 Domain", "项目 Item", "状态 Status",
                           "要求 Requirement", "适用款号 Styles"], 1):
        cell(s, 2, c, h, bold=True, bg=_NAVY, white=True, center=True)
    for c, w in zip(range(1, 6), (14, 24, 20, 64, 30)):
        s.column_dimensions[chr(64 + c)].width = w

    def _gkey(k):
        # groups that are N/A for every style sink to the bottom
        na_only = all(status == "not_applicable" for status, _ in groups[k])
        return (na_only, k[0], k[1])

    r = 3
    for key in sorted(groups, key=_gkey):
        domain, subtype = key
        items = sorted(groups[key].items(),
                       key=lambda kv: (_STATUS_RANK.get(kv[0][0], 9), kv[0][1]))
        for (status, value), styles in items:
            common = styles == all_styles
            styles_txt = f"全部 All ({n})" if common else "、".join(sorted(styles))
            hl = None if common else _YELLOW      # highlight style-specific rows
            cell(s, r, 1, domain)
            cell(s, r, 2, subtype, bg=hl)
            cell(s, r, 3, _STATUS_CN.get(status, status),
                 bg=_STATUS_BG.get(status), center=True)
            cell(s, r, 4, value, bg=hl)
            cell(s, r, 5, styles_txt, bg=hl, center=common)
            r += 1
    s.freeze_panes = "A3"


# ===========================================================================
# KL-style illustrated sheets (PO Index · Requirements + Pictures ·
# Requirement Matrix · Pre-pack · Actions & Confirm), auto-filled from CPRS.
# ===========================================================================

_DOMAIN_TITLE = {
    "label": "🏷️ 标 (Labels)", "care_content": "🧵 成分/洗涤/COO (Care)",
    "hangtag": "🎫 吊牌 (Hangtags)", "packaging": "👜 包装 (Packaging)",
    "carton": "◆ 外箱 (Carton)", "hanger": "🪝 衣架 (Hanger)",
    "testing": "🧪 测试 (Testing)", "shipping": "🚚 物流 (Shipping)",
}

_MATRIX_SYM = {"confirmed": "✓", "pending_input": "✓*", "conflict": "⚡",
               "not_applicable": "—", "missing_mandatory_context": "ctx"}


def _u(contexts, key):
    """Ordered distinct non-empty string values of *key* across contexts."""
    seen: list[str] = []
    for c in contexts:
        v = str(c.get(key, "") or "").strip()
        if v and v not in seen:
            seen.append(v)
    return seen


def _dest_label(ctx) -> str:
    """Short account/destination label — the Requirement Matrix column key."""
    acct = str(ctx.get("account", "") or "").strip()
    region = str(ctx.get("region", "") or "").strip()
    base = acct or "STOCK"
    return f"{base} ({region})" if region else base


def _prepack_of(results):
    """(ratio, pcs_per_box) from the winning prepack/ratio requirement, if any."""
    ratio, pcs = "", ""
    for res in results or []:
        sub = str(res.get("subtype", "")).lower()
        if "prepack" in sub or "pre_pack" in sub or "ratio" in sub:
            rj = res.get("resultJson") or {}
            ratio = ratio or str(rj.get("ratio") or rj.get("alpha")
                                 or rj.get("numeric") or "")
            pcs = pcs or str(rj.get("pieces_per_bag") or rj.get("pcs_per_carton")
                            or rj.get("units_per_pack") or "")
    return ratio, pcs


def _write_po_index_sheet(wb, contexts, cell, index=0):
    s = wb.create_sheet("PO Index", index)
    hdrs = ["PO Number", "款号 Style", "品名 Article", "数量 Units",
            "客户 Account", "仓库 Whs", "目的地 Destination", "包装 Packing",
            "MSRP", "Source File"]
    n = len(hdrs)
    for c, w in enumerate((16, 14, 20, 9, 22, 8, 26, 22, 10, 20), 1):
        s.column_dimensions[chr(64 + c)].width = w
    brand = "、".join(_u(contexts, "brand")) or "GIII"
    cell(s, 1, 1, f"{brand} · PO Requirements — {len(contexts)} PO(s)",
         bold=True, bg=_NAVY, white=True, center=True)
    s.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    for c, h in enumerate(hdrs, 1):
        cell(s, 2, c, h, bold=True, bg=_NAVY, white=True, center=True)
    total, r = 0, 3
    for ctx in contexts:
        u = int(ctx.get("units", 0) or 0)
        total += u
        vals = [ctx.get("po_number", ""), ctx.get("style", ""),
                ctx.get("article", ""), u, ctx.get("account", "") or "—",
                ctx.get("warehouse", ""), ctx.get("destination", ""),
                ctx.get("packing", ""), ctx.get("msrp", "") or "—",
                ctx.get("source_file", "")]
        for c, v in enumerate(vals, 1):
            cell(s, r, c, v, center=(c in (4, 6, 9)))
        r += 1
    cell(s, r, 3, "TOTAL", bold=True, bg=_YELLOW, center=True)
    cell(s, r, 4, total, bold=True, bg=_YELLOW, center=True)
    cell(s, r, 5, f"{len(contexts)} POs", bold=True, bg=_YELLOW)
    s.freeze_panes = "A3"


def _write_requirements_pictures_sheet(wb, contexts, cell, index=3):
    from collections import OrderedDict
    n_groups = len({_dest_label(c) for c in contexts}) or 1
    subs: "OrderedDict[tuple, dict]" = OrderedDict()
    for ctx in contexts:
        lbl = _dest_label(ctx)
        for res in ctx.get("results", []):
            if res.get("status") == "not_applicable":
                continue
            key = (res.get("domain", ""), res.get("subtype", ""))
            e = subs.setdefault(key, {"text": "", "source": "",
                                      "applies": set(), "images": [], "seen": set()})
            e["applies"].add(lbl)
            if not e["text"]:
                e["text"] = _req_text(res)
            if not e["source"]:
                e["source"] = str((res.get("resultJson") or {}).get("source", ""))
            for b in (res.get("_images") or []):
                k = (len(b), b[:48])
                if k not in e["seen"]:
                    e["seen"].add(k)
                    e["images"].append(b)

    s = wb.create_sheet("Requirements + Pictures", index)
    s.column_dimensions["A"].width = 4
    for c in "BCDEFGHIJKLM":
        s.column_dimensions[c].width = 15
    cell(s, 1, 1, "PO Requirements — illustrated with the relevant CPRS pictures",
         bold=True, bg=_NAVY, white=True)
    s.merge_cells(start_row=1, start_column=1, end_row=1, end_column=13)
    cell(s, 2, 1, "Each domain: requirement spec · CPRS source · linked manual "
                  "picture(s). ⚠ Pictures illustrate only — verify against the "
                  "manual + PO before production.", bg=_GREY)
    s.merge_cells(start_row=2, start_column=1, end_row=2, end_column=13)

    r = 4
    for (dom, sub), e in subs.items():
        title = _DOMAIN_TITLE.get(dom, dom)
        cell(s, r, 1, f"{title}   ·   {dom}/{sub}", bold=True, bg=_NAVY, white=True)
        s.merge_cells(start_row=r, start_column=1, end_row=r, end_column=13)
        r += 1
        applies = ("所有 PO / All" if len(e["applies"]) >= n_groups
                   else "、".join(sorted(e["applies"])))
        cell(s, r, 1, f"适用 / Applies to: {applies}", bg=_GREEN)
        s.merge_cells(start_row=r, start_column=1, end_row=r, end_column=13)
        r += 1
        if e["text"]:
            cell(s, r, 1, e["text"], wrap=True)
            s.merge_cells(start_row=r, start_column=1, end_row=r, end_column=13)
            r += 1
        if e["source"]:
            cell(s, r, 1, f"📄 来源 / Source: {e['source']}", bg=_GREY)
            s.merge_cells(start_row=r, start_column=1, end_row=r, end_column=13)
            r += 1
        if e["images"]:
            _embed_thumbs(s, r, 1, e["images"], thumb_h=120)
            r += 7                       # leave room for the picture block
        r += 1                           # gap before the next domain
    s.freeze_panes = "A4"


def _write_requirement_matrix_sheet(wb, contexts, cell, index=4):
    from collections import OrderedDict
    groups: "OrderedDict[str, list]" = OrderedDict()
    for ctx in contexts:
        # setdefault(..., []).extend(...) — two PO contexts can share the same
        # destination label; setdefault(..., ctx["results"]) would only keep
        # the FIRST context's results and silently drop the rest.
        groups.setdefault(_dest_label(ctx), []).extend(ctx.get("results", []))
    labels = list(groups.keys())

    subs: "OrderedDict[tuple, dict]" = OrderedDict()
    for results in groups.values():
        for res in results:
            key = (res.get("domain", ""), res.get("subtype", ""))
            if key not in subs:
                subs[key] = {"tier": res.get("appliedPriorityTier", ""),
                             "title": _req_text(res)[:90]}

    s = wb.create_sheet("Requirement Matrix", index)
    ncol = 3 + len(labels)
    cell(s, 1, 1, "Requirement Matrix — what applies to each destination (CPRS)",
         bold=True, bg=_NAVY, white=True)
    s.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    hdr = ["Domain / Subtype", "Tier"] + labels + ["Requirement (winning)"]
    for c, h in enumerate(hdr, 1):
        cell(s, 2, c, h, bold=True, bg=_NAVY, white=True, center=(c > 1))
    s.column_dimensions["A"].width = 30
    s.column_dimensions["B"].width = 6
    for i in range(len(labels)):
        s.column_dimensions[chr(67 + i)].width = 13
    s.column_dimensions[chr(67 + len(labels))].width = 58

    r = 3
    for (dom, sub), meta in subs.items():
        cell(s, r, 1, f"{dom}/{sub}")
        cell(s, r, 2, meta["tier"], center=True)
        for i, lbl in enumerate(labels):
            res = next((x for x in groups[lbl]
                        if x.get("domain") == dom and x.get("subtype") == sub), None)
            st_ = res.get("status") if res else None
            bg = (_GREEN if st_ == "confirmed" else _RED if st_ == "conflict"
                  else _YELLOW if st_ == "pending_input" else None)
            cell(s, r, 3 + i, _MATRIX_SYM.get(st_, "—"), center=True, bg=bg)
        cell(s, r, 3 + len(labels), meta["title"])
        r += 1
    cell(s, r + 1, 1, "✓ applies · ✓* needs a PO input · ⚡ conflict · "
                      "— N/A or absent · ctx needs context.", bg=_GREY)
    s.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=ncol)
    s.freeze_panes = "C3"


def _write_prepack_sheet(wb, contexts, cell, index=5):
    s = wb.create_sheet("Pre-pack", index)
    for c, w in zip("ABCDE", (16, 16, 14, 22, 14)):
        s.column_dimensions[c].width = w
    cell(s, 1, 1, "Pre-pack / Packs-per-box", bold=True, bg=_NAVY, white=True)
    s.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    cell(s, 2, 1, "PPK code on the PO = pack quantity. Ratio / pcs-per-carton "
                  "from CPRS packaging/prepack + pre_pack_ratio.", bg=_GREY)
    s.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
    for c, h in enumerate(["PO", "款号 Style", "预包 Prepack?", "比例 Ratio",
                           "每箱件数 Pcs/box"], 1):
        cell(s, 4, c, h, bold=True, bg=_NAVY, white=True, center=(c > 2))
    r = 5
    for ctx in contexts:
        ratio, pcs = _prepack_of(ctx.get("results", []))
        ip = ctx.get("is_prepack")
        prepack = "Y" if (ip or ratio) else ("N" if ip is False else "—")
        cell(s, r, 1, ctx.get("po_number", ""))
        cell(s, r, 2, ctx.get("style", ""))
        cell(s, r, 3, prepack, center=True, bg=_GREEN if prepack == "Y" else None)
        cell(s, r, 4, ratio, center=True)
        cell(s, r, 5, pcs, center=True)
        r += 1
    cell(s, r + 1, 1, "Source: CPRS packaging/prepack + pre_pack_ratio + PO "
                      "\"Prepack QTY per Size\" pages.", bg=_GREY)
    s.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=5)
    s.freeze_panes = "A5"


def _write_actions_sheet(wb, contexts, warnings, cell, index=6):
    s = wb.create_sheet("Actions & Confirm", index)
    for c, w in zip("ABCD", (5, 22, 74, 12)):
        s.column_dimensions[c].width = w
    cell(s, 1, 1, "Notes, Conflicts & Confirmations", bold=True, bg=_NAVY, white=True)
    s.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    for c, h in enumerate(["#", "项目 Item", "详情 Detail", "状态 Status"], 1):
        cell(s, 2, c, h, bold=True, bg=_NAVY, white=True, center=(c in (1, 4)))
    r = [3]
    n = [1]

    def _row(item, detail, status, bg):
        cell(s, r[0], 1, n[0], center=True)
        cell(s, r[0], 2, item)
        cell(s, r[0], 3, detail)
        cell(s, r[0], 4, status, center=True, bg=bg)
        r[0] += 1
        n[0] += 1

    for ctx in contexts:
        for res in ctx.get("results", []):
            if res.get("status") == "conflict":
                _row("Conflict 冲突",
                     f"PO {ctx['po_number']} — {res.get('domain')}/{res.get('subtype')}: "
                     f"{_req_text(res)[:120]}", "Confirm", _RED)
    seen = set()
    for ctx in contexts:
        for res in ctx.get("results", []):
            if res.get("status") == "missing_mandatory_context":
                key = (res.get("domain"), res.get("subtype"))
                if key in seen:
                    continue
                seen.add(key)
                _row("Needs context 缺少信息",
                     f"{res.get('domain')}/{res.get('subtype')} — supply the "
                     f"runtime input (e.g. DIM code) then re-resolve.",
                     "Confirm", _YELLOW)
    for w in (warnings or []):
        _row("Note 提示", w, "Info", _GREY)
    if n[0] == 1:
        _row("OK", "No conflicts or missing-context items across the selected POs.",
             "Info", _GREEN)
    s.freeze_panes = "A3"
