"""GIII Production Plan (生产计划单 / buy plan) exporter.

Generates one Excel workbook with one sheet per style.  Each sheet follows
the format used in the GIII/HHP production plan template:

  Row 1  : Factory / manufacturer name (from seller field)
  Row 2  : 生产计划单（buy plan）
  Row 4  : 供应商名称 | 日期
  Row 5+ : 面料/FIBER (one row per 款式面料表格 part; extra parts as
           面料_其他1/2/3) | 更新日期
  Next   : 品名/Description | 2ND更新日期
  Header : Two-row merged header (合同号 款号 PO号 CPO# 仓库代码 买家
            颜色英 颜色中 [sizes…] 总数量 离厂时间 红色箱贴纸 主箱唛 备注)
  Data   : one row per (PO, color), with repeated-value columns merged
  Footer : 订单要求 TTL / 溢短装要求 / 包装 / 样衣 / 主箱唛

Enrichment (all optional — the plan still generates without them):
* 面料 rows from the 款式面料表格 (``store.load_fabric_parts_for_styles``).
* 合同号 from the 大货进度表 contract maps.
* 颜色(中文) from the 进度表 EN→CN lookup (falls back to the colour DB).
* 红色箱贴纸 / 主箱唛 text and artwork from CPRS requirement resolution
  (pass *requirements* = ``{po_number: RowRequirements}``).
"""
from __future__ import annotations

import datetime
import io
import math as _math
import re

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ._excel_helpers import clean_sheet_name
from ..ui_helpers.giii_requirements import (
    brand_from_po, pack_ratio, prepack_flag, strip_ratio,
)
from ._excel_helpers import neutralise_foreign_formulas
from po_extractor.exporters._excel_helpers import HOUSE_HIGHLIGHT_HEX, HOUSE_SECTION_HEX, house_fill

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
_THIN   = Side(style="thin")
_MEDIUM = Side(style="medium")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# A form, not a table: the light section band, with the highlight colour on
# the one column the factory must check (see _excel_helpers' palette).
_HDR_FILL    = house_fill(HOUSE_SECTION_HEX)
_YELLOW_FILL = house_fill(HOUSE_HIGHLIGHT_HEX)
_NO_FILL     = PatternFill("none")

_CENTER   = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT     = Alignment(horizontal="left",   vertical="center", wrap_text=True)

_FONT_TITLE    = Font(name="微软雅黑", bold=True, size=14)
_FONT_SUBTITLE = Font(name="微软雅黑", bold=True, size=12)
_FONT_BOLD     = Font(name="微软雅黑", bold=True, size=10)
_FONT_NORMAL   = Font(name="微软雅黑", size=10)
_FONT_HDR      = Font(name="微软雅黑", size=10)
_FONT_FLAG     = Font(name="微软雅黑", bold=True, size=10, color="FFCC0000")

# Standard size ordering for GIII (add / reorder as needed)
_SIZE_ORDER = [
    "XXS", "XS", "S", "M", "L", "XL", "XXL", "3XL", "4XL",
    "1X",  "2X",  "3X",  "4X",  "5X",
    "0",  "2",  "4",  "6",  "8",  "10", "12", "14", "16", "18", "20",
    "00", "24", "25", "26", "27", "28", "29", "30", "31",
    "32", "33", "34", "36", "38", "40",
    "ONE SIZE", "OS",
]
_SIZE_RANK = {s.upper(): i for i, s in enumerate(_SIZE_ORDER)}


def _sort_sizes(sizes: list[str]) -> list[str]:
    return sorted(sizes, key=lambda s: (_SIZE_RANK.get(s.strip().upper(), 999), s))


# ---------------------------------------------------------------------------
# Low-level cell helpers
# ---------------------------------------------------------------------------

def _cell(ws, row: int, col: int, value=None,
          font=None, fill=None, border=None, align=None, num_fmt=None):
    c = ws.cell(row, col)
    if value is not None:
        c.value = value
    if font   is not None: c.font      = font
    if fill   is not None: c.fill      = fill
    if border is not None: c.border    = border
    if align  is not None: c.alignment = align
    if num_fmt:             c.number_format = num_fmt
    return c


def _merge(ws, r1: int, c1: int, r2: int, c2: int,
           value=None, font=None, fill=None, border=None, align=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    top = ws.cell(r1, c1)
    if value  is not None: top.value     = value
    if font   is not None: top.font      = font
    if fill   is not None: top.fill      = fill
    if align  is not None: top.alignment = align
    # Apply border to every cell in the merged region
    if border is not None:
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                ws.cell(r, c).border = border
    return top


def _merge_or_set(ws, r1: int, c1: int, r2: int, c2: int, **kw):
    """Merge if r1 != r2 (or c1 != c2), else set single cell."""
    if r1 == r2 and c1 == c2:
        _cell(ws, r1, c1, **kw)
    else:
        _merge(ws, r1, c1, r2, c2, **kw)


def _safe(val) -> str:
    """Return clean string; treat NaN / None / 'nan' / 'None' as empty."""
    if val is None:
        return ""
    try:
        if _math.isnan(float(val)):
            return ""
    except (TypeError, ValueError):
        pass
    s = str(val).strip()
    return "" if s.lower() in ("nan", "none", "nat") else s


def _fmt_msrp(val) -> str:
    """PO MSRP is stored as a bare number ('59.00'); show it as a price ('$59.00').
    A value already carrying a currency mark or non-numeric text passes through."""
    s = _safe(val)
    if not s:
        return ""
    if s[0] in "$€£¥":
        return s
    try:
        float(s.replace(",", ""))
    except ValueError:
        return s
    return f"${s}"


# ── explanatory comments for blank requirement cells ──────────────────────────

def _req_empty_reason(field: str, has_brand: bool, has_req: bool) -> str:
    """Bilingual explanation for WHY a requirement cell is blank."""
    if not has_brand:
        return ("PO 无品牌 → 未解析客户要求。\n"
                "No brand on the PO — client requirements were not resolved.")
    if not has_req:
        return ("CPRS 未配置，或该品牌未在 CPRS 中匹配 → 未解析要求。\n"
                "CPRS not configured, or the brand was not found in CPRS.")
    return {
        "packaging": "PO 未提供包装方式（Infor Nexus PO 常不含包装文本）。\n"
                     "No packing method on the PO.",
        "hanger": "PO 未提供衣架信息。\nNo hanger info on the PO.",
        "ppk": "PO 未提供包装信息，CPRS 也无预包比例 → 无法判断是否预包。\n"
               "No packing info on the PO and no prepack ratio from CPRS.",
        "pcs": "CPRS 与 PO 均未提供每箱件数。\n"
               "No pieces-per-carton from CPRS or the PO.",
        "wtl": "CPRS 未提供箱重限制。\nNo carton weight limit from CPRS.",
        "rfid": "CPRS 未提供该仓库的 RFID 要求（仓库可能未匹配）。\n"
                "No RFID default from CPRS (the warehouse may be unresolved).",
        "red": "CPRS 未返回红色箱贴代码（可能非预包单或不适用）。\n"
               "No red-sticker code from CPRS (not prepack / not applicable).",
        "mark": "CPRS 未返回主箱唛要求。\nNo carton-marking value from CPRS.",
        "msrp": "CPRS 未提供该仓库的 MSRP 要求（仓库可能未匹配）。\n"
                "No MSRP default from CPRS (the warehouse may be unresolved).",
    }.get(field, "未找到数据。\nNo data found.")


def _annotate_empty(ws, row, col, value, field, has_brand, has_req) -> None:
    """Attach an explanatory comment to a blank requirement cell (no-op when the
    cell has a value)."""
    if str(value or "").strip():
        return
    from openpyxl.comments import Comment
    ws.cell(row, col).comment = Comment(
        _req_empty_reason(field, has_brand, has_req), "GIII 系统")


def _annotate_msrp(ws, row, col, msrp_txt, po_msrp, has_brand, has_req) -> None:
    """MSRP note: none when an actual price is shown; when the cell is only the
    required-flag 'Y' (no price found) or blank, explain why the actual price is
    missing."""
    from openpyxl.comments import Comment
    if po_msrp:                                # actual price shown → nothing to say
        return
    txt = str(msrp_txt or "").strip()
    if txt.upper() == "Y":
        ws.cell(row, col).comment = Comment(
            "需要 MSRP，但 PO 未印价格、CPRS 未提供价格 → 无法填入实际价格。\n"
            "MSRP is required, but no price is printed on the PO and CPRS "
            "provides none — the actual price could not be filled in.",
            "GIII 系统")
    elif txt == "":
        ws.cell(row, col).comment = Comment(
            _req_empty_reason("msrp", has_brand, has_req), "GIII 系统")
    # "N" → not required; no comment needed.


_US_STATES = frozenset(
    "AL AK AZ AR CA CO CT DE FL GA HI ID IL IN IA KS KY LA ME MD MA MI MN MS "
    "MO MT NE NV NH NJ NM NY NC ND OH OK OR PA RI SC SD TN TX UT VT VA WA WV "
    "WI WY DC PR".split())
_EU_NAMES = ("NETHERLANDS", "GERMANY", "FRANCE", "ITALY", "SPAIN", "POLAND",
             "BELGIUM", "AUSTRIA", "IRELAND", "PORTUGAL", "CZECH", "SWEDEN",
             "DENMARK", "FINLAND", "GREECE", "HUNGARY", "ROMANIA", "SLOVAKIA",
             "SLOVENIA", "CROATIA", "BULGARIA", "LUXEMBOURG", "ESTONIA",
             "LATVIA", "LITHUANIA")
# 2-letter EU codes that can't be confused with US states / compass points /
# common words (so no DE=Delaware, no SE=southeast, no AT/IT/ES/FR words).
_EU_CODES = frozenset("NL PL BE PT CZ SK SI HR BG LU EE LV LT DK HU RO GR".split())
# State codes that are also English words — matched only after a comma
# ("…, IN"), never as a bare trailing token ("HOLD IN TRANSIT" is not Indiana).
_US_STATES_WORDY = frozenset("IN OR ME HI OK LA DE".split())


def _dest_country(ship_to: str) -> str:
    """目的地国家 from the ship-to text — only unambiguous markers, else ''."""
    txt = str(ship_to or "").upper()
    if not txt.strip():
        return ""
    toks = re.sub(r"[^A-Z0-9]+", " ", txt).split()
    tokset = set(toks)
    if "USA" in tokset or "UNITED STATES" in txt:
        return "US"
    if "CANADA" in tokset:
        return "CA"
    if "AUSTRALIA" in tokset:
        return "AU"
    if "UNITED KINGDOM" in txt or "ENGLAND" in txt or {"UK", "GB"} & tokset:
        return "UK"
    if any(n in txt for n in _EU_NAMES) or (tokset & _EU_CODES):
        return "EU"
    # US state + ZIP (e.g. "PERRIS,CA 92571"), or a trailing state code.
    if re.search(r"\b[A-Z]{2}\s*,?\s+\d{5}(?:-\d{4})?\b", txt):
        return "US"
    if toks and toks[-1] in (_US_STATES - _US_STATES_WORDY):
        return "US"
    if re.search(r",\s*(" + "|".join(_US_STATES_WORDY) + r")\s*$", txt):
        return "US"
    return ""


def _dest_address(ship_to: str, buyer: str) -> str:
    """目的地 shows the ADDRESS. Ship-to text usually leads with the consignee
    name ('ROSS STORES / 3404 INDIAN AVE / …'), which duplicates the 买家
    column — strip that first segment when it matches the buyer."""
    s = str(ship_to or "").strip()
    b = str(buyer or "").strip()
    if not s or not b or "/" not in s:
        return s
    def _key(x: str) -> str:
        return "".join(ch for ch in x.upper() if ch.isalnum())
    parts = [p.strip() for p in s.split("/")]
    first, bn = _key(parts[0]), _key(b)
    # Drop the segment only when it says nothing beyond the buyer name
    # (equal, or an abbreviation of it). A segment with EXTRA info
    # ("ROSS STORES DC#4") is kept — stripping it would lose the DC.
    if first and bn and (first == bn or first in bn):
        return " / ".join(p for p in parts[1:] if p)
    return s


def _ex_factory_date(etd: str) -> str:
    """离厂时间 = ETD − 10 days. The PO's ship date is the vessel ETD; goods
    must leave the factory ~10 days earlier. Unparseable dates pass through
    unchanged rather than corrupting the cell.

    ASSUMPTION (user-confirmed 2026-07): both factory_ship_date and
    xport_date hold the ETD for every current GIII PO format. If a future
    format stores a true ex-factory date, exempt that field here or the
    date shifts −10 twice."""
    if not etd:
        return ""
    try:
        dt = pd.to_datetime(etd, errors="coerce")
        if pd.isna(dt):
            return etd
        return (dt - pd.Timedelta(days=10)).strftime("%m/%d/%Y")
    except Exception:
        return etd


def _fabric_line(p) -> str:
    """One 面料 row value from a 款式面料表格 part, e.g.
    '大身：HHN-DB-YS240782 86%Polyester 14%Spandex 200gsm 有效170cm'."""
    bits = [str(getattr(p, "hhn_no", "") or "").strip(),
            str(getattr(p, "composition", "") or "").strip()]
    if getattr(p, "weight_gsm", 0):
        bits.append(f"{p.weight_gsm}gsm")
    if getattr(p, "width_cm", 0):
        bits.append(f"有效{p.width_cm}cm")
    spec = " ".join(b for b in bits if b).strip()
    body = str(getattr(p, "body_part", "") or "").strip()
    if body and spec:
        return f"{body}：{spec}"
    return spec or body


def _embed_img(ws, img_bytes, col: int, row_no: int,
               max_h: int = 76, max_w: int = 76) -> bool:
    """Anchor CPRS requirement artwork (红色箱贴纸 / 主箱唛) into a cell.
    Never fails the export — bad image bytes just leave the text value."""
    if not img_bytes:
        return False
    try:
        from openpyxl.drawing.image import Image as XLImage
        img = XLImage(io.BytesIO(img_bytes))
        scale = min(1.0, max_h / img.height if img.height else 1.0,
                    max_w / img.width if img.width else 1.0)
        img.height, img.width = int(img.height * scale), int(img.width * scale)
        ws.add_image(img, f"{get_column_letter(col)}{row_no}")
        ws.row_dimensions[row_no].height = max(
            ws.row_dimensions[row_no].height or 0, img.height * 0.75 + 6)
        return True
    except Exception:
        return False


def _embed_style_photos(ws, front_bytes, back_bytes, col_start: int) -> None:
    """Put the style's front / back photo into the blank header band (row 3).

    Uses :func:`_embed_img`, which scales to fit while KEEPING the aspect
    ratio. The Sky East sheet stretches its photos to fill a fixed merged box
    (twoCellAnchor editAs="twoCell") because its template reserves a box of
    the right shape; this layout's free header region is wide and short, so
    the same stretch would squash a garment photo. _embed_img also grows row 3
    to the image height, so the band only opens up when a photo exists.

    Row 3 is the layout's blank spacer and the columns from the first size
    column rightwards carry nothing until the two-row table header, so the
    photos sit in genuinely free space -- they never cover the 面料 / 品名
    labels on the left or the 日期 fields on the right.
    """
    _embed_img(ws, front_bytes, col_start,     3, max_h=180, max_w=150)
    _embed_img(ws, back_bytes,  col_start + 4, 3, max_h=180, max_w=150)


# ---------------------------------------------------------------------------
# Main public API
# ---------------------------------------------------------------------------

def generate_giii_production_plan(
    selected_pos: list[str],
    store,
    cn_color_lookup: dict | None = None,
    *,
    color_lookup_en: dict | None = None,
    contract_by_po: dict | None = None,
    contract_by_style: dict | None = None,
    requirements: dict | None = None,
    consumption: dict | None = None,
    style_image_map: dict | None = None,
) -> bytes:
    """Return an .xlsx workbook (bytes) with one sheet per style.

    Parameters
    ----------
    selected_pos:
        PO numbers to include (must all exist in *store*).
    store:
        A ``POStore`` instance.
    cn_color_lookup:
        Optional mapping ``{(client, brand, norm_en_color): cn_color}``
        from the colour-translation DB.  Pass ``None`` to skip CN colour.
    color_lookup_en:
        Optional ``{EN colour (upper) → CN colour}`` from the 大货进度表 —
        tried before the colour-translation DB (进度表 is the primary source).
    contract_by_po / contract_by_style:
        合同号 maps from the 大货进度表 (``build_contract_maps`` output).
    requirements:
        Optional ``{po_number: RowRequirements}`` from CPRS resolution —
        fills 红色箱贴纸 / 主箱唛 (text + artwork) and the 仓库代码 when
        the PO carries no destination code.
    style_image_map:
        Optional ``{style: [front_bytes|None, back_bytes|None]}`` — the same
        shape ``ui.shared.load_style_photo_map`` returns and the Sky East buy
        plan already consumes. A style with no entry simply gets no photo.
    """
    if not selected_pos:
        return b""

    df_all   = store.list_pos()
    df_pos   = df_all[df_all["po_number"].isin(selected_pos)].copy()
    if df_pos.empty:
        return b""

    df_sizes = store.load_size_rows(selected_pos)
    if df_sizes.empty:
        return b""

    # Normalise size-row column names (load_size_rows uses title case)
    df_sizes = df_sizes.rename(columns={
        "PO Number": "po_number",
        "Style":     "style",
        "Color":     "color",
        "Size":      "size",
        "Units":     "units",
    })

    # Attempt to build CN colour lookup if not supplied
    if cn_color_lookup is None:
        try:
            from po_extractor.store.color_translation_store import ColorTranslationStore
            from po_extractor.config import DB_PATH
            cts = ColorTranslationStore(DB_PATH)
            cn_color_lookup = cts.build_lookup_dict()
        except Exception:
            cn_color_lookup = {}

    wb = Workbook()
    wb.remove(wb.active)

    # Group selected POs by style — preserves multi-style buy plans
    styles_order = df_pos["style"].dropna().unique().tolist()

    # 面料信息 from the 款式面料表格 (style-fabric mapping), one block per style.
    try:
        parts_by_style = store.load_fabric_parts_for_styles(
            [str(s) for s in styles_order]) or {}
    except Exception:
        parts_by_style = {}

    # 单耗/排版 consumption per style (operator-uploaded table). Reconcile kg↔cm
    # per style when the buy plan is built.
    if consumption is None:
        try:
            consumption = store.load_fabric_consumption(
                [str(s) for s in styles_order]) or {}
        except Exception:
            consumption = {}

    summaries: list[dict] = []
    for style in styles_order:
        style_pos = df_pos[df_pos["style"] == style]
        style_sizes = df_sizes[df_sizes["po_number"].isin(style_pos["po_number"])]
        if style_sizes.empty:
            continue
        summary = _write_style_sheet(
            wb, str(style), style_pos, style_sizes, cn_color_lookup,
            fabric_parts=parts_by_style.get(str(style), []),
            color_lookup_en=color_lookup_en,
            contract_by_po=contract_by_po, contract_by_style=contract_by_style,
            requirements=requirements,
            consumption=consumption.get(str(style)),
            style_photos=(style_image_map or {}).get(style) or [],
        )
        if summary:
            summaries.append(summary)

    if not wb.sheetnames:
        return b""

    if summaries:
        _write_summary_sheet(wb, summaries)          # index 0
        _write_simple_summary_sheet(wb, summaries)   # index 1
        _write_upc_summary_sheet(wb, summaries)      # index 2  (pivot)
        _write_upc_detail_sheet(wb, summaries)       # index 3  (per-size list)

    buf = io.BytesIO()
    # Any '=' text that arrived in the data becomes inert here; the
    # exporter's own =SUM()/='Sheet'! formulas are left alone.
    neutralise_foreign_formulas(wb)
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Per-style sheet writer
# ---------------------------------------------------------------------------

def _write_style_sheet(
    wb: Workbook,
    style: str,
    style_df: pd.DataFrame,
    sizes_df: pd.DataFrame,
    cn_color_lookup: dict,
    fabric_parts: list | None = None,
    color_lookup_en: dict | None = None,
    contract_by_po: dict | None = None,
    contract_by_style: dict | None = None,
    requirements: dict | None = None,
    consumption: dict | None = None,
    style_photos: list | None = None,
) -> dict | None:
    """Append one sheet for *style* to *wb*; return the sheet's summary record
    (one row of the workbook's Summary 汇总 table) or None when skipped."""

    # ── Layout constants ──────────────────────────────────────────────────────
    all_sizes = _sort_sizes(
        [str(s).strip() for s in sizes_df["size"].dropna().unique() if str(s).strip()]
    )
    if not all_sizes:
        return
    n_sizes = len(all_sizes)

    # Extra fabric rows shift everything below the 面料/FIBER row down.
    fabric_parts = [p for p in (fabric_parts or [])
                    if not getattr(p, "is_empty", lambda: False)()][:4]
    fab_extra = max(0, len(fabric_parts) - 1)
    R_SUPPLIER = 4
    R_FIBER    = 5                      # first fabric row
    R_DESC     = 6 + fab_extra
    R_HDR1     = 8 + fab_extra
    R_HDR2     = 9 + fab_extra

    requirements = requirements or {}

    def _req_for(po_number):
        return requirements.get(str(po_number).strip())

    # Column index map
    C_CONTRACT  = 1          # A  合同号
    C_STYLE     = 2          # B  款号
    C_BRAND     = 3          # C  品牌
    C_PO        = 4          # D  PO号
    C_CPO       = 5          # E  CPO#
    C_WH        = 6          # F  仓库代码
    C_DEST      = 7          # G  目的地
    C_CTRY      = 8          # H  目的地国家
    C_BUYER     = 9          # I  买家
    C_COLOR_EN  = 10         # J  颜色(英文)
    C_COLOR_CN  = 11         # K  颜色(中文)
    C_SZ_START  = 12         # L  first size
    C_SZ_END    = 11 + n_sizes
    C_QTY       = C_SZ_END + 1
    C_SHIP      = C_SZ_END + 2
    C_RED       = C_SZ_END + 3
    C_MARK      = C_SZ_END + 4
    C_PACK      = C_SZ_END + 5   # 包装方式 (packing only)
    C_HANG      = C_SZ_END + 6   # 衣架
    C_PPK       = C_SZ_END + 7   # 是否预包 (+ ratio when known)
    C_PCS       = C_SZ_END + 8   # 每箱件数
    C_WTL       = C_SZ_END + 9   # 箱重限制
    C_MSRP      = C_SZ_END + 10
    C_RFID      = C_SZ_END + 11
    C_NOTE      = C_SZ_END + 12
    N_COLS      = C_NOTE

    # Safe sheet title (max 31 chars, illegal chars sanitised, unique) —
    # a bare style[:31] crashed create_sheet on styles containing / \ * ? : [ ]
    base_title = clean_sheet_name(style)
    sheet_title = base_title
    suffix = 2
    while sheet_title in wb.sheetnames:
        # Reserve exactly enough room for "_{suffix}" so the title never
        # exceeds Excel's 31-char sheet-name limit — a fixed base_title[:28]
        # slice overflows once suffix reaches 2 digits ("_100" = 4 chars).
        tail = f"_{suffix}"
        sheet_title = f"{base_title[:31 - len(tail)]}{tail}"
        suffix += 1
    ws = wb.create_sheet(title=sheet_title)

    # ── Column widths ──────────────────────────────────────────────────────────
    widths = {
        C_CONTRACT: 14, C_STYLE: 12, C_BRAND: 13, C_PO: 16, C_CPO: 10,
        C_WH: 10, C_DEST: 24, C_CTRY: 10, C_BUYER: 14, C_COLOR_EN: 16,
        C_COLOR_CN: 14,
        C_QTY: 8, C_SHIP: 12, C_RED: 12, C_MARK: 16, C_PACK: 16,
        C_HANG: 16, C_PPK: 12, C_PCS: 9, C_WTL: 14, C_MSRP: 7, C_RFID: 7,
        C_NOTE: 12,
    }
    for i in range(C_SZ_START, C_SZ_END + 1):
        widths[i] = 7
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # ── Representative row for header info ─────────────────────────────────────
    rep       = style_df.iloc[0]
    seller    = _safe(rep.get("seller"))
    fabric    = _safe(rep.get("fabric"))
    desc      = _safe(rep.get("description_code")) or _safe(rep.get("style_description"))
    today_str = datetime.date.today().strftime("%Y/%m/%d")

    # ── Row heights ────────────────────────────────────────────────────────────
    fixed_heights = [(1, 26), (2, 22), (R_SUPPLIER, 18), (R_DESC, 18),
                     (R_HDR1, 22), (R_HDR2, 18)]
    fixed_heights += [(R_FIBER + i, 18) for i in range(max(1, len(fabric_parts)))]
    for r_h, height in fixed_heights:
        ws.row_dimensions[r_h].height = height

    # ─────────────────────────────────────────────────────────────────────────
    # Row 1 — company/factory name
    # ─────────────────────────────────────────────────────────────────────────
    factory_name = seller if seller else "江苏新万新服饰有限公司"
    _merge(ws, 1, 1, 1, N_COLS, factory_name,
           font=_FONT_TITLE, align=Alignment(horizontal="center", vertical="center"))

    # ─────────────────────────────────────────────────────────────────────────
    # Row 2 — title
    # ─────────────────────────────────────────────────────────────────────────
    _merge(ws, 2, 1, 2, N_COLS, "生产计划单（buy plan）",
           font=_FONT_SUBTITLE, align=Alignment(horizontal="center", vertical="center"))

    # Row 3 — blank spacer

    # ─────────────────────────────────────────────────────────────────────────
    # Row 4 — 供应商名称 / 日期
    # ─────────────────────────────────────────────────────────────────────────
    _cell(ws, R_SUPPLIER, C_CONTRACT, "供应商名称：", font=_FONT_BOLD, align=_LEFT)
    label_end = min(C_BUYER + 2, N_COLS - 3)
    _merge_or_set(ws, R_SUPPLIER, C_STYLE, R_SUPPLIER, label_end, value=seller,
                  font=_FONT_NORMAL, align=_LEFT)
    _cell(ws, R_SUPPLIER, N_COLS - 2, "日期：",   font=_FONT_BOLD,   align=_LEFT)
    _cell(ws, R_SUPPLIER, N_COLS - 1, today_str,  font=_FONT_NORMAL, align=_LEFT)

    # ─────────────────────────────────────────────────────────────────────────
    # 面料 rows — one per 款式面料表格 part (fallback: the PO's fabric field)
    # ─────────────────────────────────────────────────────────────────────────
    if fabric_parts:
        for i, part in enumerate(fabric_parts):
            rr = R_FIBER + i
            label = "面料/FIBER:" if i == 0 else f"面料_其他{i}:"
            _cell(ws, rr, C_CONTRACT, label, font=_FONT_BOLD, align=_LEFT)
            _merge_or_set(ws, rr, C_STYLE, rr, label_end, value=_fabric_line(part),
                          font=_FONT_NORMAL, align=_LEFT)
    else:
        _cell(ws, R_FIBER, C_CONTRACT, "面料/FIBER:", font=_FONT_BOLD, align=_LEFT)
        _merge_or_set(ws, R_FIBER, C_STYLE, R_FIBER, label_end, value=fabric,
                      font=_FONT_NORMAL, align=_LEFT)
    _cell(ws, R_FIBER, N_COLS - 2, "更新日期：", font=_FONT_BOLD, align=_LEFT)

    # ─────────────────────────────────────────────────────────────────────────
    # 品名/Description / 2ND更新日期
    # ─────────────────────────────────────────────────────────────────────────
    _cell(ws, R_DESC, C_CONTRACT, "品名/Description:", font=_FONT_BOLD, align=_LEFT)
    _merge_or_set(ws, R_DESC, C_PO, R_DESC, label_end, value=desc,
                  font=_FONT_NORMAL, align=_LEFT)
    _cell(ws, R_DESC, N_COLS - 2, "2ND更新日期：", font=_FONT_BOLD, align=_LEFT)

    # ── Style photo(s) — same source as the Sky East buy plan ────────────────
    _photos = list(style_photos or [])
    _embed_style_photos(ws,
                        _photos[0] if len(_photos) > 0 else None,
                        _photos[1] if len(_photos) > 1 else None,
                        C_SZ_START)

    # ─────────────────────────────────────────────────────────────────────────
    # Two-row header
    # ─────────────────────────────────────────────────────────────────────────
    fixed_headers = [
        (C_CONTRACT, "合同号",     _HDR_FILL),
        (C_STYLE,    "款号",       _HDR_FILL),
        (C_BRAND,    "品牌",       _HDR_FILL),
        (C_PO,       "PO号",       _HDR_FILL),
        (C_CPO,      "CPO#",       _HDR_FILL),
        (C_WH,       "仓库代码",   _YELLOW_FILL),
        (C_DEST,     "目的地",     _HDR_FILL),
        (C_CTRY,     "目的地国家", _HDR_FILL),
        (C_BUYER,    "买家",       _HDR_FILL),
        (C_COLOR_EN, "颜色(英文)", _HDR_FILL),
        (C_COLOR_CN, "颜色(中文)", _HDR_FILL),
        (C_QTY,      "总数量",     _HDR_FILL),
        (C_SHIP,     "离厂时间",   _HDR_FILL),
        (C_RED,      "红色箱贴纸", _YELLOW_FILL),
        (C_MARK,     "主箱唛",     _YELLOW_FILL),
        (C_PACK,     "包装方式",   _HDR_FILL),
        (C_HANG,     "衣架",       _HDR_FILL),
        (C_PPK,      "是否预包",   _HDR_FILL),
        (C_PCS,      "每箱件数",   _HDR_FILL),
        (C_WTL,      "箱重限制",   _HDR_FILL),
        (C_MSRP,     "MSRP",       _HDR_FILL),
        (C_RFID,     "RFID",       _HDR_FILL),
        (C_NOTE,     "备注",       _HDR_FILL),
    ]
    for col_idx, label, fill in fixed_headers:
        _merge(ws, R_HDR1, col_idx, R_HDR2, col_idx, label,
               font=_FONT_HDR, fill=fill, border=_BORDER, align=_CENTER)

    # Size group header + individual sizes
    _merge(ws, R_HDR1, C_SZ_START, R_HDR1, C_SZ_END, "尺码搭配",
           font=_FONT_HDR, fill=_HDR_FILL, border=_BORDER, align=_CENTER)
    for i, sz in enumerate(all_sizes):
        _cell(ws, R_HDR2, C_SZ_START + i, sz,
              font=_FONT_HDR, fill=_HDR_FILL, border=_BORDER, align=_CENTER)

    # ─────────────────────────────────────────────────────────────────────────
    # Data rows
    # ─────────────────────────────────────────────────────────────────────────
    DATA_START = R_HDR2 + 1
    current_row = DATA_START
    grand_total = 0

    # Build a flat list of row records first so we know merge ranges
    records: list[dict] = []

    for _, po_row in style_df.iterrows():
        po_number    = po_row["po_number"]
        # Brand strictly off the PO: its division field, else the documented
        # PO-number division prefix (CS/LS/DW). Unknown → flagged, not guessed.
        brand     = _safe(po_row.get("division_name")) or brand_from_po(po_number)
        cpo       = _safe(po_row.get("cpo"))
        wh_code   = _safe(po_row.get("destination_code"))
        if not wh_code:
            # CPRS resolved the warehouse from the ship-to address.
            wh_code = str(getattr(_req_for(po_number), "warehouse", "") or "")
        buyer     = _safe(po_row.get("customer")) or _safe(po_row.get("buyer"))
        # The stored date is the ETD — 离厂时间 shows ETD − 10 days.
        ship_date = _ex_factory_date(
            _safe(po_row.get("factory_ship_date")) or _safe(po_row.get("xport_date")))
        ship_raw  = _safe(po_row.get("ship_to"))
        ship_to   = _dest_address(ship_raw, buyer)
        # 目的地国家 — the CPRS warehouse region when resolved, else parsed
        # from unambiguous markers in the address itself.
        country   = str(getattr(_req_for(po_number), "region", "") or "") \
            or _dest_country(ship_raw)
        packaging = _safe(po_row.get("packaging"))
        hanger    = _safe(po_row.get("hanger"))
        # Actual MSRP printed on the PO (KL POs carry it) — this is the real
        # retail price, shown in the MSRP column ahead of CPRS's Y/N flag.
        po_msrp   = _fmt_msrp(po_row.get("msrp"))
        # A pack ratio printed in the packing/hanger text (e.g. "HANGER
        # (1-2-2-1)") IS the prepack ratio — it moves to 是否预包 and also
        # means the order is prepack even without a PPK marker.
        po_ratio   = pack_ratio(packaging, hanger)
        is_prepack = prepack_flag(packaging, hanger)
        packaging  = strip_ratio(packaging)
        hanger     = strip_ratio(hanger)

        po_sizes = sizes_df[sizes_df["po_number"] == po_number].copy()
        if po_sizes.empty:
            continue

        # NaN colours become "" instead of being dropped — dropna() silently
        # excluded those size rows from the sheet, 总数量 and the TTL footer.
        po_sizes["color"] = po_sizes["color"].fillna("")
        colors = po_sizes["color"].unique().tolist()
        po_start_row = current_row

        for color_en in colors:
            color_str = str(color_en or "").strip()
            cs = po_sizes[po_sizes["color"] == color_en]

            # 中文颜色 — the 大货进度表 lookup is the primary source,
            # the colour-translation DB the fallback.
            color_cn = ""
            if color_lookup_en and color_str:
                color_cn = str(color_lookup_en.get(color_str.upper(), "") or "")
            if not color_cn and cn_color_lookup and color_str:
                try:
                    from po_extractor.store.color_translation_store import _normalize_color_name
                    norm = _normalize_color_name(color_str)
                    color_cn = (cn_color_lookup.get(("GIII", "", norm)) or
                                cn_color_lookup.get(("GIII", "GIII", norm)) or "")
                    if not color_cn:
                        # Scan all brands for this client as fallback
                        for (cli, _br, nc), val in cn_color_lookup.items():
                            if cli == "GIII" and nc == norm and val:
                                color_cn = val
                                break
                except Exception:
                    pass

            # Pivot size quantities; keep each size's UPC for the UPC 汇总 sheet.
            size_qty: dict[str, int] = {}
            size_upc: dict[str, str] = {}
            for _, sr in cs.iterrows():
                sz  = str(sr["size"]).strip()
                qty = int(sr["units"]) if pd.notna(sr["units"]) else 0
                size_qty[sz] = size_qty.get(sz, 0) + qty
                u = _safe(sr.get("UPC"))
                if u and not size_upc.get(sz):
                    size_upc[sz] = u
            total_qty = sum(size_qty.values())
            grand_total += total_qty

            records.append({
                "row":          current_row,
                "po_start":     po_start_row,
                "po_number":       po_number,
                "brand":        brand,
                "cpo":          cpo,
                "wh_code":      wh_code,
                "buyer":        buyer,
                "ship_date":    ship_date,
                "ship_to":      ship_to,
                "country":      country,
                "packaging":    packaging,
                "hanger":       hanger,
                "po_ratio":     po_ratio,
                "is_prepack":   is_prepack,
                "color_en":     color_str,
                "color_cn":     color_cn,
                "size_qty":     size_qty,
                "size_upc":     size_upc,
                "total_qty":    total_qty,
                "po_msrp":      po_msrp,
            })
            current_row += 1

    style_end_row = current_row - 1

    # ── Write cell values ──────────────────────────────────────────────────────
    for rec in records:
        r = rec["row"]
        ws.row_dimensions[r].height = 18

        _cell(ws, r, C_COLOR_EN, rec["color_en"],  font=_FONT_NORMAL, border=_BORDER, align=_CENTER)
        _cell(ws, r, C_COLOR_CN, rec["color_cn"],  font=_FONT_NORMAL, border=_BORDER, align=_CENTER)
        _cell(ws, r, C_QTY,      rec["total_qty"], font=_FONT_NORMAL, border=_BORDER, align=_CENTER)

        for j, sz in enumerate(all_sizes):
            qty = rec["size_qty"].get(sz)
            _cell(ws, r, C_SZ_START + j, qty,
                  font=_FONT_NORMAL, border=_BORDER, align=_CENTER)

        # Cells that will be covered by merges below still need border set
        # (备注 stays empty — packing facts have their own columns).
        for col in (C_CONTRACT, C_STYLE, C_BRAND, C_PO, C_CPO, C_WH, C_DEST,
                    C_CTRY, C_BUYER, C_SHIP, C_RED, C_MARK, C_PACK, C_HANG,
                    C_PPK, C_PCS, C_WTL, C_MSRP, C_RFID, C_NOTE):
            ws.cell(r, col).border = _BORDER

    # ── PO-level merges (合同号 / PO号 / CPO# / 仓库代码 / 买家 / 离厂时间 /
    #    红色箱贴纸 / 主箱唛) — requirement cells vary per PO, not per style ────
    from ..lookups.progress_lookup import _norm_key

    contract_by_po = contract_by_po or {}
    contract_by_style = contract_by_style or {}

    po_groups: dict[str, list[dict]] = {}
    for rec in records:
        po_groups.setdefault(rec["po_number"], []).append(rec)

    sum_contracts: list[str] = []
    sum_reds: list[str] = []
    sum_marks: list[str] = []
    sum_packs: list[str] = []
    sum_hangers: list[str] = []
    sum_ppks: list[str] = []
    sum_pcs: list[str] = []
    sum_wtl: list[str] = []
    sum_msrp: list[str] = []
    sum_rfid: list[str] = []
    contract_map: dict[str, str] = {}   # po_number → 合同号, for the UPC 汇总 sheet

    for po_number, grp in po_groups.items():
        r0, r1 = grp[0]["row"], grp[-1]["row"]
        rep    = grp[0]
        # 合同号 from the 大货进度表 (by PO, falling back to by style).
        contract = (contract_by_po.get(_norm_key(str(po_number)))
                    or contract_by_style.get(_norm_key(style)) or "")
        contract_map[po_number] = contract
        # 红色箱贴纸 / 主箱唛 / MSRP / RFID / pack-out from CPRS resolution.
        # No resolution (no brand on the PO / CPRS off) → cells stay EMPTY,
        # never a claim like 无.
        req = _req_for(po_number)
        red_txt  = str(getattr(req, "red_sticker", "") or "") if req else ""
        mark_txt = str(getattr(req, "carton_mark", "") or "") if req else ""
        # Ratio printed on the PO wins; CPRS's per-account ratio is fallback.
        ratio    = rep["po_ratio"] or (str(getattr(req, "prepack_ratio", "") or "")
                                       if req else "")
        pcs_txt  = str(getattr(req, "pcs_box", "") or "") if req else ""
        wtl_txt  = str(getattr(req, "carton_weight", "") or "") if req else ""
        # The PO's actual MSRP value wins; CPRS's Y/N "MSRP required" flag is
        # only the fallback for POs that don't print a price.
        msrp_txt = rep["po_msrp"] or (str(getattr(req, "msrp", "") or "") if req else "")
        rfid_txt = str(getattr(req, "rfid", "") or "") if req else ""
        if rep["is_prepack"] is None:
            # PO packing silent → fall back to CPRS: a resolved prepack ratio
            # means the order IS prepack.
            ppk_txt = f"Y {ratio}".strip() if ratio else ""
        elif rep["is_prepack"]:
            ppk_txt = f"Y {ratio}".strip()
        else:
            ppk_txt = "N"
        for lst, v in ((sum_contracts, contract), (sum_reds, red_txt),
                       (sum_marks, mark_txt), (sum_packs, rep["packaging"]),
                       (sum_hangers, rep["hanger"]), (sum_ppks, ppk_txt),
                       (sum_pcs, pcs_txt), (sum_wtl, wtl_txt),
                       (sum_msrp, msrp_txt), (sum_rfid, rfid_txt)):
            lst.append(v)
        _merge_or_set(ws, r0, C_CONTRACT, r1, C_CONTRACT, value=contract, font=_FONT_NORMAL, border=_BORDER, align=_CENTER)
        # 品牌 — read off the PO (division field / PO-number prefix); a PO
        # without one is flagged here, never guessed.
        _merge_or_set(ws, r0, C_BRAND, r1, C_BRAND,
                      value=rep["brand"] or "⚠ 无品牌",
                      font=_FONT_NORMAL if rep["brand"] else _FONT_FLAG,
                      border=_BORDER, align=_CENTER)
        _merge_or_set(ws, r0, C_PO,    r1, C_PO,    value=rep["po_number"],   font=_FONT_NORMAL, border=_BORDER, align=_CENTER)
        _merge_or_set(ws, r0, C_CPO,   r1, C_CPO,   value=rep["cpo"],      font=_FONT_NORMAL, border=_BORDER, align=_CENTER)
        _merge_or_set(ws, r0, C_WH,    r1, C_WH,    value=rep["wh_code"],  font=_FONT_NORMAL, border=_BORDER, align=_CENTER)
        _merge_or_set(ws, r0, C_DEST,  r1, C_DEST,  value=rep["ship_to"],  font=_FONT_NORMAL, border=_BORDER, align=_LEFT)
        _merge_or_set(ws, r0, C_CTRY,  r1, C_CTRY,  value=rep["country"],  font=_FONT_NORMAL, border=_BORDER, align=_CENTER)
        _merge_or_set(ws, r0, C_BUYER, r1, C_BUYER, value=rep["buyer"],    font=_FONT_NORMAL, border=_BORDER, align=_CENTER)
        _merge_or_set(ws, r0, C_SHIP,  r1, C_SHIP,  value=rep["ship_date"],font=_FONT_NORMAL, border=_BORDER, align=_CENTER)
        _merge_or_set(ws, r0, C_PACK,  r1, C_PACK,  value=rep["packaging"],font=_FONT_NORMAL, border=_BORDER, align=_CENTER)
        _merge_or_set(ws, r0, C_HANG,  r1, C_HANG,  value=rep["hanger"],   font=_FONT_NORMAL, border=_BORDER, align=_CENTER)
        _merge_or_set(ws, r0, C_PPK,   r1, C_PPK,   value=ppk_txt,          font=_FONT_NORMAL, border=_BORDER, align=_CENTER)
        _merge_or_set(ws, r0, C_PCS,   r1, C_PCS,   value=pcs_txt,          font=_FONT_NORMAL, border=_BORDER, align=_CENTER)
        _merge_or_set(ws, r0, C_WTL,   r1, C_WTL,   value=wtl_txt,          font=_FONT_NORMAL, border=_BORDER, align=_CENTER)
        _merge_or_set(ws, r0, C_MSRP,  r1, C_MSRP,  value=msrp_txt,         font=_FONT_NORMAL, border=_BORDER, align=_CENTER)
        _merge_or_set(ws, r0, C_RFID,  r1, C_RFID,  value=rfid_txt,         font=_FONT_NORMAL, border=_BORDER, align=_CENTER)
        _merge_or_set(ws, r0, C_RED,   r1, C_RED,   value=red_txt,          font=_FONT_NORMAL, border=_BORDER, align=_CENTER)
        _merge_or_set(ws, r0, C_MARK,  r1, C_MARK,  value=mark_txt,         font=_FONT_NORMAL, border=_BORDER, align=_CENTER)
        if req is not None:
            # Requirement artwork on top of (not instead of) the text values.
            _embed_img(ws, getattr(req, "red_img", None), C_RED, r0)
            _embed_img(ws, getattr(req, "mark_img", None), C_MARK, r0)

        # Explain every blank requirement cell (hover the red triangle to see
        # WHY it's empty). MSRP additionally notes when it's required but no
        # actual price could be found.
        _has_brand = bool(rep["brand"])
        _has_req = req is not None
        for _col, _val, _fld in (
                (C_PACK, rep["packaging"], "packaging"),
                (C_HANG, rep["hanger"], "hanger"),
                (C_PPK, ppk_txt, "ppk"),
                (C_PCS, pcs_txt, "pcs"),
                (C_WTL, wtl_txt, "wtl"),
                (C_RFID, rfid_txt, "rfid"),
                (C_RED, red_txt, "red"),
                (C_MARK, mark_txt, "mark")):
            _annotate_empty(ws, r0, _col, _val, _fld, _has_brand, _has_req)
        _annotate_msrp(ws, r0, C_MSRP, msrp_txt, rep["po_msrp"], _has_brand, _has_req)

    # ── Style-level merge (款号) — one value per style ─────────────────────────
    if records:
        _merge_or_set(ws, DATA_START, C_STYLE, style_end_row, C_STYLE,
                      value=style, font=_FONT_NORMAL, border=_BORDER, align=_CENTER)

    # 单耗/排版: reconcile kg↔cm from the uploaded consumption table — the
    # values are shown on the Summary 汇总 sheet (one row per style), not here.
    from ..ui_helpers.fabric_consumption import reconcile_consumption
    _cons = consumption or {}
    _cons_kg, _cons_cm, _cons_warn = reconcile_consumption(
        _cons.get("cons_kg"), _cons.get("cons_cm"),
        _cons.get("width_cm"), _cons.get("gsm"))

    # ─────────────────────────────────────────────────────────────────────────
    # Footer rows
    # ─────────────────────────────────────────────────────────────────────────
    fr = current_row  # first footer row

    # TTL total
    _merge(ws, fr, C_CONTRACT, fr, C_SZ_END - 1, "订单要求：",
           font=_FONT_BOLD, align=_LEFT)
    _cell(ws, fr, C_SZ_END, "TTL", font=_FONT_BOLD, border=_BORDER, align=_CENTER)
    _cell(ws, fr, C_QTY, grand_total,  font=_FONT_BOLD, border=_BORDER, align=_CENTER)
    fr += 1

    for text in [
        "溢短装要求：",
        "包装：见如上说明",
        "样衣：上线后提供照片样S码1件  船样S码2件",
        "主箱唛：",
    ]:
        _merge(ws, fr, 1, fr, N_COLS, text, font=_FONT_NORMAL, align=_LEFT)
        ws.row_dimensions[fr].height = 16
        fr += 1

    # ── 图示 — full-size 红色箱贴纸 / 主箱唛 artwork from CPRS ─────────────────
    # The in-cell thumbnails above are reminders; the readable pictures go
    # here (deduped — POs sharing artwork are listed on one label).
    for kind, attr in (("红色箱贴纸图示", "red_img"), ("主箱唛图示", "mark_img")):
        by_img: dict[bytes, list[str]] = {}
        for po_number in po_groups:
            req = _req_for(po_number)
            img = getattr(req, attr, None) if req else None
            if img:
                by_img.setdefault(img, []).append(str(po_number))
        for img, po_list in by_img.items():
            pos_txt = "、".join(po_list[:4]) + ("…" if len(po_list) > 4 else "")
            _merge(ws, fr, 1, fr, N_COLS, f"{kind}（{pos_txt}）：",
                   font=_FONT_BOLD, align=_LEFT)
            ws.row_dimensions[fr].height = 16
            fr += 1
            if _embed_img(ws, img, 2, fr, max_h=240, max_w=560):
                fr += 1

    # ── Print settings ─────────────────────────────────────────────────────────
    ws.print_area = f"A1:{get_column_letter(N_COLS)}{fr - 1}"
    ws.page_setup.orientation   = "landscape"
    ws.page_setup.fitToWidth    = 1
    ws.page_setup.fitToHeight   = 0
    ws.page_setup.paperSize     = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left   = 0.25
    ws.page_margins.right  = 0.25
    ws.page_margins.top    = 0.5
    ws.page_margins.bottom = 0.5

    def _uniq(vals) -> list[str]:
        return [v for v in dict.fromkeys(str(x).strip() for x in vals) if v]

    # Per-style and per-colour size aggregates for the summary sheets.
    size_totals: dict[str, int] = {}
    color_sizes: dict[str, dict[str, int]] = {}
    color_cn_map: dict[str, str] = {}
    for rec in records:
        key = rec["color_en"] or "—"
        acc = color_sizes.setdefault(key, {})
        for sz, q in rec["size_qty"].items():
            q = int(q or 0)
            acc[sz] = acc.get(sz, 0) + q
            size_totals[sz] = size_totals.get(sz, 0) + q
        if rec["color_cn"] and not color_cn_map.get(key):
            color_cn_map[key] = rec["color_cn"]

    # UPC-level rows for the UPC 汇总 sheet — one per (PO, colour, size).
    upc_rows: list[dict] = []
    for rec in records:
        for sz in _sort_sizes(list(rec["size_qty"].keys())):
            upc_rows.append({
                "style":    style,
                "brand":    rec["brand"],
                "po":       rec["po_number"],
                "contract": contract_map.get(rec["po_number"], ""),
                "color_en": rec["color_en"],
                "color_cn": rec["color_cn"],
                "size":     sz,
                "upc":      rec["size_upc"].get(sz, ""),
                "units":    rec["size_qty"].get(sz, 0),
            })

    return {
        "sheet": sheet_title,
        "style": style,
        "upc_rows": upc_rows,
        "desc": desc,
        "brands": _uniq(r["brand"] for r in records),
        "no_brand": any(not r["brand"] for r in records),
        "fabric": _fabric_line(fabric_parts[0]) if fabric_parts else fabric,
        "contracts": _uniq(sum_contracts),
        "pos": list(po_groups.keys()),
        "colors": _uniq(r["color_en"] for r in records),
        "colors_cn": _uniq(r["color_cn"] for r in records),
        "color_cn_map": color_cn_map,
        "sizes": list(all_sizes),
        "size_totals": size_totals,
        "color_sizes": color_sizes,
        "total": grand_total,
        "ship_dates": _uniq(r["ship_date"] for r in records),
        "warehouses": _uniq(r["wh_code"] for r in records),
        "destinations": _uniq(r["ship_to"] for r in records),
        "countries": _uniq(r["country"] for r in records),
        "packings": _uniq(sum_packs),
        "hangers": _uniq(sum_hangers),
        "prepacks": _uniq(sum_ppks),
        "pcs_cartons": _uniq(sum_pcs),
        "weights": _uniq(sum_wtl),
        "msrps": _uniq(sum_msrp),
        "rfids": _uniq(sum_rfid),
        "buyers": _uniq(r["buyer"] for r in records),
        "reds": _uniq(sum_reds),
        "marks": _uniq(sum_marks),
        # 单耗/排版 (per style) → Summary 汇总 columns
        "cons_kg": _cons_kg, "cons_cm": _cons_cm,
        "cons_util": _cons.get("util"), "cons_marker_pcs": _cons.get("marker_pcs"),
        "cons_width": _cons.get("width_cm"), "cons_gsm": _cons.get("gsm"),
    }


# ---------------------------------------------------------------------------
# Summary 汇总 sheet (first sheet — one row per style's buy plan)
# ---------------------------------------------------------------------------

_SUM_COLS = [
    ("No.",         6), ("款号",   14), ("品牌",   12), ("品名",   24),
    ("合同号",     14), ("PO数",    7), ("PO号",   24), ("颜色(英文)", 18),
    ("颜色(中文)", 14), ("尺码明细", 26), ("总数量", 10), ("离厂时间", 12),
    ("仓库代码",   10), ("目的地", 24), ("目的地国家", 11), ("买家", 14),
    ("包装方式",   16), ("衣架",   14), ("是否预包", 10), ("每箱件数", 9),
    ("箱重限制",   14), ("MSRP", 7), ("RFID", 7), ("红色箱贴纸", 12),
    ("主箱唛",     16),
    # 单耗/排版 (per style)
    ("单耗(kg)",   10), ("单耗(cm)", 10), ("排版利用率", 10), ("排版件数", 9),
    ("排版有效门幅(cm)", 14), ("排版面料克重(g/m²)", 16),
]
_SUM_TOTAL_COL = 11   # 总数量 position (for the TTL row)


def _size_breakdown_text(summary: dict) -> str:
    """'S 1400 / M 2800 / L 2800 / XL 1400' in the style's size order."""
    st = summary.get("size_totals", {})
    return " / ".join(f"{sz} {st[sz]}" for sz in summary.get("sizes", [])
                      if st.get(sz))


def _write_summary_sheet(wb: Workbook, summaries: list[dict]) -> None:
    """Insert the Summary 汇总 sheet at the front: one row per style sheet,
    with a hyperlink to it, joined PO/colour/date facts, and a TTL row."""
    title = "Summary 汇总" if "Summary 汇总" not in wb.sheetnames else "汇总"
    ws = wb.create_sheet(title=title, index=0)
    n_cols = len(_SUM_COLS)

    for i, (_, w) in enumerate(_SUM_COLS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _merge(ws, 1, 1, 1, n_cols,
           f"生产计划汇总（Summary）  {datetime.date.today().strftime('%Y/%m/%d')}",
           font=_FONT_SUBTITLE,
           align=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[1].height = 24

    for i, (label, _) in enumerate(_SUM_COLS, start=1):
        _cell(ws, 2, i, label, font=_FONT_HDR, fill=_HDR_FILL,
              border=_BORDER, align=_CENTER)
    ws.row_dimensions[2].height = 20

    link_font = Font(name="微软雅黑", size=10, color="FF0563C1", underline="single")
    j = "、".join

    r = 3
    for no, s in enumerate(summaries, start=1):
        _cell(ws, r, 1, no, font=_FONT_NORMAL, border=_BORDER, align=_CENTER)
        style_cell = _cell(ws, r, 2, s["style"], font=link_font,
                           border=_BORDER, align=_CENTER)
        if '"' not in s["sheet"] and '"' not in str(s["style"]):
            # click-through to the style's own buy-plan sheet; apostrophes
            # in sheet names must be doubled inside the quoted reference
            sheet_ref = s["sheet"].replace("'", "''")
            style_cell.value = f'=HYPERLINK("#\'{sheet_ref}\'!A1","{s["style"]}")'
        # 品牌 — brand-less POs are flagged; nothing is guessed for them.
        brand_txt = j(s["brands"])
        if s.get("no_brand"):
            brand_txt = (brand_txt + " ⚠ 无品牌").strip()
        _cell(ws, r, 3, brand_txt,
              font=_FONT_FLAG if s.get("no_brand") else _FONT_NORMAL,
              border=_BORDER, align=_CENTER)
        def _cn(v):
            return "" if v in (None, "") else v
        row_vals = [
            s["desc"], j(s["contracts"]), len(s["pos"]), j(s["pos"]),
            j(s["colors"]), j(s["colors_cn"]),
            _size_breakdown_text(s), s["total"],
            j(s["ship_dates"]), j(s["warehouses"]), j(s["destinations"]),
            j(s["countries"]), j(s["buyers"]), j(s["packings"]), j(s["hangers"]),
            j(s["prepacks"]), j(s["pcs_cartons"]), j(s["weights"]),
            j(s["msrps"]), j(s["rfids"]), j(s["reds"]), j(s["marks"]),
            _cn(s.get("cons_kg")), _cn(s.get("cons_cm")), _cn(s.get("cons_util")),
            _cn(s.get("cons_marker_pcs")), _cn(s.get("cons_width")), _cn(s.get("cons_gsm")),
        ]
        _center_cols = {6, _SUM_TOTAL_COL, 15, 19, 20, 21, 22, 23,
                        26, 27, 28, 29, 30, 31}   # 单耗/排版 are numeric → centre
        for i, v in enumerate(row_vals, start=4):
            _cell(ws, r, i, v, font=_FONT_NORMAL, border=_BORDER,
                  align=_CENTER if i in _center_cols else _LEFT)
        r += 1

    _cell(ws, r, 1, "TTL", font=_FONT_BOLD, fill=_YELLOW_FILL,
          border=_BORDER, align=_CENTER)
    _cell(ws, r, _SUM_TOTAL_COL, sum(s["total"] for s in summaries),
          font=_FONT_BOLD, fill=_YELLOW_FILL, border=_BORDER, align=_CENTER)

    ws.freeze_panes = "A3"


def _write_simple_summary_sheet(wb: Workbook, summaries: list[dict]) -> None:
    """简明汇总 — one row per (style, colour): 款号 / 颜色 / 面料 / dynamic
    per-size quantity columns / 总数量, with per-size TTL row."""
    title = "简明汇总" if "简明汇总" not in wb.sheetnames else "简明汇总2"
    ws = wb.create_sheet(title=title, index=1)

    sizes = _sort_sizes(list({sz for s in summaries
                              for sz in s.get("size_totals", {})}))
    C_NO, C_STYLE, C_COLOR, C_CN, C_FABRIC = 1, 2, 3, 4, 5
    C_SZ0 = 6
    C_TOTAL = C_SZ0 + len(sizes)
    n_cols = C_TOTAL

    widths = {C_NO: 6, C_STYLE: 14, C_COLOR: 20, C_CN: 14, C_FABRIC: 44,
              C_TOTAL: 10}
    for i in range(C_SZ0, C_SZ0 + len(sizes)):
        widths[i] = 8
    for i, w in widths.items():
        ws.column_dimensions[get_column_letter(i)].width = w

    _merge(ws, 1, 1, 1, n_cols, "简明汇总（款号 / 颜色 / 面料 / 尺码 / 数量）",
           font=_FONT_SUBTITLE,
           align=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[1].height = 24

    headers = ["No.", "款号", "颜色", "颜色(中文)", "面料"] + sizes + ["总数量"]
    for i, label in enumerate(headers, start=1):
        _cell(ws, 2, i, label, font=_FONT_HDR, fill=_HDR_FILL,
              border=_BORDER, align=_CENTER)
    ws.row_dimensions[2].height = 20

    size_ttl: dict[str, int] = {sz: 0 for sz in sizes}
    grand = 0
    r = 3
    for no, s in enumerate(summaries, start=1):
        r0 = r
        cn_map = s.get("color_cn_map", {})
        for color, qty_by_size in s.get("color_sizes", {}).items():
            _cell(ws, r, C_COLOR, color, font=_FONT_NORMAL, border=_BORDER, align=_LEFT)
            _cell(ws, r, C_CN, cn_map.get(color, ""), font=_FONT_NORMAL,
                  border=_BORDER, align=_CENTER)
            row_total = 0
            for i, sz in enumerate(sizes):
                q = int(qty_by_size.get(sz, 0) or 0)
                _cell(ws, r, C_SZ0 + i, q if q else None,
                      font=_FONT_NORMAL, border=_BORDER, align=_CENTER)
                row_total += q
                size_ttl[sz] += q
            _cell(ws, r, C_TOTAL, row_total, font=_FONT_NORMAL,
                  border=_BORDER, align=_CENTER)
            grand += row_total
            r += 1
        # No. + style + fabric merged over the style's colour rows
        _merge_or_set(ws, r0, C_NO, r - 1, C_NO, value=no,
                      font=_FONT_NORMAL, border=_BORDER, align=_CENTER)
        _merge_or_set(ws, r0, C_STYLE, r - 1, C_STYLE, value=s["style"],
                      font=_FONT_NORMAL, border=_BORDER, align=_CENTER)
        _merge_or_set(ws, r0, C_FABRIC, r - 1, C_FABRIC, value=s.get("fabric", ""),
                      font=_FONT_NORMAL, border=_BORDER, align=_LEFT)

    _cell(ws, r, C_NO, "TTL", font=_FONT_BOLD, fill=_YELLOW_FILL,
          border=_BORDER, align=_CENTER)
    _cell(ws, r, C_STYLE, None, font=_FONT_BOLD, fill=_YELLOW_FILL,
          border=_BORDER, align=_CENTER)
    for i, sz in enumerate(sizes):
        _cell(ws, r, C_SZ0 + i, size_ttl[sz], font=_FONT_BOLD,
              fill=_YELLOW_FILL, border=_BORDER, align=_CENTER)
    _cell(ws, r, C_TOTAL, grand, font=_FONT_BOLD, fill=_YELLOW_FILL,
          border=_BORDER, align=_CENTER)

    ws.freeze_panes = "A3"


# ---------------------------------------------------------------------------
# UPC 汇总 sheet (third sheet — two lines per PO/colour: 数量 then UPC per size)
# ---------------------------------------------------------------------------

# Leading columns merged across each PO/colour's two-line block; then a 项目
# column (数量 / UPC), the per-size columns, and a trailing 总数量.
_UPC_LEAD = [
    ("No.", 6), ("款号", 14), ("品牌", 12), ("合同号", 14), ("PO号", 16),
    ("颜色(英文)", 18), ("颜色(中文)", 14),
]


def _write_upc_summary_sheet(wb: Workbook, summaries: list[dict]) -> None:
    """UPC 汇总 — two lines per (PO, colour): a 数量 line with each size's units
    and a UPC line with that size's barcode directly beneath.  Leading columns
    are merged across the pair; a per-block 总数量 and a TTL row close it out.
    UPCs stay text so long barcodes never render in scientific notation; a size
    the PO doesn't carry stays blank (never fabricated)."""
    all_rows = [row for s in summaries for row in s.get("upc_rows", [])]
    if not all_rows:
        return

    sizes = _sort_sizes(list({r["size"] for r in all_rows}))

    # One block per (style, PO, colour): per-size units + per-size UPC.
    groups: dict[tuple, dict] = {}
    for row in all_rows:
        key = (row["style"], row["po"], row["color_en"])
        g = groups.get(key)
        if g is None:
            g = groups[key] = {
                "style": row["style"], "brand": row["brand"],
                "contract": row["contract"], "po": row["po"],
                "color_en": row["color_en"], "color_cn": row["color_cn"],
                "units": {}, "upc": {}, "total": 0,
            }
        units = int(row.get("units", 0) or 0)
        g["units"][row["size"]] = g["units"].get(row["size"], 0) + units
        if row["upc"] and not g["upc"].get(row["size"]):
            g["upc"][row["size"]] = row["upc"]
        g["total"] += units

    title = "UPC 汇总" if "UPC 汇总" not in wb.sheetnames else "UPC汇总2"
    ws = wb.create_sheet(title=title, index=2)

    C_ITEM = len(_UPC_LEAD) + 1                # 项目 (数量 / UPC)
    C_SZ0 = C_ITEM + 1                         # first size column
    C_TOTAL = C_SZ0 + len(sizes)              # 总数量 column
    n_cols = C_TOTAL

    for i, (_, w) in enumerate(_UPC_LEAD, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.column_dimensions[get_column_letter(C_ITEM)].width = 8
    for i in range(C_SZ0, C_SZ0 + len(sizes)):
        ws.column_dimensions[get_column_letter(i)].width = 15   # fits a UPC
    ws.column_dimensions[get_column_letter(C_TOTAL)].width = 10

    _merge(ws, 1, 1, 1, n_cols, "UPC 汇总（每个尺码：上行数量 / 下行 UPC）",
           font=_FONT_SUBTITLE,
           align=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[1].height = 24

    headers = [label for label, _ in _UPC_LEAD] + ["项目"] + sizes + ["总数量"]
    for i, label in enumerate(headers, start=1):
        _cell(ws, 2, i, label, font=_FONT_HDR, fill=_HDR_FILL,
              border=_BORDER, align=_CENTER)
    ws.row_dimensions[2].height = 20

    r = 3
    grand = 0
    for no, g in enumerate(groups.values(), start=1):
        grand += g["total"]
        ra, rb = r, r + 1                     # 数量 line, UPC line
        lead = [no, g["style"], g["brand"], g["contract"], g["po"],
                g["color_en"], g["color_cn"]]
        for i, v in enumerate(lead, start=1):
            _merge_or_set(ws, ra, i, rb, i, value=v, font=_FONT_NORMAL,
                          border=_BORDER, align=_CENTER if i == 1 else _LEFT)
        _cell(ws, ra, C_ITEM, "数量", font=_FONT_BOLD, border=_BORDER, align=_CENTER)
        _cell(ws, rb, C_ITEM, "UPC",  font=_FONT_BOLD, border=_BORDER, align=_CENTER)
        for k, sz in enumerate(sizes):
            units = g["units"].get(sz, 0)
            _cell(ws, ra, C_SZ0 + k, units or None, font=_FONT_NORMAL,
                  border=_BORDER, align=_CENTER)
            upc = g["upc"].get(sz, "")
            cell = _cell(ws, rb, C_SZ0 + k, upc or None, font=_FONT_NORMAL,
                         border=_BORDER, align=_CENTER)
            if upc:                            # keep long barcodes as text
                cell.number_format = "@"
        # 总数量 merged across the block's two rows.
        _merge_or_set(ws, ra, C_TOTAL, rb, C_TOTAL, value=g["total"],
                      font=_FONT_NORMAL, border=_BORDER, align=_CENTER)
        r += 2

    _cell(ws, r, 1, "TTL", font=_FONT_BOLD, fill=_YELLOW_FILL,
          border=_BORDER, align=_CENTER)
    for c in range(2, n_cols):
        _cell(ws, r, c, None, font=_FONT_BOLD, fill=_YELLOW_FILL, border=_BORDER)
    _cell(ws, r, C_TOTAL, grand, font=_FONT_BOLD, fill=_YELLOW_FILL,
          border=_BORDER, align=_CENTER)

    ws.freeze_panes = "A3"


# ---------------------------------------------------------------------------
# UPC 明细 sheet (fourth sheet — one row per size: attached header + size/UPC/qty)
# ---------------------------------------------------------------------------

_UPC_DETAIL_COLS = [
    ("No.", 6), ("款号", 14), ("品牌", 12), ("合同号", 14), ("PO号", 16),
    ("颜色(英文)", 18), ("颜色(中文)", 14), ("尺码", 8), ("UPC", 18), ("数量", 8),
]


def _write_upc_detail_sheet(wb: Workbook, summaries: list[dict]) -> None:
    """UPC 明细 — the flat per-size list: the identifying header columns
    (款号 / 品牌 / 合同号 / PO号 / 颜色) attached to each row, then 尺码 / UPC /
    数量, one row per (PO, colour, size), with a TTL. Complements the wide
    UPC 汇总 pivot for anyone who wants a plain size→UPC→qty table."""
    all_rows = [row for s in summaries for row in s.get("upc_rows", [])]
    if not all_rows:
        return

    title = "UPC 明细" if "UPC 明细" not in wb.sheetnames else "UPC明细2"
    ws = wb.create_sheet(title=title, index=3)
    n = len(_UPC_DETAIL_COLS)
    for i, (_, w) in enumerate(_UPC_DETAIL_COLS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _merge(ws, 1, 1, 1, n, "UPC 明细（每尺码一行：尺码 / UPC / 数量）",
           font=_FONT_SUBTITLE,
           align=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[1].height = 24
    for i, (label, _) in enumerate(_UPC_DETAIL_COLS, start=1):
        _cell(ws, 2, i, label, font=_FONT_HDR, fill=_HDR_FILL,
              border=_BORDER, align=_CENTER)
    ws.row_dimensions[2].height = 20

    r, grand = 3, 0
    for no, row in enumerate(all_rows, start=1):
        units = int(row.get("units", 0) or 0)
        grand += units
        vals = [no, row["style"], row["brand"], row["contract"], row["po"],
                row["color_en"], row["color_cn"], row["size"], row["upc"], units]
        for i, v in enumerate(vals, start=1):
            cell = _cell(ws, r, i, v, font=_FONT_NORMAL, border=_BORDER,
                         align=_CENTER if i in (1, 8, 9, 10) else _LEFT)
            if i == 9 and v:                    # UPC → text (no scientific notation)
                cell.number_format = "@"
        r += 1

    _cell(ws, r, 1, "TTL", font=_FONT_BOLD, fill=_YELLOW_FILL,
          border=_BORDER, align=_CENTER)
    for c in range(2, n):
        _cell(ws, r, c, None, font=_FONT_BOLD, fill=_YELLOW_FILL, border=_BORDER)
    _cell(ws, r, n, grand, font=_FONT_BOLD, fill=_YELLOW_FILL,
          border=_BORDER, align=_CENTER)
    ws.freeze_panes = "A3"
