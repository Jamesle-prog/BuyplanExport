"""Standard cutting-plan workbook.

Cut plans arrive from different sources in different shapes (Optitex exports
one layout, other factories hand over their own).  This exporter re-emits any
of them in a single house layout — the Optitex Cut Plan structure, which the
cutting room already reads — so every PO's plan looks the same:

    header (date / order / styles / operator / client)
    Order demands            <- from the PO, always
    per material:
        Marker Definition
        Marker Ratio
        Spreading Plies
        Solution
        Total Efficiency / Cost / cut length / Tables / Average Length
    Total Tables             <- grand total

The Order demands block is filled from the PO, which is the authority on what
was ordered.  The marker blocks come from an uploaded plan when one is linked;
without a plan they are written as an empty scaffold for the cutting room to
complete, so the layout is still the standard one.
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ._excel_helpers import apply_print_settings, clean_sheet_name

_TITLE_FILL = PatternFill("solid", fgColor="DDEBF7")
_HEAD_FILL = PatternFill("solid", fgColor="F2F2F2")
_SUM_FILL = PatternFill("solid", fgColor="FFF2CC")
_BOLD = Font(bold=True)
_TITLE_FONT = Font(bold=True, size=11)
_THIN = Side(style="thin", color="B0B0B0")
_BOX = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center")

# Fallback ordering for sizes when neither the plan nor the PO can supply one.
CANONICAL_SIZES = ["XXS", "XS", "S", "M", "L", "XL", "2XL", "3XL", "4XL"]

_METRIC_HEADERS = ["Fabric Length,m", "Efficiency,%", "Cut Length,m",
                   "Marker Length,cm", "Material Cost,CNY"]
_MARKER_DEF_HEADERS = ["Material", "N Markers", "Spreading", "Width, cm",
                       "Length, cm", "Min Plies", "Max Plies",
                       "Waste Limits, cm"]


# ---------------------------------------------------------------------------
# Small writing helpers
# ---------------------------------------------------------------------------

def _put(ws, row: int, col: int, value: Any, *, bold: bool = False,
         fill: PatternFill | None = None, box: bool = False,
         center: bool = False, fmt: str | None = None):
    cell = ws.cell(row=row, column=col, value=value)
    if bold:
        cell.font = _BOLD
    if fill is not None:
        cell.fill = fill
    if box:
        cell.border = _BOX
    if center:
        cell.alignment = _CENTER
    if fmt:
        cell.number_format = fmt
    return cell


def _section(ws, row: int, title: str) -> int:
    _put(ws, row, 1, title, bold=True, fill=_TITLE_FILL)
    ws.cell(row=row, column=1).font = _TITLE_FONT
    return row + 1


def _write_size_header(ws, row: int, first_col: int,
                       groups: list[tuple[str, list[str]]]) -> int:
    """Write the two header rows of a size matrix (styles, then sizes).

    Returns the row after the size row.  Each group's style name is merged
    across its own size columns, matching the source layout.
    """
    col = first_col
    for style, sizes in groups:
        if not sizes:
            continue
        _put(ws, row, col, style, bold=True, fill=_HEAD_FILL,
             box=True, center=True)
        if len(sizes) > 1:
            ws.merge_cells(start_row=row, start_column=col,
                           end_row=row, end_column=col + len(sizes) - 1)
            for c in range(col + 1, col + len(sizes)):
                _put(ws, row, c, None, fill=_HEAD_FILL, box=True)
        for i, size in enumerate(sizes):
            _put(ws, row + 1, col + i, size, bold=True, fill=_HEAD_FILL,
                 box=True, center=True)
        col += len(sizes)
    return row + 2


def _flat_cols(groups: list[tuple[str, list[str]]]) -> list[tuple[str, str]]:
    """Flatten the size groups into ordered (style, size) column keys."""
    return [(style, size) for style, sizes in groups for size in sizes]


def _num_fmt(value: float | None, digits: int = 2) -> float | None:
    if value is None:
        return None
    return round(float(value), digits)


# ---------------------------------------------------------------------------
# Blocks
# ---------------------------------------------------------------------------

def _write_header(ws, row: int, header: dict[str, Any]) -> int:
    _put(ws, row, 1, "Date", bold=True)
    _put(ws, row, 2, header.get("plan_date", ""))
    _put(ws, row, 4, "Time", bold=True)
    _put(ws, row, 5, header.get("plan_time", ""))
    row += 1
    _put(ws, row, 1, "Order name", bold=True)
    _put(ws, row, 2, header.get("order_name", ""))
    row += 1
    for i, style in enumerate(header.get("styles", []) or [], start=1):
        _put(ws, row, 1, f"Style file {i}", bold=True)
        _put(ws, row, 2, style.get("file", ""))
        row += 1
        _put(ws, row, 1, f"Style name {i}", bold=True)
        _put(ws, row, 2, style.get("name", ""))
        row += 1
    _put(ws, row, 1, "Cut plan operator", bold=True)
    _put(ws, row, 2, header.get("operator", ""))
    row += 1
    _put(ws, row, 1, "Client", bold=True)
    _put(ws, row, 2, header.get("client", ""))
    row += 1
    if header.get("po_summary"):
        _put(ws, row, 1, "PO No.(s)", bold=True)
        _put(ws, row, 2, header["po_summary"])
        row += 1
    if header.get("pc_summary"):
        _put(ws, row, 1, "PC No.(s)", bold=True)
        _put(ws, row, 2, header["pc_summary"])
        row += 1
    if header.get("style_summary"):
        # The PO's own style codes — the plan's CAD style names above don't
        # match them, so the cutting room needs both on the sheet.
        _put(ws, row, 1, "PO Style(s)", bold=True)
        _put(ws, row, 2, header["style_summary"])
        row += 1
    if header.get("output_folder"):
        row += 1
        _put(ws, row, 1, "Output folder path", bold=True)
        _put(ws, row, 2, header["output_folder"])
        row += 1
    return row + 1


def _write_demands(ws, row: int, groups: list[tuple[str, list[str]]],
                   colors: list[str],
                   qty: dict[tuple[str, str, str], int]) -> int:
    """Order demands: colours down, (style, size) across, with a Sum row."""
    _put(ws, row, 2, "Order demands", bold=True, fill=_TITLE_FILL)
    ws.cell(row=row, column=2).font = _TITLE_FONT
    row += 1
    _put(ws, row, 2, "Colors", bold=True, fill=_HEAD_FILL, box=True)
    _put(ws, row, 3, "Sizes", bold=True, fill=_HEAD_FILL, box=True)
    row = _write_size_header(ws, row + 1, 3, groups)

    cols = _flat_cols(groups)
    for color in colors:
        _put(ws, row, 2, color, bold=True, box=True)
        for i, (style, size) in enumerate(cols):
            _put(ws, row, 3 + i, qty.get((style, color, size), 0) or None,
                 box=True, center=True)
        row += 1
    _put(ws, row, 2, "Sum", bold=True, fill=_SUM_FILL, box=True)
    for i, (style, size) in enumerate(cols):
        total = sum(qty.get((style, c, size), 0) for c in colors)
        _put(ws, row, 3 + i, total, bold=True, fill=_SUM_FILL,
             box=True, center=True)
    return row + 2


def _material_groups(mat: dict[str, Any],
                     fallback: list[tuple[str, list[str]]]
                     ) -> list[tuple[str, list[str]]]:
    """Size groups actually used by *mat* — a lining is often cut for only
    one of the styles, so its blocks are narrower than the demand matrix.

    Prefers the column order the parser recorded from the plan's own header
    rows.  Only when that's missing are the groups rebuilt from the data
    cells, and then the sizes are put back into canonical order: cell order
    reflects whichever marker happened to be read first, and the demand
    matrix can't supply an order either because the PO's style codes differ
    from the CAD style names in the plan.
    """
    recorded = mat.get("groups")
    if recorded:
        return [(style, list(sizes)) for style, sizes in recorded]

    seen: dict[str, list[str]] = {}
    sources: Iterable[dict] = (
        [c for m in mat.get("markers", []) for c in m.get("ratio", [])]
        + [p for s in mat.get("spreads", []) for r in s.get("rows", [])
           for p in r.get("pieces", [])]
        + [c for s in mat.get("solution", []) for c in s.get("cells", [])]
    )
    for cell in sources:
        style, size = cell.get("style", ""), cell.get("size", "")
        seen.setdefault(style, [])
        if size not in seen[style]:
            seen[style].append(size)
    if not seen:
        return fallback

    style_order = {style: i for i, (style, _s) in enumerate(fallback)}
    per_style_sizes = {
        style: {s: i for i, s in enumerate(sizes)} for style, sizes in fallback
    }

    def size_key(style: str, size: str) -> tuple[int, int]:
        known = per_style_sizes.get(style, {})
        if size in known:
            return (0, known[size])
        if size in CANONICAL_SIZES:
            return (1, CANONICAL_SIZES.index(size))
        return (2, 0)

    return [
        (style, sorted(sizes, key=lambda s: size_key(style, s)))
        for style, sizes in sorted(seen.items(),
                                   key=lambda kv: style_order.get(kv[0], 99))
    ]


def _write_material(ws, row: int, mat: dict[str, Any],
                    fallback_groups: list[tuple[str, list[str]]]) -> int:
    groups = _material_groups(mat, fallback_groups)
    cols = _flat_cols(groups)

    # ── Marker Definition ────────────────────────────────────────────────
    row = _section(ws, row, "Marker Definition")
    for i, head in enumerate(_MARKER_DEF_HEADERS):
        _put(ws, row, 1 + i, head, bold=True, fill=_HEAD_FILL,
             box=True, center=True)
    row += 1
    for i, key in enumerate(["material", "n_markers", "spreading", "width_cm",
                             "length_cm", "min_plies", "max_plies",
                             "waste_limits"]):
        _put(ws, row, 1 + i, mat.get(key), box=True, center=True)
    row += 2

    # ── Marker Ratio ─────────────────────────────────────────────────────
    row = _section(ws, row, "Marker Ratio")
    _put(ws, row, 1, "Marker", bold=True, fill=_HEAD_FILL, box=True)
    _put(ws, row, 2, "File Name", bold=True, fill=_HEAD_FILL, box=True)
    _put(ws, row, 3, "Sizes", bold=True, fill=_HEAD_FILL, box=True)
    row = _write_size_header(ws, row + 1, 3, groups)
    for marker in mat.get("markers", []):
        _put(ws, row, 1, marker.get("marker_no"), box=True, center=True)
        _put(ws, row, 2, marker.get("file_name", ""), box=True)
        ratio = {(c.get("style", ""), c.get("size", "")): c.get("qty")
                 for c in marker.get("ratio", [])}
        for i, key in enumerate(cols):
            _put(ws, row, 3 + i, ratio.get(key), box=True, center=True)
        row += 1
    row += 1

    # ── Spreading Plies ──────────────────────────────────────────────────
    row = _section(ws, row, "Spreading Plies")
    metric_col = 3 + len(cols)
    for spread in mat.get("spreads", []):
        _put(ws, row, 1, f"Marker {spread.get('marker_no')}", bold=True)
        _put(ws, row, 2, spread.get("file_name", ""))
        _put(ws, row, 3, "Sizes", bold=True, fill=_HEAD_FILL, box=True)
        for i, head in enumerate(_METRIC_HEADERS):
            _put(ws, row, metric_col + i, head, bold=True, fill=_HEAD_FILL,
                 box=True, center=True)
        _put(ws, row + 1, 1, "Colors", bold=True, fill=_HEAD_FILL, box=True)
        _put(ws, row + 1, 2, "Plies number", bold=True, fill=_HEAD_FILL,
             box=True)
        row = _write_size_header(ws, row + 1, 3, groups)
        sums: dict[tuple[str, str], int] = {}
        plies_total = 0
        for line in spread.get("rows", []):
            _put(ws, row, 1, line.get("color", ""), bold=True, box=True)
            _put(ws, row, 2, line.get("plies"), box=True, center=True)
            plies_total += int(line.get("plies") or 0)
            pieces = {(c.get("style", ""), c.get("size", "")): c.get("qty")
                      for c in line.get("pieces", [])}
            for i, key in enumerate(cols):
                val = pieces.get(key)
                _put(ws, row, 3 + i, val, box=True, center=True)
                if val:
                    sums[key] = sums.get(key, 0) + int(val)
            for i, key in enumerate(["fabric_length_m", "efficiency_pct",
                                     "cut_length_m", "marker_length_cm",
                                     "cost"]):
                _put(ws, row, metric_col + i, _num_fmt(line.get(key), 4),
                     box=True, center=True, fmt="0.00")
            row += 1
        _put(ws, row, 1, "Sum", bold=True, fill=_SUM_FILL, box=True)
        _put(ws, row, 2, plies_total or None, bold=True, fill=_SUM_FILL,
             box=True, center=True)
        for i, key in enumerate(cols):
            _put(ws, row, 3 + i, sums.get(key, 0), bold=True, fill=_SUM_FILL,
                 box=True, center=True)
        row += 2

    # ── Solution ─────────────────────────────────────────────────────────
    row = _section(ws, row, "Solution")
    solution = mat.get("solution", [])
    has_kind = any(s.get("kind") for s in solution)
    size_start = 4 if has_kind else 3
    _put(ws, row, 2, "Colors", bold=True, fill=_HEAD_FILL, box=True)
    if has_kind:
        _put(ws, row, 3, "Quantity", bold=True, fill=_HEAD_FILL, box=True)
    _put(ws, row, size_start, "Sizes", bold=True, fill=_HEAD_FILL, box=True)
    sol_metric_col = size_start + len(cols)
    _put(ws, row, sol_metric_col, "Total Quantity", bold=True,
         fill=_HEAD_FILL, box=True, center=True)
    _put(ws, row, sol_metric_col + 1, "Fabric Length,m", bold=True,
         fill=_HEAD_FILL, box=True, center=True)
    row = _write_size_header(ws, row + 1, size_start, groups)
    last_color = None
    for sol in solution:
        color = sol.get("color", "")
        if color != last_color:
            _put(ws, row, 2, color, bold=True, box=True)
            last_color = color
        if has_kind:
            _put(ws, row, 3, (sol.get("kind") or "").title(), box=True,
                 center=True)
        cells = {(c.get("style", ""), c.get("size", "")): c.get("qty")
                 for c in sol.get("cells", [])}
        for i, key in enumerate(cols):
            _put(ws, row, size_start + i, cells.get(key), box=True,
                 center=True)
        _put(ws, row, sol_metric_col, sol.get("total_qty") or None,
             bold=True, box=True, center=True)
        _put(ws, row, sol_metric_col + 1, _num_fmt(sol.get("fabric_length_m"), 4),
             box=True, center=True, fmt="0.00")
        row += 1
    row += 1

    # ── Totals ───────────────────────────────────────────────────────────
    for label, key, unit, fmt in (
        ("Total Efficiency", "total_efficiency_pct", "%", "0.00"),
        ("Total Cost", "total_cost", "CNY", "0.00"),
        ("Total cut length", "total_cut_length_m", "m", "0.00"),
        ("Total Tables", "total_tables", "", "0"),
        ("Average Length", "average_length_m", "m", "0.0000"),
    ):
        _put(ws, row, 1, label, bold=True)
        _put(ws, row, 2, _num_fmt(mat.get(key), 4), fmt=fmt)
        if unit:
            _put(ws, row, 3, unit)
        row += 1
    return row + 1


def _write_empty_material_scaffold(
        ws, row: int, groups: list[tuple[str, list[str]]]) -> int:
    """The standard blocks with no data — used when no plan is linked yet."""
    return _write_material(ws, row, {
        "material": None, "n_markers": None, "spreading": None,
        "width_cm": None, "length_cm": None, "min_plies": None,
        "max_plies": None, "waste_limits": None,
        "markers": [], "spreads": [], "solution": [],
    }, groups)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def build_standard_cut_plan(*, header: dict[str, Any],
                            groups: list[tuple[str, list[str]]],
                            colors: list[str],
                            demand_qty: dict[tuple[str, str, str], int],
                            materials: list[dict[str, Any]] | None = None,
                            sheet_name: str = "Cut Plan",
                            clean: bool = False) -> bytes:
    """Render a cutting plan in the standard layout and return .xlsx bytes.

    ``groups``      ordered [(style, [size, ...]), ...] — the demand columns.
    ``colors``      ordered colour rows.
    ``demand_qty``  {(style, colour, size): qty} from the PO.
    ``materials``   parsed material blocks from a linked plan; when empty an
                    empty scaffold is written so the layout is still standard.
    ``clean``       run the house cleanup (Chinese labels, marker paths reduced
                    to bare names) — the step that used to be a hand-run Excel
                    macro.

    ``clean`` defaults to **False** on purpose: this layout is re-parseable by
    ``parsers.cutting_plan`` (which finds its blocks by English anchor text),
    and the app round-trips its own export. Cleaning is a *delivery* step —
    turn it on for the copy handed to the cutting room, not for the canonical
    workbook.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = clean_sheet_name(sheet_name, fallback="Cut Plan")

    row = _write_header(ws, 1, header)
    row = _write_demands(ws, row, groups, colors, demand_qty)

    materials = materials or []
    if materials:
        for mat in materials:
            row = _write_material(ws, row, mat, groups)
        _put(ws, row, 1, "Total Tables", bold=True)
        _put(ws, row, 2, sum(int(m.get("total_tables") or 0)
                             for m in materials) or None)
    else:
        row = _write_empty_material_scaffold(ws, row, groups)
        _put(ws, row, 1, "Total Tables", bold=True)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 34
    n_cols = 3 + len(_flat_cols(groups)) + len(_METRIC_HEADERS)
    for c in range(3, n_cols + 2):
        ws.column_dimensions[get_column_letter(c)].width = 13
    ws.freeze_panes = "C2"

    # Last step, after every cell is written and before saving — the same
    # point the Excel macro ran at.
    if clean:
        from .cutting_plan_clean import clean_workbook
        clean_workbook(wb)

    apply_print_settings(wb)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def plan_header_from_parsed(parsed: dict[str, Any]) -> dict[str, Any]:
    """Header dict for :func:`build_standard_cut_plan` from a parsed plan."""
    return {
        "plan_date": parsed.get("plan_date", ""),
        "plan_time": parsed.get("plan_time", ""),
        "order_name": parsed.get("order_name", ""),
        "styles": parsed.get("styles", []),
        "operator": parsed.get("operator", ""),
        "client": parsed.get("client", ""),
        "output_folder": parsed.get("output_folder", ""),
    }


def today_header(order_name: str, *, client: str = "", operator: str = "",
                 styles: list[dict] | None = None) -> dict[str, Any]:
    """Header dict stamped with the current date/time."""
    now = datetime.now()
    return {
        "plan_date": now.strftime("%B %d - %Y"),
        "plan_time": now.strftime("%H:%M"),
        "order_name": order_name,
        "styles": styles or [],
        "operator": operator,
        "client": client,
    }
