"""Factory progress request form — the Excel round-trip (工厂进度回报表).

``build_progress_request_xlsx``: generate a pre-filled form for one factory
listing their PO/styles with ordered qty and already-reported totals; the
factory fills in the 本次新增 (new since last report) columns + date and
sends the file back.

``parse_progress_report_xlsx``: read a returned form back into report-row
dicts ready for ``FactoryProgressStore.add_report`` — with validation
issues collected per row rather than raised, so the UI can show a preview
of exactly what will be imported and what was skipped.

The form is deliberately dumb-Excel: no macros, no validation lists that
break in WPS, fixed header row detected by cell text on re-import (so an
extra title row or renamed file doesn't matter).
"""
from __future__ import annotations

import io
from datetime import date

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from ..store._factory_progress_schema import (
    REPORT_STAGES, MILESTONE_STAGES, MILESTONE_LABELS,
)
from ..utils.normalize import cell_date_text as _cell_date

# Column layout of the request form (1-based). The 本次新增 columns are what
# the factory fills in; everything else is context we pre-fill.
_FORM_HEADERS = [
    "PO号 PO No.",              # 1
    "款号 Style",                # 2
    "订单数 Order Qty",          # 3
    "已报裁剪 Cut so far",       # 4
    "已报车缝 Sewn so far",      # 5
    "已报包装 Packed so far",    # 6
    "本次新增裁剪 New Cut",      # 7  <- factory fills
    "本次新增车缝 New Sewn",     # 8  <- factory fills
    "本次新增包装 New Packed",   # 9  <- factory fills
    "报告日期 Report Date",      # 10 <- factory fills (YYYY-MM-DD)
    "备注 Notes",                # 11 <- factory fills (optional)
]

# Column index (1-based) -> stage key for the fill-in columns.
_STAGE_COLS = {7: "cutting", 8: "sewing", 9: "packing"}
_COL_PO, _COL_STYLE, _COL_QTY = 1, 2, 3
_COL_DATE, _COL_NOTES = 10, 11

_SHEET_TITLE = "进度回报 Progress"

# Milestone sheet (sheet 2) layout — one row per (PO, style, milestone).
_MS_SHEET_TITLE = "里程碑 Milestones"
_MS_HEADERS = [
    "PO号 PO No.",                 # 1
    "款号 Style",                   # 2
    "里程碑 Milestone",             # 3  (fixed label — do not edit)
    "阶段代码 Code",                # 4  (internal stage key — do not edit)
    "预计完成 Expected Date",       # 5  <- editable (YYYY-MM-DD)
    "状态备注 Status Note",         # 6  <- editable
    "完成日期 Completed Date",      # 7  <- editable (YYYY-MM-DD; fill = done/arrived)
]
_MS_COL_PO, _MS_COL_STYLE, _MS_COL_LABEL, _MS_COL_CODE = 1, 2, 3, 4
_MS_COL_EXPECTED, _MS_COL_NOTE, _MS_COL_DONE = 5, 6, 7


def build_progress_request_xlsx(factory: str, rows: list[dict],
                                milestones: list[dict] | None = None) -> bytes:
    """Build the request form for one *factory*.

    Each row dict: ``{"po_number", "style", "order_qty", "cut", "sewn",
    "packed"}`` (the last four ints; already-reported totals may be 0).

    *milestones* (optional): dicts with ``po_number``, ``style``, ``stage``
    (a MILESTONE_STAGES key), ``expected``, ``note``, ``completed`` — one
    pre-filled row each on a second "里程碑 Milestones" sheet where the
    factory updates expected dates / status notes and marks completion.
    Returns the xlsx file as bytes (callers hand it to a download button).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _SHEET_TITLE

    ws.cell(1, 1, value=(
        f"工厂进度回报表 Factory Progress Report — {factory} — "
        f"生成日期 {date.today().isoformat()}"
    )).font = Font(bold=True, size=12)
    ws.cell(2, 1, value=(
        "请只填写右侧「本次新增」三列 + 报告日期（YYYY-MM-DD），"
        "数量为自上次回报以来新完成的件数（不是累计数）。"
        "Fill ONLY the three 'New' columns + Report Date; quantities are "
        "units completed SINCE the last report (not cumulative)."
    )).font = Font(size=9, italic=True, color="FF808080")

    header_row = 4
    for ci, h in enumerate(_FORM_HEADERS, 1):
        cell = ws.cell(header_row, ci, value=h)
        cell.fill = PatternFill(start_color="FF1F3864", end_color="FF1F3864",
                                fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
    ws.row_dimensions[header_row].height = 30

    fill_hint = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC",
                            fill_type="solid")   # light yellow = fill me in
    for ri, row in enumerate(rows, header_row + 1):
        ws.cell(ri, _COL_PO,    value=row.get("po_number", ""))
        ws.cell(ri, _COL_STYLE, value=row.get("style", ""))
        _qty = row.get("order_qty")
        ws.cell(ri, _COL_QTY,   value="" if _qty in (None, "") else _qty)
        ws.cell(ri, 4, value=row.get("cut", 0))
        ws.cell(ri, 5, value=row.get("sewn", 0))
        ws.cell(ri, 6, value=row.get("packed", 0))
        for ci in (*_STAGE_COLS, _COL_DATE, _COL_NOTES):
            ws.cell(ri, ci).fill = fill_hint

    widths = {1: 16, 2: 14, 3: 11, 4: 11, 5: 11, 6: 11, 7: 13, 8: 13, 9: 13,
              10: 14, 11: 24}
    from openpyxl.utils import get_column_letter
    for ci, w in widths.items():
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = ws.cell(header_row + 1, 1).coordinate

    # ── Sheet 2: milestones (expected date / status note / completed) ───────
    if milestones:
        ms = wb.create_sheet(_MS_SHEET_TITLE)
        ms.cell(1, 1, value=(
            "里程碑跟踪 Milestone Tracking — 请更新右侧三列：预计完成日期、状态"
            "备注、完成日期（填写完成日期即表示该里程碑已完成/已到厂）。"
            "Update the three right columns; filling Completed Date marks the "
            "milestone as done/arrived."
        )).font = Font(size=9, italic=True, color="FF808080")

        ms_header_row = 3
        for ci, h in enumerate(_MS_HEADERS, 1):
            cell = ms.cell(ms_header_row, ci, value=h)
            cell.fill = PatternFill(start_color="FF1F3864", end_color="FF1F3864",
                                    fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFFFF", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
        ms.row_dimensions[ms_header_row].height = 28

        for ri, m in enumerate(milestones, ms_header_row + 1):
            ms.cell(ri, _MS_COL_PO,       value=m.get("po_number", ""))
            ms.cell(ri, _MS_COL_STYLE,    value=m.get("style", ""))
            ms.cell(ri, _MS_COL_LABEL,
                    value=MILESTONE_LABELS.get(m.get("stage", ""), m.get("stage", "")))
            ms.cell(ri, _MS_COL_CODE,     value=m.get("stage", ""))
            ms.cell(ri, _MS_COL_EXPECTED, value=m.get("expected", "") or "")
            ms.cell(ri, _MS_COL_NOTE,     value=m.get("note", "") or "")
            ms.cell(ri, _MS_COL_DONE,     value=m.get("completed", "") or "")
            for ci in (_MS_COL_EXPECTED, _MS_COL_NOTE, _MS_COL_DONE):
                ms.cell(ri, ci).fill = fill_hint

        for ci, w in {1: 16, 2: 14, 3: 26, 4: 18, 5: 15, 6: 30, 7: 15}.items():
            ms.column_dimensions[get_column_letter(ci)].width = w
        ms.freeze_panes = ms.cell(ms_header_row + 1, 1).coordinate

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_progress_report_xlsx(content: bytes, factory: str = "") -> dict:
    """Parse a returned form. Returns::

        {"reports": [ {po_number, style, stage, units, report_date,
                       factory, notes}, ... ],
         "issues":  [ "row 6: ...", ... ],   # skipped/suspect rows
         "rows_seen": int}

    One report dict per non-empty New-column cell (a row reporting cut AND
    sewn yields two reports). Rows with no New quantities at all are simply
    unchanged rows of the form, not issues. Never raises on cell content —
    file-level problems (unreadable, header row not found) do raise
    ValueError so the UI can show a clear error.

    When the form carries a "里程碑 Milestones" sheet, its rows come back as
    ``"milestones"``: ``{po_number, style, stage, expected, note,
    completed}`` — one dict per row that has ANY of the three editable
    fields filled (the importer applies them to production_tracking).
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    try:
        ws = wb[_SHEET_TITLE] if _SHEET_TITLE in wb.sheetnames else wb.active

        # Find the header row by its first cell text (tolerates added rows).
        header_row = None
        for r in range(1, min(ws.max_row, 20) + 1):
            if str(ws.cell(r, 1).value or "").strip().startswith("PO号"):
                header_row = r
                break
        if header_row is None:
            raise ValueError(
                "Header row not found — is this the progress request form? "
                "(expected a 'PO号 PO No.' header cell)"
            )

        reports: list[dict] = []
        issues: list[str] = []
        rows_seen = 0
        for r in range(header_row + 1, ws.max_row + 1):
            po = str(ws.cell(r, _COL_PO).value or "").strip()
            style = str(ws.cell(r, _COL_STYLE).value or "").strip()
            if not po:
                continue
            rows_seen += 1

            raw_date = ws.cell(r, _COL_DATE).value
            if hasattr(raw_date, "isoformat"):          # datetime/date cell
                report_date = raw_date.date().isoformat() if hasattr(raw_date, "date") \
                    else raw_date.isoformat()
            else:
                report_date = str(raw_date or "").strip()
            notes = str(ws.cell(r, _COL_NOTES).value or "").strip()

            row_units: dict[str, int] = {}
            for ci, stage in _STAGE_COLS.items():
                raw = ws.cell(r, ci).value
                if raw in (None, ""):
                    continue
                try:
                    units = int(float(raw))
                except (TypeError, ValueError):
                    issues.append(f"row {r}: {stage} value {raw!r} is not a number — skipped")
                    continue
                if units == 0:
                    continue
                if units < 0:
                    issues.append(f"row {r}: negative {stage} value {units} — skipped "
                                  "(delete the wrong earlier report instead)")
                    continue
                row_units[stage] = units

            if not row_units:
                continue
            if not report_date:
                issues.append(f"row {r} ({po} / {style}): quantities given but "
                              f"报告日期 Report Date is empty — skipped")
                continue

            for stage, units in row_units.items():
                reports.append({
                    "po_number": po, "style": style, "stage": stage,
                    "units": units, "report_date": report_date,
                    "factory": factory, "notes": notes,
                })

        # ── Milestone sheet (optional) ──────────────────────────────────────
        milestones: list[dict] = []
        if _MS_SHEET_TITLE in wb.sheetnames:
            ms = wb[_MS_SHEET_TITLE]
            ms_header = None
            for r in range(1, min(ms.max_row, 20) + 1):
                if str(ms.cell(r, 1).value or "").strip().startswith("PO号"):
                    ms_header = r
                    break
            if ms_header is not None:
                _known_stages = {k for k, _ in MILESTONE_STAGES}


                for r in range(ms_header + 1, ms.max_row + 1):
                    po = str(ms.cell(r, _MS_COL_PO).value or "").strip()
                    stage = str(ms.cell(r, _MS_COL_CODE).value or "").strip()
                    if not po or not stage:
                        continue
                    if stage not in _known_stages:
                        issues.append(
                            f"milestone row {r}: unknown stage code {stage!r} — skipped")
                        continue
                    expected = _cell_date(ms.cell(r, _MS_COL_EXPECTED).value)
                    note = str(ms.cell(r, _MS_COL_NOTE).value or "").strip()
                    completed = _cell_date(ms.cell(r, _MS_COL_DONE).value)
                    if not (expected or note or completed):
                        continue
                    milestones.append({
                        "po_number": po,
                        "style": str(ms.cell(r, _MS_COL_STYLE).value or "").strip(),
                        "stage": stage, "expected": expected,
                        "note": note, "completed": completed,
                    })

        return {"reports": reports, "milestones": milestones,
                "issues": issues, "rows_seen": rows_seen}
    finally:
        wb.close()


# ── Returned BUY PLAN import — the Index tab as the round-trip form ─────────
#
# The intended loop: the buy plan generates with its tracking columns empty
# (they only fill once tracking data exists), the factory/merchandiser fills
# the expected dates directly in the Index tab, the returned file is imported
# here, and the NEXT buy plan reflects reality. Header text → tracking field:
_BP_INDEX_HEADER_MAP: dict[str, tuple[str, str]] = {
    "生产工厂":               ("factory",            "base"),
    "工厂交期":               ("shipping",           "stage"),
    "面料（计划）到厂时间":    ("fabric_purchase",    "stage"),
    "辅料（计划）到厂时间":    ("trim_purchase",      "stage"),
    "样衣（计划）确认时间":    ("pp_sample",          "stage"),
    "大货版（计划）完成时间":  ("base_size_pattern",  "stage"),
    "全码版（计划）完成时间":  ("full_sized_pattern", "stage"),
    "裁剪（计划）完成时间":    ("cutting",            "stage"),
    "车位（计划）完成时间":    ("sewing",             "stage"),
    "后道（计划）完成时间":    ("packing",            "stage"),
    # 裁剪计划（计划）完成时间 has no tracking stage; 裁剪数/出货数 are
    # quantities (reported via the progress form's dated log, not here).
}


def parse_buyplan_index_tracking(content: bytes) -> dict:
    """Parse the tracking columns of a returned Sky East buy plan's Index tab.

    Returns ``{"rows": [{style, pc_no, factory, planned: {stage: date}}],
    "issues": [...]}`` — one row per (style, PC No.) with at least one filled
    tracking cell; repeated Index rows for the same pair (multi-fabric
    styles) are merged, last non-empty value winning. Only EXPECTED (计划)
    dates travel on this form — completion marking stays online or on the
    progress form's 里程碑 sheet. Raises ValueError when the file has no
    recognisable Index sheet.
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    try:
        ws = wb["Index"] if "Index" in wb.sheetnames else None
        if ws is None:
            raise ValueError("No 'Index' sheet — is this a generated buy plan?")

        headers = {str(c.value or "").strip(): c.column for c in ws[1] if c.value}
        if "款号" not in headers or "客人PC NO" not in headers:
            raise ValueError(
                "Index sheet is missing the 款号 / 客人PC NO columns — "
                "is this a generated buy plan?"
            )
        col_style, col_pc = headers["款号"], headers["客人PC NO"]
        field_cols = {
            h: (headers[h], spec) for h, spec in _BP_INDEX_HEADER_MAP.items()
            if h in headers
        }


        merged: dict[tuple[str, str], dict] = {}
        issues: list[str] = []
        for r in range(2, ws.max_row + 1):
            style = str(ws.cell(r, col_style).value or "").strip()
            pc_no = str(ws.cell(r, col_pc).value or "").strip()
            if not style or not pc_no:
                continue
            row = merged.setdefault(
                (pc_no, style),
                {"style": style, "pc_no": pc_no, "factory": "", "planned": {}},
            )
            for _h, (col, (field, kind)) in field_cols.items():
                raw = ws.cell(r, col).value
                if raw in (None, ""):
                    continue
                if kind == "base":
                    row["factory"] = str(raw).strip()
                else:
                    row["planned"][field] = _cell_date(raw)

        rows = [row for row in merged.values()
                if row["factory"] or row["planned"]]
        return {"rows": rows, "issues": issues}
    finally:
        wb.close()
