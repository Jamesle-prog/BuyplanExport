"""Tracking grid ⇄ Excel round-trip.

The 🏭 Tracking grid is nine milestone dates per PO/style, in **计划 Planned**
and **实际 Actual** flavours. Typing them into the on-screen editor is fine for
a few rows; for a whole season it is faster to edit in Excel. This module is
that door:

``build_tracking_grid_xlsx(records)`` writes one row per tracked PO/style with
BOTH the planned and actual date of every milestone, so nothing is lost in the
round-trip.

``parse_tracking_grid_xlsx(content)`` reads it back into per-row field updates
ready for ``ProductionTrackingStore.update_stage_fields`` — applying dates the
same way the grid does (a filled 实际 Actual date marks that milestone Done).

Design choices that keep it safe and WPS-friendly:
  - Columns are matched on **header text**, not position, so reordering
    columns in Excel doesn't scramble the import.
  - A **blank cell never erases** a stored date — you can't wipe a season's
    plan by clearing a column. To clear one date, do it in the on-screen grid.
  - PO号 / 款号 identify the row; unrecognised or untracked rows are reported,
    never silently dropped.
"""
from __future__ import annotations

import io
from datetime import date

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..store._factory_progress_schema import MILESTONE_STAGES, MILESTONE_LABELS
from ..utils.normalize import cell_date_text as _cell_date

_SHEET_TITLE = "跟踪 Tracking"

_COL_PO, _COL_STYLE, _COL_FACTORY = 1, 2, 3
_FIRST_DATE_COL = 4                       # milestone date columns start here

# Header suffixes that distinguish the two date flavours of one milestone.
_PLANNED_SUFFIX = "｜计划 Planned"
_ACTUAL_SUFFIX  = "｜实际 Actual"


def _planned_header(label: str) -> str:
    return f"{label} {_PLANNED_SUFFIX}"


def _actual_header(label: str) -> str:
    return f"{label} {_ACTUAL_SUFFIX}"


# header text → (stage_key, kind) for re-import. Built once from the milestone
# contract so export and import can never disagree about a column's meaning.
_HEADER_TO_FIELD: dict[str, tuple[str, str]] = {}
for _stage, _label in MILESTONE_STAGES:
    _HEADER_TO_FIELD[_planned_header(_label)] = (_stage, "planned")
    _HEADER_TO_FIELD[_actual_header(_label)]  = (_stage, "actual")


def build_tracking_grid_xlsx(records: list[dict]) -> bytes:
    """Build the editable tracking workbook for *records* (as returned by
    ``ProductionTrackingStore.list_all``). Returns xlsx bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _SHEET_TITLE

    ws.cell(1, 1, value=(
        f"生产跟踪表 Production Tracking — 生成日期 {date.today().isoformat()}"
    )).font = Font(bold=True, size=12)
    ws.cell(2, 1, value=(
        "编辑「计划 Planned」和「实际 Actual」日期后上传即可更新（YYYY-MM-DD）。"
        "填写「实际」日期即表示该里程碑已完成。留空的单元格不会清除已有日期。"
        " Edit the Planned / Actual dates and re-upload. A filled Actual date "
        "marks the milestone done; a blank cell never erases a stored date."
    )).font = Font(size=9, italic=True, color="FF808080")

    header_row = 4
    headers = ["PO号 PO No.", "款号 Style", "工厂 Factory"]
    for _stage, label in MILESTONE_STAGES:
        headers.append(_planned_header(label))
        headers.append(_actual_header(label))
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(header_row, ci, value=h)
        cell.fill = PatternFill(start_color="FF1F3864", end_color="FF1F3864",
                                fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
    ws.row_dimensions[header_row].height = 34

    meta_fill    = PatternFill(start_color="FFEEF0F3", end_color="FFEEF0F3",
                               fill_type="solid")   # grey = context, read-only
    planned_fill = PatternFill(start_color="FFEAF3FB", end_color="FFEAF3FB",
                               fill_type="solid")   # light blue = planned
    actual_fill  = PatternFill(start_color="FFEAF7EE", end_color="FFEAF7EE",
                               fill_type="solid")   # light green = actual

    for ri, r in enumerate(records, header_row + 1):
        ws.cell(ri, _COL_PO,      value=r.get("po_number", "")).fill = meta_fill
        ws.cell(ri, _COL_STYLE,   value=r.get("style") or "").fill = meta_fill
        ws.cell(ri, _COL_FACTORY, value=r.get("factory") or "").fill = meta_fill
        ci = _FIRST_DATE_COL
        for stage, _label in MILESTONE_STAGES:
            p = ws.cell(ri, ci,     value=(r.get(f"{stage}_planned") or "") or None)
            p.fill = planned_fill
            a = ws.cell(ri, ci + 1, value=(r.get(f"{stage}_actual") or "") or None)
            a.fill = actual_fill
            ci += 2

    ws.column_dimensions[get_column_letter(_COL_PO)].width = 16
    ws.column_dimensions[get_column_letter(_COL_STYLE)].width = 16
    ws.column_dimensions[get_column_letter(_COL_FACTORY)].width = 22
    for ci in range(_FIRST_DATE_COL, _FIRST_DATE_COL + 2 * len(MILESTONE_STAGES)):
        ws.column_dimensions[get_column_letter(ci)].width = 15
    ws.freeze_panes = ws.cell(header_row + 1, _FIRST_DATE_COL).coordinate

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()



def parse_tracking_grid_xlsx(content: bytes) -> dict:
    """Read a returned tracking workbook.

    Returns ``{"rows": [{"po_number", "style", "fields": {col: val}}],
    "issues": [str], "rows_seen": int}``. ``fields`` is ready for
    ``update_stage_fields``: a filled Actual date also sets ``{stage}_status
    = 'Done'``. Rows with no filled dates are skipped (nothing to do). Raises
    ValueError only for file-level problems (unreadable / header not found).
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:                 # not a readable .xlsx at all
        raise ValueError(f"Not a readable Excel file: {exc}") from exc
    try:
        ws = wb[_SHEET_TITLE] if _SHEET_TITLE in wb.sheetnames else wb.active

        # Locate the header row by ANY cell starting with "PO号", so a header
        # still parses if the user reordered the columns in Excel.
        header_row = None
        for r in range(1, min(ws.max_row, 20) + 1):
            if any(str(ws.cell(r, c).value or "").strip().startswith("PO号")
                   for c in range(1, ws.max_column + 1)):
                header_row = r
                break
        if header_row is None:
            raise ValueError(
                "Header row not found — is this the tracking export? "
                "(expected a 'PO号 PO No.' header cell)"
            )

        # Map every column to its meaning by HEADER TEXT (never fixed index),
        # so reordering PO / Style / date columns in Excel can't scramble it.
        col_po = _COL_PO
        col_style = None
        date_cols: dict[int, tuple[str, str]] = {}
        for c in range(1, ws.max_column + 1):
            text = str(ws.cell(header_row, c).value or "").strip()
            if text.startswith("PO号"):
                col_po = c
            elif text.startswith("款号"):
                col_style = c
            elif text in _HEADER_TO_FIELD:
                date_cols[c] = _HEADER_TO_FIELD[text]

        rows: list[dict] = []
        issues: list[str] = []
        rows_seen = 0
        for r in range(header_row + 1, ws.max_row + 1):
            po = str(ws.cell(r, col_po).value or "").strip()
            if not po:
                continue
            rows_seen += 1
            style = str(ws.cell(r, col_style).value or "").strip() if col_style else ""

            fields: dict = {}
            for c, (stage, kind) in date_cols.items():
                iso = _cell_date(ws.cell(r, c).value)
                if not iso:                       # blank never erases
                    continue
                # Guard against a mistyped date so one bad cell doesn't poison
                # the whole import; report it and move on.
                if not _looks_like_iso(iso):
                    issues.append(
                        f"row {r} ({po} / {style}): "
                        f"{MILESTONE_LABELS.get(stage, stage)} "
                        f"date {iso!r} isn't YYYY-MM-DD — skipped")
                    continue
                fields[f"{stage}_{kind}"] = iso
                if kind == "actual":
                    fields[f"{stage}_status"] = "Done"

            if fields:
                rows.append({"po_number": po, "style": style, "fields": fields})

        return {"rows": rows, "issues": issues, "rows_seen": rows_seen}
    finally:
        wb.close()


def _looks_like_iso(s: str) -> bool:
    try:
        date.fromisoformat(s[:10])
        return True
    except (ValueError, TypeError):
        return False
