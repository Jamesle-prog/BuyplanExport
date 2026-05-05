"""HHP Template_P exporter — fabric-grouped colour × size summary.

Matches the VBA's second output:
  - One workbook per Fabric1_Code value.
  - Within each workbook, one sheet per style (Main Supplier Config SKU).
  - Each sheet: row 2 = size headers (B onward), row 3+ = colour rows.

All fabric workbooks are returned as a list of (filename, bytes) tuples
ready to be zipped and served for download.
"""
from __future__ import annotations

import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..utils.file_utils import versioned_path
from ._excel_helpers import clean_sheet_name, stable_unique

SIZES = ["XS", "S", "M", "L", "XL", "XXL"]

_HDRFILL = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
_TOTFILL = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
_ALTFILL = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")


def _thin():
    s = Side(border_style="thin", color="FFAAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


def _hdr(cell, value):
    cell.value = value
    cell.fill = _HDRFILL
    cell.font = Font(color="FFFFFFFF", bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _thin()


def _data(cell, value, *, alt=False):
    cell.value = value
    if alt:
        cell.fill = _ALTFILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _thin()
    if isinstance(value, (int, float)):
        cell.number_format = "#,##0"


def _total(cell, value):
    cell.value = value
    cell.fill = _TOTFILL
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _thin()
    if isinstance(value, (int, float)):
        cell.number_format = "#,##0"


def export_hhp_template_p(
    df: pd.DataFrame,
    output_dir: str,
) -> list[tuple[str, bytes]]:
    """Generate one workbook per distinct Fabric1_Code and return file data.

    Returns
    -------
    List of (filename, bytes) tuples — one per fabric workbook.
    An extra "ALL" workbook is also produced containing every style
    regardless of fabric (useful when Fabric1_Code is blank).
    """
    fabric_col  = "面料_面料_编号"   # Fabric1_Code internal name
    style_col   = "Main Supplier Config SKU"
    color_col   = "颜色"            # Chinese colour name
    color_en    = "Main Supplier Color Description"

    # Determine the colour column to use (prefer Chinese, fall back to English)
    use_color_col = color_col if color_col in df.columns else color_en

    # Sizes actually present
    sizes_present = [s for s in SIZES if s in df.columns]

    results: list[tuple[str, bytes]] = []

    # Group by fabric code
    has_fabric = fabric_col in df.columns and df[fabric_col].notna().any()
    if has_fabric:
        fabric_groups = [
            (str(fk).strip(), sub)
            for fk, sub in df.groupby(fabric_col, dropna=False)
            if str(fk).strip() not in ("", "nan", "None")
        ]
        # Styles with no fabric code go into an "UNKNOWN" workbook
        no_fabric = df[df[fabric_col].isna() | (df[fabric_col].astype(str).str.strip() == "")]
        if not no_fabric.empty:
            fabric_groups.append(("UNKNOWN", no_fabric))
    else:
        fabric_groups = [("ALL", df)]

    for fabric_key, fab_df in fabric_groups:
        wb = Workbook()
        wb.remove(wb.active)

        styles = stable_unique(
            fab_df[style_col].dropna().astype(str).str.strip()
            if style_col in fab_df.columns else pd.Series(dtype=str)
        )

        for style in styles:
            sub = fab_df[
                fab_df[style_col].astype(str).str.strip() == style
            ].copy()
            if sub.empty:
                continue

            sheet_name = clean_sheet_name(style)
            existing = [s.title for s in wb.worksheets]
            candidate = sheet_name
            sfx = 2
            while candidate in existing:
                candidate = sheet_name[:28] + f"_{sfx}"
                sfx += 1

            ws = wb.create_sheet(candidate)
            _write_style_sheet_p(ws, sub, sizes_present, use_color_col)

        if not wb.sheetnames:
            wb.create_sheet("Empty")

        fname = f"Zalando_面料_{fabric_key}.xlsx"
        buf = io.BytesIO()
        wb.save(buf)
        results.append((fname, buf.getvalue()))

    return results


# ── Per-style sheet ───────────────────────────────────────────────────────────

def _write_style_sheet_p(ws, sub: pd.DataFrame, sizes: list[str], color_col: str):
    n = len(sizes)

    # Row 1: merged "Color" label + "Sizes" merged header
    _hdr(ws.cell(1, 1), "Color")
    for c in range(2, 2 + n):
        _hdr(ws.cell(1, c), "Sizes" if c == 2 else None)
    if n > 1:
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=1 + n)
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    # Row 2: individual size names
    for i, sz in enumerate(sizes):
        _hdr(ws.cell(2, 2 + i), sz)

    # Group by colour, sum size quantities
    pivot = (
        sub.groupby(color_col, dropna=False)[sizes].sum()
        if color_col in sub.columns
        else pd.DataFrame()
    )

    col_totals = [0] * n
    row_idx = 3
    for alt_idx, (color, row) in enumerate(pivot.iterrows()):
        alt = alt_idx % 2 == 1
        # BUG-23: pandas NaN key becomes "nan" when passed through str(); use empty string instead
        color_str = "" if pd.isna(color) else str(color).strip()
        _data(ws.cell(row_idx, 1), color_str, alt=alt)
        for i, sz in enumerate(sizes):
            v = int(row[sz])
            _data(ws.cell(row_idx, 2 + i), v, alt=alt)
            col_totals[i] += v
        row_idx += 1

    # Total row
    _total(ws.cell(row_idx, 1), "Total")
    for i, v in enumerate(col_totals):
        _total(ws.cell(row_idx, 2 + i), v)

    # Column widths
    ws.column_dimensions["A"].width = max(
        12, max((len(str(c)) for c in pivot.index), default=6) + 3
    )
    for i in range(n):
        ws.column_dimensions[get_column_letter(2 + i)].width = max(7, len(sizes[i]) + 3)


# ── Helpers ───────────────────────────────────────────────────────────────────

