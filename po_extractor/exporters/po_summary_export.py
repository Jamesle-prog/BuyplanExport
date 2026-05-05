"""PO Summary export — two tabs:
  1. 'PO Summary'    : one row per (Style, PO Number, Color) + sizes + metadata
  2. 'Style-Color'   : one row per (Style, Color) summed across all POs, no metadata
"""
from datetime import datetime, timedelta

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..utils.file_utils import versioned_path
from ..utils.size_config import get_size_order


def _xfactory_date(xport: str | None) -> str | None:
    if not xport:
        return None
    try:
        return (datetime.strptime(str(xport), "%Y-%m-%d") - timedelta(days=10)).strftime("%Y-%m-%d")
    except ValueError:
        return None


def _thin():
    s = Side(border_style="thin", color="FF000000")
    return Border(left=s, right=s, top=s, bottom=s)


def _style_header(cell, value):
    cell.value = value
    cell.fill = PatternFill(start_color="FF000000", end_color="FF000000", fill_type="solid")
    cell.font = Font(color="FFFFFFFF", bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _thin()


def _style_data(cell, value, *, yellow=False):
    cell.value = value
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _thin()
    if yellow:
        cell.fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
        cell.font = Font(bold=True)
    if isinstance(value, (int, float)):
        cell.number_format = "#,##0"


def _auto_widths(ws, headers, flat):
    for c_idx, h in enumerate(headers, 1):
        col_vals = flat[h].dropna().astype(str) if h in flat.columns else pd.Series(dtype=str)
        body_max = max((len(v) for v in col_vals), default=0)
        width = max(len(str(h)), body_max) + 3
        ws.column_dimensions[get_column_letter(c_idx)].width = min(width, 42)


def _write_sheet(ws, flat: pd.DataFrame, headers: list[str], total_col_name: str = "Total"):
    total_col = headers.index(total_col_name) + 1 if total_col_name in headers else None
    for c_idx, h in enumerate(headers, 1):
        _style_header(ws.cell(row=1, column=c_idx), h)
    for r_idx, (_, row) in enumerate(flat.iterrows(), start=2):
        for c_idx, h in enumerate(headers, 1):
            val = row[h]
            if pd.isna(val):
                val = None
            elif isinstance(val, (int, float)) and float(val).is_integer():
                val = int(val)
            _style_data(
                ws.cell(row=r_idx, column=c_idx),
                val,
                yellow=(total_col is not None and c_idx == total_col),
            )
    _auto_widths(ws, headers, flat)


def _size_cols(df: pd.DataFrame) -> list[str]:
    size_order = get_size_order()
    present = set(df.columns)
    known   = [s for s in size_order if s in present]
    unknown = [s for s in present if s not in size_order]
    return known + unknown


def export_po_summary(df_size: pd.DataFrame, df_meta: pd.DataFrame, output_dir: str) -> str:
    path = versioned_path(output_dir, "po_summary", ".xlsx")
    wb = Workbook()

    has_cn = "Color (CN)" in df_size.columns

    # ------------------------------------------------------------------ #
    # Tab 1: PO Summary — (Style, PO Number, Color) × Size               #
    # ------------------------------------------------------------------ #
    idx1 = ["Style", "PO Number", "Color"] + (["Color (CN)"] if has_cn else [])
    pivot1 = df_size.pivot_table(
        index=idx1, columns="Size",
        values="Units", aggfunc="sum", fill_value=0,
    )
    sc1 = _size_cols(pivot1)
    pivot1 = pivot1.reindex(columns=sc1)
    pivot1["Total"] = pivot1.sum(axis=1)
    flat1 = pivot1.reset_index()
    flat1.columns.name = None

    # Attach metadata
    meta_field_map = {
        "po_number":         "PO Number",
        "company":           "Company",
        "factory":           "Factory",
        "country_of_origin": "COO",
        "xport_date":        "X-Port Date",
        "division_code":     "Division Code",
        "division_name":     "Division Name",
        "po_date":           "Issue Date",
        "issue_date":        "Issue Date",
        "version":           "Version",
        "extracted_at":      "Extracted At",
    }
    available = {k: v for k, v in meta_field_map.items() if k in df_meta.columns}
    if available:
        meta_slim = (
            df_meta[list(available.keys())]
            .copy()
            .rename(columns=available)
            .drop_duplicates(subset=["PO Number"])
        )
        if "X-Port Date" in meta_slim.columns:
            meta_slim["X-Factory Date"] = meta_slim["X-Port Date"].apply(_xfactory_date)
        flat1 = flat1.merge(meta_slim, on="PO Number", how="left")

    flat1 = flat1.sort_values(["Style", "PO Number", "Color"]).reset_index(drop=True)
    # Ensure Company column exists even if not in metadata
    if "Company" not in flat1.columns:
        flat1["Company"] = None

    extra = ["Total", "Company", "Factory", "COO", "X-Port Date", "X-Factory Date",
             "Division Code", "Division Name", "Issue Date", "Version", "Extracted At"]
    base1 = ["Company", "Style", "PO Number", "Color"] + (["Color (CN)"] if has_cn else [])
    headers1 = (
        base1
        + sc1
        + [c for c in extra if c in flat1.columns and c not in base1 + sc1]
        + [c for c in flat1.columns if c not in base1 + sc1 + extra]
    )
    flat1 = flat1[headers1]

    ws1 = wb.active
    ws1.title = "PO Summary"
    _write_sheet(ws1, flat1, headers1)

    # ------------------------------------------------------------------ #
    # Tab 2: Style-Color — (Style, Color) × Size summed across all POs   #
    # ------------------------------------------------------------------ #
    idx2 = ["Style", "Color"] + (["Color (CN)"] if has_cn else [])
    pivot2 = df_size.pivot_table(
        index=idx2, columns="Size",
        values="Units", aggfunc="sum", fill_value=0,
    )
    sc2 = _size_cols(pivot2)
    pivot2 = pivot2.reindex(columns=sc2)
    pivot2["Total"] = pivot2.sum(axis=1)
    flat2 = pivot2.reset_index()
    flat2.columns.name = None
    flat2 = flat2.sort_values(["Style", "Color"]).reset_index(drop=True)

    headers2 = idx2 + sc2 + ["Total"]
    flat2 = flat2[headers2]

    ws2 = wb.create_sheet("Style-Color")
    _write_sheet(ws2, flat2, headers2)

    wb.save(path)
    return path
