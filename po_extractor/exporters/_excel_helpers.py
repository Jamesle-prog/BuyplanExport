"""Shared Excel-export utilities.

Single source of truth for small helpers that were previously duplicated
across several exporters (sheet-name cleanup, stable de-duplication, cell
value extraction).
"""
from __future__ import annotations

from functools import lru_cache
import re
from typing import Any, Collection, Iterable

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties


# ---------------------------------------------------------------------------
# Page / print settings
# ---------------------------------------------------------------------------

def apply_print_settings(wb) -> None:
    """Apply A4 landscape "fit all columns on one page" print settings to every
    sheet in *wb*.

    Settings match the Excel Page Setup screenshot:
      • 横向    — Landscape orientation
      • A4      — 21 cm × 29.7 cm
      • 将所有列调整为一页 — fitToWidth=1, fitToHeight=0 (unlimited rows)

    Call this just before ``wb.save()`` so it applies to all sheets including
    any Index sheet and per-style copies.

    Implementation note
    -------------------
    openpyxl's ``page_setup.fitToPage`` is not a real attribute and has no
    effect.  The "fit to page" scaling mode is controlled by two separate XML
    elements:
      • ``<sheetPr><pageSetUpPr fitToPage="1"/>`` — switches Excel from
        percentage scaling to fit-to-page mode (set via
        ``ws.sheet_properties.pageSetUpPr``).
      • ``<pageSetup fitToWidth="1" fitToHeight="0"/>`` — the actual page
        counts (set via ``ws.page_setup``).
    Both must be present; missing either one leaves Excel using the default
    percentage scaling and the fit-to-page settings are silently ignored.
    """
    for ws in wb.worksheets:
        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize   = 9     # A4  (openpyxl PAPERSIZE_A4)
        ws.page_setup.fitToWidth  = 1     # all columns on one page
        ws.page_setup.fitToHeight = 0     # unlimited pages tall
        # Switch Excel to fit-to-page scaling mode.  Mutate in place when a
        # pageSetUpPr already exists in the template so other attributes
        # (e.g. autoPageBreaks) are preserved instead of silently lost.
        if ws.sheet_properties.pageSetUpPr is None:
            ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        else:
            ws.sheet_properties.pageSetUpPr.fitToPage = True
        # Margins (all values in inches):
        #   上/下 1.91 cm = 0.75 in  |  左/右 0.64 cm = 0.25 in
        #   页眉/页脚 0.76 cm = 0.30 in
        ws.page_margins = PageMargins(
            top=0.75, bottom=0.75,
            left=0.25, right=0.25,
            header=0.30, footer=0.30,
        )


# ---------------------------------------------------------------------------
# Sheet names
# ---------------------------------------------------------------------------

# Excel forbids these characters in sheet names: / \ [ ] * ? :
# Apostrophe is technically allowed but breaks formula references
# like 'sheet'!A1, so we strip it too (BUG-41 mitigation).
_ILLEGAL_SHEET_CHARS = r"/\[]*?:'"


def unique_sheet_name(base: str, taken: Collection[str]) -> str:
    """*base* (already a clean ≤31-char sheet name) or, when it is in
    *taken*, ``base_2``, ``base_3`` … — the suffix always fits inside
    Excel's 31-character limit, however many digits it grows to.

    The caller records the returned name in *taken* itself (a workbook's
    ``sheetnames`` does that on ``create_sheet``).
    """
    name, sfx = base, 2
    while name in taken:
        tail = f"_{sfx}"
        name = f"{base[:31 - len(tail)]}{tail}"
        sfx += 1
    return name


_UNSAFE_FILENAME_RE = re.compile(r'[<>:"/\\|?*\s]+')


def safe_filename(name: str | None, *, fallback: str = "unknown") -> str:
    """A file-name-safe stem: Windows-illegal characters and whitespace runs
    become ``_``; leading/trailing ``_`` trimmed; *fallback* when nothing is
    left."""
    return _UNSAFE_FILENAME_RE.sub("_", str(name or "")).strip("_") or fallback


def clean_sheet_name(name: str | None, *, fallback: str = "Sheet") -> str:
    """Return a valid Excel sheet name (≤31 chars, no illegal characters).

    Replaces every occurrence of ``/ \\ [ ] * ? :`` and apostrophe with ``_``
    and truncates to 31 characters.  Empty / None input returns *fallback*.
    """
    s = (name or "").strip()
    if not s:
        return fallback
    for ch in _ILLEGAL_SHEET_CHARS:
        s = s.replace(ch, "_")
    return s[:31] or fallback


# ---------------------------------------------------------------------------
# Internal (same-workbook) hyperlinks
# ---------------------------------------------------------------------------

def set_internal_hyperlink(cell, sheet_name: str, anchor: str = "A1") -> None:
    """Point *cell* at ``anchor`` on *sheet_name* within the same workbook.

    Do NOT do this with ``cell.hyperlink = f"#'{sheet_name}'!{anchor}"`` (a
    plain string) — openpyxl always writes a plain-string hyperlink as an
    *external* relationship (``TargetMode="External"``) with that string as
    the literal ``Target``.  Excel special-cases a target starting with
    ``"#"`` and follows it as an internal jump anyway, but that's an
    Excel-only leniency, not part of the OOXML spec — WPS (and other
    readers) take the target literally and the link does nothing.

    The portable, spec-correct way to express a same-workbook link is the
    hyperlink's ``location`` attribute with no relationship at all — pass an
    explicit ``Hyperlink(location=..., target=None)`` object instead of a
    string, which is what this helper does.
    """
    from openpyxl.worksheet.hyperlink import Hyperlink

    cell.hyperlink = Hyperlink(
        ref=cell.coordinate, location=f"'{sheet_name}'!{anchor}", target=None,
    )


# ---------------------------------------------------------------------------
# Sequence helpers
# ---------------------------------------------------------------------------

def stable_unique(values: Iterable[Any]) -> list:
    """Return values in first-seen order, with duplicates removed.

    Works on any iterable (lists, pandas Series, generators).  Values must
    be hashable.
    """
    seen: set = set()
    out: list = []
    for v in values:
        if v not in seen:
            seen.add(v)
            out.append(v)
    return out


# ---------------------------------------------------------------------------
# Cell value extraction
# ---------------------------------------------------------------------------

_NULLISH = {"", "nan", "none", "null"}


def cell_value(row, col: str) -> str | None:
    """Return a stripped string from a pandas Series / dict-like *row*.

    Returns None when the cell is empty, missing, NaN, or string "nan".
    """
    v = row.get(col) if hasattr(row, "get") else None
    if v is None:
        try:
            v = row[col]
        except (KeyError, TypeError, IndexError):
            return None
    if v is None:
        return None
    s = str(v).strip()
    return s if s and s.lower() not in _NULLISH else None


# ---------------------------------------------------------------------------
# Cell styling
# ---------------------------------------------------------------------------
# openpyxl style objects are immutable value objects (assigning one to a cell
# stores a proxy), so one shared Border / Fill per colour is safe and avoids
# rebuilding four Side objects for every cell of a large export.

@lru_cache(maxsize=None)
def thin_border(color: str = "FF000000") -> Border:
    """A thin border on all four sides in *color* (aRGB or RGB hex)."""
    s = Side(border_style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


@lru_cache(maxsize=None)
def solid_fill(color: str) -> PatternFill:
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


_ALIGN_CENTER = Alignment(horizontal="center", vertical="center")


def style_header(cell, value, *, fill: str = "FF000000",
                 font_color: str = "FFFFFFFF", border_color: str = "FF000000") -> None:
    """Bold, centred, filled header cell with a thin border."""
    cell.value = value
    cell.fill = solid_fill(fill)
    cell.font = Font(color=font_color, bold=True)
    cell.alignment = _ALIGN_CENTER
    cell.border = thin_border(border_color)


def style_total(cell, value, *, fill: str = "FFFFFF00",
                font_color: str = "FF000000", border_color: str = "FF000000") -> None:
    """Bold, centred total cell (yellow by default); numbers get ``#,##0``."""
    cell.value = value
    cell.fill = solid_fill(fill)
    cell.font = Font(color=font_color, bold=True)
    cell.alignment = _ALIGN_CENTER
    cell.border = thin_border(border_color)
    if isinstance(value, (int, float)):
        cell.number_format = "#,##0"


def style_data(cell, value, *, border_color: str = "FF000000") -> None:
    """Centred data cell with a thin border; numbers get ``#,##0``."""
    cell.value = value
    cell.alignment = _ALIGN_CENTER
    cell.border = thin_border(border_color)
    if isinstance(value, (int, float)):
        cell.number_format = "#,##0"


def write_cell(ws, r: int, c: int, v, *, bold: bool = False, bg: str | None = None,
               white: bool = False, center: bool = False, wrap: bool = True,
               num: str | None = None, border: Border | None = None,
               font_name: str = "Arial", size: int = 10):
    """Write *v* at (*r*, *c*) in the GIII document style (Arial 10, wrapped,
    optional solid *bg*, white text when *white*) and return the cell.

    Bind the per-workbook defaults once with ``functools.partial`` —
    ``cell = partial(write_cell, ws, border=thin_border(), center=True)``.
    """
    cl = ws.cell(r, c, v)
    cl.font = Font(name=font_name, size=size, bold=bold,
                   color="FFFFFFFF" if white else "FF000000")
    if bg:
        cl.fill = PatternFill("solid", fgColor=bg)
    cl.alignment = Alignment(horizontal="center" if center else "left",
                             vertical="center", wrap_text=wrap)
    if border is not None:
        cl.border = border
    if num:
        cl.number_format = num
    return cl


# ---------------------------------------------------------------------------
# Template sheets
# ---------------------------------------------------------------------------

def replace_placeholders(ws, values: dict) -> None:
    """Substitute ``{{key}}`` in every string cell of *ws*."""
    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            v = cell.value
            for key, val in values.items():
                v = v.replace(f"{{{{{key}}}}}", str(val or ""))
            cell.value = v


def clear_data_area(ws, start_row: int) -> None:
    """Unmerge and blank every cell from *start_row* downward."""
    to_unmerge = [str(r) for r in ws.merged_cells.ranges if r.min_row >= start_row]
    for r in to_unmerge:
        ws.unmerge_cells(r)
    for row in ws.iter_rows(min_row=start_row):
        for cell in row:
            cell.value = None
