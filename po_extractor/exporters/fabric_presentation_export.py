"""Exporter for fabric presentation sheets (面料推荐单 / HHN Presentation).

Reproduces the hand-maintained ``HHN Presentation`` workbook: a title block
(submission date / type / customer), one row per recommended fabric, and the
type legend at the foot.

Two things the hand-made sheet could not do are handled here:

* **Price mode** — which price column(s) to print is chosen at export time.
  The internal RMB/M cost must never reach a customer copy by accident, so
  it is opt-in per export rather than a column that is always present.
* **QR code** — each sheet carries its own, encoding the presentation's
  scan URL.  Scanning records when the sheet went out and what was on it.
"""
from __future__ import annotations

import io
import warnings

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.utils import get_column_letter

from ._excel_helpers import apply_print_settings, clean_sheet_name, thin_border
from ..utils.qr import QRUnavailable, qr_png, scan_url

# Price modes.  'usd' is the customer-facing default; 'rmb' and 'both'
# expose internal cost and are for review copies only.
PRICE_USD = "usd"
PRICE_RMB = "rmb"
PRICE_BOTH = "both"
PRICE_MODES = (PRICE_USD, PRICE_RMB, PRICE_BOTH)

PRICE_MODE_LABELS = {
    PRICE_USD:  "USD/Y only (customer copy)",
    PRICE_RMB:  "RMB/M only (internal cost)",
    PRICE_BOTH: "Both — internal review copy",
}

_HDR_FILL = "FFD9D9D9"
_TITLE_FILL = "FF1F3864"

# Columns that are always printed, in order: (heading, line-dict key, width).
_BASE_COLS: list[tuple[str, str, int]] = [
    ("NO.",               "line_no",     6),
    ("Style",             "style",       12),
    ("Season",            "season",      10),
    ("HHN_Fabric#",       "quality_no",  20),
    ("Content",           "content",     38),
    ("Fabric Description", "description", 18),
    ("Weight（gsm）",      "weight_gsm",  12),
    ("Width (inch)",      "width_in",    12),
    ("MOQ(Y)",            "moq_y",       10),
    ("MCQ(Y)",            "mcq_y",       10),
]

_LEGEND = [
    "New fabric (Client-Initiated): Fabric submissions based on client "
    "provided sample/inspiration or other requests",
    "New fabric (HHN-Initiated): Fabric recommended by HHN without client's request",
]


def _thin() -> Border:
    return thin_border("FF000000")


def _price_columns(price_mode: str) -> list[tuple[str, str, int]]:
    """The price column(s) for *price_mode*, as (heading, key, width)."""
    if price_mode == PRICE_RMB:
        return [("Price (RMB/M)\nOver MOQ", "price_rmb_m", 14)]
    if price_mode == PRICE_BOTH:
        return [("Price (RMB/M)\nOver MOQ", "price_rmb_m", 14),
                ("USD/Y", "price_usd_y", 10)]
    return [("USD/Y", "price_usd_y", 10)]


def build_presentation_workbook(presentation: dict, lines: list[dict], *,
                                price_mode: str = PRICE_USD,
                                scan_base_url: str = "") -> bytes:
    """Return the .xlsx bytes for one presentation.

    *scan_base_url* is the web_scan service's base URL (e.g.
    ``http://192.168.0.153:8502``).  When it is empty the QR code is left
    out rather than encoding a URL that resolves nowhere.
    """
    if price_mode not in PRICE_MODES:
        raise ValueError(f"price_mode must be one of {PRICE_MODES}, got {price_mode!r}")

    columns = _BASE_COLS + _price_columns(price_mode)
    n_cols = len(columns)

    wb = Workbook()
    ws = wb.active
    ws.title = clean_sheet_name(presentation.get("title") or "Presentation")

    # ── Title block ─────────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=n_cols)
    tc = ws.cell(1, 1, "HIGH HOPE NEWEST")
    tc.font = Font(bold=True, size=18, color="FFFFFFFF")
    tc.fill = PatternFill("solid", fgColor=_TITLE_FILL)
    tc.alignment = Alignment(horizontal="center", vertical="center")

    meta = [
        ("Submission date：", presentation.get("submission_date") or ""),
        ("Type:",            presentation.get("fabric_type") or ""),
        ("Customer:",        presentation.get("customer") or ""),
    ]
    for i, (label, value) in enumerate(meta, start=3):
        lc = ws.cell(i, 1, label)
        lc.font = Font(bold=True)
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)
        ws.cell(i, 2, value)

    # ── Header row ──────────────────────────────────────────────────────────
    hdr_row = 7
    for ci, (heading, _key, width) in enumerate(columns, 1):
        c = ws.cell(hdr_row, ci, heading)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor=_HDR_FILL)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = _thin()
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[hdr_row].height = 30

    # ── Data rows ───────────────────────────────────────────────────────────
    r = hdr_row + 1
    for i, line in enumerate(lines, 1):
        for ci, (_heading, key, _w) in enumerate(columns, 1):
            value = line.get(key)
            if key == "line_no":
                value = line.get("line_no") or i
            # A fabric with no price prints blank, not 0 — "we did not quote
            # this" and "we quoted nothing" are different claims.
            c = ws.cell(r, ci, value if value not in ("", None) else None)
            c.border = _thin()
            if key in ("content", "description"):
                c.alignment = Alignment(horizontal="left", vertical="center",
                                        wrap_text=True)
            else:
                c.alignment = Alignment(horizontal="center", vertical="center")
            if key == "price_usd_y" and value not in ("", None):
                c.number_format = "0.00"
            elif key == "price_rmb_m" and value not in ("", None):
                c.number_format = "0.00"
        r += 1

    # ── Legend ──────────────────────────────────────────────────────────────
    r += 1
    for text in _LEGEND:
        lc = ws.cell(r, 1, text)
        lc.font = Font(size=9, italic=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
        r += 1

    # ── QR code ─────────────────────────────────────────────────────────────
    token = presentation.get("token") or ""
    if token and scan_base_url:
        _add_qr(ws, scan_base_url, token, anchor_col=n_cols + 2)

    apply_print_settings(wb)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _add_qr(ws, scan_base_url: str, token: str, *, anchor_col: int) -> None:
    """Place the sheet's QR code and its caption to the right of the table.

    A QR failure must not cost the user their export — the sheet is still
    correct without it, so the error is surfaced as a warning and the
    workbook is returned.
    """
    url = scan_url(scan_base_url, token)
    try:
        png = qr_png(url, scale=4)
    except (QRUnavailable, ValueError) as exc:
        warnings.warn(f"[fabric presentation] QR code omitted — {exc}")
        return

    col = get_column_letter(anchor_col)
    img = XLImage(io.BytesIO(png))
    ws.add_image(img, f"{col}3")
    cap = ws.cell(1, anchor_col, f"Sheet ID: {token}")
    cap.font = Font(size=9, bold=True)
    ws.column_dimensions[col].width = 22
