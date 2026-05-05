"""Cross-check export — verifies unit totals match across buy plan, color plan, PO summary."""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..utils.file_utils import versioned_path


def _thin():
    s = Side(border_style="thin", color="FF000000")
    return Border(left=s, right=s, top=s, bottom=s)


def _header(cell, value):
    cell.value = value
    cell.fill = PatternFill(start_color="FF000000", end_color="FF000000", fill_type="solid")
    cell.font = Font(color="FFFFFFFF", bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _thin()


def _buyplan_totals(path: str) -> dict:
    """Return {style: total_units} by summing the 'Total' column in each sheet's data rows.

    Buy plan layout: info rows 1-4, header row 5, data rows 6..N-1, total row N.
    We sum the last column of data rows (excludes the total row itself).
    """
    wb = load_workbook(path, data_only=True)
    out = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        # Find Total column in row 5
        total_col = None
        for cell in ws[5]:
            if cell.value == "Total":
                total_col = cell.column
                break
        if not total_col:
            continue
        # Data rows: row 6 to max_row - 1 (last row is the in-sheet Total row)
        s = 0
        for r in range(6, ws.max_row):
            v = ws.cell(row=r, column=total_col).value
            if isinstance(v, (int, float)):
                s += int(v)
        out[sheet] = s
    return out


def _colorplan_totals(path: str) -> dict:
    """Return {style: total_units} by summing all numeric cells from row 3 onward."""
    wb = load_workbook(path, data_only=True)
    out = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        s = 0
        for r in range(3, ws.max_row + 1):
            for c in range(2, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, (int, float)):
                    s += int(v)
        out[sheet] = s
    return out


def _po_summary_totals(path: str) -> dict:
    """Return {style: total_units} by summing the 'Total' column grouped by Style."""
    wb = load_workbook(path, data_only=True)
    ws = wb["PO Summary"] if "PO Summary" in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    try:
        style_col = headers.index("Style") + 1
        total_col = headers.index("Total") + 1
    except ValueError:
        return {}
    out: dict = {}
    for r in range(2, ws.max_row + 1):
        style = ws.cell(row=r, column=style_col).value
        v = ws.cell(row=r, column=total_col).value
        if style and isinstance(v, (int, float)):
            out[style] = out.get(style, 0) + int(v)
    return out


def export_cross_check(df_size: pd.DataFrame, buyplan_path: str,
                       color_plan_path: str, po_summary_path: str,
                       output_dir: str) -> str:
    path = versioned_path(output_dir, "cross_check", ".xlsx")

    # Source-of-truth totals per style from the raw extracted size rows
    source = df_size.groupby("Style")["Units"].sum().astype(int).to_dict()
    bp = _buyplan_totals(buyplan_path)
    cp = _colorplan_totals(color_plan_path)
    ps = _po_summary_totals(po_summary_path)

    styles = sorted(set(source) | set(bp) | set(cp) | set(ps))

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Cross Check"

    headers = ["Style", "Source (df_size)", "Buy Plan", "Color Plan",
               "PO Summary", "Match?"]
    for i, h in enumerate(headers, 1):
        _header(ws.cell(row=1, column=i), h)

    border = _thin()
    good_fill = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
    bad_fill = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")

    row = 2
    src_total = bp_total = cp_total = ps_total = 0
    all_ok = True
    for style in styles:
        src = source.get(style, 0)
        b = bp.get(style[:31], bp.get(style, 0))
        c = cp.get(style[:31], cp.get(style, 0))
        p = ps.get(style, 0)
        match = (src == b == c == p)
        if not match:
            all_ok = False

        ws.cell(row=row, column=1, value=style)
        ws.cell(row=row, column=2, value=src)
        ws.cell(row=row, column=3, value=b)
        ws.cell(row=row, column=4, value=c)
        ws.cell(row=row, column=5, value=p)
        ws.cell(row=row, column=6, value="OK" if match else "MISMATCH")

        fill = good_fill if match else bad_fill
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = fill
            if col in (2, 3, 4, 5) and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"

        src_total += src; bp_total += b; cp_total += c; ps_total += p
        row += 1

    # Grand total row
    totals = [src_total, bp_total, cp_total, ps_total]
    match = len(set(totals)) == 1
    ws.cell(row=row, column=1, value="TOTAL")
    ws.cell(row=row, column=2, value=src_total)
    ws.cell(row=row, column=3, value=bp_total)
    ws.cell(row=row, column=4, value=cp_total)
    ws.cell(row=row, column=5, value=ps_total)
    ws.cell(row=row, column=6, value="OK" if match and all_ok else "MISMATCH")
    for col in range(1, 7):
        cell = ws.cell(row=row, column=col)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
        if col in (2, 3, 4, 5):
            cell.number_format = "#,##0"

    # Column widths
    widths = [24, 16, 12, 12, 14, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(path)
    return path
