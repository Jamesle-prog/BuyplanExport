"""Buy-plan Excel export (one sheet per style).

Templates are stored per-client in ``data/buyplan_templates/``.

Lookup order for a given *company*:
  1. ``data/buyplan_templates/{company}.xlsx``   (client-specific)
  2. ``data/buyplan_templates/default.xlsx``     (shared fallback)
  3. ``data/buyplan_template.xlsx``              (legacy single-file fallback)
  4. Built-in format (no template)

Template conventions
--------------------
* The **first sheet** is used as the per-style master.
* Any cell containing ``{{key}}`` is replaced at run time.
  Supported keys: factory, style, xfactory_date, xport_date, coo,
  division, created_at
* Place ``{{data_start}}`` in one cell to mark the first row of the data
  table (headers → rows → grand total).  If absent, the ``header_row``
  from ``{name}_config.json`` (default 5) is used.

Implementation detail
---------------------
Private helpers live in _buyplan_helpers.py to keep this file focused on the
public API. Import them via star-import so existing callers of private symbols
(e.g. ``from .buyplan_export import _load_template``) continue to work.
"""
from __future__ import annotations

import warnings
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook

from ..utils.file_utils import versioned_path
from ..utils.size_config import get_size_order

from ._buyplan_helpers import *  # noqa: F401,F403  (re-export all private helpers)
from ._buyplan_helpers import (
    _safe_name, _TEMPLATES_DIR,
    _load_template, _build_po_meta_cache, _lookup_meta,
    _replace_placeholders, _clear_data_area,
    _write_data_table, _write_mapped_rows, _set_col_widths,
    _build_default_sheet, _xfactory_date,
    _resolve_template_path,
)
from ._excel_helpers import clean_sheet_name
from ._image_inject import inject_style_photos
from ._photo_utils import load_photo_from_disk


# ---------------------------------------------------------------------------
# Public template-management API
# ---------------------------------------------------------------------------

def list_client_templates() -> list[dict]:
    """Return info about every installed client template.

    Each entry: {"client": str, "file": str, "size_bytes": int}
    """
    _TEMPLATES_DIR.mkdir(parents=True, exist_ok=True)
    result = []
    for p in sorted(_TEMPLATES_DIR.glob("*.xlsx")):
        name = p.stem
        result.append({
            "client":     name,
            "file":       p.name,
            "size_bytes": p.stat().st_size,
        })
    return result


def save_client_template(company: str, xlsx_bytes: bytes,
                         header_row: int = 5,
                         extra_config: dict | None = None) -> Path:
    """Save *xlsx_bytes* as the template for *company* and write its config.

    *extra_config* can contain ``write_headers``, ``column_map``,
    ``size_column_map``, ``meta_column_map`` etc.
    """
    import json
    _TEMPLATES_DIR.mkdir(parents=True, exist_ok=True)
    stem = _safe_name(company)
    tpl  = _TEMPLATES_DIR / f"{stem}.xlsx"
    cfg  = _TEMPLATES_DIR / f"{stem}_config.json"
    tpl.write_bytes(xlsx_bytes)
    cfg_data = {"header_row": header_row}
    if extra_config:
        cfg_data.update(extra_config)
    cfg.write_text(json.dumps(cfg_data, indent=2, ensure_ascii=False), encoding="utf-8")
    return tpl


def delete_client_template(company: str) -> None:
    """Remove the template (and its config) for *company*."""
    stem = _safe_name(company)
    for suffix in (".xlsx", "_config.json"):
        p = _TEMPLATES_DIR / f"{stem}{suffix}"
        if p.exists():
            p.unlink()


# ---------------------------------------------------------------------------
# Public export entry point
# ---------------------------------------------------------------------------

def export_buyplan(
    data: pd.DataFrame,
    metadata: pd.DataFrame,
    output_dir: str,
    images_dir: str = "",
) -> str:
    path = versioned_path(
        output_dir,
        "transformed_data_by_style_filtered_with_totals_and_metadata",
        ".xlsx",
    )

    has_cn     = "Color (CN)" in data.columns
    size_order = get_size_order()
    idx_cols   = ["PO Number", "Style", "Color"] + (["Color (CN)"] if has_cn else [])
    created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Detect client company from metadata (use the most common value)
    company: str | None = None
    if not metadata.empty and "company" in metadata.columns:
        vals = metadata["company"].dropna()
        if not vals.empty:
            company = str(vals.mode().iloc[0]).strip() or None

    template_wb, template_ws, tpl_cfg = _load_template(company)
    use_template  = template_wb is not None and template_ws is not None
    header_row    = tpl_cfg["header_row"]
    use_col_map   = bool(tpl_cfg.get("column_map") or tpl_cfg.get("size_column_map"))

    meta_cache = _build_po_meta_cache(metadata)

    wb = template_wb if use_template else Workbook()
    if not use_template:
        wb.remove(wb.active)

    sheet_style_map: dict[str, str] = {}   # sheet_title → original style (for photo lookup)

    for style in data["Style"].unique():
        sub = data[data["Style"] == style]

        if use_col_map:
            pivot_idx = ["PO Number", "Style", "Color"] + (["Color (CN)"] if has_cn else [])
        else:
            pivot_idx = idx_cols

        pivot = sub.pivot_table(
            index=pivot_idx, columns="Size",
            values="Units", aggfunc="sum", fill_value=0,
        )
        pivot = pivot.loc[:, (pivot != 0).any(axis=0)]
        known     = [s for s in size_order if s in pivot.columns]
        unknown   = [s for s in pivot.columns if s not in size_order]
        size_cols = known + unknown
        pivot     = pivot.reindex(columns=size_cols)
        flat      = pivot.reset_index()
        flat.columns.name = None

        sheet_name = clean_sheet_name(style)   # BUG-42: illegal chars + ≤31
        sheet_style_map[sheet_name] = style
        po_numbers = flat["PO Number"].astype(str).unique().tolist()
        m = _lookup_meta(meta_cache, po_numbers)
        factory = m.get("factory") or "N/A"

        if use_template:
            try:
                ws = wb.copy_worksheet(template_ws)
                ws.title = sheet_name
                _replace_placeholders(ws, {
                    "factory":       factory,
                    "style":         style or "",
                    "xfactory_date": _xfactory_date(m.get("xport_date")) or "N/A",
                    "xport_date":    m.get("xport_date") or "N/A",
                    "coo":           m.get("country_of_origin") or "N/A",
                    "division":      (
                        f"{m.get('division_code', '')} {m.get('division_name', '')}"
                    ).strip(),
                    "created_at":    created_at,
                })
                _clear_data_area(ws, header_row)

                if use_col_map:
                    _write_mapped_rows(ws, flat, size_cols, tpl_cfg, m)
                else:
                    _write_data_table(ws, flat, pivot_idx, size_cols, header_row)
                # Do NOT call _set_col_widths here — the template already has
                # its column widths precisely tuned; auto-sizing would break them.
            except Exception as exc:
                warnings.warn(
                    f"Template copy failed for style '{style}': {exc}. Falling back."
                )
                ws = wb.create_sheet(sheet_name)
                _build_default_sheet(ws, flat, idx_cols, size_cols, m, created_at)
        else:
            ws = wb.create_sheet(sheet_name)
            _build_default_sheet(ws, flat, idx_cols, size_cols, m, created_at)

    if use_template and template_ws in wb.worksheets:
        wb.remove(template_ws)

    if not wb.sheetnames:
        wb.create_sheet("Empty")

    wb.save(str(path))

    # Inject style photos (front / back) into each sheet.
    # openpyxl copy_worksheet() does not carry drawings — we patch the zip.
    if use_template:
        photo_regions = tpl_cfg.get("photo_regions")   # None → use defaults
        sheet_photo_map: dict[str, dict[str, bytes | None]] = {
            sheet_name: {
                "front": load_photo_from_disk(images_dir, style, "front"),
                "back":  load_photo_from_disk(images_dir, style, "back"),
            }
            for sheet_name, style in sheet_style_map.items()
        }
        inject_style_photos(
            output_path=path,
            sheet_photo_map=sheet_photo_map,
            photo_regions=photo_regions,
        )

    return path
