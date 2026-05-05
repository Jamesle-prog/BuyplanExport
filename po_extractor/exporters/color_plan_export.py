"""Color plan export — one sheet per style, Color × Size pivot summed across all POs."""
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..utils.file_utils import versioned_path
from ..utils.size_config import get_size_order
from ._excel_helpers import clean_sheet_name


def _thin_border():
    s = Side(border_style="thin", color="FF000000")
    return Border(left=s, right=s, top=s, bottom=s)


def _style_header(cell, value):
    cell.value = value
    cell.fill = PatternFill(start_color="FF000000", end_color="FF000000", fill_type="solid")
    cell.font = Font(color="FFFFFFFF", bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _thin_border()


def export_color_plan(df_size: pd.DataFrame, output_dir: str) -> str:
    path = versioned_path(output_dir, "color_plan_by_style", ".xlsx")

    # Build en→cn mapping if Color (CN) column is present
    has_cn = "Color (CN)" in df_size.columns
    cn_map: dict = {}
    if has_cn:
        cn_map = (
            df_size[["Color", "Color (CN)"]]
            .drop_duplicates(subset=["Color"])
            .set_index("Color")["Color (CN)"]
            .fillna("")
            .to_dict()
        )

    # Number of leading columns before size columns
    n_lead = 2 if has_cn else 1  # Color [, Color (CN)]

    size_order = get_size_order()

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        wrote_any = False
        for style in df_size["Style"].unique():
            sub = df_size[df_size["Style"] == style]
            pivot = sub.pivot_table(
                index="Color", columns="Size",
                values="Units", aggfunc="sum", fill_value=0,
            )
            known   = [s for s in size_order if s in pivot.columns]
            unknown = [s for s in pivot.columns if s not in size_order]
            size_cols = known + unknown
            pivot = pivot.reindex(columns=size_cols)

            sheet = clean_sheet_name(style)   # BUG-42: handles illegal chars + ≤31
            ws = writer.book.create_sheet(sheet)
            writer.sheets[sheet] = ws
            wrote_any = True

            n_sizes = len(size_cols)

            # ---- Header rows ----
            # Row 1: "Color" [| "Color (CN)"] | "Sizes" (merged across size cols)
            # Row 2: merged  [| merged        ] | individual size names
            _style_header(ws.cell(row=1, column=1), "Color")
            _style_header(ws.cell(row=2, column=1), None)
            ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

            if has_cn:
                _style_header(ws.cell(row=1, column=2), "Color (CN)")
                _style_header(ws.cell(row=2, column=2), None)
                ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)

            size_start_col = n_lead + 1
            for c in range(size_start_col, size_start_col + n_sizes):
                _style_header(ws.cell(row=1, column=c),
                              "Sizes" if c == size_start_col else None)
            for i, sz in enumerate(size_cols):
                _style_header(ws.cell(row=2, column=size_start_col + i), sz)

            if n_sizes > 1:
                ws.merge_cells(start_row=1, start_column=size_start_col,
                               end_row=1, end_column=size_start_col + n_sizes - 1)

            # ---- Data rows from row 3 ----
            border = _thin_border()
            row_idx = 3
            for color, row in pivot.iterrows():
                ws.cell(row=row_idx, column=1, value=color)
                if has_cn:
                    ws.cell(row=row_idx, column=2, value=cn_map.get(str(color), ""))
                for i, sz in enumerate(size_cols):
                    ws.cell(row=row_idx, column=size_start_col + i, value=int(row[sz]))
                for c in range(1, size_start_col + n_sizes):
                    cell = ws.cell(row=row_idx, column=c)
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if c >= size_start_col and isinstance(cell.value, (int, float)):
                        cell.number_format = "#,##0"
                row_idx += 1

            # ---- Column widths ----
            ws.column_dimensions["A"].width = max(
                10, max((len(str(c)) for c in pivot.index), default=6) + 3
            )
            if has_cn:
                cn_vals = [str(cn_map.get(str(c), "")) for c in pivot.index]
                ws.column_dimensions["B"].width = max(
                    12, max((len(v) for v in cn_vals), default=6) + 3
                )
            for i in range(n_sizes):
                ws.column_dimensions[get_column_letter(size_start_col + i)].width = max(6, len(size_cols[i]) + 3)

        if not wrote_any:
            ws = writer.book.create_sheet("Empty")
            writer.sheets["Empty"] = ws
            ws.cell(row=1, column=1, value="No data")

    return path
