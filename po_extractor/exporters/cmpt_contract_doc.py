"""CMPT contract document generation — fills the user's own Excel template.

The template is any .xlsx the user uploads (stored at
``data/cmpt_templates/default.xlsx``), edited once to contain
``{{placeholder}}`` tokens wherever contract data should go:

Header tokens (any cell, mixed with surrounding text):
    {{contract_no}} {{contract_date}} {{factory}} {{company}} {{currency}}
    {{status}} {{notes}} {{today}}
    {{total_qty}} {{total_amount}} {{total_amount_cn}}   <- 大写金额

Line tokens — put ONE prototype row anywhere in the sheet containing any of:
    {{line.no}} {{line.po}} {{line.style}} {{line.color}}
    {{line.description}} {{line.qty}} {{line.unit_price}} {{line.amount}}
That row is replicated once per contract line (rows below shift down), so
merged titles above and signature blocks below survive.

Unknown ``{{tokens}}`` are left as-is (visible in the output — easier to
spot a typo than silently swallowing it).
"""
from __future__ import annotations

import io
import re
from datetime import date

import openpyxl
from copy import copy

from ..utils.rmb_amount import rmb_capital
from ._excel_helpers import neutralise_foreign_formulas

_TOKEN_RE = re.compile(r"\{\{([a-z_.]+)\}\}")
_LINE_PREFIX = "line."


def _header_values(contract: dict) -> dict[str, str]:
    agreed = contract.get("agreed_total") or 0
    return {
        "contract_no":     str(contract.get("contract_no") or ""),
        "contract_date":   str(contract.get("contract_date") or ""),
        "factory":         str(contract.get("factory") or ""),
        "company":         str(contract.get("company") or ""),
        "currency":        str(contract.get("currency") or ""),
        "status":          str(contract.get("status") or ""),
        "notes":           str(contract.get("notes") or ""),
        "today":           date.today().isoformat(),
        "total_qty":       f"{int(contract.get('total_qty') or 0):,}",
        "total_amount":    f"{agreed:,.2f}",
        "total_amount_cn": rmb_capital(agreed),
    }


def _line_values(line: dict, no: int) -> dict[str, str]:
    return {
        "line.no":          str(no),
        "line.po":          str(line.get("po_number") or ""),
        "line.style":       str(line.get("style") or ""),
        "line.color":       str(line.get("color") or ""),
        "line.description": str(line.get("description") or ""),
        "line.qty":         f"{int(line.get('qty') or 0):,}",
        "line.unit_price":  f"{float(line.get('unit_price') or 0):,.2f}",
        "line.amount":      f"{float(line.get('amount') or 0):,.2f}",
    }


def _substitute(text: str, values: dict[str, str]) -> str:
    def repl(m: re.Match) -> str:
        return values.get(m.group(1), m.group(0))
    return _TOKEN_RE.sub(repl, text)


def generate_cmpt_contract_xlsx(template_bytes: bytes, contract: dict) -> bytes:
    """Fill *template_bytes* (the user's template .xlsx) from *contract*
    (a ``CmptContractStore.get_contract()`` dict). Returns the filled
    workbook as bytes.

    Raises ValueError when the template can't be read.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    except Exception as exc:
        raise ValueError(f"Template could not be read: {exc}") from exc

    header_vals = _header_values(contract)
    lines = contract.get("lines") or []

    for ws in wb.worksheets:
        # ── 1. Find the prototype line row (first row with a line.* token) ──
        proto_row = None
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and _LINE_PREFIX in cell.value \
                        and _TOKEN_RE.search(cell.value):
                    proto_row = cell.row
                    break
            if proto_row:
                break

        # ── 2. Replicate the prototype row once per line ────────────────────
        if proto_row is not None and lines:
            n_extra = len(lines) - 1
            if n_extra > 0:
                ws.insert_rows(proto_row + 1, amount=n_extra)
                # Copy the prototype's cell values + styles into the new rows.
                for offset in range(1, n_extra + 1):
                    for col in range(1, ws.max_column + 1):
                        src = ws.cell(proto_row, col)
                        dst = ws.cell(proto_row + offset, col)
                        dst.value = src.value
                        if src.has_style:
                            dst.font          = copy(src.font)
                            dst.fill          = copy(src.fill)
                            dst.border        = copy(src.border)
                            dst.alignment     = copy(src.alignment)
                            dst.number_format = src.number_format
                    ws.row_dimensions[proto_row + offset].height = \
                        ws.row_dimensions[proto_row].height
            for i, line in enumerate(lines):
                vals = {**header_vals, **_line_values(line, i + 1)}
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(proto_row + i, col)
                    if isinstance(cell.value, str) and "{{" in cell.value:
                        cell.value = _substitute(cell.value, vals)
        elif proto_row is not None and not lines:
            # No lines: blank the prototype tokens rather than shipping them.
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(proto_row, col)
                if isinstance(cell.value, str) and "{{" in cell.value:
                    cell.value = _substitute(
                        cell.value,
                        {k: "" for k in _line_values({}, 0)} | header_vals,
                    )

        # ── 3. Header tokens everywhere else ────────────────────────────────
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "{{" in cell.value:
                    cell.value = _substitute(cell.value, header_vals)

    buf = io.BytesIO()
    # Any '=' text that arrived in the data becomes inert here; the
    # exporter's own =SUM()/='Sheet'! formulas are left alone.
    neutralise_foreign_formulas(wb)
    wb.save(buf)
    wb.close()
    return buf.getvalue()
