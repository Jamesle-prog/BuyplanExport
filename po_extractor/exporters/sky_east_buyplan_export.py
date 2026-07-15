"""Sky East buy-plan export — Python port of VBA CreateBuyPlan_template2.

Two public entry points
-----------------------
export_sky_east_buyplan(df_items, cn_lookup, output_dir)
    → one Excel file, one sheet per style (Sky_East.xlsx template)
    → includes fabric-header cells (B2:E4), Q5 style total, Index sheet

export_sky_east_nukuryou(df_items, cn_lookup, output_dir)
    → list of Excel files, one per distinct fabric_item_no
      (Sky_East_P.xlsx template; one sheet per style within each workbook)

Both mirror the VBA subs:
    CreateBuyPlan_template2 / ProcessSKU / ProcessDataGrouping
    CreateZalando核料Workbooks

Data-row column layout (1-based, matching the Template sheet):
    A=合同号  B=Style  C=Brand  D=Article  E=PO  F=ConfigSKU
    G=ColorEN  H=ColorCN  I=主标色
    J=XS  K=S  L=M  M=L  N=XL  O=XXL
    P=船样要求 (blank — not in data model)
    Q=Total  R=EX-FTY

Fabric-header rows (rows 2-5, above data):
    B=大身 (body part)  C=HHN code  D=面料成分  E=综合标识Key
"""
from __future__ import annotations

import re
import zipfile
from datetime import datetime
from pathlib import Path
from typing import NamedTuple

import pandas as pd
from openpyxl import load_workbook

from ._sky_east_helpers import (
    _TEMPLATES_DIR, _SE_TEMPLATE, _SE_TEMPLATE_P,
    _SIZES_LC, _COL_BOAT_SAMPLE,
    _apply_config_overrides, _clean_sheet_name, _clear_data_area, _cn_color,
    _create_index_sheet, _create_overview_sheet,
    _detect_buyplan_layout, _detect_fabric_rows, _detect_nukuryou_layout,
    _embed_style_photos, _replace_placeholders, _set_sheet_column_widths,
    _strip_color_brackets, _style_data, _style_total, derive_main_label_color,
    _COLOR_NOT_FOUND, _color_miss_comment_text,
)
from ._excel_helpers import apply_print_settings
from ..utils.file_utils import versioned_path
from ..store.color_translation_store import _normalize_color_name as _nz_color
from ..lookups.progress_lookup import _norm_key
from auth.companies import COMPANY_SKY_EAST


# ── Display-formatting parameters ────────────────────────────────────────────
# Separator joining 中文颜色代码 and 中文颜色 in the BODY COLOR-CN cell, e.g.
# ``"52#|黑色"``.  Edit here to change the format everywhere.
_CN_DISPLAY_SEP = "|"

# Template for the per-colour row label in 核料 workbooks, e.g.
# ``"Black(52#|黑色)"``.  Variables: ``{en}`` (English colour) and ``{cn}``
# (the formatted Chinese display string from ``_format_body_color_cn``).
_NUKURYOU_LABEL_FMT = "{en}({cn})"

# Sentinel used as the "brand" component of brand-agnostic flat lookup keys.
# Centralised so changes propagate to every dict that follows the convention.
_BRAND_AGNOSTIC = ""

# _COLOR_NOT_FOUND is defined in _sky_east_helpers.py (shared with the
# Overview sheet builder, which attaches a diagnostic comment to any cell
# showing it) and reaches this module via the `from ._sky_east_helpers
# import *` above.

# Sentinel distinguishing "sky_east_store not passed — auto-fetch the
# canonical store" from "explicitly passed None — disable colour-miss
# logging entirely".  Callers that just want the default behaviour never
# need to know this exists; tests pass ``sky_east_store=None`` to keep
# colour-miss diagnostics (a real DB write) out of the shared dev database.
_AUTO_STORE = object()


def _se_base_style_key(style: str, known_styles: set[str]) -> str:
    """Map a trailing-"A" variant onto its base style when the base exists.

    ``"DR5302A"`` → ``"DR5302"`` only when ``"DR5302"`` is itself present in
    *known_styles*; every other style is returned unchanged.  Used to group an
    "A" variant into the same buy-plan sheet as its base style while each data
    row still carries its own style name (with / without the trailing "A").
    A standalone ``"…A"`` style with no matching base keeps its own sheet.
    """
    s = str(style or "").strip()
    if len(s) > 1 and s.endswith("A") and s[:-1] in known_styles:
        return s[:-1]
    return s


def _format_body_color_cn(
    cn_code: str, color_cn: str, sep: str = _CN_DISPLAY_SEP,
) -> str:
    """Return the BODY COLOR-CN cell text: ``"<code><sep><name>"``.

    Strips any embedded code prefix from *color_cn* to prevent doubling.
    Source data (Colors tab DB or 大货进度表) sometimes stores the color as
    ``"84#红色"`` or ``"84|红色"`` rather than the plain name ``"红色"``.
    When *cn_code* is ``"84"`` and *color_cn* is ``"84#红色"``, naïvely
    joining gives ``"84|84#红色"`` — so we detect and strip the prefix.
    """
    if cn_code and color_cn:
        # Strip any leading "{cn_code}#", "{cn_code}{sep}", or "{cn_code} "
        # prefix that source data may already embed in the Chinese color field.
        name = color_cn
        for pfx in (f"{cn_code}#", f"{cn_code}{sep}", f"{cn_code} "):
            if color_cn.startswith(pfx):
                name = color_cn[len(pfx):].lstrip()
                break
        return f"{cn_code}{sep}{name}" if name else cn_code
    return color_cn or cn_code


def _brand_keyed_get(
    lookup: dict | None, company: str, brand: str, norm_en: str,
    *, default: str = "",
) -> str:
    """Primary-only (brand-specific) lookup against a flat color dict.

    Only the exact ``(company, brand, norm_en)`` key is tried.  No
    brand-agnostic fallback — returning a value from a different brand's
    mapping would produce wrong color codes.  Returns *default* when the
    key is absent.
    """
    if not lookup:
        return default
    return lookup.get((company, brand, norm_en)) or default


def _progress_brand(brand_by_pc: dict | None, pc_no, sty_norm: str,
                    fallback: str) -> str:
    """品牌 from 大货进度表: try ``(pc_no_norm, style_norm)`` then ``style_norm``
    alone; fall back to *fallback* (the order file's own brand) when the
    progress table has no brand for this style/PC or isn't loaded."""
    if not brand_by_pc:
        return fallback
    pcn = _norm_key(str(pc_no or ""))
    return (brand_by_pc.get((pcn, sty_norm))
            or brand_by_pc.get(sty_norm)
            or fallback)


def _available_progress_colors(
    cn_by_pc_lookup: dict | None, pc_no_norm: str, sty_norm: str,
) -> list[str]:
    """Return the distinct colour names 大货进度表 actually has on file for
    (pc_no_norm, sty_norm) — used to build a diagnostic comment when the
    order-file colour doesn't match any of them, so a reviewer can see the
    client's PO colour and 大货进度表's own colour(s) side by side instead of
    having to open the source file to check for a naming mismatch (e.g.
    client says "Dark Blue", 大货进度表 says "Navy").

    Cheap by construction: only ever called on a miss (rare), not per row.
    """
    if not cn_by_pc_lookup:
        return []
    return sorted({
        key[2] for key in cn_by_pc_lookup
        if key[0] == pc_no_norm and key[1] == sty_norm and key[2]
    })


def _resolve_pc_color(
    row, sty_norm: str, color_en: str, brand: str,
    cn_lookup: dict, cn_code_lookup: dict | None,
    cn_by_pc_lookup: dict | None,
) -> tuple[str, str, str, str]:
    """Resolve (color_cn, cn_code, label_color, color_cn_display) for one row.

    The *selected* colour source is authoritative — whichever one the caller
    wired in is used exclusively, with no fallback to the other:

      • ``cn_by_pc_lookup`` is not ``None`` → 大货进度表 is the selected
        source (the user's "Chinese color mapping source" radio picked it
        and a file is loaded).  A miss on the exact (pc_no, style, color)
        key is reported as an explicit ``_COLOR_NOT_FOUND`` error — it does
        **not** fall through to the internal colour DB.
      • ``cn_by_pc_lookup`` is ``None`` → the internal colour DB
        (``cn_lookup`` / ``cn_code_lookup``) is the selected source.  A miss
        there is reported the same way.

    Silently trying a different source than the one the user explicitly
    picked would show data that looks legitimate but didn't come from where
    they expect it to — an explicit error is more useful than a
    plausible-looking value from the wrong place.

    ``label_color`` (主标颜色) is exempt from this rule: the buyplan caller
    always falls back through ``label_lookup`` → the light/dark heuristic
    regardless of source, since a label colour is always derivable from the
    English body colour and withholding it would just break wash labels.
    """
    norm_en = _nz_color(color_en)

    if cn_by_pc_lookup is not None:
        pc_match = cn_by_pc_lookup.get(
            (_norm_key(row.get("pc_no") or ""), sty_norm, norm_en)
        )
        if pc_match is None:
            return (_COLOR_NOT_FOUND, _COLOR_NOT_FOUND, "", _COLOR_NOT_FOUND)
        color_cn    = pc_match.cn_color
        cn_code     = pc_match.color_code or "NA"
        label_color = pc_match.label_color
        return color_cn, cn_code, label_color, _format_body_color_cn(cn_code, color_cn)

    color_cn = _cn_color(cn_lookup, brand, color_en)
    cn_code  = _brand_keyed_get(cn_code_lookup, COMPANY_SKY_EAST, brand, norm_en, default="")
    if not color_cn and not cn_code:
        return (_COLOR_NOT_FOUND, _COLOR_NOT_FOUND, "", _COLOR_NOT_FOUND)
    cn_code = cn_code or "NA"
    return color_cn, cn_code, "", _format_body_color_cn(cn_code, color_cn)


def _ai_retry_component(
    row, sty_norm: str, comp: str, brand: str,
    cn_lookup: dict, cn_code_lookup: dict | None,
    cn_by_pc_lookup: dict | None,
    ai_api_key: str, ai_model: str,
) -> tuple[str, str, str, str] | None:
    """AI-assisted retry for a colour *comp* that missed the local lookup.

    Two strategies, in order:

    1. **Constrained candidate match** (progress source only) — when
       ``cn_by_pc_lookup`` is the selected source and the row's 客人PC NO +
       款式 *does* have colour(s) on file, ask the API which of those exact
       recorded colours *comp* is the SAME NAME written differently — a
       typo ("Daek Blue" -> "Dark Blue") or abbreviation of the identical
       name ("DK Brown" -> "Dark Brown"). The API is explicitly told NOT to
       match different-but-related colour names (e.g. "Navy" is a different
       colour from "Dark Blue", not a synonym) — a wrong pick would silently
       assign the wrong colour code, worse than an honest miss. It also
       only ever picks among colours already on file, so it can never
       inject a fabricated value.
    2. **Open-ended recognition** — fall back to asking the API to normalise
       *comp* into candidate colour names, then retry the lookup with each.
       Covers the internal-DB source (which has no per-PC candidate set) and
       progress rows whose colour is buried in a longer description.

    Returns the first result that resolves, or ``None``.
    """
    from ..lookups.color_ai_enhance import (
        match_color_to_candidates, recognize_colors,
    )

    if cn_by_pc_lookup is not None:
        candidates = _available_progress_colors(
            cn_by_pc_lookup, _norm_key(row.get("pc_no") or ""), sty_norm,
        )
        if candidates:
            picked = match_color_to_candidates(
                comp, candidates, ai_api_key, ai_model,
            )
            if picked:
                result = _resolve_pc_color(
                    row, sty_norm, picked, brand,
                    cn_lookup, cn_code_lookup, cn_by_pc_lookup,
                )
                if result[0] != _COLOR_NOT_FOUND:
                    return result

    for cand in recognize_colors(comp, ai_api_key, ai_model):
        result = _resolve_pc_color(
            row, sty_norm, cand, brand,
            cn_lookup, cn_code_lookup, cn_by_pc_lookup,
        )
        if result[0] != _COLOR_NOT_FOUND:
            return result
    return None


def _resolve_pc_color_multi(
    row, sty_norm: str, color_en: str, brand: str,
    cn_lookup: dict, cn_code_lookup: dict | None,
    cn_by_pc_lookup: dict | None,
    *,
    ai_enhance: bool = False,
    ai_api_key: str = "",
    ai_model: str = "deepseek-chat",
) -> tuple[str, str, str, str]:
    """Same resolution as :func:`_resolve_pc_color`, but tolerant of an order
    file colour that names two colours in one string, e.g. ``"Dark Blue /
    White"`` (produced by :func:`_strip_color_brackets` from a source cell
    like ``"(dark blue)(white)"``).

    Mirrors the "detect and separate, then look up on this" approach already
    used for 大货进度表 rows (see ``progress_lookup._split_multi_color``):
    each ``" / "``-separated component is resolved independently, and the
    Chinese name / code of every component that actually resolves are
    combined with ``" / "`` — a two-tone "Dark Blue / White" item shows
    "藏青 / 白色", not just whichever component happened to resolve first.
    A component that never resolves (even after AI enhance) is silently
    omitted from the combination rather than blanking the whole result.

    A single-colour value (the common case — no ``" / "`` present) is passed
    straight through to :func:`_resolve_pc_color` unchanged.

    ``ai_enhance`` ("Local + AI Enhance" mode): whenever an individual
    component fails to resolve locally and an API key is configured, that
    component's raw text is sent to
    :func:`~po_extractor.lookups.color_ai_enhance.recognize_colors` to ask
    an LLM to identify the colour — candidate names are retried against the
    same local lookup, same as any other component. The API is never
    consulted before local resolution of that component has already failed,
    and never for anything other than a colour miss.
    """
    components = [c.strip() for c in color_en.split(" / ") if c.strip()]
    if len(components) <= 1:
        components = [color_en]

    resolved = []
    for comp in components:
        result = _resolve_pc_color(
            row, sty_norm, comp, brand,
            cn_lookup, cn_code_lookup, cn_by_pc_lookup,
        )
        if result[0] == _COLOR_NOT_FOUND and ai_enhance and ai_api_key:
            improved = _ai_retry_component(
                row, sty_norm, comp, brand,
                cn_lookup, cn_code_lookup, cn_by_pc_lookup,
                ai_api_key, ai_model,
            )
            if improved is not None:
                result = improved
        resolved.append(result)

    if len(resolved) == 1:
        return resolved[0]

    cn_names = [r[0] for r in resolved if r[0] and r[0] != _COLOR_NOT_FOUND]
    cn_codes = [r[1] for r in resolved if r[1] and r[1] not in (_COLOR_NOT_FOUND, "NA")]
    label    = next((r[2] for r in resolved if r[2]), "")

    if not cn_names and not cn_codes:
        return (_COLOR_NOT_FOUND, _COLOR_NOT_FOUND, "", _COLOR_NOT_FOUND)

    combined_cn   = " / ".join(cn_names) if cn_names else _COLOR_NOT_FOUND
    combined_code = " / ".join(cn_codes) if cn_codes else "NA"
    return combined_cn, combined_code, label, _format_body_color_cn(combined_code, combined_cn)


def _prefetch_ai_color_cache(
    df_items: pd.DataFrame,
    cn_lookup: dict, cn_code_lookup: dict | None, cn_by_pc_lookup: dict | None,
    ai_enhance: bool, ai_api_key: str, ai_model: str,
    on_progress=None,
) -> None:
    """Parallel pre-warm of every AI colour-recognition call this export will
    need, so the SERIAL per-row/per-sheet writing pass below never blocks on
    a network round-trip.

    "Local + AI Enhance" mode used to make one blocking DeepSeek API call per
    colour that missed local resolution, INLINE inside the row-writing loop
    (one style at a time) -- for a run with a dozen unresolved colours, at
    ~1s+ per call, that alone dominated total export time (19s of a 21s
    export, measured). Both AI functions cache by (raw string) /
    (string, candidate-set) (see color_ai_enhance.py), so warming the cache
    here turns the writing loop's calls into free dict lookups -- mirrors
    price_mask.mask_prices_batch's existing pattern of parallelising AI
    network calls while keeping the actual file-writing work serial.

    Best-effort and over-inclusive by design: for a colour needing AI, BOTH
    the constrained candidate-match (progress source only) and the
    open-ended recognise call are queued, even though the row logic only
    ever needs whichever one actually resolves -- deciding that upfront would
    require replicating _ai_retry_component's full branch logic here just to
    save a few parallel (not serial) calls. A prefetch failure just means the
    writing loop's own call happens serially as before; never breaks export.

    *on_progress(done, total)*, when given, is called once before any job
    starts (done=0) and once as each of the *total* parallel jobs completes
    (in COMPLETION order, not submission order) -- lets a caller show
    "Resolving colours via AI (N/total)..." for what is typically the single
    largest chunk of export time when AI enhance is on. Exceptions from the
    callback are swallowed; a broken progress UI must never break the export.
    """
    if not (ai_enhance and ai_api_key):
        return
    if "style" not in df_items.columns or "color_name" not in df_items.columns:
        return

    from ..lookups.color_ai_enhance import match_color_to_candidates, recognize_colors

    jobs: list = []
    seen: set = set()
    sty_norm_cache: dict[str, str] = {}

    for _, row in df_items.iterrows():
        style = str(row.get("style", "") or "").strip()
        if not style:
            continue
        sty_norm = sty_norm_cache.get(style)
        if sty_norm is None:
            sty_norm = sty_norm_cache[style] = _norm_key(style)
        brand = str(row.get("brand", "") or "")
        color_en = _strip_color_brackets(str(row.get("color_name", "") or "")).title()
        components = [c.strip() for c in color_en.split(" / ") if c.strip()] or [color_en]

        for comp in components:
            if not comp:
                continue
            result = _resolve_pc_color(
                row, sty_norm, comp, brand, cn_lookup, cn_code_lookup, cn_by_pc_lookup,
            )
            if result[0] != _COLOR_NOT_FOUND:
                continue   # resolves locally -- no AI needed for this component

            if cn_by_pc_lookup is not None:
                candidates = _available_progress_colors(
                    cn_by_pc_lookup, _norm_key(row.get("pc_no") or ""), sty_norm,
                )
                if candidates:
                    key = ("match", comp, tuple(sorted(candidates)))
                    if key not in seen:
                        seen.add(key)
                        jobs.append(lambda c=comp, cs=candidates:
                                    match_color_to_candidates(c, cs, ai_api_key, ai_model))

            key = ("recognize", comp)
            if key not in seen:
                seen.add(key)
                jobs.append(lambda c=comp:
                            recognize_colors(c, ai_api_key, ai_model))

    if not jobs:
        return

    total = len(jobs)
    done = 0
    if on_progress:
        try:
            on_progress(done, total)
        except Exception:
            pass

    from concurrent.futures import ThreadPoolExecutor, as_completed
    with ThreadPoolExecutor(max_workers=min(6, total)) as ex:
        futs = [ex.submit(job) for job in jobs]
        for f in as_completed(futs):
            try:
                f.result()
            except Exception:
                pass
            done += 1
            if on_progress:
                try:
                    on_progress(done, total)
                except Exception:
                    pass


def _resolve_ai_enhance_settings(
    ai_enhance: bool | None,
    ai_api_key: str | None,
    ai_model: str | None,
) -> tuple[bool, str, str]:
    """Resolve the "Local + AI Enhance" colour-recognition settings.

    Any argument left as ``None`` is read from the admin **AppSettings** store
    (``KEY_COLOR_AI_ENHANCE`` / ``KEY_DEEPSEEK_API_KEY`` / ``KEY_DEEPSEEK_MODEL``
    — the same DeepSeek key/model used for AI PO extraction).  An explicit
    non-``None`` argument always wins (used by tests / callers that override).

    Never raises: on any store/import error it warns and falls back to safe
    defaults (AI disabled, no key, ``deepseek-chat``).  Returns a fully
    resolved ``(ai_enhance, ai_api_key, ai_model)`` tuple.
    """
    if ai_enhance is not None and ai_api_key is not None and ai_model is not None:
        return bool(ai_enhance), ai_api_key, ai_model
    try:
        from ..store import get_app_settings_store
        from ..store.app_settings_store import (
            KEY_COLOR_AI_ENHANCE, KEY_DEEPSEEK_API_KEY, KEY_DEEPSEEK_MODEL,
        )
        _settings = get_app_settings_store()
        if ai_enhance is None:
            ai_enhance = _settings.get(KEY_COLOR_AI_ENHANCE, "local") == "local_ai_enhance"
        if ai_api_key is None:
            ai_api_key = _settings.get(KEY_DEEPSEEK_API_KEY, "")
        if ai_model is None:
            ai_model = _settings.get(KEY_DEEPSEEK_MODEL, "deepseek-chat")
    except Exception as exc:
        import warnings as _w
        _w.warn(f"[sky_east] AI-enhance settings fetch failed: {exc!r}")
    return bool(ai_enhance), ai_api_key or "", ai_model or "deepseek-chat"


def _apply_sky_east_compact_layout(ws, *, last_row: int) -> None:
    """User-requested compact layout overrides applied after data is filled.

      • Every populated cell uses font size 10 (preserves bold/colour).
      • Columns A–I and P–R = width 20.
      • Columns J–O (sizes) = width 6.
      • Every row 1..last_row+1 = height 28pt.
    """
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter as _gcl

    SIZE_COLS = set(range(10, 16))   # J-O
    # Pre-build two shared Font instances — reusing immutable style objects is
    # ~10x faster than constructing a new Font per cell (avoids deep-copy
    # overhead inside openpyxl for every cell on every sheet).
    _font_cache: dict[tuple, Font] = {}

    for row in ws.iter_rows(min_row=1, max_row=last_row + 1, min_col=1, max_col=18):
        for cell in row:
            f = cell.font
            try:
                key = (f.name, f.bold, f.italic, f.vertAlign,
                       f.underline, f.strike,
                       f.color.rgb if f.color and f.color.type == "rgb" else None)
                if key not in _font_cache:
                    _font_cache[key] = Font(
                        name=f.name, size=10, bold=f.bold, italic=f.italic,
                        vertAlign=f.vertAlign, underline=f.underline,
                        strike=f.strike, color=f.color,
                    )
                cell.font = _font_cache[key]
            except Exception:
                pass

    for col_idx in range(1, 19):    # A..R
        letter = _gcl(col_idx)
        ws.column_dimensions[letter].width = 6 if col_idx in SIZE_COLS else 20

    for r in range(1, max(last_row + 2, 14)):
        ws.row_dimensions[r].height = 28


# ---------------------------------------------------------------------------
# Pre-fetch caches (extracted from export_sky_east_buyplan for testability)
# ---------------------------------------------------------------------------

def _present_order_sizes(df_items: pd.DataFrame) -> list[str]:
    """Return the sizes (canonical XS→XXL order) that carry any non-zero
    quantity anywhere in *df_items* — i.e. the order's actual size range.

    Used to lay out the 核料 size columns dynamically so a size the order never
    uses isn't shown as an empty column, and a size the template's header omits
    (e.g. XS / XXL) is no longer silently dropped.  Falls back to the full
    canonical set when the frame carries no size data at all.
    """
    present: list[str] = []
    for sz in _SIZES_LC:
        if sz in df_items.columns:
            vals = pd.to_numeric(df_items[sz], errors="coerce").fillna(0)
            if bool((vals != 0).any()):
                present.append(sz)
    return present or list(_SIZES_LC)


def _prefetch_boat_sample_cache(df_items: pd.DataFrame) -> dict[str, str]:
    """Batch-fetch 船样要求 (shipping-sample requirement) text for every brand
    in *df_items*, returning ``{brand: requirement}``.  Best-effort: any store
    failure warns and yields an empty cache so the buy plan still generates.
    """
    bsr_cache: dict[str, str] = {}
    try:
        if "brand" in df_items.columns:
            all_brands = [
                str(b).strip() for b in df_items["brand"].dropna().unique() if b
            ]
            if all_brands:
                from ..store import get_boat_sample_store
                bsr_cache = get_boat_sample_store().get_batch(
                    COMPANY_SKY_EAST, all_brands
                ) or {}
    except Exception as exc:
        import warnings as _w
        _w.warn(f"[sky_east buyplan] shipping-sample lookup failed: {exc!r}")
        bsr_cache = {}
    return bsr_cache


class _FabricMasterCache:
    """Batch-fetched fabric-master enrichment, a case-insensitive index over
    it, and a running set of HHN codes not found in the DB (for the end-of-run
    diagnostic warning).

    Encapsulates what used to be three locals + a nested closure inside
    ``export_sky_east_buyplan`` so ``display_key`` can be unit-tested directly.
    """

    __slots__ = ("by_hhn", "by_hhn_lc", "misses")

    def __init__(self, by_hhn: dict | None) -> None:
        self.by_hhn: dict = by_hhn or {}
        # Whitespace/case-insensitive index so stray-cased HHN codes still match.
        self.by_hhn_lc: dict[str, dict] = {
            str(k).strip().lower(): v
            for k, v in self.by_hhn.items()
            if isinstance(k, str)
        }
        self.misses: set[str] = set()

    def display_key(self, hhn: str,
                    fallback_composition: str = "",
                    fallback_gsm=None,
                    fallback_width=None) -> str:
        """Return 综合标识Key for *hhn* (``quality_no|composition|gsm|width``).

        DB-first resolution:
          1. When the HHN exists in the fabric-master cache (even partially),
             read from it — so the value reflects current DB state, not stale
             caller data.  If ``display_key`` is populated, use it verbatim;
             otherwise rebuild the key from DB columns, consulting caller
             fallbacks only for fields that are NULL/0 in the DB record.
          2. When the HHN is not in the DB at all, fall back to the caller's
             values and record the HHN as a miss for the diagnostic warning.

        Lookup is whitespace- and case-insensitive.  Returns an empty string
        only when *hhn* itself is empty.
        """
        if not hhn:
            return ""
        h = str(hhn).strip()
        rec = self.by_hhn.get(h) or self.by_hhn_lc.get(h.lower())

        if rec is not None:
            dk = str(rec.get("display_key") or "").strip()
            if dk:
                return dk
            qno      = str(rec.get("quality_no")     or h).strip()
            db_comp  = str(rec.get("composition_en") or "").strip()
            db_gsm   = rec.get("weight_gsm")
            db_width = rec.get("cuttable_width_cm")
            comp     = db_comp  or str(fallback_composition or "").strip()
            gsm      = db_gsm   or fallback_gsm
            width    = db_width or fallback_width
            gsm_s   = str(int(gsm))   if gsm   else ""
            width_s = str(int(width)) if width else ""
            return f"{qno}|{comp}|{gsm_s}|{width_s}"

        # HHN not in fabric master — record the miss and build a best-effort
        # partial key from caller fallbacks.
        self.misses.add(h)
        comp    = str(fallback_composition or "").strip()
        gsm_s   = str(int(fallback_gsm))   if fallback_gsm   else ""
        width_s = str(int(fallback_width)) if fallback_width else ""
        if comp or gsm_s or width_s:
            return f"{h}|{comp}|{gsm_s}|{width_s}"
        return h


def _prefetch_fabric_master_cache(
    fabric_parts_by_style: dict | None, df_items: pd.DataFrame,
    fabric_version_id: int | None = None,
) -> _FabricMasterCache:
    """Collect every HHN code referenced by *fabric_parts_by_style* and
    *df_items* and batch-fetch their fabric-master enrichment in one call,
    wrapped in a :class:`_FabricMasterCache`.  Best-effort: a store failure
    warns and yields an empty cache.

    *fabric_version_id* pins the enrichment to a specific archived fabric-list
    version (see ``FabricMasterStore.list_versions``/``get_batch_enrichment``)
    instead of the live table; ``None`` (default) is "use the latest version",
    unchanged from prior behavior.
    """
    fm_cache: dict = {}
    try:
        all_hhns: set = set()
        if fabric_parts_by_style:
            for _parts in fabric_parts_by_style.values():
                for fp in _parts or []:
                    h = str(getattr(fp, "hhn_no", "") or "").strip()
                    if h:
                        all_hhns.add(h)
        if "fabric_item_no" in df_items.columns:
            for v in df_items["fabric_item_no"].dropna().astype(str):
                v = v.strip()
                if v:
                    all_hhns.add(v)
        if all_hhns:
            from ..store import get_fabric_master_store
            fm_cache = get_fabric_master_store().get_batch_enrichment(
                list(all_hhns), version_id=fabric_version_id
            ) or {}
    except Exception as exc:
        # Silent failures here have masked real bugs in the past — make the
        # cause visible (the buyplan still proceeds with an empty cache).
        import warnings as _w
        _w.warn(f"[sky_east buyplan] fabric_master lookup failed: {exc!r}")
        fm_cache = {}
    return _FabricMasterCache(fm_cache)


def _fill_fabric_header(ws, combo_parts, fabric_rows, first,
                        fm: _FabricMasterCache) -> list[tuple[str, str]]:
    """Write the fabric-header rows for one style sheet and return the
    ``[(label, 综合标识Key), ...]`` summary that feeds the Overview sheet.

    Layout per fabric-slot row: B=body part, C=HHN (cleared — the code is
    already encoded in the key), D/E=综合标识Key (``quality_no|composition|
    gsm|width``).  When *combo_parts* is provided each part fills one slot;
    otherwise a single slot is populated from the item row's ``fabric_item_no``
    fallback.  ``first`` is the representative item row supplying the
    ``fabrication`` composition fallback.
    """
    fabrication_fb = str(first.get("fabrication", "") or "")
    fabric_summary: list[tuple[str, str]] = []

    if combo_parts is not None:
        for slot_idx, fp in enumerate(combo_parts[:len(fabric_rows)]):
            frow, body_c, hhn_c, _comp_c, dk_c = fabric_rows[slot_idx]
            hhn       = str(getattr(fp, "hhn_no", "") or "")
            body_part = str(getattr(fp, "body_part", "") or "")
            comp_fb   = str(getattr(fp, "composition", "") or "") or fabrication_fb
            ws.cell(frow, body_c).value = body_part
            ws.cell(frow, hhn_c).value  = None    # cleared — HHN already in the key
            dk = fm.display_key(
                hhn,
                fallback_composition=comp_fb,
                fallback_gsm  =getattr(fp, "weight_gsm", None) or None,
                fallback_width=getattr(fp, "width_cm",  None) or None,
            )
            ws.cell(frow, dk_c).value = dk
            fabric_summary.append(
                (f"{hhn} ({body_part})" if body_part else hhn, dk)
            )
            # Defensive: clear column E in case an older config put the display
            # key there — the value belongs in D.
            if dk_c != 5:
                ws.cell(frow, 5).value = None
    else:
        # Fallback: populate the first slot from the item row's fabric_item_no.
        if fabric_rows:
            frow, _body_c, hhn_c, _comp_c, dk_c = fabric_rows[0]
            hhn_fb = str(first.get("fabric_item_no", "") or "")
            ws.cell(frow, hhn_c).value = None
            dk = fm.display_key(hhn_fb, fallback_composition=fabrication_fb)
            ws.cell(frow, dk_c).value = dk
            if hhn_fb:
                fabric_summary.append((hhn_fb, dk))
            if dk_c != 5:
                ws.cell(frow, 5).value = None
    return fabric_summary


class _RowContext(NamedTuple):
    """Immutable per-export context threaded into :func:`_fill_one_style_row`.

    Bundles the lookups, resolved AI settings, caches, and column map that are
    constant for the whole workbook, so the per-row writer takes a sane
    signature instead of ~13 positional args.  ``sty_norm_cache`` is a shared
    (mutable) memo of ``style -> _norm_key(style)`` — pure, so caching it once
    across all sheets is equivalent to (and slightly cheaper than) the old
    per-sheet cache.
    """
    col: dict
    cn_lookup: dict
    cn_code_lookup: dict | None
    cn_by_pc_lookup: dict | None
    brand_by_pc: dict | None
    label_lookup: dict | None
    ai_enhance: bool
    ai_api_key: str
    ai_model: str
    bsr_cache: dict
    boat_sample_col: int
    se_store: object
    color_source: str
    sty_norm_cache: dict
    # Mutable collector: (style, color_en, progress_label, derived_label) for
    # every row where 大货进度表's 主标颜色 disagrees with the derived heuristic.
    label_warnings: list
    # Mutable collector: (style, color_en) for every row that has NO 主标颜色 on
    # file (大货进度表 / DB) — the cell is left blank (never derived) and flagged.
    label_missing: list


def _fill_one_style_row(ws, out_row: int, g, grp_df, base_style: str,
                        sheet_title: str, front, fabric_summary,
                        ctx: _RowContext) -> tuple[int, dict]:
    """Write one grouped buy-plan data row at *out_row*; return
    ``(row_total, overview_row_dict)``.

    *g* is the group's representative row, *grp_df* the full group (summed for
    sizes), *base_style* the sheet's base style (used when a row carries no
    style of its own).  Pure per-row: it reads from g/grp_df, writes cells on
    *ws*, resolves and logs a colour miss, and returns the Overview-sheet dict.
    The caller accumulates ``style_total`` and appends the dict — this function
    owns neither the loop nor ``out_row`` advancement.
    """
    col = ctx.col
    _row_style = str(g.get("style", "") or "").strip() or base_style
    _row_sty_norm = ctx.sty_norm_cache.get(_row_style)
    if _row_sty_norm is None:
        _row_sty_norm = ctx.sty_norm_cache[_row_style] = _norm_key(_row_style)
    # Strip decorative wrapping brackets some order files store the colour in,
    # e.g. "(dark blue)" — both for a clean display and so the value can
    # exact-match a 大货进度表 / internal DB colour key (which carry no brackets).
    color_en = _strip_color_brackets(str(g.get("color_name", "") or "")).title()
    # brand = the order file's own brand (GIII data) — the key every brand-keyed
    # lookup (colour DB, label DB, shipping-sample requirement) was built against.
    # NEVER swap this for the progress-table brand: 大货进度表's BRAND text is a
    # separate, free-typed field that won't match those DB keys, so silently
    # rekeying the lookups would turn matches into misses.
    brand = str(g.get("brand", "") or "")
    # display_brand = what the 品牌 cell actually shows — 大货进度表 first (the
    # requested source), falling back to the order-file brand. Display only;
    # never used as a lookup key.
    display_brand = _progress_brand(ctx.brand_by_pc, g.get("pc_no"), _row_sty_norm, brand)
    # color_en may combine two colours (e.g. "Dark Blue / White") — the multi
    # resolver tries each component individually.
    color_cn, cn_code, _pc_label, color_cn_display = _resolve_pc_color_multi(
        g, _row_sty_norm, color_en, brand,
        ctx.cn_lookup, ctx.cn_code_lookup, ctx.cn_by_pc_lookup,
        ai_enhance=ctx.ai_enhance, ai_api_key=ctx.ai_api_key, ai_model=ctx.ai_model,
    )
    _raw_client_color = str(g.get("color_name", "") or "")
    _progress_colors = None
    if color_cn == _COLOR_NOT_FOUND:
        if ctx.cn_by_pc_lookup is not None:
            _progress_colors = _available_progress_colors(
                ctx.cn_by_pc_lookup,
                _norm_key(str(g.get("pc_no", "") or "")),
                _row_sty_norm,
            )
        if ctx.se_store is not None:
            try:
                ctx.se_store.log_color_miss(
                    pc_no=str(g.get("pc_no", "") or ""),
                    contract_no=str(g.get("contract_no", "") or ""),
                    style=_row_style,
                    po_no=str(g.get("zalando_po", "") or ""),
                    client_po_color=_raw_client_color,
                    attempted_color=color_en,
                    progress_colors=_progress_colors,
                    source=ctx.color_source,
                )
            except Exception:
                pass   # diagnostic logging must never break export

    xs  = int(grp_df["xs"].sum()  if "xs"  in grp_df.columns else 0)
    s   = int(grp_df["s"].sum()   if "s"   in grp_df.columns else 0)
    m   = int(grp_df["m"].sum()   if "m"   in grp_df.columns else 0)
    l   = int(grp_df["l"].sum()   if "l"   in grp_df.columns else 0)
    xl  = int(grp_df["xl"].sum()  if "xl"  in grp_df.columns else 0)
    xxl = int(grp_df["xxl"].sum() if "xxl" in grp_df.columns else 0)
    row_total = xs + s + m + l + xl + xxl

    _style_data(ws.cell(out_row, col["contract"]),  str(g.get("contract_no",  "") or ""))
    _style_data(ws.cell(out_row, col["style"]),     _row_style)
    _style_data(ws.cell(out_row, col["brand"]),     display_brand)
    _style_data(ws.cell(out_row, col["article"]),   str(g.get("article_name", "") or ""))
    _style_data(ws.cell(out_row, col["po"]),        str(g.get("zalando_po",   "") or ""))
    _style_data(ws.cell(out_row, col["config"]),    str(g.get("config_sku",   "") or ""))
    _style_data(ws.cell(out_row, col["color_en"]),  color_en)
    _color_cn_cell = ws.cell(out_row, col["color_cn"])
    _style_data(_color_cn_cell, color_cn_display)
    if color_cn == _COLOR_NOT_FOUND:
        from openpyxl.comments import Comment as _Comment
        _color_cn_cell.comment = _Comment(
            _color_miss_comment_text(_raw_client_color, _progress_colors),
            "PO Extractor",
        )
    # 主标颜色: read from 大货进度表 (authoritative), then the DB label lookup.
    # The value is NEVER derived — the light/dark heuristic is used only to
    # *cross-check* an on-file value:
    #   • On file + agrees with the heuristic → written, no flag.
    #   • On file + disagrees → 大货进度表's value is kept, but the cell gets a
    #     comment and the export raises an end-of-run mismatch warning.
    #   • Not on file at all → the cell is left genuinely blank (not derived)
    #     and flagged as a missing label for manual entry.
    _label_clr = _pc_label  # 大货进度表 主标颜色 (may be "" outside progress mode)
    if not _label_clr:
        _label_clr = _brand_keyed_get(
            ctx.label_lookup, COMPANY_SKY_EAST, brand, _nz_color(color_en),
        )
    _derived_label = derive_main_label_color(color_en)
    _label_cell = ws.cell(out_row, col["label_clr"])
    _style_data(_label_cell, _label_clr)   # blank when nothing is on file
    from openpyxl.comments import Comment as _Comment
    if not _label_clr:
        _label_cell.comment = _Comment(
            f"主标颜色 missing — no label colour on file for body colour "
            f"\"{color_en}\". Not derived; please enter manually.",
            "PO Extractor",
        )
        ctx.label_missing.append((_row_style, color_en))
    elif _derived_label and _label_clr != _derived_label:
        _label_cell.comment = _Comment(
            f"主标颜色 mismatch — 大货进度表: \"{_label_clr}\"; body colour "
            f"\"{color_en}\" suggests \"{_derived_label}\". Using 大货进度表's "
            f"value; please verify.",
            "PO Extractor",
        )
        ctx.label_warnings.append((_row_style, color_en, _label_clr, _derived_label))
    _style_data(ws.cell(out_row, col["xs"]),  xs)
    _style_data(ws.cell(out_row, col["s"]),   s)
    _style_data(ws.cell(out_row, col["m"]),   m)
    _style_data(ws.cell(out_row, col["l"]),   l)
    _style_data(ws.cell(out_row, col["xl"]),  xl)
    _style_data(ws.cell(out_row, col["xxl"]), xxl)
    # col P = 船样要求 — inject from BoatSampleStore if available
    _boat_req = ctx.bsr_cache.get(brand, "")
    if _boat_req:
        _style_data(ws.cell(out_row, ctx.boat_sample_col), _boat_req)
    _style_data(ws.cell(out_row, col["total"]),  row_total)
    _ex_fty = str(g.get("ex_fty_date", "") or "")
    _style_data(ws.cell(out_row, col["ex_fty"]), _ex_fty)

    overview_row = {
        "contract_no":  str(g.get("contract_no", "") or ""),
        "pc_no":        str(g.get("pc_no", "") or ""),
        "style":        _row_style,
        "sheet_name":   sheet_title,
        "color_en":     color_en,
        "color_cn":     color_cn,
        "color_code":   cn_code if cn_code != "NA" else "",
        "client_po_color": _raw_client_color,
        "progress_colors": _progress_colors,
        "label_color":  _label_clr,
        "brand":        display_brand,
        "po":           str(g.get("zalando_po", "") or ""),
        "config_sku":   str(g.get("config_sku", "") or ""),
        "article_name": str(g.get("article_name", "") or ""),
        "xs": xs, "s": s, "m": m, "l": l, "xl": xl, "xxl": xxl,
        "total":        row_total,
        "ex_fty":       _ex_fty,
        "photo":        front,
        "fabrics":      fabric_summary,
    }
    return row_total, overview_row


# ---------------------------------------------------------------------------
# Main buy plan  (Template sheet → one sheet per style)
# ---------------------------------------------------------------------------

def export_sky_east_buyplan(
    df_items: pd.DataFrame,
    cn_lookup: dict,
    output_dir: str,
    fabric_parts_by_style: dict | None = None,
    style_image_map: dict | None = None,
    label_lookup: dict | None = None,
    cn_code_lookup: dict | None = None,
    cn_by_pc_lookup: dict | None = None,
    brand_by_pc_lookup: dict | None = None,
    fabric_version_id: int | None = None,
    ai_enhance: bool | None = None,
    ai_api_key: str | None = None,
    ai_model: str | None = None,
    sky_east_store=_AUTO_STORE,
    on_progress=None,
) -> str:
    """Generate the main Sky East buy plan.

    Parameters
    ----------
    df_items             : ``SkyEastStore.list_items()`` result
    cn_lookup            : ``ColorTranslationStore.build_lookup_dict()`` result
                           — ``{(client, brand, en_color): cn_color}``.
    output_dir           : directory to write the output file
    fabric_parts_by_style: optional ``{style: [FabricPart, ...]}`` mapping from
                           ``POStore.load_fabric_parts_for_styles()``.  When
                           provided, writes the 综合标识Key (display_key) to
                           column E of each fabric-slot row in the style sheet.
                           Falls back to the single ``fabric_item_no`` field from
                           ``df_items`` when absent.
    style_image_map      : optional ``{style_name: image_bytes}`` — when provided,
                           a "图片" thumbnail column is added to the Index sheet
                           next to "款号".
    label_lookup         : optional ``{(client, brand, en_color): label_color}``
                           returned by
                           ``ColorTranslationStore.build_label_lookup_dict()``.
                           When None, the exporter fetches it from the canonical
                           DB itself.  Pass an explicit empty dict to disable
                           DB-driven label lookup entirely.
    cn_code_lookup       : optional ``{(client, brand, en_color): color_code}``
                           (中文颜色代码) returned by
                           ``ColorTranslationStore.build_cn_code_lookup_dict()``.
                           When None, fetched automatically from the DB.
    cn_by_pc_lookup : optional ``{(pc_no_norm, style_norm, en_color_norm): (cn_color, color_code)}``
                           from ``ProgressLookup.build_pc_style_color_lookups()``.
                           When provided, both the Chinese color name and code are
                           resolved in one lookup with PC No. + style + color priority
                           (more specific than the brand + color flat lookup).
                           None → skipped.
    brand_by_pc_lookup   : optional ``{(pc_no_norm, style_norm): brand, style_norm: brand}``
                           from ``ProgressLookup.build_brand_lookup()``.  When
                           provided, the 品牌 cell is sourced from 大货进度表 (PC
                           No. + style, else style alone), falling back to the
                           order file's own brand.  None → order-file brand only.
    fabric_version_id     : optional fabric-list version (see
                           ``FabricMasterStore.list_versions``) to enrich
                           against instead of the live table.  None (default)
                           → the latest version, unchanged from prior behavior.
    ai_enhance            : optional override for the "Local + AI Enhance" colour
                           recognition mode.  When None, read from the admin
                           **Color Recognition** setting.  The API is only ever
                           consulted after local resolution of a colour has
                           already failed — never for anything else.
    ai_api_key            : optional override for the DeepSeek API key used by
                           AI-enhanced colour recognition.  When None, read from
                           the admin DeepSeek settings (the same key used for AI
                           PO extraction).
    ai_model              : optional override for the DeepSeek model used by
                           AI-enhanced colour recognition.  When None, read from
                           the admin DeepSeek settings.
    sky_east_store        : optional override for the store used to log colour
                           resolution misses (see ``SkyEastStore.log_color_miss``).
                           When omitted, the canonical store is auto-fetched.
                           Pass ``None`` explicitly to disable colour-miss
                           logging entirely (e.g. in tests, to avoid writing
                           diagnostic rows into a real database).
    on_progress            : optional ``callable(frac: float, label: str)``
                           called repeatedly as this call progresses from 0.0
                           to 1.0 -- covers the parallel AI colour-resolution
                           prefetch (often the single largest chunk of export
                           time when "Local + AI Enhance" is on) and then each
                           style-group sheet as it's written. Never called if
                           there's nothing to report progress on. Exceptions
                           from the callback are swallowed.

    Returns
    -------
    ``(path, style_totals)`` — the absolute path of the saved .xlsx file and a
    ``{style: total_units}`` dict (the per-style buy-plan totals, consumed by
    :func:`build_cross_comparison`).
    """
    # Auto-fetch label_lookup and cn_code_lookup from the canonical store.
    # Single source of truth: the same DB as cn_lookup.
    if label_lookup is None or cn_code_lookup is None:
        try:
            from ..store import get_color_translation_store
            _cts = get_color_translation_store()
            if label_lookup is None:
                label_lookup = _cts.build_label_lookup_dict()
            if cn_code_lookup is None:
                cn_code_lookup = _cts.build_cn_code_lookup_dict()
        except Exception as exc:
            import warnings as _w
            _w.warn(f"[sky_east buyplan] color lookup failed: {exc!r}")
            if label_lookup is None:
                label_lookup = {}
            if cn_code_lookup is None:
                cn_code_lookup = {}

    # Auto-fetch "Local + AI Enhance" settings when not explicitly overridden.
    ai_enhance, ai_api_key, ai_model = _resolve_ai_enhance_settings(
        ai_enhance, ai_api_key, ai_model,
    )

    # Colour-miss diagnostic log — best-effort, never blocks export.
    if sky_east_store is _AUTO_STORE:
        try:
            from ..store import get_sky_east_store
            _se_store = get_sky_east_store()
        except Exception:
            _se_store = None
    else:
        _se_store = sky_east_store
    _color_source = "progress" if cn_by_pc_lookup is not None else "db"

    if not _SE_TEMPLATE.exists():
        raise FileNotFoundError(f"Sky East template not found: {_SE_TEMPLATE}")

    path       = versioned_path(output_dir, "Sky_East_BuyPlan", ".xlsx")
    created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    tpl_wb = load_workbook(str(_SE_TEMPLATE))
    tpl_ws = tpl_wb.worksheets[0]

    # ── Detect column layout and data start row from the template ─────────
    col, data_row = _detect_buyplan_layout(tpl_ws)
    fabric_rows   = _detect_fabric_rows(tpl_ws, max_row=data_row - 1)

    # ── Apply user-configured overrides (Sky_East_config.json) ─────────────
    # Configured values win over auto-detection, so admins can override the
    # template scan without rebuilding the .xlsx.
    col, data_row, fabric_rows = _apply_config_overrides(
        "sky_east_buyplan", col, data_row, fabric_rows,
    )

    # Compute the style-total anchor cell.  Canonically row 5 of the header
    # block — but data_start_row is admin-configurable, and writing the total
    # at row 5 while data starts at/above it would overwrite a live data cell
    # (which the Index sheet's 订单数合计 formula then reads).  Clamp the
    # anchor into the header block above the data region.
    from openpyxl.utils import get_column_letter as _gcl
    total_col_letter  = _gcl(col["total"])
    _anchor_row = 5 if data_row > 5 else max(1, data_row - 1)
    if _anchor_row != 5:
        import warnings as _w
        _w.warn(
            f"[sky_east_buyplan] data_start_row={data_row} overlaps the "
            f"canonical total anchor row 5 — anchor moved to row {_anchor_row}"
        )
    total_anchor      = f"{total_col_letter}{_anchor_row}"   # e.g. "Q5"

    # Style totals for cross-comparison: {style: int}
    style_totals: dict[str, int] = {}

    # ── Pre-fetch 船样要求 for all brands in one batch ────────────────────────
    # Column resolved from the detected layout (falls back to P/16 via the
    # _COL_BOAT_SAMPLE default baked into _detect_buyplan_layout).
    _BOAT_SAMPLE_COL = col.get("boat_sample", _COL_BOAT_SAMPLE)
    bsr_cache = _prefetch_boat_sample_cache(df_items)

    # ── Pre-fetch fabric_master display_key for all HHN codes in one batch ─
    _fm = _prefetch_fabric_master_cache(fabric_parts_by_style, df_items,
                                        fabric_version_id=fabric_version_id)

    def _tick(frac: float, label: str) -> None:
        if on_progress:
            try:
                on_progress(frac, label)
            except Exception:
                pass

    # ── Pre-warm AI colour recognition IN PARALLEL ──────────────────────────
    # "Local + AI Enhance" mode's colour lookups happen one style at a time in
    # the serial writing loop below -- warm the process-level cache with every
    # needed call up front (concurrently) so that loop hits cache, not network.
    # Typically the single largest chunk of export time when AI enhance is on
    # -- gets its own progress range (5%-40%) rather than a single opaque step.
    def _ai_tick(done: int, total: int) -> None:
        _tick(0.05 + 0.35 * (done / total if total else 1),
              f"Resolving colours via AI ({done}/{total})…")

    _tick(0.0, "Resolving colours…")
    _prefetch_ai_color_cache(
        df_items, cn_lookup, cn_code_lookup, cn_by_pc_lookup,
        ai_enhance, ai_api_key, ai_model,
        on_progress=_ai_tick,
    )
    _tick(0.4, "Writing style sheets…")

    # Constant-for-the-whole-workbook context threaded into the per-row writer.
    _row_ctx = _RowContext(
        col=col, cn_lookup=cn_lookup, cn_code_lookup=cn_code_lookup,
        cn_by_pc_lookup=cn_by_pc_lookup, brand_by_pc=brand_by_pc_lookup,
        label_lookup=label_lookup,
        ai_enhance=ai_enhance, ai_api_key=ai_api_key, ai_model=ai_model,
        bsr_cache=bsr_cache, boat_sample_col=_BOAT_SAMPLE_COL,
        se_store=_se_store, color_source=_color_source, sty_norm_cache={},
        label_warnings=[], label_missing=[],
    )

    # Each (style, fabric-part) combination gets its own sheet.
    # A style with N fabric parts in fabric_parts_by_style produces N sheets,
    # each carrying exactly one fabric in the header row.
    # Styles with no fabric mapping produce one sheet using the fallback
    # fabric_item_no from df_items (existing behaviour).
    # Seed the uniqueness set with sheets already in the workbook (notably
    # the master template, e.g. "ZLD060_S24DTR003_HHN-DB-YS24078").
    # Otherwise the first style whose cleaned name happens to match the
    # template's name triggers a silent auto-rename inside copy_worksheet
    # → the recorded sheet_meta points at a non-existent sheet → the
    # Index row's hyperlink and total formula both break.
    _used_sheet_names: set[str] = set(tpl_wb.sheetnames)
    _sheet_meta_list:  list[dict] = []
    # One entry per item row across the whole workbook — feeds the Overview
    # sheet, a flat cross-check table mirroring the per-style sheets.
    _overview_rows:    list[dict] = []
    # Running 1-based sheet index — matches the Index sheet's own "No."
    # column, so a tab name like "3_DR5334" lines up with Index row 3.
    _sheet_seq = 0

    # Trailing-"A" variants share a sheet with their base style (e.g. DR5302A is
    # grouped into the DR5302 sheet) — but only when the base style is itself
    # present in the data.  The group key is the base style; each data row still
    # carries its own style name (with / without the "A") — see col["style"] in
    # the row loop below.  An "A" style with no matching base keeps its own sheet.
    _known_styles = {
        str(s).strip() for s in df_items["style"].dropna() if str(s).strip()
    }
    _group_keys = df_items["style"].map(
        lambda s: _se_base_style_key(s, _known_styles)
    )

    _style_groups = list(df_items.groupby(_group_keys, sort=False))
    _total_groups = len(_style_groups)
    for _gi, (style, style_df) in enumerate(_style_groups, start=1):
        # ``style`` here is the *base* style (group representative) — used for the
        # sheet name, fabric-part lookup, header placeholder and per-tab total.
        style = str(style or "").strip()
        if not style:
            continue
        _tick(0.4 + 0.55 * (_gi / _total_groups if _total_groups else 1),
              f"Writing {style} ({_gi}/{_total_groups})…")

        first = style_df.iloc[0]
        parts = (fabric_parts_by_style or {}).get(style, []) if fabric_parts_by_style else []

        # Group fabric parts by combo_idx → {0: [fp, fp], 1: [fp, fp], …}
        # Each combination becomes one sheet.  Styles without mapping get one
        # pass with an empty combo (None sentinel) — falls back to df_items.
        combos: dict[int, list] = {}
        for _fp in parts:
            combos.setdefault(_fp.combo_idx, []).append(_fp)
        # Produce ordered list of combos; use [None] sentinel when none exist.
        combo_list: list = [combos[k] for k in sorted(combos)] if combos else [None]

        for combo_parts in combo_list:
            # ── Unique sheet name ─────────────────────────────────────────
            # "<running index>_<style>" — e.g. "3_DR5334". The index alone
            # already guarantees uniqueness even when one style has multiple
            # fabric combos (each combo bumps the counter), so the fabric
            # code no longer needs to be in the name.
            _sheet_seq += 1
            _base_sn = _clean_sheet_name(f"{_sheet_seq}_{style}")
            _sn, _sfx = _base_sn, 2
            while _sn in _used_sheet_names:
                _suffix = f"_{_sfx}"
                # Trim base to leave room for suffix so the clipped name is unique.
                _trimmed = _base_sn[:31 - len(_suffix)]
                _sn = _clean_sheet_name(f"{_trimmed}{_suffix}")
                _sfx += 1
            sheet_title = _sn
            _used_sheet_names.add(sheet_title)

            ws = tpl_wb.copy_worksheet(tpl_ws)
            ws.title = sheet_title
            # Defensive: openpyxl may auto-suffix when the title collides
            # with an existing sheet (e.g. the master template).  Always
            # use the *actual* title for downstream metadata so Index
            # hyperlinks and total formulas resolve correctly.
            sheet_title = ws.title

            # ── Fill {{placeholders}} in template header ──────────────────
            _replace_placeholders(ws, {
                "created_at": created_at,
                "style":      style,
            })

            # ── Always-on date cells ──────────────────────────────────────
            # The template doesn't ship with a {{created_at}} placeholder; the
            # 制单日期 / 修改日期 labels live in J1 / J2 with the value cells
            # K-O on the same row.  Merge K..O so the date is visible across
            # the highlighted rectangle even when J-O columns are narrow,
            # then write the value to K (top-left of the merge).
            for rng in ("K1:O1", "K2:O2"):
                if rng not in (str(m) for m in ws.merged_cells.ranges):
                    try:
                        ws.merge_cells(rng)
                    except Exception:
                        pass
            ws["K1"] = created_at
            ws["K2"] = created_at

            # ── Embed Photo1 (front) and Photo2 (back) from style_image_map ──
            # Always call _embed_style_photos so placeholder text is cleared even
            # when no image is available for this style.
            _imgs  = (style_image_map or {}).get(style) or [] if style_image_map is not None else []
            _front = _imgs[0] if len(_imgs) > 0 else None
            _back  = _imgs[1] if len(_imgs) > 1 else None
            _embed_style_photos(ws, _front, _back)

            # ── Fabric header — all fabrics in this combination ───────────
            # Writes B=body part / D=综合标识Key per fabric-slot row and returns
            # the [(label, key), ...] summary shared by every item row on this
            # sheet (they all carry the same fabric combination).
            _fabric_summary = _fill_fabric_header(
                ws, combo_parts, fabric_rows, first, _fm,
            )

            # ── Clear data area ───────────────────────────────────────────
            _clear_data_area(ws, data_row)

            # ── Group by (Style | PONo | ConfigSKU | ColorDesc) — VBA grpKey ─
            # "style" leads the key so a base style and its "A" variant sharing a
            # sheet keep separate rows (each with its own style name).  For a
            # single-style sheet "style" is constant, so this is a no-op.
            grp_cols = [c for c in ["style", "zalando_po", "config_sku", "color_name"]
                        if c in style_df.columns]
            groups = list(style_df.groupby(grp_cols, sort=False))

            out_row     = data_row
            style_total = 0

            for _key, grp_df in groups:
                # Per-row cell writing + colour resolution + miss logging lives
                # in _fill_one_style_row; the loop owns only accumulation.
                _row_total, _ov = _fill_one_style_row(
                    ws, out_row, grp_df.iloc[0], grp_df, style,
                    sheet_title, _front, _fabric_summary, _row_ctx,
                )
                style_total += _row_total
                _overview_rows.append(_ov)
                out_row += 1

            # Grand-total row (only when >1 data rows, matching VBA logic)
            if len(groups) > 1:
                _style_total(ws.cell(out_row, col["style"]), "Total")
                _style_total(ws.cell(out_row, col["total"]), style_total)

            # Style-total anchor cell (detected dynamically, e.g. "Q5")
            ws[total_anchor] = style_total
            style_totals[style] = style_total   # same total for all fabric sheets

            # ── Apply dynamic column widths & row heights ─────────────────
            _set_sheet_column_widths(ws, col, data_start_row=data_row)

            # ── User-requested compact layout overrides ───────────────────
            # font 10 / width 20 (sizes 6) / row height 20.
            _apply_sky_east_compact_layout(ws, last_row=out_row)

            # ── Collect metadata for Index sheet ─────────────────────────
            # Show the first fabric in the combination as the representative entry.
            _idx_fp = combo_parts[0] if combo_parts else None
            _idx_hhn  = (str(getattr(_idx_fp, "hhn_no", "") or "") if _idx_fp
                         else str(first.get("fabric_item_no", "") or ""))
            _idx_comp = (str(getattr(_idx_fp, "composition", "") or "") if _idx_fp
                         else str(first.get("fabrication", "") or ""))
            _idx_dk = _fm.display_key(
                _idx_hhn,
                fallback_composition=_idx_comp,
                fallback_gsm  =getattr(_idx_fp, "weight_gsm", None) if _idx_fp else None,
                fallback_width=getattr(_idx_fp, "width_cm",   None) if _idx_fp else None,
            )
            _sheet_meta_list.append({
                "style":       style,
                "sheet_name":  sheet_title,
                "brand":       _progress_brand(
                    brand_by_pc_lookup, first.get("pc_no"),
                    _norm_key(str(style or "")), str(first.get("brand", "") or "")),
                "body_part":   str(getattr(_idx_fp, "body_part", "") or "") if _idx_fp else "",
                "hhn_no":      _idx_hhn,
                "display_key": _idx_dk,
                "ex_fty_date": str(first.get("ex_fty_date", "") or ""),
                "pc_no":       str(first.get("pc_no",       "") or ""),
            })

    # ── Remove master template sheet ─────────────────────────────────────
    if tpl_ws in tpl_wb.worksheets:
        tpl_wb.remove(tpl_ws)

    # ── Index sheet (VBA CreateIndexSheet) ───────────────────────────────
    _create_index_sheet(tpl_wb, df_items, total_anchor=total_anchor,
                        style_image_map=style_image_map,
                        sheet_meta_list=_sheet_meta_list)

    # ── Overview sheet — flat item-level cross-check table ────────────────
    if _overview_rows:
        _create_overview_sheet(tpl_wb, _overview_rows, n_fabric_slots=len(fabric_rows))

    if not tpl_wb.sheetnames:
        tpl_wb.create_sheet("Empty")

    apply_print_settings(tpl_wb)
    tpl_wb.save(str(path))

    # ── 综合key diagnostic — surface HHNs missing from fabric_master ──────
    # The buyplan reads the 综合key directly from the fabric_master DB.
    # When some HHNs aren't there yet, the produced key only has whatever
    # the caller could supply (often weight + width are blank → trailing
    # ``||``).  Warn loudly so the user knows to import those fabrics.
    if _fm.misses:
        import warnings as _w
        preview = ", ".join(sorted(_fm.misses)[:8])
        if len(_fm.misses) > 8:
            preview += f" … +{len(_fm.misses) - 8} more"
        _w.warn(
            f"[sky_east buyplan] 综合key partial — {len(_fm.misses)} HHN code(s) "
            f"not in fabric_master DB: {preview}"
        )

    # ── 主标颜色 cross-check — surface 大货进度表 label-colour disagreements ──
    # 大货进度表's 主标颜色 is used verbatim, but where it disagrees with the
    # light/dark heuristic derived from the body colour the row is flagged so a
    # reviewer can confirm the 大货进度表 entry isn't a data-entry slip.
    if _row_ctx.label_warnings:
        import warnings as _w
        n = len(_row_ctx.label_warnings)
        preview = "; ".join(
            f"{sty} \"{ce}\": 大货进度表={lc} vs derived={dv}"
            for sty, ce, lc, dv in _row_ctx.label_warnings[:6]
        )
        if n > 6:
            preview += f" … +{n - 6} more"
        _w.warn(
            f"[sky_east buyplan] 主标颜色 cross-check — {n} item(s) where "
            f"大货进度表's label colour disagrees with the derived value "
            f"(大货进度表 kept; please verify): {preview}"
        )

    # ── 主标颜色 missing — no label colour on file (blank cells, not derived) ──
    if _row_ctx.label_missing:
        import warnings as _w
        n = len(_row_ctx.label_missing)
        preview = "; ".join(
            f"{sty} \"{ce}\"" for sty, ce in _row_ctx.label_missing[:6]
        )
        if n > 6:
            preview += f" … +{n - 6} more"
        _w.warn(
            f"[sky_east buyplan] 主标颜色 missing — {n} item(s) have no label "
            f"colour on file (left blank, not derived; enter manually): {preview}"
        )
    _tick(1.0, "Buy plan complete")
    return str(path), style_totals


# ---------------------------------------------------------------------------
# 核料 workbooks  (Template_P sheet → one workbook per fabric)
# ---------------------------------------------------------------------------

def check_nukuryou_ready(df_items: pd.DataFrame) -> str | None:
    """Return None if 核料 workbooks can be generated, or a human-readable reason string.

    Checks three conditions in order:
    1. Template_P file exists on disk
    2. ``fabric_item_no`` column is present in ``df_items``
    3. At least one row has a non-empty ``fabric_item_no``
    """
    if not _SE_TEMPLATE_P.exists():
        return (
            f"Template_P (Sky_East_P.xlsx) not found at:\n{_SE_TEMPLATE_P}\n"
            "Upload it via Admin → Templates → Sky East Template_P."
        )
    fabric_col = "fabric_item_no"
    if fabric_col not in df_items.columns:
        return "fabric_item_no column is missing from items data — please re-import the Sky East contracts."
    has_fabric = df_items[fabric_col].fillna("").astype(str).str.strip().ne("").any()
    if not has_fabric:
        return (
            "No styles have a fabric HHN code filled in.\n"
            "Open the Sky East → Items tab, fill in the 面料编号 (HHN No.) for each style, "
            "then regenerate."
        )
    return None


def export_sky_east_nukuryou(
    df_items: pd.DataFrame,
    cn_lookup: dict,
    output_dir: str,
    cn_code_lookup: dict | None = None,
    cn_by_pc_lookup: dict | None = None,
    ai_enhance: bool | None = None,
    ai_api_key: str | None = None,
    ai_model: str | None = None,
    sky_east_store=_AUTO_STORE,
) -> list[str]:
    """Generate 核料 (material-allocation) workbooks — one per distinct fabric.

    Each workbook:
      • one sheet per style that uses that fabric
      • Row 2 = size headers (B-G → XS-XXL)
      • Row 3+ = one row per color (col A = color name, B-G = XS-XXL quantities)

    Parameters
    ----------
    cn_code_lookup  : optional ``{(client, brand, en_color): color_code}``
                      (中文颜色代码).  When None, fetched automatically from the DB.
    cn_by_pc_lookup : optional ``{(pc_no_norm, style_norm, en_color_norm): (cn_color, color_code)}``
                      from ``ProgressLookup.build_pc_style_color_lookups()``.
                      Both values resolved in one lookup.  None → skipped.
    ai_enhance, ai_api_key, ai_model : same "Local + AI Enhance" overrides as
                      ``export_sky_east_buyplan`` — auto-fetched from the admin
                      Color Recognition / DeepSeek settings when None.
    sky_east_store  : optional override for the store used to log colour
                      resolution misses (see ``SkyEastStore.log_color_miss``),
                      mirroring ``export_sky_east_buyplan``.  When omitted, the
                      canonical store is auto-fetched.  Pass ``None`` explicitly
                      to disable colour-miss logging (e.g. in tests).

    Returns list of saved file paths (empty if Template_P not found or no fabric codes).
    """
    # Auto-fetch cn_code_lookup from canonical store when not supplied.
    if cn_code_lookup is None:
        try:
            from ..store import get_color_translation_store as _get_cts
            cn_code_lookup = _get_cts().build_cn_code_lookup_dict()
        except Exception as _exc:
            import warnings as _w
            _w.warn(f"[sky_east nukuryou] cn_code_lookup fetch failed: {_exc!r}")
            cn_code_lookup = {}

    ai_enhance, ai_api_key, ai_model = _resolve_ai_enhance_settings(
        ai_enhance, ai_api_key, ai_model,
    )

    # Colour-miss diagnostic log — best-effort, never blocks export.  Mirrors
    # export_sky_east_buyplan so a colour that fails to resolve is logged (and
    # commented) here too, not just in the main buy plan.
    if sky_east_store is _AUTO_STORE:
        try:
            from ..store import get_sky_east_store
            _se_store = get_sky_east_store()
        except Exception:
            _se_store = None
    else:
        _se_store = sky_east_store
    _color_source = "progress" if cn_by_pc_lookup is not None else "db"

    if not _SE_TEMPLATE_P.exists():
        return []

    fabric_col = "fabric_item_no"
    if fabric_col not in df_items.columns:
        return []

    output_paths: list[str] = []

    # ── Read the template bytes ONCE and detect layout ONCE outside the loop ─
    # Each fabric group previously re-parsed the template from disk (ZIP+XML
    # decompression) — with N groups that was N × 100-500 ms.  We now read the
    # file into memory once and re-open it per iteration from that buffer.
    #
    # NOTE: an earlier version used copy.deepcopy(_master_wb) here for speed,
    # but deepcopy produces a workbook whose saved stylesheet has a named style
    # pointing at a non-existent fontId — openpyxl (and Excel's repair check)
    # reject it on reload.  load_workbook(BytesIO(bytes)) yields a clean,
    # reloadable copy for only ~1 ms more per fabric while still avoiding disk.
    import io as _io
    _master_bytes = _SE_TEMPLATE_P.read_bytes()
    _master_wb = load_workbook(_io.BytesIO(_master_bytes))
    _master_ws = _master_wb.worksheets[0]

    color_col, size_col_map, nuk_data_row = _detect_nukuryou_layout(_master_ws)
    # Apply user-configured overrides (Sky_East_P_config.json)
    _size_cfg_explicit = False   # admin pinned specific size columns → keep static
    try:
        from . import template_config as _tc
        _nuk_cfg = _tc.load_config("sky_east_nukuryou")
        _cfg_size_map = _nuk_cfg.get("size_column_map") or {}
        _size_cfg_explicit = bool(_cfg_size_map)
        for sz, raw in _cfg_size_map.items():
            k = str(sz).strip().lower()
            k = "xxl" if k in ("2xl", "xxl") else k
            if k in {"xs", "s", "m", "l", "xl", "xxl"}:
                try:
                    size_col_map[k] = _tc.column_letter_to_int(raw)
                except Exception:
                    pass
        cm = _nuk_cfg.get("column_map") or {}
        for k_raw, raw in cm.items():
            if str(k_raw).strip().lower() in {"color", "color name", "colordesc", "颜色"}:
                try:
                    color_col = _tc.column_letter_to_int(raw)
                except Exception:
                    pass
        if _nuk_cfg.get("data_start_row"):
            try:
                nuk_data_row = int(_nuk_cfg["data_start_row"])
            except (TypeError, ValueError):
                pass
    except Exception:
        pass

    # ── Dynamic size columns — follow the order's actual size range ───────────
    # By default the 核料 size columns are laid out from the sizes that actually
    # appear in the order (canonical XS→XXL order), anchored at the template's
    # first size column.  This stops the template's fixed header (often only
    # S/M/L/XL) from silently dropping XS / XXL quantities, and hides sizes the
    # order never uses.  An explicit size_column_map in the config still wins.
    if not _size_cfg_explicit:
        _first_size_col = min(size_col_map.values()) if size_col_map else (color_col + 1)
        size_col_map = {
            sz: _first_size_col + i
            for i, sz in enumerate(_present_order_sizes(df_items))
        }

    nuk_header_row = nuk_data_row - 1   # row where size headers are written
    created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for fabric_no, fabric_df in df_items.groupby(fabric_col, sort=False):
        fabric_no = str(fabric_no or "").strip()
        if not fabric_no:
            continue

        # Fresh clean copy per fabric from the cached template bytes — avoids
        # re-reading from disk and (unlike deepcopy) reloads/opens correctly.
        tpl_wb = load_workbook(_io.BytesIO(_master_bytes))
        tpl_ws = tpl_wb.worksheets[0]

        _replace_placeholders(tpl_ws, {"created_at": created_at, "fabric": fabric_no})

        for style, style_df in fabric_df.groupby("style", sort=False):
            style = str(style or "").strip()
            if not style:
                continue

            ws = tpl_wb.copy_worksheet(tpl_ws)
            ws.title = _clean_sheet_name(style)

            # Clear data area (keep template header rows intact)
            _clear_data_area(ws, nuk_data_row)

            # Clear any stale template size headers first (the template may ship
            # with a fixed S/M/L/XL header row), then write the dynamic set — so
            # trimming to fewer sizes leaves no orphaned "L"/"XL" labels.
            if size_col_map:
                _first_sc = min(size_col_map.values())
                for _c in range(_first_sc, _first_sc + len(_SIZES_LC)):
                    ws.cell(nuk_header_row, _c).value = None
            # Write size headers at the detected header row
            for sz_key, sz_col in size_col_map.items():
                ws.cell(nuk_header_row, sz_col).value = sz_key.upper().replace("XXL", "2XL")

            # Aggregate sizes by color (style invariant — normalise once).
            # Resolve color display labels once per distinct (color_name, brand)
            # pair, then aggregate sizes vectorised — avoids iterrows overhead.
            _sty_norm = _norm_key(style)
            color_totals: dict[str, dict[str, int]] = {}
            # Build display-label map for distinct color+brand combos
            _color_display_map: dict[tuple, str] = {}
            # Displays whose colour failed to resolve → (client_colour,
            # progress_colours) for the diagnostic comment attached at write time.
            _miss_info: dict[str, tuple[str, list | None]] = {}
            for item in style_df.itertuples(index=False):
                color_en = _strip_color_brackets(
                    str(getattr(item, "color_name", "") or "")
                ).title()
                brand    = str(getattr(item, "brand", "") or "")
                key      = (color_en, brand)
                if key not in _color_display_map:
                    color_cn, _, _, color_cn_display = _resolve_pc_color_multi(
                        item._asdict(), _sty_norm, color_en, brand,
                        cn_lookup, cn_code_lookup, cn_by_pc_lookup,
                        ai_enhance=ai_enhance, ai_api_key=ai_api_key, ai_model=ai_model,
                    )
                    if color_cn == _COLOR_NOT_FOUND:
                        # Colour miss — show the English colour alone (never
                        # embed 未找到 into the label), and record a diagnostic
                        # comment + colour-miss-log entry mirroring the buy plan.
                        display = color_en
                        _raw_client_color = str(getattr(item, "color_name", "") or "")
                        _progress_colors = None
                        if cn_by_pc_lookup is not None:
                            _progress_colors = _available_progress_colors(
                                cn_by_pc_lookup,
                                _norm_key(str(getattr(item, "pc_no", "") or "")),
                                _sty_norm,
                            )
                        _miss_info[display] = (_raw_client_color, _progress_colors)
                        if _se_store is not None:
                            try:
                                _se_store.log_color_miss(
                                    pc_no=str(getattr(item, "pc_no", "") or ""),
                                    contract_no=str(getattr(item, "contract_no", "") or ""),
                                    style=style,
                                    po_no=str(getattr(item, "zalando_po", "") or ""),
                                    client_po_color=_raw_client_color,
                                    attempted_color=color_en,
                                    progress_colors=_progress_colors,
                                    source=_color_source,
                                )
                            except Exception:
                                pass   # diagnostic logging must never break export
                    else:
                        display = (
                            _NUKURYOU_LABEL_FMT.format(en=color_en, cn=color_cn_display)
                            if color_cn_display else color_en
                        )
                    _color_display_map[key] = display
                display = _color_display_map[key]
                if display not in color_totals:
                    color_totals[display] = {sz: 0 for sz in _SIZES_LC}
                for sz in _SIZES_LC:
                    color_totals[display][sz] += int(getattr(item, sz, 0) or 0)

            # Write color rows starting at the detected data row
            out_row = nuk_data_row
            for color_display, sizes in color_totals.items():
                _color_cell = ws.cell(out_row, color_col)
                _color_cell.value = color_display
                if color_display in _miss_info:
                    from openpyxl.comments import Comment as _Comment
                    _cc, _pc = _miss_info[color_display]
                    _color_cell.comment = _Comment(
                        _color_miss_comment_text(_cc, _pc), "PO Extractor",
                    )
                for sz_key, sz_col in size_col_map.items():
                    ws.cell(out_row, sz_col).value = sizes.get(sz_key, 0)
                out_row += 1

        # Remove master template sheet
        if tpl_ws in tpl_wb.worksheets:
            tpl_wb.remove(tpl_ws)

        if not tpl_wb.sheetnames:
            tpl_wb.create_sheet("Empty")

        safe = re.sub(r'[<>:"/\\|?*\s]+', "_", fabric_no).strip("_") or "unknown"
        save_path = versioned_path(output_dir, f"Sky_East_核料_{safe}", ".xlsx")
        apply_print_settings(tpl_wb)
        tpl_wb.save(str(save_path))
        output_paths.append(str(save_path))

    return output_paths


# ---------------------------------------------------------------------------
# Cross-comparison (mirrors VBA CreateCrossComparisonSheet)
# ---------------------------------------------------------------------------

def build_cross_comparison(
    style_totals_buyplan: dict[str, int],
    df_items: pd.DataFrame,
) -> pd.DataFrame:
    """Compare style totals between buy plan and 核料 data.

    Parameters
    ----------
    style_totals_buyplan : {style: total_units} from export_sky_east_buyplan
    df_items             : raw items DataFrame (used to recompute 核料 totals)

    Returns
    -------
    DataFrame with columns:
        Style | Total (Buy Plan) | Total (核料) | Match
    """
    # Compute 核料 totals directly from df_items.  Fold trailing-"A" variants
    # onto their base style (DR5302A → DR5302) so the comparison matches the
    # buy plan, which now groups those variants into one base-style sheet.
    _known_styles = {
        str(s).strip() for s in df_items["style"].dropna() if str(s).strip()
    }
    nukuryou_totals: dict[str, int] = {}
    for style, grp in df_items.groupby("style", sort=False):
        total = sum(
            int(grp[sz].sum()) for sz in _SIZES_LC if sz in grp.columns
        )
        _key = _se_base_style_key(style, _known_styles)
        nukuryou_totals[_key] = nukuryou_totals.get(_key, 0) + total

    all_styles = sorted(
        set(style_totals_buyplan) | set(nukuryou_totals)
    )
    rows = []
    for style in all_styles:
        bp  = style_totals_buyplan.get(style, 0)
        nk  = nukuryou_totals.get(style, 0)
        rows.append({
            "Style":            style,
            "Total (Buy Plan)": bp,
            "Total (核料)":     nk,
            "Match":            "✅ OK" if bp == nk else "❌ Mismatch",
        })
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Public template-management API for Admin UI
# ---------------------------------------------------------------------------

# Configuration file that holds Sky East column-header overrides.
SE_CONFIG_PATH = _TEMPLATES_DIR / "Sky_East_config.json"

# Catalog of Sky East templates managed by the Admin UI.
# Each entry: kind → (path, human-friendly label, description)
SE_TEMPLATE_CATALOG = {
    "main": (
        _SE_TEMPLATE,
        "Main buy plan (Template)",
        "Per-style sheet workbook with fabric header rows, size totals, and Index sheet.",
    ),
    "nukuryou": (
        _SE_TEMPLATE_P,
        "核料 workbooks (Template_P)",
        "One workbook per fabric — Color × Size pivot per style.",
    ),
}


def list_sky_east_templates() -> list[dict]:
    """Return metadata for each managed Sky East template (and the config file)."""
    result: list[dict] = []
    for kind, (path, label, desc) in SE_TEMPLATE_CATALOG.items():
        exists = path.exists()
        result.append({
            "kind":        kind,
            "label":       label,
            "description": desc,
            "file":        path.name,
            "path":        str(path),
            "exists":      exists,
            "size_bytes":  path.stat().st_size if exists else 0,
            "modified":    (datetime.fromtimestamp(path.stat().st_mtime)
                            .strftime("%Y-%m-%d %H:%M:%S") if exists else ""),
        })
    return result


def read_sky_east_template(kind: str) -> bytes:
    """Return the raw bytes of the named Sky East template ('main' or 'nukuryou')."""
    if kind not in SE_TEMPLATE_CATALOG:
        raise ValueError(f"Unknown Sky East template kind: {kind!r}")
    path = SE_TEMPLATE_CATALOG[kind][0]
    if not path.exists():
        raise FileNotFoundError(f"Sky East template not found: {path}")
    return path.read_bytes()


def replace_sky_east_template(kind: str, xlsx_bytes: bytes) -> Path:
    """Overwrite the named Sky East template file with *xlsx_bytes*. Returns its path."""
    if kind not in SE_TEMPLATE_CATALOG:
        raise ValueError(f"Unknown Sky East template kind: {kind!r}")
    if not xlsx_bytes or xlsx_bytes[:2] != b"PK":
        raise ValueError("Uploaded data does not look like a valid .xlsx file (missing PK header).")
    path = SE_TEMPLATE_CATALOG[kind][0]
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(xlsx_bytes)
    return path


def read_sky_east_config_text() -> str:
    """Return the Sky_East_config.json contents as text (or empty string if absent)."""
    if not SE_CONFIG_PATH.exists():
        return ""
    return SE_CONFIG_PATH.read_text(encoding="utf-8")


def write_sky_east_config_text(text: str) -> Path:
    """Write *text* to Sky_East_config.json after verifying it parses as JSON."""
    import json as _json
    # Empty text → delete the config (revert to template defaults)
    if not text.strip():
        if SE_CONFIG_PATH.exists():
            SE_CONFIG_PATH.unlink()
        return SE_CONFIG_PATH
    # Validate JSON before writing so we never persist a broken config
    _json.loads(text)
    SE_CONFIG_PATH.parent.mkdir(parents=True, exist_ok=True)
    SE_CONFIG_PATH.write_text(text, encoding="utf-8")
    return SE_CONFIG_PATH
