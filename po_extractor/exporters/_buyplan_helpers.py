"""Private helpers for buyplan_export.py — constants, styling, template, and writers."""
from __future__ import annotations

import json
import re
import warnings
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..utils.normalize import normalize_header
from ..utils.size_config import get_size_order  # noqa: F401 (re-exported for export_buyplan)
from ..config import EXCEL_PALETTE as _P

# ---------------------------------------------------------------------------
# Path constants
# ---------------------------------------------------------------------------

_DATA_DIR      = Path(__file__).parent.parent.parent / "data"
_TEMPLATES_DIR = _DATA_DIR / "buyplan_templates"
# Legacy single-file paths (still consulted as a last resort)
_TEMPLATE_PATH     = _DATA_DIR / "buyplan_template.xlsx"
_TEMPLATE_CFG_PATH = _DATA_DIR / "buyplan_template_config.json"

# ---------------------------------------------------------------------------
# Dynamic column-detection for template header rows
# ---------------------------------------------------------------------------

# Known non-size field aliases: field_name → list of recognised header strings
_FIELD_ALIASES: list[tuple[str, list[str]]] = [
    ("PO Number",   ["po number", "po no", "po no.", "po#", "p.o. number",
                     "order number", "order no", "订单号", "po"]),
    ("Style",       ["style", "style no", "style no.", "style number",
                     "款号", "款式", "style code"]),
    ("Color",       ["color", "colour", "color name", "colour name",
                     "颜色", "color/colorway", "colorway"]),
    ("Color (CN)",  ["color (cn)", "colour (cn)", "color cn", "colour cn",
                     "中文颜色", "颜色(中文)", "颜色（中文）", "color(cn)"]),
    ("Total",       ["total", "total qty", "total quantity", "grand total",
                     "合计", "总计", "总数量"]),
]

# Regex that matches common garment / footwear sizes (do NOT match generic words)
_SIZE_RE = re.compile(
    r"""^(
        # numeric-only: 0, 00, 2, 4 … 22, 24, 26, 28 … 36, 38, 40 … 54 (even clothing)
        0{1,2}|[2-9]|1[0-9]|2[0-9]|3[0-9]|4[0-9]|5[0-4]
        # waist sizes with W/L suffix: 28W, 30, 32L
        |[2-5][0-9][WL]?
        # letter sizes with optional modifiers
        |XXS|XS|S|M|L|XL|XXL|2XL|3XL|4XL|5XL|XXXL|XXXXL
        # alpha-numeric combos: 1XL, 0X, 1X, 2X, 3X
        |[0-5]X[SL]?|[0-5]XL
        # US number + letter: 6P, 8P, 10P, 12P
        |[0-9]{1,2}P
        # kids/youth: 2T, 3T, 4T
        |[2-6]T
        # ONE SIZE
        |OS|ONE\s*SIZE
    )$""",
    re.VERBOSE | re.IGNORECASE,
)

# ---------------------------------------------------------------------------
# Column auto-detection
# ---------------------------------------------------------------------------

def _auto_detect_columns(
    ws, header_row: int
) -> tuple[dict[str, str], dict[str, str]]:
    """Scan *ws* at *header_row* and return (column_map, size_column_map).

    *column_map*      maps field names (e.g. "PO Number") → column letter.
    *size_column_map* maps size labels (e.g. "S", "M", "XL") → column letter.

    Only returns non-empty dicts when at least one recognised header is found.
    """
    # Build alias → field lookup
    alias_to_field: dict[str, str] = {}
    for field, aliases in _FIELD_ALIASES:
        for alias in aliases:
            alias_to_field.setdefault(normalize_header(alias), field)

    column_map:      dict[str, str] = {}
    size_column_map: dict[str, str] = {}

    try:
        header_cells = ws[header_row]
    except Exception:
        return column_map, size_column_map

    for cell in header_cells:
        if cell.value is None:
            continue
        raw   = str(cell.value).strip()
        norm  = normalize_header(raw)
        col_l = get_column_letter(cell.column)

        if norm in alias_to_field:
            field = alias_to_field[norm]
            column_map.setdefault(field, col_l)
            continue

        if _SIZE_RE.match(raw):
            size_column_map.setdefault(raw, col_l)

    return column_map, size_column_map


def _safe_name(company: str) -> str:
    """Sanitise a company name for use as a filename."""
    return re.sub(r'[<>:"/\\|?*\s]+', '_', company).strip('_') or "unknown"

# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------

def _xfactory_date(xport: str | None) -> str | None:
    if not xport:
        return None
    try:
        return (
            datetime.strptime(str(xport), "%Y-%m-%d") - timedelta(days=10)
        ).strftime("%Y-%m-%d")
    except ValueError:
        return None


def _thin():
    s = Side(border_style="thin", color=_P["black"])
    return Border(left=s, right=s, top=s, bottom=s)


def _style_header(cell, value):
    cell.value = value
    cell.fill = PatternFill(start_color=_P["black"], end_color=_P["black"], fill_type="solid")
    cell.font = Font(color=_P["white"], bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _thin()


def _style_total(cell, value):
    cell.value = value
    cell.fill = PatternFill(start_color=_P["yellow"], end_color=_P["yellow"], fill_type="solid")
    cell.font = Font(color=_P["black"], bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _thin()
    if isinstance(value, (int, float)):
        cell.number_format = "#,##0"


def _style_data(cell, value):
    cell.value = value
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _thin()
    if isinstance(value, (int, float)):
        cell.number_format = "#,##0"

# ---------------------------------------------------------------------------
# Metadata helpers
# ---------------------------------------------------------------------------

def _build_po_meta_cache(metadata: pd.DataFrame) -> dict[str, dict]:
    """Pre-build a {po_number: record_dict} index for O(1) lookup inside style loops."""
    if metadata.empty or "po_number" not in metadata.columns:
        return {}
    cache: dict[str, dict] = {}
    for rec in metadata.to_dict("records"):
        pn = str(rec.get("po_number") or "").strip()
        if pn and pn not in cache:
            cache[pn] = rec
    return cache


def _lookup_meta(cache: dict[str, dict], po_numbers: list[str]) -> dict:
    """Return first non-null value per key across *po_numbers* from pre-built *cache*."""
    out: dict = {}
    for pn in po_numbers:
        rec = cache.get(str(pn).strip(), {})
        for k, v in rec.items():
            if k not in out and v is not None and str(v).lower() != "nan":
                out[k] = v
    return out

# ---------------------------------------------------------------------------
# Template helpers
# ---------------------------------------------------------------------------

def _resolve_template_path(company: str | None) -> Path | None:
    """Return the best matching template path for *company*, or None."""
    _TEMPLATES_DIR.mkdir(parents=True, exist_ok=True)
    if company:
        p = _TEMPLATES_DIR / f"{_safe_name(company)}.xlsx"
        if p.exists():
            return p
    p = _TEMPLATES_DIR / "default.xlsx"
    if p.exists():
        return p
    if _TEMPLATE_PATH.exists():
        return _TEMPLATE_PATH
    return None


def _load_template_config(company: str | None) -> dict:
    """Load and return the full config dict for *company* (merged with defaults)."""
    defaults = {"header_row": 5, "write_headers": True,
                "column_map": {}, "size_column_map": {}, "meta_column_map": {}}
    _TEMPLATES_DIR.mkdir(parents=True, exist_ok=True)
    candidates = []
    if company:
        candidates.append(_TEMPLATES_DIR / f"{_safe_name(company)}_config.json")
    candidates.append(_TEMPLATES_DIR / "default_config.json")
    candidates.append(_TEMPLATE_CFG_PATH)   # legacy
    for p in candidates:
        if p.exists():
            try:
                loaded = json.loads(p.read_text(encoding="utf-8"))
                defaults.update(loaded)
                return defaults
            except Exception:
                pass
    return defaults


def _load_template(company: str | None = None) -> tuple:
    """Return (workbook, template_worksheet, config_dict) or (None, None, default_cfg)."""
    default_cfg = {"header_row": 5, "write_headers": True,
                   "column_map": {}, "size_column_map": {}, "meta_column_map": {}}

    tpl_path = _resolve_template_path(company)
    if not tpl_path:
        return None, None, default_cfg

    cfg = _load_template_config(company)
    try:
        wb = load_workbook(str(tpl_path))
        ws = wb.worksheets[0]

        _found_marker = False
        for row in ws.iter_rows():
            for cell in row:
                if (
                    isinstance(cell.value, str)
                    and "{{data_start}}" in cell.value.lower()
                ):
                    cfg["header_row"] = cell.row
                    cfg["write_headers"] = True
                    _found_marker = True
                    break
            if _found_marker:
                break

        if not cfg.get("column_map") and not cfg.get("size_column_map"):
            auto_col_map, auto_sz_map = _auto_detect_columns(ws, cfg["header_row"])
            if auto_col_map or auto_sz_map:
                cfg["column_map"]      = auto_col_map
                cfg["size_column_map"] = auto_sz_map
                cfg["_auto_detected"]  = True

        return wb, ws, cfg
    except Exception as exc:
        warnings.warn(
            f"buyplan_export: failed to load template for '{company}' "
            f"({tpl_path.name}) — {exc}. Using built-in format."
        )
        return None, None, default_cfg


def _replace_placeholders(ws, values: dict) -> None:
    """Substitute ``{{key}}`` in every string cell."""
    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            v = cell.value
            for key, val in values.items():
                v = v.replace(f"{{{{{key}}}}}", str(val or ""))
            cell.value = v


def _clear_data_area(ws, header_row: int) -> None:
    """Remove values and unmerge merged cells at or below *header_row*."""
    to_unmerge = [
        str(r) for r in ws.merged_cells.ranges if r.min_row >= header_row
    ]
    for r in to_unmerge:
        ws.unmerge_cells(r)
    for row in ws.iter_rows(min_row=header_row):
        for cell in row:
            cell.value = None

# ---------------------------------------------------------------------------
# Data-table writers
# ---------------------------------------------------------------------------

def _write_data_table(
    ws, flat: pd.DataFrame, idx_cols: list[str],
    size_cols: list[str], header_row: int,
) -> None:
    """Write column-header row, data rows, and grand-total row."""
    n_idx = len(idx_cols)
    headers = idx_cols + size_cols + ["Total"]

    for c_idx, h in enumerate(headers, 1):
        _style_header(ws.cell(row=header_row, column=c_idx), h)

    col_totals = [0] * len(size_cols)
    for r_offset, (_, row_data) in enumerate(flat.iterrows()):
        r = header_row + 1 + r_offset
        for ci, col in enumerate(idx_cols):
            _style_data(ws.cell(row=r, column=ci + 1), row_data[col])
        row_total = 0
        for i, sz in enumerate(size_cols):
            v = int(row_data[sz])
            _style_data(ws.cell(row=r, column=n_idx + 1 + i), v)
            col_totals[i] += v
            row_total += v
        _style_total(ws.cell(row=r, column=n_idx + 1 + len(size_cols)), row_total)

    total_row = header_row + 1 + len(flat)
    _style_total(ws.cell(row=total_row, column=1), "Total")
    for ci in range(1, n_idx):
        _style_total(ws.cell(row=total_row, column=ci + 1), "")
    for i, v in enumerate(col_totals):
        _style_total(ws.cell(row=total_row, column=n_idx + 1 + i), int(v))
    _style_total(
        ws.cell(row=total_row, column=n_idx + 1 + len(size_cols)),
        int(sum(col_totals)),
    )


def _col_num(col: str | int) -> int:
    """Convert column letter (e.g. 'J') or number to 1-based int."""
    from .template_config import column_letter_to_int
    return column_letter_to_int(col)


def _write_mapped_rows(
    ws, flat: pd.DataFrame, size_cols: list[str],
    cfg: dict, m: dict,
) -> None:
    """Write data rows using the column_map / size_column_map from *cfg*.

    No header row is written — the template's own headers are preserved.
    """
    data_row     = cfg["header_row"]
    col_map      = cfg.get("column_map", {})
    sz_col_map   = cfg.get("size_column_map", {})
    meta_col_map = cfg.get("meta_column_map", {})

    xfactory = _xfactory_date(m.get("xport_date")) or ""

    col_totals: dict[str, int] = {}

    for _, row_data in flat.iterrows():
        for field, col in col_map.items():
            c = _col_num(col)
            if field == "Total":
                val = sum(int(row_data.get(sz, 0)) for sz in size_cols)
            else:
                val = row_data.get(field, "")
            _style_data(ws.cell(row=data_row, column=c), val)

        for sz, col in sz_col_map.items():
            c   = _col_num(col)
            qty = int(row_data.get(sz, 0))
            _style_data(ws.cell(row=data_row, column=c), qty)
            col_totals[sz] = col_totals.get(sz, 0) + qty

        for meta_field, col in meta_col_map.items():
            c   = _col_num(col)
            val = xfactory if meta_field == "xfactory_date" else ""
            _style_data(ws.cell(row=data_row, column=c), val)

        data_row += 1

    if len(flat) > 1:
        if "Total" in col_map:
            _style_total(
                ws.cell(row=data_row, column=_col_num(col_map["Total"])),
                sum(col_totals.values()),
            )
        for sz, col in sz_col_map.items():
            _style_total(
                ws.cell(row=data_row, column=_col_num(col)),
                col_totals.get(sz, 0),
            )
        _style_total(ws.cell(row=data_row, column=1), "Total")


def _set_col_widths(
    ws, flat: pd.DataFrame, idx_cols: list[str],
    size_cols: list[str], factory: str,
) -> None:
    """Auto-size every column that contains data.

    Strategy
    --------
    * For each column in the sheet, measure the maximum content width across
      all cells (header + data rows) and apply it with a small padding.
    * Minimum widths are enforced per column type so narrow columns stay
      readable.
    * Existing column widths from a copied template are used as a *floor*
      so the template designer's intent is always respected.
    """
    from openpyxl.utils import get_column_letter

    def _vis_len(text: str) -> int:
        """Visual width in Excel character units.

        CJK / fullwidth characters render ~2× wider than ASCII in a
        proportional font (Calibri).  We count each such code-point as 2
        units so the column is wide enough to display them without clipping.
        """
        width = 0
        for ch in text:
            cp = ord(ch)
            # CJK Unified Ideographs + common fullwidth / wide ranges
            if (
                0x1100 <= cp <= 0x115F   # Hangul Jamo
                or 0x2E80 <= cp <= 0x303F  # CJK Radicals / Kangxi / Punctuation
                or 0x3040 <= cp <= 0x33FF  # Hiragana, Katakana, bopomofo …
                or 0x3400 <= cp <= 0x4DBF  # CJK Extension A
                or 0x4E00 <= cp <= 0x9FFF  # CJK Unified Ideographs (main block)
                or 0xA000 <= cp <= 0xA4CF  # Yi Syllables
                or 0xA960 <= cp <= 0xA97F  # Hangul Jamo Extended-A
                or 0xAC00 <= cp <= 0xD7FF  # Hangul Syllables
                or 0xF900 <= cp <= 0xFAFF  # CJK Compatibility Ideographs
                or 0xFE10 <= cp <= 0xFE1F  # Vertical forms
                or 0xFE30 <= cp <= 0xFE4F  # CJK Compatibility Forms
                or 0xFF00 <= cp <= 0xFF60  # Fullwidth Latin / Halfwidth Katakana
                or 0xFFE0 <= cp <= 0xFFE6  # Fullwidth signs
                or 0x20000 <= cp <= 0x2A6DF  # CJK Extension B
            ):
                width += 2
            else:
                width += 1
        return width

    # Build {col_letter: max_visual_width} from all non-empty cells
    col_max: dict[str, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            cl = get_column_letter(cell.column)
            val_len = _vis_len(str(cell.value))
            col_max[cl] = max(col_max.get(cl, 0), val_len)

    for cl, content_len in col_max.items():
        # Preserve existing template width as a floor
        existing = ws.column_dimensions[cl].width or 0
        desired  = content_len + 3          # +3 char padding
        new_w    = max(existing, desired, 8) # never narrower than 8
        # Cap very wide columns (e.g. fabric description) at 45
        new_w    = min(new_w, 45)
        ws.column_dimensions[cl].width = new_w


def _build_default_sheet(
    ws, flat: pd.DataFrame, idx_cols: list[str],
    size_cols: list[str], m: dict, created_at: str,
) -> None:
    """Populate a fresh sheet with the built-in 4-row metadata header + table."""
    factory  = m.get("factory") or "N/A"
    xport    = m.get("xport_date")
    xfactory = _xfactory_date(xport) or "N/A"
    coo      = m.get("country_of_origin") or "N/A"
    div_code = m.get("division_code") or ""
    div_name = m.get("division_name") or ""
    style    = flat["Style"].iloc[0] if not flat.empty else ""

    ws.cell(row=1, column=1, value="工厂信息:").alignment = Alignment(horizontal="right")
    ws.cell(row=1, column=2, value=factory)
    ws.cell(row=2, column=1, value="款号:").alignment = Alignment(horizontal="right")
    ws.cell(row=2, column=2, value=style)
    ws.cell(row=3, column=1, value="面料信息:").alignment = Alignment(horizontal="right")
    ws.cell(row=4, column=1, value="出厂日期:").alignment = Alignment(horizontal="right")
    ws.cell(row=4, column=2, value=xfactory)
    ws.cell(row=4, column=4, value="Orig X-Port Date:")
    ws.cell(row=4, column=5, value=xport or "N/A")
    ws.cell(row=4, column=7, value="COO:")
    ws.cell(row=4, column=8, value=coo)
    ws.cell(row=4, column=10, value="Division:")
    ws.cell(row=4, column=11, value=f"{div_code}  {div_name}".strip())
    ws.cell(row=1, column=13, value="创建时间:").alignment = Alignment(horizontal="right")
    ws.cell(row=1, column=14, value=created_at)

    _write_data_table(ws, flat, idx_cols, size_cols, header_row=5)
    _set_col_widths(ws, flat, idx_cols, size_cols, factory)
