"""House cleanup for cut-plan output — the post-process the cutting room expects.

This replaces the Excel macro that was run by hand on every cut plan
(``cleanCuttingPlanOutput`` = ``translate`` + ``RemovePathAndExtension``):

* English labels become the Chinese the cutting room reads.
* Marker cells arrive as full Windows paths from the marker software
  (``D:\\Markers\\SS26\\ABC-1234.mrk``); only the marker name is wanted.

Both passes run over **every** cell, exactly as the macro did — the labels are
not the only thing that needs it, since marker paths live in data cells.

Faithfulness notes (the macro's behaviour is deliberately preserved):

* Replacements are **case-insensitive substring** matches (``LookAt:=xlPart``),
  so ``"Marker Length,cm"`` becomes ``"版长,cm"`` and ``"N Markers"`` becomes
  ``"N 版s"``. That is what the macro produced and what the cutting room is
  used to reading.
* Order matters and is preserved: ``Marker Length`` and ``Marker Ratio`` must
  be replaced **before** the bare ``Marker`` rule, or the short rule would
  consume the prefix and the specific labels would never match.
* Every replacement is Chinese and every search term is ASCII, so no rule can
  re-match an earlier rule's output — the sequence cannot cascade.
"""
from __future__ import annotations

import re
from typing import Any

# (search, replacement) — ORDER IS SIGNIFICANT, see module docstring.
TRANSLATIONS: list[tuple[str, str]] = [
    # ── Block titles and fixed headings ──────────────────────────────────
    # These come FIRST because several contain a word a later rule also
    # matches ("Marker Definition" vs "Marker", "Total cut length" vs "Cut
    # Length", "Spreading Plies" vs "Spreading"). Longest/most specific wins
    # by running earlier — the same reason the original macro ordered its own
    # rules the way it did.
    ("Order demands",     "订单需求"),
    ("Marker Definition", "版定义"),
    ("Spreading Plies",   "铺布层数"),
    ("Total Efficiency",  "总利用率"),
    ("Total Quantity",    "总数量"),
    ("Total Tables",      "总台数"),
    ("Total Cost",        "总成本"),
    ("Average Length",    "平均版长"),
    ("Waste Limits, cm",  "损耗上限,cm"),
    ("Cut plan operator", "裁剪计划员"),
    ("Style file",        "款式文件"),
    ("Style name",        "款式名称"),
    ("Order name",        "订单名称"),
    ("N Markers",         "版数"),
    ("Min Plies",         "最少层数"),
    ("Max Plies",         "最多层数"),
    ("File Name",         "文件名"),
    ("Width, cm",         "幅宽,cm"),
    ("Length, cm",        "长度,cm"),
    ("Efficiency,%",      "利用率,%"),
    ("PO Style(s)",       "订单款号"),
    ("Spreading",         "铺布"),
    ("Solution",          "排版结果"),
    ("Material Cost",     "材料成本"),   # MUST precede the bare "Material"
    ("Material",          "面料"),
    ("Quantity",          "数量"),
    ("Sizes",             "尺码"),
    ("Client",            "客户"),
    ("Sum",               "合计"),
    ("Date",              "日期"),
    ("Time",              "时间"),

    # ── The original macro's own rules, in its order ─────────────────────
    ("Plies number",  "层数"),
    ("Marker Length", "版长"),
    ("Colors",        "颜色"),
    ("Order",         "订单数"),
    ("Real",          "裁剪数"),
    ("Marker Ratio",  "裁剪配比"),
    ("Marker",        "版"),
    ("Fabric Length", "面料长度"),
    ("Fabric Weight", "面料重量"),
    # Added beyond the original macro. "Total cut length" MUST precede
    # "Cut Length" — the sheet has both, and the short rule would otherwise
    # leave the total reading "Total 裁剪长度".
    ("Total cut length", "总裁剪长度"),
    ("Cut Length",       "裁剪长度"),
]

_COMPILED: list[tuple[re.Pattern, str]] = [
    (re.compile(re.escape(src), re.IGNORECASE), dst) for src, dst in TRANSLATIONS
]

MARKER_EXT = ".mrk"

_CJK = re.compile(r"[一-鿿]")


def has_chinese(text: str) -> bool:
    """True when *text* already contains a Chinese character."""
    return bool(_CJK.search(text))


def _norm_color(text: str) -> str:
    """Case/space-insensitive colour key — matches the colour store's own
    normalisation ("NAVY", "navy", " Navy " all collapse together)."""
    return " ".join(str(text).split()).casefold()


def build_color_map(lookup: dict, client: str = "") -> dict[str, str]:
    """Flatten the colour store's ``{(client, brand, en): cn}`` into
    ``{normalised_en: cn}`` for whole-cell colour replacement.

    A colour can be mapped differently per client/brand, so rows for *client*
    are applied last and win; everything else is a fallback. Blank Chinese
    names are skipped so a half-filled row can't blank a colour out.
    """
    out: dict[str, str] = {}
    for scope in (False, True):          # others first, then the client's own
        for key, cn in (lookup or {}).items():
            if not cn:
                continue
            row_client = key[0] if isinstance(key, tuple) and key else ""
            en = key[2] if isinstance(key, tuple) and len(key) > 2 else ""
            if not en:
                continue
            is_client = bool(client) and str(row_client) == client
            if is_client == scope:
                out[_norm_color(en)] = cn
    return out


def translate_text(text: str) -> str:
    """Apply the label replacements to *text* (case-insensitive, substring)."""
    for pattern, repl in _COMPILED:
        # lambda, not the raw string: a replacement is inserted literally so
        # backslashes/group refs in it can never be interpreted.
        text = pattern.sub(lambda _m, r=repl: r, text)
    return text


def strip_path_and_ext(text: str) -> str:
    """Reduce a marker path to its bare name.

    ``D:\\Markers\\SS26\\ABC-1234.mrk`` -> ``ABC-1234``. Keeps text without a
    path or extension untouched. The extension check is case-insensitive
    (the macro's was not — ``.MRK`` slipped through it).
    """
    if "\\" in text:
        text = text.rsplit("\\", 1)[-1]
    if text.lower().endswith(MARKER_EXT):
        text = text[: -len(MARKER_EXT)]
    return text


def build_reverse_color_map(lookup: dict) -> dict[str, str]:
    """``{normalised chinese: normalised english}`` across every client/brand.

    Used to recognise a Chinese colour the *plan* already carries so it can be
    compared against what the PO calls that colour.
    """
    out: dict[str, str] = {}
    for key, cn in (lookup or {}).items():
        if not cn:
            continue
        en = key[2] if isinstance(key, tuple) and len(key) > 2 else ""
        if en:
            out.setdefault(_norm_color(cn), _norm_color(en))
    return out


def find_color_conflicts(wb, color_map: dict[str, str],
                         reverse_map: dict[str, str]) -> list[dict]:
    """Report cells where the plan's own Chinese colour disagrees with the PO's.

    **Read-only — changes nothing.** A cut plan that already names a colour in
    Chinese may be using a different house spelling than the PO (e.g. the plan
    says 深蓝 where the PO says 藏青 for Navy). Those are surfaced for a human
    to decide on rather than silently rewritten.

    Returns one entry per distinct disagreement::

        {"plan_cn", "po_cn", "english", "sheet", "cell"}
    """
    seen: set[tuple[str, str]] = set()
    out: list[dict] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not isinstance(v, str) or not has_chinese(v):
                    continue
                key = _norm_color(v)
                en = reverse_map.get(key)
                if not en:
                    continue          # not a colour we know — leave it be
                po_cn = (color_map or {}).get(en)
                if not po_cn or _norm_color(po_cn) == key:
                    continue          # agrees, or the PO has no name for it
                sig = (key, _norm_color(po_cn))
                if sig in seen:
                    continue
                seen.add(sig)
                out.append({"plan_cn": v.strip(), "po_cn": po_cn,
                            "english": en, "sheet": ws.title,
                            "cell": cell.coordinate})
    return out


def clean_value(value: Any, color_map: dict[str, str] | None = None,
                cn_overrides: dict[str, str] | None = None) -> Any:
    """Clean one cell value; non-text and formulas pass through untouched.

    When *color_map* is given, a cell whose **whole** value is a known English
    colour becomes its Chinese name from the PO colour data. Whole-cell only —
    a substring rule here would corrupt any text that merely contains a colour
    word.

    A colour that is **already Chinese** is left alone unless *cn_overrides*
    maps it (``{normalised plan colour: replacement}``) — that only happens
    once the user has explicitly approved the change.
    """
    if not isinstance(value, str):
        return value
    if value.startswith("="):        # never rewrite a formula
        return value
    key = _norm_color(value)
    if cn_overrides:
        repl = cn_overrides.get(key)
        if repl:
            return repl              # user-approved colour rename
    if color_map and not has_chinese(value):
        cn = color_map.get(key)
        if cn:
            return cn                # a colour cell is only ever the colour
    return strip_path_and_ext(translate_text(value))


# Header rows the cutting room doesn't use. Dropped from the delivered copy
# only — the canonical sheet keeps them because parsers.cutting_plan reads
# Order name / Date / Time back out of it. Matched on the label in column A,
# before translation, so these are the English originals.
# Kept on the sheet: Date, Order name, Style file N, Cut plan operator, Client
# — the cutting room uses those to identify the plan.
HEADER_ROWS_DROPPED = (
    "output folder path",
)


def _is_dropped_header(label: str) -> bool:
    s = " ".join(str(label).split()).casefold()
    # "Style name 1", "Style name 2", … — the "Style file N" rows kept above
    # already carry the style, so this is a duplicate of it.
    return s in HEADER_ROWS_DROPPED or s.startswith("style name")


# Metric columns the cutting room doesn't use. Matched on the heading before
# translation, so these are the English originals as the marker software wrote
# them. The cost column also repeats the fabric length rather than a cost.
COLUMNS_DROPPED = ("cut length", "material cost")


def _is_dropped_column(label: str) -> bool:
    s = " ".join(str(label).split()).casefold()
    return any(s.startswith(h) for h in COLUMNS_DROPPED)


def _has_text(value) -> bool:
    return isinstance(value, str) and bool(value.strip())


def _run_bounds(ws, row: int, col: int) -> tuple[int, int]:
    """The contiguous run of labelled columns in *row* that *col* belongs to.

    A block's metric headings sit shoulder to shoulder at the right-hand end of
    its top row, so the run is bounded by the first empty cell on either side.
    The wide merged ``Sizes`` heading to their left reads empty on every cell
    but its first, which is what stops the walk.
    """
    lo = col
    while lo > 1 and _has_text(ws.cell(row, lo - 1).value):
        lo -= 1
    hi = col
    while hi < ws.max_column and _has_text(ws.cell(row, hi + 1).value):
        hi += 1
    return lo, hi


def _run_last_row(ws, header_row: int, lo: int, hi: int) -> int:
    """Last row of the block headed by the run at *header_row*.

    The block runs down to the row before the next label in the same columns —
    that label is the next block's own heading.
    """
    for r in range(header_row + 1, ws.max_row + 1):
        if any(_has_text(ws.cell(r, c).value) for c in range(lo, hi + 1)):
            return r - 1
    return ws.max_row


def _pack_run(ws, *, top: int, bot: int, lo: int, hi: int,
              keep: list[int]) -> None:
    """Left-pack the *keep* columns into ``lo…`` over rows *top*–*bot*.

    Only the run's own rectangle is touched — values, styles and the merges
    contained in it. That is the whole point: a sheet column carries a metric
    in one block and a size in another, so emptying or deleting the column
    outright can only ever come out right for one of them.
    """
    from copy import copy

    snap = {(r, c): (ws.cell(r, c).value, copy(ws.cell(r, c)._style))
            for r in range(top, bot + 1) for c in range(lo, hi + 1)}
    inside = [m for m in list(ws.merged_cells.ranges)
              if lo <= m.min_col and m.max_col <= hi
              and top <= m.min_row and m.max_row <= bot]
    for m in inside:
        ws.unmerge_cells(str(m))

    moved = {src: lo + i for i, src in enumerate(keep)}
    for r in range(top, bot + 1):
        for src, dst in moved.items():
            value, style = snap[(r, src)]
            cell = ws.cell(r, dst)
            cell.value = value
            cell._style = style
        # The tail the pack vacated. Its borders and fill go too — an empty
        # bordered cell is exactly the gap being closed.
        for c in range(lo + len(keep), hi + 1):
            cell = ws.cell(r, c)
            cell.value = None
            cell._style = None

    for m in inside:
        if m.min_col not in moved or m.max_col not in moved:
            continue                               # a dropped column's own
        a, b = moved[m.min_col], moved[m.max_col]
        if a == b and m.min_row == m.max_row:
            continue                               # a 1x1 "merge" is not one
        ws.merge_cells(start_row=m.min_row, start_column=a,
                       end_row=m.max_row, end_column=b)


def compact_dropped_columns(ws) -> int:
    """Remove the unwanted metric columns from every block. Returns the count.

    Each block writes its metrics immediately to the right of *its own* size
    columns, so two fabrics with different size counts put ``Cut Length,m`` on
    two different sheet columns — and the column one block uses for a metric
    the other uses for a size. Clearing the cells and then deleting the empty
    sheet column therefore only ever worked for the widest block; every
    narrower one kept a blank gap where its metric had been.

    So the gap is closed inside each block instead: the block's run of metric
    columns is packed to the left over the block's own rows, leaving the rest
    of the sheet — including whatever another block keeps in those columns —
    exactly where it was.
    """
    removed = 0
    for r in range(1, ws.max_row + 1):
        done: set[int] = set()
        for c in range(1, ws.max_column + 1):
            if c in done:
                continue
            value = ws.cell(r, c).value
            if not isinstance(value, str) or not _is_dropped_column(value):
                continue
            lo, hi = _run_bounds(ws, r, c)
            done.update(range(lo, hi + 1))
            keep = [k for k in range(lo, hi + 1)
                    if not _is_dropped_column(str(ws.cell(r, k).value or ""))]
            _pack_run(ws, top=r, bot=_run_last_row(ws, r, lo, hi),
                      lo=lo, hi=hi, keep=keep)
            removed += (hi - lo + 1) - len(keep)
    return removed


def _delete_line(ws, *, at: int, axis: str) -> None:
    """Delete one row/column and carry the merged ranges with it.

    openpyxl's ``delete_rows``/``delete_cols`` move the CELLS but leave merged
    ranges on their old coordinates — which is what slid the marker blocks out
    from under their headings the first time this was tried. The ranges have to
    be released BEFORE the delete (afterwards the cells they name no longer
    exist and unmerge raises), then re-laid on the new coordinates: ranges after
    the deleted line move back one, ranges spanning it shrink by one, ranges
    before it are untouched.
    """
    from openpyxl.utils import get_column_letter
    lo_a, hi_a = ("min_row", "max_row") if axis == "row" else ("min_col", "max_col")

    saved = [dict(min_row=r.min_row, max_row=r.max_row,
                  min_col=r.min_col, max_col=r.max_col)
             for r in list(ws.merged_cells.ranges)]
    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))

    (ws.delete_rows if axis == "row" else ws.delete_cols)(at)

    for b in saved:
        if b[hi_a] >= at:
            if b[lo_a] > at:
                b[lo_a] -= 1
            b[hi_a] -= 1
        if b["min_row"] > b["max_row"] or b["min_col"] > b["max_col"]:
            continue                               # collapsed to nothing
        if b["min_row"] == b["max_row"] and b["min_col"] == b["max_col"]:
            continue                               # a 1x1 "merge" is not one
        ws.merge_cells(
            f"{get_column_letter(b['min_col'])}{b['min_row']}:"
            f"{get_column_letter(b['max_col'])}{b['max_row']}")


def find_fabric_blocks(ws) -> list[tuple[str, int, int]]:
    """Locate each fabric's section: ``[(material, first_row, last_row), …]``.

    The sheet repeats one section per fabric — ``Marker Definition`` (whose
    table names the material), then Marker Ratio, Spreading Plies, Solution
    and that fabric's own totals, ending at its ``Total Tables`` row. The very
    last ``Total Tables`` on the sheet is the plan-wide grand total and belongs
    to no fabric, so it is never part of a block.

    Matched on the English anchors, i.e. before translation.
    """
    def label(r: int) -> str:
        return " ".join(str(ws.cell(r, 1).value or "").split()).casefold()

    starts = [r for r in range(1, ws.max_row + 1)
              if label(r) == "marker definition"]
    totals = [r for r in range(1, ws.max_row + 1) if label(r) == "total tables"]
    if not starts:
        return []

    # The sheet carries one Total Tables per fabric plus a plan-wide one at the
    # end. When there is one more than there are fabrics, that last one is the
    # grand total and must stay out of every block — otherwise removing the
    # final fabric takes the plan's own total with it.
    grand = totals[-1] if len(totals) > len(starts) else None

    blocks: list[tuple[str, int, int]] = []
    for i, first in enumerate(starts):
        limit = starts[i + 1] - 1 if i + 1 < len(starts) else ws.max_row
        if grand is not None and first < grand <= limit:
            limit = grand - 1
        ends = [t for t in totals if first < t <= limit and t != grand]
        last = ends[-1] if ends else limit
        # A fabric's own "Average Length" is printed one row BELOW its Total
        # Tables, so the rule above stops just short of it. Left out of the
        # block it survives its fabric's removal and prints as a stray figure
        # under the plan totals.
        for r in range(last + 1, limit + 1):
            lbl = label(r)
            if not lbl and ws.cell(r, 2).value is None:
                continue                  # blank spacer — keep looking
            if lbl.startswith("average length"):
                last = r
                continue
            break
        # material name: first non-empty column-A cell below the block's own
        # header row that isn't the header label itself
        name = ""
        for r in range(first + 1, last + 1):
            v = ws.cell(r, 1).value
            if v is None:
                continue
            s = " ".join(str(v).split())
            if s.casefold() in ("material", "marker definition"):
                continue
            name = s
            break
        blocks.append((name or "—", first, last))
    return blocks


def keep_only_fabrics(ws, keep: set[str]) -> int:
    """Remove the sections of fabrics not in *keep*. Returns blocks removed.

    Rows are deleted through :func:`_delete_line`, so merged ranges move with
    them. Bottom-up, so the earlier blocks' row numbers stay valid. Anything
    outside the fabric sections — the header, Order demands and the grand
    total — is untouched.
    """
    blocks = find_fabric_blocks(ws)
    if not blocks or not keep:
        return 0
    drop = [b for b in blocks if b[0] not in keep]
    if len(drop) == len(blocks):
        return 0                      # never blank the sheet entirely
    for _name, first, last in sorted(drop, key=lambda b: b[1], reverse=True):
        for _ in range(last - first + 1):
            _delete_line(ws, at=first, axis="row")

    # The plan-wide Total Tables counted every fabric. Leaving it would print
    # "6" on a sheet showing one fabric's four tables, so re-add the totals of
    # the blocks that are actually still on the sheet.
    def _label(r):
        return " ".join(str(ws.cell(r, 1).value or "").split()).casefold()
    totals = [r for r in range(1, ws.max_row + 1) if _label(r) == "total tables"]
    kept = find_fabric_blocks(ws)
    if totals and len(totals) > len(kept):
        per_block = [ws.cell(t, 2).value for t in totals[:-1]]
        nums = [v for v in per_block if isinstance(v, (int, float))
                and not isinstance(v, bool)]
        if nums:
            ws.cell(totals[-1], 2).value = sum(nums)
    return len(drop)


# Block titles that should each start on their own, one blank row clear of
# whatever precedes them. Matched after translation, so these are the Chinese.
_BLOCK_TITLES = ("订单需求", "版定义", "裁剪配比", "铺布层数", "排版结果")


def space_out_blocks(ws) -> int:
    """Put exactly one blank row above each block title. Returns rows inserted.

    Trimming and fabric removal can leave a block butted straight up against
    the one before it, which reads as one run-on table. Rows are inserted
    bottom-up so earlier row numbers stay valid, and merged ranges are shifted
    with them — openpyxl's insert_rows moves the cells but not the merges,
    the same trap as deleting.
    """
    from openpyxl.utils import get_column_letter

    def title_row(r: int) -> bool:
        # Some blocks are indented — "Order demands" starts in column B — so
        # look at the first couple of columns, not just column A.
        return any(" ".join(str(ws.cell(r, c).value or "").split()) in _BLOCK_TITLES
                   for c in (1, 2))

    targets = [r for r in range(2, ws.max_row + 1)
               if title_row(r) and not _is_blank_line(ws, r - 1, "row")]
    for r in sorted(targets, reverse=True):
        saved = [dict(min_row=g.min_row, max_row=g.max_row,
                      min_col=g.min_col, max_col=g.max_col)
                 for g in list(ws.merged_cells.ranges)]
        for g in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(g))
        ws.insert_rows(r)
        for b in saved:
            if b["min_row"] >= r:
                b["min_row"] += 1
            if b["max_row"] >= r:
                b["max_row"] += 1
            ws.merge_cells(
                f"{get_column_letter(b['min_col'])}{b['min_row']}:"
                f"{get_column_letter(b['max_col'])}{b['max_row']}")
    return len(targets)


def _is_blank_line(ws, idx: int, axis: str) -> bool:
    """True when the row/column holds no value AND no merged range needs it.

    A line can be empty of values and still be structural — the second half of
    a merged heading carries no value of its own. Removing it would close a gap
    that isn't there.
    """
    for rng in ws.merged_cells.ranges:
        lo, hi = ((rng.min_row, rng.max_row) if axis == "row"
                  else (rng.min_col, rng.max_col))
        # A merge confined to this line alone is the line's OWN heading merge —
        # a dropped column keeps one after its text is cleared, and treating
        # that as structural pinned the empty column open. Only a merge that
        # reaches beyond the line is holding the layout.
        if lo <= idx <= hi and not (lo == hi == idx):
            return False
    if axis == "row":
        return all(ws.cell(idx, c).value is None
                   for c in range(1, ws.max_column + 1))
    return all(ws.cell(r, idx).value is None
               for r in range(1, ws.max_row + 1))


def close_blanked_gaps(ws, *, header_limit: int = 30) -> tuple[int, int]:
    """Close the gaps the blanking left behind. Returns (rows, cols) removed.

    Only fully empty lines are removed, and only ones this cleanup emptied:
    columns anywhere (the dropped metric columns), rows only within the header
    block — a blank row further down separates the marker blocks and is part of
    the layout. Deleting bottom-up/right-to-left so earlier indices stay valid,
    and merged ranges are re-laid after each one.
    """
    cols = [c for c in range(ws.max_column, 0, -1)
            if _is_blank_line(ws, c, "col")]
    for c in cols:
        _delete_line(ws, at=c, axis="col")
    rows = [r for r in range(min(ws.max_row, header_limit), 0, -1)
            if _is_blank_line(ws, r, "row")]
    for r in rows:
        _delete_line(ws, at=r, axis="row")
    return len(rows), len(cols)


def drop_header_rows(ws) -> int:
    """Blank the header rows the cutting room doesn't read. Returns the count.

    **Clears cells; never deletes rows.** Deleting a row pulls everything below
    it up one, which slid the marker blocks out from under their own headings
    and wrecked the layout. Emptying the cells leaves every other cell exactly
    where it was — the row is simply blank.

    Only scans the top of the sheet, so a data cell that happens to read "Date"
    further down is left alone.
    """
    limit = min(ws.max_row, 30)
    cleared = 0
    for r in range(1, limit + 1):
        if not _is_dropped_header(ws.cell(r, 1).value or ""):
            continue
        for c in range(1, ws.max_column + 1):
            if ws.cell(r, c).value is not None:
                ws.cell(r, c).value = None
        cleared += 1
    return cleared


def clean_worksheet(ws, color_map: dict[str, str] | None = None,
                    cn_overrides: dict[str, str] | None = None) -> int:
    """Clean every cell of *ws* in place. Returns the number of cells changed."""
    # Blank rows and columns FIRST, while the labels are still their English
    # originals — the matchers key on the English text.
    changed = drop_header_rows(ws) + compact_dropped_columns(ws)
    for row in ws.iter_rows():
        for cell in row:
            old = cell.value
            if not isinstance(old, str):
                continue
            new = clean_value(old, color_map, cn_overrides)
            if new != old:
                cell.value = new
                changed += 1
    return changed


def clean_workbook(wb, color_map: dict[str, str] | None = None,
                   cn_overrides: dict[str, str] | None = None) -> int:
    """Clean every worksheet of *wb* in place. Returns total cells changed."""
    total = 0
    for ws in wb.worksheets:
        total += clean_worksheet(ws, color_map, cn_overrides)
        # Close the empty column/rows the blanking left, then give each block
        # one clear row above it. Order matters: closing first collapses
        # everything, spacing then re-opens exactly one row per block.
        close_blanked_gaps(ws)
        space_out_blocks(ws)
    return total


# ── working on an already-built workbook (the UI's approve-then-apply path) ──

def conflicts_from_bytes(data: bytes, color_map: dict[str, str],
                         reverse_map: dict[str, str]) -> list[dict]:
    """:func:`find_color_conflicts` against built .xlsx bytes."""
    import openpyxl
    from io import BytesIO
    return find_color_conflicts(openpyxl.load_workbook(BytesIO(data)),
                                color_map, reverse_map)


def apply_color_overrides(data: bytes, overrides: dict[str, str]) -> bytes:
    """Return *data* with approved colour renames applied.

    Re-cleans the finished workbook instead of rebuilding it — the overrides
    only ever affect the cleanup pass, so nothing else can shift.
    """
    import openpyxl
    from io import BytesIO
    wb = openpyxl.load_workbook(BytesIO(data))
    clean_workbook(wb, color_map=None, cn_overrides=overrides)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
