"""Fabric consumption (单耗 / 排版) data — reconciliation + upload template.

The consumption figures live in their own table, populated by the operator via
an Excel template (download blank → fill → upload). The buy plan reads them
per style and appends them as the trailing columns.

kg ↔ cm derivation
------------------
Consumption in **cm** is the length of fabric one garment uses on the marker;
consumption in **kg** is its weight. Weight is billed on the **gross** width
(毛门幅), which is the usable/effective marker width (排版有效门幅) plus the
selvage — a fixed ``毛门幅 = 有效门幅 + 5cm``. So the conversion uses the gross
width and the fabric weight (克重, g/m²)::

    毛门幅 = 有效门幅 + 5
    kg = cm × 毛门幅 × gsm / 1e7           # area(m²) × gsm(g/m²) / 1000
    cm = kg × 1e7 / (毛门幅 × gsm)

:func:`reconcile_consumption` (1) checks the two agree when both are given and
(2) fills whichever is missing from the other — but only when the effective
width and gsm are present to convert; otherwise it leaves the gap and reports
why.
"""
from __future__ import annotations

import io

# 毛门幅 (gross, billed width) = 排版有效门幅 (effective marker width) + selvage.
GROSS_WIDTH_MARGIN_CM = 5.0

# Template column order (Chinese headers, matching the buy plan) → record key.
CONSUMPTION_COLUMNS: list[tuple[str, str]] = [
    ("款号",              "style"),
    ("单耗(kg)",          "cons_kg"),
    ("单耗(cm)",          "cons_cm"),
    ("排版利用率(%)",     "util"),
    ("排版件数",          "marker_pcs"),
    ("排版有效门幅(cm)",  "width_cm"),
    ("排版面料克重(g/m²)", "gsm"),
]
_CONS_HEADERS = [h for h, _ in CONSUMPTION_COLUMNS]
_CONS_KEYS = [k for _, k in CONSUMPTION_COLUMNS]

_KG_TOL = 0.05   # 5% consistency tolerance between provided and derived kg


def _num(v):
    """Parse a numeric cell → float, or None when blank/non-numeric."""
    if v is None:
        return None
    s = str(v).replace(",", "").replace("%", "").strip()
    if not s or s.lower() in ("nan", "none", "-"):
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _fmt(v, nd) -> str:
    if v is None:
        return ""
    r = round(v, nd)
    return str(int(r)) if r == int(r) else str(r)


def reconcile_consumption(cons_kg, cons_cm, width_cm, gsm, tol: float = _KG_TOL):
    """Return ``(kg, cm, warning)`` — kg/cm reconciled, warning '' when clean.

    * Both given + convertible → keep both, warn if they disagree by > *tol*.
    * One given + convertible   → derive the other.
    * One given, not convertible (no width/gsm) → leave the gap, warn.
    """
    kg, cm = _num(cons_kg), _num(cons_cm)
    w, g = _num(width_cm), _num(gsm)
    convertible = bool(w and g and w > 0 and g > 0)

    if convertible:
        gw = w + GROSS_WIDTH_MARGIN_CM   # 毛门幅 — weight is billed on the gross width
        if kg is not None and cm is not None:
            derived = cm * gw * g / 1e7
            if derived > 0 and abs(kg - derived) / derived > tol:
                return kg, cm, (
                    f"单耗不一致：提供 {_fmt(kg,4)}kg，按 {_fmt(cm,1)}cm×毛门幅"
                    f"{_fmt(gw,1)}cm×{_fmt(g,1)}g/m² 应为 {_fmt(derived,4)}kg（相差>{int(tol*100)}%）")
            return kg, cm, ""
        if cm is not None and kg is None:
            return round(cm * gw * g / 1e7, 4), cm, ""
        if kg is not None and cm is None:
            return kg, round(kg * 1e7 / (gw * g), 1), ""
        return kg, cm, ""   # neither given
    # not convertible
    if (kg is None) != (cm is None):
        return kg, cm, "缺少 门幅/克重，无法由单一单耗推算另一单耗"
    return kg, cm, ""


def reconcile_record(rec: dict) -> tuple[dict, str]:
    """Reconcile one consumption record in place; return (record, warning)."""
    kg, cm, warn = reconcile_consumption(
        rec.get("cons_kg"), rec.get("cons_cm"),
        rec.get("width_cm"), rec.get("gsm"))
    out = dict(rec)
    out["cons_kg"], out["cons_cm"] = kg, cm
    return out, warn


# ── template download / upload ────────────────────────────────────────────────

def consumption_template_bytes(rows: list[dict] | None = None) -> bytes:
    """Blank template (or current data) as an .xlsx the operator fills in.

    An example row shows the expected shape; passing *rows* exports the stored
    data instead (for download-edit-reupload)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "单耗排版"
    hdr_fill = PatternFill("solid", fgColor="BDD7EE")
    for i, h in enumerate(_CONS_HEADERS, start=1):
        c = ws.cell(1, i, h)
        c.font = Font(bold=True)
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[c.column_letter].width = max(12, len(h) + 2)

    if rows:
        for r, rec in enumerate(rows, start=2):
            for i, k in enumerate(_CONS_KEYS, start=1):
                ws.cell(r, i, rec.get(k))
    else:
        # one illustrative example (kg omitted on purpose — derived from cm)
        example = {"style": "S5DTN67A", "cons_kg": None, "cons_cm": 165,
                   "util": 82, "marker_pcs": 24, "width_cm": 150, "gsm": 200}
        for i, k in enumerate(_CONS_KEYS, start=1):
            ws.cell(2, i, example.get(k))

    buf = io.BytesIO()
    # Any '=' text that arrived in the data becomes inert here; the
    # exporter's own =SUM()/='Sheet'! formulas are left alone.
    # Imported here, not at module scope: po_extractor.exporters'
    # __init__ imports back into this package, so a module-level
    # import makes a cycle and whichever side loads first fails.
    from ..exporters._excel_helpers import neutralise_foreign_formulas
    neutralise_foreign_formulas(wb)
    wb.save(buf)
    return buf.getvalue()


def parse_consumption_upload(data: bytes) -> tuple[list[dict], list[str]]:
    """Parse an uploaded consumption template → (records, warnings).

    Maps by header name (order-independent), keyed by 款号; rows without a
    style are skipped. Each record is reconciled (kg↔cm) and its warning
    collected."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), data_only=True)
    ws = wb.active
    header_map: dict[int, str] = {}
    header_row = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    hdr_to_key = dict(CONSUMPTION_COLUMNS)
    for ci, h in enumerate(header_row, start=1):
        key = hdr_to_key.get(str(h).strip()) if h is not None else None
        if key:
            header_map[ci] = key
    if "style" not in header_map.values():
        return [], ["模板缺少『款号』列 — 无法导入。"]

    records: list[dict] = []
    warnings: list[str] = []
    seen: set[str] = set()
    for r in range(2, ws.max_row + 1):
        rec = {k: None for k in _CONS_KEYS}
        for ci, key in header_map.items():
            rec[key] = ws.cell(r, ci).value
        style = str(rec.get("style") or "").strip()
        if not style:
            continue
        if style in seen:
            warnings.append(f"款号 {style} 出现多次 — 使用最后一行。")
        seen.add(style)
        rec["style"] = style
        reconciled, warn = reconcile_record(rec)
        if warn:
            warnings.append(f"款号 {style}：{warn}")
        records.append(reconciled)
    return records, warnings
