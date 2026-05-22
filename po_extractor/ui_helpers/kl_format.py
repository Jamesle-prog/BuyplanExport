"""KL-format PO Summary Excel generator.

Produces a two-sheet workbook matching the KL reference format:
  PO Detail  — one row per PO / Style / Color / Size (21 columns)
  Summary    — Part A: pivot per PO; Part B: style rollup

Input:
  df_meta  — from POStore.list_pos() filtered to the desired PO numbers
  df_sizes — from POStore.load_size_rows() for the same PO numbers

See docs/GIII_PO_Summary_KL_Format.md for the full specification.
"""
from __future__ import annotations

import io
import math
from math import gcd
from functools import reduce

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Size ordering ──────────────────────────────────────────────────────────────
KL_SIZE_ORDER: list[str] = [
    "PXS", "PS", "PM", "PL", "PXL", "P1X", "P2X", "P3X",
    "XXS", "XS", "S", "M", "L", "XL", "XXL", "XXXL",
    "0X", "1X", "2X", "3X", "4X",
]

# ── Known ship-to addresses by customer name ───────────────────────────────────
_KNOWN_SHIP_TO: dict[str, str] = {
    "ROSS STORES": (
        "ROSS STORES / 3404 INDIAN AVE / DISTRIBUTION CENTER / PERRIS,CA 92571"
    ),
}

# ── Openpyxl style constants ───────────────────────────────────────────────────
_HDR_FILL   = PatternFill("solid", start_color="366092", end_color="366092")
_HDR_FONT   = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
_ALT_FILL   = PatternFill("solid", start_color="DCE6F1", end_color="DCE6F1")
_DATA_FONT  = Font(name="Calibri", size=10)
_GRAND_FILL = PatternFill("solid", start_color="FFC000", end_color="FFC000")
_GRAND_FONT = Font(bold=True, name="Calibri", size=10)
_THIN = Border(
    left=Side(style="thin", color="AAAAAA"),
    right=Side(style="thin", color="AAAAAA"),
    top=Side(style="thin", color="AAAAAA"),
    bottom=Side(style="thin", color="AAAAAA"),
)

# ── Detail sheet column widths (matches 21-col KL layout) ─────────────────────
_DETAIL_COL_WIDTHS = [14, 12, 22, 6, 8, 15, 16, 8, 14, 6, 10, 10,
                      14, 55, 38, 10, 14, 8, 26, 30, 24]


# ── Internal helpers ──────────────────────────────────────────────────────────

def _s(val) -> str:
    """Safe string: converts None / NaN to ''."""
    if val is None:
        return ""
    if isinstance(val, float):
        import math
        if math.isnan(val):
            return ""
    return str(val).strip()


def _etd_display(raw: str) -> str:
    """'8/01/2026' or '8/01/26' → '8/1'."""
    if not raw:
        return ""
    parts = raw.split("/")
    try:
        return f"{int(parts[0])}/{int(parts[1])}"
    except (IndexError, ValueError):
        return raw


def _pack_ratio(sizes: pd.DataFrame, size_order: list[str]) -> str:
    """Return '1-2-2-1' style ratio from a slice of size rows for one PO.

    Handles both capitalised column names ('Size'/'Units') and lower-case
    ('size'/'units') since the caller renames them to lower-case after loading.
    """
    size_col  = "size"  if "size"  in sizes.columns else "Size"
    units_col = "units" if "units" in sizes.columns else "Units"
    counts = dict(zip(sizes[size_col], sizes[units_col].astype(int)))
    ordered = [counts[s] for s in size_order if s in counts]
    if not ordered:
        return ""
    g = reduce(gcd, ordered)
    return "-".join(str(v // g) for v in ordered)


def _resolve_ship_to(ship_to_db: str, customer: str) -> str:
    """Use stored ship_to if available; else fall back to known-address dict."""
    if ship_to_db and ship_to_db.strip():
        return ship_to_db.strip()
    if customer:
        cust_upper = customer.strip().upper()
        for key, addr in _KNOWN_SHIP_TO.items():
            if key in cust_upper or cust_upper in key:
                return addr
    return _s(customer)


def _write_header(ws, headers: list, row: int = 1) -> None:
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font, c.fill, c.border = _HDR_FONT, _HDR_FILL, _THIN
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 28


def _write_row(ws, row_idx: int, values: list, *, is_grand: bool = False, alt: bool = False) -> None:
    fill = _GRAND_FILL if is_grand else (_ALT_FILL if alt else None)
    font = _GRAND_FONT if is_grand else _DATA_FONT
    for ci, val in enumerate(values, 1):
        c = ws.cell(row=row_idx, column=ci, value=val)
        c.font, c.border = font, _THIN
        if fill:
            c.fill = fill
        c.alignment = Alignment(
            horizontal="right" if isinstance(val, (int, float)) else "left",
            vertical="center",
        )


# ── Public API ────────────────────────────────────────────────────────────────

def generate_kl_format_excel(
    df_meta: pd.DataFrame,
    df_sizes: pd.DataFrame,
) -> bytes:
    """Build the KL-format two-sheet Excel workbook and return its bytes.

    Parameters
    ----------
    df_meta:
        Metadata DataFrame from ``POStore.list_pos()``.  Expected columns include
        po_number, style, unit_cost, xport_date, issue_date, factory, seller,
        style_group, hanger, description_code, customer, ship_to, packaging.
    df_sizes:
        Size rows from ``POStore.load_size_rows()``.  Expected columns:
        PO Number, Style, Color, Size, Units, UPC.

    Returns
    -------
    bytes
        Raw .xlsx content ready to send as a Streamlit download or write to disk.
    """
    if df_meta is None or df_meta.empty:
        return b""

    # ── Normalise column names (list_pos uses snake_case) ────────────────────
    meta = df_meta.copy()
    sizes = df_sizes.copy() if df_sizes is not None and not df_sizes.empty else pd.DataFrame()

    # Rename size-rows columns to match internal usage
    if not sizes.empty:
        sizes = sizes.rename(columns={
            "PO Number": "po_number",
            "Style": "style",
            "Color": "color",
            "Size": "size",
            "Units": "units",
            "UPC": "upc",
        })

    # ── Build detail rows ─────────────────────────────────────────────────────
    detail_rows: list[dict] = []

    for _, m in meta.iterrows():
        pn     = _s(m.get("po_number"))
        # NaN is truthy in Python: `NaN or 0` returns NaN, not 0.
        # Use pd.notna() to guard before converting so fob is always a
        # real float (0.0 when the price is missing/unconfirmed).
        fob = 0.0
        try:
            raw = m.get("unit_cost")
            if pd.notna(raw) and raw is not None:
                fob = float(raw)
        except (TypeError, ValueError):
            pass

        etd = _etd_display(_s(m.get("xport_date")))
        po_date = _s(m.get("issue_date"))
        desc    = _s(m.get("description_code"))
        hanger  = _s(m.get("hanger"))
        customer = _s(m.get("customer"))
        ship_to  = _resolve_ship_to(_s(m.get("ship_to")), customer)
        cpo      = _s(m.get("cpo")) or "TBA"
        msrp_val = _s(m.get("msrp"))

        # Size rows for this PO
        po_sizes = sizes[sizes["po_number"] == pn] if not sizes.empty else pd.DataFrame()
        pack_ratio = _pack_ratio(po_sizes, KL_SIZE_ORDER) if not po_sizes.empty else ""

        for _, r in po_sizes.iterrows():
            units = int(r.get("units", 0))
            detail_rows.append({
                "PO Number":        pn,
                "Style":            _s(r.get("style") or m.get("style")),
                "Color":            _s(r.get("color")),
                "Size":             _s(r.get("size")),
                "Units":            units,
                "UPC":              _s(r.get("upc")),
                "Unit Price (FOB)": fob,
                "MSRP":             msrp_val,
                "Line Total ($)":   round(units * fob, 2),
                "ETD":              etd,
                "PO Date":          po_date,
                "Ship Date":        po_date,
                "Customer Name":    customer or "ROSS STORES",
                "Ship To":          ship_to,
                "Hanger Info":      hanger,
                "Pack Ratio":       pack_ratio,
                "HTS#":             _s(m.get("style_group")),
                "CPO":              cpo,
                "Description":      desc,
                "Factory":          _s(m.get("factory")),
                "Vendor":           _s(m.get("seller")),
            })

    if not detail_rows:
        return b""

    df = pd.DataFrame(detail_rows)

    # Sort both sheets by PO Number so rows are always in a predictable order.
    _sz_order = {s: i for i, s in enumerate(KL_SIZE_ORDER)}
    df["_size_rank"] = df["Size"].map(lambda s: _sz_order.get(s, 999))
    df = df.sort_values(["PO Number", "Color", "_size_rank"]).reset_index(drop=True)
    df = df.drop(columns=["_size_rank"])

    # ── Workbook ──────────────────────────────────────────────────────────────
    wb = Workbook()

    # ════════════════════════════════════════════════════════════════════════
    # Sheet 1: PO Detail
    # ════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "PO Detail"

    detail_cols = [
        "PO Number", "Style", "Color", "Size", "Units", "UPC",
        "Unit Price (FOB)", "MSRP", "Line Total ($)", "ETD",
        "PO Date", "Ship Date", "Customer Name", "Ship To",
        "Hanger Info", "Pack Ratio", "HTS#", "CPO",
        "Description", "Factory", "Vendor",
    ]
    _write_header(ws1, detail_cols)

    for ri, row in df.iterrows():
        vals = [row[c] for c in detail_cols]
        _write_row(ws1, ri + 2, vals, alt=(ri % 2 == 1))

    # Grand total
    gt = ["GRAND TOTAL", "", "", "", int(df["Units"].sum()), "",
          "", "", round(df["Line Total ($)"].sum(), 2),
          "", "", "", "", "", "", "", "", "", "", "", ""]
    _write_row(ws1, len(df) + 2, gt, is_grand=True)

    for ci, w in enumerate(_DETAIL_COL_WIDTHS[: len(detail_cols)], 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w
    ws1.freeze_panes = "A2"

    # ════════════════════════════════════════════════════════════════════════
    # Sheet 2: Summary
    # ════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Summary")

    # ── Part A: pivot per PO ─────────────────────────────────────────────
    # Defensive: ensure all index columns are clean strings so the pivot
    # doesn't treat empty / NaN values as distinct categories.
    # (FOB is already guarded by pd.notna() above; the rest go through _s().)
    for _col in ["PO Number", "Style", "Description", "Color", "ETD", "MSRP"]:
        if _col in df.columns:
            df[_col] = df[_col].fillna("").astype(str)

    pivot = df.pivot_table(
        index=["PO Number", "Style", "Description", "Color", "ETD",
               "Unit Price (FOB)", "MSRP"],
        columns="Size", values="Units", aggfunc="sum", fill_value=0,
        # NOTE: dropna=False was removed — it causes a cartesian-product
        # explosion when index columns have many distinct values (pandas
        # creates rows for every combination of unique values, not just
        # the ones that actually appear in the data).  The pd.notna() guard
        # on FOB and the fillna("") above already prevent any NaN from
        # entering the index, so dropna=True (the default) is correct here.
    ).reset_index()
    pivot.columns.name = None
    sc = [s for s in KL_SIZE_ORDER if s in pivot.columns]
    pivot["Total Units"] = pivot[sc].sum(axis=1)
    pivot = pivot.rename(columns={"Unit Price (FOB)": "FOB Price"})
    sum_cols = ["PO Number", "Style", "Description", "Color", "ETD",
                "FOB Price", "MSRP"] + sc + ["Total Units"]
    for col in sum_cols:
        if col not in pivot.columns:
            pivot[col] = ""
    pivot = pivot[sum_cols].sort_values(["PO Number", "Color"]).reset_index(drop=True)

    _write_header(ws2, sum_cols, row=1)
    for ri, row in pivot.iterrows():
        vals = [row[c] for c in sum_cols]
        _write_row(ws2, ri + 2, vals, alt=(ri % 2 == 1))

    gt_a = (["GRAND TOTAL", "", "", "", "", "", ""]
            + [int(pivot[s].sum()) for s in sc]
            + [int(pivot["Total Units"].sum())])
    _write_row(ws2, len(pivot) + 2, gt_a, is_grand=True)

    # ── Part B: style rollup ─────────────────────────────────────────────
    sep_row = len(pivot) + 4
    style_cols = ["Style", "Description", "FOB Price", "MSRP", "Total Units", "PO Count"]
    _write_header(ws2, style_cols, row=sep_row)

    style_grp = (
        df.groupby(["Style", "Description", "Unit Price (FOB)"])
        .agg(total=("Units", "sum"), po_count=("PO Number", "nunique"))
        .reset_index()
        .rename(columns={
            "Unit Price (FOB)": "FOB Price",
            "total": "Total Units",
            "po_count": "PO Count",
        })
    )
    style_grp["MSRP"] = ""

    for ri, row in style_grp.iterrows():
        vals = [row["Style"], row["Description"], row["FOB Price"],
                row["MSRP"], int(row["Total Units"]), int(row["PO Count"])]
        _write_row(ws2, sep_row + 1 + ri, vals, alt=(ri % 2 == 1))

    gt_b = ["GRAND TOTAL", "", "", "",
            int(style_grp["Total Units"].sum()), int(style_grp["PO Count"].sum())]
    _write_row(ws2, sep_row + 1 + len(style_grp), gt_b, is_grand=True)

    # Column widths for Summary
    s2_widths = [14, 12, 32, 22, 7, 10, 8] + [8] * len(sc) + [12]
    for ci, w in enumerate(s2_widths[: len(sum_cols)], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.freeze_panes = "A2"

    # ── Serialise ─────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
