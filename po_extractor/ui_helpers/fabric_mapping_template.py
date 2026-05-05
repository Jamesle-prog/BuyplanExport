"""Fabric-mapping Excel template generator (pure openpyxl, no Streamlit)."""
from __future__ import annotations

import io

BODY_PART_LIST = [
    "Main Body / 大身",
    "Upper Body / 上身",
    "Lower Body / 下身",
    "Lining / 里布",
    "Sleeve / 袖子",
    "Sleeve Lining / 袖里布",
    "Collar / 领子",
    "Cuff / 袖口",
    "Hood / 帽子",
    "Pocket / 口袋布",
    "Pocket Lining / 口袋里布",
    "Pocket Mesh / 网眼布",
    "Waistband / 腰头",
    "Front Panel / 前片",
    "Back Panel / 后片",
    "Facing / 贴边",
    "Interlining / 衬布",
    "Piping / 嵌条",
    "Trim / 辅料",
]


def generate_fabric_mapping_template() -> bytes:
    """Shared style-fabric mapping template (GIII and Sky East).

    Layout -- one bilingual header row + two example rows:
      A: Style No. / 款式号
      B: Fabric 1 Body Part / 面料1部位  (dropdown)   C: Fabric 1 Code / 面料1编号
      D: Fabric 2 Body Part / 面料2部位  (dropdown)   E: Fabric 2 Code / 面料2编号
      F: Fabric 3 Body Part / 面料3部位  (dropdown)   G: Fabric 3 Code / 面料3编号
      H: Fabric 4 Body Part / 面料4部位  (dropdown)   I: Fabric 4 Code / 面料4编号
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Style Fabric Mapping"

    headers = [
        "Style No.\n款式号",
        "Fabric 1 Body Part\n面料1部位",  "Fabric 1 Code\n面料1编号",
        "Fabric 2 Body Part\n面料2部位",  "Fabric 2 Code\n面料2编号",
        "Fabric 3 Body Part\n面料3部位",  "Fabric 3 Code\n面料3编号",
        "Fabric 4 Body Part\n面料4部位",  "Fabric 4 Code\n面料4编号",
    ]
    col_widths = [20, 24, 18, 24, 18, 24, 18, 24, 18]

    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)

    for ci, (hdr, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=ci, value=hdr)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = width

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    examples = [
        ["S25DDR2036", "Main Body / 大身",   "HHN-JA-01715", "Lining / 里布",        "HHN-MS-01794", "",                   "",             "", ""],
        ["S25JKT1042", "Upper Body / 上身",  "HHN-JA-02301", "Pocket Mesh / 网眼布", "HHN-PO-00891", "Sleeve / 袖子",       "HHN-JA-00712", "", ""],
    ]
    ex_font = Font(color="808080", italic=True, size=10)
    for ri, row in enumerate(examples, start=2):
        for ci, val in enumerate(row, start=1):
            ws.cell(row=ri, column=ci, value=val).font = ex_font

    note = ws.cell(
        row=4, column=1,
        value="↑ 请替换示例行 / Replace example rows.  "
              "Body Part 部位 -- select from dropdown or leave blank if single fabric.  "
              "Fabric Code = HHN编号 (e.g. HHN-JA-01715).  "
              "Composition 成分 is looked up automatically from the Fabric DB.",
    )
    note.font = Font(color="C00000", size=9, italic=True)
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=9)

    # Excel's inline list formula is limited to 255 characters.  The body-part
    # list exceeds that limit, so we write the values to a hidden helper sheet
    # and reference those cells instead — no character limit applies.
    ws_lists = wb.create_sheet("_Lists")
    ws_lists.sheet_state = "hidden"
    for i, item in enumerate(BODY_PART_LIST, start=1):
        ws_lists.cell(row=i, column=1, value=item)

    n = len(BODY_PART_LIST)
    dv_formula = f"_Lists!$A$1:$A${n}"
    for col_letter in ("B", "D", "F", "H"):
        dv = DataValidation(
            type="list",
            formula1=dv_formula,
            allow_blank=True,
            showDropDown=False,
            showErrorMessage=True,
            errorTitle="Invalid entry",
            error="Please select from the dropdown list or leave blank.",
        )
        dv.sqref = f"{col_letter}2:{col_letter}1000"
        ws.add_data_validation(dv)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
