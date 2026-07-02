"""Wash-label Excel generator (pure openpyxl, no Streamlit)."""
from __future__ import annotations

import io
from typing import Any

import pandas as pd

from po_extractor.config import EXCEL_PALETTE as _P


def write_wash_label_excel(
    df_enriched: Any,
    image_cache: dict,
    fabric_parts_by_style: dict | None = None,
    styles: list[str] | None = None,
) -> bytes:
    """Generate a Wash Label Excel file.

    Columns: Style | Photo | Seq | Body Part | Fabric Code | Composition.
    One group per (style, fabric-combination).  Styles with multiple fabric
    combinations produce one group each, so BL4047 with 2 mapping rows appears
    as two separate groups in the output.

    When fabric_parts_by_style is None or missing for a style, falls back to the
    single fabric_item_no from df_enriched.

    Photo is embedded in the first row for each style group (subsequent rows blank).

    Parameters
    ----------
    df_enriched
        DataFrame with at minimum 'style' column. Optional columns:
        'picture_id', 'fabric_item_no', 'composition_en'.
    image_cache
        Mapping of picture_id → image bytes (PNG/JPG).
    fabric_parts_by_style
        Optional mapping of style → list of FabricPart-like objects with
        attributes ``combo_idx``, ``seq``, ``body_part``, ``hhn_no``,
        ``composition``.  Parts with the same ``combo_idx`` belong to one
        fabric combination and are rendered as a single group.
    styles
        Optional explicit list of styles to include in the output, in order.
        When provided this overrides the style list derived from ``df_enriched``,
        so styles that have no DB items (but do have fabric mapping data) are
        still included.

    Returns
    -------
    bytes
        Encoded .xlsx file content ready for download.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XLImage

    wb = Workbook()
    ws = wb.active
    ws.title = "Wash Labels"

    headers    = ["Style", "Photo", "Seq", "Body Part", "Fabric Code", "Composition"]
    col_widths = [20,       12,      6,     22,           20,            40]
    hdr_fill = PatternFill("solid", start_color=_P["wash_hdr"], end_color=_P["wash_hdr"])
    hdr_font = Font(bold=True, color=_P["white"], size=11)
    thin   = Side(style="thin", color=_P["border_grey"])
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, (hdr, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=ci, value=hdr)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 22

    # ── Build lookup dicts with vectorised pandas ops (replaces 3 iterrows) ──
    style_to_pid: dict[str, str] = {}
    if "picture_id" in df_enriched.columns and "style" in df_enriched.columns:
        _pid_df = (df_enriched[["style", "picture_id"]]
                   .assign(style=lambda d: d["style"].fillna("").astype(str).str.strip(),
                           picture_id=lambda d: d["picture_id"].fillna("").astype(str).str.strip()))
        _pid_df = _pid_df[(_pid_df["style"] != "") & (_pid_df["picture_id"] != "")]
        if not _pid_df.empty:
            style_to_pid = (
                _pid_df.drop_duplicates(subset="style", keep="first")
                .set_index("style")["picture_id"].to_dict()
            )

    if styles is not None:
        # Caller supplied an explicit list — use it directly (preserves order,
        # includes styles that may have no rows in df_enriched).
        styles_ordered = [s for s in styles if s]
    elif "style" in df_enriched.columns:
        styles_ordered = (
            df_enriched["style"].dropna()
            .astype(str).str.strip()
            .replace("", pd.NA).dropna()
            .drop_duplicates().tolist()
        )
    else:
        styles_ordered = []

    if not fabric_parts_by_style:
        fabric_parts_by_style = {}

    fallback_hhn: dict[str, tuple[str, str]] = {}
    if "fabric_item_no" in df_enriched.columns and "style" in df_enriched.columns:
        _fb_cols = ["style", "fabric_item_no"]
        if "composition_en" in df_enriched.columns:
            _fb_cols.append("composition_en")
        _fb = (df_enriched[_fb_cols]
               .assign(style=lambda d: d["style"].fillna("").astype(str).str.strip(),
                       fabric_item_no=lambda d: d["fabric_item_no"].fillna("").astype(str).str.strip()))
        _fb = _fb[(_fb["style"] != "") & (_fb["fabric_item_no"] != "")]
        _fb = _fb.drop_duplicates(subset="style", keep="first")
        for row in _fb.itertuples(index=False):
            c_raw = getattr(row, "composition_en", "") or ""
            c = "" if str(c_raw).lower() in ("nan", "none", "") else str(c_raw).strip()
            fallback_hhn[row.style] = (row.fabric_item_no, c)

    # ------------------------------------------------------------------ #
    # Build the render list: one entry per (style, combo_idx) group.
    # Each entry is (style_label, combo_parts_or_None).
    # Styles with multiple combinations produce multiple entries; the style
    # label is shown on every group's first row so the user can tell them apart.
    # ------------------------------------------------------------------ #
    render_list: list[tuple[str, list | None]] = []

    for style in styles_ordered:
        all_parts = fabric_parts_by_style.get(style)
        if all_parts:
            # Group parts by combo_idx, preserving the sorted order
            combos: dict[int, list] = {}
            for p in all_parts:
                cidx = getattr(p, "combo_idx", 0) or 0
                combos.setdefault(cidx, []).append(p)
            for cidx in sorted(combos):
                render_list.append((style, combos[cidx]))
        else:
            # No fabric-mapping parts — use fallback or blank row
            render_list.append((style, None))

    # ------------------------------------------------------------------ #
    # Write rows
    # ------------------------------------------------------------------ #
    even_fill = PatternFill("solid", start_color=_P["wash_alt"], end_color=_P["wash_alt"])
    data_font = Font(size=10)
    photo_row_h  = 70
    normal_row_h = 20

    ri = 2
    fill_toggle = False

    for style, combo_parts in render_list:
        # Build the sub-rows for this combo group.
        # Parts with no HHN code (or only whitespace) are skipped — they have
        # nothing meaningful to print and would render as a blank row that
        # confuses the reader.
        if combo_parts is not None:
            valid_parts = [p for p in combo_parts
                           if str(getattr(p, "hhn_no", "") or "").strip()]
        else:
            valid_parts = []

        if valid_parts:
            rows_for_group = [
                (
                    i + 1,
                    p.body_part or "",
                    p.hhn_no or "",
                    "" if (not p.composition or str(p.composition).lower() in ("nan", "none"))
                    else str(p.composition),
                )
                for i, p in enumerate(valid_parts)
            ]
        elif style in fallback_hhn:
            hhn, comp = fallback_hhn[style]
            rows_for_group = [(1, "", hhn, comp)]
        else:
            rows_for_group = [(1, "", "", "")]

        if not rows_for_group:
            continue

        fill = even_fill if fill_toggle else None
        fill_toggle = not fill_toggle

        pid       = style_to_pid.get(style, "")
        photo_set = False

        for sub_idx, (seq, body_part, hhn, comp) in enumerate(rows_for_group):
            style_val = style if sub_idx == 0 else ""
            for ci, val in enumerate(
                [style_val, "", str(seq), body_part, hhn, comp], start=1
            ):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font      = data_font
                cell.alignment = Alignment(
                    vertical="center",
                    wrap_text=(ci in (4, 6)),
                )
                cell.border = border
                if fill:
                    cell.fill = fill

            if not photo_set and pid:
                img_bytes = image_cache.get(pid)
                if img_bytes:
                    try:
                        xl_img = XLImage(io.BytesIO(img_bytes))
                        xl_img.width, xl_img.height = 55, 68
                        ws.add_image(xl_img, ws.cell(row=ri, column=2).coordinate)
                        ws.row_dimensions[ri].height = photo_row_h
                        photo_set = True
                    except Exception:
                        ws.row_dimensions[ri].height = normal_row_h
                else:
                    ws.row_dimensions[ri].height = normal_row_h
            else:
                ws.row_dimensions[ri].height = normal_row_h

            ri += 1

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
