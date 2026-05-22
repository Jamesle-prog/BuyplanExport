"""Pure-logic Excel report builders (color plan + PO summary).

These take a pandas DataFrame and return bytes — no Streamlit, no DB.
The Streamlit-side caller in app.py loads the DataFrame, then delegates here.
"""
from __future__ import annotations

import io
from typing import Callable

import pandas as pd

from po_extractor.ui_helpers.excel_format import write_excel_header_row

# Standard size ordering for color-plan pivots.
SIZE_ORDER: list[str] = [
    "PXS", "PS", "PM", "PL", "PXL", "P1X", "P2X", "P3X", "P2XL", "P3XL",
    "XXS", "XS", "S", "M", "L", "XL", "XXL", "XXXL",
    "0X", "1X", "2X", "3X", "4X",
    "ONE SIZE", "OS",
]


def generate_color_plan_excel(df_size_rows: pd.DataFrame) -> bytes:
    """Pivot size rows into a color plan: one row per (PO, Style, Color), sizes as columns.

    Input columns required: PO Number, Style, Color, Size, Units.
    Returns b'' for empty input.
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill

    if df_size_rows is None or df_size_rows.empty:
        return b""

    pivot = df_size_rows.pivot_table(
        index=["PO Number", "Style", "Color"],
        columns="Size",
        values="Units",
        aggfunc="sum",
        fill_value=0,
        dropna=False,
    ).reset_index()
    pivot.columns.name = None

    size_cols = [s for s in SIZE_ORDER if s in pivot.columns]
    extra = [c for c in pivot.columns
             if c not in ["PO Number", "Style", "Color"] + size_cols]
    ordered = ["PO Number", "Style", "Color"] + size_cols + extra
    pivot = pivot[[c for c in ordered if c in pivot.columns]]
    pivot["Total"] = pivot[size_cols + extra].sum(axis=1)

    wb = Workbook()
    ws = wb.active
    ws.title = "Color Plan"
    write_excel_header_row(ws, list(pivot.columns))

    alt_fill = PatternFill("solid", start_color="EEF2FF", end_color="EEF2FF")
    for ri, row in enumerate(pivot.itertuples(index=False), start=2):
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            if ri % 2 == 0:
                cell.fill = alt_fill

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Rich PO Summary  (two-sheet format)
# ---------------------------------------------------------------------------

def generate_po_summary_excel(
    df_pos: pd.DataFrame,
    df_sizes: pd.DataFrame | None = None,
    label_for: Callable[[str, str], str] | None = None,
) -> bytes:
    """Rich two-sheet PO Summary workbook.

    Sheet 1 – "PO Summary":
      • Header block   : title + key-value rows (Vendor, Factory, COO, Incoterm,
                         PO Date, ETD, Season, Fabric, Pack, Discount, HTS, FOB)
      • Detail table   : PO Number | Style | Color Code | Color Name | FOB |
                         Cost P/U | <sizes…> | Total Units | Extended Cost
      • TOTAL row

    Sheet 2 – "Size Breakdown":
      Style | Color Code | Color Name | <sizes…> | Total

    ``df_sizes`` must have columns: PO Number, Style, Color, Size, Units.
    When None or empty, the size columns are omitted from the detail table.
    """
    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side,
    )
    from openpyxl.utils import get_column_letter

    if label_for is None:
        def label_for(db_col, fallback):  # noqa: E306
            return fallback

    if df_pos is None or df_pos.empty:
        return b""

    def _v(row, col):
        """Safe column getter — returns '' when column missing or NaN."""
        import math
        val = row.get(col, "")
        if val is None:
            return ""
        if isinstance(val, float) and math.isnan(val):
            return ""
        return str(val).strip() if val else ""

    # ── Helper styles ────────────────────────────────────────────────────────
    HDR_FILL   = PatternFill("solid", start_color="4472C4", end_color="4472C4")
    HDR_FONT   = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    TITLE_FONT = Font(bold=True, name="Calibri", size=13)
    LABEL_FONT = Font(bold=True, name="Calibri", size=10)
    DATA_FONT  = Font(name="Calibri", size=10)
    TOTAL_FILL = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")
    TOTAL_FONT = Font(bold=True, name="Calibri", size=10)
    ALT_FILL   = PatternFill("solid", start_color="F2F7FF", end_color="F2F7FF")
    THIN_BORDER = Border(
        left=Side(style="thin", color="AAAAAA"),
        right=Side(style="thin", color="AAAAAA"),
        top=Side(style="thin", color="AAAAAA"),
        bottom=Side(style="thin", color="AAAAAA"),
    )

    def _style(cell, font=None, fill=None, align=None, border=None, num_fmt=None):
        if font:   cell.font      = font
        if fill:   cell.fill      = fill
        if align:  cell.alignment = align
        if border: cell.border    = border
        if num_fmt: cell.number_format = num_fmt

    # ── Determine sizes present in the data ─────────────────────────────────
    po_list = df_pos["po_number"].tolist() if "po_number" in df_pos.columns else []

    size_pivot: pd.DataFrame | None = None
    active_sizes: list[str] = []
    if df_sizes is not None and not df_sizes.empty:
        # Normalise column names (may come as "PO Number" or "po_number")
        rename = {}
        for c in df_sizes.columns:
            lc = c.lower().replace(" ", "_")
            if lc == "po_number":   rename[c] = "po_number"
            elif lc == "style":     rename[c] = "style"
            elif lc == "color":     rename[c] = "color"
            elif lc == "size":      rename[c] = "size"
            elif lc == "units":     rename[c] = "units"
        df_sz = df_sizes.rename(columns=rename)
        # Filter to selected POs
        if po_list and "po_number" in df_sz.columns:
            df_sz = df_sz[df_sz["po_number"].isin(po_list)]
        if not df_sz.empty and "size" in df_sz.columns:
            size_pivot = df_sz.pivot_table(
                index=["po_number", "style", "color"],
                columns="size",
                values="units",
                aggfunc="sum",
                fill_value=0,
                dropna=False,
            ).reset_index()
            size_pivot.columns.name = None
            active_sizes = [s for s in SIZE_ORDER if s in size_pivot.columns]
            # Keep unknown sizes in the order they appear
            extra_sizes = [c for c in size_pivot.columns
                           if c not in ["po_number", "style", "color"] + active_sizes]
            active_sizes += extra_sizes

    # ── Gather metadata from first PO (for header block) ─────────────────────
    first = df_pos.iloc[0] if not df_pos.empty else {}

    def _meta(col):
        return _v(first, col)

    # Build header block key-value pairs
    header_kv = [
        ("Vendor",    _meta("seller") or _meta("company")),
        ("Factory",   _meta("factory")),
        ("Ship To",   _meta("ship_to") or _meta("destination_code")),
        ("COO",       _meta("country_of_origin")),
        ("Incoterm",  _meta("incoterm")),
        ("PO Date",   _meta("po_date") or _meta("issue_date")),
        ("ETD/Ship",  _meta("xport_date") or _meta("factory_ship_date")),
        ("Season",    _meta("season")),
        ("Fabric",    _meta("style_description")),
        ("Pack",      _meta("packaging")),
        ("Discount",  _meta("discount")),
        ("HTS",       _meta("style_group")),
    ]
    # Filter out blank values but keep all 12 slots for structure
    # (user expects to see the labels even when value is blank)

    # ── Detail table columns ─────────────────────────────────────────────────
    detail_cols = (
        ["PO Number", "Style", "Color Code", "Color Name", "FOB", "Cost P/U"]
        + active_sizes
        + ["Total Units", "Extended Cost"]
    )

    # ── Workbook ─────────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "PO Summary"

    # ── SECTION 1: Title ─────────────────────────────────────────────────────
    total_cols = len(detail_cols)
    TITLE_ROW = 1
    ws.merge_cells(
        start_row=TITLE_ROW, start_column=1,
        end_row=TITLE_ROW, end_column=max(total_cols, 6),
    )
    _vendor = _meta("company") or _meta("seller") or "GIII"
    _season = _meta("season")
    title_val = f"PO Summary - {_vendor}" + (f"  |  Season {_season}" if _season else "")
    tc = ws.cell(row=TITLE_ROW, column=1, value=title_val)
    _style(tc, font=TITLE_FONT, align=Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[TITLE_ROW].height = 22

    # ── SECTION 2: Header block (key-value rows) ─────────────────────────────
    HB_START = 2
    for i, (label, value) in enumerate(header_kv):
        row_idx = HB_START + i
        lc = ws.cell(row=row_idx, column=1, value=label + ":")
        _style(lc, font=LABEL_FONT, align=Alignment(horizontal="right"))
        vc = ws.cell(row=row_idx, column=2, value=value)
        _style(vc, font=DATA_FONT, align=Alignment(horizontal="left"))
    HB_END = HB_START + len(header_kv) - 1

    # Spacer row
    DETAIL_HDR_ROW = HB_END + 2

    # ── SECTION 3: Detail table header ───────────────────────────────────────
    for ci, col in enumerate(detail_cols, start=1):
        c = ws.cell(row=DETAIL_HDR_ROW, column=ci, value=col)
        _style(c, font=HDR_FONT, fill=HDR_FILL,
               align=Alignment(horizontal="center", vertical="center", wrap_text=True),
               border=THIN_BORDER)
    ws.row_dimensions[DETAIL_HDR_ROW].height = 30

    # ── SECTION 4: Data rows ─────────────────────────────────────────────────
    DATA_START = DETAIL_HDR_ROW + 1
    grand_total_units = 0
    grand_ext_cost    = 0.0
    ri = DATA_START

    for _, po_row in df_pos.iterrows():
        po_num = _v(po_row, "po_number")
        style  = _v(po_row, "style")
        fob    = _v(po_row, "unit_cost")
        cpu    = _v(po_row, "line_extended_cost")

        # Try to parse cost-per-unit as float for Extended Cost calculation
        try:
            cpu_f = float(str(cpu).replace(",", "")) if cpu else 0.0
        except ValueError:
            cpu_f = 0.0
        try:
            fob_f = float(str(fob).replace(",", "")) if fob else 0.0
        except ValueError:
            fob_f = 0.0

        # Get size data for this PO
        if size_pivot is not None and not size_pivot.empty:
            po_sizes = size_pivot[size_pivot["po_number"] == po_num]
        else:
            po_sizes = pd.DataFrame()

        if po_sizes.empty:
            # No size breakdown — one summary row per PO
            total_u = int(_v(po_row, "total_units") or 0)
            ext_c = round(total_u * cpu_f, 2)

            row_data = [po_num, style, "", "", fob_f or fob, cpu_f or cpu]
            row_data += [0] * len(active_sizes)
            row_data += [total_u, ext_c]

            alt = (ri - DATA_START) % 2 == 1
            for ci, val in enumerate(row_data, start=1):
                c = ws.cell(row=ri, column=ci, value=val)
                _style(c, font=DATA_FONT,
                       fill=ALT_FILL if alt else None,
                       border=THIN_BORDER,
                       align=Alignment(horizontal="right" if isinstance(val, (int, float)) else "left"))
            grand_total_units += total_u
            grand_ext_cost    += ext_c
            ri += 1
        else:
            for _, sz_row in po_sizes.iterrows():
                color_full = str(_v(sz_row, "color") or "")
                # Split "CODE/COLOR NAME" or keep as-is for color code + name
                if "/" in color_full:
                    parts = color_full.split("/", 1)
                    color_code = parts[0].strip()
                    color_name = parts[1].strip()
                else:
                    color_code = color_full
                    color_name = ""

                size_vals = [int(sz_row.get(s, 0) or 0) for s in active_sizes]
                total_u   = sum(size_vals)
                ext_c     = round(total_u * cpu_f, 2)

                row_data = [po_num, style, color_code, color_name,
                            fob_f or fob, cpu_f or cpu]
                row_data += size_vals
                row_data += [total_u, ext_c]

                alt = (ri - DATA_START) % 2 == 1
                for ci, val in enumerate(row_data, start=1):
                    c = ws.cell(row=ri, column=ci, value=val)
                    _style(c, font=DATA_FONT,
                           fill=ALT_FILL if alt else None,
                           border=THIN_BORDER,
                           align=Alignment(horizontal="right" if isinstance(val, (int, float)) else "left"))
                grand_total_units += total_u
                grand_ext_cost    += ext_c
                ri += 1

    # ── SECTION 5: TOTAL row ─────────────────────────────────────────────────
    total_col_idx = detail_cols.index("Total Units") + 1
    ext_col_idx   = detail_cols.index("Extended Cost") + 1

    for ci in range(1, len(detail_cols) + 1):
        c = ws.cell(row=ri, column=ci)
        if ci == 1:
            c.value = "TOTAL"
        elif ci == total_col_idx:
            c.value = grand_total_units
        elif ci == ext_col_idx:
            c.value = round(grand_ext_cost, 2)
        _style(c, font=TOTAL_FONT, fill=TOTAL_FILL, border=THIN_BORDER,
               align=Alignment(horizontal="right" if ci > 4 else "left"))

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = {
        1: 16,   # PO Number
        2: 14,   # Style
        3: 10,   # Color Code
        4: 20,   # Color Name
        5: 9,    # FOB
        6: 9,    # Cost P/U
    }
    for i, s in enumerate(active_sizes, start=7):
        col_widths[i] = 6
    col_widths[len(detail_cols) - 1] = 12   # Total Units
    col_widths[len(detail_cols)]     = 14   # Extended Cost

    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    # Also widen label column in header block
    ws.column_dimensions["A"].width = max(16, col_widths.get(1, 16))
    ws.column_dimensions["B"].width = max(22, col_widths.get(2, 22))

    ws.freeze_panes = ws.cell(row=DETAIL_HDR_ROW + 1, column=1)

    # =========================================================================
    # Sheet 2: Size Breakdown
    # =========================================================================
    ws2 = wb.create_sheet("Size Breakdown")
    sb_cols = ["Style", "Color Code", "Color Name"] + active_sizes + ["Total"]
    for ci, col in enumerate(sb_cols, start=1):
        c = ws2.cell(row=1, column=ci, value=col)
        _style(c, font=HDR_FONT, fill=HDR_FILL,
               align=Alignment(horizontal="center", vertical="center"),
               border=THIN_BORDER)
    ws2.row_dimensions[1].height = 22

    if size_pivot is not None and not size_pivot.empty:
        # Aggregate across all POs: group by (style, color)
        agg_cols = ["style", "color"] + active_sizes
        avail_agg = [c for c in agg_cols if c in size_pivot.columns]
        grp = size_pivot[avail_agg].groupby(
            ["style", "color"], sort=False
        )[active_sizes].sum().reset_index() if active_sizes else (
            size_pivot[["style", "color"]].drop_duplicates().reset_index(drop=True)
        )

        for ri2, row in enumerate(grp.itertuples(index=False), start=2):
            style_v = getattr(row, "style", "")
            color_full = str(getattr(row, "color", "") or "")
            if "/" in color_full:
                parts = color_full.split("/", 1)
                cc, cn = parts[0].strip(), parts[1].strip()
            else:
                cc, cn = color_full, ""
            size_vals2 = [int(getattr(row, s, 0) or 0) for s in active_sizes]
            total2 = sum(size_vals2)
            sb_row = [style_v, cc, cn] + size_vals2 + [total2]
            alt = (ri2 - 2) % 2 == 1
            for ci, val in enumerate(sb_row, start=1):
                c = ws2.cell(row=ri2, column=ci, value=val)
                _style(c, font=DATA_FONT,
                       fill=ALT_FILL if alt else None,
                       border=THIN_BORDER,
                       align=Alignment(horizontal="right" if isinstance(val, (int, float)) else "left"))

    for ci, s in enumerate(sb_cols, start=1):
        width = 14 if ci <= 3 else (6 if ci <= 3 + len(active_sizes) else 10)
        ws2.column_dimensions[get_column_letter(ci)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
