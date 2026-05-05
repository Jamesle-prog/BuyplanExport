"""Buy-plan template helpers — generation and header-row detection."""
from __future__ import annotations

import io


def make_sample_buyplan_template() -> bytes:
    """Generate a downloadable sample buy-plan template."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"

    def _hdr(cell, val):
        cell.value = val
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2E4053", end_color="2E4053", fill_type="solid")
        cell.alignment = Alignment(horizontal="right", vertical="center")

    thin = Side(border_style="thin", color="AAAAAA")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

    _hdr(ws["A1"], "工厂信息:")
    ws["B1"] = "{{factory}}"
    ws["M1"] = "创建时间:"
    ws["N1"] = "{{created_at}}"
    ws["N1"].font = Font(italic=True)

    _hdr(ws["A2"], "款号:")
    ws["B2"] = "{{style}}"

    _hdr(ws["A3"], "面料信息:")

    _hdr(ws["A4"], "出厂日期:")
    ws["B4"] = "{{xfactory_date}}"
    ws["D4"] = "Orig X-Port Date:"
    ws["E4"] = "{{xport_date}}"
    ws["G4"] = "COO:"
    ws["H4"] = "{{coo}}"
    ws["J4"] = "Division:"
    ws["K4"] = "{{division}}"

    ws["A5"] = "{{data_start}}"
    ws["A5"].font = Font(italic=True, color="888888")
    ws["A5"].alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 22
    for col in ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N"]:
        ws.column_dimensions[col].width = 14

    for row in ws.iter_rows(min_row=1, max_row=5, min_col=1, max_col=14):
        for cell in row:
            cell.border = bdr

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def detect_template_header_row(xlsx_bytes: bytes) -> int | None:
    """Return the row containing {{data_start}}, or None if not found."""
    try:
        import openpyxl as _xl
        wb = _xl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True)
        ws = wb.worksheets[0]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "{{data_start}}" in cell.value.lower():
                    wb.close()
                    return cell.row
        wb.close()
    except Exception:
        pass
    return None
