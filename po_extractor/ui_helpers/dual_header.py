"""Dual-row header Excel writer for Sky East exports.

Row 1 = Sky East client's own column name (their reference).
Row 2 = our standard label resolved from the live schema.
Row 3+ = data rows.
"""
from __future__ import annotations

import io
from typing import Any, Callable

# (db_col, row1_sky_east_client_name, label_db_col_for_row2)
DUAL_HEADER_STATIC: list[tuple[str, str, str]] = [
    ("pc_no",              "客名字",                            "pc_no"),
    ("style",              "Main Supplier Config SKU",          "style"),
    ("zalando_po",         "Purchase Order Number",             "po_number"),
    ("config_sku",         "Config SKU",                        "config_sku"),
    ("article_name",       "Article Name",                      "article_name"),
    ("brand",              "Brand",                             "brand"),
    ("color_name",         "Main Supplier Color Description",   "color_name"),
    ("colour_code",        "Colour Code",                       "colour_code"),
    ("xs",                 "XS",                                "xs"),
    ("s",                  "S",                                 "s"),
    ("m",                  "M",                                 "m"),
    ("l",                  "L",                                 "l"),
    ("xl",                 "XL",                                "xl"),
    ("xxl",                "2XL",                               "xxl"),
    ("total_qty",          "TOTAL QUANTITY",                    "total_qty"),
    ("ex_fty_date",        "EX-FTY",                            "ex_fty_date"),
    ("fabric_item_no",     "Fabric Item Number",                "fabric_item_no"),
    ("composition_en",     "Composition (EN)",                  "composition_en"),
    ("shrinkage_rate",     "Shrinkage",                         "shrinkage_rate"),
    ("short_rate",         "Short Rate",                        "short_rate"),
]


def get_dual_header(
    label_for: Callable[[str, str], str],
) -> list[tuple[str, str, str]]:
    """Return [(db_col, row1_client_name, row2_standard_label), ...].

    Parameters
    ----------
    label_for
        Callable accepting ``(db_col, fallback)`` and returning the standard
        label for that field — typically a thin wrapper around the live schema.
    """
    return [
        (db_col, row1, label_for(label_key, row1))
        for db_col, row1, label_key in DUAL_HEADER_STATIC
    ]


def write_dual_header_excel(
    df_enriched: Any,
    sheet_name: str,
    writer: Any,
    image_cache: dict | None = None,
    label_for: Callable[[str, str], str] | None = None,
) -> None:
    """Write a sheet with row-1=client headers, row-2=final headers, row-3+=data.

    ``writer`` must expose ``writer.book`` (an openpyxl Workbook).
    If ``image_cache`` ({picture_id: bytes}) is supplied and df_enriched has a
    'picture_id' column, a Photo column is inserted right after Style No.
    If ``label_for`` is None, row-2 falls back to the static row-1 client names.
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.drawing.image import Image as XLImage

    if label_for is None:
        # Fallback: use row1 name as row2 label too.
        label_for = lambda _db, fallback: fallback  # noqa: E731

    present = set(df_enriched.columns)
    cols = [(db, r1, r2) for db, r1, r2 in get_dual_header(label_for) if db in present]
    if not cols:
        cols = [(c, c, c) for c in df_enriched.columns]

    has_photos = bool(image_cache and "picture_id" in df_enriched.columns)
    photo_ci = None
    if has_photos:
        style_positions = [i for i, (db, _, _) in enumerate(cols) if db == "style"]
        insert_at = style_positions[0] + 1 if style_positions else 0
        cols.insert(insert_at, ("__photo__", "Sample_Pic", "Sample_Pic"))

    db_cols   = [c[0] for c in cols]
    row1_hdrs = [c[1] for c in cols]
    row2_hdrs = [c[2] for c in cols]

    ws = writer.book.create_sheet(title=sheet_name)

    header_fill = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")
    row2_fill   = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
    bold = Font(bold=True)

    for ci, (r1, r2) in enumerate(zip(row1_hdrs, row2_hdrs), start=1):
        c1 = ws.cell(row=1, column=ci, value=r1)
        c1.font = bold
        c1.fill = header_fill
        c1.alignment = Alignment(horizontal="center", wrap_text=True)

        c2 = ws.cell(row=2, column=ci, value=r2)
        c2.font = bold
        c2.fill = row2_fill
        c2.alignment = Alignment(horizontal="center", wrap_text=True)

        if r2 == "Sample_Pic":
            photo_ci = ci

    style_to_pid: dict[str, str] = {}
    if has_photos:
        for _, r in df_enriched.iterrows():
            s   = str(r.get("style", "")).strip()
            pid = str(r.get("picture_id", "")).strip()
            if s and pid and s not in style_to_pid:
                style_to_pid[s] = pid

    real_db_cols = [c for c in db_cols if c != "__photo__"]
    row_records  = df_enriched[real_db_cols].to_dict("records")

    for ri, rec in enumerate(row_records, start=3):
        ci = 1
        for db_col in db_cols:
            if db_col == "__photo__":
                ci += 1
                continue
            ws.cell(row=ri, column=ci, value=rec.get(db_col, ""))
            ci += 1

        if has_photos and photo_ci:
            style_val = str(rec.get("style", "")).strip()
            pid       = style_to_pid.get(style_val, "")
            img_bytes = image_cache.get(pid) if pid else None
            if img_bytes:
                try:
                    xl_img = XLImage(io.BytesIO(img_bytes))
                    xl_img.width  = 60
                    xl_img.height = 75
                    ws.add_image(xl_img, ws.cell(row=ri, column=photo_ci).coordinate)
                    ws.row_dimensions[ri].height = 58
                except Exception:
                    pass

    for ci, hdr in enumerate(row2_hdrs, start=1):
        col_letter = ws.cell(row=1, column=ci).column_letter
        ws.column_dimensions[col_letter].width = 9 if hdr == "Photo" else max(12, len(hdr) + 2)
