"""Production progress lookup from 大货进度表--Angel 2026.xlsx.

Sheet  "2026 Zalando"  — the master production tracker
  Col 1  : 序号        item number
  Col 2  : 合同号      HHN contract number   ← what we need
  Col 3  : 所在PO      Sky East PC No.
  Col 4  : IMAGE       DISPIMG formula → style photo
  Col 5  : 款式        Style No.              (join key)
  Col 7  : 颜色        Color description (e.g. "52# NAVY")
  Col 10 : PO离厂日期  Ex-factory date
  Col 11 : 数量        Quantity
  Col 12 : PO#         Zalando PO number
  Col 13 : BRAND
  Col 14 : FABRICDETAIL

Lookup keys:
  (style_no, color_rough)  → HHN contract no., image_id, ex_fty, zalando_po
  style_no                 → list of all colour rows
"""
from __future__ import annotations

import re
from datetime import datetime
from typing import NamedTuple
import openpyxl


class PCColorMatch(NamedTuple):
    """Result of a PC No.-keyed color lookup from 大货进度表.

    All three fields can independently be empty strings — callers should
    treat empty as "no value, fall back to the next tier".
    """
    cn_color:    str   # 中文颜色
    color_code:  str   # 中文颜色代码 (e.g. "52#")
    label_color: str   # 主标颜色


def _norm_key(s) -> str:
    """Strip whitespace, remove non-alphanumeric chars, uppercase — for all key lookups."""
    return re.sub(r'[^A-Za-z0-9]', '', str(s).strip()).upper()

_DISPIMG_RE = re.compile(r'DISPIMG\("(ID_[0-9A-Fa-f]+)"', re.IGNORECASE)

# A cached formula error in the source file (e.g. a 中文颜色代码 VLOOKUP that
# couldn't extract a numeric code from a colour cell with no code suffix,
# such as "BLACK 黑色" with nothing after it) is stored by openpyxl as the
# literal text "#N/A" — not None. Treating that string as a real value would
# leak "#N/A" straight into match keys and the generated buy plan cell.
_EXCEL_ERROR_RE = re.compile(
    r'^#(N/A|REF!|VALUE!|DIV/0!|NAME\?|NULL!|NUM!|SPILL!|CALC!)$', re.IGNORECASE,
)


def _v(val) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if _EXCEL_ERROR_RE.match(s):
        # Blank it out so callers fall back to their next-tier match — e.g.
        # the colour name alone is enough; a code isn't required.
        return ""
    return s


def _dispimg_id(val) -> str:
    m = _DISPIMG_RE.search(_v(val))
    return m.group(1) if m else ""


def _normalise_color(color: str) -> str:
    """Canonical colour form used for matching (delegates to clean_color_for_lookup)."""
    cleaned, _ = clean_color_for_lookup(color)
    return cleaned


def clean_color_for_lookup(color: str) -> tuple[str, list[str]]:
    """
    Clean an English colour string for contract lookup, returning the
    cleaned value and a list of human-readable steps describing what
    was removed.

    Steps are recorded only when something actually changed, so callers
    can show a concise log to the user.

    Examples
    --------
    >>> clean_color_for_lookup("52# NAVY")
    ('NAVY', ["stripped leading code '52#'"])
    >>> clean_color_for_lookup("NAVY 52#")
    ('NAVY', ["stripped trailing code ' 52#'"])
    >>> clean_color_for_lookup("blue")
    ('BLUE', ['uppercased'])
    """
    if not color:
        return "", []

    steps: list[str] = []
    original = str(color)
    c = original

    # 1. Uppercase
    upper = c.upper()
    if upper != c:
        steps.append("uppercased")
    c = upper

    # 2. Collapse newlines / multi-whitespace
    if re.search(r'[\r\n]|  +', c):
        c = re.sub(r'\s+', ' ', c).strip()
        steps.append("collapsed whitespace/newlines")
    else:
        c = c.strip()

    # 3. Strip non-ASCII (Chinese annotations like "深蓝色")
    if re.search(r'[^\x00-\x7F]', c):
        before = c
        c = re.sub(r'[^\x00-\x7F]+', '', c).strip()
        removed = "".join(ch for ch in before if ord(ch) > 127)
        steps.append(f"removed non-ASCII chars '{removed}'")

    # 4. Strip leading numeric colour code: "52# NAVY" → "NAVY"
    m = re.match(r'^(#?\d+#?\s*)', c)
    if m:
        prefix = m.group(1)
        c = c[len(prefix):]
        steps.append(f"stripped leading code '{prefix.strip()}'")

    # 5. Strip trailing numeric colour code with whitespace separator:
    #    "NAVY 52#" → "NAVY",  "BURGUNDY #74" → "BURGUNDY"
    m = re.search(r'(\s+#?\d+#?\s*)$', c)
    if m:
        suffix = m.group(1)
        c = c[:-len(suffix)]
        steps.append(f"stripped trailing code '{suffix.strip()}'")

    # 6. Strip trailing numeric code attached without whitespace:
    #    "NAVY#52" → "NAVY",  "DK NAVY#52" → "DK NAVY"
    m = re.search(r'(#\d+#?)\s*$', c)
    if m:
        suffix = m.group(1)
        c = c[:-len(suffix)]
        steps.append(f"stripped attached code '{suffix.strip()}'")

    # 7. Strip trailing punctuation left over from comments (",", ";", ".")
    if re.search(r'[,;.]\s*$', c):
        c = re.sub(r'[,;.\s]+$', '', c)
        steps.append("stripped trailing punctuation")

    # 8. Final whitespace cleanup
    c = re.sub(r'\s+', ' ', c).strip()

    return c, steps


def _clean_cn_color(s: str) -> str:
    """
    Strip numeric prefix/suffix from Chinese colour names so the user sees
    a clean Chinese name in the buy plan.

    Examples
    --------
    "58#浅蓝"   → "浅蓝"
    "72# 蓝色"  → "蓝色"
    "浅蓝"      → "浅蓝"
    """
    if not s:
        return ""
    c = str(s).strip()
    # Strip leading "58#" / "58# " / "#58 "
    c = re.sub(r'^#?\d+#?\s*', '', c)
    # Strip trailing " 58#" / " #58"
    c = re.sub(r'\s*#?\d+#?$', '', c)
    return c.strip()


def _extract_color_code(color: str) -> str:
    """
    Extract the numeric colour code so:
      "52# NAVY"      → "52"
      "NAVY 52#"      → "52"
      "BURGUNDY #74"  → "74"
      "NAVY"          → ""
    """
    if not color:
        return ""
    c = str(color).strip()
    # Try leading "52#" or "#52"
    m = re.match(r'^#?(\d+)#?', c)
    if m:
        return m.group(1)
    # Try trailing " 52#" or " #74"
    m = re.search(r'\s#?(\d+)#?\s*$', c)
    return m.group(1) if m else ""


# A two-colour code cell looks like "52#/白色 3#" or "2#/白色3#": the primary
# colour's own code, then a "/", then the *second* colour's Chinese name and
# its own code (sometimes space-separated, sometimes not).
_MULTI_COLOR_CODE_RE = re.compile(
    r'^\s*([^/]+?)\s*/\s*([一-鿿]+)\s*(.+?)\s*$'
)


def _extract_second_color_en(color_summary: str, en1: str, cn1: str) -> str:
    """Best-effort extraction of the second colour's English name from the
    combined 色汇总 cell, e.g. ``"Dark BlUE /White 藏青 52#/白色 3#"`` with
    ``en1="Dark BlUE"``, ``cn1="藏青"`` → ``"White"``.

    Not always possible: some rows describe an accent/trim colour rather
    than a second body colour (e.g. ``"BLACK WITH WHITE STRAP..."``), where
    the combined cell has no separate English token for it — returns ``""``.
    """
    if not color_summary:
        return ""
    s = color_summary
    if en1 and s.lower().startswith(en1.lower()):
        s = s[len(en1):]
    s = s.lstrip(" /\t")
    if cn1 and cn1 in s:
        s = s[:s.index(cn1)]
    return s.strip(" /\t")


def _split_multi_color(
    color_en: str, cn_color: str, color_code: str, color_summary: str,
) -> list[dict]:
    """Split a two-colour progress row into its individual colour components.

    Some 大货进度表 rows describe a two-tone style with both colours crammed
    into the 中文颜色代码 cell, e.g. ``"52#/白色 3#"`` — the primary colour's
    own code (``"52#"``), then the *second* colour's Chinese name
    (``"白色"``) and its own code (``"3#"``). Left as one record, the second
    colour's code is unreachable — an order-file item whose colour is
    "White" would never match this row on an exact colour lookup, and the
    combined field is unusable garbage in the buy plan's Color Code cell.

    Splitting into one record per colour lets the existing
    ``(pc_no, style, colour)`` lookup key resolve each colour independently —
    no special-casing needed anywhere else in ``ProgressLookup`` or the
    persistent store (row identity is already keyed by colour, so two
    colours for one style/PC naturally become two distinct rows).

    Returns a list of ``{"color", "cn_color", "color_code"}`` dicts — one
    element when no second colour is detected (the common case), two when a
    second colour's code is found. When the second colour has no isolable
    English name (an accent/trim colour, not a second body colour), that
    component's ``"color"`` is ``""`` — it still carries its own Chinese
    name and code, just can't be matched by English name alone.
    """
    m = _MULTI_COLOR_CODE_RE.match(color_code) if color_code else None
    if not m:
        return [{"color": color_en, "cn_color": cn_color, "color_code": color_code}]

    code1, cn2, code2 = m.group(1).strip(), m.group(2).strip(), m.group(3).strip()
    en2 = _extract_second_color_en(color_summary, color_en, cn_color)
    return [
        {"color": color_en, "cn_color": cn_color, "color_code": code1},
        {"color": en2, "cn_color": cn2, "color_code": code2},
    ]


# Descriptive-only columns — never used for matching, so (unlike the core
# aliases below) they get no positional fallback: a header miss just leaves
# the field blank rather than risking a wrong-column guess.
_EXTRA_COL_NAMES: dict[str, set[str]] = {
    "test_note":     {"测试"},
    "color_summary": {"色汇总英文中文色号色卡本", "色汇总", "颜色汇总",
                      "色汇总 英文 中文 色号 色卡本",
                      "颜色汇总 英文 中文 色号 色卡本"},
    "launch_date":   {"launch date"},
    "remarks":       {"备注"},
}


def parse_progress_rows(source, sheet_name: str | None = None) -> list[dict]:
    """Parse a 大货进度表 file into a flat list of record dicts.

    *source* may be a file path (str), raw bytes, or an already-open
    file-like/BytesIO object. This is the single source of truth for what
    a 合同号/款式/颜色 row means — shared by ``ProgressLookup`` (loads a
    session-uploaded file transiently, via ``_load()``/``_index_records()``)
    and the persistent progress-records store
    (``po_extractor/store/_po_store_progress.py``, saves a durable copy so
    the file doesn't need re-uploading every run).

    Each record carries both ``style`` (the normalised key, e.g.
    ``"ZLD060S24DTR003"`` — used internally for matching) and
    ``style_display`` (the original text, e.g. ``"ZLD060/S24DTR003"`` — for
    human-facing display and DB storage).
    """
    import io as _io

    import openpyxl as _opxl
    import pandas as pd

    if isinstance(source, (bytes, bytearray, memoryview)):
        src: str | _io.BytesIO = _io.BytesIO(bytes(source))
    elif hasattr(source, "seek") and hasattr(source, "read"):
        source.seek(0)
        src = source
    else:
        src = source   # file path string

    wb_meta = _opxl.load_workbook(src, read_only=True, data_only=True)
    resolved_sheet = None
    if sheet_name and sheet_name in wb_meta.sheetnames:
        resolved_sheet = sheet_name
    else:
        for name in wb_meta.sheetnames:
            if "zalando" in name.lower() or "2026" in name.lower():
                resolved_sheet = name
                break
        if resolved_sheet is None:
            resolved_sheet = wb_meta.sheetnames[0]
    wb_meta.close()

    if hasattr(src, "seek"):
        src.seek(0)

    df = pd.read_excel(
        src, sheet_name=resolved_sheet, header=None, dtype=str, engine="openpyxl"
    )
    df = df.fillna("")

    col_names = {**ProgressLookup._COL_NAMES, **_EXTRA_COL_NAMES}
    col: dict[str, int] = {}
    for ci, val in enumerate(df.iloc[0]):
        # Collapse embedded newlines/multi-spaces before alias matching — some
        # headers wrap onto two lines in the source file (e.g. "色汇总\n英文
        # 中文 色号 色卡本"), which would otherwise never match any alias.
        raw = re.sub(r'\s+', ' ', _v(val).lower()).strip()
        for key, names in col_names.items():
            if raw in names and key not in col:
                col[key] = ci
                break
    for k, v in ProgressLookup._COL_DEFAULTS.items():
        col.setdefault(k, v - 1)   # 1-based default -> 0-based

    def _cv(row, key):
        c = col.get(key)
        if c is not None and c < len(row):
            return row.iloc[c]
        return None

    # See the identical guard in ProgressLookup._load()'s history — some
    # 大货进度表 exports never populate 序号 at all; enforcing the check
    # unconditionally would then discard every row in the file.
    seq_col_used = False
    if "seq" in col:
        for v in df.iloc[1:, col["seq"]]:
            try:
                int(float(str(v)))
                seq_col_used = True
                break
            except (ValueError, TypeError):
                continue

    records: list[dict] = []
    for _, row in df.iloc[1:].iterrows():
        style_display = _v(_cv(row, "style"))
        style = _norm_key(style_display)
        if not style:
            continue
        if seq_col_used:
            seq_raw = _cv(row, "seq")
            try:
                int(float(str(seq_raw)))
            except (ValueError, TypeError):
                continue

        color_raw     = _v(_cv(row, "color"))
        cn_code_raw   = _v(_cv(row, "cn_code"))
        cn_color_raw  = _clean_cn_color(_v(_cv(row, "cn_color")))
        color_summary = _v(_cv(row, "color_summary"))
        # color_code priority: explicit "中文颜色代码" column > parsed from en_color string
        color_code = cn_code_raw or _extract_color_code(color_raw)

        shared = {
            "contract_no":   _v(_cv(row, "contract_no")),
            "pc_no":         _v(_cv(row, "pc_no")),
            "image_id":      _dispimg_id(_cv(row, "image")),
            "style":         style,
            "style_display": style_display,
            "label_color":   _v(_cv(row, "label_color")),
            "ex_fty":        _v(_cv(row, "ex_fty")),
            "qty":           _cv(row, "qty"),
            "zalando_po":    _v(_cv(row, "zalando_po")),
            "brand":         _v(_cv(row, "brand")),
            "fabric":        _v(_cv(row, "fabric")),
            "test_note":     _v(_cv(row, "test_note")),
            "color_summary": color_summary,
            "launch_date":   _v(_cv(row, "launch_date")),
            "remarks":       _v(_cv(row, "remarks")),
        }
        # A two-tone style (e.g. "52#/白色 3#" in the code cell) becomes one
        # record per colour, so each colour is independently look-up-able —
        # see _split_multi_color(). The common single-colour case returns
        # exactly one component, unchanged.
        for comp in _split_multi_color(color_raw, cn_color_raw, color_code, color_summary):
            records.append({
                **shared,
                "color":      comp["color"],
                "color_norm": _normalise_color(comp["color"]),
                "color_code": comp["color_code"],
                "cn_color":   comp["cn_color"],
            })
    return records


class ProgressLookup:
    """
    Lazy-loading production progress lookup.

    Parameters
    ----------
    path : str
        Full path to 大货进度表 xlsx file.
    sheet_name : str
        Sheet to read (default "2026 Zalando ").
    """

    def __init__(self, path: str | bytes | None = None,
                 sheet_name: str | None = None, *,
                 data: bytes | None = None):
        """
        Parameters
        ----------
        path : str | bytes | None
            Path to the 大货进度表 xlsx file **or** the raw file bytes.
            Pass bytes (or use the *data* keyword) to avoid writing a temp
            file — openpyxl / pandas both accept BytesIO directly.
        data : bytes | None
            Alternative keyword-only way to supply raw bytes.
        sheet_name : str | None
            Sheet to read (``None`` → auto-detect first matching sheet).
        """
        import io as _io
        _raw: bytes | None = data if data is not None else (
            path if isinstance(path, (bytes, bytearray, memoryview)) else None
        )
        if _raw is not None:
            # Store a BytesIO; _path is unused in this mode
            self._path   = None
            self._source = _io.BytesIO(bytes(_raw))
        else:
            self._path   = path     # str path — original behaviour
            self._source = None
        self._sheet_name  = sheet_name   # None = auto-detect first matching sheet
        # PRIMARY:  (pc_no_norm, style_norm, color_norm) → record
        #           Sky East PC No (所在PO col 2) — most reliable join key in 大货进度表
        self._by_pc_style_color: dict[tuple[str, str, str], dict] = {}
        # FALLBACK1: (zalando_po_norm, style_norm, color_norm) → record
        #            Zalando PO# (col 12) — often blank in 大货进度表
        self._by_po_style_color: dict[tuple[str, str, str], dict] = {}
        # FALLBACK2: (style_norm, color_norm) → first record
        self._by_style_color: dict[tuple[str, str], dict] = {}
        # FALLBACK3: (pc_no_norm, style_norm, color_code) → record
        self._by_pc_style_code: dict[tuple[str, str, str], dict] = {}
        # FALLBACK4: (style_norm, color_code) → first record
        self._by_style_code: dict[tuple[str, str], dict] = {}
        # FALLBACK5: (pc_no_norm, style_norm) → first record
        #            When color value is too messy to match — PC+style alone is
        #            usually unique enough in 大货进度表
        self._by_pc_style: dict[tuple[str, str], dict] = {}
        # FALLBACK6: style_norm → [record, ...]
        self._by_style: dict[str, list[dict]] = {}
        self._loaded = False

    # ── Column map helpers ─────────────────────────────────────────────────────

    _COL_NAMES = {
        "seq":        {"序号"},
        "contract_no":{"合同号"},
        # PC No column header has been renamed across file versions, but the
        # data is always the same: Sky East PC No. (e.g. HHPPC040).
        # In some versions the header reads "客人PO" — but the values are still
        # HHPPC… Sky East PC numbers, NOT the buyer's actual PO number.
        "pc_no":      {"所在po", "客人pc no", "客人pcno", "客人pc",
                       "客人pc no.", "客人 pc no", "pc no", "pc no.",
                       "客人po"},
        "image":      {"image"},
        "style":      {"款式"},
        "color":      {"颜色", "英文颜色"},
        "label_color":{"主标颜色"},
        "cn_color":   {"中文颜色"},
        "cn_code":    {"中文颜色代码", "颜色代码"},
        "ex_fty":     {"po离厂日期"},
        "qty":        {"数量"},
        # Buyer's actual PO number (PO2294110C style) — column may be absent
        "zalando_po": {"po#", "po号", "buyer po", "客户po"},
        "brand":      {"brand"},
        "fabric":     {"fabricdetail"},
    }
    # Default column positions (1-based) used as fallback when the header is
    # absent from the workbook.  These match the legacy layout where 主标颜色
    # is followed immediately by PO离厂日期 / 数量 / PO# / BRAND / FABRICDETAIL
    # at cols 9-13 and there are NO cn_color / cn_code columns.
    #
    # Notes on the optional columns:
    #   • ``cn_color`` (中文颜色) and ``cn_code`` (中文颜色代码) are intentionally
    #     OMITTED from the defaults.  They were added in newer file revisions
    #     between 主标颜色 (col 8) and PO离厂日期; if a workbook lacks those
    #     headers we must NOT default them to cols 9/10 — that would silently
    #     read ex-factory dates and quantities into the Chinese-colour fields,
    #     producing corrupt buy-plan output.  Missing → return ``None`` from
    #     ``_cv`` → record fields end up as empty strings, which the consumer
    #     treats as "no PC-keyed match" and falls back to the Internal DB.
    #   • ``ex_fty`` / ``qty`` defaults reflect the *new* layout (with the two
    #     extra columns).  In legacy files the headers "PO离厂日期" / "数量"
    #     are present and override these defaults to cols 9/10 anyway.
    _COL_DEFAULTS = {
        "seq": 1, "contract_no": 2, "pc_no": 3, "image": 4,
        "style": 5, "color": 7, "label_color": 8,
        "ex_fty": 11, "qty": 12, "zalando_po": 13, "brand": 14, "fabric": 15,
    }

    def _map_cols(self, ws) -> dict[str, int]:
        col_map = {}
        # Read only header row via iter_rows (safe in read_only mode)
        for row_vals in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            for ci, val in enumerate(row_vals, start=1):
                raw = _v(val).lower().strip()
                for key, names in self._COL_NAMES.items():
                    if raw in names and key not in col_map:
                        col_map[key] = ci
                        break
        for k, v in self._COL_DEFAULTS.items():
            col_map.setdefault(k, v)
        return col_map

    # ── Lazy load ─────────────────────────────────────────────────────────────

    def _load(self):
        if self._loaded:
            return
        records = parse_progress_rows(
            self._source if self._source is not None else self._path,
            self._sheet_name,
        )
        self._index_records(records)
        self._loaded = True

    def _index_records(self, records: list[dict]) -> None:
        """Populate every lookup index from already-parsed records.

        Shared by ``_load()`` (session-uploaded file, parsed via
        ``parse_progress_rows()``) and ``from_records()`` (records already
        loaded from the persistent DB store — no file parsing needed).
        """
        for record in records:
            style = record["style"]
            pcn   = _norm_key(record["pc_no"])
            zpo   = _norm_key(record["zalando_po"])
            cnorm = record["color_norm"]
            # Index codes in the same normalised form _lookup() uses for the
            # caller's input ("52#" / "52" / "#52" all collapse to "52").
            # Without this, codes stored as "52#" never match callers passing
            # "52" and the (style, color_code) fallback tier silently misses.
            ccode_key = _norm_key(record["color_code"])

            # Build all lookup indexes (lookup priority is enforced in _lookup())
            if pcn and cnorm:
                self._by_pc_style_color.setdefault((pcn, style, cnorm), record)
            if pcn and ccode_key:
                self._by_pc_style_code.setdefault((pcn, style, ccode_key), record)
            if zpo and cnorm:
                self._by_po_style_color.setdefault((zpo, style, cnorm), record)
            if cnorm:
                self._by_style_color.setdefault((style, cnorm), record)
            if ccode_key:
                self._by_style_code.setdefault((style, ccode_key), record)
            if pcn:
                self._by_pc_style.setdefault((pcn, style), record)
            self._by_style.setdefault(style, []).append(record)

    @classmethod
    def from_records(cls, records: list[dict]) -> "ProgressLookup":
        """Build a ready-to-query ``ProgressLookup`` directly from
        already-parsed records — no file I/O.

        Used to serve buy-plan generation from the persistent progress-records
        DB store (``po_extractor/store/_po_store_progress.py``) instead of
        requiring the 大货进度表 file to be re-uploaded every run.
        """
        obj = cls()
        obj._index_records(records)
        obj._loaded = True
        return obj

    # ── Public API ────────────────────────────────────────────────────────────

    @staticmethod
    def _skey(style_no: str) -> str:
        return _norm_key(style_no)

    def _lookup(self, style_no: str, color: str, zalando_po: str = "",
                color_code: str = "", pc_no: str = "") -> dict | None:
        """
        Core record lookup with multi-tier fallback.  Priority:
          1. (pc_no, style, color_name)      — PRIMARY: Sky East PC No (most reliable)
          2. (zalando_po, style, color_name) — Zalando PO# (often blank in 大货进度表)
          3. (style, color_name)             — when no PO/PC matches
          4. (pc_no, style, color_code)      — code fallback with PC
          5. (style, color_code)             — code fallback alone
          6. (pc_no, style)                  — colour value too messy to match
          7. first row for style             — last resort

        PC scoping
        ----------
        When *pc_no* is supplied the result is constrained to that PC: the
        PC-less fallback tiers (style+color, style+code, style-only) are
        skipped to avoid leaking rows from a *different* PC contract.
        Without this gate, a caller asking for ``HHPPC040`` could silently
        receive a record from ``HHPPC041`` whenever the requested PC's row
        was missing — wrong contract, wrong dates, wrong qtys.
        """
        self._load()
        s     = self._skey(style_no)
        cnorm = _normalise_color(color) if color else ""
        zpo   = _norm_key(zalando_po)   if zalando_po else ""
        ccode = _norm_key(color_code)   if color_code else ""
        pcn   = _norm_key(pc_no)        if pc_no else ""

        # PRIMARY: (pc_no, style, color_name)
        if pcn and s and cnorm:
            rec = self._by_pc_style_color.get((pcn, s, cnorm))
            if rec:
                return rec

        # FALLBACK1: (zalando_po, style, color_name)
        if zpo and s and cnorm:
            rec = self._by_po_style_color.get((zpo, s, cnorm))
            if rec:
                return rec

        # FALLBACK3: (pc_no, style, color_code)
        if pcn and s and ccode:
            rec = self._by_pc_style_code.get((pcn, s, ccode))
            if rec:
                return rec

        # FALLBACK5: (pc_no, style) — color too messy, but PC+style is unique
        if pcn and s:
            rec = self._by_pc_style.get((pcn, s))
            if rec:
                return rec

        # When the caller specified a PC, do NOT cross into other PCs'
        # rows — return None so they can decide what to do with the miss.
        if pcn:
            return None

        # FALLBACK2: (style, color_name)  — only when no PC was given
        if s and cnorm:
            rec = self._by_style_color.get((s, cnorm))
            if rec:
                return rec

        # FALLBACK4: (style, color_code)
        if s and ccode:
            rec = self._by_style_code.get((s, ccode))
            if rec:
                return rec

        # FALLBACK6: style only — last resort
        rows = self._by_style.get(s, [])
        return rows[0] if rows else None

    def get_contract_no(self, style_no: str, color: str = "",
                        zalando_po: str = "", color_code: str = "",
                        pc_no: str = "") -> str:
        """
        Return HHN 合同号.

        Lookup priority:
          1. (pc_no, style, color)       — PRIMARY: Sky East PC No (most reliable)
          2. (zalando_po, style, color)  — Zalando PO# (often blank in 大货进度表)
          3. (style, color)              — when no PO/PC matches
          4. (pc_no, style, color_code)  — code fallback with PC
          5. (style, color_code)         — code fallback alone
          6. first row for style         — last resort
        """
        rec = self._lookup(style_no, color, zalando_po, color_code, pc_no)
        return rec["contract_no"] if rec else ""

    def get_image_id(self, style_no: str, color: str = "",
                     zalando_po: str = "", color_code: str = "",
                     pc_no: str = "") -> str:
        """Return DISPIMG image ID from the progress table."""
        rec = self._lookup(style_no, color, zalando_po, color_code, pc_no)
        if rec and rec["image_id"]:
            return rec["image_id"]
        # image_id may be blank on the matched row — scan all rows for style
        self._load()
        for r in self._by_style.get(self._skey(style_no), []):
            if r["image_id"]:
                return r["image_id"]
        return ""

    def get_record(self, style_no: str, color: str = "",
                   zalando_po: str = "", color_code: str = "",
                   pc_no: str = "") -> dict | None:
        """Return the full record dict or None."""
        return self._lookup(style_no, color, zalando_po, color_code, pc_no)

    def get_cn_color(self, style_no: str, color: str = "",
                     zalando_po: str = "", color_code: str = "",
                     pc_no: str = "") -> str:
        """Return 中文颜色 (Chinese colour name) from the matched row."""
        rec = self._lookup(style_no, color, zalando_po, color_code, pc_no)
        return rec["cn_color"] if rec else ""

    def get_color_code(self, style_no: str, color: str = "",
                       zalando_po: str = "", color_code: str = "",
                       pc_no: str = "") -> str:
        """Return 中文颜色代码 (CN colour code) from the matched row."""
        rec = self._lookup(style_no, color, zalando_po, color_code, pc_no)
        return rec["color_code"] if rec else ""

    def get_label_color(self, style_no: str, color: str = "",
                        zalando_po: str = "", color_code: str = "",
                        pc_no: str = "") -> str:
        """Return 主标颜色 (main label colour) from the matched row."""
        rec = self._lookup(style_no, color, zalando_po, color_code, pc_no)
        return rec["label_color"] if rec else ""

    def get_all_for_style(self, style_no: str) -> list[dict]:
        """Return all colour rows for a style."""
        self._load()
        return self._by_style.get(self._skey(style_no), [])

    def all_styles(self) -> list[str]:
        self._load()
        return list(self._by_style.keys())

    def _build_color_lookup(self, company: str, field: str) -> dict:
        """Build a ``{(company, brand, en_color_norm): value}`` dict from records.

        Compatible with the buyplan exporter format produced by
        ``ColorTranslationStore.build_lookup_dict()``.  Two keys are inserted per
        record — one for the cleaned colour (numeric prefix stripped) and one
        for the raw colour — so lookups succeed whether the item's
        ``color_name`` column contains "NAVY" or "52# NAVY".

        Parameters
        ----------
        company : str
            Client/company name (e.g. ``COMPANY_SKY_EAST``) for the dict key.
        field : str
            Record field to read as the value — ``"cn_color"`` or
            ``"label_color"``.
        """
        from po_extractor.store.color_translation_store import _normalize_color_name
        self._load()
        result: dict[tuple[str, str, str], str] = {}
        seen: set[tuple[str, str, str, str]] = set()
        for records in self._by_style.values():
            for rec in records:
                value = (rec.get(field) or "").strip()
                if not value:
                    continue
                brand    = (rec.get("brand") or "").strip()
                raw_col  = rec.get("color", "")
                norm_col = rec.get("color_norm", "")
                dedup_key = (brand, raw_col, norm_col, value)
                if dedup_key in seen:
                    continue
                seen.add(dedup_key)
                # Key using cleaned colour (numeric code stripped) and raw colour
                # — covers both "NAVY" and "52# NAVY" forms in items.color_name.
                for col_form in (norm_col, raw_col):
                    result.setdefault(
                        (company, brand, _normalize_color_name(col_form)), value
                    )
        return result

    def build_cn_and_code_lookups(self, company: str) -> tuple[dict, dict]:
        """Build both cn_lookup and cn_code_lookup in a single pass.

        Same key format as ``_build_color_lookup`` — two results instead of one
        so callers avoid iterating the data twice for consecutive lookups.

        Returns
        -------
        cn_lookup : dict
            ``{(company, brand, en_color_norm): cn_color}``
        cn_code_lookup : dict
            ``{(company, brand, en_color_norm): color_code}``
        """
        from po_extractor.store.color_translation_store import _normalize_color_name
        self._load()
        cn_result:   dict = {}
        code_result: dict = {}
        seen: set = set()
        for records in self._by_style.values():
            for rec in records:
                cn_color = (rec.get("cn_color")   or "").strip()
                code     = (rec.get("color_code") or "").strip()
                if not (cn_color or code):
                    continue
                brand    = (rec.get("brand")       or "").strip()
                raw_col  = rec.get("color",      "")
                norm_col = rec.get("color_norm", "")
                dedup = (brand, raw_col, norm_col, cn_color, code)
                if dedup in seen:
                    continue
                seen.add(dedup)
                for col_form in (norm_col, raw_col):
                    if not col_form:
                        continue
                    nk = _normalize_color_name(col_form)
                    if not nk:
                        continue
                    key = (company, brand, nk)
                    if cn_color:
                        cn_result.setdefault(key, cn_color)
                    if code:
                        code_result.setdefault(key, code)
        return cn_result, code_result

    def build_pc_style_color_lookups(self) -> dict:
        """Build a combined PC-No.-keyed lookup for cn_color, color_code, and label_color.

        Returns ``{(pc_no_norm, style_norm, color_norm): PCColorMatch}``
        so the exporter can retrieve all three values in a single ``.get()`` call.

        Using PC No. + style + color gives the most accurate match because it
        is scoped to the exact Sky East contract — the same colour name can
        carry different codes across contracts.

        Two color forms per record (raw + ``color_norm``) are stored so lookups
        succeed whether ``color_name`` in items is ``"blue"`` or ``"79# blue"``.
        """
        from po_extractor.store.color_translation_store import (
            _normalize_color_name as _nz,
        )
        self._load()
        result: dict = {}
        seen:   set  = set()
        for records in self._by_style.values():
            for rec in records:
                pcn         = _norm_key(rec.get("pc_no")      or "")
                sn          = _norm_key(rec.get("style")       or "")
                cn_color    = (rec.get("cn_color")    or "").strip()
                code        = (rec.get("color_code")  or "").strip()
                label_color = (rec.get("label_color") or "").strip()
                if not (pcn and sn and (cn_color or code or label_color)):
                    continue
                for col_form in (rec.get("color", ""), rec.get("color_norm", "")):
                    if not col_form:
                        continue
                    nk = _nz(col_form)
                    if not nk:
                        continue
                    dedup = (pcn, sn, nk)
                    if dedup not in seen:
                        seen.add(dedup)
                        result.setdefault(
                            (pcn, sn, nk),
                            PCColorMatch(cn_color, code, label_color),
                        )
        return result

    def build_all_color_lookups(
        self, company: str,
    ) -> tuple[dict, dict, dict, dict]:
        """Build all four color lookups in a SINGLE pass over the data.

        Roughly 3× faster than calling ``build_cn_and_code_lookups`` +
        ``build_label_lookup`` + ``build_pc_style_color_lookups`` separately —
        each of those iterates ``self._by_style.values()`` independently and
        re-normalises the same colour strings.

        Returns
        -------
        cn_lookup : dict
            ``{(company, brand, en_color_norm): cn_color}``
        cn_code_lookup : dict
            ``{(company, brand, en_color_norm): color_code}``
        label_lookup : dict
            ``{(company, brand, en_color_norm): label_color}``
        pc_style_lookup : dict
            ``{(pc_no_norm, style_norm, en_color_norm): PCColorMatch}``
        """
        from po_extractor.store.color_translation_store import _normalize_color_name
        self._load()
        cn_result:    dict = {}
        code_result:  dict = {}
        label_result: dict = {}
        pc_result:    dict = {}
        flat_seen: set = set()   # dedupe brand-keyed inserts
        pc_seen:   set = set()   # dedupe PC-keyed inserts

        for records in self._by_style.values():
            for rec in records:
                cn_color    = (rec.get("cn_color")    or "").strip()
                code        = (rec.get("color_code")  or "").strip()
                label_color = (rec.get("label_color") or "").strip()
                if not (cn_color or code or label_color):
                    continue

                brand    = (rec.get("brand")      or "").strip()
                raw_col  = rec.get("color",       "")
                norm_col = rec.get("color_norm",  "")
                pcn      = _norm_key(rec.get("pc_no") or "")
                sn       = _norm_key(rec.get("style") or "")

                flat_dedup = (brand, raw_col, norm_col, cn_color, code, label_color)
                do_flat = flat_dedup not in flat_seen
                if do_flat:
                    flat_seen.add(flat_dedup)
                do_pc = bool(pcn and sn)
                if not (do_flat or do_pc):
                    continue

                for col_form in (norm_col, raw_col):
                    if not col_form:
                        continue
                    nk = _normalize_color_name(col_form)
                    if not nk:
                        continue
                    if do_flat:
                        flat_key = (company, brand, nk)
                        if cn_color:
                            cn_result.setdefault(flat_key, cn_color)
                        if code:
                            code_result.setdefault(flat_key, code)
                        if label_color:
                            label_result.setdefault(flat_key, label_color)
                    if do_pc:
                        pc_dedup = (pcn, sn, nk)
                        if pc_dedup not in pc_seen:
                            pc_seen.add(pc_dedup)
                            pc_result[(pcn, sn, nk)] = PCColorMatch(
                                cn_color, code, label_color,
                            )
        return cn_result, code_result, label_result, pc_result

    # ── Backward-compatible thin wrappers ────────────────────────────────────

    def build_cn_lookup(self, company: str) -> dict:
        """Return cn_lookup only.  Prefer ``build_all_color_lookups`` for new code."""
        cn, _ = self.build_cn_and_code_lookups(company)
        return cn

    def build_cn_code_lookup(self, company: str) -> dict:
        """Return cn_code_lookup only.  Prefer ``build_all_color_lookups`` for new code."""
        _, code = self.build_cn_and_code_lookups(company)
        return code

    def build_label_lookup(self, company: str) -> dict:
        """Build a label_lookup dict (主标颜色).  See ``_build_color_lookup``."""
        return self._build_color_lookup(company, "label_color")

    def __len__(self) -> int:
        self._load()
        return len(self._by_style_color)
