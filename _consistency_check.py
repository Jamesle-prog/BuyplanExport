"""One-off consistency report on 大货进度表--Angel 2026.xlsx.

Scans 颜色 / 主标颜色 / 中文颜色 / Brand columns and reports:
  1. Same (brand, EN) mapped to multiple Chinese translations
  2. Same (brand, EN) mapped to multiple 主标颜色 values
  3. Manual 主标颜色 entries that disagree with the light/dark keyword rule
  4. Coverage stats (% rows with each column filled, per brand)
  5. Same Chinese colour name reused for multiple English colours
"""
import sys
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import collections
from openpyxl import load_workbook
from po_extractor.store.color_translation_store import _normalize_color_name
from po_extractor.exporters._sky_east_helpers import (
    _LIGHT_BODY_KEYWORDS, _DARK_BODY_KEYWORDS,
)

PATH = "zalandoFile/大货进度表--Angel 2026.xlsx"

# ── Read all rows ────────────────────────────────────────────────────────────
wb = load_workbook(PATH, read_only=True, data_only=True)
rows_data = []
for sn in wb.sheetnames:
    try:
        ws = wb[sn]
    except Exception:
        continue
    cols = {}
    header_row = None
    for ri, row in enumerate(ws.iter_rows(max_row=5, values_only=True), 1):
        for ci, v in enumerate(row, 1):
            if not isinstance(v, str):
                continue
            v = v.strip()
            if v in ("颜色", "Color") and "color" not in cols:
                cols["color"] = ci
            elif v == "主标颜色":
                cols["main"] = ci
            elif v in ("中文颜色", "颜色(中文)", "颜色（中文）"):
                cols["cn"] = ci
            elif v in ("Brand", "BRAND", "客户品牌", "品牌"):
                cols.setdefault("brand", ci)
            elif v in ("款式", "Style", "款号", "style"):
                cols.setdefault("style", ci)
            elif v in ("PO#", "PO号"):
                cols.setdefault("po", ci)
        if "color" in cols and "cn" in cols:
            header_row = ri
            break
    if not header_row:
        continue
    for ri, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True),
                              header_row + 1):
        if cols["color"] > len(row):
            continue
        ce_raw = row[cols["color"] - 1]
        if not ce_raw or not str(ce_raw).strip():
            continue
        en = _normalize_color_name(ce_raw)

        def cell(key):
            ci = cols.get(key)
            if not ci or ci > len(row):
                return ""
            v = row[ci - 1]
            return str(v).strip() if v is not None else ""

        cm = cell("main")
        cc = cell("cn")
        br = cell("brand")
        po = cell("po")
        sty = cell("style")
        # Filter out 'None'/'nan' literal strings
        cm = "" if cm.lower() in ("none", "nan", "") else cm
        cc = "" if cc.lower() in ("none", "nan", "") else cc

        rows_data.append({
            "sheet": sn, "row": ri, "po": po,
            "brand": br, "style": sty, "en": en,
            "main": cm, "cn": cc,
        })
wb.close()

print(f"Total non-empty rows scanned: {len(rows_data)}")
print()


def classify_shade(en: str) -> str:
    t = en.lower()
    light_pos = min((t.find(k) for k in _LIGHT_BODY_KEYWORDS if k in t), default=-1)
    dark_pos  = min((t.find(k) for k in _DARK_BODY_KEYWORDS  if k in t), default=-1)
    if light_pos == -1 and dark_pos == -1:
        return "unknown"
    if dark_pos == -1:
        return "light"
    if light_pos == -1:
        return "dark"
    return "dark" if dark_pos < light_pos else "light"


# ── 1. Same (brand, EN) → multiple CN ────────────────────────────────────────
print("=" * 78)
print("CHECK 1: Same (brand, EN) → multiple distinct CN translations")
print("=" * 78)
ct_map = collections.defaultdict(set)
ct_ex  = collections.defaultdict(list)
for r in rows_data:
    if r["cn"]:
        ct_map[(r["brand"], r["en"])].add(r["cn"])
        ct_ex[(r["brand"], r["en"])].append(r)

inconsistent_cn = [(k, v) for k, v in ct_map.items() if len(v) > 1]
print(f"Found {len(inconsistent_cn)} (brand, EN) keys with multiple CN values:")
for (brand, en), cns in inconsistent_cn:
    print(f"  {(brand or '(blank)')[:25]:25s} | {en[:30]:30s} → {sorted(cns)}")

# ── 2. Same (brand, EN) → multiple 主标颜色 ─────────────────────────────────
print()
print("=" * 78)
print("CHECK 2: Same (brand, EN) → multiple distinct 主标颜色")
print("=" * 78)
label_map = collections.defaultdict(set)
for r in rows_data:
    if r["main"]:
        label_map[(r["brand"], r["en"])].add(r["main"])
inconsistent_label = [(k, v) for k, v in label_map.items() if len(v) > 1]
print(f"Found {len(inconsistent_label)} (brand, EN) keys with multiple 主标颜色:")
for (brand, en), labels in inconsistent_label:
    print(f"  {(brand or '(blank)')[:25]:25s} | {en[:30]:30s} → {sorted(labels)}")

# ── 3. Rule violations (manual ≠ keyword classifier) ─────────────────────────
print()
print("=" * 78)
print("CHECK 3: Manual 主标颜色 vs keyword rule (light→白色, dark→黑色)")
print("=" * 78)
violations = []
unknown = []
for r in rows_data:
    if not r["main"]:
        continue
    shade = classify_shade(r["en"])
    if shade == "unknown":
        unknown.append(r)
        continue
    expected = "黑色" if shade == "dark" else "白色"
    if r["main"] != expected:
        violations.append((r, shade, expected))

print(f"Rule violations: {len(violations)}")
unique_v = sorted({(r["en"], r["main"], shade, exp) for r, shade, exp in violations})
for en, main, shade, exp in unique_v:
    n_examples = sum(1 for r, s, e in violations if r["en"] == en and r["main"] == main)
    print(f"  EN={en[:32]:32s} | shade={shade:6s} | manual={main!r:6} | rule says {exp!r:6} | ×{n_examples}")

print()
print(f"Colours the keyword rule cannot classify: "
      f"{len({(r['en'], r['main']) for r in unknown})} unique")
unique_unknown = sorted({(r["en"], r["main"]) for r in unknown})
for en, main in unique_unknown:
    print(f"  EN={en[:50]:50s} | manual main={main!r}")

# ── 4. Coverage stats ────────────────────────────────────────────────────────
print()
print("=" * 78)
print("CHECK 4: Coverage stats (% rows with each column filled)")
print("=" * 78)
total = len(rows_data)
have_cn   = sum(1 for r in rows_data if r["cn"])
have_main = sum(1 for r in rows_data if r["main"])
have_both = sum(1 for r in rows_data if r["cn"] and r["main"])
print(f"  Total rows:                    {total}")
print(f"  With 中文颜色 filled:           {have_cn:3d}  ({100*have_cn//total if total else 0}%)")
print(f"  With 主标颜色 filled:           {have_main:3d}  ({100*have_main//total if total else 0}%)")
print(f"  With BOTH filled:              {have_both:3d}  ({100*have_both//total if total else 0}%)")
print()
print("  By brand:")
brand_stats = collections.defaultdict(lambda: [0, 0, 0])
for r in rows_data:
    s = brand_stats[r["brand"]]
    s[0] += 1
    if r["cn"]:
        s[1] += 1
    if r["main"]:
        s[2] += 1
for brand, (n, ncn, nmain) in sorted(brand_stats.items()):
    label = brand or "(blank)"
    print(f"    {label[:22]:22s}: {n:3d} rows | CN={ncn:3d} ({100*ncn//n:3d}%) "
          f"| main={nmain:3d} ({100*nmain//n:3d}%)")

# ── 5. Same CN → multiple EN ─────────────────────────────────────────────────
print()
print("=" * 78)
print("CHECK 5: Same CN colour → multiple EN names (likely fine, FYI)")
print("=" * 78)
cn_to_en = collections.defaultdict(set)
for r in rows_data:
    if r["cn"]:
        cn_to_en[r["cn"]].add(r["en"])
multi_en = sorted([(cn, ens) for cn, ens in cn_to_en.items() if len(ens) > 1])
print(f"{len(multi_en)} CN colours appear under multiple EN names:")
for cn, ens in multi_en:
    print(f"  CN={cn[:30]:30s} → {sorted(ens)}")
