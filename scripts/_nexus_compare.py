"""One-off: extract all Nexus PDFs and compare with legacy DB data."""
import sys, os, re
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
import pdfplumber
from po_extractor.store.po_store import POStore
from po_extractor.config import DB_PATH
import pandas as pd

NEXUS_DIR = (
    r'C:\Users\Administrator\Mountain Duck'
    r'\873c48f1-d867-4501-87f7-c1b938827cf4'
    r'\业务三部一组\4.SUITS\SUITS-CK\PO\HOL 26 ROSS\Nexus po'
)


def extract_nexus(fpath):
    with pdfplumber.open(fpath) as pdf:
        text = pdf.pages[0].extract_text() or ''

    # "Order Number Issue Date Version\nCSKHHN015R 2026-05-18 ..."
    po_m    = re.search(r'Order Number\s+Issue Date\s+Version\s+(CSKHH\w+)', text)
    if not po_m:                                    # fallback: first CSKHH token
        po_m = re.search(r'\b(CSKHHN\d{3,4}R)\b', text)
    # FOB stated in Special Instructions: "FOB: $3.60" (dollar sign literal in text)
    fob_m   = re.search(r'FOB:\s+\$\s*([\d.]+)', text)
    if not fob_m:
        fob_m = re.search(r'FOB:\s*\$([\d.]+)', text)
    # Line item row: "1 STYLE CCODE COLORNAME - EA EA UNITS ..."
    # Color code may contain digits (e.g. C3R).  Cost P/U may have column-
    # overflow spaces (e.g. "2 . 8 5") so we get it from the summary line instead.
    li_m    = re.compile(
        r'^1\s+([A-Z0-9]{6,10})\s+([A-Z][A-Z0-9]{1,3})\s+([\w\s/()\-+.]+?)\s+-\s+EA\s+EA\s+(\d+)',
        re.MULTILINE
    ).search(text)

    # "Total Qty  Unit Cost  Extended Cost\n8,400  3.24 USD  27,216.00 USD"
    # This row is always clean (no column-overflow issues).
    tq_m    = re.search(
        r'Total Qty\s+Unit Cost\s+Extended Cost\s+([\d,]+)\s+([\d.]+)\s+USD',
        text
    )

    # Size / Qty block
    size_m = re.search(r'Size:\s+([\S ]+)', text)
    qty_m  = re.search(r'Qty:\s+([\S ,]+)', text)

    sizes_dict = {}
    if size_m and qty_m:
        sz_tok  = size_m.group(1).split()
        qty_tok = qty_m.group(1).replace(',', '').split()
        for s, q in zip(sz_tok, qty_tok):
            if s != '-' and q != '-':
                try:
                    sizes_dict[s] = int(q)
                except ValueError:
                    pass

    return {
        'po_number':   po_m.group(1)                          if po_m  else '?',
        'style_nexus': li_m.group(1)                          if li_m  else '?',
        'color_code':  li_m.group(2)                          if li_m  else '?',
        'color_name':  li_m.group(3).strip()                  if li_m  else '?',
        'fob_stated':  float(fob_m.group(1))                  if fob_m else None,
        'cost_pu':     float(tq_m.group(2))                   if tq_m  else None,
        'total_qty':   int(tq_m.group(1).replace(',', ''))    if tq_m  else (int(li_m.group(4)) if li_m else 0),
        'sizes':       sizes_dict,
    }


# ── Parse all 16 Nexus PDFs ──────────────────────────────────────────────────
files = sorted(f for f in os.listdir(NEXUS_DIR)
               if f.lower().endswith('.pdf') and f.lower().startswith('cskhha'))

nexus_rows = []
for fname in files:
    r = extract_nexus(os.path.join(NEXUS_DIR, fname))
    r['file'] = fname
    nexus_rows.append(r)

# ── Load legacy from DB ──────────────────────────────────────────────────────
store     = POStore(DB_PATH)
po_nums   = [r['po_number'] for r in nexus_rows if r['po_number'] != '?']
df_meta   = store.list_pos()
df_meta   = df_meta[df_meta['po_number'].isin(po_nums)].copy()
df_sizes  = store.load_size_rows(po_nums)

legacy = {}
for _, row in df_meta.iterrows():
    pn  = row['po_number']
    sz  = df_sizes[df_sizes['PO Number'] == pn]
    legacy[pn] = {
        'style_legacy': row.get('style', ''),
        'fob_legacy':   row.get('unit_cost'),
        'total_qty_legacy': int(sz['Units'].sum()) if not sz.empty else 0,
        'sizes_legacy': dict(zip(sz['Size'], sz['Units'].astype(int))) if not sz.empty else {},
        'color_legacy': sz['Color'].iloc[0] if not sz.empty else '',
    }

# ── Build comparison table ───────────────────────────────────────────────────
compare_rows = []
for r in nexus_rows:
    pn  = r['po_number']
    leg = legacy.get(pn, {})

    fob_legacy  = leg.get('fob_legacy')
    fob_stated  = r['fob_stated']
    cost_pu     = r['cost_pu']

    # Discount % = 1 - cost_pu / fob_stated
    disc_pct = round((1 - cost_pu / fob_stated) * 100, 1) if fob_stated and cost_pu else None

    qty_nexus  = r['total_qty']
    qty_legacy = leg.get('total_qty_legacy', 0)

    style_match = (r['style_nexus'] == leg.get('style_legacy', '')) if r['style_nexus'] != '?' else None
    qty_match   = (qty_nexus == qty_legacy)
    fob_match   = (abs(fob_stated - float(fob_legacy)) < 0.01) if (fob_stated and fob_legacy) else None

    flags = []
    if not qty_match:           flags.append('QTY')
    if fob_match is False:      flags.append('FOB')
    if style_match is False:    flags.append('STYLE')

    compare_rows.append({
        'PO Number':       pn,
        'Style (Nexus)':   r['style_nexus'],
        'Style (Legacy)':  leg.get('style_legacy', 'N/A'),
        'Color (Nexus)':   f"{r['color_code']}/{r['color_name']}",
        'Color (Legacy)':  leg.get('color_legacy', 'N/A'),
        'FOB Stated ($)':  fob_stated,
        'Cost P/U ($)':    cost_pu,
        'Discount %':      disc_pct,
        'FOB Legacy ($)':  float(fob_legacy) if fob_legacy else None,
        'Qty (Nexus)':     qty_nexus,
        'Qty (Legacy)':    qty_legacy,
        'Sizes (Nexus)':   str(r['sizes']),
        'Sizes (Legacy)':  str(leg.get('sizes_legacy', {})),
        'Flags':           ', '.join(flags) if flags else 'OK',
    })

df = pd.DataFrame(compare_rows)

# ── Print to console ─────────────────────────────────────────────────────────
pd.set_option('display.max_columns', 20)
pd.set_option('display.width', 200)
pd.set_option('display.max_colwidth', 40)
print(df[[
    'PO Number', 'Style (Nexus)', 'Style (Legacy)',
    'Color (Nexus)', 'Color (Legacy)',
    'FOB Stated ($)', 'Cost P/U ($)', 'Discount %', 'FOB Legacy ($)',
    'Qty (Nexus)', 'Qty (Legacy)', 'Flags'
]].to_string(index=False))

print()
print(f'Total Nexus units  : {df["Qty (Nexus)"].sum():,}')
print(f'Total Legacy units : {df["Qty (Legacy)"].sum():,}')

# ── Export comparison Excel ──────────────────────────────────────────────────
out = r'C:\Users\Administrator\Desktop\CSKHHN_Nexus_vs_Legacy_Comparison.xlsx'
with pd.ExcelWriter(out, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Comparison', index=False)

    # Auto-width
    ws = writer.sheets['Comparison']
    from openpyxl.styles import PatternFill, Font
    RED  = PatternFill('solid', start_color='FFCCCC', end_color='FFCCCC')
    YELL = PatternFill('solid', start_color='FFF2CC', end_color='FFF2CC')
    GRN  = PatternFill('solid', start_color='E2EFDA', end_color='E2EFDA')
    BOLD = Font(bold=True)

    flag_col = df.columns.get_loc('Flags') + 1
    for row_idx in range(2, len(df) + 2):
        flags_val = ws.cell(row=row_idx, column=flag_col).value or ''
        fill = RED if flags_val != 'OK' else GRN
        for col in range(1, len(df.columns) + 1):
            ws.cell(row=row_idx, column=col).fill = fill
    for cell in ws[1]:
        cell.font = BOLD
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 50)

print(f'Saved -> {out}')
