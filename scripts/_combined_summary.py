"""Combine CK HOL26 / DU HOL26 / LS FALL26 Summary sheets into one Excel."""
import pandas as pd, re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SIZE_COLS = ['S', 'M', 'L', 'XL', '1X', '2X', '3X']

SOURCES = {
    'CK HOL26':  r'D:\GIII_PO\CSKHHN_VendFaxCopy_Summary_new.xlsx',
    'DU HOL26':  r'D:\GIII_PO\DUKHHA_HOL26_PO_Summary.xlsx',
    'LS FALL26': r'D:\GIII_PO\LSKHHN_FALL26_KL_POs.xlsx',
}
OUT_PATH = r'D:\GIII_PO\HHN_Combined_Summary.xlsx'

PO_RE = re.compile(r'^[A-Z]{3,}[0-9]+R$')

# ── Chinese colour map (keyed on the colour-code prefix before '/') ───────
CN_COLOR_MAP = {
    'ESP':  '深咖啡色',
    'EGG':  '茄紫色',
    'JVS':  '纯黑色',
    'C3R':  '樱桃红',
    'PNE':  '松绿色',
    'MLV':  '军绿色',
    'Q2C':  '浅天蓝/深咖啡',
    'Q16':  '花瓣粉/深咖啡',
    'Q2D':  '双色深咖啡',
    'Q15':  '浅暖灰/米黄',
    'JBF':  '纯黑/米黄',
}

def _cn_color(color_str: str) -> str:
    """Return Chinese colour name for a Color value like 'ESP/ESPRESSO'."""
    code = str(color_str).split('/')[0].strip().upper()
    return CN_COLOR_MAP.get(code, '')

# ── Load & normalise ──────────────────────────────────────────────────────
frames = []
for label, fpath in SOURCES.items():
    df = pd.read_excel(fpath, sheet_name='Summary')
    df = df[df['PO Number'].apply(lambda v: bool(PO_RE.match(str(v).strip())))].copy()
    df.insert(0, 'Program', label)
    for s in SIZE_COLS:
        if s not in df.columns:
            df[s] = 0
    df[SIZE_COLS]     = df[SIZE_COLS].fillna(0).astype(int)
    df['Total Units'] = df['Total Units'].fillna(0).astype(int)
    df['FOB Price']   = pd.to_numeric(df['FOB Price'], errors='coerce')
    df['MSRP']        = pd.to_numeric(df['MSRP'],      errors='coerce')
    df['中文颜色']     = df['Color'].apply(_cn_color)
    frames.append(df)
    print(f'{label}: {len(df)} POs, {int(df["Total Units"].sum()):,} units')

combined = pd.concat(frames, ignore_index=True)
print(f'TOTAL : {len(combined)} POs, {int(combined["Total Units"].sum()):,} units')

# ── Styles ────────────────────────────────────────────────────────────────
HDR_FILL = PatternFill('solid', fgColor='1F5C8B')
SEC_FILL = PatternFill('solid', fgColor='2E75B6')
SUB_FILL = PatternFill('solid', fgColor='BDD7EE')
TOT_FILL = PatternFill('solid', fgColor='1F5C8B')
ALT_FILL = PatternFill('solid', fgColor='DEEAF1')

WHITE = Font(bold=True, color='FFFFFF')
BOLD  = Font(bold=True)
NORM  = Font(bold=False)

thin_s  = Side(style='thin',   color='B8CCE4')
med_s   = Side(style='medium', color='2E75B6')
BORDER  = Border(bottom=thin_s)
SUB_BRD = Border(top=med_s, bottom=med_s)

CENTER = Alignment(horizontal='center', vertical='center')
LEFT   = Alignment(horizontal='left',   vertical='center')

COL_W = {
    'Program': 12, 'PO Number': 14, 'Style': 12, 'Description': 32,
    'Color': 18, '中文颜色': 16, 'ETD': 7, 'FOB Price': 10, 'MSRP': 8,
    'S': 7, 'M': 7, 'L': 7, 'XL': 7, '1X': 7, '2X': 7, '3X': 7,
    'Total Units': 12, 'PO Count': 9,
}


def _hdr(ws, cols, row=1):
    for ci, col in enumerate(cols, 1):
        c = ws.cell(row, ci, col)
        c.fill = HDR_FILL
        c.font = WHITE
        c.alignment = CENTER
    ws.row_dimensions[row].height = 18


def _set_widths(ws, cols):
    for ci, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = COL_W.get(col, 10)


wb = Workbook()

# ══════════════════════════════════════════════════════════════════════════
# Sheet 1 — PO Summary
# ══════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = 'PO Summary'
ws1.freeze_panes = 'A2'

PO_COLS = (['Program', 'PO Number', 'Style', 'Description', 'Color', '中文颜色', 'ETD',
             'FOB Price', 'MSRP'] + SIZE_COLS + ['Total Units'])
_hdr(ws1, PO_COLS)

row = 2
for prog, grp in combined.groupby('Program', sort=False):
    # Section header
    c = ws1.cell(row, 1, prog)
    c.fill = SEC_FILL
    c.font = WHITE
    c.alignment = LEFT
    ws1.merge_cells(start_row=row, start_column=1,
                    end_row=row, end_column=len(PO_COLS))
    ws1.row_dimensions[row].height = 16
    row += 1

    for ri, (_, rd) in enumerate(grp.iterrows()):
        fill = ALT_FILL if ri % 2 == 1 else None
        for ci, col in enumerate(PO_COLS, 1):
            val = rd.get(col, '')
            if pd.isna(val):
                val = ''
            c = ws1.cell(row, ci, val)
            c.font = NORM
            c.alignment = CENTER if ci > 5 else LEFT
            c.border = BORDER
            if fill:
                c.fill = fill
        ws1.row_dimensions[row].height = 15
        row += 1

    # Subtotal
    sub = grp[SIZE_COLS + ['Total Units']].sum()
    for ci, col in enumerate(PO_COLS, 1):
        c = ws1.cell(row, ci)
        c.fill = SUB_FILL
        c.font = BOLD
        c.alignment = LEFT if ci == 1 else CENTER
        c.border = SUB_BRD
        if ci == 1:
            c.value = f'{prog}  —  Sub-Total'
        elif col in SIZE_COLS or col == 'Total Units':
            c.value = int(sub[col])
    ws1.merge_cells(start_row=row, start_column=1,
                    end_row=row, end_column=5)
    ws1.row_dimensions[row].height = 16
    row += 2  # subtotal + blank spacer

# Grand Total
gt = combined[SIZE_COLS + ['Total Units']].sum()
for ci, col in enumerate(PO_COLS, 1):
    c = ws1.cell(row, ci)
    c.fill = TOT_FILL
    c.font = WHITE
    c.alignment = LEFT if ci == 1 else CENTER
    if ci == 1:
        c.value = 'GRAND TOTAL'
    elif col in SIZE_COLS or col == 'Total Units':
        c.value = int(gt[col])
ws1.merge_cells(start_row=row, start_column=1,
                end_row=row, end_column=5)
ws1.row_dimensions[row].height = 18

_set_widths(ws1, PO_COLS)

# ══════════════════════════════════════════════════════════════════════════
# Sheet 2 — Style Rollup
# ══════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('Style Rollup')
ws2.freeze_panes = 'A2'

RL_COLS = (['Program', 'Style', 'Description', 'FOB Price', 'MSRP']
           + SIZE_COLS + ['Total Units', 'PO Count'])
# Style Rollup doesn't have a per-row colour column (it groups by style)
_hdr(ws2, RL_COLS)

row2 = 2
for prog, grp in combined.groupby('Program', sort=False):
    c = ws2.cell(row2, 1, prog)
    c.fill = SEC_FILL
    c.font = WHITE
    c.alignment = LEFT
    ws2.merge_cells(start_row=row2, start_column=1,
                    end_row=row2, end_column=len(RL_COLS))
    ws2.row_dimensions[row2].height = 16
    row2 += 1

    sg = (grp.groupby(['Style', 'Description', 'FOB Price'], dropna=False)
             .agg({**{s: 'sum' for s in SIZE_COLS},
                   'Total Units': 'sum', 'PO Number': 'nunique',
                   'MSRP': 'first'})
             .reset_index()
             .rename(columns={'PO Number': 'PO Count'}))

    for ri, (_, rd) in enumerate(sg.iterrows()):
        fill = ALT_FILL if ri % 2 == 1 else None
        for ci, col in enumerate(RL_COLS, 1):
            val = rd.get(col, '')
            if pd.isna(val):
                val = ''
            if col in SIZE_COLS or col in ('Total Units', 'PO Count'):
                val = int(val) if val != '' else 0
            c = ws2.cell(row2, ci, val)
            c.font = NORM
            c.alignment = CENTER if ci > 2 else LEFT
            c.border = BORDER
            if fill:
                c.fill = fill
        ws2.row_dimensions[row2].height = 15
        row2 += 1

    sub = sg[SIZE_COLS + ['Total Units', 'PO Count']].sum()
    for ci, col in enumerate(RL_COLS, 1):
        c = ws2.cell(row2, ci)
        c.fill = SUB_FILL
        c.font = BOLD
        c.alignment = LEFT if ci == 1 else CENTER
        c.border = SUB_BRD
        if ci == 1:
            c.value = f'{prog}  —  Sub-Total'
        elif col in SIZE_COLS or col in ('Total Units', 'PO Count'):
            c.value = int(sub[col])
    ws2.merge_cells(start_row=row2, start_column=1,
                    end_row=row2, end_column=3)
    ws2.row_dimensions[row2].height = 16
    row2 += 2

gt2   = combined[SIZE_COLS + ['Total Units']].sum()
gt_po = combined['PO Number'].nunique()
for ci, col in enumerate(RL_COLS, 1):
    c = ws2.cell(row2, ci)
    c.fill = TOT_FILL
    c.font = WHITE
    c.alignment = LEFT if ci == 1 else CENTER
    if ci == 1:
        c.value = 'GRAND TOTAL'
    elif col in SIZE_COLS or col == 'Total Units':
        c.value = int(gt2[col])
    elif col == 'PO Count':
        c.value = gt_po
ws2.merge_cells(start_row=row2, start_column=1,
                end_row=row2, end_column=3)
ws2.row_dimensions[row2].height = 18

_set_widths(ws2, RL_COLS)

# ── Save ──────────────────────────────────────────────────────────────────
wb.save(OUT_PATH)
print(f'\nSaved -> {OUT_PATH}')
