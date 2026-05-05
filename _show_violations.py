"""Show the full source-file rows for the 3 rule-violating colours
(Fushia, Pink, Purple) plus every related/similar row in the same file."""
import sys
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from openpyxl import load_workbook
from po_extractor.store.color_translation_store import _normalize_color_name

PATH = "zalandoFile/大货进度表--Angel 2026.xlsx"

# Read every column header so we can dump the full row in human-readable form
wb = load_workbook(PATH, read_only=True, data_only=True)
ws = wb["2026 Zalando "]

rows_iter = ws.iter_rows(values_only=True)
header = list(next(rows_iter))
n_cols = len(header)

# Find the key columns
col = {}
for ci, h in enumerate(header, 1):
    if not isinstance(h, str): continue
    s = h.strip()
    if s in ("颜色", "Color"):       col.setdefault("color_en", ci)
    elif s == "主标颜色":             col["main"] = ci
    elif s in ("中文颜色", "颜色(中文)"): col["color_cn"] = ci
    elif s in ("BRAND", "Brand"):    col.setdefault("brand", ci)
    elif s in ("款式", "style"):      col.setdefault("style", ci)
    elif s == "PO#":                 col.setdefault("po", ci)
    elif s == "数量":                col.setdefault("qty", ci)
    elif s == "合同号":              col.setdefault("contract", ci)

print(f"Header detected at row 1; key columns: {col}\n")

# Collect rows
rows = []
for ri, row in enumerate(rows_iter, 2):
    if not row or all(v is None for v in row):
        continue
    en_raw = row[col["color_en"] - 1] if col["color_en"] <= len(row) else None
    if not en_raw or not str(en_raw).strip():
        continue
    en = _normalize_color_name(en_raw)
    rows.append({
        "_row": ri,
        "en":          en,
        "en_raw":      str(en_raw).strip(),
        "main":        (str(row[col["main"]-1]).strip() if col.get("main") and col["main"]<=len(row) and row[col["main"]-1] else ""),
        "cn":          (str(row[col["color_cn"]-1]).strip() if col.get("color_cn") and col["color_cn"]<=len(row) and row[col["color_cn"]-1] else ""),
        "brand":       (str(row[col["brand"]-1]).strip() if col.get("brand") and col["brand"]<=len(row) and row[col["brand"]-1] else ""),
        "style":       (str(row[col["style"]-1]).strip() if col.get("style") and col["style"]<=len(row) and row[col["style"]-1] else ""),
        "po":          (str(row[col["po"]-1]).strip() if col.get("po") and col["po"]<=len(row) and row[col["po"]-1] else ""),
        "qty":         row[col["qty"]-1] if col.get("qty") and col["qty"]<=len(row) else "",
        "contract":    (str(row[col["contract"]-1]).strip() if col.get("contract") and col["contract"]<=len(row) and row[col["contract"]-1] else ""),
    })
wb.close()

# Filter "None"-strings
for r in rows:
    if r["main"].lower() in ("none", "nan"): r["main"] = ""
    if r["cn"].lower() in ("none", "nan"):   r["cn"] = ""

VIOLATIONS = ["Fushia", "Pink", "Purple"]

# Define which colours are "related" / similar to each violator
RELATED = {
    "Fushia": ["Fushia", "Fuchsia", "Pink", "Magenta", "Hot Pink"],
    "Pink":   ["Pink", "Fushia", "Fuchsia", "Hot Pink", "Rose", "Blush",
               "Coral", "Salmon", "Peach"],
    "Purple": ["Purple", "Violet", "Lilac", "Lavender", "Plum", "Eggplant",
               "Aubergine", "Burgundy", "Magenta"],
}

def matches_any(en: str, keywords: list[str]) -> bool:
    el = en.lower()
    return any(kw.lower() in el for kw in keywords)

def dump_rows(label: str, matched: list[dict]) -> None:
    print(f"\n{'='*100}\n{label}  —  {len(matched)} row(s)\n{'='*100}")
    if not matched:
        print("  (none found)")
        return
    for r in matched:
        print(
            f"  row {r['_row']:3d} | brand={r['brand']:14s} | style={r['style']:18s} | "
            f"po={r['po']:14s} | qty={r['qty']!s:>5} | contract={r['contract']:14s}"
        )
        print(
            f"             EN(raw)  = {r['en_raw']!r}\n"
            f"             EN(norm) = {r['en']!r}\n"
            f"             主标颜色  = {r['main']!r}\n"
            f"             中文颜色  = {r['cn']!r}"
        )

for vio in VIOLATIONS:
    # Exact violator rows
    exact = [r for r in rows if r["en"] == vio and r["main"]]
    dump_rows(f"VIOLATING ROW(S) — EN normalised to {vio!r}", exact)
    # Related rows (same family)
    related_all = [r for r in rows if matches_any(r["en"], RELATED[vio]) and r["en"] != vio]
    # Only keep rows where 主标颜色 is filled — those are useful for comparison
    related_with_main = [r for r in related_all if r["main"]]
    dump_rows(
        f"OTHER ROWS WITH RELATED COLOUR (same family as {vio!r}, with manual 主标颜色)",
        related_with_main,
    )
    related_no_main = [r for r in related_all if not r["main"]]
    if related_no_main:
        print(
            f"\n  (+{len(related_no_main)} more {vio}-related row(s) WITHOUT a manual 主标颜色 — not shown)"
        )
