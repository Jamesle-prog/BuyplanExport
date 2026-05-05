import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side


def read_csv_files(data_path, metadata_path):
    """Read the CSV files."""
    if not os.path.exists(data_path):
        raise FileNotFoundError(f"Data file not found: {data_path}")
    if not os.path.exists(metadata_path):
        raise FileNotFoundError(f"Metadata file not found: {metadata_path}")

    data = pd.read_csv(data_path)
    metadata = pd.read_csv(metadata_path)
    return data, metadata


def transform_data(data):
    """Transform the data by pivoting and reordering."""
    size_order = ['PS', 'PM', 'PL', 'PXL', 'XXS', 'XS', 'S', 'M', 'L', 'XL', '1X', '2X', '3X']
    unique_styles = data['Style'].unique()
    style_data = {}

    for style in unique_styles:
        style_specific_data = data[data['Style'] == style]
        pivot_data = style_specific_data.pivot_table(index=['PO Number', 'Style', 'Color'], columns='Size',
                                                     values='Units', aggfunc='sum', fill_value=0)
        pivot_data = pivot_data.loc[:, (pivot_data != 0).any(axis=0)]
        pivot_data = pivot_data.reindex(columns=[size for size in size_order if size in pivot_data.columns])
        pivot_data['Total'] = pivot_data.sum(axis=1)

        # Add total row for each size
        total_row = pivot_data.sum(axis=0)
        total_row.name = ('Total', style, '')
        pivot_data = pd.concat([pivot_data, pd.DataFrame(total_row).T])

        style_data[style] = pivot_data

    return style_data


def get_output_file_path(base_path, extension='.xlsx'):
    """Generate an output file path with versioning."""
    version = 1
    output_file_path = base_path + extension
    while os.path.isfile(output_file_path):
        version += 1
        output_file_path = f"{base_path}_v{version}{extension}"
    return output_file_path


def adjust_column_widths(worksheet):
    """Adjust column widths to fit the content."""
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = adjusted_width


def format_table(worksheet):
    """Format the table with black boxes, highlight total rows and columns, and add borders."""
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    total_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    total_font = Font(color="000000", bold=True)
    border_style = Side(border_style="thin", color="000000")
    border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

    for row in worksheet.iter_rows(min_row=5, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if cell.row == 5:
                cell.fill = header_fill
                cell.font = header_font
            if cell.row == worksheet.max_row or cell.column == worksheet.max_column:
                cell.fill = total_fill
                cell.font = total_font


def write_to_excel(style_data, metadata, output_file_path):
    """Write the transformed data to an Excel file with additional information."""
    excel_writer = pd.ExcelWriter(output_file_path, engine='openpyxl')

    for style, df in style_data.items():
        sheet_name = style[:31]
        df.to_excel(excel_writer, sheet_name=sheet_name, startrow=4)
        style_metadata = metadata[metadata['Style'] == style]
        factory_info = style_metadata['Factory'].values[0] if not style_metadata.empty else "N/A"
        workbook = excel_writer.book
        worksheet = excel_writer.sheets[sheet_name]
        worksheet.cell(row=1, column=1, value='工厂信息：')
        worksheet.cell(row=1, column=2, value=factory_info)
        worksheet.cell(row=2, column=1, value='款号：')
        worksheet.cell(row=2, column=2, value=style)
        worksheet.cell(row=3, column=1, value='面料信息：')

        # Add and format the headers
        headers = ['PO Number', 'Style', 'Color'] + df.columns.tolist()
        border_style = Side(border_style="thin", color="000000")
        border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=5, column=col_num, value=header)
            cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        adjust_column_widths(worksheet)
        format_table(worksheet)

    excel_writer.close()
    print(f'File saved to {output_file_path}')


def main():
    data_file_path = 'D:\\Users\\Desktop\\data\\all_extracted_data_by_size_color_v1.csv'
    metadata_file_path = 'D:\\Users\\Desktop\\data\\all_extracted_metadata_v1.csv'
    base_output_file_path = 'D:\\Users\\Desktop\\data\\transformed_data_by_style_filtered_with_totals_and_metadata'

    data, metadata = read_csv_files(data_file_path, metadata_file_path)
    style_data = transform_data(data)
    output_file_path = get_output_file_path(base_output_file_path)
    write_to_excel(style_data, metadata, output_file_path)


if __name__ == "__main__":
    main()
